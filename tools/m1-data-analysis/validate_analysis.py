from __future__ import annotations

import csv
import hashlib
import json
import re
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[2]
INPUT_ROOT = ROOT / "data" / "real-bills"
PUBLIC_ROOT = ROOT / "docs" / "analysis" / "m1-real-bills"
PRIVATE_ROOT = ROOT / "data" / "m1-real-bills-private"
SUMMARY_PATH = PUBLIC_ROOT / "summary.json"

EXPECTED_REPORTS = [f"{index:02d}-" for index in range(1, 17)]
REQUIRED_REPORT_SECTIONS = [
    "## 已确认的数据事实",
    "## 候选规则（未启用）",
    "## 无法单靠账单确认的事项",
    "## 需要运营确认的样本",
    "## REQ 与 AT",
    "## PENDING-DATA 状态",
]
HTML_CONTRACT_SECTIONS = [
    "title",
    "technical-summary",
    "key-findings",
    "scope-data-and-metric-definitions",
    "methodology",
    "limitations-uncertainty-and-robustness-checks",
    "recommended-next-steps",
    "further-questions",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def independent_xlsx_profile(path: Path) -> tuple[int, Decimal, set[str], set[str], set[str], set[str]]:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    row_count = 0
    amount_total = Decimal("0")
    work_names: set[str] = set()
    channel_names: set[str] = set()
    work_ids: set[str] = set()
    channel_ids: set[str] = set()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    for ws in workbook.worksheets:
        header_row = None
        for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
            if list(row[:7]) == ["年月", "渠道ID", "文学库渠道名称", "授权分类", "我方作品ID", "作品名称", "实销金额"]:
                header_row = row_number
                break
        if header_row is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, max_col=7, values_only=True):
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in row):
                continue
            row_count += 1
            if row[1] is not None:
                channel_ids.add(str(row[1]).strip())
            if row[2] is not None:
                channel_names.add(str(row[2]).strip())
            if row[4] is not None:
                work_ids.add(str(row[4]).strip())
            if row[5] is not None:
                work_names.add(str(row[5]).strip())
    workbook.close()

    with ZipFile(path) as archive:
        shared = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(ns + "si"):
                shared.append("".join(node.text or "" for node in item.iter(ns + "t")))
        for name in sorted(item for item in archive.namelist() if item.startswith("xl/worksheets/sheet") and item.endswith(".xml")):
            with archive.open(name) as handle:
                for _, cell in ET.iterparse(handle, events=("end",)):
                    if cell.tag != ns + "c":
                        continue
                    reference = cell.attrib.get("r", "")
                    if not re.fullmatch(r"G\d+", reference) or reference == "G1":
                        cell.clear()
                        continue
                    cell_type = cell.attrib.get("t", "n")
                    value = cell.find(ns + "v")
                    if value is None or value.text is None:
                        cell.clear()
                        continue
                    token = shared[int(value.text)] if cell_type == "s" else value.text
                    try:
                        amount_total += Decimal(token.strip())
                    except Exception:
                        pass
                    cell.clear()
    return row_count, amount_total, work_names, channel_names, work_ids, channel_ids


def main() -> None:
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    source_paths = sorted(INPUT_ROOT.glob("*.xlsx")) + sorted(INPUT_ROOT.glob("*.xlsm"))
    independent_rows = 0
    independent_total = Decimal("0")
    work_names: set[str] = set()
    channel_names: set[str] = set()
    work_ids: set[str] = set()
    channel_ids: set[str] = set()
    source_hashes = []
    for source in source_paths:
        rows, total, names, channels, works, channel_keys = independent_xlsx_profile(source)
        independent_rows += rows
        independent_total += total
        work_names |= names
        channel_names |= channels
        work_ids |= works
        channel_ids |= channel_keys
        source_hashes.append({"file_id": f"F{len(source_hashes)+1:03d}", "sha256": sha256_file(source), "size": source.stat().st_size})

    topic_reports = sorted(path for path in PUBLIC_ROOT.glob("*.md") if re.match(r"\d{2}-", path.name))
    missing_report_prefixes = [prefix for prefix in EXPECTED_REPORTS if not any(path.name.startswith(prefix) for path in topic_reports)]
    section_failures = []
    for report in topic_reports:
        text = report.read_text(encoding="utf-8")
        missing = [section for section in REQUIRED_REPORT_SECTIONS if section not in text]
        if missing:
            section_failures.append({"report": report.name, "missing": missing})

    html_path = PUBLIC_ROOT / "report.html"
    html_text = html_path.read_text(encoding="utf-8")
    missing_html_sections = [section for section in HTML_CONTRACT_SECTIONS if f'data-contract-section="{section}"' not in html_text]

    image_checks = []
    for image_path in sorted((PUBLIC_ROOT / "assets").glob("*.png")):
        with Image.open(image_path) as image:
            image.verify()
        with Image.open(image_path) as image:
            image_checks.append({"file": image_path.name, "width": image.width, "height": image.height, "bytes": image_path.stat().st_size})

    public_text_by_file = {
        path.name: path.read_text(encoding="utf-8", errors="ignore")
        for path in PUBLIC_ROOT.rglob("*")
        if path.is_file() and path.suffix.lower() in {".md", ".html", ".json"}
    }
    leak_rows = []
    for value_type, values in [("work_name", work_names), ("channel_name", channel_names)]:
        for value in values:
            if len(value) < 4:
                continue
            for filename, text in public_text_by_file.items():
                if value in text:
                    leak_rows.append({"value_type": value_type, "value": value, "public_file": filename})

    # IDs are checked only when long enough to avoid false matches against counts and dates.
    for value_type, values in [("work_id", work_ids), ("channel_id", channel_ids)]:
        for value in values:
            # Pure six-digit values collide frequently with aggregate row counts;
            # scan identifiers that are structurally distinctive enough to avoid false positives.
            if not (len(value) >= 7 or any(char.isalpha() for char in value) or (len(value) > 1 and value.startswith("0"))):
                continue
            pattern = re.compile(rf"(?<![A-Za-z0-9]){re.escape(value)}(?![A-Za-z0-9])")
            for filename, text in public_text_by_file.items():
                if pattern.search(text):
                    leak_rows.append({"value_type": value_type, "value": value, "public_file": filename})

    leak_path = PRIVATE_ROOT / "public-leak-review.csv"
    with leak_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["value_type", "value", "public_file"])
        writer.writeheader()
        writer.writerows(leak_rows)

    row_match = independent_rows == int(summary["row_count"])
    total_match = independent_total == Decimal(summary["total_amount"])
    source_unchanged = bool(summary.get("source_unchanged"))
    report_complete = not missing_report_prefixes and not section_failures
    html_complete = not missing_html_sections and len(image_checks) >= 4
    privacy_pass = len(leak_rows) == 0

    requirement_files = [
        ROOT / "docs/prd/10-data-foundation/bill-import.md",
        ROOT / "docs/prd/10-data-foundation/data-quality.md",
        ROOT / "docs/prd/10-data-foundation/work-master-data.md",
        ROOT / "docs/prd/10-data-foundation/channel-master-data.md",
        ROOT / "docs/prd/10-data-foundation/classification-and-tags.md",
        ROOT / "docs/prd/40-platform/platform-baseline.md",
    ]
    requirement_ids = []
    for path in requirement_files:
        requirement_ids.extend(re.findall(r"(?m)^## (REQ-[A-Z]+(?:-[A-Z]+)?-\d{3})", path.read_text(encoding="utf-8")))
    trace_text = (ROOT / "docs/prd/00-governance/traceability.md").read_text(encoding="utf-8")
    acceptance_text = (ROOT / "docs/prd/70-acceptance/M1.md").read_text(encoding="utf-8")
    traced_requirements = re.findall(r"REQ-[A-Z]+(?:-[A-Z]+)?-\d{3}", trace_text)
    traced_acceptance = re.findall(r"AT-M1-\d{3}", trace_text)
    acceptance_ids = re.findall(r"(?m)^### (AT-M1-\d{3})", acceptance_text)
    requirement_consistency = (
        len(requirement_ids) == len(set(requirement_ids))
        and set(requirement_ids) == set(traced_requirements)
        and len(acceptance_ids) == len(set(acceptance_ids))
        and set(acceptance_ids) == set(traced_acceptance)
    )
    at_007_match = re.search(r"### AT-M1-007[\s\S]*?(?=\n### |\n## )", acceptance_text)
    at_007_text = at_007_match.group(0) if at_007_match else ""
    at_007_boundary_pass = (
        "M1 不验收评估失效或自动重评" in at_007_text
        and "受影响标准作品集合" in at_007_text
        and "幂等影响记录" in at_007_text
    )
    design_text = (ROOT / "docs/technical-design/M1-技术设计草案-v0.2.md").read_text(encoding="utf-8")
    design_gate_pass = "APPROVED FOR DATA ANALYSIS" in design_text and "物理数据库设计 | NOT APPROVED" in design_text
    ignore_text = (ROOT / ".gitignore").read_text(encoding="utf-8")
    ignore_pass = "/data/real-bills/" in ignore_text and "/data/m1-real-bills-private/" in ignore_text

    overall = (
        row_match
        and total_match
        and source_unchanged
        and report_complete
        and html_complete
        and privacy_pass
        and requirement_consistency
        and at_007_boundary_pass
        and design_gate_pass
        and ignore_pass
    )

    result = {
        "overall": "PASS" if overall else "NEEDS_REVIEW",
        "independent_row_count": independent_rows,
        "row_count_matches": row_match,
        "independent_total_amount": format(independent_total, "f"),
        "total_matches": total_match,
        "source_unchanged": source_unchanged,
        "source_hashes": source_hashes,
        "missing_report_prefixes": missing_report_prefixes,
        "report_section_failures": section_failures,
        "missing_html_sections": missing_html_sections,
        "image_checks": image_checks,
        "public_sensitive_value_matches": len(leak_rows),
        "privacy_pass": privacy_pass,
        "requirement_count": len(requirement_ids),
        "traceability_count": len(traced_requirements),
        "acceptance_count": len(acceptance_ids),
        "requirement_traceability_consistency": requirement_consistency,
        "at_m1_007_boundary_pass": at_007_boundary_pass,
        "data_analysis_approval_gate_pass": design_gate_pass,
        "sensitive_paths_ignored": ignore_pass,
        "browser_render_qa": "BLOCKED: in-app browser runtime lacked sandbox configuration; static HTML and PNG QA completed instead",
    }
    (PUBLIC_ROOT / "validation.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    assessment = "可分享（保留业务待确认项）" if overall else "需修订"
    report = f"""# 分析验证报告

## 总体评估：{assessment}

## 方法与计算复核

- 独立解析行数：{independent_rows:,}；与主分析一致：{'是' if row_match else '否'}。
- 独立读取 XLSX XML 计算总实销：{independent_total:,.12f}；与主分析精确一致：{'是' if total_match else '否'}。
- 原始文件运行前后未变化：{'是' if source_unchanged else '否'}。

## 报告完整性

- 16 份专题报告齐全：{'是' if not missing_report_prefixes else '否'}。
- 每份专题报告均包含事实、候选规则、账单局限、运营样本、REQ/AT 和 PENDING-DATA：{'是' if not section_failures else '否'}。
- HTML 技术报告必需章节齐全：{'是' if not missing_html_sections else '否'}。
- PNG 图表数量：{len(image_checks)}；图片均可解码。

## 文档与追踪一致性

- REQ、追踪矩阵、AT 数量：{len(requirement_ids)} / {len(traced_requirements)} / {len(acceptance_ids)}；一一对应：{'是' if requirement_consistency else '否'}。
- AT-M1-007 仅验收撤销、重算、受影响作品集合和幂等影响记录：{'是' if at_007_boundary_pass else '否'}。
- M1 技术设计仅批准真实数据分析、未批准物理数据库设计：{'是' if design_gate_pass else '否'}。
- 真实账单与私有样本目录已加入 `.gitignore`：{'是' if ignore_pass else '否'}。

## 敏感信息检查

- 公开 Markdown、HTML 和 JSON 中命中的原始作品名、渠道名或长原始 ID：{len(leak_rows)}。
- 详细扫描结果仅保存在 Git 忽略目录 `data/m1-real-bills-private/public-leak-review.csv`。

## 已知限制

- 浏览器渲染 QA 因本地浏览器运行环境缺少沙箱配置而受阻；已完成 HTML 结构、图片解码、尺寸和源文件静态检查。
- 该验证确认计算与交付完整性，不替代运营对渠道、作品、重复、冲抵、分册、首次实销和主数据来源的确认。
"""
    (PUBLIC_ROOT / "validation-report.md").write_text(report, encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
