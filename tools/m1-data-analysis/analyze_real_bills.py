from __future__ import annotations

import csv
import hashlib
import html
import json
import math
import os
import posixpath
import re
import statistics
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

try:
    import psutil
except ImportError:  # Optional: the analysis remains valid without process RSS telemetry.
    psutil = None


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools" / "m1-master-data-analysis"))
from work_id_rules import parse_raw_work_id  # noqa: E402
INPUT_ROOT = ROOT / "data" / "real-bills"
PUBLIC_ROOT = ROOT / "docs" / "analysis" / "m1-real-bills"
ASSET_ROOT = PUBLIC_ROOT / "assets"
PRIVATE_ROOT = ROOT / "data" / "m1-real-bills-private"

EXPECTED_COLUMNS = [
    "年月",
    "渠道ID",
    "文学库渠道名称",
    "授权分类",
    "我方作品ID",
    "作品名称",
    "实销金额",
]

REQ_AT = {
    "file": "REQ-DATA-IMPORT-001~007；AT-M1-001~007",
    "quality": "REQ-DQ-001~003；AT-M1-010~012",
    "work": "REQ-WORK-001~011；AT-M1-020~029、AT-M1-031",
    "channel": "REQ-CHANNEL-001；AT-M1-030",
    "class": "REQ-CLASS-001~002；AT-M1-040~041",
    "platform": "REQ-PLATFORM-001~003；AT-M1-050~052",
}

TOKENS = {
    "surface": "#FCFCFD",
    "panel": "#FFFFFF",
    "ink": "#1F2430",
    "muted": "#6F768A",
    "grid": "#E6E8F0",
    "axis": "#D7DBE7",
    "blue": "#A3BEFA",
    "blue_dark": "#2E4780",
    "orange": "#F0986E",
    "orange_dark": "#804126",
    "gold": "#FFE15B",
    "gold_dark": "#736422",
    "olive": "#A3D576",
    "olive_dark": "#386411",
    "pink": "#F390CA",
    "pink_dark": "#8A3A6F",
}


@dataclass(frozen=True)
class SourceSnapshot:
    path: str
    size: int
    mtime_ns: int
    sha256: str


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def snapshot_sources(paths: Iterable[Path]) -> list[SourceSnapshot]:
    snapshots = []
    for path in paths:
        stat = path.stat()
        snapshots.append(
            SourceSnapshot(
                path=str(path.resolve()),
                size=stat.st_size,
                mtime_ns=stat.st_mtime_ns,
                sha256=sha256_file(path),
            )
        )
    return snapshots


def extract_xlsx_amount_tokens(path: Path) -> dict[tuple[int, int], tuple[str, str, bool]]:
    """Read exact column-G cell tokens from XLSX XML without float conversion."""
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    office_rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    result: dict[tuple[int, int], tuple[str, str, bool]] = {}
    with ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(main_ns + "si"):
                shared_strings.append("".join(node.text or "" for node in item.iter(main_ns + "t")))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target_by_id = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(package_rel_ns + "Relationship")
        }
        sheets = workbook.find(main_ns + "sheets")
        if sheets is None:
            return result
        for sheet_index, sheet in enumerate(sheets.findall(main_ns + "sheet"), start=1):
            relationship_id = sheet.attrib[office_rel_ns + "id"]
            target = target_by_id[relationship_id].replace("\\", "/")
            if target.startswith("/"):
                sheet_path = target.lstrip("/")
            else:
                sheet_path = posixpath.normpath(posixpath.join("xl", target))
            with archive.open(sheet_path) as handle:
                for _, cell in ET.iterparse(handle, events=("end",)):
                    if cell.tag != main_ns + "c":
                        continue
                    reference = cell.attrib.get("r", "")
                    match = re.fullmatch(r"G(\d+)", reference)
                    if not match:
                        cell.clear()
                        continue
                    row_number = int(match.group(1))
                    storage_type = cell.attrib.get("t", "n")
                    formula = cell.find(main_ns + "f") is not None
                    value_node = cell.find(main_ns + "v")
                    inline_node = cell.find(main_ns + "is")
                    if storage_type == "s" and value_node is not None and value_node.text is not None:
                        token = shared_strings[int(value_node.text)]
                    elif storage_type == "inlineStr" and inline_node is not None:
                        token = "".join(node.text or "" for node in inline_node.iter(main_ns + "t"))
                    else:
                        token = value_node.text if value_node is not None and value_node.text is not None else ""
                    result[(sheet_index, row_number)] = (token, storage_type, formula)
                    cell.clear()
    return result


def assert_sources_unchanged(before: list[SourceSnapshot]) -> None:
    after = snapshot_sources(Path(item.path) for item in before)
    if before != after:
        raise RuntimeError("原始账单在分析期间发生变化，分析已中止。")


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value)) or (isinstance(value, str) and not value.strip())


def raw_text(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def norm_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", raw_text(value)).strip()


def norm_name(value: Any) -> str:
    text = norm_text(value).lower()
    return re.sub(r"\s+", " ", text)


def sample_code(prefix: str, *values: Any) -> str:
    payload = "\x1f".join(raw_text(v) for v in values)
    return f"{prefix}-{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:10].upper()}"


def parse_month(value: Any) -> tuple[str | None, str, bool]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m"), "excel_datetime", value.day == 1
    if isinstance(value, date):
        return value.strftime("%Y-%m"), "excel_date", value.day == 1
    text = norm_text(value)
    if not text:
        return None, "blank", False
    patterns = [
        (r"^(\d{4})-(\d{2})-(\d{2})$", "text_yyyy_mm_dd"),
        (r"^(\d{4})/(\d{1,2})/(\d{1,2})$", "text_yyyy_slash_mm_dd"),
        (r"^(\d{4})-(\d{1,2})$", "text_yyyy_mm"),
        (r"^(\d{4})(\d{2})$", "text_yyyymm"),
    ]
    for pattern, label in patterns:
        match = re.match(pattern, text)
        if not match:
            continue
        year, month = int(match.group(1)), int(match.group(2))
        day = int(match.group(3)) if len(match.groups()) == 3 else 1
        try:
            parsed = date(year, month, day)
        except ValueError:
            return None, f"invalid_{label}", False
        return parsed.strftime("%Y-%m"), label, day == 1
    return None, f"unparsed_{type(value).__name__}", False


def parse_decimal(value: Any) -> tuple[Decimal | None, str, int | None]:
    if value is None:
        return None, "blank", None
    if isinstance(value, bool):
        return None, "boolean", None
    if isinstance(value, int):
        return Decimal(value), "integer", 0
    if isinstance(value, float):
        if not math.isfinite(value):
            return None, "non_finite_float", None
        dec = Decimal(str(value))
        return dec, "float", max(0, -dec.as_tuple().exponent)
    text = norm_text(value)
    if not text:
        return None, "blank", None
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        try:
            dec = Decimal(text)
            return dec, "decimal_text", max(0, -dec.as_tuple().exponent)
        except InvalidOperation:
            pass
    if re.fullmatch(r"[+-]?(?:\d+(?:\.\d+)?|\.\d+)[eE][+-]?\d+", text):
        try:
            dec = Decimal(text)
            return dec, "scientific_text", max(0, -dec.as_tuple().exponent)
        except InvalidOperation:
            return None, "invalid_scientific_text", None
    if any(char in text for char in [",", "¥", "￥", "(", ")"]):
        return None, "decorated_text", None
    return None, "invalid_text", None


def classify_work_id(value: Any) -> tuple[str, str | None, str | None, str | None]:
    parsed = parse_raw_work_id(value)
    if parsed.valid:
        if parsed.format == "pure_digits":
            category = "text_digits_leading_zero" if len(parsed.raw) > 1 and parsed.raw.startswith("0") else "text_digits"
        else:
            body = parsed.raw[1:]
            category = "text_Ydigits_leading_zero" if len(body) > 1 and body.startswith("0") else "text_Ydigits"
        return category, parsed.normalized_raw, parsed.standard_id, parsed.business_form
    category_by_format = {
        "blank": "blank",
        "lowercase_y_prefix": "text_lower_y_digits",
        "whitespace_variant": "text_whitespace_variant",
        "decimal_integer_text": "text_decimal_integer",
    }
    text = raw_text(value)
    if re.fullmatch(r"[+-]?(?:\d+(?:\.\d+)?|\.\d+)[eE][+-]?\d+", text):
        return "text_scientific", None, None, None
    return category_by_format.get(parsed.format, "other_text"), None, None, None


def decimal_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    return format(value, "f")


def excel_display_scale(number_format: str) -> int | None:
    if not number_format or number_format.lower() == "general":
        return None
    positive_section = number_format.split(";", 1)[0]
    if "." not in positive_section:
        return 0
    decimal_part = positive_section.split(".", 1)[1]
    decimal_part = re.split(r"[^0#?]", decimal_part, maxsplit=1)[0]
    return sum(char in "0#?" for char in decimal_part)


def format_decimal(value: Decimal | None) -> str:
    if value is None:
        return "不可计算"
    return f"{value:,.3f}".rstrip("0").rstrip(".")


def percent(count: int, total: int) -> str:
    return "0.00%" if total == 0 else f"{count / total:.2%}"


def md_table(headers: list[str], rows: list[list[Any]]) -> str:
    def esc(value: Any) -> str:
        return str(value).replace("|", "\\|").replace("\n", " ")

    lines = ["| " + " | ".join(map(esc, headers)) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    lines.extend("| " + " | ".join(esc(v) for v in row) + " |" for row in rows)
    return "\n".join(lines)


def write_markdown_report(
    filename: str,
    title: str,
    facts: list[str],
    candidates: list[str],
    cannot_confirm: list[str],
    review_samples: list[str],
    req_at: str,
    pending: str,
    extra: str = "",
) -> None:
    def bullets(items: list[str]) -> str:
        return "\n".join(f"- {item}" for item in items) if items else "- 无。"

    text = f"""# {title}

## 已确认的数据事实

{bullets(facts)}

## 候选规则（未启用）

{bullets(candidates)}

## 无法单靠账单确认的事项

{bullets(cannot_confirm)}

## 需要运营确认的样本

{bullets(review_samples)}

## REQ 与 AT

{req_at}

## PENDING-DATA 状态

{pending}
"""
    if extra:
        text += "\n" + extra.strip() + "\n"
    (PUBLIC_ROOT / filename).write_text(text, encoding="utf-8")


def write_private_csv(name: str, fieldnames: list[str], rows: Iterable[dict[str, Any]], limit: int = 5000) -> int:
    path = PRIVATE_ROOT / name
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: raw_text(row.get(key, "")) for key in fieldnames})
            count += 1
            if count >= limit:
                break
    return count


def use_chart_theme() -> None:
    sns.set_theme(
        style="whitegrid",
        rc={
            "figure.facecolor": TOKENS["surface"],
            "axes.facecolor": TOKENS["panel"],
            "axes.edgecolor": TOKENS["axis"],
            "axes.labelcolor": TOKENS["ink"],
            "grid.color": TOKENS["grid"],
            "font.family": "sans-serif",
            "font.sans-serif": ["Microsoft YaHei", "Segoe UI", "DejaVu Sans", "Arial"],
            "axes.spines.top": False,
            "axes.spines.right": False,
        },
    )


def add_chart_header(fig: plt.Figure, ax: plt.Axes, title: str, subtitle: str) -> None:
    ax.set_title("")
    fig.subplots_adjust(top=0.80, left=0.12, right=0.96, bottom=0.16)
    left = ax.get_position().x0
    fig.text(left, 0.96, title, ha="left", va="top", fontsize=14, fontweight="semibold", color=TOKENS["ink"])
    fig.text(left, 0.90, subtitle, ha="left", va="top", fontsize=9, color=TOKENS["muted"])
    sns.despine(ax=ax)


def save_chart(fig: plt.Figure, filename: str) -> str:
    path = ASSET_ROOT / filename
    fig.savefig(path, dpi=160, bbox_inches="tight", facecolor=TOKENS["surface"])
    plt.close(fig)
    return f"assets/{filename}"


def row_private_dict(row: pd.Series) -> dict[str, Any]:
    business_values = {column: row[column] for column in EXPECTED_COLUMNS}
    business_values["实销金额"] = row["_amount_raw_token"]
    return {
        "样本编号": row["_sample_id"],
        "文件编号": row["_file_id"],
        "工作表编号": row["_sheet_id"],
        "原始行号": row["_row_num"],
        **business_values,
    }


def main() -> None:
    started = time.perf_counter()
    PUBLIC_ROOT.mkdir(parents=True, exist_ok=True)
    ASSET_ROOT.mkdir(parents=True, exist_ok=True)
    PRIVATE_ROOT.mkdir(parents=True, exist_ok=True)

    paths = sorted(path for path in INPUT_ROOT.rglob("*") if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"})
    if not paths:
        raise RuntimeError("data/real-bills/ 中没有可分析的 XLSX/XLSM 文件。")
    before = snapshot_sources(paths)

    records: list[dict[str, Any]] = []
    structures: list[dict[str, Any]] = []
    file_public: list[dict[str, Any]] = []
    content_hashes: dict[str, dict[str, str]] = {}
    parse_started = time.perf_counter()

    for file_index, path in enumerate(paths, start=1):
        file_id = f"F{file_index:03d}"
        snap = next(item for item in before if item.path == str(path.resolve()))
        file_public.append({"file_id": file_id, "size": snap.size, "sha256": snap.sha256})
        exact_amount_tokens = extract_xlsx_amount_tokens(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        order_digest = hashlib.sha256()
        sorted_row_hashes: list[str] = []
        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            sheet_id = f"{file_id}-S{sheet_index:03d}"
            header_row = None
            header_values = None
            for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
                values = [norm_text(value) for value in row]
                if values[: len(EXPECTED_COLUMNS)] == EXPECTED_COLUMNS:
                    header_row = row_num
                    header_values = values
                    break
            structures.append(
                {
                    "file_id": file_id,
                    "sheet_id": sheet_id,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "header_row": header_row,
                    "header_match": header_values == EXPECTED_COLUMNS if header_values else False,
                }
            )
            if header_row is None:
                continue
            for row_num, cells in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=len(EXPECTED_COLUMNS)), start=header_row + 1):
                values = [cell.value for cell in cells]
                if all(is_blank(value) for value in values):
                    continue
                record = {column: value for column, value in zip(EXPECTED_COLUMNS, values)}
                amount_token, amount_storage_type, amount_formula = exact_amount_tokens.get((sheet_index, row_num), ("", "missing", False))
                record.update(
                    {
                        "_file_id": file_id,
                        "_sheet_id": sheet_id,
                        "_row_num": row_num,
                        "_cell_types": [cell.data_type for cell in cells],
                        "_number_formats": [cell.number_format for cell in cells],
                        "_amount_raw_token": amount_token,
                        "_amount_storage_type": amount_storage_type,
                        "_amount_formula": amount_formula,
                    }
                )
                canonical_values = [raw_text(value) for value in values]
                canonical_values[6] = amount_token
                canonical = "\x1f".join(canonical_values)
                row_hash = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
                order_digest.update((row_hash + "\n").encode("ascii"))
                sorted_row_hashes.append(row_hash)
                record["_row_hash"] = row_hash
                record["_sample_id"] = sample_code("ROW", file_id, sheet_id, row_num, row_hash)
                records.append(record)
        wb.close()
        set_digest = hashlib.sha256()
        for row_hash in sorted(sorted_row_hashes):
            set_digest.update((row_hash + "\n").encode("ascii"))
        content_hashes[file_id] = {"order_sensitive": order_digest.hexdigest(), "order_insensitive": set_digest.hexdigest()}

    parse_seconds = time.perf_counter() - parse_started
    df = pd.DataFrame.from_records(records)
    row_count = len(df)
    if row_count == 0:
        raise RuntimeError("未解析到数据行。")

    analysis_started = time.perf_counter()
    df["_month"], df["_month_format"], df["_month_first_day"] = zip(*df["年月"].map(parse_month))
    decimal_parts = df["_amount_raw_token"].map(parse_decimal)
    df["_amount"] = [part[0] for part in decimal_parts]
    df["_amount_format"] = [part[1] for part in decimal_parts]
    df["_amount_scale"] = [part[2] for part in decimal_parts]
    work_parts = df["我方作品ID"].map(classify_work_id)
    df["_work_id_format"] = [part[0] for part in work_parts]
    df["_work_id_norm"] = [part[1] for part in work_parts]
    df["_standard_work_id"] = [part[2] for part in work_parts]
    df["_business_form"] = [part[3] for part in work_parts]
    df["_channel_id_norm"] = df["渠道ID"].map(norm_text)
    df["_channel_name_norm"] = df["文学库渠道名称"].map(norm_name)
    df["_work_name_norm"] = df["作品名称"].map(norm_name)
    df["_auth_norm"] = df["授权分类"].map(norm_text)
    df["_amount_text"] = df["_amount"].map(decimal_str)
    df["_amount_number_format"] = df["_number_formats"].map(lambda values: values[6])
    df["_amount_display_scale"] = df["_amount_number_format"].map(excel_display_scale)

    valid_amounts = [value for value in df["_amount"] if value is not None]
    total_amount = sum(valid_amounts, Decimal("0"))
    positive_count = sum(value > 0 for value in valid_amounts)
    zero_count = sum(value == 0 for value in valid_amounts)
    negative_count = sum(value < 0 for value in valid_amounts)
    amount_invalid_count = row_count - len(valid_amounts)
    max_scale = max((int(scale) for scale in df["_amount_scale"] if scale is not None and not pd.isna(scale)), default=None)
    max_integer_digits = max((len(str(abs(int(value)))) for value in valid_amounts), default=0)
    amount_scale_counts = df["_amount_scale"].dropna().astype(int).value_counts().sort_index()
    hidden_precision_mask = (
        df["_amount_scale"].notna()
        & df["_amount_display_scale"].notna()
        & (df["_amount_scale"].astype(float) > df["_amount_display_scale"].astype(float))
    )
    hidden_precision_count = int(hidden_precision_mask.sum())
    display_rounded_total = Decimal("0")
    rounded_2_total = Decimal("0")
    rounded_3_total = Decimal("0")
    for amount, display_scale in zip(df["_amount"], df["_amount_display_scale"]):
        if amount is None:
            continue
        rounded_2_total += amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        rounded_3_total += amount.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        if display_scale is None or pd.isna(display_scale):
            display_rounded_total += amount
        else:
            quantum = Decimal("1").scaleb(-int(display_scale))
            display_rounded_total += amount.quantize(quantum, rounding=ROUND_HALF_UP)
    amount_storage_counts = df["_amount_storage_type"].value_counts(dropna=False)
    amount_formula_count = int(df["_amount_formula"].sum())

    valid_month_df = df[df["_month"].notna()].copy()
    months = sorted(valid_month_df["_month"].unique())
    min_month = months[0] if months else None
    max_month = months[-1] if months else None
    month_invalid_count = row_count - len(valid_month_df)
    month_non_first_day = int((valid_month_df["_month_first_day"] == False).sum())
    if months:
        expected_months = pd.period_range(min_month, max_month, freq="M").astype(str).tolist()
        missing_calendar_months = [month for month in expected_months if month not in set(months)]
    else:
        missing_calendar_months = []

    field_profile = []
    for column in EXPECTED_COLUMNS:
        missing = int(df[column].map(is_blank).sum())
        distinct = int(df[column].map(raw_text).nunique(dropna=False))
        max_length = int(df[column].map(lambda value: len(raw_text(value))).max())
        type_counts = Counter(type(value).__name__ for value in df[column])
        field_profile.append(
            {
                "field": column,
                "missing": missing,
                "missing_rate": missing / row_count,
                "distinct": distinct,
                "max_length": max_length,
                "types": dict(type_counts),
            }
        )

    channel_id_to_names = df.groupby("_channel_id_norm")["_channel_name_norm"].nunique()
    channel_name_to_ids = df.groupby("_channel_name_norm")["_channel_id_norm"].nunique()
    channel_multi_name_ids = set(channel_id_to_names[channel_id_to_names > 1].index)
    channel_multi_id_names = set(channel_name_to_ids[channel_name_to_ids > 1].index)
    channel_conflict_rows = df[
        df["_channel_id_norm"].isin(channel_multi_name_ids) | df["_channel_name_norm"].isin(channel_multi_id_names)
    ]

    work_id_to_names = df[df["_work_id_norm"].notna()].groupby("_work_id_norm")["_work_name_norm"].nunique()
    work_id_to_auth = df[df["_work_id_norm"].notna()].groupby("_work_id_norm")["_auth_norm"].nunique()
    work_name_conflict_ids = set(work_id_to_names[work_id_to_names > 1].index)
    work_auth_conflict_ids = set(work_id_to_auth[work_id_to_auth > 1].index)
    work_conflict_rows = df[
        df["_work_id_norm"].isin(work_name_conflict_ids | work_auth_conflict_ids)
    ]

    exact_keys = ["_file_id", "_sheet_id"] + EXPECTED_COLUMNS[:-1] + ["_amount_raw_token"]
    exact_dup_mask = df.duplicated(exact_keys, keep=False)
    exact_dup_rows = df[exact_dup_mask]
    exact_group_sizes = exact_dup_rows.groupby(exact_keys, dropna=False).size() if not exact_dup_rows.empty else pd.Series(dtype=int)
    exact_excess_rows = int((exact_group_sizes - 1).clip(lower=0).sum()) if len(exact_group_sizes) else 0
    exact_excess_amount = Decimal("0")
    if len(exact_group_sizes):
        for key, size in exact_group_sizes.items():
            amount = parse_decimal(key[-1])[0]
            if amount is not None:
                exact_excess_amount += amount * (int(size) - 1)

    normalized_dup_keys = ["_file_id", "_sheet_id", "_month", "_channel_id_norm", "_channel_name_norm", "_auth_norm", "_work_id_norm", "_work_name_norm", "_amount_text"]
    normalized_valid = df[df["_month"].notna() & df["_work_id_norm"].notna() & df["_amount"].notna()]
    normalized_dup_mask = normalized_valid.duplicated(normalized_dup_keys, keep=False)
    normalized_dup_rows = normalized_valid[normalized_dup_mask]
    normalized_groups = normalized_dup_rows.groupby(normalized_dup_keys, dropna=False).size() if not normalized_dup_rows.empty else pd.Series(dtype=int)
    normalized_excess_rows = int((normalized_groups - 1).clip(lower=0).sum()) if len(normalized_groups) else 0

    offset_groups = []
    amount_group_df = df[df["_month"].notna() & df["_work_id_norm"].notna() & df["_amount"].notna()].copy()
    for key, group in amount_group_df.groupby(["_month", "_channel_id_norm", "_work_id_norm"], dropna=False):
        amounts = list(group["_amount"])
        if any(value > 0 for value in amounts) and any(value < 0 for value in amounts):
            total = sum(amounts, Decimal("0"))
            offset_groups.append(
                {
                    "sample_id": sample_code("OFFSET", *key),
                    "month": key[0],
                    "channel_id": key[1],
                    "work_id": key[2],
                    "row_count": len(group),
                    "sum": total,
                    "exact_zero": total == 0,
                    "rows": group,
                }
            )

    amount_group_df["_abs_amount_text"] = amount_group_df["_amount"].map(lambda value: decimal_str(abs(value)))
    cross_month_offset_groups = 0
    cross_month_offset_details = []
    for key, group in amount_group_df.groupby(["_channel_id_norm", "_work_id_norm", "_abs_amount_text"], dropna=False):
        amounts = list(group["_amount"])
        if group["_month"].nunique() > 1 and any(value > 0 for value in amounts) and any(value < 0 for value in amounts):
            cross_month_offset_groups += 1
            cross_month_offset_details.append((sample_code("XOFFSET", *key), group))

    name_to_work_ids = (
        df[(df["_work_name_norm"] != "") & df["_work_id_norm"].notna()]
        .groupby("_work_name_norm")["_standard_work_id"]
        .agg(lambda values: sorted(set(values)))
    )
    same_name_multi_ids = name_to_work_ids[name_to_work_ids.map(len) > 1]
    volume_pattern = re.compile(r"(?:第?[0-9一二三四五六七八九十百]+[册卷部季集]|[（(][0-9一二三四五六七八九十]+[）)]|[上下中]册|[上下中]部)$")
    volume_marker_rows = df[df["_work_name_norm"].map(lambda value: bool(volume_pattern.search(value)))]
    same_name_candidate_rows = df[df["_work_name_norm"].isin(set(same_name_multi_ids.index))]
    volume_review_rows = pd.concat([same_name_candidate_rows, volume_marker_rows]).drop_duplicates(subset=["_file_id", "_sheet_id", "_row_num"])

    launch_df = df[df["_standard_work_id"].notna() & df["_month"].notna() & df["_amount"].notna()].copy()
    launch_records = []
    for standard_id, group in launch_df.groupby("_standard_work_id"):
        first_month = group["_month"].min()
        first = group[group["_month"] == first_month]
        amounts = list(first["_amount"])
        total = sum(amounts, Decimal("0"))
        signs = {"positive" if value > 0 else "negative" if value < 0 else "zero" for value in amounts}
        launch_records.append(
            {
                "standard_work_id": standard_id,
                "sample_id": sample_code("LAUNCH", standard_id),
                "first_month": first_month,
                "row_count": len(first),
                "total": total,
                "signs": "+".join(sorted(signs)),
                "has_nonpositive": any(value <= 0 for value in amounts),
                "nonpositive_total": total <= 0,
            }
        )
    launch_summary = pd.DataFrame(launch_records)
    launch_nonpositive_rows = launch_summary[launch_summary["has_nonpositive"]] if not launch_summary.empty else launch_summary
    launch_nonpositive_total = launch_summary[launch_summary["nonpositive_total"]] if not launch_summary.empty else launch_summary

    form_launch_records = []
    for (standard_id, form), group in launch_df.groupby(["_standard_work_id", "_business_form"], dropna=False):
        first_month = group["_month"].min()
        first = group[group["_month"] == first_month]
        amounts = list(first["_amount"])
        total = sum(amounts, Decimal("0"))
        form_launch_records.append(
            {
                "standard_work_id": standard_id,
                "business_form": form,
                "sample_id": sample_code("FORM-LAUNCH", standard_id, form),
                "first_month": first_month,
                "total": total,
                "has_nonpositive": any(value <= 0 for value in amounts),
                "nonpositive_total": total <= 0,
            }
        )
    form_launch_summary = pd.DataFrame(form_launch_records)

    work_month = (
        launch_df.groupby(["_standard_work_id", "_month"])["_amount"]
        .agg(lambda values: sum(values, Decimal("0")))
        .reset_index()
    )
    status_records = []
    max_period = pd.Period(max_month, freq="M") if max_month else None
    for standard_id, group in work_month.groupby("_standard_work_id"):
        sorted_months = sorted(group["_month"].unique())
        periods = [pd.Period(month, freq="M") for month in sorted_months]
        last_period = periods[-1]
        months_since = int(max_period.ordinal - last_period.ordinal) if max_period is not None else None
        gaps = [int(periods[i].ordinal - periods[i - 1].ordinal - 1) for i in range(1, len(periods))]
        last_total = group.loc[group["_month"] == sorted_months[-1], "_amount"].iloc[0]
        status_records.append(
            {
                "standard_work_id": standard_id,
                "sample_id": sample_code("STATUS", standard_id),
                "last_month": sorted_months[-1],
                "months_since_last_record": months_since,
                "max_missing_gap": max(gaps, default=0),
                "reappeared_after_gap": any(gap > 0 for gap in gaps),
                "last_month_sign": "positive" if last_total > 0 else "negative" if last_total < 0 else "zero",
            }
        )
    status_df = pd.DataFrame(status_records)

    issue_masks: dict[str, pd.Series] = {
        "年月缺失或无法解析": df["_month"].isna(),
        "金额缺失或无法精确解析": df["_amount"].isna(),
        "作品ID无法按当前规则识别": df["_work_id_norm"].isna(),
        "渠道ID缺失": df["渠道ID"].map(is_blank),
        "渠道名称缺失": df["文学库渠道名称"].map(is_blank),
        "授权分类缺失": df["授权分类"].map(is_blank),
        "作品名称缺失": df["作品名称"].map(is_blank),
        "完全重复候选": exact_dup_mask,
        "渠道ID/名称多重关系待确认": df.index.isin(channel_conflict_rows.index),
        "作品ID对应多名称或多授权分类": df.index.isin(work_conflict_rows.index),
    }
    any_issue = pd.Series(False, index=df.index)
    issue_counts = []
    for issue_type, mask in issue_masks.items():
        affected = int(mask.sum())
        any_issue |= mask
        issue_counts.append({"issue_type": issue_type, "affected_rows": affected, "rate": affected / row_count})
    issue_rows_count = int(any_issue.sum())

    auth_values = df["_auth_norm"].value_counts(dropna=False)
    work_format_counts = df["_work_id_format"].value_counts(dropna=False)
    amount_format_counts = df["_amount_format"].value_counts(dropna=False)
    month_format_counts = df["_month_format"].value_counts(dropna=False)
    monthly_counts = valid_month_df.groupby("_month").size().sort_index()
    latest_month_rows = int(monthly_counts.iloc[-1]) if not monthly_counts.empty else 0
    prior_month_rows = int(monthly_counts.iloc[-2]) if len(monthly_counts) >= 2 else None
    prior_12_month_median_rows = float(monthly_counts.iloc[-13:-1].median()) if len(monthly_counts) >= 13 else None
    sign_counts = pd.Series({"正数": positive_count, "零值": zero_count, "负数": negative_count, "无效": amount_invalid_count})

    analysis_seconds = time.perf_counter() - analysis_started
    rss_mb = psutil.Process(os.getpid()).memory_info().rss / 1024 / 1024 if psutil is not None else None

    # Private operator review samples. These files are intentionally under a Git-ignored directory.
    private_manifest = {}
    base_private_fields = ["样本编号", "文件编号", "工作表编号", "原始行号"] + EXPECTED_COLUMNS
    private_manifest["channel-review.csv"] = write_private_csv(
        "channel-review.csv", base_private_fields, (row_private_dict(row) for _, row in channel_conflict_rows.iterrows())
    )
    private_manifest["work-name-auth-review.csv"] = write_private_csv(
        "work-name-auth-review.csv", base_private_fields, (row_private_dict(row) for _, row in work_conflict_rows.iterrows())
    )
    private_manifest["duplicate-candidates.csv"] = write_private_csv(
        "duplicate-candidates.csv", base_private_fields, (row_private_dict(row) for _, row in exact_dup_rows.iterrows())
    )
    offset_private_rows = []
    for item in offset_groups:
        for _, row in item["rows"].iterrows():
            offset_private_rows.append({"冲抵候选类型": "同月正负并存", "冲抵候选编号": item["sample_id"], **row_private_dict(row)})
    for offset_id, group in cross_month_offset_details:
        for _, row in group.iterrows():
            offset_private_rows.append({"冲抵候选类型": "跨月同绝对值正负", "冲抵候选编号": offset_id, **row_private_dict(row)})
    private_manifest["offset-candidates.csv"] = write_private_csv(
        "offset-candidates.csv", ["冲抵候选类型", "冲抵候选编号"] + base_private_fields, offset_private_rows
    )
    private_manifest["volume-candidates.csv"] = write_private_csv(
        "volume-candidates.csv", base_private_fields, (row_private_dict(row) for _, row in volume_review_rows.iterrows())
    )

    launch_private_rows = []
    for _, item in launch_nonpositive_rows.head(5000).iterrows():
        raw_rows = launch_df[(launch_df["_standard_work_id"] == item["standard_work_id"]) & (launch_df["_month"] == item["first_month"])]
        for _, row in raw_rows.iterrows():
            result = row_private_dict(row)
            result["首次月份样本编号"] = item["sample_id"]
            result["首次月份"] = item["first_month"]
            result["首次月份合计"] = item["total"]
            launch_private_rows.append(result)
    private_manifest["launch-sign-review.csv"] = write_private_csv(
        "launch-sign-review.csv",
        ["首次月份样本编号", "首次月份", "首次月份合计"] + base_private_fields,
        launch_private_rows,
    )

    invalid_mask = issue_masks["年月缺失或无法解析"] | issue_masks["金额缺失或无法精确解析"] | issue_masks["作品ID无法按当前规则识别"]
    private_manifest["invalid-required-fields.csv"] = write_private_csv(
        "invalid-required-fields.csv", base_private_fields, (row_private_dict(row) for _, row in df[invalid_mask].iterrows())
    )
    private_manifest["amount-precision-review.csv"] = write_private_csv(
        "amount-precision-review.csv",
        ["原始小数位", "Excel显示小数位", "Excel数字格式"] + base_private_fields,
        (
            {
                "原始小数位": row["_amount_scale"],
                "Excel显示小数位": row["_amount_display_scale"],
                "Excel数字格式": row["_amount_number_format"],
                **row_private_dict(row),
            }
            for _, row in df[hidden_precision_mask | (df["_amount_format"] == "scientific_text")].iterrows()
        ),
    )

    status_private_rows = []
    if not status_df.empty:
        for _, item in status_df.sort_values(["months_since_last_record", "max_missing_gap"], ascending=False).head(2000).iterrows():
            work_rows = df[df["_standard_work_id"] == item["standard_work_id"]].head(1)
            if work_rows.empty:
                continue
            row = work_rows.iloc[0]
            result = row_private_dict(row)
            result.update(
                {
                    "状态样本编号": item["sample_id"],
                    "最后出现月份": item["last_month"],
                    "距数据截止月份数": item["months_since_last_record"],
                    "最大缺月间隔": item["max_missing_gap"],
                    "最后月份金额符号": item["last_month_sign"],
                }
            )
            status_private_rows.append(result)
    private_manifest["product-status-review.csv"] = write_private_csv(
        "product-status-review.csv",
        ["状态样本编号", "最后出现月份", "距数据截止月份数", "最大缺月间隔", "最后月份金额符号"] + base_private_fields,
        status_private_rows,
    )

    representative_categories = {
        "NEGATIVE_AMOUNT": df[df["_amount"].map(lambda value: value is not None and value < 0)],
        "ZERO_AMOUNT": df[df["_amount"].map(lambda value: value is not None and value == 0)],
        "EXACT_DUPLICATE": exact_dup_rows,
        "CHANNEL_RELATION": channel_conflict_rows,
        "WORK_NAME_AUTH_CONFLICT": work_conflict_rows,
        "WORK_ID_INVALID": df[df["_work_id_norm"].isna()],
        "VOLUME_MARKER": volume_marker_rows,
    }
    representative_rows = []
    for category, part in representative_categories.items():
        for _, row in part.sort_values("_sample_id").head(10).iterrows():
            result = row_private_dict(row)
            result["验收样本类别"] = category
            representative_rows.append(result)
    private_manifest["representative-acceptance-rows.csv"] = write_private_csv(
        "representative-acceptance-rows.csv", ["验收样本类别"] + base_private_fields, representative_rows
    )

    private_manifest_path = PRIVATE_ROOT / "manifest.json"
    private_manifest_path.write_text(json.dumps(private_manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    # Public report data contains aggregate statistics and anonymized sample codes only.
    summary = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "source_file_count": len(paths),
        "sheet_count": len(structures),
        "row_count": row_count,
        "month_min": min_month,
        "month_max": max_month,
        "month_count": len(months),
        "total_amount": decimal_str(total_amount),
        "positive_count": positive_count,
        "zero_count": zero_count,
        "negative_count": negative_count,
        "invalid_amount_count": amount_invalid_count,
        "max_amount_scale": max_scale,
        "max_amount_integer_digits": max_integer_digits,
        "amount_hidden_precision_count": hidden_precision_count,
        "amount_formula_count": amount_formula_count,
        "display_rounded_total": decimal_str(display_rounded_total),
        "display_rounding_difference": decimal_str(display_rounded_total - total_amount),
        "row_rounded_2_total": decimal_str(rounded_2_total),
        "row_rounded_2_difference": decimal_str(rounded_2_total - total_amount),
        "row_rounded_3_total": decimal_str(rounded_3_total),
        "row_rounded_3_difference": decimal_str(rounded_3_total - total_amount),
        "invalid_month_count": month_invalid_count,
        "non_first_day_month_count": month_non_first_day,
        "missing_calendar_month_count": len(missing_calendar_months),
        "latest_month_rows": latest_month_rows,
        "prior_month_rows": prior_month_rows,
        "prior_12_month_median_rows": prior_12_month_median_rows,
        "latest_month_completeness": "PENDING_OPERATION_CONFIRMATION",
        "distinct_channel_ids": int(df["_channel_id_norm"].nunique()),
        "distinct_channel_names": int(df["_channel_name_norm"].nunique()),
        "channel_ids_with_multiple_names": len(channel_multi_name_ids),
        "channel_names_with_multiple_ids": len(channel_multi_id_names),
        "distinct_work_ids": int(df["_work_id_norm"].dropna().nunique()),
        "distinct_standard_work_ids": int(df["_standard_work_id"].dropna().nunique()),
        "work_ids_with_multiple_names": len(work_name_conflict_ids),
        "work_ids_with_multiple_auth": len(work_auth_conflict_ids),
        "exact_duplicate_affected_rows": int(exact_dup_mask.sum()),
        "exact_duplicate_excess_rows": exact_excess_rows,
        "exact_duplicate_excess_amount_candidate": decimal_str(exact_excess_amount),
        "normalized_duplicate_affected_rows": int(normalized_dup_mask.sum()),
        "normalized_duplicate_excess_rows": normalized_excess_rows,
        "offset_candidate_groups": len(offset_groups),
        "exact_zero_offset_groups": sum(item["exact_zero"] for item in offset_groups),
        "cross_month_opposite_amount_groups": cross_month_offset_groups,
        "same_name_multiple_id_groups": int(len(same_name_multi_ids)),
        "volume_marker_rows": int(len(volume_marker_rows)),
        "launch_work_count": int(len(launch_summary)),
        "launch_with_nonpositive_row_count": int(len(launch_nonpositive_rows)),
        "launch_nonpositive_total_count": int(len(launch_nonpositive_total)),
        "form_launch_count": int(len(form_launch_summary)),
        "form_launch_with_nonpositive_row_count": int(form_launch_summary["has_nonpositive"].sum()) if not form_launch_summary.empty else 0,
        "issue_affected_rows": issue_rows_count,
        "issue_affected_rate": issue_rows_count / row_count,
        "parse_seconds": parse_seconds,
        "analysis_seconds": analysis_seconds,
        "rss_mb": rss_mb,
        "source_unchanged": True,
    }
    (PUBLIC_ROOT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    # Charts: aggregated counts only; no raw names, IDs, or row amounts.
    use_chart_theme()
    chart_map = []
    monthly_chart = None
    if len(monthly_counts) >= 8:
        fig, ax = plt.subplots(figsize=(10, 4.8))
        plot_df = monthly_counts.rename("rows").reset_index().rename(columns={"_month": "month"})
        plot_df["month_dt"] = pd.to_datetime(plot_df["month"] + "-01")
        sns.lineplot(data=plot_df, x="month_dt", y="rows", ax=ax, color=TOKENS["blue"], linewidth=1.2)
        ax.set_xlabel("账单月份")
        ax.set_ylabel("记录数")
        ax.yaxis.set_major_formatter(mticker.StrMethodFormatter("{x:,.0f}"))
        add_chart_header(fig, ax, "各月账单记录规模", f"{min_month} 至 {max_month}；最新月份完整性尚未确认")
        monthly_chart = save_chart(fig, "monthly-row-count.png")
        chart_map.append({"section": "数据范围", "chart": monthly_chart, "family": "Trend line", "claim": "月度记录量分布和覆盖变化"})

    fig, ax = plt.subplots(figsize=(8.5, 4.6))
    sign_df = sign_counts.reset_index()
    sign_df.columns = ["sign", "rows"]
    colors = [TOKENS["blue"], TOKENS["gold"], TOKENS["orange"], "#C5CAD3"]
    bars = ax.bar(sign_df["sign"], sign_df["rows"], color=colors, edgecolor=TOKENS["ink"], linewidth=0.8)
    ax.bar_label(bars, labels=[f"{value:,.0f}" for value in sign_df["rows"]], padding=3, fontsize=8)
    ax.set_xlabel("")
    ax.set_ylabel("记录数")
    add_chart_header(fig, ax, "实销金额符号分布", f"全量 {row_count:,} 行；金额只做符号与精度统计，不展示行级明细")
    sign_chart = save_chart(fig, "amount-sign-count.png")
    chart_map.append({"section": "金额质量", "chart": sign_chart, "family": "Bar", "claim": "正数、零值、负数和无效金额记录规模"})

    issue_chart_df = pd.DataFrame(issue_counts).sort_values("affected_rows", ascending=True)
    fig_height = max(5.0, 0.42 * len(issue_chart_df) + 2.0)
    fig, ax = plt.subplots(figsize=(10, fig_height))
    bars = ax.barh(issue_chart_df["issue_type"], issue_chart_df["affected_rows"], color=TOKENS["orange"], edgecolor=TOKENS["orange_dark"], linewidth=0.8)
    ax.bar_label(bars, labels=[f"{value:,.0f}" for value in issue_chart_df["affected_rows"]], padding=3, fontsize=8)
    ax.set_xlabel("受影响记录数（问题类型可重叠）")
    ax.set_ylabel("")
    add_chart_header(fig, ax, "数据问题候选规模", "问题类型可重叠；所有规则均未自动启用")
    issue_chart = save_chart(fig, "issue-candidate-count.png")
    chart_map.append({"section": "阻断问题", "chart": issue_chart, "family": "Ranked horizontal bar", "claim": "各类待确认问题的受影响行数"})

    work_format_df = work_format_counts.rename("rows").reset_index().rename(columns={"_work_id_format": "format"}).sort_values("rows")
    fig, ax = plt.subplots(figsize=(10, max(4.5, 0.38 * len(work_format_df) + 2)))
    bars = ax.barh(work_format_df["format"], work_format_df["rows"], color=TOKENS["olive"], edgecolor=TOKENS["olive_dark"], linewidth=0.8)
    ax.bar_label(bars, labels=[f"{value:,.0f}" for value in work_format_df["rows"]], padding=3, fontsize=8)
    ax.set_xlabel("记录数")
    ax.set_ylabel("")
    add_chart_header(fig, ax, "原始作品 ID 存储与格式", "按原始单元格类型和当前可识别格式分类")
    work_id_chart = save_chart(fig, "work-id-format.png")
    chart_map.append({"section": "作品 ID", "chart": work_id_chart, "family": "Ranked horizontal bar", "claim": "当前账单中作品 ID 的物理存储与格式分布"})

    (PUBLIC_ROOT / "chart-map.json").write_text(json.dumps(chart_map, ensure_ascii=False, indent=2), encoding="utf-8")

    # Public topic reports.
    structure_table = md_table(
        ["文件编号", "工作表编号", "最大行", "最大列", "表头行", "七字段精确匹配"],
        [[item["file_id"], item["sheet_id"], item["max_row"], item["max_column"], item["header_row"], "是" if item["header_match"] else "否"] for item in structures],
    )
    write_markdown_report(
        "01-file-workbook-structure.md",
        "文件及工作表结构报告",
        [f"共分析 {len(paths)} 个工作簿、{len(structures)} 个工作表、{row_count:,} 条数据行。", "分析前后文件 SHA-256、大小和修改时间一致，原始账单未发生变化。"],
        ["当前只观察到一种工作簿结构；后续新增结构必须重新进行表头和解析兼容性验证。"],
        ["一个文件不足以确认长期是否存在多模板、多工作表或加密/损坏文件。"],
        ["无原始行公开样本；文件结构只使用 F/S 编号。"],
        REQ_AT["file"],
        "文件类型、最大文件规模和多模板兼容范围仍为 PENDING-DATA。",
        "## 聚合结构\n\n" + structure_table,
    )

    field_table = md_table(
        ["字段", "缺失数", "缺失率", "去重值数", "最大文本长度", "原始类型"],
        [[item["field"], item["missing"], f"{item['missing_rate']:.2%}", item["distinct"], item["max_length"], ", ".join(f"{k}:{v}" for k, v in sorted(item["types"].items()))] for item in field_profile],
    )
    write_markdown_report(
        "02-seven-fields-domain-missingness.md",
        "七字段值域和缺失情况",
        ["检测到的表头与 PRD 七字段一致。", f"共有 {issue_rows_count:,} 行命中至少一种质量问题候选，问题类型可以重叠。"],
        ["稳定的非空、类型和长度约束应以全量字段分布为依据形成，不自动启用。"],
        ["字段业务值的合法枚举不能只靠当前账单频次确定。"],
        ["详细异常行位于本地 `invalid-required-fields.csv`。"],
        f"{REQ_AT['file']}；{REQ_AT['quality']}",
        "异常值域处理仍为 PENDING-DATA；七字段名称及唯一金额口径已确认。",
        "## 字段剖析\n\n" + field_table,
    )

    month_table = md_table(["原始格式分类", "记录数"], [[key, int(value)] for key, value in month_format_counts.items()])
    write_markdown_report(
        "03-year-month-format.md",
        "年月格式报告",
        [f"可解析月份范围为 {min_month} 至 {max_month}，共有 {len(months)} 个不同月份。", f"无法解析 {month_invalid_count:,} 行；日期不在当月第一日 {month_non_first_day:,} 行；连续范围内缺少 {len(missing_calendar_months)} 个日历月。", f"最新月份 {max_month} 有 {latest_month_rows:,} 行，前一月份有 {prior_month_rows:,} 行，之前 12 个月月度行数中位数为 {prior_12_month_median_rows:,.1f}；只陈述覆盖差异，不自行判定结算是否完整。"],
        ["若当前格式在其他真实文件中保持一致，可作为首个解析器适配类型。"],
        ["仅一个文件不能证明未来账单不会出现文本年月、数字年月或其他日期格式。", "最大月份是否为完整结算月份不能只靠出现记录判断。"],
        ["无法解析年月的详细行位于本地 `invalid-required-fields.csv`；最新月份完整性需要运营确认。"],
        f"{REQ_AT['file']}；{REQ_AT['quality']}",
        "最终允许格式、非法年月处理及最新月份完整性仍为 PENDING-DATA。",
        "## 格式分布\n\n" + month_table,
    )

    amount_table = md_table(["金额解析分类", "记录数"], [[key, int(value)] for key, value in amount_format_counts.items()])
    scale_table_public = md_table(["原始数值小数位", "记录数"], [[int(key), int(value)] for key, value in amount_scale_counts.items()])
    storage_table_public = md_table(["XLSX 原始存储类型", "记录数"], [[key, int(value)] for key, value in amount_storage_counts.items()])
    rounding_table_public = md_table(
        ["口径（仅敏感性检查）", "合计", "相对原始存储值差异"],
        [
            ["原始 XLSX 数值文本", format_decimal(total_amount), "0"],
            ["按每格 Excel 显示格式舍入", format_decimal(display_rounded_total), format_decimal(display_rounded_total - total_amount)],
            ["逐行保留 2 位小数", format_decimal(rounded_2_total), format_decimal(rounded_2_total - total_amount)],
            ["逐行保留 3 位小数", format_decimal(rounded_3_total), format_decimal(rounded_3_total - total_amount)],
        ],
    )
    write_markdown_report(
        "04-amount-precision-zero-negative.md",
        "金额精度、零值和负值报告",
        [f"基于 XLSX 内部 XML 原始数值文本计算的总实销为 {format_decimal(total_amount)}；正数 {positive_count:,} 行、零值 {zero_count:,} 行、负数 {negative_count:,} 行、无法精确解析 {amount_invalid_count:,} 行。", f"当前观察到的原始数值最大小数位为 {max_scale}，最大整数位数为 {max_integer_digits}；{hidden_precision_count:,} 行的原始小数位超过 Excel 显示格式。", f"金额列公式单元格为 {amount_formula_count:,}；科学计数法文本按 Decimal 原样解析，不做浮点转换。"],
        ["物理金额字段必须使用精确十进制；需先确认隐藏精度是业务有效精度还是历史计算尾差，才能冻结具体精度。"],
        ["负数和零值的具体业务原因不能单靠金额符号确认；不得自动删除。", "Excel 显示精度不能证明应当对原始存储值执行舍入。"],
        ["负数、零值代表样本位于本地 `representative-acceptance-rows.csv`；隐藏精度和科学计数法样本位于 `amount-precision-review.csv`。"],
        f"{REQ_AT['file']}；{REQ_AT['quality']}",
        "金额精度、隐藏尾差处理及零/负值业务解释仍为 PENDING-DATA。",
        "## 解析分布\n\n" + amount_table + "\n\n## XLSX 原始存储类型\n\n" + storage_table_public + "\n\n## 原始小数位分布\n\n" + scale_table_public + "\n\n## 舍入敏感性检查\n\n" + rounding_table_public,
    )

    write_markdown_report(
        "05-channel-id-name-relations.md",
        "渠道 ID 与名称关系报告",
        [f"共有 {summary['distinct_channel_ids']:,} 个规范化渠道 ID、{summary['distinct_channel_names']:,} 个规范化渠道名称。", f"{len(channel_multi_name_ids):,} 个渠道 ID 对应多个名称；{len(channel_multi_id_names):,} 个渠道名称对应多个 ID。"],
        ["可以将多名称或多 ID 关系作为渠道别名候选，但在运营确认前不能自动映射。"],
        ["账单不能区分主体更名、业务线拆分、录入差异或真实多渠道关系。"],
        [f"渠道关系样本使用匿名编号；详细 {private_manifest['channel-review.csv']:,} 行位于本地 `channel-review.csv`。"],
        REQ_AT["channel"],
        "渠道别名唯一键、标准化规则和有效期仍为 PENDING-DATA。",
    )

    work_format_table = md_table(["格式分类", "记录数"], [[key, int(value)] for key, value in work_format_counts.items()])
    write_markdown_report(
        "06-work-id-format.md",
        "作品 ID 格式报告",
        [f"识别到 {summary['distinct_work_ids']:,} 个原始作品 ID、{summary['distinct_standard_work_ids']:,} 个标准作品数字主体。", f"无法按当前纯数字或 Y+数字规则识别的记录为 {int(df['_work_id_norm'].isna().sum()):,} 行。"],
        ["作品 ID 物理存储应采用文本语义，避免前导零或长数字精度损失；规范化细节不自动启用。"],
        ["当前文件是否覆盖两种业务形态、异常前缀是否具有业务含义，需要运营确认。"],
        ["异常作品 ID 详细行位于本地 `invalid-required-fields.csv`。"],
        f"{REQ_AT['work']}；{REQ_AT['quality']}",
        "前导零、大小写、科学计数法和异常形式处理仍为 PENDING-DATA。",
        "## 格式分布\n\n" + work_format_table,
    )

    auth_table = md_table(["授权分类匿名编号", "记录数"], [[sample_code("AUTH", key), int(value)] for key, value in auth_values.items()])
    write_markdown_report(
        "07-name-authorization-conflicts.md",
        "名称与授权分类冲突报告",
        [f"{len(work_name_conflict_ids):,} 个作品 ID 对应多个规范化作品名称；{len(work_auth_conflict_ids):,} 个作品 ID 对应多个授权分类。", f"授权分类共有 {len(auth_values):,} 个原始值。"],
        ["多名称或多授权分类关系可作为阻断候选，需区分更名、错误和真实业务变化。"],
        ["标准作品名称选择、授权分类与业务形态的关系不能仅按频次决定。"],
        [f"详细 {private_manifest['work-name-auth-review.csv']:,} 行位于本地 `work-name-auth-review.csv`。"],
        f"{REQ_AT['work']}；{REQ_AT['quality']}",
        "名称规范化和冲突处理规则仍为 PENDING-DATA。",
        "## 授权分类聚合（匿名）\n\n" + auth_table,
    )

    write_markdown_report(
        "08-duplicate-offset-candidates.md",
        "重复与合法冲抵候选报告",
        [f"完全重复候选影响 {int(exact_dup_mask.sum()):,} 行，候选多余行 {exact_excess_rows:,} 行；若删除多余行，候选金额影响为 {format_decimal(exact_excess_amount)}，该金额仅用于风险量化。", f"规范化后重复候选影响 {int(normalized_dup_mask.sum()):,} 行；同月同渠道同作品同时含正负金额的组为 {len(offset_groups):,} 个，其中组内净额为零 {sum(item['exact_zero'] for item in offset_groups):,} 个；跨月出现同绝对值正负金额的候选组为 {cross_month_offset_groups:,} 个。"],
        ["完全相同记录可进入重复候选清单；正负抵消组只作为合法冲抵候选，均不得自动处理。"],
        ["账单本身不能证明一组相同行是重复还是多笔合法交易，也不能证明正负配对的业务原因。"],
        [f"重复候选详细 {private_manifest['duplicate-candidates.csv']:,} 行位于本地 `duplicate-candidates.csv`；冲抵候选 {private_manifest['offset-candidates.csv']:,} 行位于 `offset-candidates.csv`。"],
        f"{REQ_AT['quality']}；{REQ_AT['file']}",
        "完整重复模式、自动删除范围和冲抵识别仍为 PENDING-DATA。",
    )

    write_markdown_report(
        "09-historical-volume-candidates.md",
        "历史分册候选报告",
        [f"同一规范化作品名称对应多个作品 ID 的候选组为 {len(same_name_multi_ids):,} 个。", f"名称包含分册/卷/部/季等启发式标记的记录为 {len(volume_marker_rows):,} 行。"],
        ["名称完全一致或含分册标记可以用于生成候选，不足以自动归并。"],
        ["当前七字段缺少作者、完整分类和版权期限，无法验证分册归并的全部必需证据。"],
        [f"分册标记详细 {private_manifest['volume-candidates.csv']:,} 行位于本地 `volume-candidates.csv`。"],
        f"REQ-WORK-005~006、REQ-DQ-001；AT-M1-024~025、AT-M1-010",
        "候选发现特征和主 ID 推荐顺序仍为 PENDING-DATA。",
    )

    launch_table = md_table(
        ["指标", "标准作品数"],
        [
            ["可计算首次月份", len(launch_summary)],
            ["首次月份含零值或负值记录", len(launch_nonpositive_rows)],
            ["首次月份合计小于或等于零", len(launch_nonpositive_total)],
        ],
    )
    write_markdown_report(
        "10-rb-work-launch-001.md",
        "RB-WORK-LAUNCH-001 首次实销金额符号验证",
        [f"可计算 {len(launch_summary):,} 个标准作品的最早账单月份。", f"其中 {len(launch_nonpositive_rows):,} 个作品的最早月份包含零值或负值记录，{len(launch_nonpositive_total):,} 个作品的最早月份合计不大于零。"],
        ["当前分析只提供样本分布，不建议自动排除零值或负值。"],
        ["第一笔入账、第一笔正收入和上线业务事实是否等价，不能单靠账单确认。"],
        [f"详细 {private_manifest['launch-sign-review.csv']:,} 行位于本地 `launch-sign-review.csv`。"],
        "REQ-WORK-003~004；AT-M1-022~023",
        "仍为 PENDING-DATA，必须由运营确认后才能决定首次实销计算规则。",
        "## 聚合结果\n\n" + launch_table,
    )

    duplicate_binary_groups = defaultdict(list)
    duplicate_business_groups = defaultdict(list)
    for item in file_public:
        duplicate_binary_groups[item["sha256"]].append(item["file_id"])
        duplicate_business_groups[content_hashes[item["file_id"]]["order_insensitive"]].append(item["file_id"])
    binary_dup_sets = [group for group in duplicate_binary_groups.values() if len(group) > 1]
    business_dup_sets = [group for group in duplicate_business_groups.values() if len(group) > 1]
    write_markdown_report(
        "11-fingerprint-content-duplicates.md",
        "文件指纹和业务内容重复报告",
        [f"文件 SHA-256 重复组 {len(binary_dup_sets)} 个；忽略行顺序后的业务内容重复组 {len(business_dup_sets)} 个。", "文件名未作为重复判断依据。"],
        ["二进制指纹用于普通上传阻断；业务内容指纹只作为风险提示，不能替代正式规则。"],
        ["当前只有一个文件，无法验证跨文件重复、格式变化但内容相同等情况。"],
        ["无须公开原始文件名或内容。"],
        "REQ-DATA-IMPORT-002、REQ-DATA-IMPORT-007；AT-M1-002、AT-M1-007",
        "跨文件内容重复规则仍为 PENDING-DATA。",
    )

    scale_table = md_table(
        ["指标", "结果"],
        [
            ["文件数", len(paths)],
            ["输入字节", f"{sum(item.size for item in before):,}"],
            ["数据行", f"{row_count:,}"],
            ["只读解析秒", f"{parse_seconds:.3f}"],
            ["分析秒", f"{analysis_seconds:.3f}"],
            ["解析吞吐（行/秒）", f"{row_count / parse_seconds:,.0f}" if parse_seconds else "-"],
            ["分析进程 RSS（MB）", f"{rss_mb:,.1f}" if rss_mb is not None else "当前环境未提供"],
        ],
    )
    write_markdown_report(
        "12-scale-performance-baseline.md",
        "数据规模及性能基线",
        [f"当前全量输入为 {len(paths)} 个文件、{row_count:,} 行、{sum(item.size for item in before):,} 字节。", f"本次只读解析耗时 {parse_seconds:.3f} 秒，分析耗时 {analysis_seconds:.3f} 秒；这是本机单次基线，不是生产 SLA。"],
        ["该结果可作为后续分块和容量测试起点，不应直接成为正式性能阈值。"],
        ["只有一个文件和一台机器，无法确定并发、峰值、恢复和长期增长性能。"],
        ["无需运营确认原始行；需要技术团队在目标部署环境复测。"],
        f"{REQ_AT['file']}；{REQ_AT['platform']}",
        "文件上限、分块大小、并发和正式性能指标仍为 PENDING-DATA。",
        "## 本次基线\n\n" + scale_table,
    )

    missing_master_fields = ["作者", "一级分类", "二级分类", "三级分类", "版权开始日期", "版权到期日期", "辅助标签", "特殊属性标签"]
    gap_table = md_table(["M1 所需信息", "七字段账单是否提供", "结论"], [[field, "否", "需要补全表或其他权威来源"] for field in missing_master_fields])
    write_markdown_report(
        "13-master-data-source-gaps.md",
        "基础信息来源缺口报告",
        ["当前七字段账单不包含作者、完整三级分类、版权期限或标签。", "作品名称存在于账单，但不能自动认定为最终标准作品名称。"],
        ["账单可用于导出待补全标准作品集合，不能自动生成缺失基础信息。"],
        ["作者消歧、分类路径、版权期限及两种业务形态期限反例均需要额外权威数据或运营确认。"],
        ["本报告不公开作品名称；补全源需要另行提供。"],
        "REQ-WORK-008~011、REQ-CLASS-001~002；AT-M1-027~029、AT-M1-031、AT-M1-040~041",
        "最终分类树、标签库、作者消歧和版权期限反例验证仍为 PENDING-DATA。",
        "## 缺口矩阵\n\n" + gap_table,
    )

    inactivity_counts = status_df["months_since_last_record"].value_counts().sort_index() if not status_df.empty else pd.Series(dtype=int)
    status_table = md_table(
        ["描述性指标", "结果"],
        [
            ["标准作品数", len(status_df)],
            ["最后一次记录距截止月中位数", f"{status_df['months_since_last_record'].median():.0f} 个月" if not status_df.empty else "-"],
            ["最后一次记录距截止月最大值", f"{status_df['months_since_last_record'].max():.0f} 个月" if not status_df.empty else "-"],
            ["出现后中断并再次出现", int(status_df["reappeared_after_gap"].sum()) if not status_df.empty else 0],
            ["最后出现月份净额为零", int((status_df["last_month_sign"] == "zero").sum()) if not status_df.empty else 0],
            ["最后出现月份净额为负", int((status_df["last_month_sign"] == "negative").sum()) if not status_df.empty else 0],
        ],
    )
    write_markdown_report(
        "14-product-status-distribution.md",
        "产品状态数据分布",
        [f"以账单最后出现月份描述 {len(status_df):,} 个标准作品；该指标不等同于已下架。", f"有 {int(status_df['reappeared_after_gap'].sum()) if not status_df.empty else 0:,} 个作品曾在缺月后再次出现记录。"],
        ["最后出现月份、缺月长度和恢复收入可以作为疑似状态的分析特征，但不自动产生产品状态。"],
        ["账单缺行不一定代表零收入；已下架和重新上架必须由运营确认。"],
        [f"详细 {private_manifest['product-status-review.csv']:,} 行位于本地 `product-status-review.csv`。"],
        "REQ-WORK-007；AT-M1-026",
        "疑似下架观察周期、基本无收入和归零条件仍为 PENDING-DATA。",
        "## 描述性分布\n\n" + status_table,
    )

    issue_table = md_table(
        ["问题候选", "受影响行", "占全部行"],
        [[item["issue_type"], f"{item['affected_rows']:,}", f"{item['rate']:.2%}"] for item in sorted(issue_counts, key=lambda item: item["affected_rows"], reverse=True)],
    )
    write_markdown_report(
        "15-issue-volume-return-candidates.md",
        "数据问题数量及退回规则候选",
        [f"至少命中一个问题候选的记录为 {issue_rows_count:,} 行，占 {issue_rows_count / row_count:.2%}。", "问题类型允许重叠，不能将各类型行数直接相加作为问题总行数。"],
        ["退回判断可考虑受影响行比例、问题类型数、金额影响和是否存在系统性模式，但本次不设置阈值。"],
        ["只有一个文件，无法比较逐项处理与整份退回在不同文件上的运营成本。"],
        ["各类详细样本分别保存在本地私有目录。"],
        f"{REQ_AT['quality']}；{REQ_AT['file']}",
        "错误过多时整份退回的判定规则仍为 PENDING-DATA。",
        "## 问题候选分布\n\n" + issue_table,
    )

    acceptance_rows = [
        ["AT-M1-001", "零值、负值、七字段格式", "可用", "representative-acceptance-rows.csv"],
        ["AT-M1-002", "文件指纹", "单文件，仅可验证指纹生成", "无原始内容公开"],
        ["AT-M1-003~004", "大文件与严格对账", "可用于脚本级基线；系统原子性需实现后验证", "summary.json"],
        ["AT-M1-010~012", "缺失、冲突、重复和本地修正", "问题样本可用；本地修正流程需实现后验证", "多个私有 CSV"],
        ["AT-M1-020~025", "作品 ID、形态、上线时间、分册", "部分可用；作者/版权资料缺失", "launch/volume 私有 CSV"],
        ["AT-M1-026", "产品状态", "仅描述性候选，不可确认状态", "product-status-review.csv"],
        ["AT-M1-027~031", "基础信息、作者、版权期限", "账单字段不足", "需额外权威来源"],
        ["AT-M1-030", "渠道更名", "关系候选可用，需运营确认", "channel-review.csv"],
        ["AT-M1-040~041", "分类和标签", "账单字段不足", "需补全数据"],
        ["AT-M1-050~052", "任务、恢复、AI", "不能由静态账单验证", "实现后验收"],
    ]
    write_markdown_report(
        "16-representative-acceptance-dataset-catalog.md",
        "代表验收数据集清单",
        [f"本地私有目录生成 {len(representative_rows):,} 条代表行，按问题类别抽样。", "Git 中只保存样本类别、可用性和私有文件名，不保存原始行。"],
        ["代表样本可作为后续验收输入候选，但不能替代完整或接近完整真实数据验收。"],
        ["系统行为类 AT 必须在实现后验证，静态账单只能提供输入样本。"],
        ["详细代表行位于本地 `representative-acceptance-rows.csv`。"],
        "全部 M1 REQ/AT；具体映射见追踪矩阵。",
        "代表案例最终选择和运营确认仍为 PENDING-DATA。",
        "## 验收样本覆盖\n\n" + md_table(["验收编号", "样本主题", "当前可用性", "证据位置"], acceptance_rows),
    )

    blocking_findings = [item for item in issue_counts if item["affected_rows"] > 0]
    blocking_findings.append(
        {
            "issue_type": f"最新月份 {max_month} 完整性待确认（{latest_month_rows:,} 行；前月 {prior_month_rows:,} 行）",
            "affected_rows": latest_month_rows,
            "rate": latest_month_rows / row_count,
        }
    )
    directly_freezable = [
        "当前账单表头为七个已冻结业务字段，实销金额是唯一金额字段。",
        "正式收入事实必须保留原始渠道 ID、原始渠道名称、授权分类、原始作品 ID 和作品名称，并使用文本语义保存 ID。",
        "金额必须使用精确十进制语义；当前观察精度和范围已量化，但最终数据库精度仍需留出边界并评审。",
        "原始事实与渠道/作品映射投影必须分离，所有候选映射均不能自动启用。",
    ]
    operator_questions = [
        "确认当前不同渠道 ID/名称组合中是否存在同一主体的历史更名或别名。",
        "确认同一作品 ID 多名称、多授权分类的业务解释。",
        "确认同名多 ID 和分册标记候选是否属于历史分册归并。",
        "确认零金额或负金额能否决定首次实销月份。",
        "确认金额隐藏精度是有效业务精度还是历史计算尾差。",
        f"确认最新月份 {max_month} 是否为完整结算月份。",
        "提供作者、完整分类、版权期限和标签的权威补全来源。",
    ]
    if exact_excess_rows:
        operator_questions.insert(2, "确认完全重复候选是否确为重复。")
    if cross_month_offset_groups or offset_groups:
        operator_questions.insert(2, "确认正负金额候选是否为合法冲抵。")
    physical_ready = False
    readiness_reason = "未满足：金额隐藏精度和最新月份完整性尚未确认，仍存在需运营确认的数据关系，且作者、完整分类、版权期限等物理模型关键来源缺失；首次实销金额符号规则仍为 PENDING-DATA。"

    summary_md = f"""# M1 真实账单分析汇总

## 技术结论

- 已只读分析 **{len(paths)}** 个文件、**{len(structures)}** 个工作表、**{row_count:,}** 条数据行。
- 月份范围为 **{min_month} 至 {max_month}**，总实销为 **{format_decimal(total_amount)}**。
- 原始文件前后哈希、大小和修改时间一致；详细敏感样本仅保存在被 Git 忽略的本地目录。
- 当前**尚未满足进入物理数据库设计的条件**。{readiness_reason}

## 阻断性问题

{chr(10).join(f'- {item["issue_type"]}：{item["affected_rows"]:,} 行（{item["rate"]:.2%}）' for item in blocking_findings) if blocking_findings else '- 未发现阻断性问题。'}

## 可以直接冻结的字段语义

{chr(10).join(f'- {item}' for item in directly_freezable)}

## 仍需运营确认

{chr(10).join(f'- {item}' for item in operator_questions)}

## 报告索引

专题报告位于本目录 `01` 至 `16`；聚合结果见 `summary.json`，图表见 `assets/`，方法与复现信息见 `source-notes.md`。详细原始样本位于本地 `data/m1-real-bills-private/`，不得提交 Git。
"""
    (PUBLIC_ROOT / "README.md").write_text(summary_md, encoding="utf-8")

    source_notes = {
        "source_inventory": [
            {"file_id": item["file_id"], "size": item["size"], "sha256": item["sha256"]} for item in file_public
        ],
        "script": "tools/m1-data-analysis/analyze_real_bills.py",
        "input_policy": "read-only; source size, mtime and SHA-256 verified before and after",
        "public_output_policy": "aggregates and anonymized sample codes only",
        "private_output_policy": "raw review samples under Git-ignored data/m1-real-bills-private",
        "definitions": {
            "row_count": "non-empty rows after the exact seven-column header",
            "total_amount": "sum of exact 实销金额 tokens read from XLSX XML, including zero and negative values; no display rounding",
            "duplicate_candidate": "same file/sheet and identical seven raw cell values; not confirmed duplicate",
            "launch_candidate": "earliest parseable month per standard-work body; zero/negative eligibility remains pending",
        },
        "chart_map": chart_map,
        "omissions": [
            "No raw names, channel names, work IDs or row-level amounts in Git-trackable reports.",
            "No status, mapping, cleaning, merge or deduplication rule was enabled.",
            "No production SLA inferred from a single-machine run.",
        ],
    }
    (PUBLIC_ROOT / "source-notes.md").write_text(
        "# 分析来源与复现说明\n\n```json\n" + json.dumps(source_notes, ensure_ascii=False, indent=2) + "\n```\n",
        encoding="utf-8",
    )

    # Technical HTML report: single durable report surface. Supporting markdown files remain audit evidence.
    chart_html = ""
    if monthly_chart:
        chart_html += f'<p><strong>月度记录量展示了数据覆盖的时间形状。</strong>最新月份只有 {latest_month_rows:,} 行，明显低于前月 {prior_month_rows:,} 行；这只证明覆盖差异，是否完整必须由运营确认。</p><figure><img src="{monthly_chart}" alt="各月账单记录规模"><figcaption>{min_month} 至 {max_month}；最新月份完整性尚未确认。</figcaption></figure>'
    chart_html += f'<p><strong>零值和负值在真实账单中客观存在。</strong>它们必须保留；金额符号的业务解释和首次实销资格仍需确认。</p><figure><img src="{sign_chart}" alt="实销金额符号分布"><figcaption>全量 {row_count:,} 行，按金额符号分类。</figcaption></figure>'
    chart_html += f'<p><strong>问题候选集中在映射关系和重复识别，而不是自动可修复错误。</strong>各类型可重叠，不能直接相加。</p><figure><img src="{issue_chart}" alt="数据问题候选规模"><figcaption>所有候选规则均未启用。</figcaption></figure>'

    html_report = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>M1 真实账单数据质量分析</title>
  <style>
    body {{ font-family: "Segoe UI", "Microsoft YaHei", sans-serif; margin:0; background:#f8fafc; color:#0f172a; }}
    main {{ max-width:980px; margin:0 auto; padding:40px 22px 64px; }}
    header, section {{ margin-bottom:34px; }} h1,h2 {{ line-height:1.2; margin:0 0 14px; }}
    p,li {{ line-height:1.72; }} .summary {{ background:#fff; border:1px solid #e2e8f0; border-radius:14px; padding:20px 24px; }}
    .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:12px; margin:18px 0; }}
    .kpi {{ background:#fff; border:1px solid #e2e8f0; border-radius:12px; padding:14px; }}
    .kpi strong {{ display:block; font-size:22px; margin-top:6px; }}
    figure {{ margin:22px 0 30px; }} img {{ width:100%; height:auto; background:#fff; border-radius:10px; }}
    figcaption {{ color:#475569; font-size:14px; margin-top:8px; }}
    table {{ width:100%; border-collapse:collapse; background:#fff; }} th,td {{ border-bottom:1px solid #e2e8f0; padding:10px; text-align:left; vertical-align:top; }}
    .blocked {{ border-left:5px solid #F0986E; padding-left:16px; }} code {{ background:#e2e8f0; padding:2px 5px; border-radius:5px; }}
  </style>
</head>
<body><main data-report-audience="technical">
  <header data-contract-section="title"><h1>M1 真实账单数据质量分析</h1></header>
  <section data-contract-section="technical-summary" class="summary">
    <h2>技术摘要</h2>
    <p>已对当前完整账单目录执行只读全量分析。源文件未改变，七字段表头一致，但数据中存在需要运营确认的渠道关系、作品名称/授权分类关系、重复候选、分册候选以及首次实销金额符号问题。</p>
    <div class="kpis"><div class="kpi">文件<strong>{len(paths)}</strong></div><div class="kpi">数据行<strong>{row_count:,}</strong></div><div class="kpi">月份<strong>{min_month}—{max_month}</strong></div><div class="kpi">总实销<strong>{html.escape(format_decimal(total_amount))}</strong></div></div>
    <p class="blocked"><strong>物理数据库设计尚未获准。</strong>{html.escape(readiness_reason)}</p>
  </section>
  <section data-contract-section="key-findings"><h2>真实数据确认了输入规模，也暴露了必须先处理的关系歧义</h2>{chart_html}</section>
  <section data-contract-section="scope-data-and-metric-definitions">
    <h2>范围、数据与指标定义</h2>
    <p>范围为 <code>data/real-bills/</code> 下全部 XLSX/XLSM 文件。数据行为七字段表头之后的非空行；总实销为所有可精确解析实销金额之和，包含零值和负值；重复只表示候选，不代表已确认删除。</p>
  </section>
  <section data-contract-section="methodology">
    <h2>方法保证原始账单只读且结论可复现</h2>
    <p>脚本使用只读工作簿模式逐行解析，运行前后校验 SHA-256、大小和修改时间。公开报告只写聚合统计和匿名样本编号；原始确认样本写入被 Git 忽略的本地目录。所有候选规则保持禁用。</p>
  </section>
  <section data-contract-section="limitations-uncertainty-and-robustness-checks">
    <h2>单一账单文件无法完成主数据和阈值冻结</h2>
    <ul><li>七字段不含作者、完整分类、版权期限和标签。</li><li>一个文件无法覆盖跨文件重复和模板漂移。</li><li>渠道多重关系、重复、冲抵、分册和产品状态均需要运营确认。</li><li>单机耗时是容量基线，不是生产 SLA。</li></ul>
  </section>
  <section data-contract-section="recommended-next-steps">
    <h2>先完成运营确认和补充数据，再冻结物理模型</h2>
    <ol><li>按本地私有样本完成渠道、作品名称/授权分类、重复和分册确认。</li><li>确认 RB-WORK-LAUNCH-001 的零/负金额规则。</li><li>提供作者、分类、版权期限和标签的权威补充来源。</li><li>确认后更新 REQ/AT 或 ADR，再决定物理字段、索引和分区。</li></ol>
  </section>
  <section data-contract-section="further-questions"><h2>仍需回答的业务问题</h2><ul>{''.join(f'<li>{html.escape(item)}</li>' for item in operator_questions)}</ul></section>
</main></body></html>"""
    (PUBLIC_ROOT / "report.html").write_text(html_report, encoding="utf-8")

    assert_sources_unchanged(before)
    elapsed = time.perf_counter() - started
    summary["total_elapsed_seconds"] = elapsed
    summary["source_unchanged"] = True
    (PUBLIC_ROOT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
