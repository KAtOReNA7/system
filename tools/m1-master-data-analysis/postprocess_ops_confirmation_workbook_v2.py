from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OPS_ROOT = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.xlsx"


FREEZE_PANES = {
    "确认进度总览": "A5",
    "正式导入阻断确认": "C2",
    "多ID归并候选": "D2",
    "标准作品基础信息补全": "B2",
    "台账真实冲突": "C2",
    "版权期限反例": "C2",
    "非阻断观察": "D2",
}

STATUS_HEADERS = {
    "运营确认结果",
    "是否解除阻断",
    "是否归并",
    "是否解除冲突",
    "补全状态",
    "是否标记为异常",
    "是否升级为阻断",
}


def is_amount_display_column(header: str) -> bool:
    if "完整精度" in header:
        return False
    return header in {"累计实销", "金额"} or header.endswith("金额")


def is_date_column(header: str) -> bool:
    if "候选" in header or "全部" in header:
        return False
    return "日期" in header or header in {"生成时间", "确认版权开始日期", "确认版权到期日期"}


def postprocess() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)

    wait_fill = PatternFill(fill_type="solid", fgColor="FFF4C2")
    blocked_fill = PatternFill(fill_type="solid", fgColor="FFEDDE")
    done_fill = PatternFill(fill_type="solid", fgColor="D8ECBD")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = FREEZE_PANES.get(worksheet.title, "B2")

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            if worksheet.title == "确认进度总览":
                worksheet.auto_filter.ref = f"A5:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            else:
                worksheet.auto_filter.ref = worksheet.dimensions

        header_row = 5 if worksheet.title == "确认进度总览" else 1
        headers = [worksheet.cell(header_row, col).value for col in range(1, worksheet.max_column + 1)]

        for row in worksheet.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment

        for col_idx, header_value in enumerate(headers, start=1):
            if header_value is None:
                continue
            header = str(header_value)
            col_letter = get_column_letter(col_idx)

            if "完整精度" in header:
                worksheet.column_dimensions[col_letter].hidden = True
                for cell in worksheet[col_letter]:
                    cell.number_format = "@"
                    cell.font = Font(color="666666")
                continue

            if is_amount_display_column(header):
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=header_row + 1, max_row=worksheet.max_row):
                    for item in cell:
                        item.number_format = '#,##0.00'

            if is_date_column(header):
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=header_row + 1, max_row=worksheet.max_row):
                    for item in cell:
                        item.number_format = "yyyy-mm-dd"

            if header in STATUS_HEADERS and worksheet.max_row > header_row:
                target_range = f"{col_letter}{header_row + 1}:{col_letter}{worksheet.max_row}"
                worksheet.conditional_formatting.add(
                    target_range,
                    FormulaRule(formula=[f'OR({col_letter}{header_row + 1}="",{col_letter}{header_row + 1}="待确认")'], fill=wait_fill),
                )
                worksheet.conditional_formatting.add(
                    target_range,
                    FormulaRule(formula=[f'OR({col_letter}{header_row + 1}="否",{col_letter}{header_row + 1}="需补充资料")'], fill=blocked_fill),
                )
                worksheet.conditional_formatting.add(
                    target_range,
                    FormulaRule(formula=[f'OR({col_letter}{header_row + 1}="是",{col_letter}{header_row + 1}="已补齐")'], fill=done_fill),
                )

        if worksheet.title == "确认进度总览":
            worksheet["B3"].number_format = "yyyy-mm-dd hh:mm:ss"

    workbook.save(WORKBOOK_PATH)
    workbook.close()


if __name__ == "__main__":
    postprocess()
