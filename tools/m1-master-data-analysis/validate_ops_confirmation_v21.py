from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
OPS_ROOT = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.1.xlsx"
OLD_WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.xlsx"
PUBLIC_VALIDATION_PATH = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2.1-validation.json"
PUBLIC_SUMMARY_PATH = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2.1-summary.md"


EXPECTED_SHEETS = [
    "确认进度总览",
    "正式导入阻断确认",
    "多ID归并候选",
    "标准作品基础信息补全",
    "台账真实冲突",
    "版权期限反例",
    "非阻断观察",
]


def headers(ws: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> list[str]:
    return [str(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]


def header_index(ws: openpyxl.worksheet.worksheet.Worksheet, name: str, row: int = 1) -> int:
    values = headers(ws, row)
    return values.index(name) + 1


def count_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    return max(ws.max_row - 1, 0)


def public_sensitive_scan() -> int:
    public_root = ROOT / "docs" / "analysis" / "m1-master-data"
    text = ""
    for path in public_root.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".md", ".json", ".html"}:
            text += "\n" + path.read_text(encoding="utf-8", errors="ignore")
    sensitive_patterns = [
        r"版权期限候选:\n\d{4}-\d{2}-\d{2}",
        r"作者候选:\n.+\n.+\n.+",
    ]
    return sum(1 for pattern in sensitive_patterns if re.search(pattern, text))


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
    workbook_open = True
    expected_sheets_ok = wb.sheetnames == EXPECTED_SHEETS

    overview = wb["确认进度总览"]
    non_block_formula_ok = (
        'COUNTIFS(\'非阻断观察\'!L2:L91,"是",\'非阻断观察\'!M2:M91,"<>已解除")' in str(overview["F11"].value)
        and '"否")' not in str(overview["F11"].value)
        and "COUNTBLANK" not in str(overview["F11"].value)
    )
    non_block_completion_ok = str(overview["I11"].value) == '=IF((C11+D11+E11+F11)=0,1,C11/(C11+D11+E11+F11))'

    nb = wb["非阻断观察"]
    nb_headers = headers(nb)
    release_status_present = "阻断解除状态" in nb_headers
    mark_col = header_index(nb, "是否标记为异常")
    escalate_col = header_index(nb, "是否升级为阻断")
    release_col = header_index(nb, "阻断解除状态")
    default_no_not_prefilled = all(
        nb.cell(row, mark_col).value in (None, "") and nb.cell(row, escalate_col).value in (None, "")
        for row in range(2, nb.max_row + 1)
    )
    no_false_unresolved_count = all(
        not (nb.cell(row, escalate_col).value in (None, "", "否") and nb.cell(row, release_col).value not in (None, ""))
        for row in range(2, nb.max_row + 1)
    )

    ledger = wb["台账真实冲突"]
    ledger_headers = headers(ledger)
    copyright_task_cols_ok = "关联版权任务ID" in ledger_headers and "版权填写提示" in ledger_headers
    std_col = header_index(ledger, "标准作品ID")
    field_col = header_index(ledger, "冲突字段")
    value_col = header_index(ledger, "各候选值")
    task_col = header_index(ledger, "关联版权任务ID")
    duplicate_ids = {"162214", "163360"}
    ledger_copyright_removed = True
    ledger_has_task_links = True
    for row in range(2, ledger.max_row + 1):
        field = str(ledger.cell(row, field_col).value or "")
        value = str(ledger.cell(row, value_col).value or "")
        if "版权期限" in field or "版权期限候选" in value:
            ledger_copyright_removed = False
        if str(ledger.cell(row, std_col).value or "") in duplicate_ids and not str(ledger.cell(row, task_col).value or "").startswith("CP-"):
            ledger_has_task_links = False

    copyright = wb["版权期限反例"]
    copyright_rows_ok = count_rows(copyright) == 2
    copyright_ids = {str(copyright.cell(row, 2).value) for row in range(2, copyright.max_row + 1)}
    copyright_ids_ok = duplicate_ids == copyright_ids

    row_counts = {
        "确认进度总览": 6,
        "正式导入阻断确认": count_rows(wb["正式导入阻断确认"]),
        "多ID归并候选": count_rows(wb["多ID归并候选"]),
        "标准作品基础信息补全": count_rows(wb["标准作品基础信息补全"]),
        "台账真实冲突": count_rows(ledger),
        "版权期限反例": count_rows(copyright),
        "非阻断观察": count_rows(nb),
    }
    old_workbook_kept = OLD_WORKBOOK_PATH.exists()
    public_sensitive_matches = public_sensitive_scan()
    public_privacy_ok = public_sensitive_matches == 0

    validation = {
        "workbook_open": workbook_open,
        "expected_sheets_ok": expected_sheets_ok,
        "row_counts": row_counts,
        "old_workbook_kept": old_workbook_kept,
        "non_blocking_unresolved_formula_ok": non_block_formula_ok,
        "non_blocking_completion_formula_ok": non_block_completion_ok,
        "non_blocking_release_status_present": release_status_present,
        "non_blocking_default_no_not_prefilled": default_no_not_prefilled,
        "non_blocking_false_values_not_unresolved": no_false_unresolved_count,
        "ledger_copyright_task_columns_ok": copyright_task_cols_ok,
        "ledger_copyright_fields_removed": ledger_copyright_removed,
        "ledger_duplicate_rows_link_to_copyright_tasks": ledger_has_task_links,
        "copyright_conflict_rows_ok": copyright_rows_ok,
        "copyright_conflict_ids_ok": copyright_ids_ok,
        "public_sensitive_matches": public_sensitive_matches,
        "public_privacy_ok": public_privacy_ok,
    }
    validation["overall"] = all(validation[key] for key in validation if key not in {"row_counts", "public_sensitive_matches"})

    PUBLIC_VALIDATION_PATH.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = f"""# M1 运营确认包 v2.1 聚合摘要

## 输出

- 本地私有工作簿：`data/m1-master-data-private/ops-confirmation/M1-运营确认包-v2.1.xlsx`
- 公开验证结果：`docs/analysis/m1-master-data/ops-confirmation-v2.1-validation.json`

## 工作表行数

| 工作表 | 行数 |
|---|---:|
{chr(10).join(f"| {name} | {count} |" for name, count in row_counts.items())}

## 本次小修复

- 非阻断观察的“未解除阻断”只统计 `是否升级为阻断=是` 且 `阻断解除状态<>已解除` 的记录；默认空值或默认否不计为未解除阻断。
- 非阻断观察默认不预填“否”，避免默认值被误认为人工确认结果。
- `162214` 和 `163360` 的版权日期只在“版权期限反例”填写；“台账真实冲突”仅保留非版权冲突，并展示关联版权任务 ID。
- 版权期限反例仍为 2 组，未重复计算为台账版权确认任务。

## 验证结论

- 验证状态：{'通过' if validation['overall'] else '未通过'}
- 旧 v2 工作簿保留：{'是' if old_workbook_kept else '否'}
- 公开报告敏感明细扫描命中数：{public_sensitive_matches}
"""
    PUBLIC_SUMMARY_PATH.write_text(summary, encoding="utf-8")
    wb.close()

    if not validation["overall"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
