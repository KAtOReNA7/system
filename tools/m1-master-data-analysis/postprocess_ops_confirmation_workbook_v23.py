from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = (
    ROOT
    / "data"
    / "m1-master-data-private"
    / "ops-confirmation"
    / "M1-运营确认包-v2.3.xlsx"
)

FREEZE_PANES = {
    "确认进度总览": "A13",
    "正式导入阻断确认": "B2",
    "多ID归并候选": "D2",
    "标准作品基础信息补全": "C2",
    "台账真实冲突": "C2",
    "版权期限反例": "C2",
    "非阻断观察": "D2",
}


def _header_column(ws, header: str) -> str:
    for cell in ws[1]:
        if cell.value == header:
            return get_column_letter(cell.column)
    raise KeyError(f"sheet {ws.title!r} missing header {header!r}")


def _range(ws, header: str) -> str:
    column = _header_column(ws, header)
    last_row = max(ws.max_row, 2)
    sheet = ws.title.replace("'", "''")
    return f"'{sheet}'!${column}$2:${column}${last_row}"


def _count_nonblank_excluding_need_more(range_ref: str) -> str:
    return (
        f'=COUNTIF({range_ref},"<>")'
        f'-COUNTIF({range_ref},"需补充资料")'
    )


def _set_overview_formulas(wb) -> None:
    ws = wb["确认进度总览"]

    normal_rows = [
        (13, "正式导入阻断确认", "运营确认结果", "是否解除阻断"),
        (14, "多ID归并候选", "是否归并", "是否解除阻断"),
        (15, "标准作品基础信息补全", "补全状态", None),
        (16, "台账真实冲突", "运营确认值", "是否解除冲突"),
        (17, "版权期限反例", "确认版权开始日期", "是否解除阻断"),
    ]

    for row, sheet_name, status_header, unblock_header in normal_rows:
        task_ws = wb[sheet_name]
        status_range = _range(task_ws, status_header)
        ws.cell(row, 2).value = max(task_ws.max_row - 1, 0)
        ws.cell(row, 3).value = _count_nonblank_excluding_need_more(status_range)
        ws.cell(row, 4).value = f'=COUNTBLANK({status_range})'
        ws.cell(row, 5).value = f'=COUNTIF({status_range},"需补充资料")'
        if unblock_header:
            unblock_range = _range(task_ws, unblock_header)
            ws.cell(row, 6).value = (
                f'=COUNTIF({unblock_range},"否")'
                f'+COUNTBLANK({unblock_range})'
            )
        else:
            ws.cell(row, 6).value = ""
        ws.cell(row, 7).value = f"=B{row}" if row in (13, 14, 16, 17) else ""
        ws.cell(row, 8).value = f"=B{row}" if row == 15 else ""
        ws.cell(row, 9).value = f'=IF(B{row}=0,1,C{row}/B{row})'

    nonblocking_ws = wb["非阻断观察"]
    mark_range = _range(nonblocking_ws, "是否标记为异常")
    upgrade_range = _range(nonblocking_ws, "是否升级为阻断")
    release_range = _range(nonblocking_ws, "阻断解除状态")
    row = 18
    ws.cell(row, 2).value = max(nonblocking_ws.max_row - 1, 0)
    ws.cell(row, 3).value = f'=COUNTIF({mark_range},"是")'
    ws.cell(row, 4).value = (
        f'=COUNTIFS({upgrade_range},"是",{release_range},"")'
    )
    ws.cell(row, 5).value = (
        f'=COUNTIF({mark_range},"需补充资料")'
        f'+COUNTIF({release_range},"需补充资料")'
    )
    ws.cell(row, 6).value = (
        f'=COUNTIFS({upgrade_range},"是",{release_range},"<>已解除")'
    )
    ws.cell(row, 7).value = ""
    ws.cell(row, 8).value = ""
    ws.cell(row, 9).value = f'=IF(B{row}=0,1,C{row}/B{row})'


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    for ws in wb.worksheets:
        ws.freeze_panes = FREEZE_PANES.get(ws.title, "A2")
        if ws.title == "确认进度总览":
            ws.auto_filter.ref = "A12:J18"
        else:
            ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(1, col).value or "")
            if "完整精度" in header:
                letter = get_column_letter(col)
                ws.column_dimensions[letter].hidden = True
                ws.column_dimensions[letter].width = 3

    _set_overview_formulas(wb)
    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    main()
