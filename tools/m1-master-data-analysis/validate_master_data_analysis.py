from __future__ import annotations

import csv
import hashlib
import json
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[2]
BILL_INPUT_ROOT = ROOT / "data" / "real-bills"
MASTER_INPUT_ROOT = ROOT / "data" / "master-data"
PUBLIC_ROOT = ROOT / "docs" / "analysis" / "m1-master-data"
PRIVATE_ROOT = ROOT / "data" / "m1-master-data-private"
OPS_ROOT = PRIVATE_ROOT / "ops-confirmation"
SUMMARY_PATH = PUBLIC_ROOT / "summary.json"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_unchanged(summary: dict) -> bool:
    for group in ("bill", "master"):
        for item in summary["source_snapshots"][group]:
            path = Path(item["path"])
            stat = path.stat()
            if stat.st_size != item["size"] or stat.st_mtime_ns != item["mtime_ns"] or sha256_file(path) != item["sha256"]:
                return False
    return True


def recompute_master_shape() -> tuple[int, int, int]:
    files = sorted(MASTER_INPUT_ROOT.glob("*.xlsx")) + sorted(MASTER_INPUT_ROOT.glob("*.xlsm"))
    sheet_count = 0
    row_count = 0
    column_count = 0
    for path in files:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_count += len(workbook.sheetnames)
        for worksheet in workbook.worksheets:
            header_row = None
            headers = []
            for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
                current = ["" if value is None else str(value).strip() for value in row]
                if "作品ID" in current:
                    header_row = row_number
                    headers = current
                    break
            if header_row is None:
                continue
            column_count = max(column_count, len(headers))
            for row in worksheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
                if not all(value is None or (isinstance(value, str) and not value.strip()) for value in row):
                    row_count += 1
        workbook.close()
    return sheet_count, row_count, column_count


def scan_public_for_sensitive_values() -> list[dict[str, str]]:
    values: set[tuple[str, str]] = set()
    sensitive_headers = ["出版书名", "合同书名", "作者署名", "作者原名", "授权方", "合同编号", "作品ID", "书号"]
    for path in sorted(MASTER_INPUT_ROOT.glob("*.xlsx")) + sorted(MASTER_INPUT_ROOT.glob("*.xlsm")):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            header_row = None
            headers = []
            for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
                current = ["" if value is None else str(value).strip() for value in row]
                if "作品ID" in current:
                    header_row = row_number
                    headers = current
                    break
            if header_row is None:
                continue
            indices = [headers.index(header) for header in sensitive_headers if header in headers]
            for row in worksheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
                for index in indices:
                    value = row[index]
                    if value is None:
                        continue
                    header = headers[index]
                    text = str(value).strip()
                    if text.isdigit() and len(text) < 7:
                        continue
                    if len(text) >= 4:
                        values.add((header, text))
        workbook.close()

    public_text = {}
    for path in PUBLIC_ROOT.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".md", ".html", ".json"}:
            public_text[path.name] = path.read_text(encoding="utf-8", errors="ignore")
    matches = []
    for header, value in values:
        pattern = re.compile(rf"(?<![A-Za-z0-9]){re.escape(value)}(?![A-Za-z0-9])") if value.isalnum() else re.compile(re.escape(value))
        for filename, text in public_text.items():
            if pattern.search(text):
                matches.append({"field": header, "value": value, "public_file": filename})
    return matches


def check_req_at() -> dict[str, object]:
    requirement_files = [
        ROOT / "docs/prd/10-data-foundation/bill-import.md",
        ROOT / "docs/prd/10-data-foundation/data-quality.md",
        ROOT / "docs/prd/10-data-foundation/work-master-data.md",
        ROOT / "docs/prd/10-data-foundation/channel-master-data.md",
        ROOT / "docs/prd/10-data-foundation/classification-and-tags.md",
        ROOT / "docs/prd/40-platform/platform-baseline.md",
    ]
    requirement_ids = []
    for path in requirement_files:
        requirement_ids.extend(re.findall(r"(?m)^## (REQ-[A-Z]+(?:-[A-Z]+)?-\d{3})", path.read_text(encoding="utf-8")))
    trace_text = (ROOT / "docs/prd/00-governance/traceability.md").read_text(encoding="utf-8")
    acceptance_text = (ROOT / "docs/prd/70-acceptance/M1.md").read_text(encoding="utf-8")
    traced_requirements = re.findall(r"REQ-[A-Z]+(?:-[A-Z]+)?-\d{3}", trace_text)
    traced_acceptance = re.findall(r"AT-M1-\d{3}", trace_text)
    acceptance_ids = re.findall(r"(?m)^### (AT-M1-\d{3})", acceptance_text)
    return {
        "requirement_count": len(requirement_ids),
        "trace_requirement_count": len(traced_requirements),
        "acceptance_count": len(acceptance_ids),
        "trace_acceptance_count": len(traced_acceptance),
        "consistent": len(requirement_ids) == len(set(requirement_ids))
        and set(requirement_ids) == set(traced_requirements)
        and len(acceptance_ids) == len(set(acceptance_ids))
        and set(acceptance_ids) == set(traced_acceptance),
    }


def main() -> None:
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    master_sheet_count, master_row_count, master_column_count = recompute_master_shape()
    source_ok = source_unchanged(summary)
    shape_ok = (
        master_sheet_count == summary["master_sheet_count"]
        and master_row_count == summary["master_row_count"]
        and master_column_count == summary["master_column_count"]
    )

    expected_reports = [f"{index:02d}-" for index in range(1, 15)]
    reports = sorted(path.name for path in PUBLIC_ROOT.glob("*.md") if re.match(r"\d{2}-", path.name))
    report_complete = all(any(name.startswith(prefix) for name in reports) for prefix in expected_reports)

    html_text = (PUBLIC_ROOT / "report.html").read_text(encoding="utf-8")
    html_sections = [
        "title",
        "technical-summary",
        "key-findings",
        "scope-data-and-metric-definitions",
        "methodology",
        "limitations-uncertainty-and-robustness-checks",
        "recommended-next-steps",
        "further-questions",
    ]
    html_ok = all(f'data-contract-section="{section}"' in html_text for section in html_sections)

    image_checks = []
    for image_path in sorted((PUBLIC_ROOT / "assets").glob("*.png")):
        with Image.open(image_path) as image:
            image.verify()
        with Image.open(image_path) as image:
            image_checks.append({"file": image_path.name, "width": image.width, "height": image.height, "bytes": image_path.stat().st_size})
    image_ok = len(image_checks) >= 3 and all(item["width"] > 100 and item["height"] > 100 for item in image_checks)

    workbook_path = OPS_ROOT / "M1-运营确认包.xlsx"
    workbook_ok = workbook_path.exists() and workbook_path.stat().st_size > 0
    workbook_sheets = []
    if workbook_ok:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        workbook_sheets = workbook.sheetnames
        workbook_ok = len(workbook_sheets) >= 9 and "README" in workbook_sheets
        workbook.close()

    leak_rows = scan_public_for_sensitive_values()
    leak_path = PRIVATE_ROOT / "public-sensitive-leak-review.csv"
    with leak_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["field", "value", "public_file"])
        writer.writeheader()
        writer.writerows(leak_rows)
    privacy_ok = len(leak_rows) == 0

    gitignore_text = (ROOT / ".gitignore").read_text(encoding="utf-8", errors="ignore")
    ignored_ok = all(
        item in gitignore_text
        for item in ["/data/real-bills/", "/data/master-data/", "/data/m1-real-bills-private/", "/data/m1-master-data-private/"]
    )
    req_at = check_req_at()

    validation = {
        "generated_at": __import__("datetime").datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_unchanged": source_ok,
        "master_shape_recomputed": {
            "sheet_count": master_sheet_count,
            "row_count": master_row_count,
            "column_count": master_column_count,
            "matches_summary": shape_ok,
        },
        "report_complete": report_complete,
        "html_contract_ok": html_ok,
        "image_ok": image_ok,
        "workbook_ok": workbook_ok,
        "workbook_sheets": workbook_sheets,
        "privacy_ok": privacy_ok,
        "public_sensitive_matches": len(leak_rows),
        "ignored_paths_ok": ignored_ok,
        "req_at": req_at,
        "overall": all([source_ok, shape_ok, report_complete, html_ok, image_ok, workbook_ok, privacy_ok, ignored_ok, req_at["consistent"]]),
    }
    (PUBLIC_ROOT / "validation.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    report = f"""# M1 数字版权台账分析验证报告

## Overall Assessment: {'Ready to share with caveats' if validation['overall'] else 'Needs revision'}

### Methodology Review

- 原始账单和台账文件运行后校验：{'通过' if source_ok else '失败'}。
- 台账规模独立复算：{'通过' if shape_ok else '失败'}。
- 公开报告完整性：{'通过' if report_complete else '失败'}。
- HTML 报告结构：{'通过' if html_ok else '失败'}。
- 图表 PNG 校验：{'通过' if image_ok else '失败'}。
- 运营确认包工作簿：{'通过' if workbook_ok else '失败'}。
- 公开报告敏感值扫描：{'通过' if privacy_ok else '失败'}，匹配数 {len(leak_rows)}。
- REQ/AT 追踪一致性：{'通过' if req_at['consistent'] else '失败'}。

### Calculation Spot-Checks

- 台账行数：复算 {master_row_count}，报告 {summary['master_row_count']}。
- 台账字段数：复算 {master_column_count}，报告 {summary['master_column_count']}。
- 台账覆盖率：报告 {summary['master_covered_standard_work_count']} / {summary['bill_standard_work_count']}。

### Required Caveats for Stakeholders

- 台账只有一个输入文件，未来新增模板需要重新分析。
- 分类树、标签库、版权开始日期权威字段和冲突解除仍需运营确认。
- 金额 `NUMERIC(32,18)` 是候选精度，物理模型阶段最终确认。
"""
    (PUBLIC_ROOT / "validation-report.md").write_text(report, encoding="utf-8")
    if not validation["overall"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
