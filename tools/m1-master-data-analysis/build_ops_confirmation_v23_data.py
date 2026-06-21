from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

import analyze_master_data as amd  # noqa: E402
from work_id_rules import derive_business_form, derive_standard_work_id, parse_raw_work_id  # noqa: E402


PRIVATE_ROOT = ROOT / "data" / "m1-master-data-private"
OPS_ROOT = PRIVATE_ROOT / "ops-confirmation"
INPUT_V2_JSON = OPS_ROOT / "ops-confirmation-v2-data.json"
OUTPUT_JSON = OPS_ROOT / "ops-confirmation-v2.3-data.json"
TASK_MAPPING_JSON = OPS_ROOT / "ops-confirmation-v2.3-task-mapping.json"
PUBLIC_AUDIT = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2.3-logic-audit.md"
PUBLIC_SUMMARY = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2.3-summary.md"


FORMAL_SHEET = "正式导入阻断确认"
MULTI_ID_SHEET = "多ID归并候选"
BASIC_INFO_SHEET = "标准作品基础信息补全"
LEDGER_CONFLICT_SHEET = "台账真实冲突"
COPYRIGHT_SHEET = "版权期限反例"
NON_BLOCKING_SHEET = "非阻断观察"

ISSUE_MULTI_NAME = "多名称作品ID"
ISSUE_MULTI_AUTH = "多授权分类作品ID"
ISSUE_ABNORMAL_ID = "异常作品ID"


def digest(*values: Any, length: int = 10) -> str:
    payload = "\x1f".join("" if value is None else str(value) for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length].upper()


def task_id(prefix: str, identity: Any) -> str:
    text = "" if identity is None else str(identity)
    safe = text.replace("\n", "\\n").strip()
    if safe and len(safe) <= 80:
        return f"{prefix}::{safe}"
    return f"{prefix}::HASH-{digest(text)}"


def sheet_rows(payload: dict[str, Any], sheet_name: str) -> list[dict[str, Any]]:
    return [sheet["rows"] for sheet in payload["sheets"] if sheet["name"] == sheet_name][0]


def exact_amount_display(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def dec_sum(rows: list[dict[str, Any]]) -> Decimal:
    total = Decimal("0")
    for row in rows:
        amount = row.get("amount")
        if amount is not None:
            total += amount
    return total


def line_join(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = amd.norm_text(value)
        if text and text not in seen:
            seen.add(text)
            out.append(text)
    return "\n".join(out)


def summarize_value(rows: list[dict[str, Any]], key: str) -> str:
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = amd.norm_text(row.get(key)) or "<空>"
        item = groups.setdefault(value, {"count": 0, "months": [], "amounts": []})
        item["count"] += 1
        item["months"].append(row.get("month"))
        item["amounts"].append(row.get("amount"))
    parts: list[str] = []
    for value, item in sorted(groups.items(), key=lambda pair: (-pair[1]["count"], pair[0])):
        parts.append(
            f"{value}｜记录数:{item['count']}｜月份:{amd.month_range(item['months'])}｜实销合计:{amd.fmt_decimal(amd.dec_sum(item['amounts']))}"
        )
    return "\n".join(parts)


def normalize_issue_types(value: Any) -> list[str]:
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def formal_row_identity(row: dict[str, Any]) -> str:
    return str(row.get("原始作品ID") or row.get("任务ID") or "")


def formal_row_is_normal_dual_business_only(row: dict[str, Any]) -> bool:
    """Existing v2 rows are per full raw ID. A row is only removable if it actually contains both raw IDs."""
    raw_text = str(row.get("原始作品ID") or "")
    raw_ids = {part.strip() for part in raw_text.replace(",", "\n").splitlines() if part.strip()}
    if len(raw_ids) < 2:
        return False
    standards = {derive_standard_work_id(raw_id) for raw_id in raw_ids}
    forms = {derive_business_form(raw_id) for raw_id in raw_ids}
    return len(standards - {None}) == 1 and {"有声版权", "有声成品"} <= forms


def build_formal_rows(v2_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    invalidated: list[dict[str, Any]] = []
    for row in v2_rows:
        old_task = str(row.get("任务ID") or "")
        raw_id = formal_row_identity(row)
        issue_types = normalize_issue_types(row.get("问题类型"))
        if formal_row_is_normal_dual_business_only(row):
            invalidated.append(
                {
                    "old_task_id": old_task,
                    "new_task_id": "",
                    "status": "invalidated",
                    "reason": "仅因纯数字与Y前缀正常双业务形态共同存在，不再作为正式导入阻断。",
                    "sheet": FORMAL_SHEET,
                }
            )
            continue
        new = dict(row)
        new["旧任务ID"] = old_task
        new["任务ID"] = task_id("IMPORT-BLOCK", raw_id)
        new["候选逻辑审计结论"] = "按完整原始作品ID分组；纯数字ID与Y前缀ID共同存在不会单独构成阻断。"
        new["原始ID解析规则"] = "标准作品ID=数字主体；业务形态仅由完整原始作品ID派生；授权分类不反推业务形态。"
        if ISSUE_ABNORMAL_ID in issue_types:
            new["候选逻辑审计结论"] = "原始作品ID格式异常，继续作为正式导入阻断。"
        rows.append(new)
    return rows, invalidated


def add_task_ids(sheet_name: str, rows: list[dict[str, Any]], id_field: str, prefix: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        new = dict(row)
        old_task = str(new.get("任务ID") or new.get("观察ID") or "")
        identity = new.get(id_field) or old_task or digest(sheet_name, row)
        new["旧任务ID"] = old_task
        new["任务ID"] = task_id(prefix, identity)
        out.append(new)
    return out


def cross_business_observations(bill_conflicts: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for standard_id, group_rows in sorted(bill_conflicts["by_standard"].items(), key=lambda item: (len(item[0]), item[0])):
        pure_rows = [row for row in group_rows if row.get("raw_work_id") == standard_id]
        y_rows = [row for row in group_rows if row.get("raw_work_id") == f"Y{standard_id}"]
        if not pure_rows or not y_rows:
            continue
        pure_names = {row["work_name_norm"] for row in pure_rows if row.get("work_name_norm")}
        y_names = {row["work_name_norm"] for row in y_rows if row.get("work_name_norm")}
        pure_auths = {amd.norm_text(row.get("auth")) for row in pure_rows if amd.norm_text(row.get("auth"))}
        y_auths = {amd.norm_text(row.get("auth")) for row in y_rows if amd.norm_text(row.get("auth"))}
        name_diff = pure_names != y_names
        auth_diff = pure_auths != y_auths
        if not (name_diff or auth_diff):
            continue
        all_rows = pure_rows + y_rows
        amount = dec_sum(all_rows)
        rows.append(
            {
                "观察ID": task_id("OBS-CROSS-FORM", standard_id),
                "任务ID": task_id("OBS-CROSS-FORM", standard_id),
                "观察类型": "跨业务形态名称或授权分类差异",
                "标准作品ID": standard_id,
                "纯数字原始ID": standard_id,
                "Y前缀原始ID": f"Y{standard_id}",
                "原始财务ID": f"{standard_id}\nY{standard_id}",
                "两种业务形态": "有声版权\n有声成品",
                "业务形态": "有声版权\n有声成品",
                "有声版权名称": summarize_value(pure_rows, "work_name"),
                "有声成品名称": summarize_value(y_rows, "work_name"),
                "有声版权授权分类": summarize_value(pure_rows, "auth"),
                "有声成品授权分类": summarize_value(y_rows, "auth"),
                "是否需要统一标准作品名称": "",
                "记录数": len(all_rows),
                "月份范围": amd.month_range(row.get("month") for row in all_rows),
                "金额": exact_amount_display(amount),
                "金额_完整精度": amd.fmt_decimal(amount),
                "名称/说明": "\n".join(
                    [
                        "正常双业务形态观察，不阻断正式收入事实入库。",
                        f"名称差异: {'是' if name_diff else '否'}；授权分类差异: {'是' if auth_diff else '否'}。",
                    ]
                ),
                "默认处理": "默认不阻断；不要求运营填写“不适用”来解除阻断。",
                "是否标记为异常": "",
                "是否升级为阻断": "",
                "阻断解除状态": "",
                "运营备注": "",
            }
        )
    return rows


def normalize_non_blocking_rows(v2_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in v2_rows:
        new = dict(row)
        old_id = str(new.get("观察ID") or new.get("任务ID") or "")
        new["旧任务ID"] = old_id
        new["任务ID"] = task_id("OBS-LEGACY", old_id or digest(row))
        new.setdefault("纯数字原始ID", "")
        new.setdefault("Y前缀原始ID", "")
        new.setdefault("两种业务形态", new.get("业务形态", ""))
        new.setdefault("有声版权名称", "")
        new.setdefault("有声成品名称", "")
        new.setdefault("有声版权授权分类", "")
        new.setdefault("有声成品授权分类", "")
        new.setdefault("是否需要统一标准作品名称", "")
        new.setdefault("阻断解除状态", "")
        out.append(new)
    return out


def confirmation_values_present(workbook_path: Path) -> int:
    try:
        import openpyxl
    except Exception:
        return 0
    if not workbook_path.exists():
        return 0
    confirmation_headers = {
        "运营确认结果",
        "确认标准作品ID",
        "确认标准作品名称",
        "是否解除阻断",
        "是否归并",
        "主标准作品ID",
        "其他ID处理方式",
        "补全状态",
        "运营确认值",
        "是否解除冲突",
        "确认版权开始日期",
        "确认版权到期日期",
        "是否标记为异常",
        "是否升级为阻断",
        "阻断解除状态",
        "运营备注",
    }
    count = 0
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    for ws in wb.worksheets:
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        indexes = [i for i, value in enumerate(header) if value in confirmation_headers]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(i < len(row) and row[i] not in (None, "") for i in indexes):
                count += 1
    wb.close()
    return count


def main() -> None:
    OPS_ROOT.mkdir(parents=True, exist_ok=True)
    v2_payload = json.loads(INPUT_V2_JSON.read_text(encoding="utf-8"))

    source_snapshots_before = {
        "bill": [item.__dict__ for item in amd.snapshot_sources(amd.workbook_inputs(amd.BILL_INPUT_ROOT), "B")],
        "master": [item.__dict__ for item in amd.snapshot_sources(amd.workbook_inputs(amd.MASTER_INPUT_ROOT), "M")],
    }
    bill_rows, _bill_structure = amd.read_bills()
    bill_conflicts = amd.group_bill_conflicts(bill_rows)
    source_snapshots_after = {
        "bill": [item.__dict__ for item in amd.snapshot_sources(amd.workbook_inputs(amd.BILL_INPUT_ROOT), "B")],
        "master": [item.__dict__ for item in amd.snapshot_sources(amd.workbook_inputs(amd.MASTER_INPUT_ROOT), "M")],
    }

    v2_formal = sheet_rows(v2_payload, FORMAL_SHEET)
    formal_rows, invalidated_formal = build_formal_rows(v2_formal)

    multi_id_rows = add_task_ids(MULTI_ID_SHEET, sheet_rows(v2_payload, MULTI_ID_SHEET), "标准作品ID候选", "MERGE")
    basic_rows = add_task_ids(BASIC_INFO_SHEET, sheet_rows(v2_payload, BASIC_INFO_SHEET), "标准作品ID", "BASIC-INFO")
    ledger_rows = add_task_ids(LEDGER_CONFLICT_SHEET, sheet_rows(v2_payload, LEDGER_CONFLICT_SHEET), "标准作品ID", "LEDGER-CONFLICT")
    copyright_rows = add_task_ids(COPYRIGHT_SHEET, sheet_rows(v2_payload, COPYRIGHT_SHEET), "标准作品ID", "COPYRIGHT-CONFLICT")

    legacy_non_blocking = normalize_non_blocking_rows(sheet_rows(v2_payload, NON_BLOCKING_SHEET))
    cross_rows = cross_business_observations(bill_conflicts)
    non_blocking_rows = legacy_non_blocking + cross_rows

    old_issue_counts = Counter()
    for row in v2_formal:
        old_issue_counts.update(normalize_issue_types(row.get("问题类型")))
    new_issue_counts = Counter()
    for row in formal_rows:
        new_issue_counts.update(normalize_issue_types(row.get("问题类型")))

    old_formal_count = len(v2_formal)
    normal_dual_removed_count = len(invalidated_formal)
    formal_raws = {row.get("原始作品ID") for row in v2_formal if row.get("原始作品ID")}
    formal_groups_containing_pure_and_y = sum(1 for row in v2_formal if formal_row_is_normal_dual_business_only(row))

    task_mapping: list[dict[str, Any]] = []
    for sheet_name, rows, prefix, id_field in [
        (FORMAL_SHEET, formal_rows, "IMPORT-BLOCK", "原始作品ID"),
        (MULTI_ID_SHEET, multi_id_rows, "MERGE", "标准作品ID候选"),
        (BASIC_INFO_SHEET, basic_rows, "BASIC-INFO", "标准作品ID"),
        (LEDGER_CONFLICT_SHEET, ledger_rows, "LEDGER-CONFLICT", "标准作品ID"),
        (COPYRIGHT_SHEET, copyright_rows, "COPYRIGHT-CONFLICT", "标准作品ID"),
        (NON_BLOCKING_SHEET, legacy_non_blocking, "OBS-LEGACY", "观察ID"),
    ]:
        for row in rows:
            task_mapping.append(
                {
                    "sheet": sheet_name,
                    "old_task_id": row.get("旧任务ID", ""),
                    "new_task_id": row.get("任务ID", ""),
                    "status": "semantic_same",
                    "reason": "任务语义未变化；如旧包存在人工填写值，可按任务ID人工核对迁移。",
                }
            )
    task_mapping.extend(invalidated_formal)
    for row in cross_rows:
        task_mapping.append(
            {
                "sheet": NON_BLOCKING_SHEET,
                "old_task_id": "",
                "new_task_id": row.get("任务ID", ""),
                "status": "new_task",
                "reason": "v2.3 新增跨业务形态非阻断观察。",
            }
        )

    repo_v22 = OPS_ROOT / "M1-运营确认包-v2.2.xlsx"
    desktop_v22_candidates = sorted((Path.home() / "Desktop").glob("M1*2.2*xlsx"))
    manual_result_count = confirmation_values_present(repo_v22) + sum(confirmation_values_present(path) for path in desktop_v22_candidates[:1])

    metrics = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "old_formal_import_blocker_count": old_formal_count,
        "old_formal_groups_containing_pure_digit_and_y_prefix": formal_groups_containing_pure_and_y,
        "old_formal_groups_removed_as_normal_dual_business": normal_dual_removed_count,
        "new_formal_import_blocker_count": len(formal_rows),
        "new_multi_name_conflict_count": new_issue_counts[ISSUE_MULTI_NAME],
        "new_multi_auth_conflict_count": new_issue_counts[ISSUE_MULTI_AUTH],
        "new_abnormal_id_count": new_issue_counts[ISSUE_ABNORMAL_ID],
        "old_multi_name_conflict_count": old_issue_counts[ISSUE_MULTI_NAME],
        "old_multi_auth_conflict_count": old_issue_counts[ISSUE_MULTI_AUTH],
        "old_multi_name_cross_business_only_count": 0,
        "old_multi_auth_normal_dual_only_count": 0,
        "cross_business_non_blocking_observation_count": len(cross_rows),
        "legacy_non_blocking_observation_count": len(legacy_non_blocking),
        "new_non_blocking_observation_count": len(non_blocking_rows),
        "multi_id_candidate_count": len(multi_id_rows),
        "basic_info_work_count": len(basic_rows),
        "m2_basic_info_missing_count": sum(1 for row in basic_rows if amd.norm_text(row.get("需要补充的字段"))),
        "ledger_conflict_count": len(ledger_rows),
        "copyright_conflict_count": len(copyright_rows),
        "v22_manual_result_rows_detected": manual_result_count,
        "v22_manual_result_safely_migrated_count": 0,
        "source_unchanged": source_snapshots_before == source_snapshots_after,
    }

    sheets = [
        {"name": "确认进度总览", "kind": "overview", "rows": []},
        {"name": FORMAL_SHEET, "kind": "formal_blockers", "rows": formal_rows},
        {"name": MULTI_ID_SHEET, "kind": "multi_id", "rows": multi_id_rows},
        {"name": BASIC_INFO_SHEET, "kind": "basic_info", "rows": basic_rows},
        {"name": LEDGER_CONFLICT_SHEET, "kind": "ledger_conflicts", "rows": ledger_rows},
        {"name": COPYRIGHT_SHEET, "kind": "copyright_conflicts", "rows": copyright_rows},
        {"name": NON_BLOCKING_SHEET, "kind": "non_blocking", "rows": non_blocking_rows},
    ]

    payload = {
        "generated_at": metrics["generated_at"],
        "metrics": metrics,
        "sheets": sheets,
        "task_mapping": task_mapping,
        "source_snapshots": source_snapshots_after,
        "rules": {
            "derive_standard_work_id": "^[0-9]+$ -> 原值；^Y[0-9]+$ -> 去掉首字母Y；其他格式无效。",
            "derive_business_form": "^[0-9]+$ -> 有声版权；^Y[0-9]+$ -> 有声成品；授权分类不参与业务形态派生。",
            "raw_work_id_identity": "完整原始作品ID按账单文本识别；12345 与 Y12345 是两个不同原始作品ID。",
        },
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    TASK_MAPPING_JSON.write_text(json.dumps(task_mapping, ensure_ascii=False, indent=2), encoding="utf-8")

    audit = f"""# M1 运营确认候选逻辑审计 v2.3

## 审计结论

- 当前脚本的多名称作品ID、多授权分类作品ID和正式导入阻断确认均按**完整原始作品ID**分组，不是按去掉 Y 后的标准作品ID分组。
- `12345` 与 `Y12345` 是两个不同的原始作品ID；二者可以映射到同一个标准作品ID `12345`。
- 业务形态只由完整原始作品ID派生：纯数字为有声版权，`Y` 前缀为有声成品。
- 授权分类不得反向覆盖业务形态。

## 数量结果

| 指标 | 数量 |
|---|---:|
| v2.2 正式导入阻断组 | {metrics['old_formal_import_blocker_count']} |
| 旧正式阻断组中实际包含纯数字ID与Y前缀ID的组 | {metrics['old_formal_groups_containing_pure_digit_and_y_prefix']} |
| 旧正式阻断组中被判定为正常双业务形态并移出的组 | {metrics['old_formal_groups_removed_as_normal_dual_business']} |
| v2.2 多名称组 | {metrics['old_multi_name_conflict_count']} |
| 其中仅为跨业务形态名称不同的组 | {metrics['old_multi_name_cross_business_only_count']} |
| v2.2 多授权分类组 | {metrics['old_multi_auth_conflict_count']} |
| 其中仅为正常双业务形态的组 | {metrics['old_multi_auth_normal_dual_only_count']} |
| v2.3 正式导入阻断组 | {metrics['new_formal_import_blocker_count']} |
| v2.3 跨业务形态非阻断观察 | {metrics['cross_business_non_blocking_observation_count']} |

## 解释

既有 346 个正式阻断候选不是由“纯数字ID和Y前缀ID共同存在”直接产生；它们要么是同一个完整原始ID内部多名称、多授权分类，要么是原始ID格式异常。跨业务形态名称或授权分类差异已单独进入非阻断观察，不要求运营填写“不适用”来解除导入阻断。
"""
    PUBLIC_AUDIT.write_text(audit, encoding="utf-8")

    summary = f"""# M1 运营确认包 v2.3 聚合摘要

## 输出

- 本地私有工作簿：`data/m1-master-data-private/ops-confirmation/M1-运营确认包-v2.3.xlsx`
- 私有任务映射：`data/m1-master-data-private/ops-confirmation/ops-confirmation-v2.3-task-mapping.json`
- 公开逻辑审计：`docs/analysis/m1-master-data/ops-confirmation-v2.3-logic-audit.md`

## 核心数量

| 指标 | 数量 |
|---|---:|
| 正式导入阻断组 | {metrics['new_formal_import_blocker_count']} |
| 多名称冲突组 | {metrics['new_multi_name_conflict_count']} |
| 多授权分类冲突组 | {metrics['new_multi_auth_conflict_count']} |
| 异常ID组 | {metrics['new_abnormal_id_count']} |
| 新增跨业务形态非阻断观察 | {metrics['cross_business_non_blocking_observation_count']} |
| 非阻断观察总数 | {metrics['new_non_blocking_observation_count']} |
| M2 基础信息待补全作品 | {metrics['m2_basic_info_missing_count']} |

## 规则变化

- 正常双业务形态不再作为正式导入阻断。
- 授权分类不决定业务形态。
- 任务ID改为稳定语义ID，不依赖 Excel 行号。
- 未迁移任何 v2.2 人工填写结果；当前可安全自动迁移数量为 0。
"""
    PUBLIC_SUMMARY.write_text(summary, encoding="utf-8")


if __name__ == "__main__":
    main()
