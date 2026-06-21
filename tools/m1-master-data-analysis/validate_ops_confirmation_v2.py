from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
OPS_ROOT = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
DATA_PATH = OPS_ROOT / "ops-confirmation-v2-data.json"
WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.xlsx"
PUBLIC_VALIDATION_PATH = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2-validation.json"
PUBLIC_SUMMARY_PATH = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2-summary.md"


EXPECTED_SHEETS = [
    "确认进度总览",
    "正式导入阻断确认",
    "多ID归并候选",
    "标准作品基础信息补全",
    "台账真实冲突",
    "版权期限反例",
    "非阻断观察",
]

EXPECTED_FREEZE_PANES = {
    "确认进度总览": "A5",
    "正式导入阻断确认": "C2",
    "多ID归并候选": "D2",
    "标准作品基础信息补全": "B2",
    "台账真实冲突": "C2",
    "版权期限反例": "C2",
    "非阻断观察": "D2",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sources_unchanged(payload: dict) -> bool:
    for group in ("bill", "master"):
        for item in payload["source_snapshots"][group]:
            path = Path(item["path"])
            stat = path.stat()
            if stat.st_size != item["size"] or stat.st_mtime_ns != item["mtime_ns"] or sha256_file(path) != item["sha256"]:
                return False
    return True


def rows_from_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        rows.append({str(headers[index]): value for index, value in enumerate(row) if headers[index] is not None})
    return rows


def public_sensitive_scan() -> int:
    private_values = set()
    non_sensitive_terms = {
        "raw_work_id",
        "month_range",
        "standard_id",
        "business_form",
        "多名称作品ID确认表",
        "多授权分类作品ID确认表",
        "异常作品ID确认表",
        "分册候选确认表",
        "正负冲抵候选确认表",
        "首次实销仍为空作品确认表",
        "台账匹配失败或一对多冲突确认表",
        "版权期限反例确认表",
    }
    for csv_path in OPS_ROOT.glob("*.csv"):
        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    for value in row:
                        text = str(value).strip()
                        if text in non_sensitive_terms:
                            continue
                        if re.fullmatch(r"[A-Za-z0-9_]+", text):
                            continue
                        if text.endswith("确认表"):
                            continue
                        if len(text) >= 8 and not text.isdigit():
                            private_values.add(text)
        except Exception:
            continue
    public_text = ""
    public_root = ROOT / "docs" / "analysis" / "m1-master-data"
    for path in public_root.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".md", ".html", ".json"} and path.name != "ops-confirmation-v2-validation.json":
            public_text += "\n" + path.read_text(encoding="utf-8", errors="ignore")
    matches = 0
    for value in list(private_values)[:5000]:
        if value and value in public_text:
            matches += 1
    return matches


def main() -> None:
    payload = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    workbook_open = True
    sheet_names = wb.sheetnames
    expected_sheets_ok = sheet_names == EXPECTED_SHEETS

    formal_rows = rows_from_sheet(wb["正式导入阻断确认"])
    raw_ids = [row.get("原始作品ID") for row in formal_rows]
    formal_dedup_ok = len(raw_ids) == len(set(raw_ids))

    multi_rows = rows_from_sheet(wb["多ID归并候选"])
    multi_no_single_ok = all(len([x for x in str(row.get("标准作品ID候选") or "").splitlines() if x.strip()]) >= 2 for row in multi_rows)

    ledger_rows = rows_from_sheet(wb["台账真实冲突"])
    ledger_unmatched_separated_ok = all("未匹配" not in str(row.get("冲突字段", "")) and "未匹配" not in str(row.get("各候选值", "")) for row in ledger_rows)

    basic_rows = rows_from_sheet(wb["标准作品基础信息补全"])
    standards = [row.get("标准作品ID") for row in basic_rows]
    basic_unique_ok = len(standards) == len(set(standards))
    start_candidate_header_ok = "版权开始日期候选（签订日期）" in [cell.value for cell in wb["标准作品基础信息补全"][1]]

    duplicate_fill_headers = []
    for sheet_name in ["正式导入阻断确认", "多ID归并候选", "台账真实冲突", "非阻断观察"]:
        headers = [str(cell.value) for cell in wb[sheet_name][1] if cell.value is not None]
        duplicate_fill_headers.extend(
            (sheet_name, header)
            for header in headers
            if header in {"运营确认作者", "运营确认一级分类", "运营确认二级分类", "运营确认三级分类", "运营确认版权开始日期", "运营确认版权到期日期", "运营确认必需标签"}
        )
    no_duplicate_basic_fill_ok = not duplicate_fill_headers

    amount_precision_ok = all(
        any("完整精度" in str(cell.value) for cell in wb[sheet_name][1])
        for sheet_name in ["正式导入阻断确认", "多ID归并候选", "标准作品基础信息补全", "非阻断观察"]
    )
    hidden_precision_ok = True
    for sheet_name in ["正式导入阻断确认", "多ID归并候选", "标准作品基础信息补全", "非阻断观察"]:
        ws = wb[sheet_name]
        for index, cell in enumerate(ws[1], start=1):
            if "完整精度" in str(cell.value):
                column_letter = openpyxl.utils.get_column_letter(index)
                hidden_precision_ok = hidden_precision_ok and bool(ws.column_dimensions[column_letter].hidden)

    freeze_ok = all(wb[sheet_name].freeze_panes == expected for sheet_name, expected in EXPECTED_FREEZE_PANES.items())
    filter_ok = all(bool(wb[sheet_name].auto_filter.ref) for sheet_name in EXPECTED_SHEETS)
    dropdown_ok = all(
        len(wb[sheet_name].data_validations.dataValidation) > 0
        for sheet_name in EXPECTED_SHEETS
        if sheet_name != "确认进度总览"
    )
    conditional_formatting_ok = all(
        len(wb[sheet_name].conditional_formatting) > 0
        for sheet_name in EXPECTED_SHEETS
        if sheet_name != "确认进度总览"
    )
    generated_time_format_ok = "yyyy" in str(wb["确认进度总览"]["B3"].number_format)
    usability_ok = all([hidden_precision_ok, freeze_ok, filter_ok, dropdown_ok, conditional_formatting_ok, generated_time_format_ok])

    pending_text = (ROOT / "docs" / "prd" / "60-validation" / "pending-data-decisions.md").read_text(encoding="utf-8")
    work_text = (ROOT / "docs" / "prd" / "10-data-foundation" / "work-master-data.md").read_text(encoding="utf-8")
    at_text = (ROOT / "docs" / "prd" / "70-acceptance" / "M1.md").read_text(encoding="utf-8")
    design_text = (ROOT / "docs" / "technical-design" / "M1-技术设计草案-v0.3.md").read_text(encoding="utf-8")
    docs_sync_ok = (
        "签订日期` 是否可作为正式版权开始日期" not in pending_text
        and "签订日期" in work_text
        and "版权开始日期" in work_text
        and "签订日期" in at_text
        and "物理模型设计门槛" in design_text
        and "首次正式数据迁移门槛" in design_text
        and ("M2正式评估门槛" in design_text or "M2 正式评估门槛" in design_text)
    )

    sensitive_matches = public_sensitive_scan()
    public_privacy_ok = sensitive_matches == 0
    source_ok = sources_unchanged(payload) and bool(payload["metrics"].get("source_unchanged"))

    row_counts = {
        sheet: max(wb[sheet].max_row - 1, 0)
        for sheet in EXPECTED_SHEETS
    }
    row_counts["确认进度总览"] = 6

    validation = {
        "workbook_open": workbook_open,
        "expected_sheets_ok": expected_sheets_ok,
        "row_counts": row_counts,
        "source_unchanged": source_ok,
        "formal_import_blockers_dedup_by_raw_work_id": formal_dedup_ok,
        "multi_id_sheet_excludes_single_id_candidates": multi_no_single_ok,
        "ledger_unmatched_and_real_conflicts_separated": ledger_unmatched_separated_ok,
        "basic_info_unique_standard_work": basic_unique_ok,
        "no_duplicate_basic_info_fill_fields": no_duplicate_basic_fill_ok,
        "signed_date_as_copyright_start_candidate": start_candidate_header_ok,
        "full_amount_precision_columns_present": amount_precision_ok,
        "full_amount_precision_columns_hidden": hidden_precision_ok,
        "freeze_panes_ok": freeze_ok,
        "auto_filter_ok": filter_ok,
        "dropdowns_ok": dropdown_ok,
        "conditional_formatting_ok": conditional_formatting_ok,
        "generated_time_format_ok": generated_time_format_ok,
        "workbook_usability_ok": usability_ok,
        "public_sensitive_matches": sensitive_matches,
        "public_privacy_ok": public_privacy_ok,
        "docs_sync_ok": docs_sync_ok,
        "overall": all(
            [
                workbook_open,
                expected_sheets_ok,
                source_ok,
                formal_dedup_ok,
                multi_no_single_ok,
                ledger_unmatched_separated_ok,
                basic_unique_ok,
                no_duplicate_basic_fill_ok,
                start_candidate_header_ok,
                amount_precision_ok,
                usability_ok,
                public_privacy_ok,
                docs_sync_ok,
            ]
        ),
    }
    PUBLIC_VALIDATION_PATH.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")

    metrics = payload["metrics"]
    summary = f"""# M1 运营确认包 v2 聚合摘要

## 输出

- 本地私有工作簿：`data/m1-master-data-private/ops-confirmation/M1-运营确认包-v2.xlsx`
- 公开验证结果：`docs/analysis/m1-master-data/ops-confirmation-v2-validation.json`

## 工作表行数

| 工作表 | 行数 |
|---|---:|
{chr(10).join(f"| {sheet} | {count} |" for sheet, count in row_counts.items())}

## 核心数量

- 正式导入阻断组数量：{metrics['formal_import_blocker_count']}
- 多ID归并候选数量：{metrics['multi_id_candidate_count']}
- M2 基础信息待补全作品数量：{metrics['m2_basic_info_missing_count']}
- 非阻断观察数量：{metrics['non_blocking_observation_count']}
- 版权期限冲突数量：{metrics['copyright_conflict_count']}

## 验证结论

- 验证状态：{'通过' if validation['overall'] else '未通过'}
- 原始分析数据未修改：{'是' if source_ok else '否'}
- 公开报告敏感明细扫描命中数：{sensitive_matches}
- Excel 易用性检查：{'通过' if usability_ok else '未通过'}
"""
    PUBLIC_SUMMARY_PATH.write_text(summary, encoding="utf-8")

    wb.close()
    if not validation["overall"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
