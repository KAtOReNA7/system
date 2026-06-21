from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
sys.path.insert(0, str(SCRIPT_DIR))

from work_id_rules import derive_business_form, derive_standard_work_id, parse_raw_work_id  # noqa: E402


OPS_ROOT = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.3.xlsx"
OLD_WORKBOOK_PATH = OPS_ROOT / "M1-运营确认包-v2.2.xlsx"
DATA_PATH = OPS_ROOT / "ops-confirmation-v2.3-data.json"
PUBLIC_VALIDATION_PATH = ROOT / "docs" / "analysis" / "m1-master-data" / "ops-confirmation-v2.3-validation.json"

EXPECTED_SHEETS = [
    "确认进度总览",
    "正式导入阻断确认",
    "多ID归并候选",
    "标准作品基础信息补全",
    "台账真实冲突",
    "版权期限反例",
    "非阻断观察",
]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def headers(ws: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> list[str]:
    return [str(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]


def rows_from_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    head = headers(ws)
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({head[index]: value for index, value in enumerate(values) if index < len(head)})
    return rows


def public_sensitive_scan() -> int:
    public_root = ROOT / "docs" / "analysis" / "m1-master-data"
    text = ""
    for path in public_root.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".md", ".json", ".html"}:
            text += "\n" + path.read_text(encoding="utf-8", errors="ignore")
    sensitive_patterns = [
        r"历史作品名称及记录数.+\n.+｜记录数",
        r"有声版权名称.+\n.+｜记录数",
        r"作者候选.+\n.+",
        r"版权开始日期候选.+\d{4}-\d{2}-\d{2}",
    ]
    return sum(1 for pattern in sensitive_patterns if re.search(pattern, text))


def sources_unchanged(payload: dict[str, Any]) -> bool:
    for group in ("bill", "master"):
        for item in payload["source_snapshots"][group]:
            path = Path(item["path"])
            if not path.exists():
                return False
            stat = path.stat()
            if stat.st_size != item["size"] or stat.st_mtime_ns != item["mtime_ns"]:
                return False
            if file_sha256(path) != item["sha256"]:
                return False
    return True


def no_sql_files_generated() -> bool:
    blocked_roots = [ROOT / "migrations", ROOT / "db" / "migrations"]
    if any(path.exists() for path in blocked_roots):
        return False
    sql_files = [
        path
        for path in ROOT.rglob("*.sql")
        if ".analysis-" not in str(path) and "node_modules" not in str(path)
    ]
    return not sql_files


def main() -> None:
    payload = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)

    workbook_open = True
    expected_sheets_ok = wb.sheetnames == EXPECTED_SHEETS

    row_counts = {sheet: max(wb[sheet].max_row - 1, 0) for sheet in EXPECTED_SHEETS}
    row_counts["确认进度总览"] = 17

    formal_rows = rows_from_sheet(wb["正式导入阻断确认"])
    non_blocking_rows = rows_from_sheet(wb["非阻断观察"])

    parse_12345 = parse_raw_work_id("12345")
    parse_y12345 = parse_raw_work_id("Y12345")
    parse_invalid = parse_raw_work_id("y12345")
    id_rule_ok = (
        parse_12345.normalized_raw == "12345"
        and parse_y12345.normalized_raw == "Y12345"
        and parse_12345.normalized_raw != parse_y12345.normalized_raw
        and derive_standard_work_id("12345") == "12345"
        and derive_standard_work_id("Y12345") == "12345"
        and derive_business_form("12345") == "有声版权"
        and derive_business_form("Y12345") == "有声成品"
        and derive_business_form("12345") == "有声版权"
        and not parse_invalid.valid
    )
    auth_not_business_form_ok = derive_business_form("12345") == "有声版权" and derive_business_form("Y12345") == "有声成品"

    formal_headers = headers(wb["正式导入阻断确认"])
    task_id_present = "任务ID" in formal_headers
    semantic_task_ids_ok = all(str(row.get("任务ID", "")).startswith("IMPORT-BLOCK::") for row in formal_rows)
    raw_ids = [row.get("原始作品ID") for row in formal_rows]
    formal_dedup_ok = len(raw_ids) == len(set(raw_ids))
    true_multi_name_issue_present = any("多名称作品ID" in str(row.get("问题类型", "")) for row in formal_rows)
    true_multi_auth_issue_present = any("多授权分类作品ID" in str(row.get("问题类型", "")) for row in formal_rows)
    invalid_issue_present = any("异常作品ID" in str(row.get("问题类型", "")) for row in formal_rows)

    cross_rows = [row for row in non_blocking_rows if row.get("观察类型") == "跨业务形态名称或授权分类差异"]
    cross_non_blocking_ok = bool(cross_rows) and all(
        row.get("纯数字原始ID") and row.get("Y前缀原始ID") and row.get("是否升级为阻断") in (None, "")
        for row in cross_rows
    )
    non_blocking_no_default_release_required = all(
        not (
            row.get("观察类型") == "跨业务形态名称或授权分类差异"
            and row.get("阻断解除状态") not in (None, "")
        )
        for row in non_blocking_rows
    )

    v22_kept = OLD_WORKBOOK_PATH.exists()
    sources_ok = sources_unchanged(payload) and payload["metrics"].get("source_unchanged") is True
    public_sensitive_matches = public_sensitive_scan()
    public_report_ok = public_sensitive_matches == 0
    no_sql = no_sql_files_generated()

    freeze_panes_ok = all(wb[sheet].freeze_panes for sheet in EXPECTED_SHEETS)
    filters_ok = all(wb[sheet].auto_filter.ref for sheet in EXPECTED_SHEETS if sheet != "确认进度总览")
    dropdowns_ok = all(
        len(wb[sheet].data_validations.dataValidation) > 0
        for sheet in EXPECTED_SHEETS
        if sheet != "确认进度总览"
    )
    complete_precision_hidden = True
    for sheet in EXPECTED_SHEETS:
        if sheet == "确认进度总览":
            continue
        ws = wb[sheet]
        for index, header in enumerate(headers(ws), start=1):
            if "完整精度" in header:
                letter = openpyxl.utils.get_column_letter(index)
                complete_precision_hidden = complete_precision_hidden and bool(ws.column_dimensions[letter].hidden or ws.column_dimensions[letter].width <= 3.5)

    overview = wb["确认进度总览"]
    nonblocking_completion_formula = str(overview.cell(18, 9).value or "")
    nonblocking_pending_formula = str(overview.cell(18, 4).value or "")
    nonblocking_unresolved_formula = str(overview.cell(18, 6).value or "")
    overview_nonblocking_formula_ok = (
        nonblocking_completion_formula == "=IF(B18=0,1,C18/B18)"
        and "COUNTIFS" in nonblocking_pending_formula
        and "是否升级为阻断" not in nonblocking_completion_formula
        and "COUNTIFS" in nonblocking_unresolved_formula
        and '"<>已解除"' in nonblocking_unresolved_formula
    )
    overview_completion_formulas_ok = all(
        str(overview.cell(row, 9).value or "") == f"=IF(B{row}=0,1,C{row}/B{row})"
        for row in range(13, 19)
    )

    validation = {
        "workbook_open": workbook_open,
        "expected_sheets_ok": expected_sheets_ok,
        "row_counts": row_counts,
        "id_rule_12345_y12345_distinct_raw_same_standard": id_rule_ok,
        "auth_category_does_not_override_business_form": auth_not_business_form_ok,
        "same_full_raw_id_multi_name_generates_issue": true_multi_name_issue_present,
        "same_full_raw_id_multi_auth_generates_issue": true_multi_auth_issue_present,
        "invalid_id_format_blocks": invalid_issue_present and not parse_invalid.valid,
        "formal_blockers_dedup_by_full_raw_id": formal_dedup_ok,
        "stable_task_ids_present": task_id_present and semantic_task_ids_ok,
        "cross_business_diff_is_non_blocking_observation": cross_non_blocking_ok,
        "cross_business_observation_no_default_release_required": non_blocking_no_default_release_required,
        "v22_workbook_kept": v22_kept,
        "source_files_unchanged": sources_ok,
        "public_report_no_sensitive_detail_patterns": public_report_ok,
        "public_sensitive_matches": public_sensitive_matches,
        "no_sql_files_or_migration_dirs_generated": no_sql,
        "freeze_panes_ok": freeze_panes_ok,
        "filters_ok": filters_ok,
        "dropdowns_ok": dropdowns_ok,
        "complete_precision_columns_hidden_or_narrow": complete_precision_hidden,
        "overview_nonblocking_formula_ok": overview_nonblocking_formula_ok,
        "overview_completion_formulas_ok": overview_completion_formulas_ok,
    }
    validation["overall"] = all(value for key, value in validation.items() if key not in {"row_counts", "public_sensitive_matches"})

    PUBLIC_VALIDATION_PATH.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    wb.close()
    if not validation["overall"]:
        raise SystemExit(json.dumps(validation, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
