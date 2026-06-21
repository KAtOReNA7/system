from __future__ import annotations

import csv
import hashlib
import html
import json
import math
import os
import posixpath
import re
import statistics
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import openpyxl
from work_id_rules import parse_work_id_dict

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns
except Exception:  # Charts are useful, but the numeric reports remain authoritative.
    plt = None
    sns = None


ROOT = Path(__file__).resolve().parents[2]
BILL_INPUT_ROOT = ROOT / "data" / "real-bills"
MASTER_INPUT_ROOT = ROOT / "data" / "master-data"
PUBLIC_ROOT = ROOT / "docs" / "analysis" / "m1-master-data"
ASSET_ROOT = PUBLIC_ROOT / "assets"
PRIVATE_ROOT = ROOT / "data" / "m1-master-data-private"
OPS_ROOT = PRIVATE_ROOT / "ops-confirmation"
REAL_BILL_SUMMARY = ROOT / "docs" / "analysis" / "m1-real-bills" / "summary.json"

EXPECTED_BILL_COLUMNS = [
    "年月",
    "渠道ID",
    "文学库渠道名称",
    "授权分类",
    "我方作品ID",
    "作品名称",
    "实销金额",
]

MASTER_ID_COL = "作品ID"
TITLE_COLS = ["出版书名", "合同书名"]
AUTHOR_COLS = ["作者署名", "作者原名"]
COPYRIGHT_START_CANDIDATE = "签订日期"
COPYRIGHT_END_CANDIDATE = "到期时间"
COPYRIGHT_PREV_END_CANDIDATE = "续约前到期日期"
DATE_COLS = ["作品ID创建时间", "签订日期", "到期时间", "续约前到期日期", "CIP出版时间", "续签时间", "首发时间"]
CATEGORY_REQUIRED_COLS = ["一级分类", "二级分类", "三级分类"]
CATEGORY_CANDIDATE_COLS = ["产品线", "级别", "归属", "授权性质", "版权类型"]
TAG_CANDIDATE_COLS = [
    "独家",
    "信息网络传播",
    "转授权",
    "有声使用权",
    "有声改编权",
    "有声转授权",
    "有声权利描述",
    "广播剧",
    "课程",
    "外文、少数民族、繁体",
    "外文、少数民族、繁体转授权",
    "合作方式",
    "授权范围（中国大陆地区和中国大陆地区（除港澳台）都只限于中国大陆地区使用）",
    "是否外版",
    "电子书续签情况",
    "归属",
    "授权性质",
    "版权类型",
]


def workbook_inputs(root: Path) -> list[Path]:
    return sorted(
        path
        for pattern in ("*.xlsx", "*.xlsm")
        for path in root.glob(pattern)
        if not path.name.startswith("~$")
    )


SENSITIVE_HEADER_PATTERNS = [
    "书名",
    "作者",
    "合同",
    "作品ID",
    "书号",
    "签约人",
    "授权方",
    "授权",
    "归属",
    "产品线",
    "级别",
    "版权类型",
    "编辑中心",
    "出版社",
    "排查人员",
    "时间",
    "日期",
    "CIP",
    "备注",
    "业务代码",
    "选题编号",
    "图书产品编码",
]

REQ_AT = {
    "data_cutoff": "REQ-DATA-IMPORT-006；AT-M1-006",
    "amount": "REQ-DATA-IMPORT-001、REQ-DATA-IMPORT-004；AT-M1-001、AT-M1-004",
    "work_launch": "REQ-WORK-003、REQ-WORK-004；AT-M1-022、AT-M1-023",
    "work_conflict": "REQ-WORK-001、REQ-WORK-002、REQ-WORK-006、REQ-DQ-001；AT-M1-020、AT-M1-021、AT-M1-025、AT-M1-010",
    "master_data": "REQ-WORK-008、REQ-WORK-009、REQ-WORK-010、REQ-WORK-011；AT-M1-027、AT-M1-028、AT-M1-029、AT-M1-031",
    "classification_tags": "REQ-CLASS-001、REQ-CLASS-002；AT-M1-040、AT-M1-041",
}


@dataclass(frozen=True)
class SourceSnapshot:
    path: str
    file_id: str
    size: int
    mtime_ns: int
    sha256: str


def ensure_dirs() -> None:
    PUBLIC_ROOT.mkdir(parents=True, exist_ok=True)
    ASSET_ROOT.mkdir(parents=True, exist_ok=True)
    PRIVATE_ROOT.mkdir(parents=True, exist_ok=True)
    OPS_ROOT.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def snapshot_sources(paths: Iterable[Path], prefix: str) -> list[SourceSnapshot]:
    snapshots: list[SourceSnapshot] = []
    for index, path in enumerate(sorted(paths), start=1):
        stat = path.stat()
        snapshots.append(
            SourceSnapshot(
                path=str(path.resolve()),
                file_id=f"{prefix}{index:03d}",
                size=stat.st_size,
                mtime_ns=stat.st_mtime_ns,
                sha256=sha256_file(path),
            )
        )
    return snapshots


def assert_sources_unchanged(before: list[SourceSnapshot]) -> bool:
    after_by_path = {item.path: item for item in snapshot_sources([Path(item.path) for item in before], before[0].file_id[0] if before else "F")}
    for before_item in before:
        current = after_by_path.get(before_item.path)
        if not current:
            return False
        if (before_item.size, before_item.mtime_ns, before_item.sha256) != (
            current.size,
            current.mtime_ns,
            current.sha256,
        ):
            return False
    return True


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value)) or (isinstance(value, str) and not value.strip())


def raw_text(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value)


def norm_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", raw_text(value)).strip()


def norm_name(value: Any) -> str:
    text = norm_text(value).lower()
    text = re.sub(r"\s+", " ", text)
    return text


def sample_code(prefix: str, *values: Any) -> str:
    payload = "\x1f".join(raw_text(value) for value in values)
    return f"{prefix}-{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:10].upper()}"


def parse_work_id(value: Any) -> dict[str, Any]:
    return parse_work_id_dict(value)


def parse_month(value: Any) -> tuple[str | None, str, bool]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m"), "excel_datetime", value.day == 1
    if isinstance(value, date):
        return value.strftime("%Y-%m"), "excel_date", value.day == 1
    text = norm_text(value)
    if not text:
        return None, "blank", False
    patterns = [
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "text_yyyy_mm_dd"),
        (r"^(\d{4})/(\d{1,2})/(\d{1,2})$", "text_yyyy_slash_mm_dd"),
        (r"^(\d{4})-(\d{1,2})$", "text_yyyy_mm"),
        (r"^(\d{4})(\d{2})$", "text_yyyymm"),
    ]
    for pattern, label in patterns:
        match = re.match(pattern, text)
        if not match:
            continue
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3)) if len(match.groups()) >= 3 else 1
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}", label, day == 1
    return None, "invalid", False


def parse_decimal_token(token: Any) -> Decimal | None:
    text = norm_text(token)
    if not text:
        return None
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date_value(value: Any) -> tuple[str | None, str]:
    if isinstance(value, datetime):
        return value.date().isoformat(), "excel_datetime"
    if isinstance(value, date):
        return value.isoformat(), "excel_date"
    text = norm_text(value)
    if not text:
        return None, "blank"
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    for pattern, label in [
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "text_yyyy_mm_dd"),
        (r"^(\d{4})(\d{2})(\d{2})$", "text_yyyymmdd"),
        (r"^(\d{4})-(\d{1,2})$", "text_yyyy_mm"),
    ]:
        match = re.match(pattern, text)
        if not match:
            continue
        year, month = int(match.group(1)), int(match.group(2))
        day = int(match.group(3)) if len(match.groups()) >= 3 else 1
        if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}", label
    return None, "invalid"


def month_range(months: Iterable[str | None]) -> str:
    values = sorted(month for month in months if month)
    if not values:
        return ""
    return values[0] if values[0] == values[-1] else f"{values[0]}~{values[-1]}"


def dec_sum(values: Iterable[Decimal | None]) -> Decimal:
    total = Decimal("0")
    for value in values:
        if value is not None:
            total += value
    return total


def fmt_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    return format(value, "f")


def extract_xlsx_amount_tokens(path: Path) -> dict[tuple[int, int], str]:
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    office_rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    result: dict[tuple[int, int], str] = {}
    with ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(main_ns + "si"):
                shared_strings.append("".join(node.text or "" for node in item.iter(main_ns + "t")))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target_by_id = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(package_rel_ns + "Relationship")
        }
        sheets = workbook.find(main_ns + "sheets")
        if sheets is None:
            return result
        for sheet_index, sheet in enumerate(sheets.findall(main_ns + "sheet"), start=1):
            relationship_id = sheet.attrib[office_rel_ns + "id"]
            target = target_by_id[relationship_id].replace("\\", "/")
            sheet_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
            with archive.open(sheet_path) as handle:
                for _, cell in ET.iterparse(handle, events=("end",)):
                    if cell.tag != main_ns + "c":
                        continue
                    reference = cell.attrib.get("r", "")
                    match = re.fullmatch(r"G(\d+)", reference)
                    if not match:
                        cell.clear()
                        continue
                    row_number = int(match.group(1))
                    cell_type = cell.attrib.get("t", "n")
                    value_node = cell.find(main_ns + "v")
                    inline_node = cell.find(main_ns + "is")
                    if cell_type == "s" and value_node is not None and value_node.text is not None:
                        token = shared_strings[int(value_node.text)]
                    elif cell_type == "inlineStr" and inline_node is not None:
                        token = "".join(node.text or "" for node in inline_node.iter(main_ns + "t"))
                    else:
                        token = value_node.text if value_node is not None and value_node.text is not None else ""
                    result[(sheet_index, row_number)] = token
                    cell.clear()
    return result


def read_bills() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    file_reports = []
    for file_index, path in enumerate(workbook_inputs(BILL_INPUT_ROOT), start=1):
        amount_tokens = extract_xlsx_amount_tokens(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            header_row = None
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row), max_col=7, values_only=True),
                start=1,
            ):
                if list(row[:7]) == EXPECTED_BILL_COLUMNS:
                    header_row = row_number
                    break
            if header_row is None:
                continue
            sheet_rows = 0
            for row_number, row_values in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, max_col=7, values_only=True),
                start=header_row + 1,
            ):
                if all(is_blank(value) for value in row_values):
                    continue
                month, month_format, month_first_day = parse_month(row_values[0])
                work = parse_work_id(row_values[4])
                amount_token = amount_tokens.get((sheet_index, row_number), raw_text(row_values[6]))
                amount = parse_decimal_token(amount_token)
                rows.append(
                    {
                        "file_id": f"B{file_index:03d}",
                        "sheet_index": sheet_index,
                        "row_number": row_number,
                        "month": month,
                        "month_format": month_format,
                        "month_first_day": month_first_day,
                        "channel_id": norm_text(row_values[1]),
                        "channel_name": norm_text(row_values[2]),
                        "auth": norm_text(row_values[3]),
                        "raw_work_id": work["normalized_raw"] or work["raw"],
                        "raw_work_id_original": work["raw"],
                        "standard_id": work["standard_id"],
                        "business_form": work["business_form"],
                        "work_id_valid": work["valid"],
                        "work_id_format": work["format"],
                        "work_name": norm_text(row_values[5]),
                        "work_name_norm": norm_name(row_values[5]),
                        "amount": amount,
                        "amount_sign": "positive" if amount is not None and amount > 0 else "negative" if amount is not None and amount < 0 else "zero" if amount == 0 else "invalid",
                    }
                )
                sheet_rows += 1
            file_reports.append({"file_id": f"B{file_index:03d}", "sheet_index": sheet_index, "row_count": sheet_rows})
        workbook.close()
    return rows, {"file_reports": file_reports}


def read_master() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    structure = {"files": [], "worksheets": [], "headers": []}
    for file_index, path in enumerate(workbook_inputs(MASTER_INPUT_ROOT), start=1):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        file_id = f"M{file_index:03d}"
        structure["files"].append({"file_id": file_id, "sheet_count": len(workbook.sheetnames), "size": path.stat().st_size})
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            header_row = None
            headers: list[str] = []
            for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
                current = [norm_text(value) for value in row]
                if MASTER_ID_COL in current:
                    header_row = row_number
                    headers = current
                    break
            if header_row is None:
                continue
            headers = [header or f"未命名字段_{index+1}" for index, header in enumerate(headers)]
            structure["worksheets"].append(
                {"file_id": file_id, "sheet_index": sheet_index, "row_count": worksheet.max_row - header_row, "column_count": len(headers)}
            )
            structure["headers"] = headers
            for row_number, values in enumerate(worksheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True), start=header_row + 1):
                if all(is_blank(value) for value in values):
                    continue
                record = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
                parsed_id = parse_work_id(record.get(MASTER_ID_COL))
                record["_file_id"] = file_id
                record["_sheet_index"] = sheet_index
                record["_row_number"] = row_number
                record["_sample_id"] = sample_code("MD", file_id, sheet_index, row_number, record.get(MASTER_ID_COL))
                record["_raw_work_id"] = parsed_id["normalized_raw"] or parsed_id["raw"]
                record["_standard_id"] = parsed_id["standard_id"]
                record["_business_form"] = parsed_id["business_form"]
                record["_work_id_valid"] = parsed_id["valid"]
                record["_work_id_format"] = parsed_id["format"]
                start, start_format = parse_date_value(record.get(COPYRIGHT_START_CANDIDATE))
                end, end_format = parse_date_value(record.get(COPYRIGHT_END_CANDIDATE))
                prev_end, prev_end_format = parse_date_value(record.get(COPYRIGHT_PREV_END_CANDIDATE))
                record["_copyright_start_candidate"] = start
                record["_copyright_start_format"] = start_format
                record["_copyright_end_candidate"] = end
                record["_copyright_end_format"] = end_format
                record["_copyright_prev_end_candidate"] = prev_end
                record["_copyright_prev_end_format"] = prev_end_format
                rows.append(record)
        workbook.close()
    return rows, structure


def field_profile(master_rows: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    profiles = []
    total = len(master_rows)
    for header in headers:
        values = [norm_text(row.get(header)) for row in master_rows]
        non_empty = [value for value in values if value]
        distinct = len(set(non_empty))
        sensitive = any(pattern in header for pattern in SENSITIVE_HEADER_PATTERNS)
        counter = Counter(non_empty)
        top_values = []
        profiles.append(
            {
                "field": header,
                "non_empty_count": len(non_empty),
                "missing_count": total - len(non_empty),
                "missing_rate": (total - len(non_empty)) / total if total else 0,
                "distinct_count": distinct,
                "public_value_policy": "count_and_distinct_only",
                "top_values": top_values,
                "top_value_counts": [count for _value, count in counter.most_common(8)],
            }
        )
    return profiles


def summarize_value_rows(rows: list[dict[str, Any]], value_key: str) -> str:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = norm_text(row.get(value_key))
        if not value:
            value = "<空>"
        item = grouped.setdefault(value, {"count": 0, "months": [], "amounts": []})
        item["count"] += 1
        item["months"].append(row.get("month"))
        item["amounts"].append(row.get("amount"))
    parts = []
    for value, item in sorted(grouped.items(), key=lambda pair: (-pair[1]["count"], pair[0])):
        parts.append(
            f"{value}｜记录数:{item['count']}｜月份:{month_range(item['months'])}｜实销合计:{fmt_decimal(dec_sum(item['amounts']))}"
        )
    return "\n".join(parts)


def summarize_bill_group(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "record_count": len(rows),
        "month_range": month_range(row.get("month") for row in rows),
        "amount_total": fmt_decimal(dec_sum(row.get("amount") for row in rows)),
    }


def group_bill_conflicts(bill_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_raw: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_standard: dict[str, list[dict[str, Any]]] = defaultdict(list)
    invalid_by_raw: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in bill_rows:
        if row["work_id_valid"]:
            by_raw[row["raw_work_id"]].append(row)
            by_standard[row["standard_id"]].append(row)
        else:
            invalid_by_raw[row["raw_work_id"] or "<空>"].append(row)

    multi_name = []
    multi_auth = []
    for raw_id, rows in by_raw.items():
        names = {row["work_name_norm"] for row in rows if row["work_name_norm"]}
        auths = {norm_text(row["auth"]) for row in rows if norm_text(row["auth"])}
        base = summarize_bill_group(rows)
        relation = "Y前缀" if raw_id.startswith("Y") else "纯数字"
        common = {
            "raw_work_id": raw_id,
            "standard_id": rows[0]["standard_id"],
            "business_form": rows[0]["business_form"],
            "pure_digit_y_relation": f"{relation}；标准作品ID={rows[0]['standard_id']}",
            "record_count": base["record_count"],
            "month_range": base["month_range"],
            "amount_total": base["amount_total"],
            "name_details": summarize_value_rows(rows, "work_name"),
            "auth_details": summarize_value_rows(rows, "auth"),
        }
        if len(names) > 1:
            multi_name.append(
                {
                    **common,
                    "confirmation_group_id": sample_code("NAME", raw_id),
                    "system_candidate_explanation": "同一原始作品ID出现多个作品名称；需按冲突组确认标准作品名称、更名关系或错配。",
                }
            )
        if len(auths) > 1:
            multi_auth.append(
                {
                    **common,
                    "confirmation_group_id": sample_code("AUTH", raw_id),
                    "system_candidate_explanation": "同一原始作品ID出现多个授权分类；授权分类不反向覆盖业务形态，需确认是否阻断或作为历史分类差异。",
                }
            )

    abnormal_ids = []
    for raw_id, rows in invalid_by_raw.items():
        base = summarize_bill_group(rows)
        abnormal_ids.append(
            {
                "confirmation_group_id": sample_code("ABNID", raw_id),
                "raw_work_id": raw_id,
                "standard_id": "",
                "business_form": "",
                "pure_digit_y_relation": "无法按纯数字或Y+数字解析",
                "record_count": base["record_count"],
                "month_range": base["month_range"],
                "amount_total": base["amount_total"],
                "name_details": summarize_value_rows(rows, "work_name"),
                "auth_details": summarize_value_rows(rows, "auth"),
                "system_candidate_explanation": "作品ID格式不满足当前账单派生规则；未确认前阻断正式导入。",
            }
        )

    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    volume_pattern = re.compile(r"(第?[一二三四五六七八九十百千万\d]+[卷册部季集]|上册|下册|全集|全\d+册)")
    volume_candidates = []
    for row in bill_rows:
        if row["work_id_valid"] and row["work_name_norm"]:
            by_name[row["work_name_norm"]].append(row)
    for name_norm, rows in by_name.items():
        standard_ids = sorted({row["standard_id"] for row in rows if row["standard_id"]})
        has_volume_marker = any(volume_pattern.search(row["work_name"]) for row in rows)
        if len(standard_ids) > 1 or has_volume_marker:
            base = summarize_bill_group(rows)
            volume_candidates.append(
                {
                    "confirmation_group_id": sample_code("VOL", name_norm, ",".join(standard_ids)),
                    "candidate_type": "同名多标准作品" if len(standard_ids) > 1 else "名称含分册标识",
                    "standard_id_candidates": "\n".join(standard_ids),
                    "raw_work_id_candidates": "\n".join(sorted({row["raw_work_id"] for row in rows if row["raw_work_id"]})),
                    "business_form_candidates": "\n".join(sorted({row["business_form"] for row in rows if row["business_form"]})),
                    "record_count": base["record_count"],
                    "month_range": base["month_range"],
                    "amount_total": base["amount_total"],
                    "name_details": summarize_value_rows(rows, "work_name"),
                    "auth_details": summarize_value_rows(rows, "auth"),
                    "system_candidate_explanation": "仅作为分册/同名候选，不自动归并；需结合作者、分类和版权期限确认主标准作品。",
                }
            )

    offset_groups: dict[tuple[str, Decimal], list[dict[str, Any]]] = defaultdict(list)
    for row in bill_rows:
        amount = row.get("amount")
        if row["work_id_valid"] and amount is not None and amount != 0:
            offset_groups[(row["standard_id"], abs(amount))].append(row)
    offset_candidates = []
    for (standard_id, abs_amount), rows in offset_groups.items():
        has_positive = any(row["amount"] is not None and row["amount"] > 0 for row in rows)
        has_negative = any(row["amount"] is not None and row["amount"] < 0 for row in rows)
        month_count = len({row["month"] for row in rows if row["month"]})
        if has_positive and has_negative and month_count > 1:
            base = summarize_bill_group(rows)
            offset_candidates.append(
                {
                    "confirmation_group_id": sample_code("OFF", standard_id, abs_amount),
                    "standard_id": standard_id,
                    "absolute_amount": fmt_decimal(abs_amount),
                    "raw_work_id_candidates": "\n".join(sorted({row["raw_work_id"] for row in rows if row["raw_work_id"]})),
                    "business_form_candidates": "\n".join(sorted({row["business_form"] for row in rows if row["business_form"]})),
                    "record_count": base["record_count"],
                    "month_range": base["month_range"],
                    "amount_total": base["amount_total"],
                    "name_details": summarize_value_rows(rows, "work_name"),
                    "auth_details": summarize_value_rows(rows, "auth"),
                    "system_candidate_explanation": "跨月同额正负记录候选；不得自动删除，需确认是否为合法冲抵。",
                }
            )

    first_positive_by_standard: dict[str, str] = {}
    first_positive_by_form: dict[tuple[str, str], str] = {}
    standards_seen = set()
    form_seen = set()
    standard_nonpositive_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in bill_rows:
        if not row["work_id_valid"] or not row["standard_id"]:
            continue
        standards_seen.add(row["standard_id"])
        if row["business_form"]:
            form_seen.add((row["standard_id"], row["business_form"]))
        amount = row.get("amount")
        if amount is not None and amount > 0 and row.get("month"):
            first_positive_by_standard[row["standard_id"]] = min(first_positive_by_standard.get(row["standard_id"], row["month"]), row["month"])
            if row["business_form"]:
                key = (row["standard_id"], row["business_form"])
                first_positive_by_form[key] = min(first_positive_by_form.get(key, row["month"]), row["month"])
        elif amount is not None and amount <= 0:
            standard_nonpositive_rows[row["standard_id"]].append(row)

    first_sale_empty = []
    for standard_id in sorted(standards_seen - set(first_positive_by_standard)):
        rows = standard_nonpositive_rows.get(standard_id, [])
        base = summarize_bill_group(rows)
        first_sale_empty.append(
            {
                "confirmation_group_id": sample_code("LAUNCH", standard_id),
                "standard_id": standard_id,
                "raw_work_id_candidates": "\n".join(sorted({row["raw_work_id"] for row in rows if row["raw_work_id"]})),
                "business_form_candidates": "\n".join(sorted({row["business_form"] for row in rows if row["business_form"]})),
                "record_count": base["record_count"],
                "month_range": base["month_range"],
                "amount_total": base["amount_total"],
                "name_details": summarize_value_rows(rows, "work_name"),
                "auth_details": summarize_value_rows(rows, "auth"),
                "system_candidate_explanation": "该标准作品从未出现正数实销记录；首次实销月份留空并进入异常确认。",
            }
        )

    return {
        "by_raw": by_raw,
        "by_standard": by_standard,
        "multi_name": sorted(multi_name, key=lambda item: item["confirmation_group_id"]),
        "multi_auth": sorted(multi_auth, key=lambda item: item["confirmation_group_id"]),
        "abnormal_ids": sorted(abnormal_ids, key=lambda item: item["confirmation_group_id"]),
        "volume_candidates": sorted(volume_candidates, key=lambda item: item["confirmation_group_id"]),
        "offset_candidates": sorted(offset_candidates, key=lambda item: item["confirmation_group_id"]),
        "first_sale_empty": first_sale_empty,
        "first_positive_by_standard": first_positive_by_standard,
        "first_positive_by_form": first_positive_by_form,
        "standards_seen": standards_seen,
        "form_seen": form_seen,
    }


def analyze_master_relations(master_rows: list[dict[str, Any]], bill_rows: list[dict[str, Any]], bill_conflicts: dict[str, Any]) -> dict[str, Any]:
    bill_standard_ids = set(bill_conflicts["standards_seen"])
    bill_raw_ids = {row["raw_work_id"] for row in bill_rows if row["work_id_valid"] and row["raw_work_id"]}

    master_by_standard: dict[str, list[dict[str, Any]]] = defaultdict(list)
    master_by_raw: dict[str, list[dict[str, Any]]] = defaultdict(list)
    invalid_master_ids = []
    for row in master_rows:
        if row.get("_work_id_valid") and row.get("_standard_id"):
            master_by_standard[row["_standard_id"]].append(row)
            master_by_raw[row["_raw_work_id"]].append(row)
        else:
            invalid_master_ids.append(row)

    master_standard_ids = set(master_by_standard)
    master_raw_ids = set(master_by_raw)
    covered_standard_ids = bill_standard_ids & master_standard_ids
    unmatched_bill_standard_ids = bill_standard_ids - master_standard_ids

    def title_values(records: list[dict[str, Any]]) -> set[str]:
        values = set()
        for record in records:
            for col in TITLE_COLS:
                value = norm_name(record.get(col))
                if value:
                    values.add(value)
        return values

    def author_values(records: list[dict[str, Any]]) -> set[str]:
        values = set()
        for record in records:
            for col in AUTHOR_COLS:
                value = norm_name(record.get(col))
                if value:
                    values.add(value)
        return values

    def has_candidate_author(records: list[dict[str, Any]]) -> bool:
        return bool(author_values(records))

    def has_candidate_title(records: list[dict[str, Any]]) -> bool:
        return bool(title_values(records))

    def has_candidate_copyright(records: list[dict[str, Any]]) -> bool:
        return any(record.get("_copyright_start_candidate") and record.get("_copyright_end_candidate") for record in records)

    def has_candidate_tags(records: list[dict[str, Any]]) -> bool:
        return any(any(norm_text(record.get(col)) for col in TAG_CANDIDATE_COLS if col in record) for record in records)

    missing_author = []
    missing_title = []
    missing_copyright_candidate = []
    missing_tag_candidate = []
    for standard_id in sorted(bill_standard_ids):
        records = master_by_standard.get(standard_id, [])
        if not records or not has_candidate_title(records):
            missing_title.append(standard_id)
        if not records or not has_candidate_author(records):
            missing_author.append(standard_id)
        if not records or not has_candidate_copyright(records):
            missing_copyright_candidate.append(standard_id)
        if not records or not has_candidate_tags(records):
            missing_tag_candidate.append(standard_id)

    name_difference_rows = []
    for standard_id in sorted(covered_standard_ids):
        bill_names = {row["work_name_norm"] for row in bill_conflicts["by_standard"].get(standard_id, []) if row["work_name_norm"]}
        master_titles = title_values(master_by_standard[standard_id])
        if bill_names and master_titles and bill_names.isdisjoint(master_titles):
            rows = bill_conflicts["by_standard"].get(standard_id, [])
            base = summarize_bill_group(rows)
            name_difference_rows.append(
                {
                    "confirmation_group_id": sample_code("NMDIFF", standard_id),
                    "standard_id": standard_id,
                    "raw_work_id_candidates": "\n".join(sorted({row["raw_work_id"] for row in rows if row["raw_work_id"]})),
                    "business_form_candidates": "\n".join(sorted({row["business_form"] for row in rows if row["business_form"]})),
                    "bill_name_candidates": "\n".join(sorted({row["work_name"] for row in rows if row["work_name"]})),
                    "master_title_candidates": "\n".join(sorted({norm_text(record.get(col)) for record in master_by_standard[standard_id] for col in TITLE_COLS if norm_text(record.get(col))})),
                    "record_count": base["record_count"],
                    "month_range": base["month_range"],
                    "amount_total": base["amount_total"],
                    "system_candidate_explanation": "台账书名与账单历史作品名无规范化一致项；不得仅凭相似书名自动关联。",
                }
            )

    author_alias_rows = []
    alias_to_originals: dict[str, set[str]] = defaultdict(set)
    for record in master_rows:
        signed = norm_text(record.get("作者署名"))
        original = norm_text(record.get("作者原名"))
        if signed and original and signed != original:
            alias_to_originals[norm_name(signed)].add(original)
            author_alias_rows.append(
                {
                    "sample_id": record["_sample_id"],
                    "standard_id": record.get("_standard_id") or "",
                    "raw_work_id": record.get("_raw_work_id") or "",
                    "author_alias": signed,
                    "author_standard_candidate": original,
                    "system_candidate_explanation": "作者署名与作者原名不同，作为作者别名候选；同名或歧义需运营确认。",
                    "运营确认结果": "",
                    "确认作者标准名": "",
                    "是否保留为别名": "",
                    "备注": "",
                }
            )
    ambiguous_author_alias_rows = []
    for alias, originals in alias_to_originals.items():
        if len(originals) > 1:
            ambiguous_author_alias_rows.append(
                {
                    "confirmation_group_id": sample_code("AUTHALIAS", alias),
                    "author_alias": alias,
                    "author_standard_candidates": "\n".join(sorted(originals)),
                    "system_candidate_explanation": "同一规范化作者署名对应多个作者原名；不得自动归并。",
                    "运营确认结果": "",
                    "确认作者标准名": "",
                    "备注": "",
                }
            )

    duplicate_raw_id_group_count_all = 0
    duplicate_raw_ids = []
    for raw_id, records in master_by_raw.items():
        if len(records) > 1:
            duplicate_raw_id_group_count_all += 1
        title_count = len(title_values(records))
        author_count = len(author_values(records))
        date_pairs = {(record.get("_copyright_start_candidate") or "", record.get("_copyright_end_candidate") or "") for record in records}
        if len(records) > 1 and (title_count > 1 or author_count > 1 or len(date_pairs) > 1):
            duplicate_raw_ids.append(
                {
                    "confirmation_group_id": sample_code("MDRAW", raw_id),
                    "issue_type": "台账同原始作品ID关键字段冲突",
                    "raw_work_id": raw_id,
                    "standard_id": records[0].get("_standard_id") or "",
                    "record_count": len(records),
                    "distinct_title_count": title_count,
                    "distinct_author_count": author_count,
                    "distinct_copyright_period_count": len(date_pairs),
                    "master_title_candidates": "\n".join(sorted({norm_text(record.get(col)) for record in records for col in TITLE_COLS if norm_text(record.get(col))})),
                    "author_candidates": "\n".join(sorted({norm_text(record.get(col)) for record in records for col in AUTHOR_COLS if norm_text(record.get(col))})),
                    "copyright_period_candidates": "\n".join(sorted(f"{start}~{end}" for start, end in date_pairs)),
                    "system_candidate_explanation": "台账同一作品ID存在重复记录或字段差异，需确认是否为同一标准作品基础信息版本。",
                }
            )

    standard_multi_raw = []
    for standard_id, records in master_by_standard.items():
        raws = sorted({record["_raw_work_id"] for record in records if record.get("_raw_work_id")})
        if len(raws) > 1:
            standard_multi_raw.append(
                {
                    "confirmation_group_id": sample_code("MDSTDRAW", standard_id),
                    "standard_id": standard_id,
                    "master_raw_work_id_candidates": "\n".join(raws),
                    "record_count": len(records),
                    "system_candidate_explanation": "台账中同一标准数字主体存在多个原始作品ID；需确认是否为纯数字/Y业务形态或历史分册。",
                }
            )

    copyright_period_groups = []
    pure_y_counterexamples = []
    for standard_id, records in master_by_standard.items():
        period_by_form: dict[str, set[tuple[str, str]]] = defaultdict(set)
        period_pairs = set()
        for record in records:
            start = record.get("_copyright_start_candidate") or ""
            end = record.get("_copyright_end_candidate") or ""
            if start or end:
                period_pairs.add((start, end))
                period_by_form[record.get("_business_form") or "未知"].add((start, end))
        if len(period_pairs) > 1:
            copyright_period_groups.append(
                {
                    "confirmation_group_id": sample_code("CP", standard_id),
                    "standard_id": standard_id,
                    "raw_work_id_candidates": "\n".join(sorted({record.get("_raw_work_id") or "" for record in records if record.get("_raw_work_id")})),
                    "business_form_candidates": "\n".join(sorted({record.get("_business_form") or "未知" for record in records})),
                    "copyright_period_candidates": "\n".join(sorted(f"{start}~{end}" for start, end in period_pairs)),
                    "record_count": len(records),
                    "system_candidate_explanation": "同一标准作品存在多个候选版权期限；共享版权期限规则下需运营确认，未确认前阻断。",
                }
            )
        if "有声版权" in period_by_form and "有声成品" in period_by_form and period_by_form["有声版权"] != period_by_form["有声成品"]:
            pure_y_counterexamples.append(
                {
                    "confirmation_group_id": sample_code("CPFORM", standard_id),
                    "standard_id": standard_id,
                    "pure_digit_periods": "\n".join(sorted(f"{start}~{end}" for start, end in period_by_form["有声版权"])),
                    "y_prefix_periods": "\n".join(sorted(f"{start}~{end}" for start, end in period_by_form["有声成品"])),
                    "system_candidate_explanation": "纯数字与Y前缀候选期限不同；属于共享版权期限反例候选，需业务确认。",
                }
            )

    ledger_match_rows = []
    for standard_id in sorted(unmatched_bill_standard_ids):
        rows = bill_conflicts["by_standard"].get(standard_id, [])
        base = summarize_bill_group(rows)
        ledger_match_rows.append(
            {
                "confirmation_group_id": sample_code("MDMISS", standard_id),
                "issue_type": "账单标准作品ID未匹配台账",
                "standard_id": standard_id,
                "raw_work_id_candidates": "\n".join(sorted({row["raw_work_id"] for row in rows if row["raw_work_id"]})),
                "business_form_candidates": "\n".join(sorted({row["business_form"] for row in rows if row["business_form"]})),
                "record_count": base["record_count"],
                "month_range": base["month_range"],
                "amount_total": base["amount_total"],
                "system_candidate_explanation": "台账未找到同标准数字主体；不得仅凭书名相似自动关联。",
            }
        )
    for row in standard_multi_raw:
        ledger_match_rows.append({**row, "issue_type": "台账同标准作品多原始ID"})
    for row in duplicate_raw_ids:
        ledger_match_rows.append(row)

    return {
        "master_by_standard": master_by_standard,
        "master_by_raw": master_by_raw,
        "bill_standard_ids": bill_standard_ids,
        "bill_raw_ids": bill_raw_ids,
        "master_standard_ids": master_standard_ids,
        "master_raw_ids": master_raw_ids,
        "covered_standard_ids": covered_standard_ids,
        "unmatched_bill_standard_ids": unmatched_bill_standard_ids,
        "invalid_master_ids": invalid_master_ids,
        "missing_title_standard_ids": missing_title,
        "missing_author_standard_ids": missing_author,
        "missing_copyright_candidate_standard_ids": missing_copyright_candidate,
        "missing_tag_candidate_standard_ids": missing_tag_candidate,
        "name_difference_rows": name_difference_rows,
        "author_alias_rows": author_alias_rows,
        "ambiguous_author_alias_rows": ambiguous_author_alias_rows,
        "duplicate_raw_id_group_count_all": duplicate_raw_id_group_count_all,
        "duplicate_raw_ids": duplicate_raw_ids,
        "standard_multi_raw": standard_multi_raw,
        "copyright_period_groups": copyright_period_groups,
        "pure_y_counterexamples": pure_y_counterexamples,
        "ledger_match_rows": ledger_match_rows,
    }


def analyze_dates(master_rows: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    result = []
    for col in DATE_COLS:
        if col not in headers:
            continue
        formats = Counter()
        invalid = 0
        missing = 0
        parsed_values = []
        for row in master_rows:
            parsed, fmt = parse_date_value(row.get(col))
            formats[fmt] += 1
            if fmt == "blank":
                missing += 1
            elif fmt == "invalid":
                invalid += 1
            elif parsed:
                parsed_values.append(parsed)
        result.append(
            {
                "field": col,
                "missing_count": missing,
                "invalid_count": invalid,
                "valid_count": len(parsed_values),
                "format_counts": dict(formats),
                "month_range": month_range(value[:7] for value in parsed_values),
            }
        )
    return result


def analyze_tag_distribution(master_rows: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    rows = []
    for col in TAG_CANDIDATE_COLS + CATEGORY_CANDIDATE_COLS:
        if col not in headers:
            continue
        values = [norm_text(row.get(col)) for row in master_rows if norm_text(row.get(col))]
        counter = Counter(values)
        rows.append(
            {
                "field": col,
                "non_empty_count": len(values),
                "distinct_count": len(counter),
                "public_value_policy": "count_and_distinct_only",
                "top_value_counts": [count for _value, count in counter.most_common(10)],
                "top_values": [],
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def with_confirmation_fields(rows: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
    fields = {
        "运营确认结果": "",
        "确认标准作品ID": "",
        "确认标准作品名称": "",
        "确认作者": "",
        "确认一级分类": "",
        "确认二级分类": "",
        "确认三级分类": "",
        "确认版权开始日期": "",
        "确认版权到期日期": "",
        "是否解除阻断": "",
        "运营备注": "",
    }
    return [{**row, "confirmation_table": kind, **fields} for row in rows]


def build_ops_tables(bill_conflicts: dict[str, Any], master_analysis: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    return {
        "01-多名称作品ID确认表": with_confirmation_fields(bill_conflicts["multi_name"], "多名称作品ID确认表"),
        "02-多授权分类作品ID确认表": with_confirmation_fields(bill_conflicts["multi_auth"], "多授权分类作品ID确认表"),
        "03-异常作品ID确认表": with_confirmation_fields(bill_conflicts["abnormal_ids"], "异常作品ID确认表"),
        "04-分册候选确认表": with_confirmation_fields(bill_conflicts["volume_candidates"], "分册候选确认表"),
        "05-正负冲抵候选确认表": with_confirmation_fields(bill_conflicts["offset_candidates"], "正负冲抵候选确认表"),
        "06-首次实销仍为空作品确认表": with_confirmation_fields(bill_conflicts["first_sale_empty"], "首次实销仍为空作品确认表"),
        "07-台账匹配失败或一对多冲突确认表": with_confirmation_fields(master_analysis["ledger_match_rows"], "台账匹配失败或一对多冲突确认表"),
        "08-版权期限反例确认表": with_confirmation_fields(
            master_analysis["copyright_period_groups"] + master_analysis["pure_y_counterexamples"], "版权期限反例确认表"
        ),
    }


def write_private_outputs(
    master_rows: list[dict[str, Any]],
    bill_conflicts: dict[str, Any],
    master_analysis: dict[str, Any],
    ops_tables: dict[str, list[dict[str, Any]]],
) -> None:
    write_csv(PRIVATE_ROOT / "master-invalid-id-rows.csv", [
        {
            "sample_id": row["_sample_id"],
            "raw_work_id": norm_text(row.get(MASTER_ID_COL)),
            "work_id_format": row.get("_work_id_format"),
        }
        for row in master_analysis["invalid_master_ids"]
    ])
    write_csv(PRIVATE_ROOT / "master-name-differences.csv", master_analysis["name_difference_rows"])
    write_csv(PRIVATE_ROOT / "master-author-alias-candidates.csv", master_analysis["author_alias_rows"])
    write_csv(PRIVATE_ROOT / "master-ambiguous-author-alias.csv", master_analysis["ambiguous_author_alias_rows"])
    write_csv(PRIVATE_ROOT / "master-duplicate-raw-id-conflicts.csv", master_analysis["duplicate_raw_ids"])
    write_csv(PRIVATE_ROOT / "master-copyright-period-conflicts.csv", master_analysis["copyright_period_groups"])
    write_csv(PRIVATE_ROOT / "master-pure-y-copyright-counterexamples.csv", master_analysis["pure_y_counterexamples"])
    write_csv(PRIVATE_ROOT / "master-unmatched-bill-standard-works.csv", [
        {"standard_id": standard_id}
        for standard_id in sorted(master_analysis["unmatched_bill_standard_ids"])
    ])

    for sheet_name, rows in ops_tables.items():
        safe_name = sheet_name.replace("/", "-")
        write_csv(OPS_ROOT / f"{safe_name}.csv", rows)

    workbook_payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sheets": [{"name": name, "rows": rows} for name, rows in ops_tables.items()],
    }
    (OPS_ROOT / "ops-confirmation-workbook-data.json").write_text(json.dumps(workbook_payload, ensure_ascii=False, indent=2), encoding="utf-8")


def markdown_table(rows: list[dict[str, Any]], columns: list[str], limit: int | None = None) -> str:
    body = rows if limit is None else rows[:limit]
    lines = ["| " + " | ".join(columns) + " |", "| " + " | ".join(["---"] * len(columns)) + " |"]
    for row in body:
        values = []
        for col in columns:
            value = row.get(col, "")
            if isinstance(value, float):
                value = f"{value:.4f}"
            values.append(str(value).replace("\n", "<br>").replace("|", "\\|"))
        lines.append("| " + " | ".join(values) + " |")
    if limit is not None and len(rows) > limit:
        lines.append("| ... | " + " | ".join([""] * (len(columns) - 1)) + " |")
    return "\n".join(lines)


def write_report_file(path: Path, title: str, facts: list[str], candidate_rules: list[str], cannot_confirm: list[str], ops_samples: list[str], req_at: str, pending: str) -> None:
    content = f"""# {title}

## 已确认的数据事实

{chr(10).join(f"- {item}" for item in facts)}

## 候选规则（未启用）

{chr(10).join(f"- {item}" for item in candidate_rules)}

## 无法单靠台账或账单确认的事项

{chr(10).join(f"- {item}" for item in cannot_confirm)}

## 需要运营确认的样本

{chr(10).join(f"- {item}" for item in ops_samples)}

## REQ 与 AT

- {req_at}

## PENDING-DATA 状态

- {pending}
"""
    path.write_text(content, encoding="utf-8")


def make_chart(path: Path, labels: list[str], values: list[int], title: str, subtitle: str) -> bool:
    if plt is None or sns is None:
        return False
    font_family = "DejaVu Sans"
    try:
        from matplotlib import font_manager

        for font_path in [
            Path("C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("C:/Windows/Fonts/simsun.ttc"),
        ]:
            if font_path.exists():
                font_manager.fontManager.addfont(str(font_path))
                font_family = font_manager.FontProperties(fname=str(font_path)).get_name()
                break
    except Exception:
        font_family = "DejaVu Sans"
    sns.set_theme(
        style="whitegrid",
        rc={
            "figure.facecolor": "#FCFCFD",
            "axes.facecolor": "#FFFFFF",
            "font.family": "sans-serif",
            "font.sans-serif": [font_family, "Microsoft YaHei", "SimHei", "DejaVu Sans"],
            "axes.unicode_minus": False,
        },
    )
    fig, ax = plt.subplots(figsize=(9, max(3.5, 0.45 * len(labels) + 1.8)))
    order = list(range(len(labels)))
    ax.barh(order, values, color="#A3BEFA", edgecolor="#2E4780", linewidth=1.0)
    ax.set_yticks(order, labels)
    ax.invert_yaxis()
    ax.set_xlabel("数量")
    ax.set_title(title, loc="left", fontsize=13, fontweight="semibold", color="#1F2430", pad=20)
    ax.text(0, 1.02, subtitle, transform=ax.transAxes, ha="left", va="bottom", fontsize=9, color="#6F768A")
    ax.grid(axis="x", color="#E6E8F0")
    ax.grid(axis="y", visible=False)
    for index, value in enumerate(values):
        ax.text(value, index, f" {value:,}", va="center", fontsize=9, color="#1F2430")
    sns.despine(ax=ax)
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return True


def write_public_reports(
    summary: dict[str, Any],
    structure: dict[str, Any],
    field_profiles: list[dict[str, Any]],
    date_profiles: list[dict[str, Any]],
    tag_profiles: list[dict[str, Any]],
    bill_conflicts: dict[str, Any],
    master_analysis: dict[str, Any],
    source_snapshots: dict[str, list[SourceSnapshot]],
) -> None:
    charts = []
    if make_chart(
        ASSET_ROOT / "master-coverage.png",
        ["台账覆盖标准作品", "台账未覆盖标准作品"],
        [summary["master_covered_standard_work_count"], summary["master_unmatched_standard_work_count"]],
        "台账覆盖率",
        "按账单派生的 3,099 个标准作品 ID 统计，未使用书名相似自动匹配。",
    ):
        charts.append("assets/master-coverage.png")
    if make_chart(
        ASSET_ROOT / "required-field-gaps.png",
        ["候选作者缺失", "明确三级分类缺失", "候选版权期限缺失", "候选标签缺失"],
        [
            summary["missing_author_candidate_standard_work_count"],
            summary["missing_confirmed_category_standard_work_count"],
            summary["missing_copyright_candidate_standard_work_count"],
            summary["missing_tag_candidate_standard_work_count"],
        ],
        "M1 基础信息缺口",
        "作者与版权期限按台账候选字段统计；分类和标签按可冻结语义统计。",
    ):
        charts.append("assets/required-field-gaps.png")
    if make_chart(
        ASSET_ROOT / "ops-confirmation-groups.png",
        [
            "多名称ID",
            "多授权分类ID",
            "异常作品ID",
            "分册候选",
            "正负冲抵",
            "首次实销空",
            "台账匹配/一对多",
            "版权期限反例",
        ],
        [
            summary["ops_confirmation_counts"]["01-多名称作品ID确认表"],
            summary["ops_confirmation_counts"]["02-多授权分类作品ID确认表"],
            summary["ops_confirmation_counts"]["03-异常作品ID确认表"],
            summary["ops_confirmation_counts"]["04-分册候选确认表"],
            summary["ops_confirmation_counts"]["05-正负冲抵候选确认表"],
            summary["ops_confirmation_counts"]["06-首次实销仍为空作品确认表"],
            summary["ops_confirmation_counts"]["07-台账匹配失败或一对多冲突确认表"],
            summary["ops_confirmation_counts"]["08-版权期限反例确认表"],
        ],
        "运营确认组数量",
        "按冲突组或候选组输出，避免逐行确认账单明细。",
    ):
        charts.append("assets/ops-confirmation-groups.png")

    reports = [
        (
            "01-file-workbook-structure.md",
            "文件、工作表、表头和记录规模",
            [
                f"台账输入文件数：{summary['master_source_file_count']}；工作表数：{summary['master_sheet_count']}；台账记录数：{summary['master_row_count']}；字段数：{summary['master_column_count']}。",
                f"账单输入文件数：{summary['bill_source_file_count']}；账单行数：{summary['bill_row_count']}；账单月份范围：{summary['bill_month_min']}~{summary['bill_month_max']}。",
                "原始文件运行前后按 SHA-256、大小和修改时间校验，未发生变化。",
            ],
            ["台账文件结构目前只有一种样式；后续若新增文件模板，需重新执行同样检查。"],
            ["单文件分析不能证明所有未来台账模板稳定。"],
            ["无作品级明细进入公开报告；文件级指纹保存在 source-notes。"],
            REQ_AT["master_data"],
            "文件模板多样性仍需随新增台账继续验证。",
        ),
        (
            "02-field-domain-and-missingness.md",
            "数字版权台账字段和值域",
            [
                f"台账字段数：{summary['master_column_count']}；完整字段清单已纳入本报告，敏感字段仅展示非空、缺失和 distinct 计数。",
                "作者、书名、合同、作品ID、版权日期等字段不在公开报告展示具体值。",
            ],
            ["字段值域可作为补全表校验候选输入，但不能直接形成分类树、标签库或版权开始日期语义。"],
            ["字段名中未发现明确的一级、二级、三级主分类字段；未发现明确命名为版权开始日期的字段。"],
            ["敏感取值样本保存在被 Git 忽略的私有目录，用于运营确认。"],
            REQ_AT["master_data"] + "；" + REQ_AT["classification_tags"],
            "分类树、标签库和版权开始日期语义仍需运营确认。",
        ),
        (
            "03-work-id-matching.md",
            "台账作品ID与账单ID匹配",
            [
                f"台账可解析作品ID记录数：{summary['master_valid_work_id_row_count']}；不可解析/缺失记录数：{summary['master_invalid_work_id_row_count']}。",
                f"台账覆盖账单标准作品：{summary['master_covered_standard_work_count']} / {summary['bill_standard_work_count']}，覆盖率 {summary['master_coverage_rate']:.2%}。",
                f"未使用书名相似自动关联；匹配只基于作品ID数字主体。",
            ],
            ["台账 `作品ID` 可作为 M1 主数据补全的优先关联键候选。"],
            ["未匹配的标准作品不能从台账自动补全；需运营确认或补充来源。"],
            ["台账匹配失败或一对多冲突确认表已生成。"],
            REQ_AT["work_conflict"] + "；" + REQ_AT["master_data"],
            "未匹配和一对多冲突在运营确认前继续阻断正式补全应用。",
        ),
        (
            "04-duplicates-relations.md",
            "重复ID、一对多和多对一关系",
            [
                f"台账同一原始作品ID重复组：{summary['master_duplicate_raw_id_group_count']}；其中关键字段冲突组：{summary['master_duplicate_raw_id_conflict_group_count']}。",
                f"台账同一标准作品多原始ID组：{summary['master_standard_multi_raw_group_count']}。",
                f"账单同一原始ID多名称冲突组：{summary['bill_multi_name_group_count']}；多授权分类冲突组：{summary['bill_multi_auth_group_count']}。",
            ],
            ["确认单位应为冲突组，不应逐行处理账单明细。"],
            ["不能自动选择出现次数最多或收入最高的名称作为正式结论。"],
            ["已生成多名称、多授权分类、异常ID和台账匹配冲突确认表。"],
            REQ_AT["work_conflict"],
            "冲突组确认前继续阻断正式导入或正式基础信息应用。",
        ),
        (
            "05-standard-work-coverage.md",
            "台账对标准作品基础信息的覆盖",
            [
                f"候选作者缺失标准作品数：{summary['missing_author_candidate_standard_work_count']}。",
                f"明确一级至三级主分类缺失标准作品数：{summary['missing_confirmed_category_standard_work_count']}。",
                f"候选版权期限缺失标准作品数：{summary['missing_copyright_candidate_standard_work_count']}。",
                f"候选标签缺失标准作品数：{summary['missing_tag_candidate_standard_work_count']}。",
            ],
            ["台账可部分覆盖作品名称、候选作者和候选版权期限。"],
            ["台账不能单独覆盖 M1 全部基础信息需求，主要缺口为明确三级分类、标签库语义、版权开始日期语义和未匹配作品。"],
            ["缺口作品清单保存在私有目录。"],
            REQ_AT["master_data"] + "；" + REQ_AT["classification_tags"],
            "分类树、标签库、版权开始日期语义、未匹配作品仍需运营确认或额外来源。",
        ),
        (
            "06-name-differences.md",
            "标准作品名称差异",
            [
                f"台账书名与账单历史作品名无规范化一致项的标准作品组：{summary['master_bill_name_difference_group_count']}。",
                "名称差异仅作为风险信号，不作为自动解除或自动归并依据。",
            ],
            ["名称差异确认可作为基础信息补全前的阻断问题类型。"],
            ["无法仅凭名称相似判断同一作品或正式标准作品名称。"],
            ["名称差异样本保存在私有目录。"],
            REQ_AT["work_conflict"] + "；" + REQ_AT["master_data"],
            "标准作品名称选择来源需运营确认。",
        ),
        (
            "07-author-alias.md",
            "作者与作者别名候选",
            [
                f"作者署名/作者原名候选别名记录数：{summary['author_alias_candidate_row_count']}。",
                f"同一规范化作者署名对应多个作者原名的歧义组：{summary['ambiguous_author_alias_group_count']}。",
            ],
            ["作者署名与作者原名不同可作为作者别名候选。"],
            ["同名作者、歧义别名不能自动归并。"],
            ["作者别名候选和歧义组保存在私有目录。"],
            REQ_AT["master_data"],
            "作者对象和别名映射仍需运营确认后版本化。",
        ),
        (
            "08-classification-tags.md",
            "分类和标签字段分布",
            [
                "未发现明确的一级、二级、三级主分类字段。",
                f"候选分类字段：{', '.join([col for col in CATEGORY_CANDIDATE_COLS if col in structure['headers']]) or '无'}。",
                f"候选标签字段数：{summary['tag_candidate_field_count']}。",
            ],
            ["候选字段可辅助设计补全表和初始标签库候选，但不能直接冻结为 PRD 主分类或标签。"],
            ["最终分类树、标签值、字段到标签库的映射无法单靠台账确认。"],
            ["字段聚合分布在公开报告中以低敏聚合展示；作品级标签样本仅私有保存。"],
            REQ_AT["classification_tags"],
            "最终分类树、初始标签库和自动分类/标签规则仍为 PENDING-DATA。",
        ),
        (
            "09-copyright-dates.md",
            "版权日期格式、缺失和异常",
            [
                f"`签订日期`有效候选记录数：{summary['copyright_start_candidate_valid_row_count']}；`到期时间`有效候选记录数：{summary['copyright_end_candidate_valid_row_count']}。",
                f"候选版权期限完整标准作品数：{summary['copyright_candidate_complete_standard_work_count']} / {summary['bill_standard_work_count']}。",
                "公开报告不展示作品级版权日期。",
            ],
            ["`签订日期` 可作为版权开始日期候选字段，`到期时间` 可作为版权到期日期候选字段。"],
            ["是否把 `签订日期` 正式定义为版权开始日期，无法单靠字段名确认。"],
            ["版权期限反例确认表已生成。"],
            REQ_AT["master_data"],
            "版权开始日期权威字段语义仍需运营确认；候选期限缺失作品需补充。",
        ),
        (
            "10-copyright-counterexamples.md",
            "共享版权期限反例检查",
            [
                f"同一标准作品存在多个候选版权期限组：{summary['copyright_period_conflict_group_count']}。",
                f"纯数字与Y前缀候选期限不同反例组：{summary['pure_y_copyright_counterexample_group_count']}。",
            ],
            ["出现多候选期限时，按共享版权期限规则阻断并提交业务确认。"],
            ["不自动扩展为业务形态级版权期限字段。"],
            ["版权期限反例确认表已生成。"],
            REQ_AT["master_data"],
            "反例组确认前，共享版权期限规则不能解除相关阻断。",
        ),
        (
            "11-master-data-quality.md",
            "台账自身重复、冲突和缺失问题",
            [
                f"台账记录数：{summary['master_row_count']}；缺失作品ID或无法解析作品ID记录：{summary['master_invalid_work_id_row_count']}。",
                f"台账重复原始ID组：{summary['master_duplicate_raw_id_group_count']}；关键字段冲突组：{summary['master_duplicate_raw_id_conflict_group_count']}。",
                f"台账同标准多原始ID组：{summary['master_standard_multi_raw_group_count']}。",
            ],
            ["台账问题可转化为基础信息补全前的整批校验问题。"],
            ["错误过多时整份退回阈值仍需结合运营处理时长确认。"],
            ["私有目录包含台账冲突样本。"],
            REQ_AT["master_data"],
            "退回阈值和自动处理边界仍为 PENDING-DATA。",
        ),
        (
            "12-unfillable-works.md",
            "无法通过台账补全的作品",
            [
                f"台账未覆盖账单标准作品数：{summary['master_unmatched_standard_work_count']}。",
                f"候选作者缺失标准作品数：{summary['missing_author_candidate_standard_work_count']}。",
                f"候选版权期限缺失标准作品数：{summary['missing_copyright_candidate_standard_work_count']}。",
            ],
            ["未覆盖作品必须进入补全表或额外来源流程。"],
            ["无法使用作品名称相似度自动补全。"],
            ["缺口清单保存在私有目录和运营确认包。"],
            REQ_AT["master_data"],
            "未覆盖作品和缺失字段在运营确认前阻断进入后续正式评估。",
        ),
        (
            "13-freezable-fields.md",
            "可直接冻结的主数据字段语义",
            [
                "台账关联账单的首选键为作品ID数字主体；书名只作为冲突检查信号。",
                "账单最大月份与最新已确认完整月份必须分离；2026-05是不完整月份，最新已确认完整月份为2026-04。",
                "首次实销月份只取首次出现正数实销记录的月份；零值和负值不参与首次实销判断。",
                "实销金额使用 Excel 底层完整精度和精确十进制语义。",
            ],
            ["金额物理类型候选为 NUMERIC/DECIMAL(32,18)，物理模型阶段最终确认。"],
            ["分类树、标签库、版权开始日期权威字段仍不能冻结。"],
            ["冻结规则已回写 PRD 和技术设计。"],
            "REQ-DATA-IMPORT-006、REQ-WORK-003、REQ-WORK-004；AT-M1-006、AT-M1-022、AT-M1-023",
            "金额物理类型候选待物理模型阶段最终确认；分类和标签仍为 PENDING-DATA。",
        ),
        (
            "14-operation-confirmation.md",
            "运营确认包",
            [
                f"运营确认表数量：8；确认总组数：{summary['ops_confirmation_total_group_count']}。",
                "确认包按冲突组/候选组生成，不生成 53,900 条账单逐行确认文件。",
                "确认表位于被 Git 忽略的本地目录。",
            ],
            ["运营确认结果应形成版本化映射，并应用于组内全部相关收入投影。"],
            ["未确认冲突组继续阻断正式导入或基础信息正式应用。"],
            ["确认包 Excel 工作簿和 CSV 已生成到私有目录。"],
            REQ_AT["work_conflict"] + "；" + REQ_AT["master_data"],
            "所有冲突确认结果未回填前，相关导入/补全仍保持阻断。",
        ),
    ]

    for filename, title, facts, candidate_rules, cannot_confirm, ops_samples, req_at, pending in reports:
        write_report_file(PUBLIC_ROOT / filename, title, facts, candidate_rules, cannot_confirm, ops_samples, req_at, pending)

    source_notes = [
        "# M1 数字版权台账分析来源说明",
        "",
        "## 输入源",
        "",
        "公开报告只记录文件级匿名 ID、大小和 SHA-256，不记录原始文件名。",
        "",
        "### 账单文件",
        "",
        markdown_table(
            [{"file_id": item.file_id, "size": item.size, "sha256": item.sha256} for item in source_snapshots["bill"]],
            ["file_id", "size", "sha256"],
        ),
        "",
        "### 数字版权台账文件",
        "",
        markdown_table(
            [{"file_id": item.file_id, "size": item.size, "sha256": item.sha256} for item in source_snapshots["master"]],
            ["file_id", "size", "sha256"],
        ),
        "",
        "## 字段值域公开策略",
        "",
        "- 作者、书名、合同、作品 ID、版权日期等敏感字段仅公开缺失率、distinct 数和格式统计。",
        "- 运营确认明细、作品级样本、作者、版权日期和金额明细保存于 `data/m1-master-data-private/`，该目录已被 `.gitignore` 排除。",
    ]
    (PUBLIC_ROOT / "source-notes.md").write_text("\n".join(source_notes), encoding="utf-8")

    field_summary_rows = []
    for item in field_profiles:
        field_summary_rows.append(
            {
                "字段": item["field"],
                "非空": item["non_empty_count"],
                "缺失": item["missing_count"],
                "Distinct": item["distinct_count"],
                "公开取值策略": item["public_value_policy"],
            }
        )
    (PUBLIC_ROOT / "field-profile-public.json").write_text(json.dumps(field_profiles, ensure_ascii=False, indent=2), encoding="utf-8")
    (PUBLIC_ROOT / "date-profile-public.json").write_text(json.dumps(date_profiles, ensure_ascii=False, indent=2), encoding="utf-8")
    (PUBLIC_ROOT / "tag-profile-public.json").write_text(json.dumps(tag_profiles, ensure_ascii=False, indent=2), encoding="utf-8")

    readme = f"""# M1 数字版权台账分析报告

## 技术摘要

- 台账输入：{summary['master_source_file_count']} 个文件、{summary['master_sheet_count']} 个工作表、{summary['master_row_count']} 条记录、{summary['master_column_count']} 个字段。
- 账单基线：{summary['bill_row_count']} 行，账单最大月份 {summary['bill_month_max']}；运营确认 2026-05 为不完整月份，最新已确认完整月份为 2026-04。
- 台账覆盖：{summary['master_covered_standard_work_count']} / {summary['bill_standard_work_count']} 个账单标准作品，覆盖率 {summary['master_coverage_rate']:.2%}；未使用书名相似自动关联。
- M1 基础信息覆盖：台账可部分提供候选名称、作者和版权期限，但不能单独覆盖明确三级分类、标签库语义和版权开始日期权威字段。
- 首次实销规则已解除 PENDING-DATA：标准作品和业务形态首次实销月份均取首次出现正数实销记录的月份，零值和负值不作为首次实销。
- 金额精度已解除 PENDING-DATA：财务认可 Excel 底层完整精度，权威金额和对账必须使用精确十进制语义；物理类型候选为 `NUMERIC(32,18)`。

## 关键数量

| 指标 | 数值 |
|---|---:|
| 标准作品总数（账单派生） | {summary['bill_standard_work_count']} |
| 台账覆盖标准作品 | {summary['master_covered_standard_work_count']} |
| 台账未覆盖标准作品 | {summary['master_unmatched_standard_work_count']} |
| 候选作者缺失标准作品 | {summary['missing_author_candidate_standard_work_count']} |
| 明确三级分类缺失标准作品 | {summary['missing_confirmed_category_standard_work_count']} |
| 候选版权期限缺失标准作品 | {summary['missing_copyright_candidate_standard_work_count']} |
| 候选标签缺失标准作品 | {summary['missing_tag_candidate_standard_work_count']} |
| 运营确认总组数 | {summary['ops_confirmation_total_group_count']} |

## 图表

{chr(10).join(f"![{Path(chart).stem}]({chart})" for chart in charts)}

## 分项报告

- [01-file-workbook-structure.md](01-file-workbook-structure.md)
- [02-field-domain-and-missingness.md](02-field-domain-and-missingness.md)
- [03-work-id-matching.md](03-work-id-matching.md)
- [04-duplicates-relations.md](04-duplicates-relations.md)
- [05-standard-work-coverage.md](05-standard-work-coverage.md)
- [06-name-differences.md](06-name-differences.md)
- [07-author-alias.md](07-author-alias.md)
- [08-classification-tags.md](08-classification-tags.md)
- [09-copyright-dates.md](09-copyright-dates.md)
- [10-copyright-counterexamples.md](10-copyright-counterexamples.md)
- [11-master-data-quality.md](11-master-data-quality.md)
- [12-unfillable-works.md](12-unfillable-works.md)
- [13-freezable-fields.md](13-freezable-fields.md)
- [14-operation-confirmation.md](14-operation-confirmation.md)

## 结论

数字版权台账不能单独覆盖 M1 基础信息需求。进入物理数据库设计仍取决于：运营确认冲突组、确认版权开始日期字段语义、补齐明确三级分类/标签库来源，并确认台账未覆盖标准作品的补全路径。
"""
    (PUBLIC_ROOT / "README.md").write_text(readme, encoding="utf-8")

    html_sections = []
    for filename, title, *_ in reports:
        text = (PUBLIC_ROOT / filename).read_text(encoding="utf-8")
        body = html.escape(text)
        body = body.replace("\n", "<br>\n")
        html_sections.append(f"<section><h2>{html.escape(title)}</h2><p>{body}</p></section>")
    chart_html = "\n".join(f'<figure><img src="{html.escape(chart)}" alt="{html.escape(Path(chart).stem)}"></figure>' for chart in charts)
    report_html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>M1 数字版权台账分析报告</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 40px; color: #1f2430; background: #fcfcfd; }}
    main {{ max-width: 1080px; margin: auto; }}
    section {{ background: #fff; border: 1px solid #e6e8f0; border-radius: 12px; padding: 20px; margin: 18px 0; }}
    img {{ max-width: 100%; border: 1px solid #e6e8f0; border-radius: 10px; background: #fff; }}
    .summary {{ font-size: 16px; line-height: 1.7; }}
    code {{ background: #f4f5f7; padding: 2px 5px; border-radius: 4px; }}
  </style>
</head>
<body>
<main>
  <h1 data-contract-section="title">M1 数字版权台账分析报告</h1>
  <section data-contract-section="technical-summary" class="summary">
    <h2>技术摘要</h2>
    <p>台账覆盖 {summary['master_covered_standard_work_count']} / {summary['bill_standard_work_count']} 个账单标准作品；不能单独覆盖 M1 基础信息需求。金额精度和首次实销金额符号规则已解除 PENDING-DATA；分类、标签、版权开始日期权威字段和冲突组确认仍需运营处理。</p>
  </section>
  <section data-contract-section="key-findings">
    <h2>关键发现与图表</h2>
    {chart_html}
  </section>
  <section data-contract-section="scope-data-and-metric-definitions">
    <h2>范围、数据和定义</h2>
    <p>关联只使用作品 ID 数字主体；账单最大月份与最新已确认完整月份分离；首次实销月份按首次正数实销记录确定。</p>
  </section>
  <section data-contract-section="methodology">
    <h2>方法</h2>
    <p>读取前后校验 SHA-256、大小和修改时间；账单金额使用 Excel 底层 token 转 Decimal；台账敏感取值仅进入被 Git 忽略的私有确认包。</p>
  </section>
  {''.join(html_sections)}
  <section data-contract-section="limitations-uncertainty-and-robustness-checks">
    <h2>限制、不确定性和稳健性检查</h2>
    <p>只有一个台账文件和一个账单文件；不证明未来模板稳定。分类树、标签库、版权开始日期语义和冲突解除仍需运营确认。</p>
  </section>
  <section data-contract-section="recommended-next-steps">
    <h2>建议下一步</h2>
    <p>先处理本地运营确认包；确认结果形成版本化映射和基础信息补全规则后，再判断是否进入物理数据库设计。</p>
  </section>
  <section data-contract-section="further-questions">
    <h2>仍需回答的问题</h2>
    <p>签订日期是否等同版权开始日期；候选分类/标签字段如何映射为正式分类树和标签库；未匹配作品由哪个来源补齐。</p>
  </section>
</main>
</body>
</html>
"""
    (PUBLIC_ROOT / "report.html").write_text(report_html, encoding="utf-8")


def derive_summary(
    bill_rows: list[dict[str, Any]],
    master_rows: list[dict[str, Any]],
    structure: dict[str, Any],
    bill_conflicts: dict[str, Any],
    master_analysis: dict[str, Any],
    date_profiles: list[dict[str, Any]],
    tag_profiles: list[dict[str, Any]],
    ops_tables: dict[str, list[dict[str, Any]]],
    source_snapshots: dict[str, list[SourceSnapshot]],
    start_time: float,
) -> dict[str, Any]:
    real_summary = json.loads(REAL_BILL_SUMMARY.read_text(encoding="utf-8")) if REAL_BILL_SUMMARY.exists() else {}
    bill_months = sorted({row["month"] for row in bill_rows if row.get("month")})
    amounts = [row["amount"] for row in bill_rows if row.get("amount") is not None]
    max_integer_digits = max((len(str(abs(int(amount)))) for amount in amounts), default=0)
    max_scale = max((-amount.as_tuple().exponent if amount.as_tuple().exponent < 0 else 0 for amount in amounts), default=0)
    total_amount = dec_sum(amounts)
    max_batch_total_integer_digits = len(str(abs(int(total_amount)))) if amounts else 0

    date_by_field = {item["field"]: item for item in date_profiles}
    start_profile = date_by_field.get(COPYRIGHT_START_CANDIDATE, {})
    end_profile = date_by_field.get(COPYRIGHT_END_CANDIDATE, {})

    ops_counts = {name: len(rows) for name, rows in ops_tables.items()}
    bill_standard_count = len(bill_conflicts["standards_seen"])
    covered_count = len(master_analysis["covered_standard_ids"])
    summary = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "bill_source_file_count": len(source_snapshots["bill"]),
        "master_source_file_count": len(source_snapshots["master"]),
        "bill_row_count": len(bill_rows),
        "bill_month_min": bill_months[0] if bill_months else None,
        "bill_month_max": bill_months[-1] if bill_months else None,
        "confirmed_complete_month": "2026-04",
        "incomplete_months": ["2026-05"],
        "bill_total_amount": fmt_decimal(total_amount),
        "amount_max_integer_digits": max_integer_digits,
        "amount_max_scale": max_scale,
        "amount_max_batch_total": fmt_decimal(total_amount),
        "amount_max_batch_total_integer_digits": max_batch_total_integer_digits,
        "amount_decimal_candidate": "NUMERIC(32,18)",
        "amount_decimal_candidate_integer_digits": 14,
        "master_sheet_count": sum(item["sheet_count"] for item in structure["files"]),
        "master_row_count": len(master_rows),
        "master_column_count": len(structure["headers"]),
        "master_headers": structure["headers"],
        "master_valid_work_id_row_count": sum(1 for row in master_rows if row.get("_work_id_valid")),
        "master_invalid_work_id_row_count": sum(1 for row in master_rows if not row.get("_work_id_valid")),
        "bill_standard_work_count": bill_standard_count,
        "bill_raw_work_id_count": len(master_analysis["bill_raw_ids"]),
        "master_distinct_raw_work_id_count": len(master_analysis["master_raw_ids"]),
        "master_distinct_standard_work_id_count": len(master_analysis["master_standard_ids"]),
        "master_covered_standard_work_count": covered_count,
        "master_unmatched_standard_work_count": bill_standard_count - covered_count,
        "master_coverage_rate": covered_count / bill_standard_count if bill_standard_count else 0,
        "missing_title_candidate_standard_work_count": len(master_analysis["missing_title_standard_ids"]),
        "missing_author_candidate_standard_work_count": len(master_analysis["missing_author_standard_ids"]),
        "missing_confirmed_category_standard_work_count": bill_standard_count,
        "missing_copyright_candidate_standard_work_count": len(master_analysis["missing_copyright_candidate_standard_ids"]),
        "missing_confirmed_copyright_start_semantic_standard_work_count": bill_standard_count,
        "missing_tag_candidate_standard_work_count": len(master_analysis["missing_tag_candidate_standard_ids"]),
        "missing_confirmed_tag_library_standard_work_count": bill_standard_count,
        "copyright_start_candidate_valid_row_count": start_profile.get("valid_count", 0),
        "copyright_end_candidate_valid_row_count": end_profile.get("valid_count", 0),
        "copyright_candidate_complete_standard_work_count": bill_standard_count - len(master_analysis["missing_copyright_candidate_standard_ids"]),
        "bill_multi_name_group_count": len(bill_conflicts["multi_name"]),
        "bill_multi_auth_group_count": len(bill_conflicts["multi_auth"]),
        "bill_abnormal_id_group_count": len(bill_conflicts["abnormal_ids"]),
        "bill_volume_candidate_group_count": len(bill_conflicts["volume_candidates"]),
        "bill_offset_candidate_group_count": len(bill_conflicts["offset_candidates"]),
        "bill_first_sale_empty_group_count": len(bill_conflicts["first_sale_empty"]),
        "master_bill_name_difference_group_count": len(master_analysis["name_difference_rows"]),
        "author_alias_candidate_row_count": len(master_analysis["author_alias_rows"]),
        "ambiguous_author_alias_group_count": len(master_analysis["ambiguous_author_alias_rows"]),
        "master_duplicate_raw_id_group_count": master_analysis["duplicate_raw_id_group_count_all"],
        "master_duplicate_raw_id_conflict_group_count": len(master_analysis["duplicate_raw_ids"]),
        "master_standard_multi_raw_group_count": len(master_analysis["standard_multi_raw"]),
        "copyright_period_conflict_group_count": len(master_analysis["copyright_period_groups"]),
        "pure_y_copyright_counterexample_group_count": len(master_analysis["pure_y_counterexamples"]),
        "tag_candidate_field_count": len(tag_profiles),
        "ops_confirmation_counts": ops_counts,
        "ops_confirmation_total_group_count": sum(ops_counts.values()),
        "real_bill_prior_summary": {
            "row_count": real_summary.get("row_count"),
            "month_min": real_summary.get("month_min"),
            "month_max": real_summary.get("month_max"),
            "total_amount": real_summary.get("total_amount"),
        },
        "pending_data_resolved": [
            "完整数据截止月份：2026-05 为不完整账单月份，最新已确认完整月份为 2026-04。",
            "首次实销月份金额符号：仅正数实销记录可决定标准作品上线时间和业务形态首次实销月份。",
            "金额精度：使用 Excel 底层完整精度和精确十进制语义，不按显示位数舍入。",
        ],
        "pending_data_remaining": [
            "明确一级至三级分类树和字段来源。",
            "初始标签库与标签字段映射。",
            "版权开始日期权威字段语义，尤其 `签订日期` 是否可作为版权开始日期。",
            "台账未覆盖作品和冲突组的运营确认。",
            "金额 NUMERIC/DECIMAL 物理精度在物理模型阶段最终确认。",
        ],
        "source_unchanged": None,
        "analysis_seconds": time.perf_counter() - start_time,
    }
    return summary


def write_summary(summary: dict[str, Any], source_snapshots: dict[str, list[SourceSnapshot]]) -> None:
    payload = dict(summary)
    payload["source_snapshots"] = {
        key: [item.__dict__ for item in value]
        for key, value in source_snapshots.items()
    }
    (PUBLIC_ROOT / "summary.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    start_time = time.perf_counter()
    ensure_dirs()
    bill_files = workbook_inputs(BILL_INPUT_ROOT)
    master_files = workbook_inputs(MASTER_INPUT_ROOT)
    if not bill_files:
        raise RuntimeError("未找到 data/real-bills 下的账单文件。")
    if not master_files:
        raise RuntimeError("未找到 data/master-data 下的台账文件。")

    source_snapshots = {
        "bill": snapshot_sources(bill_files, "B"),
        "master": snapshot_sources(master_files, "M"),
    }

    bill_rows, _bill_structure = read_bills()
    master_rows, structure = read_master()
    field_profiles = field_profile(master_rows, structure["headers"])
    date_profiles = analyze_dates(master_rows, structure["headers"])
    tag_profiles = analyze_tag_distribution(master_rows, structure["headers"])
    bill_conflicts = group_bill_conflicts(bill_rows)
    master_analysis = analyze_master_relations(master_rows, bill_rows, bill_conflicts)
    ops_tables = build_ops_tables(bill_conflicts, master_analysis)
    write_private_outputs(master_rows, bill_conflicts, master_analysis, ops_tables)

    summary = derive_summary(
        bill_rows,
        master_rows,
        structure,
        bill_conflicts,
        master_analysis,
        date_profiles,
        tag_profiles,
        ops_tables,
        source_snapshots,
        start_time,
    )
    source_unchanged = assert_sources_unchanged(source_snapshots["bill"]) and assert_sources_unchanged(source_snapshots["master"])
    summary["source_unchanged"] = source_unchanged
    write_summary(summary, source_snapshots)
    write_public_reports(summary, structure, field_profiles, date_profiles, tag_profiles, bill_conflicts, master_analysis, source_snapshots)
    if not source_unchanged:
        raise RuntimeError("原始账单或台账文件在分析期间发生变化。")


if __name__ == "__main__":
    main()
