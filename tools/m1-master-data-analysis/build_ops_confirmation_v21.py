from __future__ import annotations

import json
from copy import copy
from datetime import datetime
from pathlib import Path
from shutil import copy2

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
OPS_ROOT = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
SOURCE_WORKBOOK = OPS_ROOT / "M1-运营确认包-v2.xlsx"
OUTPUT_WORKBOOK = OPS_ROOT / "M1-运营确认包-v2.1.xlsx"

COPYRIGHT_TASK_BY_STANDARD_ID = {
    "163360": "CP-26D16E3713",
    "162214": "CP-C972D52413",
}


def header_map(ws: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> dict[str, int]:
    return {str(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1) if ws.cell(row, col).value is not None}


def remove_copyright_lines(value: object) -> str | None:
    if value is None:
        return None
    lines = str(value).splitlines()
    kept: list[str] = []
    skipping = False
    for line in lines:
        if line.startswith("版权期限候选"):
            skipping = True
            continue
        if skipping:
            continue
        kept.append(line)
    text = "\n".join(line for line in kept if line.strip()).strip()
    return text or None


def remove_conflict_token(value: object, token: str) -> str | None:
    if value is None:
        return None
    parts = [part.strip() for part in str(value).splitlines() if part.strip() and part.strip() != token]
    return "\n".join(parts) if parts else None


def add_list_validation(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str, values: list[str]) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def style_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
    ws.auto_filter.ref = ws.dimensions


def update_ledger_conflicts(wb: openpyxl.Workbook) -> None:
    ws = wb["台账真实冲突"]
    headers = header_map(ws)

    if "关联版权任务ID" not in headers:
        insert_at = headers["来源记录数量"] + 1
        ws.insert_cols(insert_at, amount=2)
        ws.cell(1, insert_at).value = "关联版权任务ID"
        ws.cell(1, insert_at + 1).value = "版权填写提示"

    headers = header_map(ws)
    for row in range(2, ws.max_row + 1):
        standard_id = str(ws.cell(row, headers["标准作品ID"]).value or "")
        ws.cell(row, headers["冲突字段"]).value = remove_conflict_token(ws.cell(row, headers["冲突字段"]).value, "版权期限")
        ws.cell(row, headers["各候选值"]).value = remove_copyright_lines(ws.cell(row, headers["各候选值"]).value)
        if standard_id in COPYRIGHT_TASK_BY_STANDARD_ID:
            ws.cell(row, headers["关联版权任务ID"]).value = COPYRIGHT_TASK_BY_STANDARD_ID[standard_id]
            ws.cell(row, headers["版权填写提示"]).value = "版权开始/到期日期仅在“版权期限反例”填写；本表只处理名称、作者或其他非版权字段冲突。"

    add_list_validation(ws, f"{get_column_letter(headers['是否解除冲突'])}2:{get_column_letter(headers['是否解除冲突'])}{ws.max_row}", ["是", "否", "需补充资料"])
    ws.freeze_panes = "C2"
    style_sheet(ws)


def update_non_blocking_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb["非阻断观察"]
    headers = header_map(ws)

    if "阻断解除状态" not in headers:
        insert_at = headers["运营备注"]
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "阻断解除状态"

    headers = header_map(ws)
    for row in range(2, ws.max_row + 1):
        mark_cell = ws.cell(row, headers["是否标记为异常"])
        escalate_cell = ws.cell(row, headers["是否升级为阻断"])
        release_cell = ws.cell(row, headers["阻断解除状态"])
        if mark_cell.value == "否":
            mark_cell.value = None
        if escalate_cell.value == "否":
            escalate_cell.value = None
        if release_cell.value == "否":
            release_cell.value = None

    add_list_validation(ws, f"{get_column_letter(headers['是否标记为异常'])}2:{get_column_letter(headers['是否标记为异常'])}{ws.max_row}", ["是", "否", "需补充资料"])
    add_list_validation(ws, f"{get_column_letter(headers['是否升级为阻断'])}2:{get_column_letter(headers['是否升级为阻断'])}{ws.max_row}", ["是", "否"])
    add_list_validation(ws, f"{get_column_letter(headers['阻断解除状态'])}2:{get_column_letter(headers['阻断解除状态'])}{ws.max_row}", ["处理中", "已解除", "需补充资料", "不适用"])

    wait_fill = PatternFill(fill_type="solid", fgColor="FFF4C2")
    blocked_fill = PatternFill(fill_type="solid", fgColor="FFEDDE")
    done_fill = PatternFill(fill_type="solid", fgColor="D8ECBD")
    for header in ["是否标记为异常", "是否升级为阻断", "阻断解除状态"]:
        col = get_column_letter(headers[header])
        target = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(target, FormulaRule(formula=[f'{col}2="需补充资料"'], fill=blocked_fill))
        ws.conditional_formatting.add(target, FormulaRule(formula=[f'{col}2="处理中"'], fill=wait_fill))
        ws.conditional_formatting.add(target, FormulaRule(formula=[f'OR({col}2="已解除",{col}2="否")'], fill=done_fill))

    ws.freeze_panes = "D2"
    style_sheet(ws)


def update_overview(wb: openpyxl.Workbook) -> None:
    ws = wb["确认进度总览"]
    ws["A1"] = "M1 运营确认包 v2.1 - 确认进度总览"
    ws["B3"] = datetime.now()
    ws["B3"].number_format = "yyyy-mm-dd hh:mm:ss"

    ledger_last = wb["台账真实冲突"].max_row
    ws["C9"] = f'=COUNTIF(\'台账真实冲突\'!I2:I{ledger_last},"<>")-COUNTIF(\'台账真实冲突\'!I2:I{ledger_last},"需补充资料")'
    ws["D9"] = f'=COUNTBLANK(\'台账真实冲突\'!I2:I{ledger_last})'
    ws["E9"] = f'=COUNTIF(\'台账真实冲突\'!I2:I{ledger_last},"需补充资料")'
    ws["F9"] = f'=COUNTIF(\'台账真实冲突\'!J2:J{ledger_last},"否")+COUNTBLANK(\'台账真实冲突\'!J2:J{ledger_last})'

    nb_last = wb["非阻断观察"].max_row
    ws["C11"] = (
        f'=COUNTIF(\'非阻断观察\'!K2:K{nb_last},"否")'
        f'+COUNTIFS(\'非阻断观察\'!K2:K{nb_last},"是",\'非阻断观察\'!L2:L{nb_last},"<>是")'
        f'+COUNTIFS(\'非阻断观察\'!L2:L{nb_last},"是",\'非阻断观察\'!M2:M{nb_last},"已解除")'
    )
    ws["D11"] = (
        f'=COUNTIFS(\'非阻断观察\'!K2:K{nb_last},"是",\'非阻断观察\'!L2:L{nb_last},"")'
        f'+COUNTIFS(\'非阻断观察\'!L2:L{nb_last},"是",\'非阻断观察\'!M2:M{nb_last},"")'
        f'+COUNTIFS(\'非阻断观察\'!L2:L{nb_last},"是",\'非阻断观察\'!M2:M{nb_last},"处理中")'
    )
    ws["E11"] = (
        f'=COUNTIF(\'非阻断观察\'!K2:K{nb_last},"需补充资料")'
        f'+COUNTIFS(\'非阻断观察\'!L2:L{nb_last},"是",\'非阻断观察\'!M2:M{nb_last},"需补充资料")'
    )
    ws["F11"] = f'=COUNTIFS(\'非阻断观察\'!L2:L{nb_last},"是",\'非阻断观察\'!M2:M{nb_last},"<>已解除")'
    ws["I11"] = '=IF((C11+D11+E11+F11)=0,1,C11/(C11+D11+E11+F11))'
    ws["J11"] = "默认不要求确认；只有人工标记异常或升级阻断后才进入进度统计。默认空值不计为待确认或未解除阻断。"

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A5:J11"


def postprocess_formats(wb: openpyxl.Workbook) -> None:
    for ws in wb.worksheets:
        style_sheet(ws)
        header_row = 5 if ws.title == "确认进度总览" else 1
        headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
        for col_idx, header_value in enumerate(headers, start=1):
            if header_value is None:
                continue
            header = str(header_value)
            col_letter = get_column_letter(col_idx)
            if "完整精度" in header:
                ws.column_dimensions[col_letter].hidden = True
                for cell in ws[col_letter]:
                    cell.number_format = "@"
                    cell.font = Font(color="666666")
            if header in {"累计实销", "金额"}:
                for column in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=header_row + 1, max_row=ws.max_row):
                    for cell in column:
                        cell.number_format = "#,##0.00"
            if "日期" in header and "候选" not in header and "全部" not in header:
                for column in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=header_row + 1, max_row=ws.max_row):
                    for cell in column:
                        cell.number_format = "yyyy-mm-dd"


def main() -> None:
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(SOURCE_WORKBOOK)
    copy2(SOURCE_WORKBOOK, OUTPUT_WORKBOOK)
    wb = openpyxl.load_workbook(OUTPUT_WORKBOOK)
    update_ledger_conflicts(wb)
    update_non_blocking_sheet(wb)
    update_overview(wb)
    postprocess_formats(wb)
    wb.save(OUTPUT_WORKBOOK)
    wb.close()

    payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": str(SOURCE_WORKBOOK),
        "output": str(OUTPUT_WORKBOOK),
    }
    (OPS_ROOT / "ops-confirmation-v2.1-build.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
