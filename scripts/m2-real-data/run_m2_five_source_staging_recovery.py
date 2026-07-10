from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from m2_five_source_staging_contract import (  # noqa: E402
    CORE_FIELDS,
    SCHEMA,
    STATUS_FIELDS,
    validate_staging_payload,
)


PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
DEFAULT_FOUNDATION = PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.xlsx"
DEFAULT_OUTPUT = PRIVATE_DIR / "M2-five-source-local-staging-apply-result-cn-v1.json"
MASTER_DIR = ROOT / "data" / "master-data"
BILL_PATH = ROOT / "data" / "real-bills" / "渠道实销汇总 -修改版.xlsx"
DIGITAL_LEDGER = MASTER_DIR / "数字版权台账.xlsx"
ORIGINAL_LIBRARY = MASTER_DIR / "原创全库.xlsx"
ORIGINAL_LIBRARY_2 = MASTER_DIR / "原创全库2.xlsx"
AUTHORIZATION_SUMMARY = MASTER_DIR / "授权汇总台账.xlsx"
AUTHORIZATION_RELATIONSHIP = MASTER_DIR / "授权关系仪表板.xlsx"
DUAL_SOURCE_STAGING = (
    ROOT
    / "data"
    / "private-output"
    / "m1-master-data"
    / "M1-dual-source-limited-staging-table-v1.json"
)
MAPPING_PAYLOAD = (
    ROOT
    / "data"
    / "m1-master-data-private"
    / "mapping-candidate"
    / "M1-formal-mapping-version-candidate-v0.1-detail-payload.json"
)
MAPPING_OVERLAY = (
    ROOT
    / "experiments"
    / "m1-mapping-version-import-candidate"
    / "G07-mapping-strategy-overlay-v0.2.json"
)


FOUNDATION_HEADERS = [
    "序号",
    "书名",
    "作品编号",
    "作者",
    "一级分类(最终采用)",
    "二级分类(最终采用)",
    "三级分类(最终采用)",
    "辅助标签(最终采用)",
    "备注",
]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def canonical_work_id(value) -> str:
    text = clean(value)
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return str(int(match.group(0))) if match else text


def normalize_author(value) -> str:
    text = re.sub(r"\s+", "", clean(value))
    return text.replace("，", ",").replace("；", ";")


def normalize_title(value) -> str:
    text = clean(value)
    text = re.sub(r"[《》“”\"'‘’（）()\[\]【】、，,。:：;；\s]", "", text)
    text = re.sub(r"(新版|修订版|珍藏版|套装|全集|增订版|纪念版|典藏版)", "", text)
    return text.lower()


def author_tokens(value) -> set[str]:
    parts = re.split(r"[、，,;；/／\s]+| and | 和 | 与 ", clean(value))
    return {
        re.sub(r"(著|编著|主编|作者|译|编)", "", part).strip().lower()
        for part in parts
        if part.strip()
    }


def normalize_date(value) -> str:
    if value is None or clean(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        parsed = date(1899, 12, 30) + timedelta(days=int(round(float(value))))
        return parsed.isoformat()
    text = clean(value)
    if re.search(r"无限期|无期限|永久|长期有效|版权保护期满", text):
        return "无限期"
    matches = re.findall(
        r"(20\d{2}|19\d{2})[-/.年](\d{1,2})(?:[-/.月](\d{1,2}))?", text
    )
    if not matches:
        return ""
    year, month, day = matches[-1]
    return f"{int(year):04d}-{int(month):02d}-{int(day or 1):02d}"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_output(path: Path) -> Path:
    resolved = path.resolve()
    private_root = (ROOT / "data" / "private-output").resolve()
    if private_root != resolved and private_root not in resolved.parents:
        raise SystemExit("Recovery output must stay under the ignored private-output root.")
    return resolved


def read_rows(path: Path, sheet_name: str, max_columns: int | None = None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Required source sheet is missing for role {path.stem}.")
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True, max_col=max_columns)
        headers = [clean(value) for value in next(iterator)]
        for values in iterator:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                if header not in row or (not clean(row[header]) and clean(value)):
                    row[header] = value
            if any(clean(value) for value in row.values()):
                yield row
    finally:
        workbook.close()


def load_foundation(path: Path) -> dict[str, dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: dict[str, dict] = {}
    try:
        for sheet_name in ["01_出版物", "02_网文"]:
            if sheet_name not in workbook.sheetnames:
                raise SystemExit("Foundation workbook is missing a required main sheet.")
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True, max_col=len(FOUNDATION_HEADERS))
            headers = [clean(value) for value in next(iterator)]
            if headers != FOUNDATION_HEADERS:
                raise SystemExit("Foundation workbook headers differ from the confirmed contract.")
            for values in iterator:
                row = dict(zip(headers, values))
                if not any(clean(value) for value in row.values()):
                    continue
                work_id = canonical_work_id(row.get("作品编号"))
                if not work_id or work_id in records:
                    raise SystemExit("Foundation workbook has a missing or duplicate work ID.")
                records[work_id] = {
                    "作品编号": work_id,
                    "书名": clean(row.get("书名")),
                    "作者": clean(row.get("作者")),
                    "一级分类": clean(row.get("一级分类(最终采用)")),
                    "二级分类": clean(row.get("二级分类(最终采用)")),
                    "三级分类": clean(row.get("三级分类(最终采用)")),
                    "辅助标签": clean(row.get("辅助标签(最终采用)")) or "无",
                    "备注": clean(row.get("备注")),
                }
    finally:
        workbook.close()
    if len(records) != 3053:
        raise SystemExit(f"Foundation workbook must contain 3053 works, found {len(records)}.")
    required = ["书名", "作者", "一级分类", "二级分类", "三级分类"]
    if any(not record[field] for record in records.values() for field in required):
        raise SystemExit("Foundation workbook contains an empty required field.")
    return records


def load_standard_mapping(final_ids: set[str]) -> dict[str, str]:
    if not MAPPING_PAYLOAD.exists():
        return {}
    payload = json.loads(MAPPING_PAYLOAD.read_text(encoding="utf-8-sig"))
    mapping = {}
    for row in payload.get("effective_mapping_snapshot", []):
        if row.get("layer") != "historical_volume" or row.get("effective_status") != "effective_candidate":
            continue
        source = canonical_work_id(row.get("historical_standard_work_id"))
        target = canonical_work_id(row.get("target_standard_work_id"))
        if source and target in final_ids:
            mapping[source] = target
    if MAPPING_OVERLAY.exists():
        overlay = json.loads(MAPPING_OVERLAY.read_text(encoding="utf-8-sig"))
        for row in overlay.get("changes", []):
            if row.get("to") != "historical_volume_mapping":
                continue
            source = canonical_work_id(row.get("raw_work_id"))
            target = canonical_work_id(row.get("target_standard_work_id"))
            if source and target in final_ids:
                mapping[source] = target
    return mapping


def build_candidate(
    source_role: str,
    value: str,
    match_method: str,
    confirmation_mode: str,
) -> dict:
    return {
        "sourceRole": source_role,
        "value": value,
        "matchMethod": match_method,
        "confirmationMode": confirmation_mode,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Regenerate a private M2 five-source staging recovery candidate."
    )
    parser.add_argument("--foundation-workbook", type=Path, default=DEFAULT_FOUNDATION)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--as-of", default="2026-07-10")
    parser.add_argument("--write-incomplete-candidate", action="store_true")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    source_paths = {
        "billing_scope": BILL_PATH,
        "digital_ledger": DIGITAL_LEDGER,
        "original_library": ORIGINAL_LIBRARY,
        "original_library_2": ORIGINAL_LIBRARY_2,
        "authorization_summary": AUTHORIZATION_SUMMARY,
        "authorization_relationship": AUTHORIZATION_RELATIONSHIP,
    }
    required = [args.foundation_workbook, *source_paths.values()]
    missing = [path.name for path in required if not path.exists()]
    if missing:
        raise SystemExit("Required authorized private inputs are missing: " + ", ".join(missing))

    as_of = date.fromisoformat(args.as_of)
    foundation = load_foundation(args.foundation_workbook)
    final_ids = set(foundation)
    standard_mapping = load_standard_mapping(final_ids)
    candidates: dict[str, dict[str, list[dict]]] = defaultdict(
        lambda: defaultdict(list)
    )
    source_entries: list[dict] = []
    explicit_status: dict[str, set[str]] = defaultdict(set)

    def target_id(value) -> str:
        work_id = canonical_work_id(value)
        if work_id not in final_ids and standard_mapping.get(work_id) in final_ids:
            return standard_mapping[work_id]
        return work_id

    def add(
        raw_id,
        field: str,
        value,
        source_role: str,
        normalizer,
        match_method: str = "exact_id",
        confirmation_mode: str = "regenerated_direct_source",
    ) -> None:
        work_id = target_id(raw_id)
        normalized = normalizer(value)
        if work_id not in final_ids or not normalized:
            return
        candidate = build_candidate(
            source_role, normalized, match_method, confirmation_mode
        )
        if candidate not in candidates[work_id][field]:
            candidates[work_id][field].append(candidate)

    def register(
        raw_id,
        titles,
        author,
        start,
        end,
        source_role: str,
    ) -> None:
        add(raw_id, "作者", author, source_role, normalize_author)
        add(raw_id, "版权开始", start, source_role, normalize_date)
        add(raw_id, "版权到期", end, source_role, normalize_date)
        source_entries.append(
            {
                "sourceRole": source_role,
                "titles": {
                    normalize_title(title)
                    for title in titles
                    if normalize_title(title)
                },
                "authorTokens": author_tokens(author),
                "版权开始": normalize_date(start),
                "版权到期": normalize_date(end),
            }
        )

    for work_id, record in foundation.items():
        add(
            work_id,
            "作者",
            record["作者"],
            "user_confirmed_foundation",
            normalize_author,
            "confirmed_foundation_id",
            "recovered_from_user_foundation",
        )

    for row in read_rows(DIGITAL_LEDGER, "Sheet1", 7):
        register(
            row.get("作品ID"),
            [row.get("出版书名"), row.get("合同书名")],
            row.get("作者署名"),
            row.get("签订日期"),
            row.get("到期时间"),
            "digital_ledger",
        )

    for row in read_rows(ORIGINAL_LIBRARY, "全库排查", 11):
        register(
            row.get("作品ID"),
            [row.get("书名初"), row.get("书名更")],
            row.get("作者笔名"),
            row.get("授权时间"),
            row.get("结束时间"),
            "original_library",
        )

    for sheet_name, start_header, end_header in [
        ("有声读物改编权", "授权时间", "结束时间"),
        ("醉唐有声作品", "授权开始日期", "授权结束时间"),
        ("有声读物改编权的转授权", "授权时间", "结束时间"),
    ]:
        for row in read_rows(ORIGINAL_LIBRARY_2, sheet_name, 9):
            register(
                row.get("作品ID"),
                [row.get("书名初"), row.get("书名更"), row.get("作品名称")],
                row.get("作者笔名"),
                row.get(start_header),
                row.get(end_header),
                f"original_library_2:{sheet_name}",
            )

    for row in read_rows(AUTHORIZATION_SUMMARY, "授权表汇总", 20):
        work_id = target_id(row.get("作品ID"))
        register(
            row.get("作品ID"),
            [row.get("平台作品名称")],
            row.get("作者"),
            row.get("授权开始时间"),
            row.get("授权结束时间"),
            "authorization_summary:detail",
        )
        status = clean(row.get("是否下架"))
        if work_id in final_ids and status:
            explicit_status[work_id].add(status)

    for row in read_rows(AUTHORIZATION_SUMMARY, "授权汇总表-新", 6):
        register(
            row.get("作品ID"),
            [row.get("平台作品名称")],
            row.get("作者"),
            None,
            row.get("我方版权到期时间"),
            "authorization_summary:new",
        )

    for row in read_rows(AUTHORIZATION_RELATIONSHIP, "有声专用", 6):
        register(
            row.get("我方作品ID"),
            [row.get("我方作品名称")],
            row.get("作者名称"),
            None,
            None,
            "authorization_relationship",
        )

    if DUAL_SOURCE_STAGING.exists():
        payload = json.loads(DUAL_SOURCE_STAGING.read_text(encoding="utf-8-sig"))
        field_map = {
            "authorName": ("作者", normalize_author),
            "copyrightStartDate": ("版权开始", normalize_date),
            "copyrightEndDate": ("版权到期", normalize_date),
        }
        for row in payload.get("records", []):
            mapped = field_map.get(clean(row.get("fieldName")))
            if not mapped:
                continue
            field, normalizer = mapped
            add(
                row.get("standardWorkId"),
                field,
                row.get("applyValue"),
                (
                    "legacy_dual_source_user_override"
                    if row.get("resolvedByUserOverride")
                    else "legacy_dual_source_staging"
                ),
                normalizer,
                "legacy_confirmed_identity",
                (
                    "legacy_user_confirmed"
                    if row.get("resolvedByUserOverride")
                    else "legacy_local_staging"
                ),
            )

    title_index: dict[str, list[dict]] = defaultdict(list)
    for entry in source_entries:
        for title in entry["titles"]:
            title_index[title].append(entry)
    for work_id, record in foundation.items():
        title = normalize_title(record["书名"])
        expected_authors = author_tokens(record["作者"])
        exact_entries = [
            entry
            for entry in title_index.get(title, [])
            if not expected_authors
            or not entry["authorTokens"]
            or expected_authors.intersection(entry["authorTokens"])
        ]
        for field in ["版权开始", "版权到期"]:
            if candidates[work_id][field]:
                continue
            values = {entry[field] for entry in exact_entries if entry[field]}
            if len(values) != 1:
                continue
            value = next(iter(values))
            roles = sorted(
                {
                    entry["sourceRole"]
                    for entry in exact_entries
                    if entry[field] == value
                }
            )
            add(
                work_id,
                field,
                value,
                "+".join(roles),
                clean,
                "title_author_exact",
                "regenerated_direct_source",
            )

    priority = [
        "user_confirmed_foundation",
        "legacy_dual_source_user_override",
        "legacy_dual_source_staging",
        "digital_ledger",
        "original_library_2:有声读物改编权",
        "original_library_2:醉唐有声作品",
        "original_library",
        "original_library_2:有声读物改编权的转授权",
        "authorization_summary:new",
        "authorization_summary:detail",
        "authorization_relationship",
    ]

    def choose(work_id: str, field: str) -> tuple[str, dict]:
        values = candidates[work_id][field]
        for role in priority:
            matches = [candidate for candidate in values if candidate["sourceRole"] == role]
            if matches:
                return matches[0]["value"], matches[0]
        if values:
            return values[0]["value"], values[0]
        return "", {
            "sourceRole": "missing",
            "value": "",
            "matchMethod": "unresolved",
            "confirmationMode": "unresolved",
        }

    records = []
    for work_id in sorted(
        final_ids,
        key=lambda value: (0, int(value)) if value.isdigit() else (1, value),
    ):
        foundation_row = foundation[work_id]
        author, author_source = choose(work_id, "作者")
        start, start_source = choose(work_id, "版权开始")
        end, end_source = choose(work_id, "版权到期")
        raw_statuses = explicit_status.get(work_id, set())
        if raw_statuses.intersection({"已下架", "下架", "是", "已解约"}):
            work_status = "已下架"
            work_status_source = build_candidate(
                "authorization_summary:detail",
                work_status,
                "explicit_status",
                "regenerated_direct_source",
            )
        elif "未下架" in raw_statuses:
            work_status = "已上架"
            work_status_source = build_candidate(
                "authorization_summary:detail",
                work_status,
                "explicit_status",
                "regenerated_direct_source",
            )
        else:
            work_status = ""
            work_status_source = build_candidate(
                "missing", "", "unresolved", "unresolved"
            )

        if end == "无限期":
            audio_status = "无限期"
        elif end:
            audio_status = "版权有效" if date.fromisoformat(end) >= as_of else "版权已到期"
        else:
            audio_status = ""
        audio_source = build_candidate(
            end_source["sourceRole"] if end else "missing",
            audio_status,
            "derived_from_confirmed_expiry" if end else "unresolved",
            (
                end_source["confirmationMode"]
                if end
                else "unresolved"
            ),
        )

        unique_values = {
            field: sorted({candidate["value"] for candidate in candidates[work_id][field]})
            for field in CORE_FIELDS
        }
        records.append(
            {
                **foundation_row,
                "作者": author,
                "版权开始": start,
                "版权到期": end,
                "作品状态": work_status,
                "音频版权状态": audio_status,
                "字段来源": {
                    "作者": author_source,
                    "版权开始": start_source,
                    "版权到期": end_source,
                    "作品状态": work_status_source,
                    "音频版权状态": audio_source,
                },
                "候选值数量": {
                    field: len(values) for field, values in unique_values.items()
                },
                "存在冲突": {
                    field: len(values) > 1 for field, values in unique_values.items()
                },
            }
        )

    bill_ids = set()
    for row in read_rows(BILL_PATH, "年月+渠道+作品-渠道月度实销", 7):
        work_id = target_id(row.get("我方作品ID"))
        if work_id:
            bill_ids.add(work_id)

    source_manifest = [
        {"role": role, "sha256": sha256(path)}
        for role, path in source_paths.items()
    ]
    source_manifest.append(
        {"role": "user_confirmed_foundation", "sha256": sha256(args.foundation_workbook)}
    )
    if DUAL_SOURCE_STAGING.exists():
        source_manifest.append(
            {"role": "legacy_dual_source_staging", "sha256": sha256(DUAL_SOURCE_STAGING)}
        )

    payload = {
        "schema": SCHEMA,
        "status": "blocked_incomplete_recovery",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "asOfDate": as_of.isoformat(),
        "artifactRole": "private_local_recovery_candidate",
        "sourceManifest": source_manifest,
        "scope": {
            "foundationWorkCount": len(final_ids),
            "billCanonicalWorkCount": len(bill_ids),
            "billAndFoundationScopeEqual": bill_ids == final_ids,
        },
        "verification": {
            "verified": False,
            "reason": "Content must pass the complete staging contract; presence alone is never sufficient.",
        },
        "records": records,
        "prohibitedActionsConfirmed": {
            "databaseWritten": False,
            "formalMasterDataWritten": False,
            "mappingActivated": False,
            "m3FormalExecutionStarted": False,
        },
    }
    validation = validate_staging_payload(payload, final_ids, require_verified=False)
    payload["verification"]["validation"] = validation
    payload["summary"] = {
        "workCount": len(records),
        "missingByField": validation["missingByField"],
        "workStatusDistribution": validation["workStatusDistribution"],
        "audioRightsStatusDistribution": validation[
            "audioRightsStatusDistribution"
        ],
        "confirmationModeDistribution": validation[
            "confirmationModeDistribution"
        ],
        "issues": validation["issues"],
    }

    output = safe_output(args.output)
    if output.exists() and not args.force:
        raise SystemExit("Private recovery output already exists; use --force only after review.")
    if validation["issues"] and not args.write_incomplete_candidate:
        print(
            json.dumps(
                {
                    "status": "blocked",
                    "written": False,
                    "summary": payload["summary"],
                },
                ensure_ascii=False,
            )
        )
        raise SystemExit(2)

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "status": payload["status"],
                "written": True,
                "artifactRole": payload["artifactRole"],
                "summary": payload["summary"],
                "databaseWritten": False,
                "formalMasterDataWritten": False,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
