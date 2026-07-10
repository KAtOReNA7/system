from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import warnings
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing spreadsheet dependency. Install openpyxl in a temporary dependency "
        "directory and expose it through PYTHONPATH before rerunning."
    ) from exc

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
)


ROOT = Path(__file__).resolve().parents[2]
PRIVATE_ROOT = ROOT / "data" / "private-output"
PRIVATE_DIR = PRIVATE_ROOT / "m2-readiness"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"

TAXONOMY_PATH = (
    ROOT
    / "src"
    / "domain"
    / "oldProductEvaluation"
    / "classificationTaxonomy.v1.json"
)
CORE_STAGING_PATH = PRIVATE_DIR / "M2-five-source-local-staging-apply-result-cn-v1.json"
CLASSIFICATION_STAGING_V1_PATH = (
    PRIVATE_DIR / "M2-classification-aux-tag-local-staging-apply-result-cn-v1.json"
)
CLASSIFICATION_STAGING_V2_PATH = (
    PRIVATE_DIR / "M2-classification-aux-tag-local-staging-apply-result-cn-v2.json"
)
OLD_FILL_PACK_PATH = PRIVATE_DIR / "M2-classification-aux-tag-fill-pack-cn-v2.xlsx"
LATEST_FILL_PACK_PATH = PRIVATE_DIR / "M2-classification-aux-tag-fill-pack-cn-v3.xlsx"
COUNTRY_REVIEW_PATH = PRIVATE_DIR / "M2-country-aux-tag-review-pack-cn-v1.xlsx"
SUPERSEDED_FOUNDATION_PACK_PATH = (
    PRIVATE_DIR / "M2-foundation-data-final-human-review-pack-cn-v1.xlsx"
)

OUTPUT_WORKBOOK_PATH = (
    PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.xlsx"
)
OUTPUT_BASELINE_PATH = (
    PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.json"
)
PUBLIC_REPORT_MD_PATH = (
    DOCS_DIR / "M2-classification-tag-final-foundation-table-precheck-v1.md"
)
PUBLIC_REPORT_JSON_PATH = (
    DOCS_DIR / "M2-classification-tag-final-foundation-table-precheck-v1.json"
)

MAIN_HEADERS = [
    "序号",
    "书名",
    "作品编号",
    "作者",
    "一级分类（最终采用）",
    "二级分类（最终采用）",
    "三级分类（最终采用）",
    "辅助标签（最终采用）",
    "备注",
]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepare the M2 classification/tag final foundation workbook."
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite an existing private workbook. Never use after user filling without a backup.",
    )
    return parser.parse_args()


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_work_id(value) -> str:
    text = clean(value)
    if not text:
        return ""
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def split_tags(value) -> list[str]:
    text = clean(value)
    if not text or text == "无":
        return []
    result = []
    for item in re.split(r"[；;、，,|/]+", text):
        tag = clean(item)
        if tag and tag != "无" and tag not in result:
            result.append(tag)
    return result


def relative(path: Path) -> str:
    return str(path.relative_to(ROOT)).replace("\\", "/")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def work_sort_key(work_id: str) -> tuple[int, int | str]:
    return (0, int(work_id)) if work_id.isdigit() else (1, work_id)


def rows_from_sheet(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Required sheet is missing from {path.name}: {sheet_name}")
        sheet = workbook[sheet_name]
        first_row = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
        )
        headers = [clean(value) for value in first_row]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {
                header: clean(values[index] if index < len(values) else "")
                for index, header in enumerate(headers)
                if header
            }
            if any(row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def load_taxonomy() -> dict:
    return json.loads(TAXONOMY_PATH.read_text(encoding="utf-8"))


def normalize_classification_path(
    taxonomy: dict, level1: str, level2: str, level3: str
) -> tuple[str, str, str]:
    level1 = clean(level1)
    level2 = clean(level2)
    level3 = taxonomy.get("classificationAliases", {}).get(clean(level3), clean(level3))
    for expected_level1, branches in taxonomy["classificationTree"].items():
        if level2 in branches:
            level1 = expected_level1
            break
    return level1, level2, level3


def validate_classification_path(
    taxonomy: dict, level1: str, level2: str, level3: str
) -> bool:
    branches = taxonomy["classificationTree"].get(level1, {})
    return level2 in branches and level3 in branches[level2]


def load_core_identifiers() -> dict[str, dict]:
    payload = json.loads(CORE_STAGING_PATH.read_text(encoding="utf-8"))
    works: dict[str, dict] = {}
    for row in payload.get("records", []):
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        work = works.setdefault(
            work_id,
            {"作品编号": work_id, "书名": "", "作者": ""},
        )
        if clean(row.get("书名")):
            work["书名"] = clean(row.get("书名"))
        if clean(row.get("字段")) == "作者":
            work["作者"] = clean(row.get("采用值"))

    missing_titles = sum(not work["书名"] for work in works.values())
    missing_authors = sum(not work["作者"] for work in works.values())
    if missing_titles or missing_authors:
        raise SystemExit(
            "Core staging identifiers are incomplete: "
            f"missing titles={missing_titles}, missing authors={missing_authors}."
        )
    return works


def load_latest_manual_inputs() -> tuple[dict, dict, dict]:
    classification_rows = rows_from_sheet(
        LATEST_FILL_PACK_PATH, "01_分类需人工确认"
    )
    auxiliary_rows = rows_from_sheet(LATEST_FILL_PACK_PATH, "02_辅助标签需核对")

    classifications = {}
    for row in classification_rows:
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in classifications:
            raise SystemExit("Latest v3 classification pack contains duplicate work IDs.")
        values = (
            clean(row.get("一级分类_请填写")),
            clean(row.get("二级分类_请填写")),
            clean(row.get("三级分类_请填写")),
        )
        if not all(values):
            raise SystemExit("Latest v3 classification pack still has incomplete rows.")
        classifications[work_id] = values

    auxiliary = {}
    for row in auxiliary_rows:
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in auxiliary:
            raise SystemExit("Latest v3 auxiliary-tag pack contains duplicate work IDs.")
        auxiliary[work_id] = split_tags(row.get("辅助标签_请填写"))

    stats = {
        "classificationRows": len(classification_rows),
        "classificationWorks": len(classifications),
        "auxiliaryRows": len(auxiliary_rows),
        "auxiliaryWorks": len(auxiliary),
        "sha256": file_hash(LATEST_FILL_PACK_PATH),
    }
    return classifications, auxiliary, stats


def load_old_pack_stats() -> dict:
    classification_rows = rows_from_sheet(OLD_FILL_PACK_PATH, "01_分类需人工确认")
    auxiliary_rows = rows_from_sheet(OLD_FILL_PACK_PATH, "02_辅助标签需核对")
    classification_filled_ids = {
        normalize_work_id(row.get("作品编号"))
        for row in classification_rows
        if any(
            clean(row.get(field))
            for field in [
                "一级分类_请填写",
                "二级分类_请填写",
                "三级分类_请填写",
            ]
        )
    }
    auxiliary_prefilled_ids = {
        normalize_work_id(row.get("作品编号"))
        for row in auxiliary_rows
        if clean(row.get("辅助标签_请填写"))
    }
    classification_filled_ids.discard("")
    auxiliary_prefilled_ids.discard("")
    return {
        "classificationRows": len(classification_rows),
        "classificationRowsWithAnyFilledValue": len(classification_filled_ids),
        "auxiliaryRows": len(auxiliary_rows),
        "auxiliaryRowsWithPrefilledValue": len(auxiliary_prefilled_ids),
        "supersededBy": LATEST_FILL_PACK_PATH.name,
        "usedAsFinalInput": False,
    }


def load_country_decisions(taxonomy: dict) -> tuple[dict[str, list[str]], dict]:
    rows = rows_from_sheet(COUNTRY_REVIEW_PATH, "01_国家标签需核对")
    country_tags = set(taxonomy["auxiliaryTagGroups"]["国家"])
    adopted_by_work: dict[str, list[str]] = {}
    decision_distribution = Counter()
    adopted_distribution = Counter()

    for row in rows:
        work_id = normalize_work_id(row.get("作品编号"))
        decision = clean(row.get("是否采用（请填写）"))
        if not work_id or decision not in {"采用", "不采用"}:
            raise SystemExit("Country review pack contains an incomplete or invalid decision.")
        decision_distribution[decision] += 1
        if decision == "采用":
            adopted = split_tags(
                row.get("采用值（如需修改）") or row.get("国家标签候选")
            )
            if not adopted:
                raise SystemExit("An adopted country decision has no adopted value.")
            unknown = [tag for tag in adopted if tag not in country_tags]
            if unknown:
                raise SystemExit(
                    "Country review contains values outside the current controlled list: "
                    + "、".join(unknown)
                )
            adopted_by_work[work_id] = adopted
            adopted_distribution.update(adopted)

    return adopted_by_work, {
        "rowCount": len(rows),
        "decisionDistribution": dict(decision_distribution),
        "adoptedWorkCount": len(adopted_by_work),
        "adoptedTagCount": sum(len(tags) for tags in adopted_by_work.values()),
        "adoptedDistribution": dict(adopted_distribution),
        "sha256": file_hash(COUNTRY_REVIEW_PATH),
    }


def load_base_classification_records() -> dict[str, dict]:
    payload = json.loads(CLASSIFICATION_STAGING_V1_PATH.read_text(encoding="utf-8"))
    records = {}
    for row in payload.get("records", []):
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in records:
            raise SystemExit("Classification staging v1 contains duplicate work IDs.")
        records[work_id] = row
    return records


def verify_combined_staging(records: dict[str, dict]) -> None:
    if not CLASSIFICATION_STAGING_V2_PATH.exists():
        return
    payload = json.loads(CLASSIFICATION_STAGING_V2_PATH.read_text(encoding="utf-8"))
    existing = {
        normalize_work_id(row.get("作品编号")): row
        for row in payload.get("records", [])
        if normalize_work_id(row.get("作品编号"))
    }
    if set(existing) != set(records):
        raise SystemExit("Classification staging v2 work scope differs from the rebuilt scope.")
    for work_id, row in records.items():
        old = existing[work_id]
        for field in ["一级分类", "二级分类", "三级分类"]:
            if clean(old.get(field)) != clean(row.get(field)):
                raise SystemExit(f"Classification staging v2 mismatch at {field}.")
        if set(split_tags(old.get("已核对辅助标签"))) != set(
            split_tags(row.get("已核对辅助标签"))
        ):
            raise SystemExit("Classification staging v2 auxiliary-tag mismatch.")


def audit_private_workbooks() -> dict:
    inspected = []
    input_candidates = []
    inspection_errors = []
    for path in sorted(PRIVATE_ROOT.rglob("*.xlsx")):
        if path.name.startswith("~$") or path == OUTPUT_WORKBOOK_PATH:
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            matching_sheets = []
            try:
                for sheet in workbook.worksheets:
                    values = next(
                        sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
                    )
                    headers = {clean(value) for value in values if clean(value)}
                    has_classification_input = bool(
                        headers
                        & {
                            "一级分类_请填写",
                            "二级分类_请填写",
                            "三级分类_请填写",
                            "辅助标签_请填写",
                        }
                    )
                    has_country_input = {
                        "国家标签候选",
                        "是否采用（请填写）",
                    }.issubset(headers)
                    if has_classification_input or has_country_input:
                        matching_sheets.append(sheet.title)
            finally:
                workbook.close()
            inspected.append(relative(path))
            if matching_sheets:
                input_candidates.append(
                    {"path": relative(path), "matchingSheets": matching_sheets}
                )
        except Exception as exc:  # pragma: no cover - environment/file corruption guard
            inspection_errors.append(
                {"path": relative(path), "errorType": type(exc).__name__}
            )

    if inspection_errors:
        raise SystemExit(
            "One or more prior private workbooks could not be inspected: "
            + json.dumps(inspection_errors, ensure_ascii=False)
        )
    return {
        "workbooksInspected": len(inspected),
        "classificationOrTagInputCandidates": input_candidates,
        "inspectionErrors": inspection_errors,
        "latestFinalInputs": [
            relative(LATEST_FILL_PACK_PATH),
            relative(COUNTRY_REVIEW_PATH),
        ],
        "supersededInputs": [
            relative(OLD_FILL_PACK_PATH),
            relative(SUPERSEDED_FOUNDATION_PACK_PATH),
        ],
    }


def build_records() -> tuple[list[dict], dict]:
    taxonomy = load_taxonomy()
    core = load_core_identifiers()
    base = load_base_classification_records()
    manual_classification, manual_auxiliary, latest_stats = load_latest_manual_inputs()
    country_by_work, country_stats = load_country_decisions(taxonomy)
    old_pack_stats = load_old_pack_stats()
    workbook_audit = audit_private_workbooks()

    if set(core) != set(base):
        raise SystemExit("Core staging and classification staging do not cover the same works.")

    all_allowed_tags = {
        tag
        for tags in taxonomy["auxiliaryTagGroups"].values()
        for tag in tags
    }
    records = []
    carried_manual_classification = 0
    carried_manual_auxiliary = 0
    issues = []

    for work_id in sorted(core, key=work_sort_key):
        source = base[work_id]
        level1 = clean(source.get("一级分类"))
        level2 = clean(source.get("二级分类"))
        level3 = clean(source.get("三级分类"))
        if not validate_classification_path(taxonomy, level1, level2, level3):
            issues.append({"type": "invalidClassificationPath", "workId": work_id})

        if work_id in manual_classification:
            normalized_manual = normalize_classification_path(
                taxonomy, *manual_classification[work_id]
            )
            if normalized_manual != (level1, level2, level3):
                issues.append({"type": "manualClassificationNotCarried", "workId": work_id})
            else:
                carried_manual_classification += 1

        tags = split_tags(source.get("已核对辅助标签"))
        if work_id in manual_auxiliary:
            if set(tags) != set(manual_auxiliary[work_id]):
                issues.append({"type": "manualAuxiliaryNotCarried", "workId": work_id})
            else:
                carried_manual_auxiliary += 1
        for tag in country_by_work.get(work_id, []):
            if tag not in tags:
                tags.append(tag)
        invalid_tags = [tag for tag in tags if tag not in all_allowed_tags]
        if invalid_tags:
            issues.append({"type": "invalidAuxiliaryTag", "workId": work_id})

        records.append(
            {
                "作品编号": work_id,
                "书名": core[work_id]["书名"],
                "作者": core[work_id]["作者"],
                "一级分类": level1,
                "二级分类": level2,
                "三级分类": level3,
                "辅助标签": "；".join(tags) if tags else "无",
                "备注": "",
            }
        )

    if issues:
        counts = dict(Counter(issue["type"] for issue in issues))
        raise SystemExit(
            "Final classification/tag foundation validation failed: "
            + json.dumps(counts, ensure_ascii=False)
        )

    rebuilt_for_v2_check = {}
    for record in records:
        rebuilt_for_v2_check[record["作品编号"]] = {
            "作品编号": record["作品编号"],
            "一级分类": record["一级分类"],
            "二级分类": record["二级分类"],
            "三级分类": record["三级分类"],
            "已核对辅助标签": record["辅助标签"],
        }
    verify_combined_staging(rebuilt_for_v2_check)

    tag_lists = [split_tags(record["辅助标签"]) for record in records]
    summary = {
        "schema": "m2.classification_tag_final_foundation_precheck.v1",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "taxonomyVersion": taxonomy["version"],
        "scope": "classification_and_auxiliary_tags_only",
        "workCount": len(records),
        "level1Distribution": dict(Counter(row["一级分类"] for row in records)),
        "validClassificationPaths": sum(
            validate_classification_path(
                taxonomy, row["一级分类"], row["二级分类"], row["三级分类"]
            )
            for row in records
        ),
        "taggedWorkCount": sum(bool(tags) for tags in tag_lists),
        "tagAssignmentCount": sum(len(tags) for tags in tag_lists),
        "manualCarryForward": {
            "latestClassificationWorks": latest_stats["classificationWorks"],
            "classificationWorksCarried": carried_manual_classification,
            "latestAuxiliaryWorks": latest_stats["auxiliaryWorks"],
            "auxiliaryWorksCarried": carried_manual_auxiliary,
            "countryReviewRows": country_stats["rowCount"],
            "countryDecisionsComplete": sum(
                country_stats["decisionDistribution"].values()
            )
            == country_stats["rowCount"],
            "countryTagsAdopted": country_stats["adoptedTagCount"],
        },
        "sourceAudit": workbook_audit,
        "oldV2Pack": old_pack_stats,
        "quality": {
            "uniqueWorkIds": len({row["作品编号"] for row in records}),
            "missingTitles": sum(not row["书名"] for row in records),
            "missingAuthors": sum(not row["作者"] for row in records),
            "missingClassificationFields": sum(
                not all([row["一级分类"], row["二级分类"], row["三级分类"]])
                for row in records
            ),
            "invalidClassificationPaths": 0,
            "invalidAuxiliaryTags": 0,
            "priorConfirmedItemsReopened": 0,
        },
        "boundaries": {
            "copyrightFieldsIncluded": False,
            "authorFieldEditableAsMasterData": False,
            "statusFieldsIncluded": False,
            "formalMasterDataWritten": False,
            "databaseWritten": False,
            "m3FormalExecutionStarted": False,
            "privateOutputsGitignored": True,
        },
        "outputs": {
            "privateWorkbook": relative(OUTPUT_WORKBOOK_PATH),
            "privateBaselineJson": relative(OUTPUT_BASELINE_PATH),
            "publicReportMarkdown": relative(PUBLIC_REPORT_MD_PATH),
            "publicReportJson": relative(PUBLIC_REPORT_JSON_PATH),
        },
        "inputs": {
            "latestFillPackSha256": latest_stats["sha256"],
            "countryReviewSha256": country_stats["sha256"],
            "classificationStagingSha256": file_hash(CLASSIFICATION_STAGING_V1_PATH),
            "coreIdentifierStagingSha256": file_hash(CORE_STAGING_PATH),
        },
    }
    return records, summary


def style_sheet(sheet, editable_columns: set[int] | None = None) -> None:
    editable_columns = editable_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    fixed_fill = PatternFill("solid", fgColor="F2F2F2")
    border_side = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=border_side, right=border_side, top=border_side, bottom=border_side
        )
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side,
            )
            if cell.column in editable_columns:
                cell.fill = editable_fill
            elif sheet.title in {"01_出版物", "02_网文"}:
                cell.fill = fixed_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def write_workbook(records: list[dict], summary: dict) -> None:
    taxonomy = load_taxonomy()
    workbook = Workbook()
    readme = workbook.active
    readme.title = "00_填写说明"
    readme_rows = [
        ["项目", "说明"],
        ["这张表做什么", "一次性确认 3053 部作品的一级、二级、三级分类和辅助标签。"],
        ["先看哪里", "先看 01_出版物，再看 02_网文；可用筛选和搜索定位作品。"],
        ["如何修改", "分类或标签有误时，直接修改黄色列中的最终采用值；没有辅助标签填写“无”。"],
        ["辅助标签格式", "多个标签请用中文分号“；”分隔；只填写 04_分类标签选项 中允许的值。"],
        ["如何补遗漏", "直接在辅助标签最终采用列补写；备注列可说明原因。"],
        ["如何完成", "全部检查后，在 03_整体确认 填写“确认作为后续基础表格”。"],
        ["已自动带入", "此前已经人工确认的分类、辅助标签和国家标签结果均已带入，不需要重复抄写。"],
        ["不在本表范围", "版权、作者、作品状态、音频版权状态均已收口，本表不重新核对。作者只用于识别作品。"],
        ["当前边界", "本表是本地私有基础候选，不写正式主数据，不进入下一阶段正式开发。"],
    ]
    for row in readme_rows:
        readme.append(row)
    style_sheet(readme)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 90
    readme.freeze_panes = "A2"

    level2_values = [
        level2
        for branches in taxonomy["classificationTree"].values()
        for level2 in branches
    ]

    for level1, sheet_name in [("出版物", "01_出版物"), ("网文", "02_网文")]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(MAIN_HEADERS)
        subset = [row for row in records if row["一级分类"] == level1]
        for index, row in enumerate(subset, 1):
            sheet.append(
                [
                    index,
                    row["书名"],
                    row["作品编号"],
                    row["作者"],
                    row["一级分类"],
                    row["二级分类"],
                    row["三级分类"],
                    row["辅助标签"],
                    row["备注"],
                ]
            )
        level1_validation = DataValidation(
            type="list", formula1='"出版物,网文"', allow_blank=False
        )
        level2_validation = DataValidation(
            type="list",
            formula1='"' + ",".join(level2_values) + '"',
            allow_blank=False,
        )
        sheet.add_data_validation(level1_validation)
        sheet.add_data_validation(level2_validation)
        level1_validation.add(f"E2:E{sheet.max_row}")
        level2_validation.add(f"F2:F{sheet.max_row}")
        style_sheet(sheet, editable_columns={5, 6, 7, 8, 9})
        widths = {
            "A": 8,
            "B": 34,
            "C": 16,
            "D": 22,
            "E": 18,
            "F": 20,
            "G": 24,
            "H": 34,
            "I": 30,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.auto_filter.ref = f"A1:I{sheet.max_row}"
        sheet.print_title_rows = "1:1"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

    confirm = workbook.create_sheet("03_整体确认")
    confirm.append(["确认项目", "当前状态", "请填写最终结论", "说明"])
    confirm.append(
        [
            "分类与标签基础表整体确认",
            (
                f"共 {summary['workCount']} 部：出版物 "
                f"{summary['level1Distribution'].get('出版物', 0)} 部，网文 "
                f"{summary['level1Distribution'].get('网文', 0)} 部"
            ),
            "",
            "全部核对完成后填写：确认作为后续基础表格；若还没完成填写：仍需修改。",
        ]
    )
    conclusion_validation = DataValidation(
        type="list",
        formula1='"确认作为后续基础表格,仍需修改"',
        allow_blank=True,
    )
    confirm.add_data_validation(conclusion_validation)
    conclusion_validation.add("C2")
    style_sheet(confirm, editable_columns={3})
    for column, width in {"A": 30, "B": 46, "C": 28, "D": 72}.items():
        confirm.column_dimensions[column].width = width

    options = workbook.create_sheet("04_分类标签选项")
    options.append(["一级分类", "二级分类", "三级分类", "辅助标签组", "辅助标签"])
    classification_rows = []
    for level1, branches in taxonomy["classificationTree"].items():
        for level2, level3_values in branches.items():
            for level3 in level3_values:
                classification_rows.append([level1, level2, level3])
    auxiliary_rows = []
    for group, tags in taxonomy["auxiliaryTagGroups"].items():
        for tag in tags:
            auxiliary_rows.append([group, tag])
    row_count = max(len(classification_rows), len(auxiliary_rows))
    for index in range(row_count):
        classification = classification_rows[index] if index < len(classification_rows) else ["", "", ""]
        auxiliary = auxiliary_rows[index] if index < len(auxiliary_rows) else ["", ""]
        options.append(classification + auxiliary)
    style_sheet(options)
    for column, width in {"A": 18, "B": 22, "C": 28, "D": 20, "E": 28}.items():
        options.column_dimensions[column].width = width

    workbook.properties.title = "M2 分类与标签最终基础大表"
    workbook.properties.subject = "出版物与网文分类、辅助标签一次性人工确认"
    workbook.properties.creator = "KAtOReNA7/system local development"
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_WORKBOOK_PATH)


def write_private_baseline(records: list[dict], summary: dict) -> None:
    payload = {
        "summary": summary,
        "records": records,
        "completion": {
            "workbookGlobalConfirmation": "",
            "acceptedAsFutureLocalFoundation": False,
        },
    }
    OUTPUT_BASELINE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def write_public_report(summary: dict) -> None:
    public_payload = {
        key: summary[key]
        for key in [
            "schema",
            "generatedAt",
            "taxonomyVersion",
            "scope",
            "workCount",
            "level1Distribution",
            "validClassificationPaths",
            "taggedWorkCount",
            "tagAssignmentCount",
            "manualCarryForward",
            "sourceAudit",
            "oldV2Pack",
            "quality",
            "boundaries",
            "outputs",
        ]
    }
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_REPORT_JSON_PATH.write_text(
        json.dumps(public_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    source_audit = summary["sourceAudit"]
    carry = summary["manualCarryForward"]
    quality = summary["quality"]
    markdown = f"""# M2 分类与标签最终基础大表预检 v1

## 结论

- 本轮只处理分类与辅助标签，不重新打开作者、版权、作品状态或音频版权状态。
- 已生成一张覆盖 `{summary['workCount']}` 部作品的完整基础大表，按 `出版物` 和 `网文` 分为两个主阅读页。
- 出版物 `{summary['level1Distribution'].get('出版物', 0)}` 部，网文 `{summary['level1Distribution'].get('网文', 0)}` 部；分类路径有效 `{summary['validClassificationPaths']}/{summary['workCount']}`。
- 旧包中已完成的人工结果已自动前置带入；没有把已确认内容重新生成成待补项。
- 当前表是 private 本地基础候选；用户整体确认前，不写正式主数据，也不进入 M3 formal execution。

## 历史人工结果继承

- 最新 v3 人工分类：`{carry['classificationWorksCarried']}/{carry['latestClassificationWorks']}` 部已带入。
- 最新 v3 辅助标签：`{carry['auxiliaryWorksCarried']}/{carry['latestAuxiliaryWorks']}` 部已带入。
- 国家标签核对：`{carry['countryReviewRows']}` 条决定均完整，采用标签 `{carry['countryTagsAdopted']}` 个。
- 旧 v2 是中间候选/预填版本，已由 v3 取代，不作为最终人工输入，也不会重新制造待办。
- 旧“基础数据最终核对包”已废弃，本轮不会读取其中的版权待办作为当前阻断。

## 全量表质量

- 唯一作品编号：`{quality['uniqueWorkIds']}`。
- 缺书名：`{quality['missingTitles']}`；缺作者：`{quality['missingAuthors']}`。
- 缺分类字段：`{quality['missingClassificationFields']}`；无效分类路径：`{quality['invalidClassificationPaths']}`。
- 含辅助标签作品：`{summary['taggedWorkCount']}` 部；标签赋值：`{summary['tagAssignmentCount']}` 个。
- 重新打开既有人工确认项：`{quality['priorConfirmedItemsReopened']}`。

## 历史文件检查

- 已检查 private 历史 Excel：`{source_audit['workbooksInspected']}` 个。
- 最终采用来源仅为最新完整 v3 分类/标签核对包与已完成的国家标签核对包。
- 其他旧包仅用于版本追溯，不覆盖较新的人工决定。

## 用户填写方式

1. 在 `01_出版物` 和 `02_网文` 中直接修改黄色的一级、二级、三级分类、辅助标签或备注。
2. 没有辅助标签时保留“无”；多个辅助标签使用中文分号分隔。
3. 全部核对后，在 `03_整体确认` 选择“确认作为后续基础表格”。
4. 下一轮以该表和私有基线做差异校验；确认通过后，分类与标签人工字段可在本地候选层收口。

## 边界

- 作者只作为作品识别信息展示，不在本表修改主数据。
- 本表不包含版权日期、收入、作品状态、音频版权状态。
- private Excel/JSON 位于 Git 忽略范围，禁止提交。
- 本轮未写数据库、未写正式主数据、未进入 M3。
"""
    PUBLIC_REPORT_MD_PATH.write_text(markdown, encoding="utf-8")


def main() -> None:
    args = parse_arguments()
    required = [
        TAXONOMY_PATH,
        CORE_STAGING_PATH,
        CLASSIFICATION_STAGING_V1_PATH,
        OLD_FILL_PACK_PATH,
        LATEST_FILL_PACK_PATH,
        COUNTRY_REVIEW_PATH,
    ]
    missing = [relative(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit("Required local inputs are missing: " + ", ".join(missing))
    if OUTPUT_WORKBOOK_PATH.exists() and not args.force:
        raise SystemExit(
            "The private foundation workbook already exists. Refusing to overwrite possible "
            "user edits; use --force only when an explicit regeneration is intended."
        )

    records, summary = build_records()
    write_workbook(records, summary)
    write_private_baseline(records, summary)
    write_public_report(summary)
    print(
        json.dumps(
            {
                "状态": "已生成",
                "覆盖作品": summary["workCount"],
                "出版物": summary["level1Distribution"].get("出版物", 0),
                "网文": summary["level1Distribution"].get("网文", 0),
                "分类路径有效": summary["validClassificationPaths"],
                "已有人工结果重复待办": summary["quality"]["priorConfirmedItemsReopened"],
                "输出表格": relative(OUTPUT_WORKBOOK_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
