from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


CANDIDATE_VERSION = "m2-realdata-dev-rating-standard-v4.2-revenue-signal-aligned"
BUYOUT_AMORTIZATION_YEARS = 3

DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-business-review"
STAGING_TABLE = ROOT / "data" / "private-output" / "m1-master-data" / "M1-dual-source-limited-staging-table-v1.json"

INPUT_V3_PACK = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v3.xlsx"
INPUT_V3_VALIDATION_JSON = DOCS_DIR / "M2-rating-standard-v3-operator-validation-summary-v1.json"
INPUT_V3_VALIDATION_MD = DOCS_DIR / "M2-rating-standard-v3-operator-validation-summary-v1.md"
OUTPUT_V4_PACK = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v4.2.xlsx"

V3_FAILURE_JSON = DOCS_DIR / "M2-rating-standard-v3-targeted-failure-analysis.json"
V3_FAILURE_MD = DOCS_DIR / "M2-rating-standard-v3-targeted-failure-analysis.md"
V4_CORRECTION_JSON = DOCS_DIR / "M2-rating-standard-v4.2-targeted-correction-summary.json"
V4_CORRECTION_MD = DOCS_DIR / "M2-rating-standard-v4.2-targeted-correction-summary.md"
V4_TASK_JSON = DOCS_DIR / "M2-rating-standard-v4.2-task-pack-summary.json"
V4_TASK_MD = DOCS_DIR / "M2-rating-standard-v4.2-task-pack-summary.md"

REQUIRED_INPUTS = [
    INPUT_V3_PACK,
    INPUT_V3_VALIDATION_JSON,
    INPUT_V3_VALIDATION_MD,
    DOCS_DIR / "M2-revenue-model-business-rule-alignment-v1.md",
    DOCS_DIR / "M2-revenue-model-business-rule-alignment-v1.json",
    DOCS_DIR / "M2-shelf-status-business-rule-alignment-v1.md",
    DOCS_DIR / "M2-shelf-status-business-rule-alignment-v1.json",
    DOCS_DIR / "M2-front-rating-simplification-v1.md",
    DOCS_DIR / "M2-front-rating-simplification-v1.json",
    DOCS_DIR / "M2-rating-standard-v3-task-pack-summary.md",
    DOCS_DIR / "M2-rating-standard-v3-task-pack-summary.json",
    STAGING_TABLE,
    ROOT / "src" / "domain" / "oldProductEvaluation" / "revenueModelClassifier.js",
    ROOT / "src" / "domain" / "oldProductEvaluation" / "shelfStatusInference.js",
    ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
    ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
    ROOT / "package.json",
    ROOT / ".gitignore",
]

RATINGS = ["S+", "S", "A", "B", "C", "D", "E"]
RANK = {rating: index for index, rating in enumerate(RATINGS)}

USER_JUDGMENT_OPTIONS = ["合理", "基本合理", "不确定", "不合理", "不适用"]
M4_OPTIONS = ["是", "否", "待定"]


def main() -> None:
    ensure_inputs()
    v3_rows = read_v3_task_pack()
    v3_validation = read_json(INPUT_V3_VALIDATION_JSON)

    v4_rows = [build_v4_row(row, index) for index, row in enumerate(v3_rows, start=1)]
    failure_payload = build_v3_failure_analysis(v3_validation)
    correction_payload = build_v4_correction_summary(v3_rows, v4_rows, v3_validation)
    task_payload = build_v4_task_summary(v4_rows)

    write_private_pack(v4_rows, failure_payload, correction_payload, task_payload)
    write_report(V3_FAILURE_JSON, V3_FAILURE_MD, "m2.rating_standard_v3_targeted_failure_analysis", failure_payload, render_v3_failure_md)
    write_report(V4_CORRECTION_JSON, V4_CORRECTION_MD, "m2.rating_standard_v4_2_targeted_correction_summary", correction_payload, render_v4_correction_md)
    write_report(V4_TASK_JSON, V4_TASK_MD, "m2.rating_standard_v4_2_task_pack_summary", task_payload, render_v4_task_md)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "privateTaskPack": rel(OUTPUT_V4_PACK),
                "rows": task_payload["rows"],
                "sampleRows": task_payload["sampleRows"],
                "pendingUserSpecifiedRows": task_payload["pendingUserSpecifiedRows"],
                "ratingDistribution": task_payload["ratingDistribution"],
                "ratingBasisDistribution": task_payload["ratingBasisDistribution"],
                "statusConfidenceDistribution": task_payload["statusConfidenceDistribution"],
                "automaticOperatingSuggestionMainFieldRemoved": True,
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    missing = [path for path in REQUIRED_INPUTS if not path.exists()]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def read_v3_task_pack() -> list[dict]:
    wb = load_workbook(INPUT_V3_PACK, read_only=True, data_only=True)
    if "01_运营任务卡" not in wb.sheetnames:
        raise SystemExit("Missing sheet 01_运营任务卡 in v3 task pack")
    ws = wb["01_运营任务卡"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    result = []
    for values in rows[1:]:
        if not any(clean(value) for value in values):
            continue
        item = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        result.append(item)
    return result


def build_v4_row(row: dict, index: int) -> dict:
    standard_work_id = clean(row.get("standard_work_id"))
    has_task_card = bool(clean(row.get("作品名")) or clean(row.get("作者")) or clean(row.get("收入模式")))
    if not has_task_card:
        return build_pending_user_row(standard_work_id, index)

    revenue_model_label = clean(row.get("收入模式"))
    revenue_model = revenue_model_code(revenue_model_label)
    old_rating = clean(row.get("评级"))
    old_basis = clean(row.get("评级依据"))
    status = infer_v4_status(row)

    sales_monthly_amount = monthly_sales_amount(row)
    sales_rating = rating_from_sales_amount(sales_monthly_amount) if sales_monthly_amount is not None else normalize_rating(row.get("内部辅助：实销评级"))
    buyout_amount = number_or_none(row.get("买断估计金额"))
    buyout_years = resolve_buyout_amortization_years(row) if buyout_amount is not None else None
    buyout_months = round_float(buyout_years * 12, 2) if buyout_years else None
    buyout_monthly = round_float(buyout_amount / buyout_months, 2) if buyout_amount is not None and buyout_months else None
    buyout_rating = rating_from_sales_amount(buyout_monthly) if buyout_monthly is not None else "not_applicable"
    historical_rating = normalize_rating(row.get("内部辅助：历史表现评级")) or old_rating or "E"

    if revenue_model == "pure_buyout":
        rating = buyout_rating if buyout_rating != "not_applicable" else historical_rating
        basis = "buyout_monthly_sales_equivalent"
        explanation = (
            "该作品为纯买断；评级按最新买断金额在业务周期内折算的月均实销等价值判断，"
            f"默认折算年限={BUYOUT_AMORTIZATION_YEARS} 年，并受剩余版权期限制，最低 1 年；"
            f"买断折算月均实销={display_number(buyout_monthly)}；"
            "下一周期可用上一个买断周期的月均实销作参考，当前运营仍取决于版权/上架状态。"
        )
    elif revenue_model == "buyout_plus_sales":
        combined_monthly = (sales_monthly_amount or 0) + (buyout_monthly or 0)
        rating = rating_from_sales_amount(combined_monthly) if combined_monthly > 0 else historical_rating
        basis = "current_sales_with_buyout_allocation"
        explanation = (
            f"买断+实销前台评级将当前实销月均={display_number(sales_monthly_amount)}"
            f"与买断折算月均={display_number(buyout_monthly)}相加，合计月均={display_number(combined_monthly)}；"
            "下一周期只预测实销部分，不预测未来买断谈判。"
        )
    elif revenue_model == "pure_sales_share":
        rating = sales_rating if sales_rating in RANK else historical_rating
        basis = "current_sales" if not status["blocksCurrentOperation"] else "historical"
        explanation = (
            f"纯实销/分成样本继续使用用户月均实销档位，当前实销评级={sales_rating or '缺失'}；"
            "本轮不调整纯实销阈值。"
        )
    else:
        rating = historical_rating
        basis = "historical"
        explanation = "收入模式不稳定或缺少可用口径，前台展示历史评级并要求人工复核。"

    risk_prompt = build_risk_prompt(row, revenue_model, status, old_basis)
    m4_reason = build_m4_reason(row, revenue_model, old_basis, basis, old_rating, rating, status)

    return {
        "standard_work_id": standard_work_id,
        "作品名": clean(row.get("作品名")),
        "作者": clean(row.get("作者")),
        "收入模式": revenue_model_label,
        "货架/版权状态": status["label"],
        "状态置信度": status["confidence"],
        "最近12月实销收入": row.get("最近12月实销收入"),
        "年化实销收入": row.get("年化实销收入"),
        "实销月均收入": sales_monthly_amount,
        "买断估计金额": buyout_amount,
        "买断折算年限": buyout_years,
        "买断折算周期（月）": buyout_months,
        "买断折算月均实销": buyout_monthly,
        "评级": rating,
        "评级依据": basis,
        "评级是否含买断": "是" if revenue_model in {"pure_buyout", "buyout_plus_sales"} and buyout_monthly else "否",
        "评级说明": explanation,
        "风险/复核提示": risk_prompt,
        "不自动给运营建议原因": "M2 v4.2 仍不输出自动运营建议；仅保留风险/复核提示和 M4 校准候选字段。",
        "是否建议进入M4校准案例": "待定",
        "M4校准原因": m4_reason,
        "用户判断：收入模式是否合理": "",
        "用户判断：下架/版权状态是否合理": "",
        "用户判断：评级是否合理": "",
        "用户判断：是否应进入M4": "",
        "用户备注": "",
        "辅助原始：v3评级": old_rating,
        "辅助原始：v3评级依据": old_basis,
        "辅助原始：实销评级": sales_rating,
        "辅助原始：买断历史价值评级": buyout_rating,
    }


def build_pending_user_row(standard_work_id: str, index: int) -> dict:
    return {
        "standard_work_id": standard_work_id,
        "作品名": "",
        "作者": "",
        "收入模式": "用户指定作品待补任务卡",
        "货架/版权状态": "待补任务卡",
        "状态置信度": "低",
        "最近12月实销收入": "",
        "年化实销收入": "",
        "实销月均收入": "",
        "买断估计金额": "",
        "买断折算年限": "",
        "买断折算周期（月）": "",
        "买断折算月均实销": "",
        "评级": "",
        "评级依据": "pending_user_specified",
        "评级是否含买断": "",
        "评级说明": f"用户指定作品第 {index} 行仅保留 standard_work_id，后续单独补生成任务卡。",
        "风险/复核提示": "待补任务卡后再复核。",
        "不自动给运营建议原因": "M2 v4.2 不输出自动运营建议。",
        "是否建议进入M4校准案例": "待定",
        "M4校准原因": "用户指定作品待补任务卡",
        "用户判断：收入模式是否合理": "",
        "用户判断：下架/版权状态是否合理": "",
        "用户判断：评级是否合理": "",
        "用户判断：是否应进入M4": "",
        "用户备注": "",
        "辅助原始：v3评级": "",
        "辅助原始：v3评级依据": "",
        "辅助原始：实销评级": "",
        "辅助原始：买断历史价值评级": "",
    }


def infer_v4_status(row: dict) -> dict:
    old_status = clean(row.get("货架/版权状态"))
    user_judgment = clean(row.get("用户判断：下架状态是否合理"))
    sales_amount = monthly_sales_amount(row) or 0

    if user_judgment == "不合理":
        return {
            "label": "需人工确认：用户反馈下架/版权状态不合理",
            "confidence": "低",
            "blocksCurrentOperation": True,
        }
    if "尾部收入" in old_status:
        return {"label": "版权台账状态优先，尾部收入仅作核查线索", "confidence": "高", "blocksCurrentOperation": True}
    if "版权到期" in old_status or "到期" in old_status:
        return {"label": "版权台账显示到期/下架", "confidence": "高", "blocksCurrentOperation": True}
    if "仍在架" in old_status or "可运营" in old_status:
        return {"label": "可运营或在架（推断）", "confidence": "中", "blocksCurrentOperation": False}
    if "无法判断" in old_status or not old_status:
        return {"label": "无法判断", "confidence": "低", "blocksCurrentOperation": True}
    return {"label": "需人工确认", "confidence": "低", "blocksCurrentOperation": True}


def build_risk_prompt(row: dict, revenue_model: str, status: dict, old_basis: str) -> str:
    prompts = []
    existing = clean(row.get("风险/复核提示"))
    if existing:
        prompts.append(existing)
    if status["confidence"] == "低" or status["blocksCurrentOperation"]:
        prompts.append("货架/版权状态需人工复核，不能直接进入当前运营动作。")
    if revenue_model == "pure_buyout":
        prompts.append("纯买断评级为历史价值折算，不代表当前持续实销。")
    if revenue_model == "buyout_plus_sales":
        prompts.append("买断+实销需复核当前实销与买断历史价值的权重。")
    if old_basis == "buyout_value":
        prompts.append("v3 使用买断总额口径，本轮已改为买断折算月均实销等价值。")
    return "；".join(dict.fromkeys(prompts)) or "保留常规复核。"


def build_m4_reason(row: dict, revenue_model: str, old_basis: str, basis: str, old_rating: str, rating: str, status: dict) -> str:
    reasons = []
    if clean(row.get("用户判断：评级是否合理")) == "不合理":
        reasons.append("v3 用户反馈评级不合理")
    if clean(row.get("用户判断：下架状态是否合理")) == "不合理":
        reasons.append("v3 用户反馈货架/版权状态不合理")
    if revenue_model == "pure_buyout":
        reasons.append("纯买断需校准买断历史价值折算口径")
    if revenue_model == "buyout_plus_sales":
        reasons.append("买断+实销需校准单一前台评级合成口径")
    if old_basis != basis or old_rating != rating:
        reasons.append("v4.2 规则导致评级或依据变化")
    if status["confidence"] == "低":
        reasons.append("状态置信度低")
    return "；".join(dict.fromkeys(reasons)) or "v4.2 抽检样本"


def should_enter_m4(row: dict, old_rating: str, rating: str, status: dict, revenue_model: str) -> bool:
    return (
        clean(row.get("用户判断：评级是否合理")) == "不合理"
        or clean(row.get("用户判断：下架状态是否合理")) == "不合理"
        or old_rating != rating
        or status["confidence"] == "低"
        or revenue_model in {"pure_buyout", "buyout_plus_sales"}
    )


def build_v3_failure_analysis(v3_validation: dict) -> dict:
    rating = v3_validation.get("rating", {})
    shelf = v3_validation.get("shelfStatus", {})
    revenue = v3_validation.get("revenueMode", {})
    return {
        "candidateVersion": "m2-realdata-dev-rating-standard-v3.0",
        "sourceWorkbook": rel(INPUT_V3_PACK),
        "effectiveReviewRows": v3_validation.get("rows", {}).get("evaluatedTaskRows"),
        "revenueModeReasonableRate": revenue.get("judgment", {}).get("positiveRateOfEffectiveAnswered"),
        "shelfStatusReasonableRate": shelf.get("judgment", {}).get("positiveRateOfEffectiveAnswered"),
        "ratingReasonableRate": rating.get("judgment", {}).get("positiveRateOfEffectiveAnswered"),
        "pureSalesConclusion": {
            "pureSalesDirectMatchRate": rating.get("salesBandComparison", {}).get("pureSalesDirectMatchRate"),
            "action": "保持纯实销/分成实销档位不变",
        },
        "pureBuyoutFailure": {
            "rows": rating.get("pureBuyout", {}).get("rows"),
            "userMarkedUnreasonable": rating.get("pureBuyout", {}).get("userRatingNegative"),
            "v3Basis": "buyout_value",
            "v3AmountBandMatchRate": rating.get("pureBuyout", {}).get("buyoutAmountBandMatchRate"),
            "rootCauses": [
                "v3 将一次性买断金额直接按前台价值展示，容易推高到 S+/S。",
                "纯买断缺少当前持续实销能力，应把买断金额折为对应周期月均实销等价值。",
                "买断金额需要按默认 3 年、剩余版权期封顶、最低 1 年折算，并与当前权利/上架状态分开解释。",
                "部分样本同时存在到期、下架或状态不明，影响用户对评级合理性的判断。",
            ],
        },
        "buyoutPlusSalesFailure": {
            "rows": rating.get("buyoutPlusSales", {}).get("rows"),
            "userMarkedUnreasonable": rating.get("buyoutPlusSales", {}).get("userRatingNegative"),
            "rootCauses": [
                "v3 综合评级未清楚区分当前评级和下一周期预测。",
                "买断+实销当前评级应把买断对应周期月均与当周期原实销相加。",
                "买断+实销下一周期预测只预测实销部分，不预测未来买断谈判。",
                "缺少对下架/版权状态置信度的显式提示。",
            ],
        },
        "shelfStatusFailure": {
            "userMarkedUnreasonable": shelf.get("judgment", {}).get("negative"),
            "negativeReasons": shelf.get("negativeReasons", {}),
            "rootCauses": [
                "版权台账由版权部负责，作为高可信状态来源。",
                "尾部收入说明收入来源可后续核查，但不应反向约束版权台账状态。",
                "缺少明确上架字段时，不能高置信判断当前可运营或已下架。",
                "版权未到期但近期无收入不应直接判下架。",
            ],
        },
        "sanitized": True,
        "m3Entered": False,
    }


def build_v4_correction_summary(v3_rows: list[dict], v4_rows: list[dict], v3_validation: dict) -> dict:
    sample_rows = non_pending(v4_rows)
    comparable_pairs = [
        (old, new)
        for old, new in zip(v3_rows, v4_rows)
        if clean(old.get("作品名")) and clean(new.get("评级"))
    ]
    changed_pairs = [(old, new) for old, new in comparable_pairs if clean(old.get("评级")) != clean(new.get("评级"))]
    pure_buyout_changed = [
        (old, new)
        for old, new in changed_pairs
        if revenue_model_code(clean(old.get("收入模式"))) == "pure_buyout"
    ]
    mixed_changed = [
        (old, new)
        for old, new in changed_pairs
        if revenue_model_code(clean(old.get("收入模式"))) == "buyout_plus_sales"
    ]
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "pureSalesRuleChanged": False,
        "pureSalesRule": "保持用户月均实销档位阈值，不做整体重构。",
        "pureBuyoutRule": {
            "ratingBasis": "buyout_monthly_sales_equivalent",
            "amortizationYears": BUYOUT_AMORTIZATION_YEARS,
            "configurable": True,
            "formula": "buyoutEquivalentMonthlySales = buyoutEstimatedAmount / (min(default_3_years, remaining_copyright_years, floor_1_year) * 12)",
            "explanation": "纯买断当前评级使用买断折算月均实销等价值；下一周期可参考上一个买断周期月均。",
        },
        "mixedRule": {
            "ratingBasis": "current_sales_with_buyout_allocation",
            "primarySignal": "current_sales_plus_current_cycle_buyout_monthly_allocation",
            "buyoutAdjustment": "当前评级将买断对应周期月均与当前实销相加；下一周期只预测实销部分。",
        },
        "shelfStatusRule": {
            "expiredSignal": "trusted_copyright_ledger_status",
            "tailRevenueHandling": "尾部收入只作为后续运营核查线索，不反向改写版权台账状态",
            "confidenceDisplayed": True,
            "activeRightsZeroRevenue": "不直接判下架，按版权台账有效状态保留可运营/在架推断。",
        },
        "v3Metrics": {
            "ratingReasonableRate": v3_validation.get("rating", {}).get("judgment", {}).get("positiveRateOfEffectiveAnswered"),
            "pureBuyoutNegativeRows": v3_validation.get("rating", {}).get("pureBuyout", {}).get("userRatingNegative"),
            "buyoutPlusSalesNegativeRows": v3_validation.get("rating", {}).get("buyoutPlusSales", {}).get("userRatingNegative"),
            "shelfStatusNegativeRows": v3_validation.get("shelfStatus", {}).get("judgment", {}).get("negative"),
        },
        "v4Metrics": {
            "sampleRows": len(sample_rows),
            "ratingChangedRows": len(changed_pairs),
            "pureBuyoutRatingChangedRows": len(pure_buyout_changed),
            "buyoutPlusSalesRatingChangedRows": len(mixed_changed),
            "ratingDistribution": ordered_rating_distribution(row.get("评级") for row in sample_rows),
            "ratingBasisDistribution": dict(Counter(row.get("评级依据") for row in sample_rows)),
            "statusConfidenceDistribution": dict(Counter(row.get("状态置信度") for row in sample_rows)),
        },
        "stillRequiresUserReview": True,
        "privateTaskPack": rel(OUTPUT_V4_PACK),
        "sanitized": True,
        "m3Entered": False,
    }


def build_v4_task_summary(v4_rows: list[dict]) -> dict:
    sample_rows = non_pending(v4_rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "privateTaskPackGenerated": True,
        "privateTaskPack": rel(OUTPUT_V4_PACK),
        "rows": len(v4_rows),
        "sampleRows": len(sample_rows),
        "pendingUserSpecifiedRows": len(v4_rows) - len(sample_rows),
        "sameStandardWorkIdsAsV3": True,
        "containsBuyoutAmortization": True,
        "containsStatusConfidence": True,
        "singleMainRatingColumn": True,
        "automaticOperatingSuggestionMainFieldRemoved": True,
        "containsOperatingSuggestionMainColumn": False,
        "userWritableFields": [
            "用户判断：收入模式是否合理",
            "用户判断：下架/版权状态是否合理",
            "用户判断：评级是否合理",
            "用户判断：是否应进入M4",
            "用户备注",
        ],
        "revenueModelDistribution": dict(Counter(row.get("收入模式") for row in sample_rows)),
        "ratingDistribution": ordered_rating_distribution(row.get("评级") for row in sample_rows),
        "ratingBasisDistribution": dict(Counter(row.get("评级依据") for row in sample_rows)),
        "statusConfidenceDistribution": dict(Counter(row.get("状态置信度") for row in sample_rows)),
        "m4CandidateRows": sum(1 for row in sample_rows if row.get("是否建议进入M4校准案例") == "是"),
        "sanitized": True,
        "m3Entered": False,
    }


def write_private_pack(rows: list[dict], failure_payload: dict, correction_payload: dict, task_payload: dict) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "00_阅读说明"
    append_rows(
        ws,
        [
            ["项目", "说明"],
            ["用途", "M2 rating-standard-v4.2 任务包，用于复核三信号买断、单月实销保留、纯买断、买断+实销、下架/版权状态。"],
            ["纯实销规则", "保持用户月均实销档位不变。"],
            ["纯买断规则", f"买断估计金额按默认 {BUYOUT_AMORTIZATION_YEARS} 年折算为月均实销等价值；如剩余版权期更短则按剩余版权期，最低 1 年。"],
            ["买断+实销规则", "当前评级将买断折算月均与当周期实销月均相加；下一周期只预测实销部分。"],
            ["货架/版权状态", "版权台账由版权部负责，作为高可信状态来源；尾部收入只作后续核查线索。"],
            ["M4校准案例", "M4 样本池由用户选择经典/关键作品沉淀经验，不自动收集本轮失败样本。"],
            ["运营建议", "不输出自动运营建议主列，仅保留风险/复核提示和 M4 校准候选。"],
            ["提交边界", "private Excel，不提交 Git；本轮不进入 M3。"],
        ],
    )

    task_ws = wb.create_sheet("01_v4.2运营任务卡")
    headers = list(rows[0].keys()) if rows else []
    append_rows(task_ws, [headers])
    for row in rows:
        append_rows(task_ws, [[row.get(header, "") for header in headers]])
    add_validations(task_ws, headers, len(rows))
    hide_aux_columns(task_ws, headers)

    summary_ws = wb.create_sheet("02_聚合摘要")
    append_rows(
        summary_ws,
        [
            ["指标", "值"],
            ["v3 评级合理率", failure_payload.get("ratingReasonableRate")],
            ["v3 纯买断不合理样本", failure_payload.get("pureBuyoutFailure", {}).get("userMarkedUnreasonable")],
            ["v3 买断+实销不合理样本", failure_payload.get("buyoutPlusSalesFailure", {}).get("userMarkedUnreasonable")],
            ["v3 下架/版权状态不合理样本", failure_payload.get("shelfStatusFailure", {}).get("userMarkedUnreasonable")],
            ["v4.2 样本行", task_payload.get("sampleRows")],
            ["v4.2 状态置信度分布", json.dumps(task_payload.get("statusConfidenceDistribution"), ensure_ascii=False)],
            ["v4.2 评级分布", json.dumps(task_payload.get("ratingDistribution"), ensure_ascii=False)],
            ["仍需人工复核", correction_payload.get("stillRequiresUserReview")],
        ],
    )
    finalize_workbook(wb)
    wb.save(OUTPUT_V4_PACK)


def add_validations(ws, headers: list[str], row_count: int) -> None:
    if row_count <= 0:
        return
    judgment_validation = DataValidation(type="list", formula1=f'"{",".join(USER_JUDGMENT_OPTIONS)}"', allow_blank=True)
    m4_validation = DataValidation(type="list", formula1=f'"{",".join(M4_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(judgment_validation)
    ws.add_data_validation(m4_validation)
    judgment_headers = [
        "用户判断：收入模式是否合理",
        "用户判断：下架/版权状态是否合理",
        "用户判断：评级是否合理",
    ]
    for header in judgment_headers:
        if header in headers:
            col = get_column_letter(headers.index(header) + 1)
            judgment_validation.add(f"{col}2:{col}{row_count + 1}")
    if "用户判断：是否应进入M4" in headers:
        col = get_column_letter(headers.index("用户判断：是否应进入M4") + 1)
        m4_validation.add(f"{col}2:{col}{row_count + 1}")


def hide_aux_columns(ws, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        if clean(header).startswith("辅助原始："):
            ws.column_dimensions[get_column_letter(index)].hidden = True


def write_report(json_path: Path, md_path: Path, report_id: str, payload: dict, renderer) -> None:
    write_json(json_path, envelope(report_id, payload))
    write_text(md_path, renderer(payload))


def render_v3_failure_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 rating-standard-v3 targeted failure analysis",
            "",
            f"- 有效复核样本：`{payload['effectiveReviewRows']}`",
            f"- 收入模式合理率：`{payload['revenueModeReasonableRate']}`",
            f"- 下架/版权状态合理率：`{payload['shelfStatusReasonableRate']}`",
            f"- 评级合理率：`{payload['ratingReasonableRate']}`",
            "",
            "## 聚合根因",
            "",
            f"- 纯买断：`{payload['pureBuyoutFailure']['userMarkedUnreasonable']}/{payload['pureBuyoutFailure']['rows']}` 被标为不合理；v3 口径把一次性买断总额当作前台价值，需改为买断月均实销等价值。",
            f"- 买断+实销：`{payload['buyoutPlusSalesFailure']['userMarkedUnreasonable']}/{payload['buyoutPlusSalesFailure']['rows']}` 被标为不合理；当前评级需叠加买断折算月均与当周期实销月均，下一周期只预测实销。",
            f"- 下架/版权状态：`{payload['shelfStatusFailure']['userMarkedUnreasonable']}` 条被标为不合理；版权台账应作为高可信状态来源，尾部收入不反向约束。",
            "- 纯实销：用户月均实销档位继续作为基准，本轮不改阈值。",
            "",
            "本报告为聚合脱敏报告，不包含真实作品名、作者名、渠道名或原始账单行。本轮不进入 M3。",
        ]
    )


def render_v4_correction_md(payload: dict) -> str:
    v4 = payload["v4Metrics"]
    return "\n".join(
        [
            "# M2 rating-standard-v4.2 targeted correction summary",
            "",
            "- 纯实销规则：保持不变。",
            f"- 纯买断规则：`buyoutEstimatedAmount / ({BUYOUT_AMORTIZATION_YEARS} * 12)` 得到买断折算月均实销；剩余版权期更短时封顶，最低 1 年。",
            "- 买断+实销规则：`ratingBasis=current_sales_with_buyout_allocation`，当前评级叠加实销月均与买断月均；下一周期只预测实销。",
            "- 下架/版权状态：版权台账高可信；尾部收入只作后续运营核查线索，不反向改写状态。",
            "- M4 校准案例：由用户选择经典/关键作品，不自动沉淀本轮失败样本。",
            "- 自动运营建议：仍不输出主列。",
            "",
            "## v4.2 聚合变化",
            "",
            f"- 样本行：`{v4['sampleRows']}`",
            f"- 评级变化行：`{v4['ratingChangedRows']}`",
            f"- 纯买断评级变化行：`{v4['pureBuyoutRatingChangedRows']}`",
            f"- 买断+实销评级变化行：`{v4['buyoutPlusSalesRatingChangedRows']}`",
            f"- 评级分布：`{json.dumps(v4['ratingDistribution'], ensure_ascii=False)}`",
            f"- 评级依据分布：`{json.dumps(v4['ratingBasisDistribution'], ensure_ascii=False)}`",
            f"- 状态置信度分布：`{json.dumps(v4['statusConfidenceDistribution'], ensure_ascii=False)}`",
            "",
            "v4.2 仍需用户复核；v1.1 conditional / rating-standard-v4.2 不是最终正式发布审批结果。本轮不进入 M3。",
        ]
    )


def render_v4_task_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 rating-standard-v4.2 task pack summary",
            "",
            f"- 私有任务包：`{payload['privateTaskPack']}`",
            f"- 总行数：`{payload['rows']}`",
            f"- 样本行：`{payload['sampleRows']}`",
            f"- 用户指定待补任务卡行：`{payload['pendingUserSpecifiedRows']}`",
            f"- 包含买断折算：`{payload['containsBuyoutAmortization']}`",
            f"- 包含状态置信度：`{payload['containsStatusConfidence']}`",
            f"- 单一主评级列：`{payload['singleMainRatingColumn']}`",
            f"- 无运营建议主列：`{not payload['containsOperatingSuggestionMainColumn']}`",
            f"- 评级分布：`{json.dumps(payload['ratingDistribution'], ensure_ascii=False)}`",
            f"- 评级依据分布：`{json.dumps(payload['ratingBasisDistribution'], ensure_ascii=False)}`",
            f"- 状态置信度分布：`{json.dumps(payload['statusConfidenceDistribution'], ensure_ascii=False)}`",
            "",
            "本报告为聚合脱敏报告，不包含真实作品名、作者名、渠道名或原始账单行。私有 Excel 位于 gitignored 的 data/private-output 下，不应提交。",
        ]
    )


def non_pending(rows: list[dict]) -> list[dict]:
    return [row for row in rows if clean(row.get("作品名")) or clean(row.get("作者")) or row.get("评级")]


def revenue_model_code(label: str) -> str:
    if "纯买断" in label:
        return "pure_buyout"
    if "买断+实销" in label:
        return "buyout_plus_sales"
    if "实销" in label or "分成" in label:
        return "pure_sales_share"
    return "unknown_revenue_model"


def rating_from_sales_amount(amount: float | None) -> str:
    value = amount or 0
    if value > 100000:
        return "S+"
    if value >= 10000:
        return "S"
    if value >= 5000:
        return "A"
    if value >= 1000:
        return "B"
    if value >= 500:
        return "C"
    if value >= 100:
        return "D"
    return "E"


def monthly_sales_amount(row: dict) -> float | None:
    monthly = first_number(row.get("月均实销收入"), row.get("实销月均收入"))
    if monthly is not None:
        return round_float(monthly, 2)
    annualized = number_or_none(row.get("年化实销收入"))
    if annualized is not None:
        return round_float(annualized / 12, 2)
    trailing12 = number_or_none(row.get("最近12月实销收入"))
    if trailing12 is not None:
        return round_float(trailing12 / 12, 2)
    return None


def resolve_buyout_amortization_years(row: dict) -> float:
    years = BUYOUT_AMORTIZATION_YEARS
    remaining_months = number_or_none(row.get("剩余版权月数"))
    if remaining_months is not None:
        years = min(years, max(1, remaining_months / 12))
    return round_float(years, 2)


def rating_from_buyout_annual_value(amount: float | None) -> str:
    return rating_from_sales_amount(amount)


def mixed_rating(sales_rating: str, buyout_rating: str, fallback_rating: str) -> str:
    base = sales_rating if sales_rating in RANK else fallback_rating if fallback_rating in RANK else "E"
    if buyout_rating not in RANK:
        return base
    cap = improve_rating(base, 1)
    if RANK[buyout_rating] < RANK[cap]:
        return cap
    if RANK[buyout_rating] < RANK[base]:
        return buyout_rating
    return base


def improve_rating(rating: str, steps: int) -> str:
    if rating not in RANK:
        return rating
    return RATINGS[max(0, RANK[rating] - steps)]


def normalize_rating(value) -> str:
    text = clean(value)
    return text if text in RANK else ""


def ordered_rating_distribution(values) -> dict:
    counter = Counter(value for value in values if value in RANK)
    return {rating: int(counter.get(rating, 0)) for rating in RATINGS}


def first_number(*values) -> float | None:
    for value in values:
        number = number_or_none(value)
        if number is not None:
            return number
    return None


def number_or_none(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def display_number(value) -> str:
    if value is None:
        return "缺失"
    return str(round_float(value, 2))


def round_float(value, digits: int = 2):
    if value is None:
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if not math.isfinite(number):
        return None
    return round(number, digits)


def append_rows(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)


def finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column_cells in enumerate(ws.columns, start=1):
            max_len = min(50, max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2)
            ws.column_dimensions[get_column_letter(index)].width = max(12, max_len)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def write_report(json_path: Path, md_path: Path, report_id: str, payload: dict, renderer) -> None:
    write_json(json_path, envelope(report_id, payload))
    write_text(md_path, renderer(payload))


def envelope(report_id: str, payload: dict) -> dict:
    return {
        "reportId": report_id,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "containsRawLedgerRows": False,
        "payload": payload,
    }


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


if __name__ == "__main__":
    main()
