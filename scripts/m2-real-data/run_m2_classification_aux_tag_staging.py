from __future__ import annotations

import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python spreadsheet dependency. Install openpyxl into a temporary dependency path, "
        "then set PYTHONPATH before rerunning."
    ) from exc


ROOT = Path(__file__).resolve().parents[2]
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
TAXONOMY_PATH = (
    ROOT
    / "src"
    / "domain"
    / "oldProductEvaluation"
    / "classificationTaxonomy.v1.json"
)
CORE_STAGING_PATH = PRIVATE_DIR / "M2-five-source-local-staging-apply-result-cn-v1.json"
BASE_CANDIDATE_PATH = PRIVATE_DIR / "M2-classification-aux-tag-candidate-pack-cn-v2.xlsx"
FILLED_PACK_PATH = PRIVATE_DIR / "M2-classification-aux-tag-fill-pack-cn-v3.xlsx"

PRIVATE_RESULT_JSON = (
    PRIVATE_DIR / "M2-classification-aux-tag-local-staging-apply-result-cn-v1.json"
)
PRIVATE_RESULT_XLSX = (
    PRIVATE_DIR / "M2-classification-aux-tag-local-staging-apply-result-cn-v1.xlsx"
)
PRIVATE_SUMMARY_JSON = (
    PRIVATE_DIR / "M2-classification-aux-tag-local-staging-summary-cn-v1.json"
)
COUNTRY_REVIEW_XLSX = PRIVATE_DIR / "M2-country-aux-tag-review-pack-cn-v1.xlsx"
PUBLIC_SUMMARY_JSON = (
    DOCS_DIR / "M2-classification-aux-tag-local-staging-summary-v1.json"
)
PUBLIC_SUMMARY_MD = (
    DOCS_DIR / "M2-classification-aux-tag-local-staging-summary-v1.md"
)

ID_HEADERS = ["作品ID", "我方作品ID", "书号", "内容ID", "项目ID", "原创ID"]
PLATFORM_TITLE_HEADERS = ["平台作品名称", "作品名称"]
AUX_SIGNAL_HEADERS = ["授权形式", "授权分类"]

COUNTRY_ABBREVIATIONS = {
    "日": "日本",
    "韩": "韩国",
    "美": "美国",
    "英": "英国",
    "法": "法国",
    "德": "德国",
    "俄": "俄罗斯",
    "加": "加拿大",
    "澳": "澳大利亚",
    "意": "意大利",
    "西": "西班牙",
    "印": "印度",
}

HISTORY_SIGNAL_RULES = {
    "先秦": ["先秦", "春秋", "战国", "夏商周"],
    "汉": ["汉代", "汉朝", "秦汉"],
    "三国": ["三国"],
    "南北朝": ["南北朝", "魏晋南北朝"],
    "隋": ["隋代", "隋朝"],
    "唐": ["唐代", "唐朝", "盛唐", "大唐"],
    "五代十国": ["五代十国"],
    "宋": ["宋代", "宋朝", "两宋", "北宋", "南宋", "大宋"],
    "元": ["元代", "元朝", "大元"],
    "明": ["明代", "明朝", "大明"],
    "清": ["清代", "清朝", "大清"],
    "近代史": ["近代史", "中国近代", "民国史"],
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_work_id(value) -> str:
    text = clean(value)
    if not text:
        return ""
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def split_tags(value) -> list[str]:
    text = clean(value)
    if not text or text == "无":
        return []
    result = []
    for item in re.split(r"[；;、，,|/]+", text):
        tag = clean(item)
        if tag and tag != "无" and tag not in result:
            result.append(tag)
    return result


def relative(path: Path) -> str:
    return str(path.relative_to(ROOT)).replace("\\", "/")


def source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_taxonomy() -> dict:
    return json.loads(TAXONOMY_PATH.read_text(encoding="utf-8"))


def rows_from_sheet(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Required sheet is missing from {path.name}: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {
            headers[index]: clean(values[index]) if index < len(values) else ""
            for index in range(len(headers))
            if headers[index]
        }
        if any(row.values()):
            rows.append(row)
    return rows


def load_core_works() -> dict[str, dict]:
    payload = json.loads(CORE_STAGING_PATH.read_text(encoding="utf-8"))
    works: dict[str, dict] = {}
    for row in payload.get("records", []):
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        work = works.setdefault(work_id, {"作品编号": work_id, "书名": ""})
        if clean(row.get("书名")):
            work["书名"] = clean(row.get("书名"))
    return works


def load_base_candidates() -> dict[str, dict]:
    rows = rows_from_sheet(BASE_CANDIDATE_PATH, "03_全量候选明细")
    result = {}
    for row in rows:
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in result:
            raise SystemExit("Base classification candidate contains duplicate work IDs.")
        result[work_id] = row
    return result


def load_user_fills() -> tuple[dict[str, dict], dict[str, list[str]], dict]:
    classification_rows = rows_from_sheet(FILLED_PACK_PATH, "01_分类需人工确认")
    auxiliary_rows = rows_from_sheet(FILLED_PACK_PATH, "02_辅助标签需核对")

    classification = {}
    for row in classification_rows:
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in classification:
            raise SystemExit("Filled classification pack contains duplicate work IDs.")
        classification[work_id] = {
            "一级分类": clean(row.get("一级分类_请填写")),
            "二级分类": clean(row.get("二级分类_请填写")),
            "三级分类": clean(row.get("三级分类_请填写")),
            "备注": clean(row.get("备注")),
        }

    auxiliary = {}
    for row in auxiliary_rows:
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id:
            continue
        if work_id in auxiliary:
            raise SystemExit("Filled auxiliary-tag pack contains duplicate work IDs.")
        auxiliary[work_id] = split_tags(row.get("辅助标签_请填写"))

    return classification, auxiliary, {
        "classificationRows": len(classification_rows),
        "classificationUniqueWorkIds": len(classification),
        "auxiliaryRows": len(auxiliary_rows),
        "auxiliaryUniqueWorkIds": len(auxiliary),
        "inputSha256": source_hash(FILLED_PACK_PATH),
    }


def collect_source_signals() -> dict[str, dict]:
    signals = defaultdict(
        lambda: {
            "productLines": [],
            "platformTitles": [],
            "auxiliarySignals": [],
            "sourceNames": set(),
        }
    )
    source_paths = list((ROOT / "data" / "master-data").glob("*.xlsx"))
    source_paths += list((ROOT / "data" / "real-bills").glob("*.xlsx"))
    for path in source_paths:
        if path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [clean(value) for value in first_row]
            id_header = next((header for header in ID_HEADERS if header in headers), "")
            if not id_header:
                continue
            indexes = {header: index for index, header in enumerate(headers) if header}
            id_index = indexes[id_header]
            relevant_headers = [
                header
                for header in ["产品线", *PLATFORM_TITLE_HEADERS, *AUX_SIGNAL_HEADERS]
                if header in indexes
            ]
            for values in sheet.iter_rows(min_row=2, values_only=True):
                work_id = normalize_work_id(values[id_index] if id_index < len(values) else "")
                if not work_id:
                    continue
                item = signals[work_id]
                item["sourceNames"].add(f"{path.name}-{sheet.title}")
                for header in relevant_headers:
                    index = indexes[header]
                    value = clean(values[index] if index < len(values) else "")
                    if not value:
                        continue
                    if header == "产品线":
                        item["productLines"].append(value)
                    elif header in PLATFORM_TITLE_HEADERS:
                        item["platformTitles"].append(value)
                    else:
                        item["auxiliarySignals"].append(value)
    return signals


def find_level1_for_level2(taxonomy: dict, level2: str) -> str:
    for level1, branches in taxonomy["classificationTree"].items():
        if level2 in branches:
            return level1
    return ""


def normalize_path(taxonomy: dict, level1: str, level2: str, level3: str) -> dict:
    level1 = clean(level1)
    level2 = clean(level2)
    level3 = clean(level3)
    reasons = []

    alias = taxonomy.get("classificationAliases", {}).get(level3)
    if alias:
        reasons.append(f"三级分类“{level3}”归一为“{alias}”")
        level3 = alias

    expected_level1 = find_level1_for_level2(taxonomy, level2)
    if expected_level1 and level1 != expected_level1:
        reasons.append(f"二级分类“{level2}”属于“{expected_level1}”，一级分类按固定树归一")
        level1 = expected_level1

    branches = taxonomy["classificationTree"].get(level1, {})
    valid = level2 in branches and level3 in branches.get(level2, [])
    return {
        "一级分类": level1,
        "二级分类": level2,
        "三级分类": level3,
        "有效": valid,
        "归一说明": "；".join(reasons),
    }


def detect_history_refinement(product_lines: list[str]) -> tuple[str, str]:
    text = "；".join(product_lines)
    found = [
        category
        for category, terms in HISTORY_SIGNAL_RULES.items()
        if any(term in text for term in terms)
    ]
    if len(found) == 1:
        return found[0], "数字版权台账产品线含唯一明确历史时期信号"
    return "", ""


def detect_country_candidates(taxonomy: dict, source: dict) -> tuple[list[str], list[str]]:
    country_order = taxonomy["auxiliaryTagGroups"]["国家"]
    texts = [
        *[("平台作品名称", value) for value in source["platformTitles"]],
        *[("授权资料", value) for value in source["auxiliarySignals"]],
    ]
    found = set()
    evidence = []
    abbreviation_pattern = re.compile(
        r"[\(\[【（](" + "|".join(map(re.escape, COUNTRY_ABBREVIATIONS)) + r")[\)\]】）]"
    )
    for source_label, value in texts:
        matched = set()
        for country in country_order:
            if country in value:
                matched.add(country)
                signal = f"{source_label}出现明确国家全称“{country}”"
                if signal not in evidence:
                    evidence.append(signal)
        for match in abbreviation_pattern.finditer(value):
            abbreviation = match.group(1)
            matched.add(COUNTRY_ABBREVIATIONS[abbreviation])
            signal = f"{source_label}出现规范国家简称“（{abbreviation}）”"
            if signal not in evidence:
                evidence.append(signal)
        if matched:
            found.update(matched)
    ordered = [country for country in country_order if country in found]
    return ordered, evidence[:3]


def build_result() -> tuple[list[dict], list[dict], list[dict], dict]:
    taxonomy = load_taxonomy()
    works = load_core_works()
    base = load_base_candidates()
    manual_classification, reviewed_auxiliary, input_stats = load_user_fills()
    source_signals = collect_source_signals()

    non_country_tags = {
        tag
        for group, tags in taxonomy["auxiliaryTagGroups"].items()
        if group != "国家"
        for tag in tags
    }
    rows = []
    issues = []
    country_review_rows = []
    alias_normalized = 0
    hierarchy_normalized = 0
    history_refined = 0

    for work_id in sorted(works, key=lambda value: (len(value), value)):
        work = works[work_id]
        candidate = base.get(work_id)
        if not candidate:
            issues.append({"问题类型": "缺少系统分类候选", "作品编号": work_id})
            candidate = {}

        manual = manual_classification.get(work_id)
        if manual:
            raw_level1 = manual["一级分类"]
            raw_level2 = manual["二级分类"]
            raw_level3 = manual["三级分类"]
            source_type = "用户v3确认"
            confidence = "人工确认"
        else:
            raw_level1 = clean(candidate.get("一级分类"))
            raw_level2 = clean(candidate.get("二级分类"))
            raw_level3 = clean(candidate.get("三级分类"))
            source_type = "系统候选自动采用"
            confidence = clean(candidate.get("分类置信度")) or "系统候选"

        refinement_reason = ""
        if (
            not manual
            and raw_level1 == "出版物"
            and raw_level2 == "历史"
            and raw_level3 == "中国通史"
        ):
            refined, refinement_reason = detect_history_refinement(
                source_signals[work_id]["productLines"]
            )
            if refined:
                raw_level3 = refined
                source_type = "系统候选按新版历史分类细化"
                history_refined += 1

        normalized = normalize_path(taxonomy, raw_level1, raw_level2, raw_level3)
        if "归一为" in normalized["归一说明"]:
            alias_normalized += 1
        if "一级分类按固定树归一" in normalized["归一说明"]:
            hierarchy_normalized += 1
        if normalized["归一说明"]:
            source_type += "；固定树归一"

        if not normalized["有效"]:
            issues.append(
                {
                    "问题类型": "分类路径不符合固定树",
                    "作品编号": work_id,
                    "一级分类": normalized["一级分类"],
                    "二级分类": normalized["二级分类"],
                    "三级分类": normalized["三级分类"],
                }
            )

        auxiliary_tags = reviewed_auxiliary.get(work_id, [])
        invalid_auxiliary = [tag for tag in auxiliary_tags if tag not in non_country_tags]
        if invalid_auxiliary:
            issues.append(
                {
                    "问题类型": "已核对辅助标签不在固定候选库",
                    "作品编号": work_id,
                    "标签": "；".join(invalid_auxiliary),
                }
            )

        countries, country_evidence = detect_country_candidates(
            taxonomy, source_signals[work_id]
        )
        if countries:
            country_review_rows.append(
                {
                    "书名": work["书名"],
                    "作品编号": work_id,
                    "国家标签候选": "；".join(countries),
                    "是否采用（请填写）": "",
                    "采用值（如需修改）": "；".join(countries),
                    "候选依据": "；".join(country_evidence),
                    "备注": "",
                }
            )

        reasons = [clean(candidate.get("分类理由")), normalized["归一说明"], refinement_reason]
        rows.append(
            {
                "作品编号": work_id,
                "书名": work["书名"],
                "一级分类": normalized["一级分类"],
                "二级分类": normalized["二级分类"],
                "三级分类": normalized["三级分类"],
                "分类有效": "是" if normalized["有效"] else "否",
                "分类采用来源": source_type,
                "分类置信度": confidence,
                "分类说明": "；".join(reason for reason in reasons if reason),
                "已核对辅助标签": "；".join(auxiliary_tags),
                "辅助标签采用来源": "用户v3确认" if work_id in reviewed_auxiliary else "无明显特殊标签",
                "待核对国家标签候选": "；".join(countries),
                "国家标签状态": "待人工核对" if countries else "无明确国家信号",
            }
        )

    expected_manual_ids = {
        work_id
        for work_id, row in base.items()
        if clean(row.get("分类处理")) != "自动采用候选"
    }
    missing_manual_ids = expected_manual_ids - set(manual_classification)
    extra_manual_ids = set(manual_classification) - set(works)
    if missing_manual_ids:
        issues.append({"问题类型": "系统要求人工确认但v3未填写", "数量": len(missing_manual_ids)})
    if extra_manual_ids:
        issues.append({"问题类型": "v3含账单作品范围外编号", "数量": len(extra_manual_ids)})

    classification_valid_count = sum(row["分类有效"] == "是" for row in rows)
    applied_auxiliary_assignments = sum(
        len(split_tags(row["已核对辅助标签"])) for row in rows
    )
    country_distribution = Counter(
        tag
        for row in country_review_rows
        for tag in split_tags(row["国家标签候选"])
    )
    source_distribution = Counter(
        "用户确认" if row["分类采用来源"].startswith("用户v3确认") else "系统自动"
        for row in rows
    )

    summary = {
        "schema": "m2.classification_aux_tag_local_staging_summary.v1",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "taxonomyVersion": taxonomy["version"],
        "candidateVersion": "m2-local-classification-aux-tag-staging-v1",
        "input": {
            "coveredWorks": len(works),
            "baseCandidateWorks": len(base),
            "manualClassificationRows": input_stats["classificationRows"],
            "manualClassificationUniqueWorks": input_stats["classificationUniqueWorkIds"],
            "reviewedAuxiliaryRows": input_stats["auxiliaryRows"],
            "reviewedAuxiliaryUniqueWorks": input_stats["auxiliaryUniqueWorkIds"],
            "filledPackSha256": input_stats["inputSha256"],
        },
        "classification": {
            "validWorks": classification_valid_count,
            "invalidOrMissingWorks": len(rows) - classification_valid_count,
            "systemAutomaticWorks": source_distribution["系统自动"],
            "userConfirmedWorks": source_distribution["用户确认"],
            "aliasNormalizedWorks": alias_normalized,
            "hierarchyNormalizedWorks": hierarchy_normalized,
            "historySpecificRefinedWorks": history_refined,
            "level1Distribution": dict(Counter(row["一级分类"] for row in rows)),
            "level2Distribution": dict(Counter(row["二级分类"] for row in rows)),
            "level3Top20": dict(Counter(row["三级分类"] for row in rows).most_common(20)),
            "localFileStagingClosed": classification_valid_count == len(rows) and not missing_manual_ids,
        },
        "auxiliaryTags": {
            "reviewedWorksApplied": len(reviewed_auxiliary),
            "reviewedTagAssignmentsApplied": applied_auxiliary_assignments,
            "countryCandidateWorksPendingReview": len(country_review_rows),
            "countryCandidateDistribution": dict(country_distribution),
            "countryTagsApplied": 0,
            "countryReviewRequired": bool(country_review_rows),
        },
        "quality": {
            "issueCount": len(issues),
            "classificationDuplicateWorkIds": input_stats["classificationRows"]
            - input_stats["classificationUniqueWorkIds"],
            "auxiliaryDuplicateWorkIds": input_stats["auxiliaryRows"]
            - input_stats["auxiliaryUniqueWorkIds"],
        },
        "outputs": {
            "privateResultJson": relative(PRIVATE_RESULT_JSON),
            "privateResultWorkbook": relative(PRIVATE_RESULT_XLSX),
            "privateCountryReviewWorkbook": relative(COUNTRY_REVIEW_XLSX),
            "publicSummaryJson": relative(PUBLIC_SUMMARY_JSON),
            "publicSummaryMarkdown": relative(PUBLIC_SUMMARY_MD),
        },
        "boundaries": {
            "formalMasterDataWritten": False,
            "databaseWritten": False,
            "privateFilesGitignored": True,
            "m3FormalExecutionStarted": False,
            "specialAttributeTagsClosed": False,
        },
    }
    return rows, country_review_rows, issues, summary


def format_workbook(workbook: Workbook, input_sheets: set[str] | None = None) -> None:
    input_sheets = input_sheets or set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if sheet.title in input_sheets and cell.row > 1:
                    cell.fill = input_fill
        for column in sheet.columns:
            letter = column[0].column_letter
            width = max(
                (len(clean(cell.value)) for cell in column[:200]),
                default=8,
            )
            sheet.column_dimensions[letter].width = min(max(width + 2, 10), 42)


def append_dict_rows(sheet, rows: list[dict], headers: list[str]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def country_review_pack_has_user_decisions() -> bool:
    if not COUNTRY_REVIEW_XLSX.exists():
        return False
    workbook = load_workbook(COUNTRY_REVIEW_XLSX, read_only=True, data_only=True)
    if "01_国家标签需核对" not in workbook.sheetnames:
        return False
    sheet = workbook["01_国家标签需核对"]
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "是否采用（请填写）" not in headers:
        return False
    decision_index = headers.index("是否采用（请填写）")
    return any(
        clean(values[decision_index] if decision_index < len(values) else "")
        for values in sheet.iter_rows(min_row=2, values_only=True)
    )


def write_private_outputs(
    rows: list[dict], country_rows: list[dict], issues: list[dict], summary: dict
) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    PRIVATE_RESULT_JSON.write_text(
        json.dumps({"summary": summary, "records": rows, "issues": issues}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    PRIVATE_SUMMARY_JSON.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    workbook = Workbook()
    readme = workbook.active
    readme.title = "00_说明"
    for item in [
        ["用途", "分类与已核对辅助标签的本地文件级 staging 结果。"],
        ["覆盖范围", f"{summary['input']['coveredWorks']} 个账单作品。"],
        ["正式主数据", "未写入。"],
        ["国家标签", "只生成候选，未应用；需填写单独的国家标签核对表。"],
    ]:
        readme.append(item)
    summary_sheet = workbook.create_sheet("01_汇总")
    summary_sheet.append(["项目", "数值"])
    for key, value in [
        ("分类有效作品数", summary["classification"]["validWorks"]),
        ("分类无效或缺失作品数", summary["classification"]["invalidOrMissingWorks"]),
        ("系统自动分类作品数", summary["classification"]["systemAutomaticWorks"]),
        ("用户确认分类作品数", summary["classification"]["userConfirmedWorks"]),
        ("固定树层级归一作品数", summary["classification"]["hierarchyNormalizedWorks"]),
        ("别名归一作品数", summary["classification"]["aliasNormalizedWorks"]),
        ("历史时期细化作品数", summary["classification"]["historySpecificRefinedWorks"]),
        ("已应用辅助标签作品数", summary["auxiliaryTags"]["reviewedWorksApplied"]),
        ("已应用辅助标签赋值数", summary["auxiliaryTags"]["reviewedTagAssignmentsApplied"]),
        ("待核对国家标签作品数", summary["auxiliaryTags"]["countryCandidateWorksPendingReview"]),
    ]:
        summary_sheet.append([key, value])
    result_sheet = workbook.create_sheet("02_全量本地结果")
    result_headers = [
        "作品编号",
        "书名",
        "一级分类",
        "二级分类",
        "三级分类",
        "分类有效",
        "分类采用来源",
        "分类置信度",
        "分类说明",
        "已核对辅助标签",
        "辅助标签采用来源",
        "待核对国家标签候选",
        "国家标签状态",
    ]
    append_dict_rows(result_sheet, rows, result_headers)
    issue_sheet = workbook.create_sheet("03_质量问题")
    issue_headers = sorted({key for issue in issues for key in issue}) or ["说明"]
    append_dict_rows(issue_sheet, issues or [{"说明": "未发现阻断性质量问题"}], issue_headers)
    format_workbook(workbook)
    workbook.save(PRIVATE_RESULT_XLSX)

    preserve_filled_country_pack = country_review_pack_has_user_decisions()
    country_workbook = Workbook()
    country_readme = country_workbook.active
    country_readme.title = "00_填写说明"
    for item in [
        ["你需要填写什么", "只填写 01_国家标签需核对 的“是否采用（请填写）”列。"],
        ["候选正确", "填写“采用”；采用值已预填，无需重复抄写。"],
        ["候选错误", "填写“不采用”；如有正确国家，可在采用值列改成中文国家名。"],
        ["没有国家标签", "填写“不采用”，并将采用值清空或填“无”。"],
        ["安全边界", "此表为 private 本地核对包，不会提交，也不写正式主数据。"],
    ]:
        country_readme.append(item)
    country_sheet = country_workbook.create_sheet("01_国家标签需核对")
    country_headers = [
        "序号",
        "书名",
        "作品编号",
        "国家标签候选",
        "是否采用（请填写）",
        "采用值（如需修改）",
        "候选依据",
        "备注",
    ]
    country_sheet.append(country_headers)
    for index, row in enumerate(country_rows, 1):
        country_sheet.append([index] + [row.get(header, "") for header in country_headers[1:]])
    if country_rows:
        validation = DataValidation(type="list", formula1='"采用,不采用"', allow_blank=True)
        country_sheet.add_data_validation(validation)
        for row_index in range(2, len(country_rows) + 2):
            validation.add(country_sheet.cell(row=row_index, column=5))
    format_workbook(country_workbook, {"01_国家标签需核对"})
    if not preserve_filled_country_pack:
        country_workbook.save(COUNTRY_REVIEW_XLSX)


def write_public_outputs(summary: dict) -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    public_summary = {key: value for key, value in summary.items() if key != "input"}
    public_summary["input"] = {
        key: value
        for key, value in summary["input"].items()
        if key != "filledPackSha256"
    }
    PUBLIC_SUMMARY_JSON.write_text(
        json.dumps(public_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    classification = summary["classification"]
    auxiliary = summary["auxiliaryTags"]
    lines = [
        "# M2 分类与辅助标签本地 staging 汇总 v1",
        "",
        "## 结论",
        "",
        f"- 已对 `{summary['input']['coveredWorks']}` 个账单作品重新校验分类路径。",
        f"- 本地文件级分类有效作品数为 `{classification['validWorks']}`，无效或缺失为 `{classification['invalidOrMissingWorks']}`。",
        f"- 用户填写的分类共 `{classification['userConfirmedWorks']}` 部，系统自动分类共 `{classification['systemAutomaticWorks']}` 部。",
        f"- 固定树层级归一 `{classification['hierarchyNormalizedWorks']}` 部，分类别名归一 `{classification['aliasNormalizedWorks']}` 部。",
        f"- 新版历史时期规则细化 `{classification['historySpecificRefinedWorks']}` 部。",
        f"- 已应用用户核对过的辅助标签 `{auxiliary['reviewedWorksApplied']}` 部、`{auxiliary['reviewedTagAssignmentsApplied']}` 个标签赋值。",
        f"- 新增国家标签候选 `{auxiliary['countryCandidateWorksPendingReview']}` 部，尚未应用，仍需人工核对。",
        "- 本轮未写正式主数据，未写数据库，未进入 M3 formal execution。",
        "",
        "## 分类分布",
        "",
        f"- 一级分类：`{json.dumps(classification['level1Distribution'], ensure_ascii=False)}`",
        f"- 二级分类：`{json.dumps(classification['level2Distribution'], ensure_ascii=False)}`",
        "",
        "## 国家标签候选",
        "",
        f"- 候选分布：`{json.dumps(auxiliary['countryCandidateDistribution'], ensure_ascii=False)}`",
        "- 候选仅来自平台作品名称或授权资料中的明确国家全称、规范括号简称。",
        "- 国家标签按 PRD 定位为辅助标签，必须人工核对后才能进入本地 staging。",
        "",
        "## 边界",
        "",
        "- 当前结果是本地文件级 staging，不是正式主数据验收结果。",
        "- private Excel/JSON 位于 Git 忽略范围，不得提交。",
        "- 特殊属性标签仍需后续单独治理。",
    ]
    PUBLIC_SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    required = [TAXONOMY_PATH, CORE_STAGING_PATH, BASE_CANDIDATE_PATH, FILLED_PACK_PATH]
    missing = [relative(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing required local input: " + ", ".join(missing))

    rows, country_rows, issues, summary = build_result()
    write_private_outputs(rows, country_rows, issues, summary)
    write_public_outputs(summary)

    print(
        json.dumps(
            {
                "分类有效作品数": summary["classification"]["validWorks"],
                "分类无效或缺失作品数": summary["classification"]["invalidOrMissingWorks"],
                "用户确认分类作品数": summary["classification"]["userConfirmedWorks"],
                "系统自动分类作品数": summary["classification"]["systemAutomaticWorks"],
                "层级归一作品数": summary["classification"]["hierarchyNormalizedWorks"],
                "别名归一作品数": summary["classification"]["aliasNormalizedWorks"],
                "已应用辅助标签作品数": summary["auxiliaryTags"]["reviewedWorksApplied"],
                "待核对国家标签作品数": summary["auxiliaryTags"]["countryCandidateWorksPendingReview"],
                "质量问题数": summary["quality"]["issueCount"],
                "国家标签核对表": relative(COUNTRY_REVIEW_XLSX),
                "正式主数据写入": False,
                "进入M3": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
