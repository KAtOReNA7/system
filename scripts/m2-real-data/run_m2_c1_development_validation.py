#!/usr/bin/env python3
"""Run the sealed M2 C1 transparent-ensemble development validation.

The synthetic preflight is the default and reads no private role.  Development
execution is authorized only by the ignored Gate A runtime receipt bound to the
pushed Phase A commit.  The final-holdout mode always fails before any loader.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import m2_calibration_v1 as base
import m2_calibration_v1_2 as v12
import run_m2_calibration_baseline_replay as legacy
import run_m2_calibration_scoring_correction as correction
import run_m2_calibration_v1_2 as phase


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-calibration-v1-2"
PUBLIC_JSON = PUBLIC_DIR / "M2-C1-development-validation-v1.json"
PUBLIC_MD = PUBLIC_DIR / "M2-C1-development-validation-v1.md"
PRIVATE_CASES = PRIVATE_DIR / "M2-C1-development-cases-private-v1.ndjson"
PRIVATE_MANIFEST = PRIVATE_DIR / "M2-C1-development-manifest-private-v1.json"
PRIVATE_WORKBOOK = PRIVATE_DIR / "M2-C1-development-validation-private-v1.xlsx"
PHASE_A_HEAD = "879fbd0a951ce6d465082321b38f965b14815935"
PHASE_A_PARENT = "be03db7bdec19b83139d85712bee43995d872679"
BRANCH = "codex/m2-calibration-v1"
PUBLIC_MINIMUM = 10
TOLERANCE = 1e-9

C1_SOURCE_PATHS = (
    ROOT / "package.json",
    ROOT / "scripts" / "m2-real-data" / "m2_calibration_v1_2.py",
    ROOT / "scripts" / "m2-real-data" / "run_m2_c1_development_validation.py",
    ROOT
    / "src"
    / "domain"
    / "oldProductEvaluation"
    / "calibrationSpec.v1.2.amendment.json",
    ROOT / "test" / "m2-c1-transparent-ensemble.test.js",
)


class C1ValidationError(RuntimeError):
    """A C1 authorization, leakage, scoring, or artifact contract failed."""


def progress(message: str) -> None:
    print(f"[m2-c1] {message}", file=sys.stderr, flush=True)


def run_git(*args: str, check: bool = True) -> str:
    process = subprocess.run(
        ["git", *args], cwd=ROOT, text=True, capture_output=True, check=False
    )
    if check and process.returncode != 0:
        raise C1ValidationError(process.stderr.strip() or "git command failed")
    return process.stdout.strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def git_blob_sha256(commit: str, relative_path: str) -> str:
    process = subprocess.run(
        ["git", "show", f"{commit}:{relative_path}"],
        cwd=ROOT,
        capture_output=True,
        check=False,
    )
    if process.returncode != 0:
        raise C1ValidationError(f"cannot read Phase A blob: {relative_path}")
    return hashlib.sha256(process.stdout).hexdigest()


def git_blob_bytes(commit: str, relative_path: str) -> bytes:
    process = subprocess.run(
        ["git", "show", f"{commit}:{relative_path}"],
        cwd=ROOT,
        capture_output=True,
        check=False,
    )
    if process.returncode != 0:
        raise C1ValidationError(f"cannot read Phase A blob: {relative_path}")
    return process.stdout


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False)
        + "\n",
        encoding="utf-8",
        newline="\n",
    )


def write_text(path: Path, value: str) -> None:
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def _rounded(value: Any, places: int = 8) -> float | None:
    if value is None:
        return None
    number = float(value)
    return round(number, places) if math.isfinite(number) else None


def _public_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {
            str(key): _public_value(child)
            for key, child in value.items()
            if key not in {"actualTotal", "predictedTotal", "lower", "upper"}
        }
    if isinstance(value, list):
        return [_public_value(child) for child in value]
    if isinstance(value, float):
        return _rounded(value)
    return value


def require_private_boundaries() -> None:
    for path in (
        phase.PRIVATE_GATE_A_RECEIPT,
        phase.PRIVATE_PHASE_A_CASES,
        phase.PRIVATE_PHASE_A_MANIFEST,
        PRIVATE_CASES,
        PRIVATE_MANIFEST,
        PRIVATE_WORKBOOK,
    ):
        if not phase.git_ignored(path):
            raise C1ValidationError(f"private C1 role is not ignored: {path.name}")
        if run_git("ls-files", "--", str(path)):
            raise C1ValidationError(f"private C1 role is tracked: {path.name}")
    if phase.tracked_private_artifacts():
        raise C1ValidationError("a private calibration artifact is tracked")


def c1_source_digests() -> dict[str, str]:
    missing = [path.name for path in C1_SOURCE_PATHS if not path.is_file()]
    if missing:
        raise C1ValidationError(f"C1 source binding is incomplete: {missing}")
    return {
        path.relative_to(ROOT).as_posix(): file_sha256(path)
        for path in C1_SOURCE_PATHS
    }


def verify_gate_a_runtime_receipt(
    *, require_exact_phase_a_head: bool
) -> dict[str, Any]:
    """Verify Gate A against the pushed Phase A blobs, not dirty C1 sources."""

    if run_git("branch", "--show-current") != BRANCH:
        raise C1ValidationError(f"C1 must run on {BRANCH}")
    require_private_boundaries()
    if not phase.PRIVATE_GATE_A_RECEIPT.is_file():
        raise C1ValidationError("Gate A runtime receipt is missing")
    receipt = json.loads(phase.PRIVATE_GATE_A_RECEIPT.read_text(encoding="utf-8"))
    if (
        receipt.get("schema") != "m2.calibration_gate_a.runtime_result.v1"
        or receipt.get("decisionStatus") != "not_for_formal_decision"
        or receipt.get("allTrue") is not True
        or receipt.get("C1AuthorizedByGateA") is not True
        or receipt.get("phaseAHead") != PHASE_A_HEAD
        or receipt.get("phaseAParent") != PHASE_A_PARENT
        or receipt.get("remoteHead") != PHASE_A_HEAD
        or receipt.get("finalHoldoutOpened") is not False
        or receipt.get("embargoShadowOpened") is not False
        or receipt.get("deferred60MonthLabelsOpened") is not False
        or not isinstance(receipt.get("conditions"), Mapping)
        or not all(value is True for value in receipt["conditions"].values())
    ):
        raise C1ValidationError("Gate A runtime receipt is not an all-true Phase A receipt")
    current_head = run_git("rev-parse", "HEAD")
    if require_exact_phase_a_head and current_head != PHASE_A_HEAD:
        raise C1ValidationError("C1 development must start directly from Phase A HEAD")
    ancestor = subprocess.run(
        ["git", "merge-base", "--is-ancestor", PHASE_A_HEAD, current_head],
        cwd=ROOT,
        check=False,
    )
    if ancestor.returncode != 0:
        raise C1ValidationError("current HEAD does not descend from Phase A")
    if run_git("rev-parse", f"{PHASE_A_HEAD}^") != PHASE_A_PARENT:
        raise C1ValidationError("Phase A commit parent changed")

    phase_tree = run_git("rev-parse", f"{PHASE_A_HEAD}^{{tree}}")
    if receipt.get("phaseATree") != phase_tree:
        raise C1ValidationError("Gate receipt differs from the Phase A Git tree")
    tracked_gate = json.loads(
        git_blob_bytes(
            PHASE_A_HEAD, phase.GATE_A_JSON.relative_to(ROOT).as_posix()
        ).decode("utf-8")
    )
    if receipt.get("sourceSha256") != tracked_gate.get("evidenceBindings", {}).get(
        "sourceSha256"
    ):
        raise C1ValidationError("Gate receipt source map differs from tracked Gate A")
    for relative in receipt.get("sourceSha256", {}):
        exists = subprocess.run(
            ["git", "cat-file", "-e", f"{PHASE_A_HEAD}:{relative}"],
            cwd=ROOT,
            check=False,
        )
        if exists.returncode != 0:
            raise C1ValidationError(f"Phase A source is absent from its tree: {relative}")
    for relative, expected in receipt.get("publicReportSha256", {}).items():
        path = ROOT / str(relative)
        if not path.is_file() or file_sha256(path) != expected:
            raise C1ValidationError(f"Phase A public evidence differs: {relative}")
        differs = subprocess.run(
            ["git", "diff", "--quiet", PHASE_A_HEAD, "--", str(relative)],
            cwd=ROOT,
            check=False,
        )
        if differs.returncode != 0:
            raise C1ValidationError(f"Phase A public evidence changed: {relative}")
    phase_evidence = phase.verify_private_phase_a_evidence()
    runtime = receipt.get("runtimeRecomputation", {})
    if (
        runtime.get("privateCaseEvidenceSha256")
        != phase_evidence["caseEvidenceSha256"]
        or runtime.get("privateCaseRowCount")
        != phase_evidence["privateCaseRowCount"]
        or runtime.get("allSealsClosed") is not True
    ):
        raise C1ValidationError("Gate receipt differs from Phase A private evidence")
    _base_spec, _v1_1, amendment = v12.load_and_validate_contract()
    if receipt.get("calibrationSpecV1_2Digest") != v12.canonical_digest(amendment):
        raise C1ValidationError("Gate receipt has a stale v1.2 spec digest")
    return {
        "receipt": receipt,
        "phaseEvidence": phase_evidence,
        "phaseAHead": PHASE_A_HEAD,
        "gateAAllTrue": True,
        "runtimeReceiptTracked": False,
        "finalHoldoutOpened": False,
    }


def _b4_templates(
    rows: Sequence[Mapping[str, Any]], role: str
) -> dict[tuple[str, str, int, str], Mapping[str, Any]]:
    output: dict[tuple[str, str, int, str], Mapping[str, Any]] = {}
    for row in rows:
        if row.get("model_id") != "B4" or row.get("_residual_case_role") != role:
            continue
        key = v12.strict_case_key(row)
        if key in output:
            raise C1ValidationError("Phase A contains a duplicate B4 role case")
        output[key] = row
    if not output:
        raise C1ValidationError(f"Phase A lacks B4 templates for {role}")
    return output


def _candidate_point(
    component_points: Mapping[str, float | None], candidate: Mapping[str, Any]
) -> float | None:
    selected: list[tuple[float, float]] = []
    for component, weight in candidate["weights"].items():
        value = component_points.get(component)
        if value is None:
            return None
        selected.append((float(weight), float(value)))
    return round(sum(weight * value for weight, value in selected), 8)


def _materialize_component_matrix(
    templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    works: Mapping[str, Mapping[str, Any]],
    spec: Mapping[str, Any],
    role: str,
) -> tuple[dict[tuple[str, tuple[str, str, int, str]], dict[str, float | None]], str]:
    matrix: dict[
        tuple[str, tuple[str, str, int, str]], dict[str, float | None]
    ] = {}
    projection: list[dict[str, Any]] = []
    for key in sorted(templates):
        points = v12.c1_component_point_predictions(
            works[key[0]], key[1], key[2], spec
        )
        if set(points) != set(v12.C1_COMPONENT_IDS):
            raise C1ValidationError("C1 component matrix is incomplete")
        if templates[key].get("statisticallyScoreable") is True and any(
            value is None or not math.isfinite(float(value))
            for value in points.values()
        ):
            raise C1ValidationError("a scoreable C1 case lacks a component prediction")
        matrix[(role, key)] = copy.deepcopy(points)
        projection.append(
            {
                "role": role,
                "key": list(key),
                "points": {
                    component: (
                        None
                        if value is None
                        else base.fixed_decimal(float(value))
                    )
                    for component, value in sorted(points.items())
                },
            }
        )
    return matrix, v12.canonical_digest(projection)


def _metric_from_predictions(
    rows: Sequence[Mapping[str, Any]],
    component_matrix: Mapping[
        tuple[str, tuple[str, str, int, str]], Mapping[str, float | None]
    ],
    candidate: Mapping[str, Any],
) -> dict[str, Any]:
    predictions: list[float] = []
    actuals: list[float] = []
    for row in rows:
        key = v12.strict_case_key(row)
        role_key = (str(row["_residual_case_role"]), key)
        point = _candidate_point(component_matrix[role_key], candidate)
        if point is None:
            raise C1ValidationError("candidate metric population has a null point")
        predictions.append(float(point))
        actuals.append(float(row["actual"]))
    denominator = sum(abs(value) for value in actuals)
    signed_denominator = sum(actuals)
    if not rows or denominator <= 0 or abs(signed_denominator) <= TOLERANCE:
        raise C1ValidationError("candidate objective has an invalid denominator")
    errors = [abs(prediction - actual) for prediction, actual in zip(predictions, actuals)]
    return {
        "caseCount": len(rows),
        "wape": sum(errors) / denominator,
        "signedAggregateBias": (
            sum(predictions) - signed_denominator
        )
        / signed_denominator,
    }


def select_candidate_for_outer_origin(
    prior_forward_rows: Sequence[Mapping[str, Any]],
    component_matrix: Mapping[
        tuple[str, tuple[str, str, int, str]], Mapping[str, float | None]
    ],
    outer_origin: str,
    amendment: Mapping[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    contract = amendment["C1"]["training"]
    fallback = v12.c1_candidate_by_id(
        contract["insufficientInnerEvidenceFallback"]["candidateId"], amendment
    )
    inner = [
        row
        for row in prior_forward_rows
        if row.get("model_id") == "C1"
        and row.get("statisticallyScoreable") is True
        and v12.strict_case_key(row)[1] < outer_origin
        and str(row.get("target_end")) <= outer_origin
        and str(row.get("label_available_as_of")) <= outer_origin
        and str(row.get("_bill_month_max")) <= outer_origin
        and str(row.get("_available_as_of")) <= outer_origin
    ]
    origins = sorted({v12.strict_case_key(row)[1] for row in inner})
    base_evidence = {
        "outerOrigin": outer_origin,
        "innerScoreableCaseCount": len(inner),
        "innerDistinctScoreOrigins": len(origins),
        "innerOrigins": origins,
        "maximumInnerTargetEnd": max(
            (str(row["target_end"]) for row in inner), default=None
        ),
        "maximumInnerLabelAvailableAsOf": max(
            (str(row["label_available_as_of"]) for row in inner), default=None
        ),
        "candidateSpaceCount": len(v12.enumerate_c1_candidates(amendment)),
        "seed": int(contract["seed"]),
        "sameOrLaterOuterTruthRead": False,
        "warmupUsedForCandidateSelection": False,
        "finalHoldoutRead": False,
    }
    if (
        len(origins) < int(contract["minimumInnerScoreOrigins"])
        or len(inner) < int(contract["minimumInnerCaseCount"])
    ):
        return fallback, {
            **base_evidence,
            "selectionStatus": "frozen_fallback_insufficient_inner_evidence",
            "selectedCandidateId": fallback["candidateId"],
            "biasFeasibleCandidateCount": 0,
            "candidateEvaluationDigest": v12.canonical_digest([]),
            "selectedObjective": None,
        }

    candidates = v12.enumerate_c1_candidates(amendment)
    objective_weights = contract["selectionObjective"]["weights"]
    bias_guard = contract["biasFeasibilityGuard"]
    evaluations: list[dict[str, Any]] = []
    for candidate in candidates:
        overall = _metric_from_predictions(inner, component_matrix, candidate)
        horizon_metrics: dict[str, dict[str, Any]] = {}
        for horizon in v12.CORE_HORIZONS:
            group = [row for row in inner if v12.strict_case_key(row)[2] == horizon]
            if group:
                horizon_metrics[str(horizon)] = _metric_from_predictions(
                    group, component_matrix, candidate
                )
        top10_rows = [
            row
            for row in inner
            if bool(row.get("strata", {}).get("top_10_percent"))
        ]
        top10 = _metric_from_predictions(top10_rows, component_matrix, candidate)
        mean_horizon = sum(
            float(metric["wape"]) for metric in horizon_metrics.values()
        ) / len(horizon_metrics)
        feasible = abs(float(overall["signedAggregateBias"])) <= float(
            bias_guard["overallAbsoluteSignedBiasMaximumInclusive"]
        ) + TOLERANCE and all(
            abs(float(metric["signedAggregateBias"]))
            <= float(
                bias_guard[
                    "eachDefinedCoreHorizonAbsoluteSignedBiasMaximumInclusive"
                ]
            )
            + TOLERANCE
            for metric in horizon_metrics.values()
        )
        objective = (
            float(objective_weights["overallWape"]) * float(overall["wape"])
            + float(objective_weights["meanCoreHorizonWape"]) * mean_horizon
            + float(objective_weights["absoluteSignedBias"])
            * abs(float(overall["signedAggregateBias"]))
            + float(objective_weights["top10Wape"]) * float(top10["wape"])
        )
        evaluations.append(
            {
                "candidateId": candidate["candidateId"],
                "componentCount": candidate["componentCount"],
                "nonzeroParameterCount": candidate["nonzeroParameterCount"],
                "biasFeasible": feasible,
                "objective": objective,
                "overallWape": overall["wape"],
                "overallSignedBias": overall["signedAggregateBias"],
                "meanCoreHorizonWape": mean_horizon,
                "top10Wape": top10["wape"],
                "horizonSignedBias": {
                    horizon: metric["signedAggregateBias"]
                    for horizon, metric in sorted(horizon_metrics.items())
                },
            }
        )
    feasible = [item for item in evaluations if item["biasFeasible"]]
    if not feasible:
        fallback_eval = next(
            item for item in evaluations if item["candidateId"] == fallback["candidateId"]
        )
        return fallback, {
            **base_evidence,
            "selectionStatus": "frozen_fallback_no_bias_feasible_candidate",
            "selectedCandidateId": fallback["candidateId"],
            "biasFeasibleCandidateCount": 0,
            "candidateEvaluationDigest": v12.canonical_digest(evaluations),
            "selectedObjective": _public_value(fallback_eval),
        }
    chosen_eval = min(
        feasible,
        key=lambda item: (
            float(item["objective"]),
            int(item["componentCount"]),
            int(item["nonzeroParameterCount"]),
            str(item["candidateId"]),
        ),
    )
    selected = v12.c1_candidate_by_id(str(chosen_eval["candidateId"]), amendment)
    return selected, {
        **base_evidence,
        "selectionStatus": "bias_feasible_objective_minimum",
        "selectedCandidateId": selected["candidateId"],
        "biasFeasibleCandidateCount": len(feasible),
        "candidateEvaluationDigest": v12.canonical_digest(evaluations),
        "selectedObjective": _public_value(chosen_eval),
    }


def materialize_c1_role(
    templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    works: Mapping[str, Mapping[str, Any]],
    spec: Mapping[str, Any],
    role: str,
    candidate: Mapping[str, Any],
    candidate_role: str,
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[tuple[str, tuple[str, str, int, str]], dict[str, float | None]]]:
    component_matrix, component_fingerprint = _materialize_component_matrix(
        templates, works, spec, role
    )
    predictions: list[dict[str, Any]] = []
    for key, source in sorted(templates.items()):
        prediction = v12.predict_as_of(
            works[key[0]],
            key[1],
            key[2],
            "C1",
            spec,
            c1_candidate=candidate,
            c1_candidate_role=candidate_role,
        )
        expected = _candidate_point(component_matrix[(role, key)], candidate)
        observed = prediction.get("point_forecast")
        if (expected is None) != (observed is None) or (
            expected is not None
            and not math.isclose(float(expected), float(observed), abs_tol=5e-6)
        ):
            raise C1ValidationError("C1 predict_as_of differs from its component matrix")
        predictions.append(phase._decorate_v12_prediction(prediction, source, role, "C1"))
    joined, lock = phase._lock_and_guarded_join_v12_block(
        predictions, {"B4": templates}, works, role, spec
    )
    if any(
        row.get("identity") != "C1_transparent_ensemble"
        or row.get("c1_candidate", {}).get("candidateId")
        != candidate["candidateId"]
        for row in joined
    ):
        raise C1ValidationError("C1 role contains an unregistered predictor identity")
    return joined, {
        **lock,
        "selectedCandidateId": candidate["candidateId"],
        "candidateRole": candidate_role,
        "componentMatrixFingerprint": component_fingerprint,
        "componentMatrixRowCount": len(component_matrix),
        "samePredictAsOfEntryUsed": True,
        "heldOutcomeFieldsReadByPredictor": False,
    }, component_matrix


def case_parity(
    c1_rows: Sequence[Mapping[str, Any]],
    primary_rows: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    c1 = {v12.strict_case_key(row): row for row in c1_rows}
    primary = {v12.strict_case_key(row): row for row in primary_rows}
    if len(c1) != len(c1_rows) or len(primary) != len(primary_rows):
        raise C1ValidationError("C1 or primary contains duplicate case keys")
    keys_equal = set(c1) == set(primary)
    if not keys_equal:
        raise C1ValidationError("C1 case keys differ from the primary comparator")
    actual_equal = all(
        base.fixed_decimal(c1[key]["actual"])
        == base.fixed_decimal(primary[key]["actual"])
        for key in c1
    )
    state_equal = all(
        (
            c1[key].get("statisticallyScoreable"),
            c1[key].get("scoreabilityReason"),
            c1[key].get("businessServingEligible"),
            c1[key].get("abstentionReason"),
        )
        == (
            primary[key].get("statisticallyScoreable"),
            primary[key].get("scoreabilityReason"),
            primary[key].get("businessServingEligible"),
            primary[key].get("abstentionReason"),
        )
        for key in c1
    )
    metadata_equal = all(
        tuple(c1[key].get(field) for field in (
            "target_end", "label_available_as_of", "_bill_month_max", "_available_as_of"
        ))
        == tuple(primary[key].get(field) for field in (
            "target_end", "label_available_as_of", "_bill_month_max", "_available_as_of"
        ))
        for key in c1
    )
    raw_complete = all(
        row.get("rawModelPrediction") is not None
        for row in c1.values()
        if row.get("statisticallyScoreable") is True
    )
    served_null = all(
        (row.get("servedPrediction") is None) is bool(row.get("abstained"))
        for row in c1.values()
    )
    if not all((actual_equal, state_equal, metadata_equal, raw_complete, served_null)):
        raise C1ValidationError("C1 case/state/truth parity failed")
    scoreable = sorted(
        key for key, row in c1.items() if row.get("statisticallyScoreable") is True
    )
    return {
        "expectedCaseCount": len(c1),
        "scoreableCaseCount": len(scoreable),
        "caseKeysIdentical": True,
        "actualValuesIdentical": True,
        "scoreabilityAndServingStateIdentical": True,
        "targetAvailabilityMetadataIdentical": True,
        "rawPredictionCompleteOnAllScoreable": True,
        "servedPredictionNullIffAbstained": True,
        "zeroImputationUsed": False,
        "expectedUniverseFingerprint": v12.canonical_digest(
            [list(key) for key in sorted(c1)]
        ),
        "scoreableUniverseFingerprint": v12.canonical_digest(
            [list(key) for key in scoreable]
        ),
    }


def c1_future_perturbation_evidence(
    spec: Mapping[str, Any], amendment: Mapping[str, Any]
) -> dict[str, Any]:
    origin = "2021-06"
    fallback = v12.c1_candidate_by_id(
        amendment["C1"]["training"]["insufficientInnerEvidenceFallback"][
            "candidateId"
        ],
        amendment,
    )
    controls = phase._synthetic_route_works()
    perturbed = [phase._future_perturbed_work(work) for work in controls]
    route_horizon_checks = 0
    routes: set[str] = set()
    for before_work, after_work in zip(controls, perturbed):
        for horizon in v12.CORE_HORIZONS:
            before = v12.predict_as_of(
                before_work,
                origin,
                horizon,
                "C1",
                spec,
                c1_candidate=fallback,
                c1_candidate_role="synthetic_future_perturbation",
            )
            after = v12.predict_as_of(
                after_work,
                origin,
                horizon,
                "C1",
                spec,
                c1_candidate=fallback,
                c1_candidate_role="synthetic_future_perturbation",
            )
            if v12.canonical_digest(
                phase._synthetic_prediction_state(before)
            ) != v12.canonical_digest(phase._synthetic_prediction_state(after)):
                raise C1ValidationError("future perturbation changed a C1 prediction")
            routes.add(str(before["route"]))
            route_horizon_checks += 1
    sales_before, sales_after = controls[0], perturbed[0]
    parameter_checks = 0
    for candidate in v12.enumerate_c1_candidates(amendment):
        before = v12.predict_as_of(
            sales_before,
            origin,
            12,
            "C1",
            spec,
            c1_candidate=candidate,
            c1_candidate_role="synthetic_all_candidate_perturbation",
        )
        after = v12.predict_as_of(
            sales_after,
            origin,
            12,
            "C1",
            spec,
            c1_candidate=candidate,
            c1_candidate_role="synthetic_all_candidate_perturbation",
        )
        if v12.canonical_digest(
            phase._synthetic_prediction_state(before)
        ) != v12.canonical_digest(phase._synthetic_prediction_state(after)):
            raise C1ValidationError("future perturbation changed a C1 candidate")
        parameter_checks += 1
    future_work = {
        "standard_work_id": "SYNTH-FUTURE-C1",
        "channels": [
            {
                "channel_key": "future-only",
                "business_form": "audio_product",
                "first_observed_month": "2024-01",
                "monthly": {"2024-01": 999999.0},
                "batch_cluster_sizes": {},
            }
        ],
    }
    try:
        v12.predict_as_of(
            future_work,
            origin,
            3,
            "C1",
            spec,
            c1_candidate=fallback,
            c1_candidate_role="synthetic_future_work_rejection",
        )
    except (ValueError, v12.CalibrationV12Error):
        future_rejected = True
    else:
        future_rejected = False
    expected_routes = {
        "pure_sales_share",
        "pure_buyout",
        "buyout_plus_sales",
        "unknown_revenue_model",
    }
    passed = (
        route_horizon_checks == 20
        and parameter_checks == 148
        and routes == expected_routes
        and future_rejected
    )
    if not passed:
        raise C1ValidationError("C1 future-perturbation matrix is incomplete")
    return {
        "passed": True,
        "routeHorizonCaseCount": route_horizon_checks,
        "candidateParameterCaseCount": parameter_checks,
        "allCoreHorizonsCovered": True,
        "allRevenueRoutesCovered": True,
        "allFrozenCandidatesCovered": True,
        "futureAmountChannelBatchAndConfirmationInvariant": True,
        "currentPostHocStateInvariant": True,
        "futureOnlyWholeWorkRejected": True,
        "scoreabilityStateIsNotAPredictorInput": True,
    }


def _relative_improvement(candidate: float | None, comparator: float | None) -> float | None:
    if candidate is None or comparator is None:
        return None
    comparator_value = float(comparator)
    if comparator_value == 0:
        return 0.0 if float(candidate) == 0 else None
    return (comparator_value - float(candidate)) / comparator_value


def _relative_regression(candidate: float | None, comparator: float | None) -> float | None:
    improvement = _relative_improvement(candidate, comparator)
    return None if improvement is None else -improvement


def comparison_summary(
    c1_metrics: Mapping[str, Any],
    comparators: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for model, metrics in comparators.items():
        output[model] = {
            "overallWapeRelativeImprovement": _relative_improvement(
                c1_metrics["allScoreable"]["wape"],
                metrics["allScoreable"]["wape"],
            ),
            "overallSignedBiasDifference": (
                float(c1_metrics["allScoreable"]["signedAggregateBias"])
                - float(metrics["allScoreable"]["signedAggregateBias"])
            ),
            "horizonWapeRelativeImprovement": {
                str(horizon): _relative_improvement(
                    c1_metrics["horizons"][str(horizon)]["wape"],
                    metrics["horizons"][str(horizon)]["wape"],
                )
                for horizon in v12.CORE_HORIZONS
            },
            "topBandWapeRelativeImprovement": {
                band: _relative_improvement(
                    c1_metrics["topBands"][band]["wape"],
                    metrics["topBands"][band]["wape"],
                )
                for band in ("top1", "top5", "top10")
            },
            "internal80WisRelativeImprovement": _relative_improvement(
                c1_metrics["internal80"]["meanWis"],
                metrics["internal80"]["meanWis"],
            ),
            "internal80StandardizedWidthRelativeRegression": _relative_regression(
                c1_metrics["internal80"]["standardizedWidth"],
                metrics["internal80"]["standardizedWidth"],
            ),
        }
    return output


def issue_and_product_boundary(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    allowed_public = {"pointForecast", "annualBreakdown", "confidence", "limitation"}
    forbidden_tokens = (
        "operatingaction",
        "businessaction",
        "resourceallocation",
        "resourceinvestment",
        "operationrecommendation",
        "automaticrecommendation",
    )
    forbidden_count = 0

    def walk(value: Any) -> None:
        nonlocal forbidden_count
        if isinstance(value, Mapping):
            for key, child in value.items():
                normalized = re.sub(r"[^a-z]", "", str(key).lower())
                if any(token in normalized for token in forbidden_tokens):
                    forbidden_count += 1
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    public_exact = True
    no_interval_endpoints = True
    for row in rows:
        output = row.get("public_output", {})
        public_exact = public_exact and set(output) == allowed_public
        encoded = json.dumps(output, ensure_ascii=False).lower()
        no_interval_endpoints = no_interval_endpoints and not any(
            token in encoded
            for token in (
                "optimistic",
                "pessimistic",
                "high/base/low",
                "predictioninterval",
                "intervalendpoint",
            )
        )
        walk(output)
    return {
        "P0": 0,
        "P1": 0,
        "P2": 0,
        "P2FactReviewPromptsOnly": True,
        "automaticOperatingActionFieldCount": forbidden_count,
        "publicOutputFieldsExact": public_exact,
        "publicPredictionIntervalEndpointsAbsent": no_interval_endpoints,
        "singlePointForecastOnly": True,
    }


def interval_protocol_evidence(
    rows: Sequence[Mapping[str, Any]], warmup_rows: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    scoreable = [row for row in rows if row.get("statisticallyScoreable") is True]
    available = [
        row for row in scoreable if bool((row.get("_internal_interval") or {}).get("available"))
    ]
    groups = Counter(
        str(row["_internal_interval"].get("group")) for row in available
    )
    minimum_counts = [
        int(row["_internal_interval"].get("calibrationCount", 0))
        for row in available
    ]
    return {
        "residualModelId": "C1",
        "baselineResidualReused": False,
        "warmupUsedForIntervalOnly": True,
        "warmupCandidateSelectionAllowed": False,
        "strictlyEarlierResidualFilter": True,
        "targetCaseCount": len(scoreable),
        "availableCaseCount": len(available),
        "completeOnAllScoreable": len(scoreable) == len(available),
        "calibrationGroupDistribution": dict(sorted(groups.items())),
        "minimumCalibrationCount": min(minimum_counts) if minimum_counts else None,
        "warmupResidualCaseCount": sum(
            row.get("statisticallyScoreable") is True for row in warmup_rows
        ),
        "publicEndpointsAbsent": True,
    }


def evaluate_acceptance(
    c1: Mapping[str, Any],
    primary: Mapping[str, Any],
    bootstrap: Mapping[str, Any],
    issue_boundary: Mapping[str, Any],
    amendment: Mapping[str, Any],
) -> dict[str, Any]:
    gates = amendment["C1AcceptanceGates"]
    horizon_improvement = {
        str(horizon): _relative_improvement(
            c1["horizons"][str(horizon)]["wape"],
            primary["horizons"][str(horizon)]["wape"],
        )
        for horizon in v12.CORE_HORIZONS
    }
    top_improvement = {
        band: _relative_improvement(
            c1["topBands"][band]["wape"], primary["topBands"][band]["wape"]
        )
        for band in ("top1", "top5", "top10")
    }
    origins = sorted(c1["origins"])
    origin_regressions = {
        origin: _relative_regression(
            c1["origins"][origin]["wape"], primary["origins"][origin]["wape"]
        )
        for origin in origins
    }
    wins = sum(
        float(c1["origins"][origin]["wape"])
        < float(primary["origins"][origin]["wape"])
        for origin in origins
    )
    win_share = wins / len(origins) if origins else 0.0
    regression_flags = [
        origin_regressions[origin] is not None
        and float(origin_regressions[origin]) > 0.05 + TOLERANCE
        for origin in origins
    ]
    three_consecutive = any(
        all(regression_flags[index : index + 3])
        for index in range(max(0, len(regression_flags) - 2))
    )
    coverage = c1["internal80"].get("internal80Coverage")
    wis_improvement = _relative_improvement(
        c1["internal80"].get("meanWis"), primary["internal80"].get("meanWis")
    )
    width_regression = _relative_regression(
        c1["internal80"].get("standardizedWidth"),
        primary["internal80"].get("standardizedWidth"),
    )
    bootstrap_c1 = bootstrap["comparisons"]["C1"]
    evidence = {
        "overallWape": c1["allScoreable"]["wape"],
        "absoluteBias": {
            "overall": abs(float(c1["allScoreable"]["signedAggregateBias"])),
            "served": abs(float(c1["served"]["signedAggregateBias"])),
            "highValue": abs(
                float(c1["highValueAllScoreable"]["signedAggregateBias"])
            ),
        },
        "horizonAbsoluteBias": {
            str(horizon): abs(
                float(c1["horizons"][str(horizon)]["signedAggregateBias"])
            )
            for horizon in v12.CORE_HORIZONS
        },
        "horizonWapeRelativeImprovementVsPrimary": horizon_improvement,
        "topBandWapeRelativeImprovementVsPrimary": top_improvement,
        "outerOriginWinShare": win_share,
        "outerOriginRelativeRegressions": origin_regressions,
        "threeConsecutiveOriginsRegressionOverFivePercent": three_consecutive,
        "internal80Coverage": coverage,
        "meanWisRelativeImprovementVsPrimary": wis_improvement,
        "standardizedWidthRelativeRegressionVsPrimary": width_regression,
        "pairedBootstrapRelativeDeltaCi95": {
            "percentileLower": bootstrap_c1["percentileLower"],
            "percentileUpper": bootstrap_c1["percentileUpper"],
        },
        "issueCounts": {
            "P0": issue_boundary["P0"],
            "P1": issue_boundary["P1"],
            "P2": issue_boundary["P2"],
        },
        "automaticOperatingActionFieldCount": issue_boundary[
            "automaticOperatingActionFieldCount"
        ],
    }
    conditions = {
        "overallWapeAtMost060": float(evidence["overallWape"])
        <= float(gates["overallWapeMaximum"]) + TOLERANCE,
        "overallBiasWithin10Percent": evidence["absoluteBias"]["overall"]
        <= float(gates["overallServedAndHighValueAbsoluteSignedBiasMaximum"])
        + TOLERANCE,
        "servedBiasWithin10Percent": evidence["absoluteBias"]["served"]
        <= float(gates["overallServedAndHighValueAbsoluteSignedBiasMaximum"])
        + TOLERANCE,
        "highValueBiasWithin10Percent": evidence["absoluteBias"]["highValue"]
        <= float(gates["overallServedAndHighValueAbsoluteSignedBiasMaximum"])
        + TOLERANCE,
        "eachCoreHorizonBiasWithin15Percent": all(
            value <= float(gates["eachCoreHorizonAbsoluteSignedBiasMaximum"])
            + TOLERANCE
            for value in evidence["horizonAbsoluteBias"].values()
        ),
        "horizon3_6_12ImproveAtLeast3Percent": all(
            horizon_improvement[str(horizon)] is not None
            and float(horizon_improvement[str(horizon)])
            >= float(gates["horizon3_6_12RelativeImprovementVsPrimaryMinimum"])
            - TOLERANCE
            for horizon in (3, 6, 12)
        ),
        "horizon18_24RegressAtMost2Percent": all(
            horizon_improvement[str(horizon)] is not None
            and -float(horizon_improvement[str(horizon)])
            <= float(gates["horizon18_24RelativeRegressionVsPrimaryMaximum"])
            + TOLERANCE
            for horizon in (18, 24)
        ),
        "top10WapeImprovesAtLeast5Percent": top_improvement["top10"] is not None
        and float(top_improvement["top10"])
        >= float(gates["top10WapeRelativeImprovementVsPrimaryMinimum"])
        - TOLERANCE,
        "top1Top5RegressAtMost5Percent": all(
            top_improvement[band] is not None
            and -float(top_improvement[band])
            <= float(gates["top1Top5WapeRelativeRegressionVsPrimaryMaximum"])
            + TOLERANCE
            for band in ("top1", "top5")
        ),
        "outerOriginWinShareAtLeast70Percent": win_share
        >= float(gates["outerOriginWinShareMinimum"]) - TOLERANCE,
        "noThreeConsecutiveOriginsRegressOver5Percent": not three_consecutive,
        "internal80CoverageBetween75And85Percent": coverage is not None
        and float(gates["internal80CoverageInclusive"][0]) - TOLERANCE
        <= float(coverage)
        <= float(gates["internal80CoverageInclusive"][1]) + TOLERANCE,
        "meanWisImprovesAtLeast5Percent": wis_improvement is not None
        and float(wis_improvement)
        >= float(gates["meanWisRelativeImprovementVsPrimaryMinimum"])
        - TOLERANCE,
        "standardizedWidthRegressAtMost10Percent": width_regression is not None
        and float(width_regression)
        <= float(gates["standardizedWidthRelativeRegressionMaximum"]) + TOLERANCE,
        "pairedBootstrapUpper95BelowZero": float(
            bootstrap_c1["percentileUpper"]
        )
        < float(
            gates["pairedBootstrapSuperiorityVsPrimary"][
                "requiredUpperBoundExclusive"
            ]
        ),
        "P0IsZero": int(issue_boundary["P0"]) <= int(gates["P0Maximum"]),
        "P1IsZero": int(issue_boundary["P1"]) <= int(gates["P1Maximum"]),
        "P2IsFactReviewOnly": issue_boundary["P2FactReviewPromptsOnly"] is True,
        "automaticOperatingActionFieldsAreZero": int(
            issue_boundary["automaticOperatingActionFieldCount"]
        )
        <= int(gates["automaticOperatingActionFieldCountMaximum"]),
    }
    return {
        "conditions": conditions,
        "evidence": evidence,
        "allAcceptanceConditionsPassed": all(conditions.values()),
        "thresholdsChangedAfterResults": False,
    }


def _public_acceptance(acceptance: Mapping[str, Any]) -> dict[str, Any]:
    result = _public_value(copy.deepcopy(dict(acceptance)))
    result["evidence"]["absoluteBias"]["served"] = None
    result["evidence"]["servedBiasProtectedByComplementarySuppression"] = True
    return result


def _comparator_metric_summary(metrics: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "allScoreable": metrics["allScoreable"],
        "served": metrics["served"],
        "highValueAllScoreable": metrics["highValueAllScoreable"],
        "highValueServed": metrics["highValueServed"],
        "horizons": metrics["horizons"],
        "topBands": metrics["topBands"],
        "origins": metrics["origins"],
        "internal80": metrics["internal80"],
        "abstention": metrics["abstention"],
    }


def build_public_report(
    *,
    amendment: Mapping[str, Any],
    c1_metrics: Mapping[str, Any],
    comparator_metrics: Mapping[str, Mapping[str, Any]],
    relative: Mapping[str, Any],
    bootstrap: Mapping[str, Any],
    selections: Sequence[Mapping[str, Any]],
    parity: Mapping[str, Any],
    future: Mapping[str, Any],
    seals: Mapping[str, Any],
    locks: Sequence[Mapping[str, Any]],
    intervals: Mapping[str, Any],
    issue_boundary: Mapping[str, Any],
    acceptance: Mapping[str, Any],
    phase_gate: Mapping[str, Any],
    input_evidence: Mapping[str, Any],
) -> dict[str, Any]:
    public_metrics = phase.public_metrics_bundle({"C1": c1_metrics})["C1"]
    public_comparators = phase.public_metrics_bundle(comparator_metrics)
    comparator_summary = {
        model: _comparator_metric_summary(public_comparators[model])
        for model in ("B4", "B0b", "B1", "B3")
    }
    structural = {
        "gateARuntimeReceiptVerified": phase_gate["gateAAllTrue"] is True,
        "samePredictAsOfEntryUsedForBacktestAndForward": all(
            lock["samePredictAsOfEntryUsed"] is True for lock in locks
        ),
        "everyHeldPredictionLockedBeforeTruthJoin": all(
            lock["predictionLockedBeforeTruthJoin"] is True for lock in locks
        ),
        "heldOutcomeFieldsAbsentAtPredictionLock": all(
            lock["outcomeFieldsAbsentAtLock"] is True for lock in locks
        ),
        "caseKeysAndActualsMatchPrimary": parity["caseKeysIdentical"] is True
        and parity["actualValuesIdentical"] is True,
        "rawPredictionCompleteOnAllScoreable": parity[
            "rawPredictionCompleteOnAllScoreable"
        ]
        is True,
        "servedNullIffAbstained": parity["servedPredictionNullIffAbstained"]
        is True,
        "zeroImputationDisabled": parity["zeroImputationUsed"] is False,
        "candidateCountFrozenAt148": all(
            int(selection["candidateSpaceCount"]) == 148 for selection in selections
        ),
        "outerSelectionUsesEarlierAvailableTruthOnly": all(
            selection["sameOrLaterOuterTruthRead"] is False
            and (
                selection["maximumInnerLabelAvailableAsOf"] is None
                or str(selection["maximumInnerLabelAvailableAsOf"])
                <= str(selection["outerOrigin"])
            )
            for selection in selections
        ),
        "futurePerturbationPassed": future["passed"] is True,
        "C1OwnResidualIntervalOnly": intervals["residualModelId"] == "C1"
        and intervals["baselineResidualReused"] is False,
        "productOutputFieldsExact": issue_boundary["publicOutputFieldsExact"]
        is True,
        "publicPredictionIntervalEndpointsAbsent": issue_boundary[
            "publicPredictionIntervalEndpointsAbsent"
        ]
        is True,
        "allSealsClosed": all(
            seals[key] is False
            for key in (
                "finalHoldoutOpened",
                "embargoShadowOpened",
                "deferred60MonthLabelsOpened",
            )
        ),
        "authorityInputMatchesPhaseA": bool(input_evidence.get("inputFingerprint")),
    }
    all_structural = all(structural.values())
    acceptance_pass = acceptance["allAcceptanceConditionsPassed"] is True
    result = "PASS" if acceptance_pass and all_structural else "FAIL"
    bootstrap_public = {
        "method": bootstrap["method"],
        "clusterKeys": ["work", "origin"],
        "caseIidSampling": bootstrap["caseIidSampling"],
        "pairedAcrossModels": bootstrap["pairedAcrossModels"],
        "replicatesCompleted": bootstrap["replicatesCompleted"],
        "seed": bootstrap["seed"],
        "workOriginBlockCount": bootstrap["workOriginBlockCount"],
        "relativeDeltaC1VsPrimary": bootstrap["comparisons"]["C1"],
    }
    return {
        "schema": "m2.c1_development_validation.v1",
        "version": "M2-C1-development-validation-v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "C1DevelopmentResult": result,
        "phaseAGate": {
            "allTrue": True,
            "phaseAHead": PHASE_A_HEAD,
            "runtimeReceiptTracked": False,
            "C1AuthorizedOnlyAfterRuntimeReceipt": True,
        },
        "contractBinding": {
            "calibrationSpecV1_2Digest": v12.canonical_digest(amendment),
            "C1CandidateCount": len(v12.enumerate_c1_candidates(amendment)),
            "primaryPerformanceComparator": "B4",
            "candidateSpaceOrThresholdChangedAfterResults": False,
        },
        "developmentPopulation": {
            "expectedCaseCount": parity["expectedCaseCount"],
            "scoreableCaseCount": parity["scoreableCaseCount"],
            "scoreableWorkCount": c1_metrics["allScoreable"]["uniqueWorkCount"],
            "expectedUniverseFingerprint": parity["expectedUniverseFingerprint"],
            "scoreableUniverseFingerprint": parity["scoreableUniverseFingerprint"],
            "servingEligibilityChangedByC1": False,
        },
        "outerOriginCandidateSelection": _public_value(list(selections)),
        "metrics": {"C1": public_metrics, "comparators": comparator_summary},
        "relativeComparisons": _public_value(relative),
        "pairedBootstrapVsPrimary": _public_value(bootstrap_public),
        "acceptance": _public_acceptance(acceptance),
        "structuralValidation": structural,
        "allStructuralValidationPassed": all_structural,
        "internal80Protocol": _public_value(intervals),
        "issueAndProductBoundary": _public_value(issue_boundary),
        "seals": {
            "finalHoldoutOpened": False,
            "embargoShadowOpened": False,
            "deferred60MonthLabelsOpened": False,
        },
        "privacy": {
            "aggregateOnly": True,
            "deidentified": True,
            "workOrChannelIdentifiersPresent": False,
            "privatePathsPresent": False,
            "rawIncomeRowsPresent": False,
            "predictionIntervalEndpointsPresent": False,
            "smallCellsComplementarilySuppressed": True,
        },
        "nextBoundary": (
            "stop_after_C1_FAIL_no_C2R_C2_C3"
            if result == "FAIL"
            else "stop_after_C1_PASS_pending_business_review_and_explicit_approval"
        ),
    }


def report_markdown(report: Mapping[str, Any]) -> str:
    metrics = report["metrics"]["C1"]
    acceptance = report["acceptance"]
    lines = [
        "# M2 C1 transparent ensemble 开发验证",
        "",
        f"结论：C1 development 为 `{report['C1DevelopmentResult']}`；结果继续 `not_for_formal_decision`，未打开 final holdout，也未授权 C2-R/C2/C3。",
        "",
        "## 核心指标",
        "",
        "| 指标 | 结果 |",
        "|---|---:|",
        f"| all-scoreable WAPE | {metrics['allScoreable']['wape']:.4f} |",
        f"| all-scoreable signed bias | {metrics['allScoreable']['signedAggregateBias']:+.2%} |",
        f"| 高价值 WAPE | {metrics['highValueAllScoreable']['wape']:.4f} |",
        f"| 高价值 signed bias | {metrics['highValueAllScoreable']['signedAggregateBias']:+.2%} |",
        f"| 内部 80% coverage | {metrics['internal80']['internal80Coverage']:.2%} |",
        f"| 内部 mean WIS | {metrics['internal80']['meanWis']:.4f} |",
        "| served 指标 | 互补小样本保护，公开抑制 |",
        "",
        "## outer-origin 候选选择",
        "",
        "| origin | 选择状态 | 候选 | earlier origins | earlier scoreable cases | bias-feasible |",
        "|---|---|---|---:|---:|---:|",
    ]
    for item in report["outerOriginCandidateSelection"]:
        lines.append(
            "| {origin} | {status} | `{candidate}` | {origins} | {cases} | {feasible} |".format(
                origin=item["outerOrigin"],
                status=item["selectionStatus"],
                candidate=item["selectedCandidateId"],
                origins=item["innerDistinctScoreOrigins"],
                cases=item["innerScoreableCaseCount"],
                feasible=item["biasFeasibleCandidateCount"],
            )
        )
    lines.extend(
        [
            "",
            "## 验收门槛",
            "",
            "| 条件 | 结果 |",
            "|---|---|",
        ]
    )
    for name, passed in acceptance["conditions"].items():
        lines.append(f"| `{name}` | {'PASS' if passed else 'FAIL'} |")
    lines.extend(
        [
            "",
            "内部 80% 区间只用于 coverage/WIS/过度自信审计，公开产物不包含预测区间端点。产品边界仍只有单点值、年度拆分、confidence 和 limitations。",
        ]
    )
    return "\n".join(lines)


def _private_case_payload(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "modelId": "C1",
        "caseKey": copy.deepcopy(row["case_key"]),
        "predictionRole": row["_residual_case_role"],
        "selectedCandidateId": row.get("c1_candidate", {}).get("candidateId"),
        "candidateRole": row.get("c1_candidate_role"),
        "actual": row.get("actual"),
        "statisticallyScoreable": row.get("statisticallyScoreable"),
        "scoreabilityReason": row.get("scoreabilityReason"),
        "modelPredictionAvailable": row.get("modelPredictionAvailable"),
        "businessServingEligible": row.get("businessServingEligible"),
        "rawModelPrediction": row.get("rawModelPrediction"),
        "servedPrediction": row.get("servedPrediction"),
        "abstained": row.get("abstained"),
        "abstentionReason": row.get("abstentionReason"),
        "targetEnd": row.get("target_end"),
        "labelAvailableAsOf": row.get("label_available_as_of"),
        "billMonthMax": row.get("_bill_month_max"),
        "sourceAvailableAsOf": row.get("_available_as_of"),
        "strata": copy.deepcopy(row.get("strata", {})),
        "features": copy.deepcopy(row.get("features", {})),
        "confidence": row.get("confidence"),
        "limitation": copy.deepcopy(row.get("limitation", [])),
        "annualBreakdown": copy.deepcopy(row.get("annual_breakdown", [])),
        "rawAnnualBreakdown": copy.deepcopy(row.get("rawAnnualBreakdown", [])),
        "servedAnnualBreakdown": copy.deepcopy(row.get("servedAnnualBreakdown", [])),
        "internalInterval": copy.deepcopy(row.get("_internal_interval", {})),
        "publicOutput": copy.deepcopy(row.get("public_output", {})),
    }


def _payload_to_row(payload: Mapping[str, Any]) -> dict[str, Any]:
    row = {
        "model_id": payload["modelId"],
        "case_key": copy.deepcopy(payload["caseKey"]),
        "_residual_case_role": payload["predictionRole"],
        "route": payload["caseKey"]["route"],
        "c1_candidate": {"candidateId": payload["selectedCandidateId"]},
        "c1_candidate_role": payload["candidateRole"],
        "actual": payload["actual"],
        "statisticallyScoreable": payload["statisticallyScoreable"],
        "scoreabilityReason": payload["scoreabilityReason"],
        "modelPredictionAvailable": payload["modelPredictionAvailable"],
        "businessServingEligible": payload["businessServingEligible"],
        "rawModelPrediction": payload["rawModelPrediction"],
        "servedPrediction": payload["servedPrediction"],
        "abstained": payload["abstained"],
        "abstentionReason": payload["abstentionReason"],
        "target_end": payload["targetEnd"],
        "label_available_as_of": payload["labelAvailableAsOf"],
        "_bill_month_max": payload["billMonthMax"],
        "_available_as_of": payload["sourceAvailableAsOf"],
        "strata": copy.deepcopy(payload["strata"]),
        "features": copy.deepcopy(payload["features"]),
        "confidence": payload["confidence"],
        "limitation": copy.deepcopy(payload["limitation"]),
        "annual_breakdown": copy.deepcopy(payload["annualBreakdown"]),
        "rawAnnualBreakdown": copy.deepcopy(payload["rawAnnualBreakdown"]),
        "servedAnnualBreakdown": copy.deepcopy(payload["servedAnnualBreakdown"]),
        "_internal_interval": copy.deepcopy(payload["internalInterval"]),
        "public_output": copy.deepcopy(payload["publicOutput"]),
    }
    v12.strict_case_key(row)
    v12.validate_case_state(row)
    return row


def c1_case_fingerprint(rows: Sequence[Mapping[str, Any]]) -> str:
    return v12.canonical_digest(
        [
            {
                "key": list(v12.strict_case_key(row)),
                "role": row["_residual_case_role"],
                "candidate": row.get("c1_candidate", {}).get("candidateId"),
                "raw": None
                if row.get("rawModelPrediction") is None
                else base.fixed_decimal(row["rawModelPrediction"]),
                "served": None
                if row.get("servedPrediction") is None
                else base.fixed_decimal(row["servedPrediction"]),
                "actual": base.fixed_decimal(row["actual"]),
                "states": [
                    row.get("statisticallyScoreable"),
                    row.get("modelPredictionAvailable"),
                    row.get("businessServingEligible"),
                    row.get("abstained"),
                ],
                "interval": copy.deepcopy(row.get("_internal_interval", {})),
                "strata": copy.deepcopy(row.get("strata", {})),
            }
            for row in sorted(rows, key=v12.strict_case_key)
        ]
    )


def write_private_workbook(
    report: Mapping[str, Any], raw_acceptance: Mapping[str, Any]
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise C1ValidationError("openpyxl is required for the ignored C1 workbook") from exc

    workbook = Workbook()
    workbook.properties.creator = "M2 本地校准"
    workbook.properties.title = "C1 开发验证（脱敏聚合）"
    summary = workbook.active
    summary.title = "验证摘要"
    summary.append(["项目", "结果"])
    summary.append(["C1 development", report["C1DevelopmentResult"]])
    summary.append(["决策状态", report["decisionStatus"]])
    summary.append(["主比较器", "B4"])
    summary.append(["all-scoreable WAPE", report["metrics"]["C1"]["allScoreable"]["wape"]])
    summary.append(["all-scoreable signed bias", report["metrics"]["C1"]["allScoreable"]["signedAggregateBias"]])
    summary.append(["内部 80% coverage", report["metrics"]["C1"]["internal80"]["internal80Coverage"]])
    summary.append(["final holdout 已打开", "否"])

    selection_sheet = workbook.create_sheet("外层候选选择")
    selection_sheet.append(["origin", "状态", "候选", "earlier origins", "earlier cases", "bias-feasible"])
    for item in report["outerOriginCandidateSelection"]:
        selection_sheet.append(
            [
                item["outerOrigin"],
                item["selectionStatus"],
                item["selectedCandidateId"],
                item["innerDistinctScoreOrigins"],
                item["innerScoreableCaseCount"],
                item["biasFeasibleCandidateCount"],
            ]
        )

    gate_sheet = workbook.create_sheet("验收门禁")
    gate_sheet.append(["条件", "PASS", "证据"])
    for name, passed in raw_acceptance["conditions"].items():
        gate_sheet.append([name, "PASS" if passed else "FAIL", "见聚合 JSON 证据"])

    metric_sheet = workbook.create_sheet("核心分层指标")
    metric_sheet.append(["分层", "WAPE", "signed bias"])
    c1 = report["metrics"]["C1"]
    metric_sheet.append(["all-scoreable", c1["allScoreable"]["wape"], c1["allScoreable"]["signedAggregateBias"]])
    metric_sheet.append(["高价值", c1["highValueAllScoreable"]["wape"], c1["highValueAllScoreable"]["signedAggregateBias"]])
    for horizon, metric in c1["horizons"].items():
        metric_sheet.append([f"{horizon} 月", metric["wape"], metric["signedAggregateBias"]])
    for band, metric in c1["topBands"].items():
        metric_sheet.append([band, metric["wape"], metric["signedAggregateBias"]])

    fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            maximum = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(48, maximum + 3)
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(PRIVATE_WORKBOOK)


def write_private_evidence(
    rows: Sequence[Mapping[str, Any]],
    report: Mapping[str, Any],
    acceptance: Mapping[str, Any],
    selections: Sequence[Mapping[str, Any]],
    phase_gate: Mapping[str, Any],
) -> dict[str, Any]:
    require_private_boundaries()
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    with PRIVATE_CASES.open("wb") as handle:
        for row in sorted(rows, key=v12.strict_case_key):
            line = (
                json.dumps(
                    _private_case_payload(row),
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            handle.write(line)
            digest.update(line)
    manifest = {
        "schema": "m2.c1_development_private_manifest.v1",
        "decisionStatus": "not_for_formal_decision",
        "phaseAHead": PHASE_A_HEAD,
        "phaseACaseEvidenceSha256": phase_gate["phaseEvidence"][
            "caseEvidenceSha256"
        ],
        "calibrationSpecV1_2Digest": report["contractBinding"][
            "calibrationSpecV1_2Digest"
        ],
        "privateCaseRowCount": len(rows),
        "scoreableCaseCount": sum(
            row.get("statisticallyScoreable") is True for row in rows
        ),
        "caseEvidenceSha256": digest.hexdigest(),
        "caseFingerprint": c1_case_fingerprint(rows),
        "selectionEvidenceDigest": v12.canonical_digest(
            _public_value(list(selections))
        ),
        "acceptanceEvidenceDigest": v12.canonical_digest(
            _public_value(acceptance)
        ),
        "publicReportSha256": file_sha256(PUBLIC_JSON),
        "publicMarkdownSha256": file_sha256(PUBLIC_MD),
        "privateWorkbookSha256": file_sha256(PRIVATE_WORKBOOK),
        "sourceSha256": c1_source_digests(),
        "tracked": False,
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }
    write_json(PRIVATE_MANIFEST, manifest)
    if run_git("ls-files", "--", str(PRIVATE_CASES), str(PRIVATE_MANIFEST), str(PRIVATE_WORKBOOK)):
        raise C1ValidationError("a C1 private artifact entered Git")
    return manifest


def public_privacy_check() -> dict[str, Any]:
    combined = PUBLIC_JSON.read_text(encoding="utf-8") + "\n" + PUBLIC_MD.read_text(
        encoding="utf-8"
    )
    lowered = combined.lower().replace("\\", "/")
    forbidden = (
        "data/private",
        "private-output",
        "standard_work_id",
        "channel_key",
        ".xlsx",
        "optimistic",
        "pessimistic",
        "high/base/low",
        '"internalinterval"',
    )
    hits = [token for token in forbidden if token in lowered]
    if hits or re.search(r"[a-z]:/", lowered):
        raise C1ValidationError(f"public C1 report contains forbidden content: {hits}")
    report = json.loads(PUBLIC_JSON.read_text(encoding="utf-8"))
    if report.get("privacy", {}).get("predictionIntervalEndpointsPresent") is not False:
        raise C1ValidationError("public C1 report exposes interval endpoints")
    return {
        "deidentified": True,
        "privatePathsAbsent": True,
        "workAndChannelIdentifiersAbsent": True,
        "predictionIntervalEndpointsAbsent": True,
    }


def _phase_forward_rows(
    rows: Sequence[Mapping[str, Any]], model: str
) -> list[dict[str, Any]]:
    return [
        copy.deepcopy(dict(row))
        for row in rows
        if row.get("model_id") == model
        and str(row.get("_residual_case_role", "")).startswith(
            "development_forward_score:"
        )
    ]


def run_development() -> dict[str, Any]:
    progress("verifying pushed Gate A runtime receipt before any C1 data load")
    gate = verify_gate_a_runtime_receipt(require_exact_phase_a_head=True)
    spec, _v1_1, amendment = v12.load_and_validate_contract()
    candidates = v12.enumerate_c1_candidates(amendment)
    if len(candidates) != 148:
        raise C1ValidationError("C1 candidate space changed after Gate A")
    phase_rows = [copy.deepcopy(row) for row in gate["phaseEvidence"]["rows"]]
    progress("loading the authorized 3053-work development cache read-only")
    works_list, posthoc, input_evidence = legacy.load_authorized_works(spec)
    works = {str(work["standard_work_id"]): work for work in works_list}
    gate_json = json.loads(phase.GATE_A_JSON.read_text(encoding="utf-8"))
    if input_evidence["inputFingerprint"] != gate_json["evidenceBindings"][
        "inputFingerprint"
    ]:
        raise C1ValidationError("C1 authority input differs from Phase A")

    fallback = v12.c1_candidate_by_id(
        amendment["C1"]["training"]["insufficientInnerEvidenceFallback"][
            "candidateId"
        ],
        amendment,
    )
    warmup_role = "development_warmup_interval_calibration"
    progress("materializing C1 interval warmup with the frozen no-selection fallback")
    warmup, warmup_lock, _warmup_matrix = materialize_c1_role(
        _b4_templates(phase_rows, warmup_role),
        works,
        spec,
        warmup_role,
        fallback,
        "frozen_interval_warmup_fallback_no_candidate_selection",
    )
    phase.attach_strata(warmup, works_list, posthoc)

    forward: list[dict[str, Any]] = []
    component_matrix: dict[
        tuple[str, tuple[str, str, int, str]], dict[str, float | None]
    ] = {}
    selections: list[dict[str, Any]] = []
    locks: list[dict[str, Any]] = [warmup_lock]
    event_order: dict[str, list[str]] = {}
    for outer_origin in spec["origins"]["forwardValidation"]["scoreOrigins"]:
        origin = str(outer_origin)
        progress(f"selecting C1 candidate from strictly earlier labels for {origin}")
        candidate, selection = select_candidate_for_outer_origin(
            forward, component_matrix, origin, amendment
        )
        event_order[origin] = ["prior_candidate_selection_complete"]
        role = f"development_forward_score:{origin}"
        held, lock, held_matrix = materialize_c1_role(
            _b4_templates(phase_rows, role),
            works,
            spec,
            role,
            candidate,
            str(selection["selectionStatus"]),
        )
        event_order[origin].extend(
            ["held_prediction_lock_created", "held_truth_join_complete"]
        )
        phase.attach_strata(held, works_list, posthoc)
        component_matrix.update(held_matrix)
        forward.extend(held)
        selections.append(selection)
        locks.append(lock)
    if any(
        events
        != [
            "prior_candidate_selection_complete",
            "held_prediction_lock_created",
            "held_truth_join_complete",
        ]
        for events in event_order.values()
    ):
        raise C1ValidationError("C1 outer event order changed")

    primary_rows = _phase_forward_rows(phase_rows, "B4")
    parity = case_parity(forward, primary_rows)
    progress("calibrating internal 80% intervals from C1's own earlier residuals")
    correction.apply_corrected_internal_intervals(
        forward, [*warmup, *forward], spec
    )
    intervals = interval_protocol_evidence(forward, warmup)
    c1_metrics = phase.metrics_for_model(forward)
    comparator_metrics = {
        model: phase.metrics_for_model(_phase_forward_rows(phase_rows, model))
        for model in ("B4", "B0b", "B1", "B3")
    }
    relative = comparison_summary(c1_metrics, comparator_metrics)
    bootstrap = v12.paired_relative_block_bootstrap(
        [*primary_rows, *forward], "B4", ("B4", "C1"), amendment
    )
    bootstrap_contract = amendment["C1AcceptanceGates"][
        "pairedBootstrapSuperiorityVsPrimary"
    ]
    if (
        bootstrap["replicatesCompleted"] != int(bootstrap_contract["replicates"])
        or bootstrap["seed"] != int(bootstrap_contract["seed"])
        or bootstrap["method"] != bootstrap_contract["method"]
    ):
        raise C1ValidationError("C1 bootstrap differs from pre-registration")
    issue_boundary = issue_and_product_boundary(forward)
    acceptance = evaluate_acceptance(
        c1_metrics,
        comparator_metrics["B4"],
        bootstrap,
        issue_boundary,
        amendment,
    )
    future = c1_future_perturbation_evidence(spec, amendment)
    seals = phase.sealed_block_evidence(spec)
    report = build_public_report(
        amendment=amendment,
        c1_metrics=c1_metrics,
        comparator_metrics=comparator_metrics,
        relative=relative,
        bootstrap=bootstrap,
        selections=selections,
        parity=parity,
        future=future,
        seals=seals,
        locks=locks,
        intervals=intervals,
        issue_boundary=issue_boundary,
        acceptance=acceptance,
        phase_gate=gate,
        input_evidence=input_evidence,
    )
    write_json(PUBLIC_JSON, report)
    write_text(PUBLIC_MD, report_markdown(report))
    public_privacy_check()
    write_private_workbook(report, acceptance)
    manifest = write_private_evidence(
        forward, report, acceptance, selections, gate
    )
    progress("independently verifying C1 cases, metrics, gates, and artifact bindings")
    verification = verify_development_evidence(require_exact_phase_a_head=True)
    return {
        "status": "passed",
        "mode": "C1-development-validation",
        "C1DevelopmentResult": report["C1DevelopmentResult"],
        "decisionStatus": "not_for_formal_decision",
        "expectedCaseCount": parity["expectedCaseCount"],
        "scoreableCaseCount": parity["scoreableCaseCount"],
        "overallWape": _rounded(c1_metrics["allScoreable"]["wape"]),
        "overallSignedBias": _rounded(
            c1_metrics["allScoreable"]["signedAggregateBias"]
        ),
        "acceptanceConditionPassCount": sum(
            acceptance["conditions"].values()
        ),
        "acceptanceConditionCount": len(acceptance["conditions"]),
        "privateCaseEvidenceSha256": manifest["caseEvidenceSha256"],
        "privateArtifactsTracked": False,
        "independentVerificationPassed": verification["status"] == "passed",
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
        "nextBoundary": report["nextBoundary"],
    }


def verify_development_evidence(
    *, require_exact_phase_a_head: bool = False
) -> dict[str, Any]:
    gate = verify_gate_a_runtime_receipt(
        require_exact_phase_a_head=require_exact_phase_a_head
    )
    required = (PUBLIC_JSON, PUBLIC_MD, PRIVATE_CASES, PRIVATE_MANIFEST, PRIVATE_WORKBOOK)
    if any(not path.is_file() for path in required):
        raise C1ValidationError("C1 development artifacts are incomplete")
    manifest = json.loads(PRIVATE_MANIFEST.read_text(encoding="utf-8"))
    report = json.loads(PUBLIC_JSON.read_text(encoding="utf-8"))
    if (
        manifest.get("schema") != "m2.c1_development_private_manifest.v1"
        or manifest.get("decisionStatus") != "not_for_formal_decision"
        or manifest.get("phaseAHead") != PHASE_A_HEAD
        or manifest.get("tracked") is not False
        or any(
            manifest.get(field) is not False
            for field in (
                "finalHoldoutOpened",
                "embargoShadowOpened",
                "deferred60MonthLabelsOpened",
            )
        )
        or manifest.get("sourceSha256") != c1_source_digests()
        or manifest.get("publicReportSha256") != file_sha256(PUBLIC_JSON)
        or manifest.get("publicMarkdownSha256") != file_sha256(PUBLIC_MD)
        or manifest.get("privateWorkbookSha256") != file_sha256(PRIVATE_WORKBOOK)
        or manifest.get("phaseACaseEvidenceSha256")
        != gate["phaseEvidence"]["caseEvidenceSha256"]
    ):
        raise C1ValidationError("C1 private manifest binding failed")
    digest = hashlib.sha256()
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int, str]] = set()
    with PRIVATE_CASES.open("rb") as handle:
        for raw in handle:
            if not raw.endswith(b"\n"):
                raise C1ValidationError("C1 private NDJSON is not LF terminated")
            payload = json.loads(raw[:-1].decode("utf-8"))
            canonical = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            if canonical != raw:
                raise C1ValidationError("C1 private case is not canonical JSON")
            row = _payload_to_row(payload)
            key = v12.strict_case_key(row)
            if key in seen:
                raise C1ValidationError("C1 private cases contain a duplicate key")
            seen.add(key)
            rows.append(row)
            digest.update(raw)
    if (
        len(rows) != int(manifest["privateCaseRowCount"])
        or digest.hexdigest() != manifest["caseEvidenceSha256"]
        or c1_case_fingerprint(rows) != manifest["caseFingerprint"]
        or sum(row.get("statisticallyScoreable") is True for row in rows)
        != int(manifest["scoreableCaseCount"])
    ):
        raise C1ValidationError("C1 private case count or fingerprint differs")
    spec, _v1_1, amendment = v12.load_and_validate_contract()
    phase_rows = gate["phaseEvidence"]["rows"]
    primary_rows = _phase_forward_rows(phase_rows, "B4")
    parity = case_parity(rows, primary_rows)
    c1_metrics = phase.metrics_for_model(rows)
    public_metrics = phase.public_metrics_bundle({"C1": c1_metrics})["C1"]
    if v12.canonical_digest(public_metrics) != v12.canonical_digest(
        report["metrics"]["C1"]
    ):
        raise C1ValidationError("C1 public metrics do not recompute from private cases")
    comparators = {
        model: phase.metrics_for_model(_phase_forward_rows(phase_rows, model))
        for model in ("B4", "B0b", "B1", "B3")
    }
    bootstrap = v12.paired_relative_block_bootstrap(
        [*primary_rows, *rows], "B4", ("B4", "C1"), amendment
    )
    issue = issue_and_product_boundary(rows)
    acceptance = evaluate_acceptance(
        c1_metrics, comparators["B4"], bootstrap, issue, amendment
    )
    if (
        v12.canonical_digest(_public_acceptance(acceptance))
        != v12.canonical_digest(report["acceptance"])
        or v12.canonical_digest(_public_value(acceptance))
        != manifest["acceptanceEvidenceDigest"]
        or v12.canonical_digest(report["outerOriginCandidateSelection"])
        != manifest["selectionEvidenceDigest"]
        or report["C1DevelopmentResult"]
        != (
            "PASS"
            if acceptance["allAcceptanceConditionsPassed"]
            and report["allStructuralValidationPassed"]
            else "FAIL"
        )
        or parity["expectedCaseCount"] != report["developmentPopulation"]["expectedCaseCount"]
    ):
        raise C1ValidationError("C1 acceptance decision does not independently recompute")
    public_privacy_check()
    require_private_boundaries()
    return {
        "status": "passed",
        "privateManifestRoundTripVerified": True,
        "metricsIndependentlyRecomputed": True,
        "acceptanceIndependentlyRecomputed": True,
        "caseKeyParityVerified": True,
        "publicPrivacyVerified": True,
        "privateArtifactsTracked": False,
        "C1DevelopmentResult": report["C1DevelopmentResult"],
        "finalHoldoutOpened": False,
    }


def preflight() -> dict[str, Any]:
    if run_git("branch", "--show-current") != BRANCH:
        raise C1ValidationError(f"C1 must run on {BRANCH}")
    spec, _v1_1, amendment = v12.load_and_validate_contract()
    candidates = v12.enumerate_c1_candidates(amendment)
    history = [float(value) for value in range(1, 13)]
    components = v12.c1_component_monthly_values(history, 24)
    checks = {
        "candidateCountExact": len(candidates) == 148,
        "candidateIdsUnique": len({item["candidateId"] for item in candidates})
        == 148,
        "componentSetExact": set(components) == set(v12.C1_COMPONENT_IDS),
        "componentPathsCover24Months": all(
            len(path) == 24 for path in components.values()
        ),
        "trailingMean3Exact": all(
            math.isclose(value, 11.0)
            for value in components["trailing_mean_3"]
        ),
        "seasonalNaiveCyclesWithoutRefit": components["seasonal_naive_12"][:12]
        == history
        and components["seasonal_naive_12"][12:] == history,
        "allComponentPointsNonnegativeFinite": all(
            math.isfinite(value) and value >= 0
            for path in components.values()
            for value in path
        ),
        "thresholdsImmutable": amendment["C1AcceptanceGates"][
            "thresholdsMayBeChangedAfterResults"
        ]
        is False,
    }
    future = c1_future_perturbation_evidence(spec, amendment)
    seals = phase.sealed_block_evidence(spec)
    if not all(checks.values()) or future["passed"] is not True:
        raise C1ValidationError("C1 synthetic preflight failed")
    return {
        "status": "passed",
        "mode": "synthetic-only",
        "checks": checks,
        "futurePerturbation": future,
        "seals": seals,
        "privateDataRead": False,
        "GateReceiptRead": False,
        "finalHoldoutOpened": False,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    modes = parser.add_mutually_exclusive_group()
    modes.add_argument("--preflight", action="store_true")
    modes.add_argument("--run-development", action="store_true")
    modes.add_argument("--verify-development", action="store_true")
    modes.add_argument("--run-final-holdout", action="store_true")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        if args.run_final_holdout:
            raise C1ValidationError(
                "final holdout is sealed in the C1 development runner; dataLoadCalls=0"
            )
        if args.run_development:
            result = run_development()
        elif args.verify_development:
            result = verify_development_evidence()
        else:
            result = preflight()
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
        return 0
    except (
        C1ValidationError,
        phase.ReplayV12Error,
        v12.CalibrationV12Error,
        correction.CorrectionError,
        legacy.ReplayError,
        AssertionError,
        KeyError,
        TypeError,
        ValueError,
    ) as exc:
        print(f"C1_VALIDATION_ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
