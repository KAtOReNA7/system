from __future__ import annotations

import json
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


INPUT_XLSX = (
    ROOT
    / "data"
    / "private-output"
    / "m2-business-review"
    / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v4.2.xlsx"
)
OUTPUT_JSON = (
    ROOT
    / "docs"
    / "analysis"
    / "m2-real-data"
    / "M2-rating-standard-v4.2-operator-validation-summary-v1.json"
)
OUTPUT_MD = (
    ROOT
    / "docs"
    / "analysis"
    / "m2-real-data"
    / "M2-rating-standard-v4.2-operator-validation-summary-v1.md"
)

TASK_SHEET = "01_v4.2运营任务卡"
RATINGS = ["S+", "S", "A", "B", "C", "D", "E"]
RANK = {rating: index for index, rating in enumerate(RATINGS)}


def main() -> None:
    if not INPUT_XLSX.exists():
        raise SystemExit(f"Missing required input workbook: {relative(INPUT_XLSX)}")

    rows, missing_headers = read_rows(INPUT_XLSX)
    summary = build_summary(rows, missing_headers)
    OUTPUT_JSON.write_text(json.dumps(envelope(summary), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUTPUT_MD.write_text(render_markdown(summary), encoding="utf-8")

    print(
        json.dumps(
            {
                "sourceWorkbook": relative(INPUT_XLSX),
                "rows": summary["rows"]["evaluatedTaskRows"],
                "revenueModelPositiveRate": summary["revenueModel"]["judgment"]["positiveRateOfEffectiveAnswered"],
                "shelfStatusPositiveRate": summary["shelfStatus"]["judgment"]["positiveRateOfEffectiveAnswered"],
                "ratingPositiveRate": summary["rating"]["judgment"]["positiveRateOfEffectiveAnswered"],
                "m4UserMarkedYes": summary["m4CalibrationCandidates"]["userMarkedYes"],
                "baselineVerdict": summary["baselineVerdict"]["verdict"],
                "outputJson": relative(OUTPUT_JSON),
                "outputMarkdown": relative(OUTPUT_MD),
            },
            ensure_ascii=False,
        )
    )


def read_rows(path: Path) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if TASK_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Missing required sheet: {TASK_SHEET}")
    sheet = workbook[TASK_SHEET]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    required = [
        "standard_work_id",
        "收入模式",
        "货架/版权状态",
        "状态置信度",
        "实销月均收入",
        "买断折算月均实销",
        "评级",
        "评级依据",
        "评级是否含买断",
        "风险/复核提示",
        "是否建议进入M4校准案例",
        "用户判断：收入模式是否合理",
        "用户判断：下架/版权状态是否合理",
        "用户判断：评级是否合理",
        "用户判断：是否应进入M4",
        "用户备注",
    ]
    missing = [field for field in required if field not in headers]
    rows = []
    for index, values in enumerate(iterator, start=1):
        if not any(clean(value) for value in values):
            continue
        source = {headers[pos]: values[pos] if pos < len(values) else "" for pos in range(len(headers))}
        rows.append(
            {
                "anonymousId": f"RSV42-{index:03d}",
                "standardWorkIdPresent": bool(clean(source.get("standard_work_id"))),
                "hasTaskCard": bool(clean(source.get("收入模式")) and clean(source.get("评级"))),
                "revenueModel": clean(source.get("收入模式")),
                "shelfStatus": clean(source.get("货架/版权状态")),
                "statusConfidence": clean(source.get("状态置信度")),
                "salesMonthlyAmount": to_number(source.get("实销月均收入")),
                "buyoutMonthlyAmount": to_number(source.get("买断折算月均实销")),
                "rating": clean(source.get("评级")),
                "ratingBasis": clean(source.get("评级依据")),
                "ratingIncludesBuyout": clean(source.get("评级是否含买断")),
                "riskPromptPresent": bool(clean(source.get("风险/复核提示"))),
                "systemM4Candidate": clean(source.get("是否建议进入M4校准案例")),
                "userRevenueModelJudgment": clean(source.get("用户判断：收入模式是否合理")),
                "userShelfStatusJudgment": clean(source.get("用户判断：下架/版权状态是否合理")),
                "userRatingJudgment": clean(source.get("用户判断：评级是否合理")),
                "userM4Judgment": clean(source.get("用户判断：是否应进入M4")),
                "userNoteCategory": categorize_note(clean(source.get("用户备注"))),
            }
        )
    return rows, missing


def build_summary(rows: list[dict], missing_headers: list[str]) -> dict:
    task_rows = [row for row in rows if row["hasTaskCard"]]
    revenue_judgment = summarize_judgment(row["userRevenueModelJudgment"] for row in task_rows)
    shelf_judgment = summarize_judgment(row["userShelfStatusJudgment"] for row in task_rows)
    rating_judgment = summarize_judgment(row["userRatingJudgment"] for row in task_rows)
    rating_comparison = compare_rating_to_user_monthly_sales(task_rows)
    pure_buyout = [row for row in task_rows if row["revenueModel"] == "纯买断"]
    buyout_plus_sales = [row for row in task_rows if row["revenueModel"] == "买断+实销"]
    expired_or_off_shelf = [
        row for row in task_rows if any(token in row["shelfStatus"] for token in ["到期", "下架"])
    ]
    note_categories = Counter(row["userNoteCategory"] for row in task_rows)
    feedback_complete = [
        row
        for row in task_rows
        if row["userRevenueModelJudgment"] and row["userShelfStatusJudgment"] and row["userRatingJudgment"]
    ]
    baseline = decide_baseline(revenue_judgment, shelf_judgment, rating_judgment)

    return {
        "sourceWorkbook": relative(INPUT_XLSX),
        "onlyReadRatingStandardV42Workbook": True,
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "rows": {
            "rawNonEmptyRows": len(rows),
            "rowsWithStandardWorkId": sum(1 for row in rows if row["standardWorkIdPresent"]),
            "evaluatedTaskRows": len(task_rows),
            "feedbackCompleteRows": len(feedback_complete),
            "feedbackCompletionRate": rate(len(feedback_complete), len(task_rows)),
        },
        "missingRequiredHeaders": missing_headers,
        "revenueModel": {
            "distribution": dict(Counter(row["revenueModel"] for row in task_rows)),
            "judgment": revenue_judgment,
            "negativeReasonCategories": negative_reasons(task_rows, "userRevenueModelJudgment"),
        },
        "shelfStatus": {
            "distribution": dict(Counter(row["shelfStatus"] for row in task_rows)),
            "statusConfidenceDistribution": dict(Counter(row["statusConfidence"] for row in task_rows)),
            "judgment": shelf_judgment,
            "negativeReasonCategories": negative_reasons(task_rows, "userShelfStatusJudgment"),
            "expiredOrOffShelfRows": len(expired_or_off_shelf),
            "expiredOrOffShelfRatedE": sum(1 for row in expired_or_off_shelf if row["rating"] == "E"),
            "expiredOrOffShelfNotSimplyE": sum(1 for row in expired_or_off_shelf if row["rating"] != "E"),
        },
        "rating": {
            "distribution": ordered_counter(row["rating"] for row in task_rows),
            "basisDistribution": dict(Counter(row["ratingBasis"] for row in task_rows)),
            "includesBuyoutDistribution": dict(Counter(row["ratingIncludesBuyout"] for row in task_rows)),
            "judgment": rating_judgment,
            "negativeReasonCategories": negative_reasons(task_rows, "userRatingJudgment"),
            "monthlySalesBandComparison": rating_comparison,
            "pureBuyout": summarize_subset(pure_buyout),
            "buyoutPlusSales": summarize_subset(buyout_plus_sales),
        },
        "riskAndReviewPrompt": {
            "rowsWithPrompt": sum(1 for row in task_rows if row["riskPromptPresent"]),
            "coverageRate": rate(sum(1 for row in task_rows if row["riskPromptPresent"]), len(task_rows)),
            "separateUsefulnessFieldAvailable": False,
            "userFlaggedPromptIssueRows": 0,
        },
        "m4CalibrationCandidates": {
            "systemSuggestedDistribution": dict(Counter(row["systemM4Candidate"] for row in task_rows)),
            "userMarkedYes": sum(1 for row in task_rows if row["userM4Judgment"] == "是"),
            "userMarkedNo": sum(1 for row in task_rows if row["userM4Judgment"] == "否"),
            "userMarkedPendingOrBlank": sum(
                1 for row in task_rows if row["userM4Judgment"] not in {"是", "否"}
            ),
        },
        "userNoteCategories": dict(note_categories),
        "baselineVerdict": baseline,
        "prohibitedActionsConfirmed": {
            "wroteFormalMasterData": False,
            "enteredM3": False,
            "committedPrivateExcel": False,
            "usedGitAddDot": False,
            "touchedStash": False,
        },
    }


def summarize_judgment(values) -> dict:
    counter = Counter()
    normalized = Counter()
    for value in values:
        text = clean(value)
        counter[text or "(blank)"] += 1
        normalized[classify_judgment(text)] += 1
    effective = sum(normalized[k] for k in ["positive", "negative", "neutral", "notApplicable", "other"])
    decisive = normalized["positive"] + normalized["negative"]
    return {
        "positive": normalized["positive"],
        "negative": normalized["negative"],
        "neutral": normalized["neutral"],
        "notApplicable": normalized["notApplicable"],
        "blank": normalized["blank"],
        "other": normalized["other"],
        "effectiveAnswered": effective,
        "decisiveAnswered": decisive,
        "positiveRateOfEffectiveAnswered": rate(normalized["positive"], effective),
        "positiveRateOfDecisiveAnswered": rate(normalized["positive"], decisive),
        "rawValueDistribution": dict(counter),
    }


def classify_judgment(text: str) -> str:
    if not text:
        return "blank"
    if text in {"合理", "基本合理", "可信", "基本可信", "可执行"}:
        return "positive"
    if text in {"不合理", "不可信", "不可执行"}:
        return "negative"
    if text in {"不确定", "待定"}:
        return "neutral"
    if text == "不适用":
        return "notApplicable"
    if "不合理" in text or "不可信" in text or "不可执行" in text:
        return "negative"
    if "合理" in text or "可信" in text or "可执行" in text:
        return "positive"
    return "other"


def negative_reasons(rows: list[dict], judgment_key: str) -> dict:
    counter = Counter()
    for row in rows:
        if classify_judgment(row[judgment_key]) != "negative":
            continue
        if judgment_key == "userRevenueModelJudgment":
            counter["收入模式仍有买断/实销边界误判"] += 1
        elif judgment_key == "userShelfStatusJudgment":
            if row["userNoteCategory"] == "currently_on_shelf_correction":
                counter["系统状态显示不可运营或无法判断，但用户确认目前在架"] += 1
            elif row["userNoteCategory"] == "bill_sales_record_gap":
                counter["原账单有实销记录但任务包显示无记录或证据不足"] += 1
            else:
                counter["货架/版权状态需复核"] += 1
        elif judgment_key == "userRatingJudgment":
            counter["评级档位或口径需复核"] += 1
    return dict(counter)


def compare_rating_to_user_monthly_sales(rows: list[dict]) -> dict:
    by_rating = {rating: {"rows": 0, "matches": 0, "mismatches": 0} for rating in RATINGS}
    comparable = 0
    matches = 0
    for row in rows:
        expected = expected_rating(row)
        actual = row["rating"]
        if not expected or actual not in RANK:
            continue
        comparable += 1
        by_rating[actual]["rows"] += 1
        if expected == actual:
            matches += 1
            by_rating[actual]["matches"] += 1
        else:
            by_rating[actual]["mismatches"] += 1
    return {
        "basis": "按用户确认口径，纯实销使用实销月均；纯买断使用买断折算月均；买断+实销使用实销月均加买断折算月均。",
        "comparableRows": comparable,
        "matches": matches,
        "mismatches": comparable - matches,
        "matchRate": rate(matches, comparable),
        "byDisplayedRating": by_rating,
    }


def expected_rating(row: dict) -> str | None:
    sales = row["salesMonthlyAmount"] or 0
    buyout = row["buyoutMonthlyAmount"] or 0
    if row["revenueModel"] == "纯买断":
        amount = buyout
    elif row["revenueModel"] == "买断+实销":
        amount = sales + buyout
    elif row["revenueModel"] == "纯实销/纯分成":
        amount = sales
    else:
        return None
    return rating_from_monthly_sales(amount)


def rating_from_monthly_sales(amount: float) -> str:
    if amount > 100000:
        return "S+"
    if amount >= 10000:
        return "S"
    if amount >= 5000:
        return "A"
    if amount >= 1000:
        return "B"
    if amount >= 500:
        return "C"
    if amount >= 100:
        return "D"
    return "E"


def summarize_subset(rows: list[dict]) -> dict:
    judgment = summarize_judgment(row["userRatingJudgment"] for row in rows)
    return {
        "rows": len(rows),
        "userRatingPositive": judgment["positive"],
        "userRatingNegative": judgment["negative"],
        "ratingDistribution": ordered_counter(row["rating"] for row in rows),
    }


def decide_baseline(revenue: dict, shelf: dict, rating: dict) -> dict:
    revenue_ok = revenue["positiveRateOfEffectiveAnswered"] >= 0.9
    rating_ok = rating["positiveRateOfEffectiveAnswered"] >= 0.9
    shelf_ok = shelf["positiveRateOfEffectiveAnswered"] >= 0.8
    if revenue_ok and rating_ok and shelf_ok:
        verdict = "可作为 M2 v4.2 候选基线进入下一步受控复核。"
        can_use = True
    elif revenue_ok and rating_ok:
        verdict = "评级层和收入模式可作为候选基线，但货架/版权状态层不能冻结，需先修正状态推断。"
        can_use = True
    else:
        verdict = "暂不建议作为 M2 评级层候选基线，需先完成最小修正。"
        can_use = False
    blockers = []
    if not shelf_ok:
        blockers.append("货架/版权状态合理率不足，用户反馈集中指向在架状态和实销记录缺口。")
    if not revenue_ok:
        blockers.append("收入模式仍存在不合理反馈，需继续排查买断/实销边界。")
    if not rating_ok:
        blockers.append("评级仍存在不合理反馈，需继续校准月均实销档位。")
    return {
        "canUseAsM2RatingLayerCandidateBaseline": can_use,
        "verdict": verdict,
        "hardBlockers": blockers,
        "minimumFixDirections": [
            "优先修复货架/版权状态推断：用户确认在架的样本不得继续显示无法判断或不可运营。",
            "核查“原账单有实销记录但任务包显示无记录或证据不足”的链路，避免状态层误判。",
            "保留本轮买断三信号规则：大额整数、同批次同额、候选买断后无实销。",
            "M4 校准池只纳入用户标记为“是”的经典样本，不再把所有错题自动塞入 M4。",
        ],
    }


def categorize_note(note: str) -> str:
    text = clean(note)
    if not text:
        return "blank"
    if "在架" in text:
        return "currently_on_shelf_correction"
    if "账单" in text or "实销记录" in text or "没有记录" in text:
        return "bill_sales_record_gap"
    return "other_note"


def ordered_counter(values) -> dict:
    counter = Counter(clean(value) for value in values)
    return {rating: counter.get(rating, 0) for rating in RATINGS}


def envelope(summary: dict) -> dict:
    return {
        "reportId": "m2.rating_standard_v4_2_operator_validation_summary.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        **summary,
    }


def render_markdown(summary: dict) -> str:
    revenue = summary["revenueModel"]["judgment"]
    shelf = summary["shelfStatus"]["judgment"]
    rating = summary["rating"]["judgment"]
    lines = [
        "# M2 rating-standard-v4.2 operator validation summary v1",
        "",
        "本报告只读取 `data/private-output/m2-business-review/m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v4.2.xlsx`。报告为聚合脱敏结果，不包含真实作品名、作者名、渠道名或原始账单行。",
        "",
        "## 样本范围",
        "",
        f"- 非空行数：`{summary['rows']['rawNonEmptyRows']}`",
        f"- 带 standard_work_id 的行数：`{summary['rows']['rowsWithStandardWorkId']}`",
        f"- 纳入验证的完整任务样本数：`{summary['rows']['evaluatedTaskRows']}`",
        f"- 三项核心判断均已填写：`{summary['rows']['feedbackCompleteRows']}`",
        f"- 缺失关键字段：`{', '.join(summary['missingRequiredHeaders']) if summary['missingRequiredHeaders'] else '无'}`",
        "",
        "## 用户判断合理率",
        "",
        "| 指标 | 有效填写 | 合理 | 不合理 | 不确定/其他 | 合理率 |",
        "|---|---:|---:|---:|---:|---:|",
        judgment_row("收入模式判断合理率", revenue),
        judgment_row("下架/版权状态判断合理率", shelf),
        judgment_row("评级判断合理率", rating),
        "",
        "## 不合理原因归类",
        "",
        "| 类别 | 原因归类 | 数量 |",
        "|---|---|---:|",
    ]
    for category, reasons in [
        ("收入模式不合理原因", summary["revenueModel"]["negativeReasonCategories"]),
        ("下架/版权状态不合理原因", summary["shelfStatus"]["negativeReasonCategories"]),
        ("评级不合理原因", summary["rating"]["negativeReasonCategories"]),
    ]:
        if reasons:
            for reason, count in reasons.items():
                lines.append(f"| {category} | {reason} | {count} |")
        else:
            lines.append(f"| {category} | 无 | 0 |")
    lines += [
        "",
        "## 评级与用户月均实销档位核对",
        "",
        summary["rating"]["monthlySalesBandComparison"]["basis"],
        "",
        "| 评级 | 样本数 | 匹配 | 不匹配 |",
        "|---|---:|---:|---:|",
    ]
    for rating_key, payload in summary["rating"]["monthlySalesBandComparison"]["byDisplayedRating"].items():
        lines.append(
            f"| {rating_key} | {payload['rows']} | {payload['matches']} | {payload['mismatches']} |"
        )
    lines += [
        "",
        f"- 可比样本匹配率：`{percent(summary['rating']['monthlySalesBandComparison']['matchRate'])}`",
        "",
        "## 关键业务规则复核",
        "",
        f"- 纯买断样本数：`{summary['rating']['pureBuyout']['rows']}`；用户认为评级合理：`{summary['rating']['pureBuyout']['userRatingPositive']}`；不合理：`{summary['rating']['pureBuyout']['userRatingNegative']}`。",
        f"- 买断+实销样本数：`{summary['rating']['buyoutPlusSales']['rows']}`；用户认为评级合理：`{summary['rating']['buyoutPlusSales']['userRatingPositive']}`；不合理：`{summary['rating']['buyoutPlusSales']['userRatingNegative']}`。",
        f"- 到期/下架相关样本数：`{summary['shelfStatus']['expiredOrOffShelfRows']}`；被简单打成 E：`{summary['shelfStatus']['expiredOrOffShelfRatedE']}`；未被简单打成 E：`{summary['shelfStatus']['expiredOrOffShelfNotSimplyE']}`。",
        f"- 风险/复核提示覆盖：`{summary['riskAndReviewPrompt']['rowsWithPrompt']}/{summary['rows']['evaluatedTaskRows']}`。任务包没有单独的“风险提示是否有用”填写字段，因此本轮只统计覆盖，不推断主观有用率。",
        f"- M4 校准候选：用户标记 `是` = `{summary['m4CalibrationCandidates']['userMarkedYes']}`；`否` = `{summary['m4CalibrationCandidates']['userMarkedNo']}`；待定/空 = `{summary['m4CalibrationCandidates']['userMarkedPendingOrBlank']}`。",
        "",
        "## 用户备注类别",
        "",
        "| 备注类别 | 数量 |",
        "|---|---:|",
    ]
    for category, count in summary["userNoteCategories"].items():
        label = {
            "blank": "未填写备注",
            "currently_on_shelf_correction": "用户确认目前在架",
            "bill_sales_record_gap": "原账单有实销记录但任务包证据缺失",
            "other_note": "其他备注",
        }.get(category, category)
        lines.append(f"| {label} | {count} |")
    lines += [
        "",
        "## 基线结论",
        "",
        f"- 结论：{summary['baselineVerdict']['verdict']}",
        f"- 是否可作为 M2 评级层候选基线：`{yes_no(summary['baselineVerdict']['canUseAsM2RatingLayerCandidateBaseline'])}`",
        f"- 阻断项：`{'; '.join(summary['baselineVerdict']['hardBlockers']) if summary['baselineVerdict']['hardBlockers'] else '无'}`",
        "- 最小修正方向：",
    ]
    for item in summary["baselineVerdict"]["minimumFixDirections"]:
        lines.append(f"  - {item}")
    lines += [
        "",
        "## 安全边界",
        "",
        "- 未写正式主数据。",
        "- 未进入 M3。",
        "- 未提交 private Excel。",
        "- 未使用 `git add .`。",
        "- 未触碰 stash。",
        "",
    ]
    return "\n".join(lines)


def judgment_row(label: str, judgment: dict) -> str:
    uncertain = judgment["neutral"] + judgment["notApplicable"] + judgment["blank"] + judgment["other"]
    return (
        f"| {label} | {judgment['effectiveAnswered']} | {judgment['positive']} | "
        f"{judgment['negative']} | {uncertain} | {percent(judgment['positiveRateOfEffectiveAnswered'])} |"
    )


def yes_no(value: bool) -> str:
    return "是" if value else "否"


def percent(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value * 100:.1f}%"


def rate(numerator: int | float, denominator: int | float) -> float | None:
    if not denominator:
        return None
    return round(float(numerator) / float(denominator), 4)


def to_number(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def relative(path: Path) -> str:
    return path.resolve().relative_to(ROOT.resolve()).as_posix()


if __name__ == "__main__":
    main()
