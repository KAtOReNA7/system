from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))

FIXTURE_SELF_TEST = "--fixture-self-test" in sys.argv
if not FIXTURE_SELF_TEST:
    from openpyxl import load_workbook  # noqa: E402

from m2_post_foundation_input_contract import (  # noqa: E402
    SCHEMA as FORMAL_INPUT_SCHEMA,
    copyright_value_type,
    validate_post_foundation_input_payload,
)


PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
WORKBOOK_PATH = (
    PRIVATE_DIR / "M2-post-foundation-review-buckets-user-confirmation-cn-v1.xlsx"
)
PACK_JSON_PATH = (
    PRIVATE_DIR / "M2-post-foundation-review-buckets-user-confirmation-cn-v1.json"
)
FOUNDATION_PATH = (
    PRIVATE_DIR / "M2-classification-tag-foundation-local-fixed-cn-v1.json"
)
CORE_STAGING_PATH = PRIVATE_DIR / "M2-five-source-local-staging-apply-result-cn-v1.json"
STATUS_STAGING_PATH = PRIVATE_DIR / "M2-status-local-staging-apply-result-cn-v1.json"
PRIVATE_DECISION_OUTPUT = (
    PRIVATE_DIR / "M2-post-foundation-review-decision-apply-private-v1.json"
)
FORMAL_INPUT_OUTPUT = PRIVATE_DIR / "M2-formal-basic-info-input-private-v1.json"
PUBLIC_JSON = (
    ROOT
    / "docs"
    / "analysis"
    / "m2-real-data"
    / "M2-post-foundation-review-decision-apply-summary-v1.json"
)
PUBLIC_MD = PUBLIC_JSON.with_suffix(".md")

MASTER_DIR = ROOT / "data" / "master-data"
DIGITAL_LEDGER = MASTER_DIR / "数字版权台账.xlsx"
ORIGINAL_LIBRARY = MASTER_DIR / "原创全库.xlsx"
ORIGINAL_LIBRARY_2 = MASTER_DIR / "原创全库2.xlsx"
AUTHORIZATION_SUMMARY = MASTER_DIR / "授权汇总台账.xlsx"
MAPPING_PAYLOAD = (
    ROOT
    / "data"
    / "m1-master-data-private"
    / "mapping-candidate"
    / "M1-formal-mapping-version-candidate-v0.1-detail-payload.json"
)
MAPPING_OVERLAY = (
    ROOT
    / "experiments"
    / "m1-mapping-version-import-candidate"
    / "G07-mapping-strategy-overlay-v0.2.json"
)

AS_OF_DATE = date(2026, 7, 13)
EXPECTED_EXPIRED = 146
EXPECTED_SPARSE = 92

EXPIRED_OPTIONS = {
    "采用系统候选",
    "确认已续约并填写新到期时间",
    "确认是结算或渠道滞后",
    "确认是权利期外收入需审计",
    "仍不确定",
    "其他（备注说明）",
}
SPARSE_OPTIONS = {
    "采用系统候选",
    "确认保持已上架",
    "确认改为已下架",
    "确认版权有效但暂停运营",
    "仍不确定",
    "其他（备注说明）",
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def canonical_work_id(value) -> str:
    text = clean(value)
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return str(int(match.group(0))) if match else text


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def normalize_date(value) -> str:
    if value is None or clean(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return (date(1899, 12, 30) + timedelta(days=int(round(float(value))))).isoformat()
    text = clean(value)
    if re.search(r"无限期|无期限|永久|长期有效|版权保护期满", text):
        return "无限期"
    matches = re.findall(
        r"(20\d{2}|19\d{2})[-/.年](\d{1,2})(?:[-/.月](\d{1,2}))?", text
    )
    if not matches:
        return ""
    year, month, day = matches[-1]
    return f"{int(year):04d}-{int(month):02d}-{int(day or 1):02d}"


def read_sheet_rows(path: Path, sheet_name: str, max_columns: int | None = None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True, max_col=max_columns)
        headers = [clean(value) for value in next(iterator)]
        for values in iterator:
            row = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            if any(clean(value) for value in row.values()):
                yield row
    finally:
        workbook.close()


def load_standard_mapping() -> dict[str, str]:
    mapping: dict[str, str] = {}
    if MAPPING_PAYLOAD.exists():
        for row in read_json(MAPPING_PAYLOAD).get("effective_mapping_snapshot", []):
            if (
                row.get("layer") == "historical_volume"
                and row.get("effective_status") == "effective_candidate"
            ):
                source = canonical_work_id(row.get("historical_standard_work_id"))
                target = canonical_work_id(row.get("target_standard_work_id"))
                if source and target:
                    mapping[source] = target
    if MAPPING_OVERLAY.exists():
        for row in read_json(MAPPING_OVERLAY).get("changes", []):
            if row.get("to") == "historical_volume_mapping":
                source = canonical_work_id(row.get("raw_work_id"))
                target = canonical_work_id(row.get("target_standard_work_id"))
                if source and target:
                    mapping[source] = target
    return mapping


def parse_filled_workbook() -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        result = []
        specifications = [
            (
                "01_到期仍有收入",
                EXPECTED_EXPIRED,
                EXPIRED_OPTIONS,
                "expired_with_revenue",
            ),
            (
                "02_版权有效收入稀疏",
                EXPECTED_SPARSE,
                SPARSE_OPTIONS,
                "active_rights_sparse_revenue",
            ),
        ]
        for sheet_name, expected_count, allowed_options, bucket in specifications:
            if sheet_name not in workbook.sheetnames:
                raise SystemExit(f"复核表缺少工作表：{sheet_name}")
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            headers = [clean(value) for value in rows[0]]
            index = {header: position for position, header in enumerate(headers)}
            required = {"作品编号", "书名", "系统候选", "你的确认", "你的备注"}
            if not required.issubset(index):
                raise SystemExit(f"{sheet_name} 表头与当前复核契约不一致。")
            parsed = []
            for values in rows[1:]:
                if not any(clean(value) for value in values):
                    continue
                decision = clean(values[index["你的确认"]])
                if decision not in allowed_options:
                    raise SystemExit(f"{sheet_name} 存在空白或无效的确认项。")
                note = clean(values[index["你的备注"]])
                new_expiry = ""
                if "新的版权到期时间（如续约）" in index:
                    new_expiry = normalize_date(
                        values[index["新的版权到期时间（如续约）"]]
                    )
                if decision == "确认已续约并填写新到期时间" and not new_expiry:
                    raise SystemExit("已续约记录必须填写可解析的新版权到期时间。")
                if decision == "其他（备注说明）" and not note:
                    raise SystemExit("选择“其他”时必须填写备注。")
                if decision == "确认是权利期外收入需审计" and not note:
                    raise SystemExit("权利期外收入审计确认必须填写归因备注。")
                parsed.append(
                    {
                        "bucket": bucket,
                        "workId": canonical_work_id(values[index["作品编号"]]),
                        "title": clean(values[index["书名"]]),
                        "currentWorkStatus": clean(
                            values[index.get("当前作品状态", -1)]
                        ),
                        "currentAudioRightsStatus": clean(
                            values[index.get("当前音频版权状态", -1)]
                        ),
                        "currentCopyrightEnd": clean(
                            values[index.get("版权到期时间", -1)]
                        ),
                        "systemCandidate": clean(values[index["系统候选"]]),
                        "userDecision": decision,
                        "newCopyrightEnd": new_expiry,
                        "userNote": note,
                    }
                )
            if len(parsed) != expected_count:
                raise SystemExit(
                    f"{sheet_name} 应有 {expected_count} 条，实际为 {len(parsed)} 条。"
                )
            result.append(parsed)
        return result[0], result[1]
    finally:
        workbook.close()


def verify_pack_identity(expired: list[dict], sparse: list[dict]) -> None:
    source = read_json(PACK_JSON_PATH)
    checks = [
        (expired, source.get("expiredWithRevenue", [])),
        (sparse, source.get("activeRightsSparseRevenue", [])),
    ]
    for workbook_rows, source_rows in checks:
        workbook_ids = [row["workId"] for row in workbook_rows]
        source_ids = [canonical_work_id(row.get("workId")) for row in source_rows]
        if len(workbook_ids) != len(set(workbook_ids)) or set(workbook_ids) != set(
            source_ids
        ):
            raise SystemExit("已填写工作簿与原始复核包的作品范围不一致。")
        source_titles = {
            canonical_work_id(row.get("workId")): clean(row.get("title"))
            for row in source_rows
        }
        if any(source_titles.get(row["workId"]) != row["title"] for row in workbook_rows):
            raise SystemExit("已填写工作簿的作品标题与原始复核包不一致。")


def load_base_records() -> tuple[dict[str, dict], dict[str, dict], dict[str, dict]]:
    foundation = {
        canonical_work_id(row.get("作品编号")): row
        for row in read_json(FOUNDATION_PATH).get("records", [])
    }
    core: dict[str, dict] = defaultdict(dict)
    for row in read_json(CORE_STAGING_PATH).get("records", []):
        work_id = canonical_work_id(row.get("作品编号"))
        field = clean(row.get("字段"))
        if work_id and field:
            core[work_id][field] = row
    status: dict[str, dict] = defaultdict(dict)
    for row in read_json(STATUS_STAGING_PATH).get("records", []):
        work_id = canonical_work_id(row.get("作品编号"))
        field = clean(row.get("字段"))
        if work_id and field:
            status[work_id][field] = row
    if len(foundation) != 3053 or len(core) != 3053 or len(status) != 3053:
        raise SystemExit("基础表、核心字段 staging 或状态 staging 的作品范围不完整。")
    if set(foundation) != set(core) or set(foundation) != set(status):
        raise SystemExit("基础表、核心字段 staging 与状态 staging 的作品范围不一致。")
    return foundation, dict(core), dict(status)


def collect_expiry_candidates(target_ids: set[str]) -> dict[str, list[dict]]:
    mapping = load_standard_mapping()
    found: dict[str, list[dict]] = {work_id: [] for work_id in target_ids}

    def target(value) -> str:
        work_id = canonical_work_id(value)
        return mapping.get(work_id, work_id)

    def scan(path, sheet, id_header, end_header, source_role, max_columns=None):
        if not path.exists():
            return
        for row in read_sheet_rows(path, sheet, max_columns):
            work_id = target(row.get(id_header))
            if work_id not in found:
                continue
            expiry = normalize_date(row.get(end_header))
            if expiry:
                candidate = {"sourceRole": source_role, "copyrightEnd": expiry}
                if candidate not in found[work_id]:
                    found[work_id].append(candidate)

    scan(DIGITAL_LEDGER, "Sheet1", "作品ID", "到期时间", "digital_ledger", 7)
    scan(
        ORIGINAL_LIBRARY,
        "全库排查",
        "作品ID",
        "结束时间",
        "original_library",
        11,
    )
    for sheet, end_header, role in [
        ("有声读物改编权", "结束时间", "original_library_2:audio_rights"),
        ("醉唐有声作品", "授权结束时间", "original_library_2:zuitang"),
        ("有声读物改编权的转授权", "结束时间", "original_library_2:sub_license"),
    ]:
        scan(ORIGINAL_LIBRARY_2, sheet, "作品ID", end_header, role, 9)
    scan(
        AUTHORIZATION_SUMMARY,
        "Sheet1",
        "作品ID",
        "我方版权到期时间",
        "authorization_summary",
        6,
    )
    for row in read_json(STATUS_STAGING_PATH).get("records", []):
        if clean(row.get("字段")) != "音频版权状态":
            continue
        work_id = target(row.get("作品编号"))
        if work_id not in found:
            continue
        status_value = clean(row.get("采用值"))
        evidence = " ".join(
            [clean(row.get("证据")), clean(row.get("处理说明"))]
        )
        if status_value == "无限期" or "无限期" in evidence:
            expiry = "无限期"
        elif status_value == "版权有效":
            dates = re.findall(r"\d{4}-\d{2}-\d{2}", evidence)
            expiry = max(dates, key=date.fromisoformat) if dates else ""
        else:
            expiry = ""
        if expiry:
            candidate = {
                "sourceRole": "status_staging_current_rights",
                "copyrightEnd": expiry,
            }
            if candidate not in found[work_id]:
                found[work_id].append(candidate)
    return found


def resolve_active_expiry(candidates: list[dict]) -> dict | None:
    future = [
        item
        for item in candidates
        if item["copyrightEnd"] == "无限期"
        or date.fromisoformat(item["copyrightEnd"]) >= AS_OF_DATE
    ]
    if not future:
        return None
    authorization = [
        item for item in future if item["sourceRole"] == "authorization_summary"
    ]
    if authorization:
        chosen = latest_copyright_end(authorization)
        agreeing_sources = sorted(
            {item["sourceRole"] for item in future if item["copyrightEnd"] == chosen}
        )
        return {
            "copyrightEnd": chosen,
            "sourceRoles": agreeing_sources,
            "resolution": (
                "two_source_exact_agreement"
                if len(agreeing_sources) >= 2
                else "current_authorization_summary_plus_user_confirmation"
            ),
        }
    values = {item["copyrightEnd"] for item in future}
    if len(values) == 1:
        chosen = next(iter(values))
        return {
            "copyrightEnd": chosen,
            "sourceRoles": sorted({item["sourceRole"] for item in future}),
            "resolution": "single_future_value_across_available_sources",
        }
    return None


def latest_copyright_end(candidates: list[dict]) -> str:
    values = {item["copyrightEnd"] for item in candidates}
    if "无限期" in values:
        return "无限期"
    return max(values, key=date.fromisoformat)


def resolve_exact_expiry(candidates: list[dict], audio_status: str) -> dict | None:
    if not candidates:
        return None
    if audio_status in {"版权有效", "无限期"}:
        active_resolution = resolve_active_expiry(candidates)
        if active_resolution:
            return active_resolution
    authorization = [
        item for item in candidates if item["sourceRole"] == "authorization_summary"
    ]
    if authorization:
        chosen = latest_copyright_end(authorization)
        return {
            "copyrightEnd": chosen,
            "sourceRoles": sorted(
                {
                    item["sourceRole"]
                    for item in candidates
                    if item["copyrightEnd"] == chosen
                }
            ),
            "resolution": "current_authorization_summary",
        }
    values = {item["copyrightEnd"] for item in candidates}
    if len(values) == 1:
        chosen = next(iter(values))
        return {
            "copyrightEnd": chosen,
            "sourceRoles": sorted({item["sourceRole"] for item in candidates}),
            "resolution": "single_exact_value_across_available_sources",
        }
    if audio_status in {"版权有效", "无限期"}:
        active_values = {
            item["copyrightEnd"]
            for item in candidates
            if item["copyrightEnd"] == "无限期"
            or date.fromisoformat(item["copyrightEnd"]) >= AS_OF_DATE
        }
        if len(active_values) == 1:
            chosen = next(iter(active_values))
            return {
                "copyrightEnd": chosen,
                "sourceRoles": sorted(
                    {
                        item["sourceRole"]
                        for item in candidates
                        if item["copyrightEnd"] == chosen
                    }
                ),
                "resolution": "status_consistent_exact_value",
            }
    return None


def needs_expiry_refresh(core_fields: dict, status_fields: dict) -> bool:
    start_value = clean(core_fields.get("版权开始", {}).get("采用值"))
    end_value = clean(core_fields.get("版权到期", {}).get("采用值"))
    end_type = copyright_value_type(end_value)
    audio_status = clean(status_fields.get("音频版权状态", {}).get("采用值"))
    if end_type == "invalid":
        return True
    if end_type == "exact_date":
        end_date = date.fromisoformat(end_value)
        if copyright_value_type(start_value) == "exact_date" and end_date < date.fromisoformat(start_value):
            return True
        if audio_status in {"版权有效", "无限期"} and end_date < AS_OF_DATE:
            return True
    if audio_status == "无限期" and end_value != "无限期":
        return True
    if audio_status == "版权有效" and end_type == "expired_unknown_date":
        return True
    return False


def apply_decisions(expired: list[dict], sparse: list[dict]) -> dict:
    foundation, core, status = load_base_records()
    expiry_refresh_ids = {
        work_id
        for work_id, fields in core.items()
        if needs_expiry_refresh(fields, status[work_id])
    }
    active_without_date_ids = {
        row["workId"]
        for row in expired
        if row["userNote"] == "版权未到期"
    }
    source_candidates = collect_expiry_candidates(
        active_without_date_ids | expiry_refresh_ids
    )
    source_resolutions = {
        work_id: resolve_active_expiry(source_candidates.get(work_id, []))
        for work_id in active_without_date_ids
    }

    updates: dict[str, dict] = defaultdict(dict)
    blockers: dict[str, list[str]] = defaultdict(list)
    advisories: dict[str, list[str]] = defaultdict(list)
    decisions = []
    generated_at = datetime.now(timezone.utc).isoformat()

    for work_id in expiry_refresh_ids:
        audio_status = clean(
            status[work_id].get("音频版权状态", {}).get("采用值")
        )
        resolution = resolve_exact_expiry(
            source_candidates.get(work_id, []), audio_status
        )
        if not resolution:
            continue
        updates[work_id]["版权到期"] = resolution["copyrightEnd"]
        updates[work_id]["版权到期解析"] = resolution
        if resolution["copyrightEnd"] == "无限期":
            updates[work_id]["音频版权状态"] = "无限期"
        elif date.fromisoformat(resolution["copyrightEnd"]) >= AS_OF_DATE:
            updates[work_id]["音频版权状态"] = "版权有效"
        else:
            updates[work_id]["音频版权状态"] = "版权已到期"

    def append_decision(row: dict, outcome: str, applied_fields: list[str]) -> None:
        work_id = row["workId"]
        decisions.append(
            {
                **row,
                "normalizedOutcome": outcome,
                "appliedFields": applied_fields,
                "formalInputBlockers": blockers[work_id],
                "advisories": advisories[work_id],
                "reviewStatus": "approved" if not blockers[work_id] else "pending",
                "audit": {
                    "eventType": "post_foundation_business_review_decision",
                    "actorType": "user_confirmed_private_pack",
                    "occurredAt": generated_at,
                    "sourceWorkbookSha256": sha256(WORKBOOK_PATH),
                },
            }
        )

    for row in expired:
        work_id = row["workId"]
        decision = row["userDecision"]
        note = row["userNote"]
        applied = []
        outcome = "confirmed_no_master_data_change"

        if "作品已下架" in note:
            updates[work_id]["作品状态"] = "已下架"
            applied.append("作品状态")

        if decision == "确认已续约并填写新到期时间":
            updates[work_id]["版权到期"] = row["newCopyrightEnd"]
            updates[work_id]["音频版权状态"] = (
                "无限期"
                if row["newCopyrightEnd"] == "无限期"
                else "版权有效"
            )
            applied.extend(["版权到期", "音频版权状态"])
            outcome = "renewal_date_applied"
        elif decision == "确认是权利期外收入需审计":
            advisories[work_id].append("到期后持续收听收入归因已确认，保留审计提示")
            outcome = "post_expiry_entitlement_revenue_confirmed_with_audit_advisory"
        elif decision == "确认是结算或渠道滞后":
            advisories[work_id].append("到期后收入已确认为结算或渠道滞后")
            outcome = "settlement_or_channel_lag_confirmed"
        elif decision == "仍不确定":
            blockers[work_id].append("到期仍有收入的业务归因未确认")
            outcome = "unresolved"
        elif decision == "其他（备注说明）":
            if "无限期" in note:
                updates[work_id]["版权到期"] = "无限期"
                updates[work_id]["音频版权状态"] = "无限期"
                applied.extend(["版权到期", "音频版权状态"])
                outcome = "perpetual_rights_applied"
            elif "续约中" in note and "没有确定" in note:
                advisories[work_id].append("续约进行中；新合同日期确定前保持当前已到期状态")
                outcome = "renewal_in_progress_current_expiry_retained"
            else:
                blockers[work_id].append("其他处理备注无法自动映射为受控状态")
                outcome = "unresolved_other"
        elif decision == "采用系统候选" and note == "版权未到期":
            resolution = source_resolutions.get(work_id)
            if resolution:
                updates[work_id]["版权到期"] = resolution["copyrightEnd"]
                updates[work_id]["音频版权状态"] = (
                    "无限期"
                    if resolution["copyrightEnd"] == "无限期"
                    else "版权有效"
                )
                updates[work_id]["版权到期解析"] = resolution
                applied.extend(["版权到期", "音频版权状态"])
                outcome = "active_expiry_resolved_from_current_sources"
            else:
                blockers[work_id].append("用户确认版权未到期，但现有来源无法唯一确定到期日")
                outcome = "active_expiry_unresolved"
        elif decision == "采用系统候选":
            outcome = "system_candidate_accepted"

        append_decision(row, outcome, list(dict.fromkeys(applied)))

    for row in sparse:
        work_id = row["workId"]
        decision = row["userDecision"]
        note = row["userNote"]
        applied = []
        outcome = "confirmed_no_master_data_change"
        if decision == "确认保持已上架":
            updates[work_id]["作品状态"] = "已上架"
            applied.append("作品状态")
            outcome = "on_shelf_confirmed"
        elif decision == "确认改为已下架":
            updates[work_id]["作品状态"] = "已下架"
            applied.append("作品状态")
            outcome = "off_shelf_confirmed"
        elif decision == "确认版权有效但暂停运营":
            updates[work_id]["作品状态"] = "已下架"
            applied.append("作品状态")
            advisories[work_id].append("版权有效但暂停运营")
            outcome = "rights_active_operation_paused"
        elif decision == "仍不确定":
            blockers[work_id].append("版权有效但收入稀疏的货架状态未确认")
            outcome = "unresolved"
        elif decision == "采用系统候选":
            candidate = row["systemCandidate"]
            if "保持已下架" in candidate:
                updates[work_id]["作品状态"] = "已下架"
                applied.append("作品状态")
            elif "保留已上架" in candidate or "保持已上架" in candidate:
                updates[work_id]["作品状态"] = "已上架"
                applied.append("作品状态")
            outcome = "system_candidate_accepted"
        else:
            blockers[work_id].append("其他处理备注无法自动映射为受控状态")
            outcome = "unresolved_other"
        if "套装书" in note and "第一册" in note:
            advisories[work_id].append("套装收入集中记在第一册，单册稀疏收入不代表下架")
        append_decision(row, outcome, applied)

    decision_by_work = {row["workId"]: row for row in decisions}
    unified_records = []
    field_updates = Counter()
    source_resolution_distribution = Counter()
    for work_id in sorted(
        foundation, key=lambda value: (0, int(value)) if value.isdigit() else (1, value)
    ):
        base = foundation[work_id]
        core_fields = core[work_id]
        status_fields = status[work_id]
        record = {
            "作品编号": work_id,
            "书名": clean(base.get("书名")),
            "作者": clean(base.get("作者")),
            "版权开始": clean(core_fields.get("版权开始", {}).get("采用值")),
            "版权到期": clean(core_fields.get("版权到期", {}).get("采用值")),
            "作品状态": clean(status_fields.get("作品状态", {}).get("采用值")),
            "音频版权状态": clean(
                status_fields.get("音频版权状态", {}).get("采用值")
            ),
            "一级分类": clean(base.get("一级分类")),
            "二级分类": clean(base.get("二级分类")),
            "三级分类": clean(base.get("三级分类")),
            "辅助标签": clean(base.get("辅助标签")) or "无",
            "字段来源": {
                "作者": "用户固定基础大表",
                "版权开始": clean(
                    core_fields.get("版权开始", {}).get("证据来源")
                ),
                "版权到期": clean(
                    core_fields.get("版权到期", {}).get("证据来源")
                ),
                "作品状态": clean(status_fields.get("作品状态", {}).get("证据")),
                "音频版权状态": clean(
                    status_fields.get("音频版权状态", {}).get("证据")
                ),
                "分类与标签": "用户固定基础大表",
            },
            "复核提示": list(advisories.get(work_id, [])),
            "正式输入阻断": list(blockers.get(work_id, [])),
            "复核决策状态": (
                decision_by_work.get(work_id, {}).get("reviewStatus", "not_required")
            ),
        }
        for field, value in updates.get(work_id, {}).items():
            if field == "版权到期解析":
                source_resolution_distribution[value["resolution"]] += 1
                record["字段来源"]["版权到期"] = "+".join(value["sourceRoles"])
                continue
            if record.get(field) != value:
                field_updates[field] += 1
            record[field] = value
            record["字段来源"][field] = "用户确认复核包或当前权威来源解析"
        record["版权到期类型"] = copyright_value_type(record["版权到期"])
        if record["版权到期类型"] == "invalid":
            record["正式输入阻断"].append("版权到期表述不在受控期限类型中")
        if (
            record["版权到期类型"] == "exact_date"
            and copyright_value_type(record["版权开始"]) == "exact_date"
            and date.fromisoformat(record["版权到期"])
            < date.fromisoformat(record["版权开始"])
        ):
            record["正式输入阻断"].append("版权到期早于版权开始")
        record["正式输入可用"] = not record["正式输入阻断"]
        unified_records.append(record)

    review_status_distribution = Counter(row["reviewStatus"] for row in decisions)
    decision_distribution = Counter(row["userDecision"] for row in decisions)
    outcome_distribution = Counter(row["normalizedOutcome"] for row in decisions)
    work_status_distribution = Counter(row["作品状态"] for row in unified_records)
    audio_status_distribution = Counter(
        row["音频版权状态"] for row in unified_records
    )
    copyright_term_type_distribution = Counter(
        row["版权到期类型"] for row in unified_records
    )
    advisory_distribution = Counter(
        advisory for row in unified_records for advisory in row["复核提示"]
    )
    blocker_distribution = Counter(
        blocker for row in unified_records for blocker in row["正式输入阻断"]
    )

    formal_input = {
        "schema": FORMAL_INPUT_SCHEMA,
        "status": "verified_complete" if not blocker_distribution else "blocked",
        "generatedAt": generated_at,
        "asOfDate": AS_OF_DATE.isoformat(),
        "artifactRole": "private_formal_basic_info_input_candidate",
        "privateOnly": True,
        "sourceManifest": [
            {"role": "confirmed_foundation", "sha256": sha256(FOUNDATION_PATH)},
            {"role": "core_staging", "sha256": sha256(CORE_STAGING_PATH)},
            {"role": "status_staging", "sha256": sha256(STATUS_STAGING_PATH)},
            {"role": "review_pack", "sha256": sha256(WORKBOOK_PATH)},
            {"role": "review_pack_manifest", "sha256": sha256(PACK_JSON_PATH)},
        ],
        "reviewDecisionSummary": {
            "total": len(decisions),
            "approved": review_status_distribution.get("approved", 0),
            "pending": review_status_distribution.get("pending", 0),
            "expiredWithRevenue": len(expired),
            "activeRightsSparseRevenue": len(sparse),
            "advisoryAssignmentCount": sum(advisory_distribution.values()),
            "auditEventCount": len(decisions),
        },
        "summary": {
            "workCount": len(unified_records),
            "fieldUpdates": dict(field_updates),
            "workStatusDistribution": dict(work_status_distribution),
            "audioRightsStatusDistribution": dict(audio_status_distribution),
            "copyrightTermTypeDistribution": dict(
                copyright_term_type_distribution
            ),
            "reviewDecisionDistribution": dict(decision_distribution),
            "normalizedOutcomeDistribution": dict(outcome_distribution),
            "advisoryDistribution": dict(advisory_distribution),
            "formalInputBlockerDistribution": dict(blocker_distribution),
            "sourceResolutionDistribution": dict(source_resolution_distribution),
        },
        "verification": {"verified": False, "issues": []},
        "records": unified_records,
        "formalAuthorization": {
            "grantedAt": "2026-07-13",
            "formalMasterDataWrite": True,
            "mappingActivation": True,
            "formalEvaluation": True,
            "formalTaskExportReleaseAudit": True,
            "m3FormalExecution": False,
        },
        "executionState": {
            "formalMasterDataWritten": False,
            "mappingActivated": False,
            "formalEvaluationExecuted": False,
            "formalReleaseCreated": False,
        },
    }
    formal_input["verification"]["verified"] = not blocker_distribution
    validation = validate_post_foundation_input_payload(
        formal_input, set(foundation), require_verified=not blocker_distribution
    )
    formal_input["verification"]["validation"] = validation
    formal_input["verification"]["issues"] = validation["issues"]
    if not validation["verified"]:
        formal_input["status"] = "blocked"

    return {
        "generatedAt": generated_at,
        "decisions": decisions,
        "formalInput": formal_input,
        "validation": validation,
        "summary": formal_input["summary"],
        "reviewDecisionSummary": formal_input["reviewDecisionSummary"],
    }


def build_public_summary(result: dict, mode: str) -> dict:
    summary = result["summary"]
    review = result["reviewDecisionSummary"]
    return {
        "schema": "m2.post_foundation_review_decision_apply_summary.v1",
        "generatedAt": result["generatedAt"],
        "mode": mode,
        "input": {
            "workbookVersion": "post_foundation_user_confirmation_cn_v1",
            "expiredWithRevenue": review["expiredWithRevenue"],
            "activeRightsSparseRevenue": review["activeRightsSparseRevenue"],
            "total": review["total"],
        },
        "result": {
            "approved": review["approved"],
            "pending": review["pending"],
            "auditEventCount": review["auditEventCount"],
            "advisoryAssignmentCount": review["advisoryAssignmentCount"],
            "fieldUpdates": summary["fieldUpdates"],
            "workStatusDistribution": summary["workStatusDistribution"],
            "audioRightsStatusDistribution": summary[
                "audioRightsStatusDistribution"
            ],
            "copyrightTermTypeDistribution": summary[
                "copyrightTermTypeDistribution"
            ],
            "normalizedOutcomeDistribution": summary[
                "normalizedOutcomeDistribution"
            ],
            "sourceResolutionDistribution": summary[
                "sourceResolutionDistribution"
            ],
            "formalInputBlockerDistribution": summary[
                "formalInputBlockerDistribution"
            ],
            "formalInputContractVerified": result["validation"]["verified"],
        },
        "privacy": {
            "containsRealTitles": False,
            "containsAuthors": False,
            "containsChannels": False,
            "containsPerWorkRevenue": False,
            "privateArtifactsCommitted": False,
        },
        "boundary": {
            "formalMasterDataWritten": False,
            "mappingActivated": False,
            "formalEvaluationExecuted": False,
            "m3Entered": False,
        },
    }


def public_markdown(summary: dict) -> str:
    result = summary["result"]
    return f"""# M2 post-foundation 两类复核决策应用摘要 v1

## 结论

- 用户填写范围：`{summary['input']['total']}` 条，其中到期仍有收入 `{summary['input']['expiredWithRevenue']}` 条、版权有效但收入稀疏 `{summary['input']['activeRightsSparseRevenue']}` 条。
- 已通过校验并形成最终复核状态：`{result['approved']}` 条；仍待确认：`{result['pending']}` 条。
- private 逐作品正式输入候选内容契约：`{'通过' if result['formalInputContractVerified'] else '未通过'}`。
- 本摘要只包含聚合信息；逐作品书名、作者、确认备注、日期和值留在 Git 忽略的 private 输出中。

## 本地 staging 变化

- 字段更新分布：`{json.dumps(result['fieldUpdates'], ensure_ascii=False)}`。
- 作品状态分布：`{json.dumps(result['workStatusDistribution'], ensure_ascii=False)}`。
- 音频版权状态分布：`{json.dumps(result['audioRightsStatusDistribution'], ensure_ascii=False)}`。
- 版权期限类型分布：`{json.dumps(result['copyrightTermTypeDistribution'], ensure_ascii=False)}`。
- 复核审计事件：`{result['auditEventCount']}` 条；复核提示赋值：`{result['advisoryAssignmentCount']}` 条。
- 当前来源自动消歧：`{json.dumps(result['sourceResolutionDistribution'], ensure_ascii=False)}`。
- 正式输入阻断：`{json.dumps(result['formalInputBlockerDistribution'], ensure_ascii=False)}`。

## 边界

- 本步骤只应用 private 文件级确认并生成可验证正式输入候选；尚未写正式主数据、激活 mapping、执行 formal evaluation 或创建正式 release。
- M2 不输出自动运营建议；复核提示仅记录事实归因、审计或后续证据要求。
- M3 formal execution 仍未开始。
"""


def fixture_self_test() -> dict:
    candidates = [
        {"sourceRole": "authorization_summary", "copyrightEnd": "2040-12-31"},
        {"sourceRole": "original_library", "copyrightEnd": "2025-12-31"},
    ]
    resolution = resolve_active_expiry(candidates)
    latest_resolution = resolve_active_expiry(
        [
            {"sourceRole": "authorization_summary", "copyrightEnd": "2021-06-23"},
            {"sourceRole": "authorization_summary", "copyrightEnd": "2026-07-30"},
        ]
    )
    return {
        "fixtureSelfTest": True,
        "currentAuthorizationWinsStaleConflict": resolution
        == {
            "copyrightEnd": "2040-12-31",
            "sourceRoles": ["authorization_summary"],
            "resolution": "current_authorization_summary_plus_user_confirmation",
        },
        "unresolvableEmptyCandidatesBlocked": resolve_active_expiry([]) is None,
        "latestCurrentAuthorizationWinsOlderAuthorization": latest_resolution
        == {
            "copyrightEnd": "2026-07-30",
            "sourceRoles": ["authorization_summary"],
            "resolution": "current_authorization_summary_plus_user_confirmation",
        },
        "noDatabaseWrite": True,
        "noMappingActivation": True,
        "noFormalEvaluation": True,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--fixture-self-test", action="store_true")
    args = parser.parse_args()
    if args.fixture_self_test:
        print(json.dumps(fixture_self_test(), ensure_ascii=False))
        return
    if args.dry_run and args.apply:
        raise SystemExit("--dry-run 与 --apply 不能同时使用。")
    mode = "apply" if args.apply else "dry_run"
    required = [
        WORKBOOK_PATH,
        PACK_JSON_PATH,
        FOUNDATION_PATH,
        CORE_STAGING_PATH,
        STATUS_STAGING_PATH,
        DIGITAL_LEDGER,
        ORIGINAL_LIBRARY,
        ORIGINAL_LIBRARY_2,
        AUTHORIZATION_SUMMARY,
    ]
    missing = [path.name for path in required if not path.exists()]
    if missing:
        raise SystemExit("缺少当前复核应用输入：" + "、".join(missing))

    expired, sparse = parse_filled_workbook()
    verify_pack_identity(expired, sparse)
    result = apply_decisions(expired, sparse)
    public = build_public_summary(result, mode)

    if args.apply:
        if not result["validation"]["verified"]:
            raise SystemExit(
                "逐作品正式输入内容契约未通过，已停止写入应用结果："
                + ", ".join(result["validation"]["issues"])
            )
        PRIVATE_DECISION_OUTPUT.write_text(
            json.dumps(
                {
                    "schema": "m2.post_foundation_review_decision_apply_private.v1",
                    "generatedAt": result["generatedAt"],
                    "privateOnly": True,
                    "summary": result["reviewDecisionSummary"],
                    "decisions": result["decisions"],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        FORMAL_INPUT_OUTPUT.write_text(
            json.dumps(result["formalInput"], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        PUBLIC_JSON.write_text(
            json.dumps(public, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        PUBLIC_MD.write_text(public_markdown(public), encoding="utf-8")

    print(
        json.dumps(
            {
                "mode": mode,
                "reviewDecisionSummary": result["reviewDecisionSummary"],
                "fieldUpdates": result["summary"]["fieldUpdates"],
                "workStatusDistribution": result["summary"][
                    "workStatusDistribution"
                ],
                "audioRightsStatusDistribution": result["summary"][
                    "audioRightsStatusDistribution"
                ],
                "formalInputContractVerified": result["validation"]["verified"],
                "formalInputBlockerCount": result["validation"][
                    "formalInputBlockerCount"
                ],
                "formalInputContractIssues": result["validation"]["issues"],
                "privateDecisionOutputWritten": bool(args.apply),
                "formalInputOutputWritten": bool(args.apply),
                "databaseWritten": False,
                "mappingActivated": False,
                "formalEvaluationExecuted": False,
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
