from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it into a local temp dependency path, "
        "for example: python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


PRIVATE_OPERATOR_XLSX = (
    ROOT
    / "data"
    / "private-output"
    / "m2-business-review"
    / "m2-v1.1-30-work-operator-task-pack-cn-after-dual-source-staging-v2.xlsx"
)
OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
REPORT_JSON = OUTPUT_DIR / "M2-v1.1-after-staging-operator-validation-summary-v1.json"
REPORT_MD = OUTPUT_DIR / "M2-v1.1-after-staging-operator-validation-summary-v1.md"

TASK_SHEET = "01_运营任务卡"
USER_RESERVED_SOURCE = "用户指定作品"

CORE_FIELDS = {
    "forecastTrust": "运营判断：预测是否可信",
    "ratingReasonable": "运营判断：评级是否合理",
    "suggestionExecutable": "运营判断：建议是否可执行",
    "issueType": "运营发现的问题类型",
    "suggestionCorrection": "运营建议修正",
    "m4CalibrationCandidate": "是否应进入M4校准案例池",
}

SUPPORT_FIELDS = {
    "standardWorkId": "standard_work_id",
    "sampleSource": "样本来源",
    "forecastOutputType": "辅助原始forecastOutputType",
    "lifecycle": "辅助原始生命周期code",
    "revenueBucket": "辅助原始收入层级code",
    "forecastabilityStatus": "辅助原始预测状态code",
    "businessActionStatus": "辅助原始业务动作状态code",
}

ISSUE_SPLIT_RE = re.compile(r"[，,、;；/]+")


def main() -> None:
    payload = build_payload()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(REPORT_JSON, payload)
    REPORT_MD.write_text(build_markdown(payload), encoding="utf-8")
    print(
        json.dumps(
            {
                "reviewableRows": payload["summary"]["reviewableRows"],
                "coreFeedbackCompletionRate": payload["summary"]["coreFeedbackCompletionRate"],
                "forecastTrustPositiveRate": payload["operatorFeedback"]["forecastTrust"]["positiveRate"],
                "ratingReasonablePositiveRate": payload["operatorFeedback"]["ratingReasonable"]["positiveRate"],
                "suggestionExecutablePositiveRate": payload["operatorFeedback"]["suggestionExecutable"]["positiveRate"],
                "m4ConfirmedCandidateCount": payload["m4Calibration"]["confirmedCandidateCount"],
                "v1_1ConditionalAcceptance": payload["v1_1ConditionalAcceptance"]["overallDecision"],
                "report": str(REPORT_MD.relative_to(ROOT)),
            },
            ensure_ascii=False,
        )
    )


def build_payload() -> dict:
    rows = load_task_rows()
    grouped = group_by_standard_work_id(rows)
    reviewable_rows = [row for row in rows if clean(row.get(SUPPORT_FIELDS["sampleSource"])) != USER_RESERVED_SOURCE]
    user_reserved_rows = [row for row in rows if clean(row.get(SUPPORT_FIELDS["sampleSource"])) == USER_RESERVED_SOURCE]
    rows_with_core_feedback = [
        row
        for row in reviewable_rows
        if all(clean(row.get(CORE_FIELDS[key])) for key in ["forecastTrust", "ratingReasonable", "suggestionExecutable"])
    ]

    operator_feedback = {
        "forecastTrust": feedback_summary(reviewable_rows, CORE_FIELDS["forecastTrust"], positive_values={"可信", "基本可信"}),
        "ratingReasonable": feedback_summary(reviewable_rows, CORE_FIELDS["ratingReasonable"], positive_values={"合理", "基本合理"}),
        "suggestionExecutable": feedback_summary(
            reviewable_rows,
            CORE_FIELDS["suggestionExecutable"],
            positive_values={"可执行", "需要人工确认", "仅供参考", "不适用"},
        ),
        "issueType": {
            "exactDistribution": distribution(reviewable_rows, CORE_FIELDS["issueType"]),
            "splitDistribution": split_issue_distribution(reviewable_rows),
        },
        "suggestionCorrection": {
            "filledCount": sum(1 for row in reviewable_rows if clean(row.get(CORE_FIELDS["suggestionCorrection"]))),
            "blankCount": sum(1 for row in reviewable_rows if not clean(row.get(CORE_FIELDS["suggestionCorrection"]))),
        },
    }
    m4 = m4_summary(rows, reviewable_rows, user_reserved_rows)
    acceptance = acceptance_decision(reviewable_rows, operator_feedback, m4)
    return {
        "schema": "m2.v1_1.after_staging_operator_validation_summary.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "currentHead": git(["rev-parse", "HEAD"]),
        "originMain": git(["rev-parse", "origin/main"]),
        "input": {
            "privateWorkbookPath": str(PRIVATE_OPERATOR_XLSX.relative_to(ROOT)),
            "versionConstraint": "after-dual-source-staging-v2 only",
            "oldAnonymousWorkbookRead": False,
            "groupedByStandardWorkId": True,
        },
        "summary": {
            "taskRows": len(rows),
            "reviewableRows": len(reviewable_rows),
            "userReservedRows": len(user_reserved_rows),
            "uniqueStandardWorkIdCount": len(grouped),
            "duplicateStandardWorkIdCount": sum(1 for items in grouped.values() if len(items) > 1),
            "rowsWithCoreFeedback": len(rows_with_core_feedback),
            "coreFeedbackCompletionRate": ratio(len(rows_with_core_feedback), len(reviewable_rows)),
            "overallCoreFeedbackCompletionRate": ratio(len(rows_with_core_feedback), len(rows)),
            "sampleSourceDistribution": distribution(rows, SUPPORT_FIELDS["sampleSource"]),
            "forecastOutputTypeDistribution": distribution(reviewable_rows, SUPPORT_FIELDS["forecastOutputType"]),
            "forecastabilityStatusDistribution": distribution(reviewable_rows, SUPPORT_FIELDS["forecastabilityStatus"]),
            "businessActionStatusDistribution": distribution(reviewable_rows, SUPPORT_FIELDS["businessActionStatus"]),
            "revenueBucketDistribution": distribution(reviewable_rows, SUPPORT_FIELDS["revenueBucket"]),
            "lifecycleDistribution": distribution(reviewable_rows, SUPPORT_FIELDS["lifecycle"]),
        },
        "operatorFeedback": operator_feedback,
        "m4Calibration": m4,
        "v1_1ConditionalAcceptance": acceptance,
        "safeOutputBoundary": {
            "sanitizedAggregateOnly": True,
            "realWorkNamesIncluded": False,
            "authorNamesIncluded": False,
            "standardWorkIdDetailsIncluded": False,
            "channelNamesIncluded": False,
            "rawLedgerRowsIncluded": False,
            "privateWorkbookCommitted": False,
            "formalMasterDataWritten": False,
            "databaseConnected": False,
            "databaseWritten": False,
            "m3Entered": False,
        },
    }


def load_task_rows() -> list[dict]:
    if not PRIVATE_OPERATOR_XLSX.exists():
        raise SystemExit(f"Missing private workbook: {PRIVATE_OPERATOR_XLSX.relative_to(ROOT)}")
    workbook = load_workbook(PRIVATE_OPERATOR_XLSX, read_only=True, data_only=True)
    sheet = workbook[TASK_SHEET] if TASK_SHEET in workbook.sheetnames else workbook.worksheets[1]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = []
    for values in iterator:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(clean(value) for value in row.values()):
            rows.append(row)
    required = set(CORE_FIELDS.values()) | set(SUPPORT_FIELDS.values())
    missing = sorted(required.difference(headers))
    if missing:
        raise SystemExit("Missing required workbook columns: " + ", ".join(missing))
    return rows


def group_by_standard_work_id(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        standard_work_id = clean(row.get(SUPPORT_FIELDS["standardWorkId"]))
        if standard_work_id:
            grouped[standard_work_id].append(row)
    return grouped


def feedback_summary(rows: list[dict], field: str, positive_values: set[str]) -> dict:
    counts = Counter(clean(row.get(field)) or "未填写" for row in rows)
    positive_count = sum(counts.get(value, 0) for value in positive_values)
    negative_count = sum(counts.values()) - positive_count - counts.get("未填写", 0)
    return {
        "distribution": dict(sorted(counts.items())),
        "positiveCount": positive_count,
        "negativeOrUncertainCount": negative_count,
        "blankCount": counts.get("未填写", 0),
        "positiveRate": ratio(positive_count, len(rows)),
    }


def distribution(rows: list[dict], field: str) -> dict:
    return dict(sorted(Counter(clean(row.get(field)) or "未填写" for row in rows).items()))


def split_issue_distribution(rows: list[dict]) -> dict:
    counts: Counter[str] = Counter()
    for row in rows:
        text = clean(row.get(CORE_FIELDS["issueType"]))
        if not text:
            counts["未填写"] += 1
            continue
        parts = [part.strip() for part in ISSUE_SPLIT_RE.split(text) if part.strip()]
        for part in parts or [text]:
            counts[part] += 1
    return dict(sorted(counts.items()))


def m4_summary(rows: list[dict], reviewable_rows: list[dict], user_reserved_rows: list[dict]) -> dict:
    counts_all = Counter(clean(row.get(CORE_FIELDS["m4CalibrationCandidate"])) or "未填写" for row in rows)
    counts_reviewable = Counter(clean(row.get(CORE_FIELDS["m4CalibrationCandidate"])) or "未填写" for row in reviewable_rows)
    issue_rows = [
        row
        for row in reviewable_rows
        if clean(row.get(CORE_FIELDS["issueType"])) and clean(row.get(CORE_FIELDS["issueType"])) != "无明显问题"
    ]
    return {
        "distributionAllRows": dict(sorted(counts_all.items())),
        "distributionReviewableRows": dict(sorted(counts_reviewable.items())),
        "confirmedCandidateCount": counts_reviewable.get("是", 0),
        "pendingCandidateCount": counts_reviewable.get("待定", 0),
        "rejectedCandidateCount": counts_reviewable.get("否", 0),
        "operatorIssueRowsThatMayNeedCalibration": len(issue_rows),
        "userReservedRowsWithoutCoreFeedback": sum(
            1
            for row in user_reserved_rows
            if not any(clean(row.get(CORE_FIELDS[key])) for key in ["forecastTrust", "ratingReasonable", "suggestionExecutable"])
        ),
        "m4Boundary": "candidate sedimentation only; no self-learning and no M3 entry in this run",
    }


def acceptance_decision(rows: list[dict], feedback: dict, m4: dict) -> dict:
    forecast_negative = feedback["forecastTrust"]["distribution"].get("不可信", 0)
    rating_unreasonable = feedback["ratingReasonable"]["distribution"].get("不合理", 0)
    suggestion_not_executable = feedback["suggestionExecutable"]["distribution"].get("不可执行", 0)
    common_sense_conflict = feedback["issueType"]["splitDistribution"].get("业务常识冲突", 0)
    forecast_issue = feedback["issueType"]["splitDistribution"].get("预测不合理", 0)
    severe_business_violation_limit = 2

    blockers = []
    warnings = []
    if common_sense_conflict > severe_business_violation_limit:
        blockers.append(
            f"业务常识冲突 {common_sense_conflict} 条，超过通过阈值 {severe_business_violation_limit} 条"
        )
    if rating_unreasonable:
        blockers.append(f"评级被运营标记为不合理 {rating_unreasonable} 条")
    if suggestion_not_executable:
        blockers.append(f"运营建议被标记为不可执行 {suggestion_not_executable} 条")
    if forecast_negative:
        warnings.append(f"预测被标记为不可信 {forecast_negative} 条")
    if forecast_issue:
        warnings.append(f"问题类型包含预测不合理 {forecast_issue} 条")
    if m4["pendingCandidateCount"]:
        warnings.append(f"M4 校准候选仍为待定 {m4['pendingCandidateCount']} 条")

    forecast_layer_decision = "usable_with_exception" if forecast_negative <= 1 else "needs_rework"
    rating_layer_decision = "needs_rework" if rating_unreasonable else "accepted"
    suggestion_layer_decision = "needs_rework" if suggestion_not_executable else "accepted"
    overall = "not_accepted_for_limited_m2_business_review_baseline" if blockers else "accepted_as_limited_m2_business_review_baseline"
    return {
        "candidateVersion": "m2-realdata-dev-disentangled-forecast-v1.1-conditional",
        "overallDecision": overall,
        "forecastLayerDecision": forecast_layer_decision,
        "ratingLayerDecision": rating_layer_decision,
        "suggestionLayerDecision": suggestion_layer_decision,
        "canBeAcceptedAsLimitedM2BusinessReviewBaseline": not blockers,
        "blockers": blockers,
        "warnings": warnings,
        "m3Allowed": False,
        "formalReleaseApproved": False,
    }


def build_markdown(payload: dict) -> str:
    summary = payload["summary"]
    feedback = payload["operatorFeedback"]
    m4 = payload["m4Calibration"]
    decision = payload["v1_1ConditionalAcceptance"]
    lines = [
        "# M2 v1.1 after-staging operator validation summary v1",
        "",
        "## Conclusion",
        f"- Input workbook: `{payload['input']['privateWorkbookPath']}`",
        f"- Reviewable rows: `{summary['reviewableRows']}`",
        f"- Core feedback completion rate: `{summary['coreFeedbackCompletionRate']}`",
        f"- v1.1 conditional accepted as limited M2 business review baseline: `{decision['canBeAcceptedAsLimitedM2BusinessReviewBaseline']}`",
        f"- Overall decision: `{decision['overallDecision']}`",
        f"- M3 entered: `{decision['m3Allowed']}`",
        "",
        "## Feedback Distributions",
        markdown_table(
            [
                {"metric": "forecast trust", **feedback["forecastTrust"]},
                {"metric": "rating reasonableness", **feedback["ratingReasonable"]},
                {"metric": "suggestion executability", **feedback["suggestionExecutable"]},
            ],
            ["metric", "distribution", "positiveCount", "negativeOrUncertainCount", "blankCount", "positiveRate"],
        ),
        "",
        "## Issue Types",
        f"- Exact distribution: `{json.dumps(feedback['issueType']['exactDistribution'], ensure_ascii=False)}`",
        f"- Split distribution: `{json.dumps(feedback['issueType']['splitDistribution'], ensure_ascii=False)}`",
        "",
        "## M4 Calibration Candidates",
        f"- Reviewable distribution: `{json.dumps(m4['distributionReviewableRows'], ensure_ascii=False)}`",
        f"- Confirmed candidates: `{m4['confirmedCandidateCount']}`",
        f"- Pending candidates: `{m4['pendingCandidateCount']}`",
        f"- Operator issue rows that may need calibration: `{m4['operatorIssueRowsThatMayNeedCalibration']}`",
        "- Boundary: candidate sedimentation only; no self-learning and no M3 entry.",
        "",
        "## Acceptance Blockers",
        *[f"- {item}" for item in decision["blockers"]],
        "",
        "## Warnings",
        *[f"- {item}" for item in decision["warnings"]],
        "",
        "## Sanitized Scope",
        "- This report is aggregate-only.",
        "- It does not include real work names, author names, standard_work_id details, channel names, or raw ledger rows.",
        "- It does not write formal master data, connect to a database, or enter M3.",
    ]
    return "\n".join(lines) + "\n"


def markdown_table(rows: list[dict], columns: list[str]) -> str:
    header = "| " + " | ".join(columns) + " |"
    divider = "| " + " | ".join("---" for _ in columns) + " |"
    body = []
    for row in rows:
        values = []
        for column in columns:
            value = row.get(column, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            values.append(str(value).replace("|", "/"))
        body.append("| " + " | ".join(values) + " |")
    return "\n".join([header, divider, *body])


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def ratio(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round(numerator / denominator, 4)


def git(args: list[str]) -> str:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True, encoding="utf-8").strip()
    except Exception:
        return ""


if __name__ == "__main__":
    main()
