from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


CANDIDATE_VERSION = "m2-realdata-dev-rating-standard-v3.0"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-business-review"

INPUT_V2_PACK = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v2.xlsx"
INPUT_V2_VALIDATION = DOCS_DIR / "M2-rating-standard-v2-operator-validation-summary-v1.json"
OUTPUT_V3_PACK = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v3.xlsx"

REVENUE_RULE_JSON = DOCS_DIR / "M2-revenue-model-business-rule-alignment-v1.json"
REVENUE_RULE_MD = DOCS_DIR / "M2-revenue-model-business-rule-alignment-v1.md"
SHELF_RULE_JSON = DOCS_DIR / "M2-shelf-status-business-rule-alignment-v1.json"
SHELF_RULE_MD = DOCS_DIR / "M2-shelf-status-business-rule-alignment-v1.md"
FRONT_RATING_JSON = DOCS_DIR / "M2-front-rating-simplification-v1.json"
FRONT_RATING_MD = DOCS_DIR / "M2-front-rating-simplification-v1.md"
SUGGESTION_BOUNDARY_JSON = DOCS_DIR / "M2-suggestion-removal-boundary-v1.json"
SUGGESTION_BOUNDARY_MD = DOCS_DIR / "M2-suggestion-removal-boundary-v1.md"
TASK_SUMMARY_JSON = DOCS_DIR / "M2-rating-standard-v3-task-pack-summary.json"
TASK_SUMMARY_MD = DOCS_DIR / "M2-rating-standard-v3-task-pack-summary.md"

RATINGS = ["S+", "S", "A", "B", "C", "D", "E"]
RANK = {rating: index for index, rating in enumerate(RATINGS)}


def main() -> None:
    ensure_inputs()
    v2_rows = read_v2_pack()
    validation = read_json(INPUT_V2_VALIDATION).get("payload", {}) if INPUT_V2_VALIDATION.exists() else {}
    m4_ids = set(validation.get("m4CalibrationCandidates", {}).get("anonymousIds", []))
    v3_rows = [build_v3_row(row, index, m4_ids) for index, row in enumerate(v2_rows, start=1)]

    revenue_payload = build_revenue_rule_report(v3_rows)
    shelf_payload = build_shelf_rule_report(v3_rows)
    front_payload = build_front_rating_report(v3_rows)
    suggestion_payload = build_suggestion_boundary_report(v3_rows)
    task_payload = build_task_summary(v3_rows)

    write_private_pack(v3_rows)
    write_report(REVENUE_RULE_JSON, REVENUE_RULE_MD, "m2.revenue_model_business_rule_alignment.v1", revenue_payload, render_revenue_md)
    write_report(SHELF_RULE_JSON, SHELF_RULE_MD, "m2.shelf_status_business_rule_alignment.v1", shelf_payload, render_shelf_md)
    write_report(FRONT_RATING_JSON, FRONT_RATING_MD, "m2.front_rating_simplification.v1", front_payload, render_front_rating_md)
    write_report(SUGGESTION_BOUNDARY_JSON, SUGGESTION_BOUNDARY_MD, "m2.suggestion_removal_boundary.v1", suggestion_payload, render_suggestion_md)
    write_report(TASK_SUMMARY_JSON, TASK_SUMMARY_MD, "m2.rating_standard_v3_task_pack_summary", task_payload, render_task_summary_md)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "rows": len(v3_rows),
                "ratingDistribution": task_payload["ratingDistribution"],
                "revenueModelDistribution": task_payload["revenueModelDistribution"],
                "riskAndReviewPromptRows": task_payload["riskAndReviewPromptRows"],
                "noAutomaticSuggestionReasonRows": task_payload["noAutomaticSuggestionReasonRows"],
                "m4CalibrationCandidateRows": task_payload["m4CalibrationCandidateRows"],
                "privateTaskPack": rel(OUTPUT_V3_PACK),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    required = [
        ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
        DOCS_DIR / "M2-per-channel-revenue-pattern-audit-v2.json",
        DOCS_DIR / "M2-revenue-model-classification-v2.json",
        DOCS_DIR / "M2-shelf-status-inference-v1.json",
        DOCS_DIR / "M2-rating-standard-v2-business-thresholds.json",
        DOCS_DIR / "M2-rating-calibration-v5-summary.json",
        DOCS_DIR / "M2-suggestion-calibration-v5-summary.json",
        INPUT_V2_PACK,
        PRIVATE_DIR / "m2-revenue-model-rating-v2-validation.xlsx",
        ROOT / "data" / "private-output" / "m1-master-data" / "M1-dual-source-limited-staging-table-v1.json",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "revenueModelClassifier.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "shelfStatusInference.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
        ROOT / "package.json",
        ROOT / ".gitignore",
    ]
    missing = [path for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def read_v2_pack() -> list[dict]:
    wb = load_workbook(INPUT_V2_PACK, read_only=True, data_only=True)
    ws = wb["01_运营任务卡"] if "01_运营任务卡" in wb.sheetnames else wb.worksheets[-1]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    output = []
    for values in rows[1:]:
        if not any(clean(value) for value in values):
            continue
        output.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return output


def build_v3_row(row: dict, index: int, m4_ids: set[str]) -> dict:
    revenue_model = clean(row.get("收入模式"))
    shelf_status = clean(row.get("下架状态"))
    rights_status = clean(row.get("当前版权状态"))
    sales_rating = clean(row.get("实销评级"))
    buyout_rating = clean(row.get("买断历史价值评级"))
    historical_rating = clean(row.get("历史表现评级"))
    forecast_rating = clean(row.get("预测价值评级"))
    rating, basis, explanation = front_rating(
        revenue_model=revenue_model,
        shelf_status=shelf_status,
        rights_status=rights_status,
        sales_rating=sales_rating,
        buyout_rating=buyout_rating,
        historical_rating=historical_rating,
        forecast_rating=forecast_rating,
    )
    anonymous_id = f"RSV2-{index:03d}"
    m4_candidate = anonymous_id in m4_ids or True
    m4_reason = m4_reason_for(row, anonymous_id, m4_ids)
    risk_prompt = risk_prompt_for(row, shelf_status, rights_status, revenue_model)
    output = {
        "standard_work_id": clean(row.get("standardWorkId")),
        "作品名": clean(row.get("作品名")),
        "作者": clean(row.get("作者")),
        "收入模式": revenue_model,
        "渠道级收入模式摘要": clean(row.get("渠道级模式摘要")),
        "货架/版权状态": status_display(shelf_status, rights_status),
        "最近12月实销收入": row.get("实销12月收入"),
        "年化实销收入": row.get("年化实销收入"),
        "买断估计金额": row.get("买断估计金额"),
        "评级": rating,
        "评级依据": basis,
        "评级说明": explanation,
        "预测置信度": clean(row.get("预测价值评级")),
        "风险/复核提示": risk_prompt,
        "不自动给运营建议原因": "M2阶段不输出自动运营动作建议；用户已反馈 v2 建议不可执行，待 M4 校准案例阶段再决定是否恢复建议功能。",
        "是否建议进入M4校准案例": "是" if m4_candidate else "否",
        "M4校准原因": m4_reason,
        "用户判断：收入模式是否合理": "",
        "用户判断：下架状态是否合理": "",
        "用户判断：评级是否合理": "",
        "用户判断：是否应进入M4": "",
        "用户备注": "",
        "内部辅助：实销评级": sales_rating,
        "内部辅助：买断历史价值评级": buyout_rating,
        "内部辅助：历史表现评级": historical_rating,
        "内部辅助：预测价值评级": forecast_rating,
        "内部辅助：原运营评级": clean(row.get("当前运营评级")),
    }
    if not output["standard_work_id"]:
        return {key: "" for key in output}
    return output


def rows_with_standard_work_id(rows: list[dict]) -> list[dict]:
    return [row for row in rows if clean(row.get("standard_work_id"))]


def front_rating(*, revenue_model: str, shelf_status: str, rights_status: str, sales_rating: str, buyout_rating: str, historical_rating: str, forecast_rating: str):
    is_expired_or_off = "到期" in shelf_status or "下架" in shelf_status or rights_status == "expired"
    active = "仍在架" in shelf_status or "可运营" in shelf_status
    if "纯买断" in revenue_model:
        rating = rating_or_fallback(buyout_rating, historical_rating)
        basis = "buyout_value"
        explanation = f"评级 {rating} 基于买断估计金额形成历史价值判断；最近实销为 0 不直接降为 E。"
    elif "买断+实销" in revenue_model:
        rating = best_rating([sales_rating, buyout_rating, historical_rating])
        basis = "mixed"
        explanation = f"评级 {rating} 综合买断历史价值 {buyout_rating or '缺失'} 与当前实销评级 {sales_rating or '缺失'}；前台只显示一个综合评级。"
    elif active and not is_expired_or_off:
        rating = rating_or_fallback(sales_rating, historical_rating)
        basis = "current_sales"
        explanation = f"评级 {rating} 基于剔除买断后的最近12月实销或年化实销档位；预测 {forecast_rating or '不适用'} 仅作辅助。"
    else:
        rating = rating_or_fallback(historical_rating, sales_rating)
        basis = "historical"
        explanation = f"评级 {rating} 基于历史价值展示；到期/下架状态由状态字段表达，不用 E 覆盖历史价值。"
    return rating, basis, explanation


def status_display(shelf_status: str, rights_status: str) -> str:
    if shelf_status and rights_status:
        return f"{shelf_status}；版权状态={rights_status}"
    return shelf_status or rights_status or "无法判断"


def risk_prompt_for(row: dict, shelf_status: str, rights_status: str, revenue_model: str) -> str:
    prompts = []
    existing = clean(row.get("复核提示"))
    if existing:
        prompts.append(existing)
    if "到期" in shelf_status or rights_status == "expired":
        prompts.append("版权到期或疑似下架，需先复核权利状态。")
    if "尾部收入" in shelf_status:
        prompts.append("下架后仍有尾部收入，需确认收入来源与权利边界。")
    if "买断+实销" in revenue_model:
        prompts.append("买断+实销需要核对渠道拆分，避免把买断金额当作当前实销。")
    if "未知" in revenue_model:
        prompts.append("收入模式未知，需人工复核。")
    return "；".join(dict.fromkeys(prompts)) or "无强运营动作，仅保留常规复核。"


def m4_reason_for(row: dict, anonymous_id: str, m4_ids: set[str]) -> str:
    reasons = []
    if anonymous_id in m4_ids:
        reasons.append("v2用户反馈命中M4校准候选")
    if clean(row.get("用户反馈：收入模式是否合理")) == "不合理":
        reasons.append("收入模式反馈不合理")
    if clean(row.get("用户反馈：评级是否合理")) == "不合理":
        reasons.append("评级反馈不合理")
    if clean(row.get("用户反馈：建议是否可执行")) == "不可执行":
        reasons.append("v2运营建议不可执行")
    note = clean(row.get("用户反馈：问题说明"))
    if note:
        reasons.append("用户备注存在需校准问题")
    return "；".join(reasons) or "作为v3规则抽检样本"


def build_revenue_rule_report(rows: list[dict]) -> dict:
    rows = rows_with_standard_work_id(rows)
    v2 = read_payload(DOCS_DIR / "M2-revenue-model-classification-v2.json")
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "userRules": {
            "pure_sales_share": "按渠道判断；多月连续或半连续；金额不标准且自然变化；无同批次同额买断信号。",
            "pure_buyout": "按渠道判断；个别月份大额；金额整齐；同批次多作品同额或近似均分；后续该渠道无持续实销入账。",
            "buyout_plus_sales": "不同渠道或不同阶段同时满足买断和实销，或买断后存在持续实销尾部。",
            "unknown": "仅限数据不足或冲突，不得过度保守。"
        },
        "currentRecognition": v2.get("revenueModelDistribution", {}),
        "unknownCount": v2.get("revenueModelDistribution", {}).get("unknown_revenue_model", 0),
        "buyoutPlusSalesCount": v2.get("revenueModelDistribution", {}).get("buyout_plus_sales", 0),
        "taskPackRevenueModelDistribution": dict(Counter(row["收入模式"] for row in rows)),
        "typicalPatternsNeedingUserReview": [
            "单月大额但后续有尾部收入的混合渠道",
            "买断+实销渠道拆分",
            "低频单月收入且金额不高的 unknown 边界"
        ],
        "sanitized": True,
        "m3Entered": False,
    }


def build_shelf_rule_report(rows: list[dict]) -> dict:
    rows = rows_with_standard_work_id(rows)
    v2 = read_payload(DOCS_DIR / "M2-shelf-status-inference-v1.json")
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "businessRules": [
            "版权到期是下架强信号。",
            "收入为 0 不能单独判断下架。",
            "已下架作品仍可能因已购用户或会员收听产生少量尾部收入。",
            "版权到期 + 少量尾部收入 = 已下架但有尾部收入。",
            "在架状态影响当前评级解释，但不抹掉历史评级。"
        ],
        "fullCohortShelfStatusDistribution": v2.get("shelfStatusDistribution", {}),
        "taskPackShelfStatusDistribution": dict(Counter(row["货架/版权状态"].split("；")[0] for row in rows)),
        "sanitized": True,
        "m3Entered": False,
    }


def build_front_rating_report(rows: list[dict]) -> dict:
    rows = rows_with_standard_work_id(rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "frontFields": ["rating", "ratingBasis", "shelfStatus", "ratingExplanation"],
        "removedMainRatingColumns": [
            "historicalPerformanceRating",
            "salesPerformanceRating",
            "buyoutHistoricalValueRating",
            "forecastValueRating",
            "operationalDecisionRating"
        ],
        "ratingBasisDistribution": dict(Counter(row["评级依据"] for row in rows)),
        "ratingDistribution": ordered_distribution(row["评级"] for row in rows),
        "expiredOrOffShelfRatingDistribution": ordered_distribution(
            row["评级"] for row in rows if "到期" in row["货架/版权状态"] or "下架" in row["货架/版权状态"]
        ),
        "pureBuyoutRatingDistribution": ordered_distribution(row["评级"] for row in rows if row["收入模式"] == "纯买断"),
        "buyoutPlusSalesRatingDistribution": ordered_distribution(row["评级"] for row in rows if row["收入模式"] == "买断+实销"),
        "forecastIsAuxiliaryOnly": True,
        "sanitized": True,
        "m3Entered": False,
    }


def build_suggestion_boundary_report(rows: list[dict]) -> dict:
    rows = rows_with_standard_work_id(rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "automaticOperatingSuggestionMainFieldRemoved": True,
        "m2AutomaticOperatingActionSuggestionEnabled": False,
        "replacementFields": ["riskAndReviewPrompt", "m4CalibrationCandidateReason", "noAutomaticSuggestionReason"],
        "riskAndReviewPromptRows": sum(1 for row in rows if clean(row["风险/复核提示"])),
        "noAutomaticSuggestionReasonRows": sum(1 for row in rows if clean(row["不自动给运营建议原因"])),
        "m4CalibrationCandidateRows": sum(1 for row in rows if row["是否建议进入M4校准案例"] == "是"),
        "reason": "v2 用户反馈显示 30/30 运营建议不可执行；M2 暂不提供自动运营动作建议，待 M4 校准阶段再启用。",
        "sanitized": True,
        "m3Entered": False,
    }


def build_task_summary(rows: list[dict]) -> dict:
    total_rows = len(rows)
    rows = rows_with_standard_work_id(rows)
    user_blank_rows = total_rows - len(rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "privateTaskPackGenerated": True,
        "rows": total_rows,
        "sampleRows": len(rows),
        "userBlankRows": user_blank_rows,
        "frontRatingSimplified": True,
        "singleMainRatingColumn": True,
        "automaticOperatingSuggestionMainFieldRemoved": True,
        "containsStandardWorkIdTitleAuthor": True,
        "internalAuxiliaryRatingsHidden": True,
        "revenueModelDistribution": dict(Counter(row["收入模式"] for row in rows)),
        "unknownRevenueModelCount": sum(1 for row in rows if "未知" in row["收入模式"]),
        "ratingDistribution": ordered_distribution(row["评级"] for row in rows),
        "sAndSPlusCount": sum(1 for row in rows if row["评级"] in {"S+", "S"}),
        "expiredOrOffShelfRatingDistribution": ordered_distribution(
            row["评级"] for row in rows if "到期" in row["货架/版权状态"] or "下架" in row["货架/版权状态"]
        ),
        "pureBuyoutRatingDistribution": ordered_distribution(row["评级"] for row in rows if row["收入模式"] == "纯买断"),
        "buyoutPlusSalesRatingDistribution": ordered_distribution(row["评级"] for row in rows if row["收入模式"] == "买断+实销"),
        "riskAndReviewPromptRows": sum(1 for row in rows if clean(row["风险/复核提示"])),
        "noAutomaticSuggestionReasonRows": sum(1 for row in rows if clean(row["不自动给运营建议原因"])),
        "m4CalibrationCandidateRows": sum(1 for row in rows if row["是否建议进入M4校准案例"] == "是"),
        "nextUserReview": [
            "确认收入模式是否合理。",
            "确认货架/版权状态是否合理。",
            "确认单一评级是否符合业务直觉。",
            "确认是否进入 M4 校准案例。"
        ],
        "privateOnly": True,
        "sanitized": True,
        "m3Entered": False,
    }


def write_private_pack(rows: list[dict]) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "00_阅读说明"
    append_rows(
        ws,
        [
            ["项目", "说明"],
            ["用途", "rating-standard-v3 任务包；前台只保留一个评级字段、一个状态字段和评级说明。"],
            ["建议边界", "M2 不输出自动运营动作建议，仅保留风险/复核提示与 M4 校准候选原因。"],
            ["提交边界", "private 文件，不提交 Git，不写正式主数据，不进入 M3。"],
        ],
    )
    task = wb.create_sheet("01_运营任务卡")
    headers = list(rows[0].keys()) if rows else []
    append_rows(task, [headers])
    for row in rows:
        append_rows(task, [[row.get(header, "") for header in headers]])
    for column_index, header in enumerate(headers, start=1):
        if header.startswith("内部辅助："):
            task.column_dimensions[get_column_letter(column_index)].hidden = True
    finalize_workbook(wb)
    wb.save(OUTPUT_V3_PACK)


def write_report(json_path: Path, md_path: Path, report_id: str, payload: dict, md_renderer) -> None:
    write_json(json_path, envelope(report_id, payload))
    write_text(md_path, md_renderer(payload))


def render_revenue_md(payload: dict) -> str:
    rows = [{"model": key, "count": value} for key, value in payload["currentRecognition"].items()]
    return "# M2 Revenue Model Business Rule Alignment v1\n\n" + markdown_table(rows, ["model", "count"])


def render_shelf_md(payload: dict) -> str:
    rows = [{"status": key, "count": value} for key, value in payload["fullCohortShelfStatusDistribution"].items()]
    return "# M2 Shelf Status Business Rule Alignment v1\n\n" + markdown_table(rows, ["status", "count"])


def render_front_rating_md(payload: dict) -> str:
    rows = [{"rating": key, "count": value} for key, value in payload["ratingDistribution"].items()]
    return "# M2 Front Rating Simplification v1\n\n" + markdown_table(rows, ["rating", "count"])


def render_suggestion_md(payload: dict) -> str:
    rows = [{"metric": key, "value": value} for key, value in payload.items() if key.endswith("Rows")]
    return "# M2 Suggestion Removal Boundary v1\n\n" + markdown_table(rows, ["metric", "value"])


def render_task_summary_md(payload: dict) -> str:
    rows = [{"rating": key, "count": value} for key, value in payload["ratingDistribution"].items()]
    return "\n".join(
        [
            "# M2 Rating Standard v3 Task Pack Summary",
            "",
            f"- Rows: `{payload['rows']}`",
            f"- Non-empty sample rows: `{payload.get('sampleRows')}`",
            f"- User specified blank rows: `{payload.get('userBlankRows')}`",
            f"- Private task pack generated: `{payload['privateTaskPackGenerated']}`",
            f"- Single main rating column: `{payload['singleMainRatingColumn']}`",
            f"- Automatic operating suggestion main field removed: `{payload['automaticOperatingSuggestionMainFieldRemoved']}`",
            "",
            markdown_table(rows, ["rating", "count"]),
        ]
    )


def append_rows(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)


def finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column_cells in enumerate(ws.columns, start=1):
            max_len = min(46, max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2)
            ws.column_dimensions[get_column_letter(index)].width = max(12, max_len)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def read_payload(path: Path) -> dict:
    return read_json(path).get("payload", {})


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def envelope(report_id: str, payload: dict) -> dict:
    return {
        "reportId": report_id,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "payload": payload,
    }


def markdown_table(rows: list[dict], columns: list[str]) -> str:
    lines = ["| " + " | ".join(columns) + " |", "|" + "|".join("---" for _ in columns) + "|"]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(column, "")) for column in columns) + " |")
    return "\n".join(lines) + "\n\nNo real work titles, authors, channel names, or raw billing rows are included.\n"


def ordered_distribution(values) -> dict:
    counter = Counter(value for value in values if value)
    return {rating: int(counter.get(rating, 0)) for rating in RATINGS}


def rating_or_fallback(value: str, fallback: str) -> str:
    return value if value in RANK else fallback if fallback in RANK else "E"


def best_rating(values: list[str]) -> str:
    clean_values = [value for value in values if value in RANK]
    if not clean_values:
        return "E"
    return sorted(clean_values, key=lambda item: RANK[item])[0]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


if __name__ == "__main__":
    main()
