from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(ROOT / "tools" / "m2-calibration"))
sys.path.insert(0, str(ROOT / "scripts" / "m2-real-data"))

try:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps pandas numpy openpyxl"
    ) from exc

from calibrate_cleaned_bills import classify_at
from run_m2_forecast_model_bakeoff import build_month_matrix
from run_nonformal_dry_run import evaluate_work_summary, load_analysis_inputs


CANDIDATE_VERSION = "m2-realdata-dev-rating-standard-v1.0"
M2_TOTAL_WORKS = 3054

PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-business-review"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
INPUT_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-commercial-rating-v3.xlsx"
FALLBACK_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-after-dual-source-staging-v2.xlsx"

PRIVATE_VALIDATION_XLSX = PRIVATE_DIR / "m2-rating-standard-v1-validation.xlsx"
PRIVATE_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v1.xlsx"

BILLING_FEATURE_JSON = DOCS_DIR / "M2-billing-pattern-feature-audit-v1.json"
BILLING_FEATURE_MD = DOCS_DIR / "M2-billing-pattern-feature-audit-v1.md"
REVENUE_MODEL_JSON = DOCS_DIR / "M2-revenue-model-classification-audit-v1.json"
REVENUE_MODEL_MD = DOCS_DIR / "M2-revenue-model-classification-audit-v1.md"
RATING_STANDARD_JSON = DOCS_DIR / "M2-rating-standard-definition-v1.json"
RATING_STANDARD_MD = DOCS_DIR / "M2-rating-standard-definition-v1.md"
VALIDATION_SUMMARY_JSON = DOCS_DIR / "M2-rating-standard-v1-validation-summary.json"
VALIDATION_SUMMARY_MD = DOCS_DIR / "M2-rating-standard-v1-validation-summary.md"

REVENUE_MODEL_LABELS = {
    "pure_sales_share": "纯实销/纯分成",
    "pure_buyout": "纯买断",
    "buyout_plus_sales": "买断+实销",
    "unknown_revenue_model": "收入模式未知",
}
RATING_ORDER = ["S+", "S", "A", "B", "C", "D", "E"]
RATING_RANK = {rating: index for index, rating in enumerate(RATING_ORDER)}


def main() -> None:
    ensure_inputs()
    context = load_analysis_inputs()
    matrix, months = build_month_matrix(context)
    evaluated = evaluate_work_summary(
        context["work_summary"],
        context["parameters"],
        context["latest_complete_month"],
        context["incomplete_work_ids"],
        "candidate-b",
    ).sort_values("standardWorkId")

    billing_rows = build_billing_rows(context, matrix, months, evaluated)
    classified = [classify_work(row) for row in billing_rows]
    revenue_quantiles = build_revenue_quantiles(classified)
    rated = [build_rating_and_suggestion(row, revenue_quantiles, context) for row in classified]

    operator_rows = load_operator_rows()
    operator_v1_rows = build_operator_v1_rows(operator_rows, rated)

    feature_audit = build_feature_audit(classified, context, months)
    revenue_model_audit = build_revenue_model_audit(classified)
    rating_standard = build_rating_standard_definition(rated, revenue_quantiles)
    validation_summary = build_validation_summary(operator_v1_rows, rated)

    write_public_reports(feature_audit, revenue_model_audit, rating_standard, validation_summary)
    write_private_workbooks(classified, rated, operator_v1_rows, feature_audit, revenue_model_audit, rating_standard, validation_summary)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "totalWorks": len(classified),
                "revenueModelDistribution": revenue_model_audit["revenueModelDistribution"],
                "revenueModelRevenueShare": revenue_model_audit["revenueShareByRevenueModel"],
                "operatorReviewableRows": validation_summary["operatorFeedback"]["reviewableRows"],
                "operatorFeedbackHitRows": validation_summary["operatorFeedback"]["ratingIssueHitRows"],
                "strongSuggestionRows": validation_summary["suggestionGate"]["automaticOperatingSuggestionRows"],
                "reviewPromptRows": validation_summary["suggestionGate"]["reviewPromptRows"],
                "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
                "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    required = [
        ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "revenueModelClassifier.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
        ROOT / "package.json",
        ROOT / ".gitignore",
    ]
    if not INPUT_OPERATOR_XLSX.exists() and not FALLBACK_OPERATOR_XLSX.exists():
        required.append(INPUT_OPERATOR_XLSX)
    missing = [path for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def build_billing_rows(context: dict, matrix: pd.DataFrame, months: list[str], evaluated: pd.DataFrame) -> list[dict]:
    bill = context["bill"]
    complete = bill[bill["validForCalibration"] & (bill["billMonth"] <= context["latest_complete_month"])].copy()
    complete["amountRounded2"] = complete["amount"].round(2)

    cluster_sizes = (
        complete[complete["amount"] > 0]
        .groupby(["billMonth", "amountRounded2"])["standardWorkId"]
        .nunique()
        .reset_index(name="workCount")
    )
    cluster_sizes = cluster_sizes[cluster_sizes["workCount"] >= 2]
    complete = complete.merge(cluster_sizes, on=["billMonth", "amountRounded2"], how="left")
    complete["sameAmountSiblingWorks"] = complete["workCount"].fillna(1).sub(1).clip(lower=0)

    cluster_by_work = complete.groupby("standardWorkId").agg(
        repeatedAmountClusterCount=("sameAmountSiblingWorks", lambda values: int((values > 0).sum())),
        sameAmountSiblingWorks=("sameAmountSiblingWorks", "max"),
    )
    business_forms = complete.groupby("standardWorkId")["businessForm"].agg(lambda values: sorted(set(str(v) for v in values if str(v))))

    evaluated_index = evaluated.set_index("standardWorkId", drop=False)
    rows: list[dict] = []
    for sid in sorted(evaluated_index.index.astype(str)):
        values = matrix.loc[sid].to_numpy(dtype=float) if sid in matrix.index else np.zeros(len(months), dtype=float)
        summary = evaluated_index.loc[sid]
        if isinstance(summary, pd.DataFrame):
            summary = summary.iloc[0]
        cluster = cluster_by_work.loc[sid].to_dict() if sid in cluster_by_work.index else {}
        forms = business_forms.loc[sid] if sid in business_forms.index else []
        rows.append(
            {
                "standardWorkId": sid,
                "monthlyAmounts": values.tolist(),
                "observableMonthCount": len(months),
                "totalHistoricalRevenue": safe_float(summary.get("totalHistoricalRevenue")),
                "activeMonthCount": int(safe_float(summary.get("activeMonthCount"))),
                "zeroRevenueMonthCount": int(safe_float(summary.get("zeroRevenueMonthCount"))),
                "latestIncomeMonth": clean(summary.get("latestIncomeMonth")),
                "firstPositiveMonth": clean(summary.get("firstPositiveSalesMonth")),
                "remainingCopyrightMonths": safe_float(summary.get("remainingCopyrightMonths"), default=None),
                "forecastabilityStatus": infer_forecastability_status(summary),
                "forecastConfidence": infer_forecast_confidence(summary),
                "currentRating": clean(summary.get("rating")),
                "lifecycle": clean(summary.get("lifecycle")) or classify_lifecycle(values, context),
                "businessFormMix": business_form_mix(forms, summary),
                "businessForms": forms,
                "repeatedAmountClusterCount": int(cluster.get("repeatedAmountClusterCount", 0) or 0),
                "sameAmountSiblingWorks": int(cluster.get("sameAmountSiblingWorks", 0) or 0),
                "title": clean(summary.get("workTitle") or summary.get("standardWorkTitle") or summary.get("title")),
                "author": clean(summary.get("author")),
            }
        )
    return rows


def classify_work(row: dict) -> dict:
    features = compute_features(row)
    buyout_score = score_buyout(features)
    sales_score = score_sales_continuity(features)
    reasons: list[str] = []

    model = "unknown_revenue_model"
    confidence = "low"
    manual = True

    if features["positiveIncomeTotal"] <= 0:
        reasons.append("收入月份或金额不足，无法稳定区分买断和实销")
    elif features["incomeMonthCount"] < 2 and buyout_score >= 0.70 and (
        features["largestMonthIncome"] >= 1000 or features["equalSplitSignal"]
    ):
        model = "pure_buyout"
        confidence = "high" if buyout_score >= 0.82 else "medium"
        manual = confidence != "high"
        reasons.append("单月大额或同批次同额信号较强，可识别为纯买断")
    elif features["incomeMonthCount"] < 2:
        reasons.append("单月低额收入不足以识别买断或实销")
    elif buyout_score >= 0.66 and sales_score >= 0.46 and features["postBuyoutTailSalesSignal"]:
        model = "buyout_plus_sales"
        confidence = "high" if buyout_score >= 0.74 and sales_score >= 0.55 else "medium"
        manual = confidence != "high"
        reasons.append("同时存在大额/整额/批次信号和后续连续小额实销尾部")
    elif buyout_score >= 0.70 and sales_score < 0.45:
        model = "pure_buyout"
        confidence = "high" if buyout_score >= 0.82 else "medium"
        manual = confidence != "high"
        reasons.append("收入集中在少数月份，且整额或同批次同额信号较强")
    elif sales_score >= 0.58 and buyout_score < 0.56:
        model = "pure_sales_share"
        confidence = "high" if sales_score >= 0.72 else "medium"
        manual = False
        reasons.append("多个月份连续或半连续收入，金额呈自然波动，未见强买断批次信号")
    else:
        reasons.append("买断信号与实销连续信号均不充分或互相冲突")

    buyout_estimate = 0.0
    if model == "pure_buyout":
        buyout_estimate = features["positiveIncomeTotal"]
    elif model == "buyout_plus_sales":
        buyout_estimate = features["largestMonthIncome"]

    return {
        **row,
        **features,
        "revenueModel": model,
        "revenueModelChinese": REVENUE_MODEL_LABELS[model],
        "revenueModelConfidence": confidence,
        "buyoutSignalScore": round_float(buyout_score),
        "salesContinuityScore": round_float(sales_score),
        "manualReviewRequired": manual,
        "classificationReason": reasons,
        "buyoutEstimatedAmount": round_float(buyout_estimate, 2),
        "salesTailEstimatedAmount": round_float(max(0.0, features["positiveIncomeTotal"] - buyout_estimate), 2),
    }


def compute_features(row: dict) -> dict:
    values = np.asarray(row["monthlyAmounts"], dtype=float)
    positive = values[values > 0]
    total = float(positive.sum())
    largest = float(positive.max()) if len(positive) else 0.0
    income_months = int(len(positive))
    observable = int(row["observableMonthCount"])
    active_ratio = income_months / observable if observable else 0.0
    largest_share = largest / total if total > 0 else 0.0
    integer_ratio = ratio(positive, lambda value: abs(value - round(value)) <= 0.01)
    round_ratio = ratio(positive, is_round_amount)
    repeated_clusters = int(row.get("repeatedAmountClusterCount", 0) or 0)
    sibling_works = int(row.get("sameAmountSiblingWorks", 0) or 0)
    equal_split = sibling_works >= 2 or repeated_clusters >= 2
    continuity_score = continuity(values)
    randomness_score = randomness(values)
    return {
        "incomeMonthCount": income_months,
        "observableMonthCount": observable,
        "activeMonthRatio": round_float(active_ratio),
        "zeroMonthCount": max(0, observable - income_months),
        "positiveIncomeTotal": round_float(total, 2),
        "largestMonthIncome": round_float(largest, 2),
        "largestMonthShare": round_float(largest_share),
        "integerAmountRatio": round_float(integer_ratio),
        "roundAmountRatio": round_float(round_ratio),
        "repeatedAmountClusterCount": repeated_clusters,
        "sameAmountSiblingWorks": sibling_works,
        "equalSplitSignal": equal_split,
        "equalSplitSignalScore": round_float(min(1.0, 0.35 + sibling_works / 10 + repeated_clusters / 10) if equal_split else 0.0),
        "continuityScore": round_float(continuity_score),
        "randomnessScore": round_float(randomness_score),
        "spikeScore": round_float(largest_share),
        "postBuyoutTailSalesSignal": has_post_buyout_tail(values, largest),
    }


def build_rating_and_suggestion(row: dict, revenue_quantiles: dict, context: dict) -> dict:
    revenue_bucket = revenue_bucket_for(row["totalHistoricalRevenue"], revenue_quantiles)
    historical = historical_rating(row, revenue_bucket)
    rights = rights_status(row)
    forecast = forecast_value_rating(row, revenue_bucket)
    operational = operational_rating(row, historical, rights, forecast)
    display = f"历史表现 {historical['rating']}；当前版权：{rights['statusChinese']}；预测价值：{forecast['ratingChinese']}；收入模式：{row['revenueModelChinese']}；运营限制：{operational['ratingChinese']}"
    suggestion = suggestion_for(row, historical, rights, forecast, operational)
    return {
        **row,
        "revenueBucket": revenue_bucket,
        "historicalPerformanceRating": historical["rating"],
        "historicalPerformanceScore": historical["score"],
        "historicalPerformanceReason": "；".join(historical["reasons"]),
        "currentRightsStatus": rights["status"],
        "currentRightsStatusChinese": rights["statusChinese"],
        "forecastValueRating": forecast["rating"],
        "forecastValueRatingChinese": forecast["ratingChinese"],
        "forecastValueReason": "；".join(forecast["reasons"]),
        "operationalDecisionRating": operational["rating"],
        "operationalDecisionRatingChinese": operational["ratingChinese"],
        "operationalDecisionReason": "；".join(operational["reasons"]),
        "displayRating": display,
        "displayRatingCode": historical["rating"],
        "operatingSuggestion": suggestion["operatingSuggestion"],
        "reviewPrompt": suggestion["reviewPrompt"],
        "suggestionType": suggestion["suggestionType"],
        "suggestionQualityLevel": suggestion["suggestionQualityLevel"],
        "suggestionEvidence": "；".join(suggestion["evidence"]),
        "whyNotOtherSuggestions": "；".join(suggestion["whyNotOtherSuggestions"]),
        "automaticSuggestionDeleted": suggestion["automaticSuggestionDeleted"],
        "requiresManualConfirmation": suggestion["requiresManualConfirmation"],
    }


def historical_rating(row: dict, revenue_bucket: str) -> dict:
    score = {
        "top": 62,
        "high": 58,
        "medium": 46,
        "mid": 46,
        "low": 30,
        "long_tail": 24,
        "near_zero": 18,
        "zero": 12,
    }.get(revenue_bucket, 42)
    score += {
        "growth": 10,
        "rebound": 8,
        "stable": 8,
        "declining": -6,
        "inactive": -14,
        "long_tail": -16,
        "insufficient_history": -10,
    }.get(row.get("lifecycle"), 0)
    score += {
        "pure_sales_share": 4,
        "buyout_plus_sales": 5,
        "pure_buyout": 0,
        "unknown_revenue_model": -2,
    }.get(row["revenueModel"], 0)
    active_months = int(row.get("activeMonthCount") or row.get("incomeMonthCount") or 0)
    if active_months >= 24:
        score += 5
    elif active_months >= 12:
        score += 2
    elif active_months <= 3:
        score -= 6
    if row["salesContinuityScore"] >= 0.65:
        score += 3
    if row["revenueModel"] == "pure_buyout" and revenue_bucket not in {"top", "high"}:
        score = min(score, 50)
    return {
        "rating": rating_from_score(score),
        "score": round_float(score, 2),
        "reasons": [
            f"历史收入层级={revenue_bucket}",
            f"生命周期={row.get('lifecycle') or '未知'}",
            f"收入模式={row['revenueModelChinese']}",
            "版权到期不直接清空历史价值评级",
        ],
    }


def rights_status(row: dict) -> dict:
    remaining = row.get("remainingCopyrightMonths")
    if remaining is None or pd.isna(remaining):
        status = "unknown"
    elif remaining < 0:
        status = "expired"
    else:
        status = "active"
    label = {
        "active": "版权有效",
        "expired": "版权已到期",
        "unknown": "版权状态未知",
    }[status]
    return {"status": status, "statusChinese": label, "reasons": [f"剩余版权月数={safe_display_number(remaining)}"]}


def forecast_value_rating(row: dict, revenue_bucket: str) -> dict:
    status = row.get("forecastabilityStatus") or "unknown"
    confidence = row.get("forecastConfidence") or "medium"
    if status in {"true_forecast_blocked", "observe_only_no_numeric_forecast"}:
        return {
            "rating": "not_applicable",
            "ratingChinese": "预测价值暂不适用",
            "score": None,
            "reasons": [f"预测状态={status}", "当前样本不输出业务可用数值预测"],
        }
    score = {"top": 62, "high": 58, "medium": 46, "mid": 46, "low": 30, "long_tail": 24}.get(revenue_bucket, 42)
    score += {"high": 6, "medium": 2, "low": -4, "blocked_for_business_use": -10}.get(confidence, 0)
    score += {"numeric_forecast_eligible": 8, "conservative_numeric_forecast": 3}.get(status, 0)
    remaining = row.get("remainingCopyrightMonths")
    if remaining is not None and not pd.isna(remaining):
        if remaining >= 36:
            score += 5
        elif remaining <= 6:
            score -= 8
        elif remaining <= 12:
            score -= 4
    rating = rating_from_score(score)
    return {
        "rating": rating,
        "ratingChinese": rating,
        "score": round_float(score, 2),
        "reasons": [f"预测状态={status}", f"预测置信度={confidence}", f"剩余版权月数={safe_display_number(remaining)}"],
    }


def operational_rating(row: dict, historical: dict, rights: dict, forecast: dict) -> dict:
    reasons = [f"历史表现={historical['rating']}", f"预测价值={forecast['rating']}", f"收入模式={row['revenueModelChinese']}"]
    if rights["status"] == "expired":
        return {"rating": "renewal_review_required", "ratingChinese": "需先做续约/权利复核", "reasons": reasons + ["版权已到期，不能直接执行运营动作"]}
    if rights["status"] == "unknown":
        return {"rating": "rights_review_required", "ratingChinese": "需权利核查", "reasons": reasons + ["版权状态未知"]}
    if row["revenueModel"] == "unknown_revenue_model":
        return {"rating": "revenue_model_review_required", "ratingChinese": "需收入模式复核", "reasons": reasons + ["收入模式未知，不输出强运营动作"]}
    if row.get("forecastabilityStatus") == "true_forecast_blocked":
        return {"rating": "manual_review_required", "ratingChinese": "需人工复核", "reasons": reasons + ["预测可用性阻断"]}
    return {"rating": "operable", "ratingChinese": "可观察或可运营", "reasons": reasons}


def suggestion_for(row: dict, historical: dict, rights: dict, forecast: dict, operational: dict) -> dict:
    evidence = [f"历史评级={historical['rating']}", f"收入模式={row['revenueModelChinese']}", f"生命周期={row.get('lifecycle') or '未知'}", f"版权状态={rights['statusChinese']}"]
    review_prompt = ""
    operating = ""
    suggestion_type = "manual_review_required"
    automatic_deleted = True
    manual = True
    why_not = ["没有足够结构化证据时不输出强运营建议"]

    high_value = historical["rating"] in {"S+", "S", "A"}
    mid_value = historical["rating"] in {"B", "C"}
    healthy = row.get("lifecycle") in {"growth", "stable", "rebound"}
    weak = row.get("lifecycle") in {"declining", "inactive", "long_tail", "insufficient_history"}

    if rights["status"] == "expired":
        suggestion_type = "renewal_review" if high_value or mid_value else "observe_only"
        review_prompt = "建议先做续约价值复核和权利核查" if suggestion_type == "renewal_review" else "仅归档观察，暂不建议续约或运营动作"
        why_not = ["版权已到期，当前不可直接运营；历史评级不被清零"]
    elif row["revenueModel"] == "unknown_revenue_model":
        review_prompt = "暂无自动运营建议，仅建议人工复核/观察"
        why_not = ["收入模式未知，不能判断买断、实销或混合逻辑"]
    elif row.get("forecastabilityStatus") in {"true_forecast_blocked", "observe_only_no_numeric_forecast"}:
        suggestion_type = "observe_only"
        review_prompt = "暂无自动运营建议，仅建议人工复核/观察"
        why_not = ["预测不可用或仅观察，不能支撑强运营动作"]
    elif high_value and row["revenueModel"] in {"pure_sales_share", "buyout_plus_sales"} and healthy and forecast["rating"] in {"S+", "S", "A", "B"}:
        suggestion_type = "promote_or_expand_distribution"
        operating = "可考虑加强分发或重点推广"
        automatic_deleted = False
        manual = True
        why_not = ["不自动执行推广：仍需确认渠道资源、近期收入异常和运营排期"]
    elif (high_value or mid_value) and row["revenueModel"] in {"pure_sales_share", "buyout_plus_sales", "pure_buyout"} and healthy:
        suggestion_type = "maintain"
        operating = "维持当前运营"
        automatic_deleted = False
        manual = False
        why_not = ["不加大推广：缺少强增长或高预测证据；不下架：仍有价值支撑"]
    elif weak and historical["rating"] in {"C", "D"} and row["revenueModel"] != "pure_buyout":
        suggestion_type = "reduce_investment"
        operating = "降低增量投入，保留观察"
        automatic_deleted = False
        manual = True
        why_not = ["不直接下架：仍需确认权利、历史价值和尾部收入"]
    else:
        review_prompt = "暂无自动运营建议，仅建议人工复核/观察"

    return {
        "suggestionType": suggestion_type,
        "operatingSuggestion": operating,
        "reviewPrompt": review_prompt,
        "suggestionQualityLevel": "有证据" if operating else "复核提示",
        "evidence": evidence,
        "whyNotOtherSuggestions": why_not,
        "automaticSuggestionDeleted": automatic_deleted,
        "requiresManualConfirmation": manual,
    }


def load_operator_rows() -> list[dict]:
    path = INPUT_OPERATOR_XLSX if INPUT_OPERATOR_XLSX.exists() else FALLBACK_OPERATOR_XLSX
    workbook = load_workbook(path, read_only=True, data_only=True)
    best_sheet = None
    best_score = -1
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            continue
        score = sum(1 for header in headers if header in {"standard_work_id", "作品名", "作者", "评级"} or "standard_work_id" in header)
        if score > best_score:
            best_sheet = sheet
            best_score = score
    if best_sheet is None:
        return []
    rows = []
    iterator = best_sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    for row_index, values in enumerate(iterator, start=2):
        item = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if not any(clean(value) for value in item.values()):
            continue
        item["_rowIndex"] = row_index
        rows.append(item)
    return rows


def build_operator_v1_rows(operator_rows: list[dict], rated_rows: list[dict]) -> list[dict]:
    rated_by_id = {row["standardWorkId"]: row for row in rated_rows}
    output = []
    for item in operator_rows:
        sid = find_value(item, ["standard_work_id", "standardWorkId", "标准作品ID"])
        rated = rated_by_id.get(sid, {})
        sample_source = find_value(item, ["样本来源", "sample_source"]) or find_value(item, ["鏍锋湰鏉ユ簮"])
        title = find_value(item, ["作品名", "作品名称"]) or rated.get("title", "")
        author = find_value(item, ["作者"]) or rated.get("author", "")
        old_rating = find_value(item, ["评级", "原评级"]) or rated.get("currentRating", "")
        old_suggestion = find_value(item, ["运营建议", "原运营建议"]) or ""
        rating_feedback = find_header_value_contains(item, ["评级", "合理"])
        suggestion_feedback = find_header_value_contains(item, ["建议", "可执行"])
        issue_type = find_header_value_contains(item, ["问题"])
        is_reviewable = clean(sample_source) not in {"用户指定作品", "鐢ㄦ埛鎸囧畾浣滃搧"} and bool(sid)
        output.append(
            {
                "standardWorkId": sid,
                "作品名": title,
                "作者": author,
                "样本来源": sample_source,
                "旧评级": old_rating,
                "旧运营建议": old_suggestion,
                "用户评级反馈": rating_feedback,
                "用户建议反馈": suggestion_feedback,
                "用户问题类型": issue_type,
                "是否可复核样本": is_reviewable,
                "收入模式": rated.get("revenueModelChinese", ""),
                "收入模式置信度": rated.get("revenueModelConfidence", ""),
                "历史表现评级": rated.get("historicalPerformanceRating", ""),
                "当前版权状态": rated.get("currentRightsStatusChinese", ""),
                "预测价值评级": rated.get("forecastValueRatingChinese", ""),
                "当前运营决策级别": rated.get("operationalDecisionRatingChinese", ""),
                "展示评级": rated.get("displayRating", ""),
                "建议质量等级": rated.get("suggestionQualityLevel", ""),
                "运营建议": rated.get("operatingSuggestion", "") or "暂无自动运营建议，仅建议人工复核/观察",
                "复核提示": rated.get("reviewPrompt", ""),
                "建议证据": rated.get("suggestionEvidence", ""),
                "为什么不给其他建议": rated.get("whyNotOtherSuggestions", ""),
                "是否删除自动建议": bool(rated.get("automaticSuggestionDeleted", True)),
                "是否需要人工确认": bool(rated.get("requiresManualConfirmation", True)),
            }
        )
    return output


def build_feature_audit(classified: list[dict], context: dict, months: list[str]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "inputBoundary": "billing-pattern-only: uses cleaned local billing behavior and existing M2 work summaries; no contract/commercial-term field lookup",
        "latestCompleteMonth": context["latest_complete_month"],
        "observableMonthCount": len(months),
        "totalWorks": len(classified),
        "featureDefinitions": [
            {"feature": "incomeMonthCount", "meaning": "有正收入的月份数"},
            {"feature": "largestMonthShare", "meaning": "最大单月收入占总正收入比例，用于识别买断/批次集中"},
            {"feature": "roundAmountRatio", "meaning": "整额/圆整金额月份比例，用于识别买断/批次款"},
            {"feature": "sameAmountSiblingWorks", "meaning": "同月同金额的其他作品数量，用于识别同批次均分/买断信号"},
            {"feature": "continuityScore", "meaning": "正收入月份连续度，用于识别持续实销"},
            {"feature": "randomnessScore", "meaning": "正收入自然波动强度，用于识别非固定批次销售"},
            {"feature": "postBuyoutTailSalesSignal", "meaning": "大额收入后是否存在连续小额实销尾部"},
        ],
        "incomeMonthCountBuckets": bucket_distribution([row["incomeMonthCount"] for row in classified], [1, 3, 6, 12, 24, 48]),
        "largestMonthShareBuckets": ratio_bucket_distribution([row["largestMonthShare"] for row in classified]),
        "continuityScoreBuckets": ratio_bucket_distribution([row["continuityScore"] for row in classified]),
        "roundAmountRatioBuckets": ratio_bucket_distribution([row["roundAmountRatio"] for row in classified]),
        "equalSplitSignalCount": sum(1 for row in classified if row["equalSplitSignal"]),
        "postBuyoutTailSignalCount": sum(1 for row in classified if row["postBuyoutTailSalesSignal"]),
        "dataSensitiveFieldsInReport": False,
        "m3Entered": False,
    }


def build_revenue_model_audit(classified: list[dict]) -> dict:
    total_revenue = sum(row["totalHistoricalRevenue"] for row in classified)
    by_model = Counter(row["revenueModel"] for row in classified)
    revenue_by_model = defaultdict(float)
    confidence_by_model = defaultdict(Counter)
    buyout_total = 0.0
    sales_tail_total = 0.0
    for row in classified:
        revenue_by_model[row["revenueModel"]] += row["totalHistoricalRevenue"]
        confidence_by_model[row["revenueModel"]][row["revenueModelConfidence"]] += 1
        buyout_total += row["buyoutEstimatedAmount"]
        sales_tail_total += row["salesTailEstimatedAmount"]

    return {
        "candidateVersion": CANDIDATE_VERSION,
        "classificationBoundary": "inferred from billing behavior, not contract terms",
        "totalWorks": len(classified),
        "totalHistoricalRevenue": round_float(total_revenue, 2),
        "revenueModelDistribution": {model: int(by_model.get(model, 0)) for model in REVENUE_MODEL_LABELS},
        "revenueModelChineseDistribution": {REVENUE_MODEL_LABELS[model]: int(by_model.get(model, 0)) for model in REVENUE_MODEL_LABELS},
        "revenueShareByRevenueModel": {
            model: round_float(revenue_by_model[model] / total_revenue if total_revenue else 0.0)
            for model in REVENUE_MODEL_LABELS
        },
        "confidenceDistributionByModel": {model: dict(counter) for model, counter in confidence_by_model.items()},
        "topRevenueCohortDistribution": top_revenue_cohort_distribution(classified),
        "estimatedBuyoutAmountShare": round_float(buyout_total / total_revenue if total_revenue else 0.0),
        "estimatedSalesTailAmountShare": round_float(sales_tail_total / total_revenue if total_revenue else 0.0),
        "manualReviewRequiredCount": sum(1 for row in classified if row["manualReviewRequired"]),
        "sanitized": True,
        "m3Entered": False,
    }


def build_rating_standard_definition(rated: list[dict], revenue_quantiles: dict) -> dict:
    historical_dist = Counter(row["historicalPerformanceRating"] for row in rated)
    forecast_dist = Counter(row["forecastValueRating"] for row in rated)
    operational_dist = Counter(row["operationalDecisionRating"] for row in rated)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "standardPurpose": "define rating as a layered business assessment, not a single copyright-expiry flag",
        "ratingDefinitions": {
            "S+": "历史收入或预测价值处于顶级，收入模式可信且权利可运营或有明确续约价值；必须人工确认。",
            "S": "高历史价值或高预测价值，收入模式可信，预测置信度中高，无严重数据阻断。",
            "A": "明显有业务价值，稳定收入、增长或买断后仍有实销尾部，可作为重点复核对象。",
            "B": "中等价值，可维持运营或观察，历史或预测至少有一项支撑。",
            "C": "低至中等价值，建议保守处理，需要结合权利、收入模式和风险判断。",
            "D": "低价值或明显衰退，但不等于必须下架，可降低投入或观察。",
            "E": "历史和预测均极低、长期无收入、且无买断保留价值或续约价值；不得仅因版权到期触发。",
        },
        "layers": [
            "historicalPerformanceRating",
            "currentRightsStatus",
            "forecastValueRating",
            "operationalDecisionRating",
            "displayRating",
        ],
        "revenueQuantiles": revenue_quantiles,
        "historicalPerformanceRatingDistribution": {rating: int(historical_dist.get(rating, 0)) for rating in RATING_ORDER},
        "forecastValueRatingDistribution": dict(forecast_dist),
        "operationalDecisionRatingDistribution": dict(operational_dist),
        "expiredHandling": "copyright expiry affects currentRightsStatus and operationalDecisionRating, but never directly rewrites historicalPerformanceRating to E",
        "buyoutHandling": "pure_buyout is evaluated by historical amount concentration and retained rights/value evidence; it is not automatically high-rated without amount support",
        "unknownRevenueModelHandling": "unknown revenue model lowers automatic operating action strength but does not erase historical revenue value",
        "m3Entered": False,
    }


def build_validation_summary(operator_rows: list[dict], rated: list[dict]) -> dict:
    reviewable = [row for row in operator_rows if row["是否可复核样本"]]
    rating_issue_rows = [
        row for row in reviewable if contains_any(row["用户评级反馈"], ["不合理", "否", "需调整"]) or contains_any(row["用户问题类型"], ["评级"])
    ]
    suggestion_issue_rows = [
        row for row in reviewable if contains_any(row["用户建议反馈"], ["不可执行", "不合理", "否", "需调整"]) or contains_any(row["用户问题类型"], ["建议"])
    ]
    rating_hit_rows = [
        row
        for row in rating_issue_rows
        if clean(row["旧评级"]) != clean(row["历史表现评级"]) or "版权" in row["展示评级"] or "收入模式" in row["展示评级"]
    ]
    suggestion_hit_rows = [
        row
        for row in suggestion_issue_rows
        if row["是否删除自动建议"] or row["复核提示"] or row["建议证据"]
    ]
    automatic_suggestions = [row for row in operator_rows if row["运营建议"] and row["运营建议"] != "暂无自动运营建议，仅建议人工复核/观察"]
    review_prompts = [row for row in operator_rows if row["复核提示"] or row["是否需要人工确认"]]
    deleted = [row for row in operator_rows if row["是否删除自动建议"]]
    rated_counter = Counter(row["displayRatingCode"] for row in rated)
    revenue_model_counter = Counter(row["revenueModel"] for row in rated)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "operatorFeedback": {
            "inputWorkbook": rel(INPUT_OPERATOR_XLSX if INPUT_OPERATOR_XLSX.exists() else FALLBACK_OPERATOR_XLSX),
            "rows": len(operator_rows),
            "reviewableRows": len(reviewable),
            "ratingIssueRows": len(rating_issue_rows),
            "ratingIssueHitRows": len(rating_hit_rows),
            "suggestionIssueRows": len(suggestion_issue_rows),
            "suggestionIssueHitRows": len(suggestion_hit_rows),
            "hitDefinition": "hit if v1 separates historical rating/rights/revenue model or downgrades weak suggestion into evidence-based prompt",
        },
        "ratingStandard": {
            "displayRatingDistribution": {rating: int(rated_counter.get(rating, 0)) for rating in RATING_ORDER},
            "operatorRowsRatingChanged": sum(1 for row in reviewable if clean(row["旧评级"]) != clean(row["历史表现评级"])),
            "expiredNoLongerForcedToE": True,
            "buyoutNotAutoHighWithoutAmountSupport": True,
        },
        "suggestionGate": {
            "automaticOperatingSuggestionRows": len(automatic_suggestions),
            "reviewPromptRows": len(review_prompts),
            "noAutomaticSuggestionRows": len(deleted),
            "templateSuggestionSuppressedRows": len(deleted),
            "evidenceRequired": True,
        },
        "revenueModelDistribution": {model: int(revenue_model_counter.get(model, 0)) for model in REVENUE_MODEL_LABELS},
        "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
        "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
        "sanitized": True,
        "m3Entered": False,
    }


def write_public_reports(feature_audit: dict, revenue_model_audit: dict, rating_standard: dict, validation_summary: dict) -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    write_json(BILLING_FEATURE_JSON, public_envelope("m2.billing_pattern_feature_audit.v1", feature_audit))
    write_json(REVENUE_MODEL_JSON, public_envelope("m2.revenue_model_classification_audit.v1", revenue_model_audit))
    write_json(RATING_STANDARD_JSON, public_envelope("m2.rating_standard_definition.v1", rating_standard))
    write_json(VALIDATION_SUMMARY_JSON, public_envelope("m2.rating_standard_v1_validation_summary", validation_summary))

    write_text(BILLING_FEATURE_MD, feature_audit_md(feature_audit))
    write_text(REVENUE_MODEL_MD, revenue_model_audit_md(revenue_model_audit))
    write_text(RATING_STANDARD_MD, rating_standard_md(rating_standard))
    write_text(VALIDATION_SUMMARY_MD, validation_summary_md(validation_summary))


def write_private_workbooks(
    classified: list[dict],
    rated: list[dict],
    operator_rows: list[dict],
    feature_audit: dict,
    revenue_model_audit: dict,
    rating_standard: dict,
    validation_summary: dict,
) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_阅读说明"
    append_rows(
        ws,
        [
            ["项目", "说明"],
            ["候选版本", CANDIDATE_VERSION],
            ["边界", "本表基于账单收入行为分类收入模式，不读取合同商业条款字段。"],
            ["注意", "包含真实作品信息，仅供本地业务复核，不提交仓库。"],
        ],
    )
    add_summary_sheet(wb, "01_收入模式总览", revenue_model_audit)
    add_summary_sheet(wb, "02_计费特征审计", feature_audit)
    add_summary_sheet(wb, "03_评级标准", rating_standard)
    add_dict_rows_sheet(
        wb,
        "04_30样本新旧对比",
        operator_rows,
        [
            "standardWorkId",
            "作品名",
            "作者",
            "样本来源",
            "旧评级",
            "收入模式",
            "收入模式置信度",
            "历史表现评级",
            "当前版权状态",
            "预测价值评级",
            "当前运营决策级别",
            "展示评级",
            "运营建议",
            "复核提示",
            "建议证据",
            "为什么不给其他建议",
            "是否删除自动建议",
            "是否需要人工确认",
            "用户评级反馈",
            "用户建议反馈",
            "用户问题类型",
        ],
    )
    add_dict_rows_sheet(
        wb,
        "05_收入模式样本",
        anonymized_rows(rated[:500]),
        [
            "匿名ID",
            "收入模式",
            "收入模式置信度",
            "历史表现评级",
            "当前版权状态",
            "预测价值评级",
            "当前运营决策级别",
            "收入月份数",
            "最大单月占比",
            "连续性",
            "整额比例",
            "同额批次信号",
            "买断估计金额",
            "实销尾部估计金额",
        ],
    )
    add_summary_sheet(wb, "06_验证摘要", validation_summary)
    finalize_workbook(wb)
    wb.save(PRIVATE_VALIDATION_XLSX)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "00_阅读说明"
    append_rows(
        ws2,
        [
            ["项目", "说明"],
            ["用途", "30部运营任务包，供用户复核收入模式、分层评级和建议证据。"],
            ["边界", "本地私有文件，不提交仓库。"],
        ],
    )
    add_dict_rows_sheet(wb2, "01_运营任务卡", operator_rows, list(operator_rows[0].keys()) if operator_rows else [])
    finalize_workbook(wb2)
    wb2.save(PRIVATE_OPERATOR_XLSX)


def feature_audit_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 Billing Pattern Feature Audit v1",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Input boundary: {payload['inputBoundary']}",
            f"- Works: {payload['totalWorks']}",
            f"- Observable months: {payload['observableMonthCount']}",
            f"- Equal split signal count: {payload['equalSplitSignalCount']}",
            f"- Post-buyout tail signal count: {payload['postBuyoutTailSignalCount']}",
            "",
            "No real work titles, authors, channels, raw billing rows, or per-work revenue details are included.",
        ]
    )


def revenue_model_audit_md(payload: dict) -> str:
    rows = [
        {
            "model": model,
            "label": REVENUE_MODEL_LABELS[model],
            "count": payload["revenueModelDistribution"].get(model, 0),
            "revenueShare": payload["revenueShareByRevenueModel"].get(model, 0),
        }
        for model in REVENUE_MODEL_LABELS
    ]
    return "\n".join(
        [
            "# M2 Revenue Model Classification Audit v1",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Boundary: {payload['classificationBoundary']}",
            f"- Total works: {payload['totalWorks']}",
            "",
            markdown_table(rows, [("model", "Model"), ("label", "Chinese label"), ("count", "Count"), ("revenueShare", "Revenue share")]),
            "",
            f"- Manual review required count: {payload['manualReviewRequiredCount']}",
            f"- Estimated buyout amount share: {payload['estimatedBuyoutAmountShare']}",
            f"- Estimated sales-tail amount share: {payload['estimatedSalesTailAmountShare']}",
            "",
            "This report is sanitized and contains no real work titles, authors, channels, raw rows, or per-work revenue details.",
        ]
    )


def rating_standard_md(payload: dict) -> str:
    rows = [{"rating": key, "definition": value} for key, value in payload["ratingDefinitions"].items()]
    return "\n".join(
        [
            "# M2 Rating Standard Definition v1",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Purpose: {payload['standardPurpose']}",
            "",
            markdown_table(rows, [("rating", "Rating"), ("definition", "Definition")]),
            "",
            f"- Expired handling: {payload['expiredHandling']}",
            f"- Buyout handling: {payload['buyoutHandling']}",
            f"- Unknown revenue model handling: {payload['unknownRevenueModelHandling']}",
            "",
            "M3 remains blocked until user/business acceptance of this M2 candidate.",
        ]
    )


def validation_summary_md(payload: dict) -> str:
    feedback = payload["operatorFeedback"]
    gate = payload["suggestionGate"]
    return "\n".join(
        [
            "# M2 Rating Standard v1 Validation Summary",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Operator rows: {feedback['rows']}",
            f"- Reviewable rows: {feedback['reviewableRows']}",
            f"- Rating issue rows: {feedback['ratingIssueRows']}",
            f"- Rating issue hit rows: {feedback['ratingIssueHitRows']}",
            f"- Suggestion issue rows: {feedback['suggestionIssueRows']}",
            f"- Suggestion issue hit rows: {feedback['suggestionIssueHitRows']}",
            f"- Automatic operating suggestion rows: {gate['automaticOperatingSuggestionRows']}",
            f"- Review prompt rows: {gate['reviewPromptRows']}",
            f"- No automatic suggestion rows: {gate['noAutomaticSuggestionRows']}",
            "",
            f"- Private validation workbook: `{payload['privateValidationWorkbook']}`",
            f"- Private operator workbook: `{payload['privateOperatorWorkbook']}`",
            "",
            "Private workbooks are gitignored and must not be committed. M3 remains blocked.",
        ]
    )


def build_revenue_quantiles(rows: list[dict]) -> dict:
    values = np.asarray([row["totalHistoricalRevenue"] for row in rows if row["totalHistoricalRevenue"] > 0], dtype=float)
    if len(values) == 0:
        return {"p40": 0, "p50": 0, "p75": 0, "p90": 0, "p95": 0, "p99": 0}
    return {f"p{int(q * 100)}": round_float(float(np.quantile(values, q)), 2) for q in [0.40, 0.50, 0.75, 0.90, 0.95, 0.99]}


def revenue_bucket_for(amount: float, q: dict) -> str:
    value = safe_float(amount)
    if value >= q["p95"]:
        return "top"
    if value >= q["p75"]:
        return "high"
    if value >= q["p40"]:
        return "medium"
    if value > 0:
        return "low"
    return "zero"


def infer_forecastability_status(summary) -> str:
    output_type = clean(summary.get("forecastOutputType"))
    if output_type == "copyright_term_forecast":
        return "numeric_forecast_eligible"
    if output_type == "operating_window_forecast":
        return "conservative_numeric_forecast"
    if output_type == "observe_only":
        return "observe_only_no_numeric_forecast"
    return clean(summary.get("forecastabilityStatus")) or "unknown"


def infer_forecast_confidence(summary) -> str:
    value = clean(summary.get("forecastConfidence") or summary.get("confidence"))
    if value:
        return value
    rating = clean(summary.get("rating"))
    if rating in {"S+", "S", "A"}:
        return "medium"
    if rating in {"D", "E"}:
        return "low"
    return "medium"


def classify_lifecycle(values: np.ndarray, context: dict) -> str:
    try:
        return classify_at(np.asarray(values, dtype=float), context["parameters"])
    except Exception:
        positive = np.count_nonzero(values > 0)
        if positive < 6:
            return "insufficient_history"
        if values[-12:].sum() <= 0:
            return "inactive"
        return "stable"


def business_form_mix(forms: list[str], summary) -> str:
    if len(set(forms)) > 1:
        return "mixed"
    if forms:
        return forms[0]
    copyright_revenue = safe_float(summary.get("audioCopyrightRevenue"))
    product_revenue = safe_float(summary.get("audioProductRevenue"))
    if copyright_revenue > 0 and product_revenue > 0:
        return "mixed"
    if copyright_revenue > 0:
        return "audio_copyright"
    if product_revenue > 0:
        return "audio_product"
    return "unknown"


def score_buyout(features: dict) -> float:
    score = 0.0
    if features["largestMonthShare"] >= 0.8:
        score += 0.35
    elif features["largestMonthShare"] >= 0.6:
        score += 0.25
    elif features["largestMonthShare"] >= 0.45:
        score += 0.12
    if features["incomeMonthCount"] <= 2:
        score += 0.22
    elif features["incomeMonthCount"] <= 4:
        score += 0.14
    if features["integerAmountRatio"] >= 0.75:
        score += 0.13
    if features["roundAmountRatio"] >= 0.5:
        score += 0.13
    if features["equalSplitSignal"]:
        score += 0.17
    if features["spikeScore"] >= 0.9:
        score += 0.10
    return min(1.0, score)


def score_sales_continuity(features: dict) -> float:
    score = 0.0
    score += features["continuityScore"] * 0.45
    score += min(1.0, features["activeMonthRatio"] * 2.2) * 0.25
    score += features["randomnessScore"] * 0.20
    if features["incomeMonthCount"] >= 12:
        score += 0.10
    elif features["incomeMonthCount"] >= 6:
        score += 0.05
    if features["largestMonthShare"] > 0.85:
        score -= 0.20
    if features["equalSplitSignal"]:
        score -= 0.08
    return max(0.0, min(1.0, score))


def continuity(values: np.ndarray) -> float:
    positive_indexes = np.where(values > 0)[0]
    if len(positive_indexes) <= 1:
        return 0.1 if len(positive_indexes) == 1 else 0.0
    adjacent = sum(1 for idx in range(1, len(positive_indexes)) if positive_indexes[idx] - positive_indexes[idx - 1] <= 1)
    return adjacent / max(1, len(positive_indexes) - 1)


def randomness(values: np.ndarray) -> float:
    positive = values[values > 0]
    if len(positive) < 3:
        return 0.0
    mean = float(np.mean(positive))
    if mean <= 0:
        return 0.0
    cv = float(np.std(positive) / mean)
    if cv < 0.05:
        return 0.1
    if cv > 2.5:
        return 0.45
    return min(1.0, cv / 1.2)


def has_post_buyout_tail(values: np.ndarray, largest: float) -> bool:
    if len(values) == 0 or largest <= 0:
        return False
    peak = int(np.argmax(values))
    if peak >= len(values) - 2:
        return False
    tail = values[peak + 1 :]
    positive_tail = tail[(tail > 0) & (tail < largest * 0.35)]
    return len(positive_tail) >= 3 and continuity(tail) >= 0.35


def top_revenue_cohort_distribution(rows: list[dict]) -> dict:
    sorted_rows = sorted(rows, key=lambda row: row["totalHistoricalRevenue"], reverse=True)
    result = {}
    for label, share in [("top1Percent", 0.01), ("top5Percent", 0.05), ("top10Percent", 0.10)]:
        count = max(1, math.ceil(len(sorted_rows) * share))
        cohort = sorted_rows[:count]
        result[label] = {
            "count": count,
            "modelDistribution": {model: int(Counter(row["revenueModel"] for row in cohort).get(model, 0)) for model in REVENUE_MODEL_LABELS},
        }
    return result


def ratio(values, predicate) -> float:
    if len(values) == 0:
        return 0.0
    return float(sum(1 for value in values if predicate(float(value))) / len(values))


def is_round_amount(value: float) -> bool:
    abs_value = abs(float(value))
    if abs_value < 10:
        return False
    return any(abs(abs_value / base - round(abs_value / base)) <= 0.01 for base in [10, 100, 1000])


def rating_from_score(score: float) -> str:
    if score >= 96:
        return "S+"
    if score >= 88:
        return "S"
    if score >= 70:
        return "A"
    if score >= 52:
        return "B"
    if score >= 38:
        return "C"
    if score >= 24:
        return "D"
    return "E"


def bucket_distribution(values: list[int], thresholds: list[int]) -> dict:
    result = Counter()
    for value in values:
        placed = False
        for threshold in thresholds:
            if value <= threshold:
                result[f"<= {threshold}"] += 1
                placed = True
                break
        if not placed:
            result[f"> {thresholds[-1]}"] += 1
    return dict(result)


def ratio_bucket_distribution(values: list[float]) -> dict:
    result = Counter()
    for value in values:
        if value < 0.2:
            result["0.00-0.20"] += 1
        elif value < 0.4:
            result["0.20-0.40"] += 1
        elif value < 0.6:
            result["0.40-0.60"] += 1
        elif value < 0.8:
            result["0.60-0.80"] += 1
        else:
            result["0.80-1.00"] += 1
    return dict(result)


def anonymized_rows(rows: list[dict]) -> list[dict]:
    output = []
    for index, row in enumerate(rows, start=1):
        output.append(
            {
                "匿名ID": f"W{index:04d}",
                "收入模式": row.get("revenueModelChinese", ""),
                "收入模式置信度": row.get("revenueModelConfidence", ""),
                "历史表现评级": row.get("historicalPerformanceRating", ""),
                "当前版权状态": row.get("currentRightsStatusChinese", ""),
                "预测价值评级": row.get("forecastValueRatingChinese", ""),
                "当前运营决策级别": row.get("operationalDecisionRatingChinese", ""),
                "收入月份数": row.get("incomeMonthCount", 0),
                "最大单月占比": row.get("largestMonthShare", 0),
                "连续性": row.get("continuityScore", 0),
                "整额比例": row.get("roundAmountRatio", 0),
                "同额批次信号": row.get("equalSplitSignal", False),
                "买断估计金额": row.get("buyoutEstimatedAmount", 0),
                "实销尾部估计金额": row.get("salesTailEstimatedAmount", 0),
            }
        )
    return output


def add_summary_sheet(wb: Workbook, title: str, payload: dict) -> None:
    ws = wb.create_sheet(title[:31])
    rows = flatten_payload(payload)
    append_rows(ws, [["指标", "值"], *rows])


def add_dict_rows_sheet(wb: Workbook, title: str, rows: list[dict], headers: list[str]) -> None:
    ws = wb.create_sheet(title[:31])
    append_rows(ws, [headers])
    for row in rows:
        append_rows(ws, [[row.get(header, "") for header in headers]])


def flatten_payload(payload: dict, prefix: str = "") -> list[list[str]]:
    rows = []
    for key, value in payload.items():
        label = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten_payload(value, label))
        elif isinstance(value, list):
            rows.append([label, json.dumps(value, ensure_ascii=False)])
        else:
            rows.append([label, value])
    return rows


def append_rows(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)


def finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column_cells in enumerate(ws.columns, start=1):
            max_len = min(48, max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2)
            ws.column_dimensions[get_column_letter(index)].width = max(12, max_len)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def public_envelope(report_id: str, payload: dict) -> dict:
    return {
        "reportId": report_id,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "payload": json_safe(payload),
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    lines = ["| " + " | ".join(label for _, label in columns) + " |"]
    lines.append("|" + "|".join("---" for _ in columns) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |")
    return "\n".join(lines)


def find_value(row: dict, preferred_headers: list[str]) -> str:
    for header in preferred_headers:
        if header in row and clean(row.get(header)):
            return clean(row.get(header))
    for key, value in row.items():
        text_key = clean(key)
        if any(header in text_key for header in preferred_headers) and clean(value):
            return clean(value)
    return ""


def find_header_value_contains(row: dict, needles: list[str]) -> str:
    for key, value in row.items():
        key_text = clean(key)
        if all(needle in key_text for needle in needles):
            return clean(value)
    return ""


def contains_any(value: str, needles: list[str]) -> bool:
    text = clean(value)
    return any(needle in text for needle in needles)


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def safe_float(value, default: float | None = 0.0) -> float | None:
    try:
        result = float(value)
    except Exception:
        return default
    if math.isfinite(result):
        return result
    return default


def safe_display_number(value) -> str:
    if value is None or pd.isna(value):
        return "未知"
    return str(round_float(float(value), 2))


def round_float(value, digits: int = 4):
    if value is None:
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if not math.isfinite(number):
        return None
    return round(number, digits)


def json_safe(value):
    if isinstance(value, dict):
        return {str(key): json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [json_safe(child) for child in value]
    if isinstance(value, tuple):
        return [json_safe(child) for child in value]
    if isinstance(value, set):
        return sorted(json_safe(child) for child in value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


if __name__ == "__main__":
    main()
