from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

CODEX_BUNDLED_PYTHON_PACKAGES = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "Lib"
    / "site-packages"
)
CODEX_BUNDLED_PYTHON_EXE = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "python.exe"
)
if CODEX_BUNDLED_PYTHON_EXE.exists() and Path(sys.executable).resolve() != CODEX_BUNDLED_PYTHON_EXE.resolve():
    os.execv(str(CODEX_BUNDLED_PYTHON_EXE), [str(CODEX_BUNDLED_PYTHON_EXE), *sys.argv])

if CODEX_BUNDLED_PYTHON_PACKAGES.exists():
    sys.path.insert(0, str(CODEX_BUNDLED_PYTHON_PACKAGES))

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools" / "m2-calibration"))

from run_nonformal_dry_run import load_analysis_inputs  # noqa: E402

OUTPUT_M1 = ROOT / "docs" / "analysis" / "m1-master-data"
OUTPUT_M2 = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m1-master-data"
PRIVATE_JSON = PRIVATE_DIR / "M1-ledger-masterdata-backfill-candidates.json"
PRIVATE_XLSX = PRIVATE_DIR / "M1-ledger-masterdata-backfill-candidates.xlsx"

STRUCTURE_MD = OUTPUT_M1 / "M1-digital-copyright-ledger-structure-audit-v1.md"
STRUCTURE_JSON = OUTPUT_M1 / "M1-digital-copyright-ledger-structure-audit-v1.json"
GAP_MD = OUTPUT_M1 / "M1-master-data-gap-ledger-backfill-audit-v1.md"
GAP_JSON = OUTPUT_M1 / "M1-master-data-gap-ledger-backfill-audit-v1.json"
SUMMARY_MD = OUTPUT_M1 / "M1-digital-copyright-ledger-masterdata-backfill-summary-v1.md"
SUMMARY_JSON = OUTPUT_M1 / "M1-digital-copyright-ledger-masterdata-backfill-summary-v1.json"
CATEGORY_MD = OUTPUT_M1 / "M1-ledger-category-author-rights-backfill-summary-v1.md"
CATEGORY_JSON = OUTPUT_M1 / "M1-ledger-category-author-rights-backfill-summary-v1.json"
IMPACT_MD = OUTPUT_M2 / "M2-ledger-backfill-impact-on-evaluation-v1.md"
IMPACT_JSON = OUTPUT_M2 / "M2-ledger-backfill-impact-on-evaluation-v1.json"

KEY_FIELDS = [
    "作品ID",
    "出版书名",
    "合同书名",
    "作者署名",
    "作者原名",
    "授权方",
    "签订日期",
    "到期时间",
    "续约前到期日期",
    "CIP出版时间",
    "书号",
    "首发时间",
    "出版社",
    "产品线",
    "CIP",
    "合同编号",
    "合同类型",
    "有声使用权",
    "有声改编权",
    "有声转授权",
    "有声权利描述",
    "信息网络传播",
    "转授权",
    "独家",
    "授权范围（中国大陆地区和中国大陆地区（除港澳台）都只限于中国大陆地区使用）",
    "是否外版",
    "合作方式",
    "电子版税",
    "有声版税",
    "电子预付",
    "有声预付",
]

FIELD_USAGE = {
    "作品ID": ["M1 标准作品身份", "M1/M2 匹配"],
    "出版书名": ["M1 标准作品基础信息", "标题匹配"],
    "合同书名": ["M1 标准作品基础信息", "标题匹配"],
    "作者署名": ["M1 作者信息", "标题作者匹配"],
    "作者原名": ["M1 作者信息", "标题作者匹配"],
    "授权方": ["M1 权利来源辅助"],
    "签订日期": ["M1 版权起始", "M2 renewal review"],
    "到期时间": ["M1 版权到期", "M2 forecast/rating/risk"],
    "续约前到期日期": ["M1 版权到期冲突复核", "M2 renewal review"],
    "CIP出版时间": ["M1 出版/首发日期", "分类辅助"],
    "书号": ["M1 身份辅助", "匹配辅助"],
    "首发时间": ["M1 出版/首发日期"],
    "出版社": ["M1 基础信息", "分类辅助"],
    "产品线": ["M1 分类候选"],
    "CIP": ["M1 分类候选"],
    "合同编号": ["匹配辅助"],
    "合同类型": ["M1 权利/业务形态辅助", "M1 分类候选"],
    "有声使用权": ["M1 有声权利/业务形态", "M2 risk"],
    "有声改编权": ["M1 有声权利/业务形态", "M2 risk"],
    "有声转授权": ["M1 有声权利/业务形态", "M2 risk"],
    "有声权利描述": ["M1 有声权利/业务形态", "M2 risk"],
    "信息网络传播": ["M1 权利/业务形态"],
    "转授权": ["M1 权利/业务形态"],
    "独家": ["M1 权利/业务形态", "M2 risk"],
    "授权范围（中国大陆地区和中国大陆地区（除港澳台）都只限于中国大陆地区使用）": ["M1 权利/业务形态"],
    "是否外版": ["M1 分类候选"],
    "合作方式": ["M1 权利/业务形态"],
    "电子版税": ["M1 权利/版税辅助"],
    "有声版税": ["M1 权利/版税辅助"],
    "电子预付": ["M1 权利/版税辅助"],
    "有声预付": ["M1 权利/版税辅助"],
}

GAP_RULES = {
    "standardWorkName": ["作品名称"],
    "authorName": ["作者"],
    "copyrightStartDate": ["版权开始日期"],
    "copyrightEndDate": ["版权到期日期"],
    "publisherName": ["出版社"],
    "classificationLevel1": ["一级至三级分类", "一级分类"],
    "classificationLevel2": ["一级至三级分类", "二级分类"],
    "classificationLevel3": ["一级至三级分类", "三级分类"],
    "requiredTags": ["必需标签", "标签"],
    "audioRightsStatus": ["有声权利"],
    "firstPublicationDate": ["出版", "首发"],
}

FIELD_SPECS = [
    ("standardWorkName", ["出版书名", "合同书名"]),
    ("authorName", ["作者署名", "作者原名"]),
    ("copyrightStartDate", ["签订日期"]),
    ("copyrightEndDate", ["到期时间", "续约前到期日期"]),
    ("publisherName", ["出版社"]),
    ("classificationLevel1", ["产品线", "CIP", "是否外版"]),
    ("classificationLevel2", ["产品线", "CIP"]),
    ("classificationLevel3", ["产品线", "CIP"]),
    ("audioRightsStatus", ["有声使用权", "有声改编权", "有声转授权", "有声权利描述"]),
    ("firstPublicationDate", ["首发时间", "CIP出版时间"]),
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--scope", choices=["all", "m1", "m2"], default="all")
    args = parser.parse_args()

    context = build_context()
    payload = build_payload(context)
    write_outputs(payload)
    print(
        json.dumps(
            {
                "scope": args.scope,
                "ledgerRows": payload["ledgerStructure"]["totalDataRows"],
                "matchedRows": payload["matchingSummary"]["matchedLedgerRows"],
                "candidateCount": payload["candidateSummary"]["total"],
                "privateJson": str(PRIVATE_JSON.relative_to(ROOT)),
                "privateXlsxExpected": str(PRIVATE_XLSX.relative_to(ROOT)),
            },
            ensure_ascii=False,
        )
    )


def build_context() -> dict:
    ledger_path = locate_ledger()
    ledger = read_ledger(ledger_path)
    analysis = load_analysis_inputs()
    work_summary = dataframe_records(analysis["work_summary"])
    bill = dataframe_records(analysis["bill"])
    mapping = load_mapping()
    ops_basic = load_ops_basic_info()
    return {
        "ledgerPath": ledger_path,
        "ledger": ledger,
        "workSummary": work_summary,
        "bill": bill,
        "mapping": mapping,
        "opsBasic": ops_basic,
        "currentHead": run_git(["rev-parse", "HEAD"]),
        "originMain": run_git(["rev-parse", "origin/main"]),
    }


def locate_ledger() -> Path:
    candidates = list((ROOT / "data" / "master-data").glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No ledger workbook was found under the private master-data directory.")
    candidates.sort(key=lambda path: path.stat().st_size, reverse=True)
    return candidates[0]


def read_ledger(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    all_rows = []
    for ws in wb.worksheets:
        iterator = ws.iter_rows(values_only=True)
        header = [clean_cell(value) for value in next(iterator)]
        rows = []
        for index, raw in enumerate(iterator, start=2):
            row = {header[col]: clean_cell(value) for col, value in enumerate(raw) if col < len(header) and header[col]}
            if any(has_value(value) for value in row.values()):
                row["_ledgerRowNumber"] = index
                row["_ledgerRowId"] = f"L{index:06d}"
                rows.append(row)
                all_rows.append(row)
        sheets.append(
            {
                "sheetName": ws.title,
                "totalRows": len(rows),
                "fieldCount": len([field for field in header if field]),
                "fields": [field for field in header if field],
            }
        )
    return {"sheets": sheets, "rows": all_rows}


def load_mapping() -> dict:
    mapping_path = ROOT / "data" / "m1-master-data-private" / "mapping-candidate" / "M1-formal-mapping-version-candidate-v0.1.json"
    if not mapping_path.exists():
        return {"rawToStandard": {}, "effectiveRows": []}
    data = json.loads(mapping_path.read_text(encoding="utf-8"))
    effective = data.get("mappings", {}).get("effective_mapping_snapshot", [])
    raw_to_standard = {}
    for row in effective:
        raw = normalize_id(row.get("raw_work_id"))
        target = normalize_id(row.get("target_standard_work_id"))
        if raw and target:
            raw_to_standard[raw] = target
    return {"rawToStandard": raw_to_standard, "effectiveRows": effective}


def load_ops_basic_info() -> dict:
    path = ROOT / "data" / "m1-master-data-private" / "ops-confirmation" / "ops-confirmation-v2.3-data.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for sheet in data.get("sheets", []):
        if sheet.get("kind") == "basic_info":
            rows = sheet.get("rows", [])
            break
    return {normalize_id(row.get("标准作品ID")): row for row in rows if normalize_id(row.get("标准作品ID"))}


def build_payload(context: dict) -> dict:
    work_index = build_work_index(context)
    parsed_rows = []
    for row in context["ledger"]["rows"]:
        parsed = parse_ledger_row(row)
        match = match_ledger_row(parsed, work_index)
        parsed_rows.append({**parsed, "match": match})

    candidate_seed = []
    for parsed in parsed_rows:
        if not parsed["match"]["matchedStandardWorkId"]:
            continue
        candidate_seed.extend(build_candidates_for_row(parsed, context, work_index))
    candidates = apply_conflicts(candidate_seed)

    ledger_structure = build_structure_audit(context["ledger"])
    matching_summary = build_matching_summary(parsed_rows, work_index)
    gap_summary = build_gap_summary(context, candidates, work_index)
    category_summary = build_category_author_rights_summary(context, candidates, parsed_rows)
    impact_summary = build_m2_impact(context, candidates, work_index)

    private_payload = {
        "schema": "m1.m2.digital_copyright_ledger.masterdata_backfill_candidates.v1",
        "generatedAt": utc_now(),
        "currentHead": context["currentHead"],
        "originMain": context["originMain"],
        "privateWorkbookExpected": str(PRIVATE_XLSX.relative_to(ROOT)),
        "safetyBoundary": {
            "containsRealWorkIds": True,
            "containsRealTitlesAuthorsLedgerSnippets": True,
            "gitignoredPrivateOutput": True,
            "databaseConnected": False,
            "dockerExecuted": False,
            "formalMasterDataWritten": False,
            "m3Entered": False,
        },
        "ledgerStructure": ledger_structure,
        "matchingSummary": matching_summary,
        "candidateSummary": summarize_candidates(candidates),
        "gapSummary": gap_summary,
        "categoryAuthorRightsSummary": category_summary,
        "m2ImpactSummary": impact_summary,
        "candidates": candidates,
        "conflicts": build_private_conflicts(candidates),
        "manualReviewItems": [item for item in candidates if item["requiresManualReview"]],
    }

    return {
        "schema": "m1.m2.digital_copyright_ledger.semantic_backfill_audit.v1",
        "generatedAt": utc_now(),
        "currentHead": context["currentHead"],
        "originMain": context["originMain"],
        "ledgerStructure": ledger_structure,
        "matchingSummary": matching_summary,
        "candidateSummary": summarize_candidates(candidates),
        "gapSummary": gap_summary,
        "categoryAuthorRightsSummary": category_summary,
        "m2ImpactSummary": impact_summary,
        "privatePayload": private_payload,
        "sanitizedBoundary": sanitized_boundary(),
    }


def build_work_index(context: dict) -> dict:
    standard_ids = {normalize_id(row.get("standardWorkId")) for row in context["workSummary"]}
    standard_ids.discard("")

    raw_to_standard = dict(context["mapping"]["rawToStandard"])
    bill_names_by_standard = defaultdict(set)
    raw_ids_by_standard = defaultdict(set)
    for row in context["bill"]:
        standard = normalize_id(row.get("standardWorkId"))
        raw = normalize_id(row.get("rawWorkId"))
        if standard and raw:
            raw_to_standard.setdefault(raw, standard)
            raw_ids_by_standard[standard].add(raw)
        title = normalize_title(row.get("作品名称"))
        if standard and title:
            bill_names_by_standard[standard].add(title)

    title_to_standard = defaultdict(set)
    author_by_standard = defaultdict(set)
    current_by_standard = {}
    for standard in sorted(standard_ids, key=id_sort_key):
        ops = context["opsBasic"].get(standard, {})
        current_by_standard[standard] = current_values_from_ops(ops)
        for field in ["账单主要名称", "台账标准名称候选", "运营确认标准作品名称"]:
            title = normalize_title(ops.get(field))
            if title:
                bill_names_by_standard[standard].add(title)
        for field in ["作者候选", "运营确认作者"]:
            for author in normalize_author_tokens(ops.get(field)):
                author_by_standard[standard].add(author)
        for title in sorted(bill_names_by_standard[standard]):
            title_to_standard[title].add(standard)

    revenue_sorted = sorted(
        context["workSummary"],
        key=lambda item: (-safe_float(item.get("totalHistoricalRevenue")), id_sort_key(normalize_id(item.get("standardWorkId")))),
    )
    revenue_rank = {normalize_id(row.get("standardWorkId")): index + 1 for index, row in enumerate(revenue_sorted)}
    total = len(revenue_sorted) or 1
    revenue_bucket = {}
    for standard, rank in revenue_rank.items():
        pct = rank / total
        if pct <= 0.01:
            bucket = "top_1_percent"
        elif pct <= 0.05:
            bucket = "top_5_percent"
        elif pct <= 0.10:
            bucket = "top_10_percent"
        else:
            bucket = "other"
        revenue_bucket[standard] = bucket

    return {
        "standardIds": standard_ids,
        "rawToStandard": raw_to_standard,
        "rawIdsByStandard": raw_ids_by_standard,
        "titleToStandard": title_to_standard,
        "titlesByStandard": bill_names_by_standard,
        "authorByStandard": author_by_standard,
        "currentByStandard": current_by_standard,
        "revenueBucketByStandard": revenue_bucket,
        "workSummaryByStandard": {normalize_id(row.get("standardWorkId")): row for row in context["workSummary"]},
    }


def current_values_from_ops(row: dict) -> dict:
    return {
        "standardWorkName": row.get("运营确认标准作品名称"),
        "authorName": row.get("运营确认作者"),
        "copyrightStartDate": row.get("运营确认版权开始日期"),
        "copyrightEndDate": row.get("运营确认版权到期日期"),
        "classificationLevel1": row.get("运营确认一级分类"),
        "classificationLevel2": row.get("运营确认二级分类"),
        "classificationLevel3": row.get("运营确认三级分类"),
        "requiredTags": row.get("运营确认必需标签"),
        "needFields": row.get("需要补充的字段", ""),
    }


def parse_ledger_row(row: dict) -> dict:
    title = first_value(row.get("出版书名"), row.get("合同书名"))
    author = first_value(row.get("作者署名"), row.get("作者原名"))
    signed = parse_date(row.get("签订日期"))
    expiry = parse_date(first_value(row.get("到期时间"), row.get("续约前到期日期")))
    first_publication = parse_date(first_value(row.get("首发时间"), row.get("CIP出版时间")))
    rights = parse_audio_rights(row)
    category = parse_category(row)
    return {
        "ledgerRowId": row["_ledgerRowId"],
        "ledgerRowNumber": row["_ledgerRowNumber"],
        "workId": normalize_id(row.get("作品ID")),
        "rawWorkId": stringify(row.get("作品ID")),
        "title": stringify(title),
        "titleNormalized": normalize_title(title),
        "contractTitleNormalized": normalize_title(row.get("合同书名")),
        "author": stringify(author),
        "authorTokens": normalize_author_tokens(author),
        "publisher": stringify(row.get("出版社")),
        "productLine": stringify(row.get("产品线")),
        "cipPresent": bool(stringify(row.get("CIP"))),
        "isbnPresent": bool(stringify(row.get("书号"))),
        "contractPresent": bool(first_value(row.get("合同编号"), row.get("合同最终码"))),
        "signedDate": signed,
        "expiryDate": expiry,
        "firstPublicationDate": first_publication,
        "rights": rights,
        "category": category,
        "raw": row,
    }


def match_ledger_row(parsed: dict, work_index: dict) -> dict:
    work_id = parsed["workId"]
    if work_id and work_id in work_index["standardIds"]:
        return match_result("matched", "exact_work_id", "high", work_id, [work_id], parsed, "台账作品ID等于标准作品ID")

    if work_id and work_id in work_index["rawToStandard"]:
        standard = work_index["rawToStandard"][work_id]
        method = "mapping_work_id" if work_id != standard else "exact_work_id"
        return match_result("matched", method, "high", standard, [work_id], parsed, "台账作品ID通过当前 raw/映射关系命中标准作品")

    title = parsed["titleNormalized"]
    authors = set(parsed["authorTokens"])
    if title:
        standards = work_index["titleToStandard"].get(title, set())
        if len(standards) == 1:
            standard = single_standard(standards)
            known_authors = work_index["authorByStandard"].get(standard, set())
            if known_authors and authors and authors.intersection(known_authors):
                return match_result("matched", "title_author_exact", "high", standard, [], parsed, "标题和作者规范化后均精确命中")
            return match_result("matched", "title_author_exact", "medium", standard, [], parsed, "标题精确命中，当前作者主数据不足，需人工确认作者")
        if len(standards) > 1:
            return match_result("conflict", "title_author_exact", "low", None, [], parsed, "标题命中多个标准作品")

        fuzzy = fuzzy_title_match(title, work_index["titleToStandard"])
        if fuzzy:
            standard, score = fuzzy
            return match_result(
                "matched",
                "title_author_fuzzy",
                "medium" if score >= 0.90 else "low",
                standard,
                [],
                parsed,
                f"标题模糊相似度 {score:.2f}，作者需人工复核",
            )

    return match_result("unmatched", "unmatched", "missing", None, [], parsed, "未通过 ID、映射或标题策略命中")


def match_result(status, method, confidence, standard, raw_ids, parsed, reason):
    return {
        "matchStatus": status,
        "matchMethod": method,
        "matchConfidence": confidence,
        "matchedStandardWorkId": standard,
        "matchedRawWorkIds": raw_ids,
        "matchedLedgerRowCount": 1,
        "selectedLedgerRowReason": reason,
        "conflictCount": 1 if status == "conflict" else 0,
        "conflictReason": reason if status == "conflict" else None,
        "requiresManualReview": confidence != "high" or status != "matched",
        "isbnOrCipAssisted": parsed["isbnPresent"] or parsed["cipPresent"],
        "contractAssisted": parsed["contractPresent"],
    }


def fuzzy_title_match(title: str, title_to_standard: dict) -> tuple[str, float] | None:
    if len(title) < 4:
        return None
    title_bigrams = bigrams(title)
    best = (None, 0.0)
    for candidate_title, standards in sorted(title_to_standard.items()):
        if not candidate_title or candidate_title[0] != title[0]:
            continue
        length_ratio = min(len(title), len(candidate_title)) / max(len(title), len(candidate_title))
        if length_ratio < 0.70:
            continue
        jaccard = overlap_ratio(title_bigrams, bigrams(candidate_title))
        if jaccard < 0.65:
            continue
        score = SequenceMatcher(None, title, candidate_title).ratio()
        if score > best[1] and len(standards) == 1:
            best = (single_standard(standards), score)
    return best if best[0] and best[1] >= 0.86 else None


def build_candidates_for_row(parsed: dict, context: dict, work_index: dict) -> list[dict]:
    standard = parsed["match"]["matchedStandardWorkId"]
    current = work_index["currentByStandard"].get(standard, {})
    candidates = []
    for field, source_fields in FIELD_SPECS:
        proposal = propose_value(field, parsed)
        if proposal["value"] in (None, ""):
            continue
        if not field_is_gap(field, current, context["opsBasic"].get(standard, {})):
            continue
        parser_status = proposal["parserStatus"]
        match_confidence = parsed["match"]["matchConfidence"]
        confidence = candidate_confidence(match_confidence, parser_status, field)
        candidates.append(
            {
                "standardWorkId": standard,
                "rawWorkId": parsed["rawWorkId"],
                "ledgerRowIds": [parsed["ledgerRowId"]],
                "fieldName": field,
                "currentValue": current.get(field),
                "proposedValue": proposal["value"],
                "proposedValueNormalized": proposal["normalized"],
                "sourceField": proposal["sourceField"] or " / ".join(source_fields),
                "sourceRawValue": proposal["rawValue"],
                "parserStatus": parser_status,
                "matchMethod": parsed["match"]["matchMethod"],
                "matchConfidence": match_confidence,
                "valueConfidence": confidence,
                "conflictStatus": "none",
                "requiresManualReview": confidence != "high" or proposal.get("requiresManualReview", False),
                "autoApplyEligible": confidence == "high" and not proposal.get("requiresManualReview", False),
                "reason": proposal["reason"],
                "auditMetadata": {
                    "ledgerRowNumber": parsed["ledgerRowNumber"],
                    "selectedLedgerRowReason": parsed["match"]["selectedLedgerRowReason"],
                    "isbnOrCipAssisted": parsed["match"]["isbnOrCipAssisted"],
                    "contractAssisted": parsed["match"]["contractAssisted"],
                    "notFormalMasterData": True,
                },
            }
        )
    return candidates


def propose_value(field: str, parsed: dict) -> dict:
    if field == "standardWorkName":
        return proposal(parsed["title"], parsed["titleNormalized"], "出版书名/合同书名", parsed["title"], "标题字段可补标准作品名称候选")
    if field == "authorName":
        return proposal(parsed["author"], "|".join(parsed["authorTokens"]), "作者署名/作者原名", parsed["author"], "作者字段可补作者候选")
    if field == "copyrightStartDate":
        return date_proposal(parsed["signedDate"], "签订日期", "签订日期可作为版权开始候选")
    if field == "copyrightEndDate":
        return date_proposal(parsed["expiryDate"], "到期时间/续约前到期日期", "到期字段可作为版权到期候选")
    if field == "publisherName":
        return proposal(parsed["publisher"], normalize_title(parsed["publisher"]), "出版社", parsed["publisher"], "出版社可补基础信息候选")
    if field == "classificationLevel1":
        return proposal("出版物", "出版物", "产品线/CIP", parsed["productLine"], "台账产品线仅支持一级分类候选，需确认")
    if field == "classificationLevel2":
        value = parsed["category"].get("level2")
        return proposal(value, normalize_title(value), "产品线", value, "产品线可作为二级分类候选，需映射确认")
    if field == "classificationLevel3":
        value = parsed["category"].get("level3")
        return proposal(value, normalize_title(value), "CIP", value, "CIP 只能辅助三级分类候选，不能自动造三级分类", requires_review=True)
    if field == "audioRightsStatus":
        value = parsed["rights"]["audioRightsStatus"]
        return proposal(value, value, "有声权利字段", parsed["rights"]["sourceSummary"], "有声权利字段可补权利状态候选", parsed["rights"]["requiresManualReview"])
    if field == "firstPublicationDate":
        return date_proposal(parsed["firstPublicationDate"], "首发时间/CIP出版时间", "首发或 CIP 出版时间可补出版日期候选")
    return proposal(None, None, None, None, "unsupported")


def proposal(value, normalized, source, raw, reason, requires_review=False):
    return {
        "value": value,
        "normalized": normalized,
        "sourceField": source,
        "rawValue": raw,
        "parserStatus": "parsed" if value not in (None, "") else "missing",
        "requiresManualReview": requires_review,
        "reason": reason,
    }


def date_proposal(parsed_date: dict, source: str, reason: str):
    if parsed_date["parserStatus"] in {"parsed", "parsed_with_condition"} and parsed_date.get("normalizedDate"):
        return proposal(parsed_date["normalizedDate"], parsed_date["normalizedDate"], source, parsed_date["rawValue"], reason, parsed_date["parserStatus"] != "parsed")
    if parsed_date["expiryType"] == "infinite":
        return proposal("无限期", "infinite", source, parsed_date["rawValue"], "字段表示无限期，需业务确认后才可进入正式主数据", True)
    if parsed_date["expiryType"] == "relative_term":
        normalized = f"{parsed_date['anchor']}+{parsed_date['years']}y"
        return proposal(normalized, normalized, source, parsed_date["rawValue"], "相对期限需要锚点日期，不可直接自动补全", True)
    return proposal(None, None, source, parsed_date.get("rawValue"), "日期缺失或无法解析", True)


def field_is_gap(field: str, current: dict, ops_row: dict) -> bool:
    if not current.get(field):
        return True
    needed = str(ops_row.get("需要补充的字段", ""))
    return any(token in needed for token in GAP_RULES.get(field, []))


def candidate_confidence(match_confidence: str, parser_status: str, field: str) -> str:
    if field in {"classificationLevel1", "classificationLevel2", "classificationLevel3"}:
        return "medium" if match_confidence == "high" else "low"
    if match_confidence == "high" and parser_status == "parsed":
        return "high"
    if match_confidence in {"high", "medium"} and parser_status in {"parsed", "parsed_with_condition"}:
        return "medium"
    return "low"


def apply_conflicts(candidates: list[dict]) -> list[dict]:
    grouped = defaultdict(set)
    for item in candidates:
        key = (item["standardWorkId"], item["fieldName"])
        grouped[key].add(str(item["proposedValueNormalized"]))
    conflict_keys = {key for key, values in grouped.items() if len(values) > 1}
    for item in candidates:
        key = (item["standardWorkId"], item["fieldName"])
        if key in conflict_keys:
            item["conflictStatus"] = "conflict"
            item["valueConfidence"] = "low"
            item["autoApplyEligible"] = False
            item["requiresManualReview"] = True
            item["reason"] = f"{item['reason']}；同一标准作品同字段存在多个候选值"
    return candidates


def build_structure_audit(ledger: dict) -> dict:
    rows = ledger["rows"]
    sheets = []
    for sheet in ledger["sheets"]:
        field_stats = []
        for field in sheet["fields"]:
            nonempty = sum(1 for row in rows if has_value(row.get(field)))
            field_stats.append(
                {
                    "fieldName": field,
                    "nonEmptyCount": nonempty,
                    "nonEmptyRate": ratio(nonempty, len(rows)),
                    "usage": FIELD_USAGE.get(field, ["辅助字段"]),
                }
            )
        sheets.append({**sheet, "fieldStats": field_stats})
    key_field_coverage = []
    fields_present = set(sheets[0]["fields"] if sheets else [])
    for field in KEY_FIELDS:
        nonempty = sum(1 for row in rows if has_value(row.get(field)))
        key_field_coverage.append(
            {
                "fieldName": field,
                "present": field in fields_present,
                "nonEmptyCount": nonempty,
                "nonEmptyRate": ratio(nonempty, len(rows)),
                "usage": FIELD_USAGE.get(field, ["辅助字段"]),
            }
        )
    return {
        "sheetCount": len(sheets),
        "totalDataRows": len(rows),
        "sheets": sheets,
        "keyFieldCoverage": key_field_coverage,
        "sanitized": True,
    }


def build_matching_summary(parsed_rows: list[dict], work_index: dict) -> dict:
    methods = Counter(row["match"]["matchMethod"] for row in parsed_rows)
    statuses = Counter(row["match"]["matchStatus"] for row in parsed_rows)
    matched_standards = {row["match"]["matchedStandardWorkId"] for row in parsed_rows if row["match"]["matchedStandardWorkId"]}
    revenue = Counter(work_index["revenueBucketByStandard"].get(standard, "other") for standard in matched_standards)
    conflict_rows = sum(1 for row in parsed_rows if row["match"]["matchStatus"] == "conflict")
    return {
        "ledgerRows": len(parsed_rows),
        "matchedLedgerRows": sum(1 for row in parsed_rows if row["match"]["matchedStandardWorkId"]),
        "matchedStandardWorkCount": len(matched_standards),
        "unmatchedLedgerRows": statuses.get("unmatched", 0),
        "conflictLedgerRows": conflict_rows,
        "matchMethodDistribution": dict(methods),
        "matchStatusDistribution": dict(statuses),
        "revenueCoverageDistribution": dict(revenue),
        "isbnOrCipAssistedRows": sum(1 for row in parsed_rows if row["match"]["isbnOrCipAssisted"]),
        "contractAssistedRows": sum(1 for row in parsed_rows if row["match"]["contractAssisted"]),
    }


def build_gap_summary(context: dict, candidates: list[dict], work_index: dict) -> dict:
    ops_rows = list(context["opsBasic"].values())
    total_work_count = len(work_index["standardIds"])
    gap_counts = {
        "missingWorkName": sum(1 for row in ops_rows if not has_value(row.get("运营确认标准作品名称"))),
        "missingAuthor": count_needed(ops_rows, "作者"),
        "missingCopyrightStart": count_needed(ops_rows, "版权开始日期"),
        "missingCopyrightEnd": count_needed(ops_rows, "版权到期日期"),
        "missingPublisher": sum(1 for row in ops_rows if not has_value(row.get("出版社"))),
        "missingClassification1": count_needed(ops_rows, "一级至三级分类"),
        "missingClassification2": count_needed(ops_rows, "一级至三级分类"),
        "missingClassification3": count_needed(ops_rows, "一级至三级分类"),
        "missingRequiredTags": count_needed(ops_rows, "必需标签"),
        "missingAudioRights": total_work_count,
        "missingFirstPublicationDate": sum(1 for row in ops_rows if not has_value(row.get("首发时间"))),
    }
    candidates_by_field = defaultdict(list)
    for item in candidates:
        candidates_by_field[item["fieldName"]].append(item)
    field_backfill = {}
    for field, items in candidates_by_field.items():
        field_backfill[field] = {
            "highConfidenceAuto": sum(1 for item in items if item["valueConfidence"] == "high" and item["autoApplyEligible"]),
            "mediumConfidenceSuggested": sum(1 for item in items if item["valueConfidence"] == "medium"),
            "lowConfidenceManual": sum(1 for item in items if item["valueConfidence"] == "low"),
            "conflict": sum(1 for item in items if item["conflictStatus"] != "none"),
            "candidateRows": len(items),
        }
    high_income = summarize_high_income_gaps(candidates, work_index)
    return {
        "totalStandardWorks": total_work_count,
        "currentGapCounts": gap_counts,
        "fieldBackfillPotential": field_backfill,
        "highIncomeImpact": high_income,
        "ledgerAlsoMissingRows": count_ledger_missing(context["ledger"]["rows"]),
    }


def build_category_author_rights_summary(context: dict, candidates: list[dict], parsed_rows: list[dict]) -> dict:
    by_field = defaultdict(list)
    for item in candidates:
        by_field[item["fieldName"]].append(item)
    return {
        "authorBackfill": summarize_field(by_field["authorName"]),
        "classificationBackfill": {
            "level1": summarize_field(by_field["classificationLevel1"]),
            "level2": summarize_field(by_field["classificationLevel2"]),
            "level3": summarize_field(by_field["classificationLevel3"]),
            "productLineCoveredRows": sum(1 for row in parsed_rows if row["productLine"]),
            "cipCoveredRows": sum(1 for row in parsed_rows if row["cipPresent"]),
            "threeLevelAutoCompletionAllowed": False,
            "reason": "产品线和 CIP 只能作为分类候选，不能伪造三级分类。",
        },
        "audioRightsBackfill": summarize_field(by_field["audioRightsStatus"]),
        "rightsDistribution": dict(Counter(row["rights"]["audioRightsStatus"] for row in parsed_rows)),
        "multiAuthorRiskRows": sum(1 for row in parsed_rows if len(row["authorTokens"]) > 1),
    }


def build_m2_impact(context: dict, candidates: list[dict], work_index: dict) -> dict:
    by_field = defaultdict(list)
    for item in candidates:
        by_field[item["fieldName"]].append(item)
    high_end = [item for item in by_field["copyrightEndDate"] if item["valueConfidence"] == "high" and item["autoApplyEligible"]]
    medium_end = [item for item in by_field["copyrightEndDate"] if item["valueConfidence"] == "medium"]
    relative_end = [item for item in by_field["copyrightEndDate"] if item["parserStatus"] != "parsed"]
    high_start = [item for item in by_field["copyrightStartDate"] if item["valueConfidence"] == "high" and item["autoApplyEligible"]]
    impacted_standards = {item["standardWorkId"] for item in high_end + medium_end}
    return {
        "copyrightTermForecastCanIncrease": len({item["standardWorkId"] for item in high_end}),
        "operatingWindowForecastPendingExpiryCanReduce": len({item["standardWorkId"] for item in high_end}),
        "relativeExpiryPendingAnchorCanReduce": len({item["standardWorkId"] for item in relative_end if item["valueConfidence"] in {"high", "medium"}}),
        "manualReviewCanReduceByHighConfidenceCandidates": len({item["standardWorkId"] for item in high_end + high_start}),
        "renewalReviewReliabilityCanImprove": len(impacted_standards),
        "ratingRemainingCopyrightAdjustmentCanImprove": len({item["standardWorkId"] for item in high_end}),
        "requiresRerunForecastOutputType": bool(high_end),
        "requiresRerun30WorkOperatorPack": bool(high_end),
        "requiresRerunBusinessReviewSampleSelection": bool(high_end),
        "affectsLimitedBusinessReviewBaseline": bool(high_end or medium_end),
        "formalCompleteAllowed": False,
        "notM3": True,
    }


def summarize_high_income_gaps(candidates: list[dict], work_index: dict) -> dict:
    rows = []
    for bucket in ["top_1_percent", "top_5_percent", "top_10_percent"]:
        standards = {standard for standard, value in work_index["revenueBucketByStandard"].items() if value == bucket}
        bucket_candidates = [item for item in candidates if item["standardWorkId"] in standards]
        rows.append(
            {
                "bucket": bucket,
                "standardWorkCount": len(standards),
                "candidateRows": len(bucket_candidates),
                "highConfidenceAuto": sum(1 for item in bucket_candidates if item["valueConfidence"] == "high" and item["autoApplyEligible"]),
                "manualReview": sum(1 for item in bucket_candidates if item["requiresManualReview"]),
            }
        )
    return rows


def count_ledger_missing(rows: list[dict]) -> dict:
    return {
        "missingTitle": sum(1 for row in rows if not has_value(first_value(row.get("出版书名"), row.get("合同书名")))),
        "missingAuthor": sum(1 for row in rows if not has_value(first_value(row.get("作者署名"), row.get("作者原名")))),
        "missingSignedDate": sum(1 for row in rows if not has_value(row.get("签订日期"))),
        "missingExpiryDate": sum(1 for row in rows if not has_value(first_value(row.get("到期时间"), row.get("续约前到期日期")))),
        "missingProductLineAndCip": sum(1 for row in rows if not has_value(row.get("产品线")) and not has_value(row.get("CIP"))),
        "missingAudioRights": sum(1 for row in rows if not any(has_value(row.get(field)) for field in ["有声使用权", "有声改编权", "有声转授权", "有声权利描述"])),
    }


def summarize_candidates(candidates: list[dict]) -> dict:
    return {
        "total": len(candidates),
        "highConfidenceAuto": sum(1 for item in candidates if item["valueConfidence"] == "high" and item["autoApplyEligible"]),
        "mediumConfidenceSuggested": sum(1 for item in candidates if item["valueConfidence"] == "medium"),
        "lowConfidenceManual": sum(1 for item in candidates if item["valueConfidence"] == "low"),
        "manualReview": sum(1 for item in candidates if item["requiresManualReview"]),
        "conflict": sum(1 for item in candidates if item["conflictStatus"] != "none"),
        "byField": dict(Counter(item["fieldName"] for item in candidates)),
        "byMatchMethod": dict(Counter(item["matchMethod"] for item in candidates)),
        "byConfidence": dict(Counter(item["valueConfidence"] for item in candidates)),
    }


def summarize_field(items: list[dict]) -> dict:
    return {
        "candidateRows": len(items),
        "highConfidenceAuto": sum(1 for item in items if item["valueConfidence"] == "high" and item["autoApplyEligible"]),
        "mediumConfidenceSuggested": sum(1 for item in items if item["valueConfidence"] == "medium"),
        "lowConfidenceManual": sum(1 for item in items if item["valueConfidence"] == "low"),
        "conflict": sum(1 for item in items if item["conflictStatus"] != "none"),
    }


def build_private_conflicts(candidates: list[dict]) -> list[dict]:
    grouped = defaultdict(list)
    for item in candidates:
        if item["conflictStatus"] != "none":
            grouped[(item["standardWorkId"], item["fieldName"])].append(item)
    return [
        {
            "standardWorkId": standard,
            "fieldName": field,
            "candidateCount": len(items),
            "ledgerRowIds": sorted({row for item in items for row in item["ledgerRowIds"]}),
            "values": sorted({str(item["proposedValue"]) for item in items}),
            "requiresManualReview": True,
        }
        for (standard, field), items in grouped.items()
    ]


def write_outputs(payload: dict) -> None:
    OUTPUT_M1.mkdir(parents=True, exist_ok=True)
    OUTPUT_M2.mkdir(parents=True, exist_ok=True)
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)

    write_json(PRIVATE_JSON, payload["privatePayload"])
    write_json(STRUCTURE_JSON, public_payload(payload, "ledgerStructure"))
    write_json(GAP_JSON, public_payload(payload, "gapSummary"))
    write_json(SUMMARY_JSON, public_payload(payload, "summary"))
    write_json(CATEGORY_JSON, public_payload(payload, "categoryAuthorRightsSummary"))
    write_json(IMPACT_JSON, public_payload(payload, "m2ImpactSummary"))

    STRUCTURE_MD.write_text(render_structure_md(payload), encoding="utf-8")
    GAP_MD.write_text(render_gap_md(payload), encoding="utf-8")
    SUMMARY_MD.write_text(render_summary_md(payload), encoding="utf-8")
    CATEGORY_MD.write_text(render_category_md(payload), encoding="utf-8")
    IMPACT_MD.write_text(render_impact_md(payload), encoding="utf-8")


def public_payload(payload: dict, section: str) -> dict:
    base = {
        "schema": f"m1.m2.digital_copyright_ledger.{section}.v1",
        "generatedAt": payload["generatedAt"],
        "currentHead": payload["currentHead"],
        "originMain": payload["originMain"],
        "sanitizedBoundary": payload["sanitizedBoundary"],
    }
    if section == "summary":
        base.update(
            {
                "ledgerStructure": summary_only(payload["ledgerStructure"]),
                "matchingSummary": payload["matchingSummary"],
                "candidateSummary": payload["candidateSummary"],
                "gapSummary": payload["gapSummary"],
                "categoryAuthorRightsSummary": payload["categoryAuthorRightsSummary"],
                "m2ImpactSummary": payload["m2ImpactSummary"],
            }
        )
    else:
        base[section] = payload[section]
    return base


def summary_only(structure: dict) -> dict:
    return {
        "sheetCount": structure["sheetCount"],
        "totalDataRows": structure["totalDataRows"],
        "keyFieldCoverage": structure["keyFieldCoverage"],
    }


def render_structure_md(payload: dict) -> str:
    structure = payload["ledgerStructure"]
    key_rows = [
        {
            "field": item["fieldName"],
            "present": yes_no(item["present"]),
            "nonEmpty": item["nonEmptyCount"],
            "rate": pct(item["nonEmptyRate"]),
            "usage": "；".join(item["usage"]),
        }
        for item in structure["keyFieldCoverage"]
    ]
    return "\n".join(
        [
            "# M1 Digital Copyright Ledger Structure Audit v1",
            "",
            "本报告仅包含字段级和聚合统计，不包含真实书名、作者名、合同号、台账原始行或完整作品明细。",
            "",
            f"- Sheet 数：`{structure['sheetCount']}`",
            f"- 数据行数：`{structure['totalDataRows']}`",
            f"- 字段数：`{structure['sheets'][0]['fieldCount'] if structure['sheets'] else 0}`",
            "",
            markdown_table(key_rows, [("field", "字段"), ("present", "存在"), ("nonEmpty", "非空数"), ("rate", "非空率"), ("usage", "可用于")]),
            "",
            "结论：台账字段覆盖不只版权到期日，还可支持作品身份、标题作者、版权起止、权利、出版社、产品线、CIP 和 M2 renewal/risk/remaining-copyright 输入候选。",
        ]
    )


def render_gap_md(payload: dict) -> str:
    gap = payload["gapSummary"]
    field_rows = [
        {"field": key, "count": value}
        for key, value in gap["currentGapCounts"].items()
    ]
    backfill_rows = [
        {"field": field, **stats}
        for field, stats in sorted(gap["fieldBackfillPotential"].items())
    ]
    return "\n".join(
        [
            "# M1 Master Data Gap Ledger Backfill Audit v1",
            "",
            "本报告为脱敏聚合审计。自动补全候选仅为本地候选，不代表正式主数据已发布。",
            "",
            f"- 当前标准作品数：`{gap['totalStandardWorks']}`",
            f"- 候选总数：`{payload['candidateSummary']['total']}`",
            f"- 高置信自动补全候选：`{payload['candidateSummary']['highConfidenceAuto']}`",
            f"- 中置信建议补全候选：`{payload['candidateSummary']['mediumConfidenceSuggested']}`",
            f"- 低置信/人工复核候选：`{payload['candidateSummary']['lowConfidenceManual']}`",
            f"- 冲突候选：`{payload['candidateSummary']['conflict']}`",
            "",
            "## 当前缺口聚合",
            markdown_table(field_rows, [("field", "缺口类型"), ("count", "数量")]),
            "",
            "## 台账可补全潜力",
            markdown_table(
                backfill_rows,
                [
                    ("field", "字段"),
                    ("candidateRows", "候选数"),
                    ("highConfidenceAuto", "高置信自动"),
                    ("mediumConfidenceSuggested", "中置信建议"),
                    ("lowConfidenceManual", "低置信人工"),
                    ("conflict", "冲突"),
                ],
            ),
            "",
            "## 高收入作品覆盖",
            markdown_table(
                gap["highIncomeImpact"],
                [
                    ("bucket", "收入桶"),
                    ("standardWorkCount", "作品数"),
                    ("candidateRows", "候选数"),
                    ("highConfidenceAuto", "高置信自动"),
                    ("manualReview", "人工复核"),
                ],
            ),
        ]
    )


def render_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M1 Digital Copyright Ledger Masterdata Backfill Summary v1",
            "",
            "本报告只保留聚合统计；真实作品ID、作品名、作者和台账摘录仅在 gitignored private 候选包中。",
            "",
            "## 结论",
            f"- 台账数据行数：`{payload['ledgerStructure']['totalDataRows']}`",
            f"- 匹配台账行数：`{payload['matchingSummary']['matchedLedgerRows']}`",
            f"- 匹配标准作品数：`{payload['matchingSummary']['matchedStandardWorkCount']}`",
            f"- 高置信自动补全候选：`{payload['candidateSummary']['highConfidenceAuto']}`",
            f"- 中置信建议补全候选：`{payload['candidateSummary']['mediumConfidenceSuggested']}`",
            f"- 仍需人工复核：`{payload['candidateSummary']['manualReview']}`",
            "- 结论：大量人工录入可以被缩减为“高置信自动候选 + 中低置信复核清单”的受控流程，但不能直接写正式主数据。",
            "- M3 状态：仍不进入 M3。",
            "",
            "## 匹配方法分布",
            markdown_table(counter_rows(payload["matchingSummary"]["matchMethodDistribution"]), [("key", "方法"), ("count", "数量")]),
            "",
            "## 候选字段分布",
            markdown_table(counter_rows(payload["candidateSummary"]["byField"]), [("key", "字段"), ("count", "数量")]),
        ]
    )


def render_category_md(payload: dict) -> str:
    summary = payload["categoryAuthorRightsSummary"]
    rows = [
        {"area": "作者", **summary["authorBackfill"]},
        {"area": "分类一级", **summary["classificationBackfill"]["level1"]},
        {"area": "分类二级", **summary["classificationBackfill"]["level2"]},
        {"area": "分类三级", **summary["classificationBackfill"]["level3"]},
        {"area": "有声权利", **summary["audioRightsBackfill"]},
    ]
    return "\n".join(
        [
            "# M1 Ledger Category Author Rights Backfill Summary v1",
            "",
            "本报告为分类、作者和权利字段的脱敏聚合摘要。",
            "",
            markdown_table(
                rows,
                [
                    ("area", "范围"),
                    ("candidateRows", "候选数"),
                    ("highConfidenceAuto", "高置信自动"),
                    ("mediumConfidenceSuggested", "中置信建议"),
                    ("lowConfidenceManual", "低置信人工"),
                    ("conflict", "冲突"),
                ],
            ),
            "",
            f"- 产品线覆盖行数：`{summary['classificationBackfill']['productLineCoveredRows']}`",
            f"- CIP 覆盖行数：`{summary['classificationBackfill']['cipCoveredRows']}`",
            f"- 多作者风险行数：`{summary['multiAuthorRiskRows']}`",
            "- 分类边界：产品线/CIP 只能形成候选和置信度，不允许伪造权威三级分类。",
            "",
            "## 权利状态分布",
            markdown_table(counter_rows(summary["rightsDistribution"]), [("key", "权利状态"), ("count", "数量")]),
        ]
    )


def render_impact_md(payload: dict) -> str:
    impact = payload["m2ImpactSummary"]
    rows = [{"metric": key, "value": value} for key, value in impact.items()]
    return "\n".join(
        [
            "# M2 Ledger Backfill Impact on Evaluation v1",
            "",
            "本报告仅评估台账补全候选对 M2 的潜在影响，不宣称 M2 formal complete。",
            "",
            markdown_table(rows, [("metric", "指标"), ("value", "值")]),
            "",
            "结论：版权到期和权利字段补全后，应重新评估 remaining-copyright、renewal review、readiness warning 和有限业务复核样本，但仍需人工确认、正式授权和后续门禁。",
        ]
    )


def parse_date(value):
    raw = value
    if value is None or stringify(value) == "":
        return date_result("missing", "missing", raw)
    if hasattr(value, "strftime"):
        return date_result("exact_date", "parsed", raw, normalized=value.strftime("%Y-%m-%d"))
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        from datetime import date, timedelta

        parsed = date(1899, 12, 30) + timedelta(days=int(round(float(value))))
        return date_result("exact_date", "parsed", raw, normalized=parsed.isoformat())
    text = stringify(value)
    if re.search(r"无限期|无期限|永久|长期有效", text):
        return date_result("infinite", "parsed", raw)
    if re.search(r"自动续约|自动延续|顺延", text):
        selected = preferred_audio_date(text) or first_date(text)
        return date_result("auto_renewal", "parsed_with_condition" if selected else "manual_review", raw, normalized=selected)
    if re.search(r"授权书|附件|另行约定", text) and not first_date(text):
        return date_result("external_reference_no_date", "manual_review", raw)
    relative = relative_date(text)
    if relative:
        return date_result("relative_term", "relative", raw, **relative)
    selected = preferred_audio_date(text) or first_date(text)
    if selected:
        return date_result("exact_date", "parsed", raw, normalized=selected, extractedDates=extract_dates(text))
    return date_result("unparseable", "unparsed", raw)


def date_result(expiry_type, parser_status, raw, **extra):
    return {
        "expiryType": expiry_type,
        "parserStatus": parser_status,
        "normalizedDate": extra.get("normalized"),
        "anchor": extra.get("anchor"),
        "years": extra.get("years"),
        "endOfYear": extra.get("endOfYear", False),
        "extractedDates": extra.get("extractedDates", []),
        "requiresManualReview": parser_status in {"relative", "manual_review", "unparsed"},
        "rawValue": stringify(raw),
    }


def relative_date(text: str) -> dict | None:
    match = re.search(r"最后一部出版之日(?:起)?(\d{1,2})年", text)
    if match:
        return {"anchor": "last_publication_date", "years": int(match.group(1)), "endOfYear": "12月31日" in text}
    match = re.search(r"出版之日(?:起)?(\d{1,2})年", text)
    if match:
        return {"anchor": "publication_date", "years": int(match.group(1)), "endOfYear": "12月31日" in text}
    return None


def preferred_audio_date(text: str) -> str | None:
    match = re.search(r"有声[^0-9]*(20\d{2}|19\d{2})[/-](\d{1,2})[/-](\d{1,2})", text)
    if match:
        return format_date(match.group(1), match.group(2), match.group(3))
    return None


def first_date(text: str) -> str | None:
    dates = extract_dates(text)
    return dates[0] if dates else None


def extract_dates(text: str) -> list[str]:
    dates = []
    for pattern in [r"(20\d{2}|19\d{2})[/-](\d{1,2})[/-](\d{1,2})", r"(20\d{2}|19\d{2})年(\d{1,2})月(\d{1,2})日?"]:
        for match in re.finditer(pattern, text):
            dates.append(format_date(match.group(1), match.group(2), match.group(3)))
    return list(dict.fromkeys(dates))


def format_date(year, month, day) -> str:
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def parse_audio_rights(row: dict) -> dict:
    fields = ["有声使用权", "有声改编权", "有声转授权", "有声权利描述", "广播剧", "课程"]
    source = "；".join(filter(None, [stringify(row.get(field)) for field in fields]))
    has_audio = bool(re.search(r"有声|听书|音频|广播剧|课程|朗读|转授权|改编", source))
    denied = bool(re.search(r"无|否|不含|未授权|没有", source))
    explicit_grant = any(re.search(r"^(是|有|Y|YES|TRUE|1)$|授权|可|拥有|包含|有声", stringify(row.get(field)), re.I) for field in ["有声使用权", "有声改编权", "有声转授权"])
    explicit_deny = any(re.search(r"^(否|无|N|NO|FALSE|0)$|不含|未授权|没有", stringify(row.get(field)), re.I) for field in ["有声使用权", "有声改编权", "有声转授权"])
    if explicit_grant or (has_audio and not denied):
        status = "granted"
    elif explicit_deny or (denied and not has_audio):
        status = "not_granted"
    elif has_audio and denied:
        status = "limited_or_conflict"
    else:
        status = "unknown"
    return {
        "audioRightsStatus": status,
        "sourceSummary": source,
        "requiresManualReview": status in {"unknown", "limited_or_conflict"} or (explicit_grant and explicit_deny),
    }


def parse_category(row: dict) -> dict:
    product = stringify(row.get("产品线"))
    cip = stringify(row.get("CIP"))
    return {
        "level1": "出版物" if product or cip else None,
        "level2": product or None,
        "level3": None,
        "classificationConfidence": "medium" if product else "low" if cip else "missing",
    }


def normalize_title(value) -> str:
    text = stringify(value)
    text = text.translate(str.maketrans({chr(code): chr(code - 0xFEE0) for code in range(0xFF01, 0xFF5F)}))
    text = text.replace("\u3000", " ")
    text = re.sub(r"[《》“”\"']", "", text)
    text = re.sub(r"[：:]", ":", text)
    text = re.sub(r"[（(].*?[）)]", "", text)
    text = re.sub(r"新版|修订版|珍藏版|套装|全集|增订版|纪念版|典藏版", "", text)
    text = re.sub(r"\s+", "", text)
    return text.strip().lower()


def normalize_author_tokens(value) -> list[str]:
    text = stringify(value)
    if not text:
        return []
    tokens = re.split(r"[、，,;；/／&]|\s+and\s+|\s+和\s+|\s+及\s+", text)
    return list(dict.fromkeys(re.sub(r"\s+", "", item).lower() for item in tokens if item.strip()))


def normalize_id(value) -> str:
    text = stringify(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def id_sort_key(value) -> tuple[int, str]:
    text = normalize_id(value)
    return (int(text), text) if re.fullmatch(r"\d+", text) else (10**18, text)


def single_standard(standards) -> str:
    return sorted(standards, key=id_sort_key)[0]


def count_needed(rows: list[dict], token: str) -> int:
    return sum(1 for row in rows if token in str(row.get("需要补充的字段", "")))


def dataframe_records(frame) -> list[dict]:
    records = []
    for row in frame.to_dict(orient="records"):
        records.append({key: json_safe(value) for key, value in row.items()})
    return records


def json_safe(value):
    if isinstance(value, dict):
        return {str(key): json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [json_safe(child) for child in value]
    if hasattr(value, "item"):
        try:
            return json_safe(value.item())
        except Exception:
            pass
    try:
        if value != value:
            return None
    except Exception:
        pass
    return value


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    lines = ["| " + " | ".join(label for _, label in columns) + " |"]
    lines.append("|" + "|".join("---" for _ in columns) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |")
    return "\n".join(lines)


def counter_rows(counter: dict) -> list[dict]:
    return [{"key": key, "count": value} for key, value in sorted(counter.items(), key=lambda item: (-item[1], item[0]))]


def bigrams(value: str) -> set[str]:
    return {value[index : index + 2] for index in range(max(0, len(value) - 1))}


def overlap_ratio(left: set[str], right: set[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left.intersection(right)) / len(left.union(right))


def ratio(part: int, total: int) -> float:
    return round(part / total, 4) if total else 0.0


def pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def yes_no(value: bool) -> str:
    return "是" if value else "否"


def safe_float(value) -> float:
    try:
        result = float(value)
    except Exception:
        return 0.0
    return result if math.isfinite(result) else 0.0


def first_value(*values):
    for value in values:
        if has_value(value):
            return value
    return ""


def has_value(value) -> bool:
    return stringify(value) != ""


def clean_cell(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value
    return str(value).strip()


def stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sanitized_boundary() -> dict:
    return {
        "sanitizedAggregateOnly": True,
        "realTitlesIncluded": False,
        "authorNamesIncluded": False,
        "channelNamesIncluded": False,
        "rawLedgerRowsIncluded": False,
        "privateDetailsStoredOnlyInGitignoredOutput": True,
        "formalMasterDataWritten": False,
        "m3Entered": False,
    }


def run_git(args: list[str]) -> str | None:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return None


if __name__ == "__main__":
    main()
