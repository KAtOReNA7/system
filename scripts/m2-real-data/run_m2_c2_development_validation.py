#!/usr/bin/env python3
"""Freeze Gate C, then run the sealed C2 development replay.

The default mode is synthetic-only.  Phase A and development modes are
explicit, branch-bound, local-only operations.  The final-holdout mode exists
only to prove fail-closed behavior before any loader can run.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import os
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

import m2_calibration_c2_v1 as c2
import m2_calibration_v1 as base
import m2_calibration_v1_2 as v12
import m2_formal_cash_comparator_v1 as formal
import run_m2_calibration_baseline_replay as legacy
import run_m2_c2r1_development_validation as c2r1_runner
import run_m2_formal_cash_comparator_replay as phase_a


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-c2-v1"
OPPORTUNITY_JSON = PUBLIC_DIR / "M2-C2-opportunity-audit-v1.json"
OPPORTUNITY_MD = PUBLIC_DIR / "M2-C2-opportunity-audit-v1.md"
DESIGN_JSON = PUBLIC_DIR / "M2-C2-model-design-v1.json"
DESIGN_MD = PUBLIC_DIR / "M2-C2-model-design-v1.md"
GATE_C_JSON = PUBLIC_DIR / "M2-calibration-gate-c-v1.json"
SEGMENT_JSON = PUBLIC_DIR / "M2-C2-activity-segment-route-manifest-v1.json"
SEGMENT_MD = PUBLIC_DIR / "M2-C2-activity-segment-route-manifest-v1.md"
RESIDUAL_JSON = PUBLIC_DIR / "M2-C2-other-new-channel-residual-audit-v1.json"
RESIDUAL_MD = PUBLIC_DIR / "M2-C2-other-new-channel-residual-audit-v1.md"
GUARD_JSON = PUBLIC_DIR / "M2-C2-high-value-guard-audit-v1.json"
GUARD_MD = PUBLIC_DIR / "M2-C2-high-value-guard-audit-v1.md"
VALIDATION_JSON = PUBLIC_DIR / "M2-C2-development-validation-v1.json"
VALIDATION_MD = PUBLIC_DIR / "M2-C2-development-validation-v1.md"
MODEL_DECISION_JSON = PUBLIC_DIR / "M2-C2-model-quality-decision-v1.json"
MODEL_DECISION_MD = PUBLIC_DIR / "M2-C2-model-quality-decision-v1.md"
BUSINESS_DECISION_JSON = PUBLIC_DIR / "M2-C2-business-coverage-decision-v1.json"
BUSINESS_DECISION_MD = PUBLIC_DIR / "M2-C2-business-coverage-decision-v1.md"
PRIVATE_PHASE_A_MANIFEST = PRIVATE_DIR / "M2-C2-phase-a-manifest-private-v1.json"
PRIVATE_VALIDATION_RECEIPT = PRIVATE_DIR / "M2-calibration-gate-c-validation-private-v1.json"
PRIVATE_PUSH_RECEIPT = PRIVATE_DIR / "M2-calibration-gate-c-push-private-v1.json"
PRIVATE_CASES = PRIVATE_DIR / "M2-C2-development-cases-private-v1.ndjson"
PRIVATE_DEVELOPMENT_MANIFEST = PRIVATE_DIR / "M2-C2-development-manifest-private-v1.json"
PRIVATE_WORKBOOK = PRIVATE_DIR / "M2-C2-中文业务抽检工作簿-private-v1.xlsx"
BRANCH = "codex/m2-c2-v1"
PHASE_A_START_HEAD = "7b315d74a29f0eb136b0827e7f2f0367db868c63"
TOLERANCE = 1e-9


PHASE_A_TRACKED_PATHS = (
    c2.SPEC_PATH,
    ROOT / "scripts" / "m2-real-data" / "m2_calibration_c2_v1.py",
    ROOT / "scripts" / "m2-real-data" / "run_m2_c2_development_validation.py",
    ROOT / "test" / "m2-c2-contract.test.js",
    ROOT / "test" / "m2-c2-development-validation.test.js",
    ROOT / "scripts" / "m2-real-data" / "run_m2_formal_cash_comparator_replay.py",
    ROOT / "test" / "m2-formal-cash-comparator.test.js",
    ROOT / "test" / "m2-formal-cash-target-report.test.js",
    ROOT / "package.json",
    OPPORTUNITY_JSON,
    OPPORTUNITY_MD,
    DESIGN_JSON,
    DESIGN_MD,
    GATE_C_JSON,
)
IMMUTABLE_PHASE_A_PATHS = PHASE_A_TRACKED_PATHS[:-1]
PUBLIC_PHASE_A_PATHS = (
    OPPORTUNITY_JSON,
    OPPORTUNITY_MD,
    DESIGN_JSON,
    DESIGN_MD,
    GATE_C_JSON,
)
PRIVATE_PATHS = (
    PRIVATE_PHASE_A_MANIFEST,
    PRIVATE_VALIDATION_RECEIPT,
    PRIVATE_PUSH_RECEIPT,
    PRIVATE_CASES,
    PRIVATE_DEVELOPMENT_MANIFEST,
    PRIVATE_WORKBOOK,
)


class C2RunnerError(RuntimeError):
    """C2 phase or replay evidence failed a frozen invariant."""


def progress(message: str) -> None:
    print(f"[m2-c2] {message}", file=sys.stderr, flush=True)


def run_git(*args: str, check: bool = True) -> str:
    process = subprocess.run(
        ["git", *args], cwd=ROOT, capture_output=True, text=True, check=False
    )
    if check and process.returncode != 0:
        raise C2RunnerError(process.stderr.strip() or "git command failed")
    return process.stdout.strip()


def file_sha256(path: Path) -> str:
    digest_builder = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest_builder.update(chunk)
    return digest_builder.hexdigest()


def digest(value: Any) -> str:
    return formal.canonical_digest(value)


def rounded(value: Any, places: int = 8) -> float | None:
    if value is None:
        return None
    number = float(value)
    return round(number, places) if math.isfinite(number) else None


def ratio(numerator: float, denominator: float) -> float | None:
    return float(numerator) / float(denominator) if float(denominator) != 0 else None


def write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False)
        + "\n",
        encoding="utf-8",
        newline="\n",
    )


def write_text(path: Path, value: str) -> None:
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def public_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {
            str(key): public_value(child)
            for key, child in value.items()
            if key
            not in {
                "actualTotal",
                "predictedTotal",
                "forecastableActualByComponent",
                "totalLedgerActualByComponent",
            }
        }
    if isinstance(value, list):
        return [public_value(child) for child in value]
    if isinstance(value, float):
        return rounded(value)
    return value


def require_named_branch() -> None:
    if run_git("branch", "--show-current") != BRANCH:
        raise C2RunnerError(f"C2 local evidence must run on {BRANCH}")
    spec = c2.load_spec()
    if any(value is not False for value in spec["seals"].values()):
        raise C2RunnerError("C2 seal is open")


def _assert_private_paths() -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    for path in PRIVATE_PATHS:
        relative = path.relative_to(ROOT).as_posix()
        ignored = subprocess.run(
            ["git", "check-ignore", "--quiet", "--", relative],
            cwd=ROOT,
            check=False,
        ).returncode
        if ignored != 0:
            raise C2RunnerError(f"private C2 role is not Git-ignored: {path.name}")
        if run_git("ls-files", "--", relative):
            raise C2RunnerError(f"private C2 role is tracked: {path.name}")
    if phase_a.phase.tracked_private_artifacts():
        raise C2RunnerError("a private calibration artifact is tracked")


def assert_public_safety(paths: Sequence[Path]) -> None:
    forbidden = (
        "data/private",
        "private-output",
        "standard_work_id",
        "channel_key",
        "rawchannel",
        ".xlsx",
        "optimistic",
        "pessimistic",
        "high/base/low",
    )
    for path in paths:
        text = path.read_text(encoding="utf-8")
        lowered = text.lower()
        if path.suffix == ".md" and not any("\u4e00" <= char <= "\u9fff" for char in text):
            raise C2RunnerError(f"public C2 markdown is not Chinese: {path.name}")
        if any(value in lowered for value in forbidden):
            raise C2RunnerError(f"public C2 artifact violates privacy: {path.name}")
        if '"lower"' in lowered or '"upper"' in lowered:
            raise C2RunnerError(f"public C2 artifact exposes interval endpoints: {path.name}")
        if any(f"{letter}:\\" in lowered for letter in "abcdefghijklmnopqrstuvwxyz"):
            raise C2RunnerError(f"public C2 artifact contains an absolute path: {path.name}")


def _case_key(payload: Mapping[str, Any]) -> tuple[str, str, int, str]:
    key = payload.get("caseKey", payload.get("case_key", {}))
    return (
        str(key["standard_work_id"]),
        str(key["origin"]),
        int(key["horizon_months"]),
        str(key["route"]),
    )


def _load_c2r1_private_cases() -> tuple[
    dict[tuple[str, str, int, str], dict[str, Any]], dict[str, Any]
]:
    manifest = json.loads(c2r1_runner.PRIVATE_MANIFEST.read_text(encoding="utf-8"))
    hasher = hashlib.sha256()
    rows: dict[tuple[str, str, int, str], dict[str, Any]] = {}
    count = 0
    with c2r1_runner.PRIVATE_CASES.open("rb") as handle:
        for raw in handle:
            if not raw.endswith(b"\n"):
                raise C2RunnerError("C2-R.1 private case is not LF-delimited")
            payload = json.loads(raw[:-1].decode("utf-8"))
            canonical = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            if canonical != raw:
                raise C2RunnerError("C2-R.1 private case is not canonical")
            hasher.update(raw)
            count += 1
            key = _case_key(payload)
            if key in rows:
                raise C2RunnerError("duplicate C2-R.1 private case key")
            rows[key] = payload
    if (
        count != int(manifest["privateCaseCount"])
        or hasher.hexdigest() != str(manifest["privateCaseSha256"])
    ):
        raise C2RunnerError("C2-R.1 private case manifest differs")
    return rows, manifest


def _metric_pair(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {
            "caseCount": 0,
            "uniqueWorkCount": 0,
            "B4": {"wape": None, "signedAggregateBias": None},
            "C2R1": {"wape": None, "signedAggregateBias": None},
            "C2R1RelativeWapeDeltaVsB4": None,
            "B4LowerAbsoluteErrorCaseCount": 0,
            "C2R1LowerAbsoluteErrorCaseCount": 0,
            "equalAbsoluteErrorCaseCount": 0,
        }
    actuals = [float(row["actual"]) for row in rows]
    b4_points = [float(row["B4"]) for row in rows]
    c2r1_points = [float(row["C2R1"]) for row in rows]
    b4_wape = base.wape(b4_points, actuals)
    c2r1_wape = base.wape(c2r1_points, actuals)
    b4_better = 0
    c2r1_better = 0
    equal = 0
    for b4_point, c2r1_point, actual in zip(b4_points, c2r1_points, actuals):
        b4_error = abs(b4_point - actual)
        c2r1_error = abs(c2r1_point - actual)
        if b4_error + TOLERANCE < c2r1_error:
            b4_better += 1
        elif c2r1_error + TOLERANCE < b4_error:
            c2r1_better += 1
        else:
            equal += 1
    return {
        "caseCount": len(rows),
        "uniqueWorkCount": len({str(row["workId"]) for row in rows}),
        "B4": {
            "wape": b4_wape,
            "signedAggregateBias": base.signed_aggregate_bias(b4_points, actuals),
        },
        "C2R1": {
            "wape": c2r1_wape,
            "signedAggregateBias": base.signed_aggregate_bias(c2r1_points, actuals),
        },
        "C2R1RelativeWapeDeltaVsB4": (
            (c2r1_wape - b4_wape) / b4_wape if b4_wape else 0.0
        ),
        "B4LowerAbsoluteErrorCaseCount": b4_better,
        "C2R1LowerAbsoluteErrorCaseCount": c2r1_better,
        "equalAbsoluteErrorCaseCount": equal,
    }


def _public_group_metric(
    rows: Sequence[Mapping[str, Any]], minimum_works: int
) -> dict[str, Any]:
    metric = _metric_pair(rows)
    if int(metric["uniqueWorkCount"]) >= minimum_works:
        return public_value(metric)
    return {
        "suppressed": True,
        "suppressionReason": "complementary_small_sample",
        "minimumWorks": minimum_works,
        "caseCount": None,
        "uniqueWorkCount": None,
        "B4": {"wape": None, "signedAggregateBias": None},
        "C2R1": {"wape": None, "signedAggregateBias": None},
        "C2R1RelativeWapeDeltaVsB4": None,
    }


def _opportunity_audit(
    *,
    spec: Mapping[str, Any],
    b4: Mapping[tuple[str, tuple[str, str, int, str]], Mapping[str, Any]],
    c2r1_rows: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    works: Sequence[Mapping[str, Any]],
    input_evidence: Mapping[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    work_by_id = {str(work["standard_work_id"]): work for work in works}
    calibration_spec, _v11, _v12 = v12.load_and_validate_contract()
    forward = {
        key: template
        for (role, key), template in b4.items()
        if role.startswith("development_forward_score:")
    }
    if set(forward) != set(c2r1_rows) or len(forward) != int(
        spec["authority"]["developmentCaseCount"]
    ):
        raise C2RunnerError("B4 and C2-R.1 opportunity case keys differ")
    audit_rows: list[dict[str, Any]] = []
    model_keys: set[tuple[str, str, int, str]] = set()
    for key, template in sorted(forward.items()):
        c2r1_row = c2r1_rows[key]
        if (
            bool(template["statisticallyScoreable"])
            and bool(template["modelPredictionAvailable"])
            and not bool(template["routeAbstained"])
        ):
            if c2r1_row.get("rawModelPrediction") is None:
                raise C2RunnerError("C2-R.1 model-population point is null")
            model_keys.add(key)
            segment = c2.segment_as_of(
                work_by_id[key[0]], key[1], calibration_spec, spec
            )
            known = {
                str(item["channel_key"])
                for item in template.get("channelComponents", [])
            }
            actual_components = template.get("forecastableActualByComponent", {}) or {}
            residual_actual = sum(
                float(value)
                for component_key, value in actual_components.items()
                if str(component_key) not in known
            )
            strata = template.get("strata", {}) or {}
            audit_rows.append(
                {
                    "workId": key[0],
                    "origin": key[1],
                    "horizon": key[2],
                    "route": key[3],
                    "segment": segment["segment"],
                    "segmentReason": segment["segmentReason"],
                    "source": str(strata.get("source", "unknown")),
                    "top1": bool(strata.get("top_1_percent")),
                    "top5": bool(strata.get("top_5_percent")),
                    "top10": bool(strata.get("top_10_percent")),
                    "residualExposure": abs(residual_actual) > TOLERANCE,
                    "residualActual": residual_actual,
                    "actual": float(template["forecastableCashActual"]),
                    "B4": float(template["rawModelPrediction"]),
                    "C2R1": float(c2r1_row["rawModelPrediction"]),
                }
            )
    if len(model_keys) != int(spec["authority"]["formalModelPopulationCaseCount"]):
        raise C2RunnerError("opportunity model population differs")
    minimum = int(spec["privacy"]["complementarySuppressionMinimumWorks"])

    def grouped(field: str, values: Sequence[Any]) -> dict[str, Any]:
        return {
            str(value): _public_group_metric(
                [row for row in audit_rows if row[field] == value], minimum
            )
            for value in values
        }

    segments = grouped("segment", c2.ACTIVITY_SEGMENTS)
    origins = grouped("origin", spec["authority"]["origins"])
    stable: dict[str, Any] = {}
    for segment in c2.ACTIVITY_SEGMENTS:
        segment_origin = {
            origin: _metric_pair(
                [
                    row
                    for row in audit_rows
                    if row["segment"] == segment and row["origin"] == origin
                ]
            )
            for origin in spec["authority"]["origins"]
        }
        eligible = [
            value
            for value in segment_origin.values()
            if int(value["caseCount"]) > 0
            and value["C2R1RelativeWapeDeltaVsB4"] is not None
        ]
        wins = sum(
            float(value["C2R1RelativeWapeDeltaVsB4"]) < 0 for value in eligible
        )
        stable[segment] = {
            "originCount": len(eligible),
            "C2R1OriginWinCount": wins,
            "C2R1OriginWinShare": ratio(wins, len(eligible)),
            "stableC2R1WinRegion": bool(
                eligible and ratio(wins, len(eligible)) >= 0.7
            ),
        }
    top_bands = {
        "top1": _public_group_metric([row for row in audit_rows if row["top1"]], minimum),
        "top5": _public_group_metric([row for row in audit_rows if row["top5"]], minimum),
        "top10": _public_group_metric([row for row in audit_rows if row["top10"]], minimum),
    }
    residual_groups = {
        "futureOrOtherChannelCashPresent": _public_group_metric(
            [row for row in audit_rows if row["residualExposure"]], minimum
        ),
        "noFutureOrOtherChannelCash": _public_group_metric(
            [row for row in audit_rows if not row["residualExposure"]], minimum
        ),
        "signedResidualCash": rounded(sum(row["residualActual"] for row in audit_rows)),
        "absoluteResidualCash": rounded(
            sum(abs(row["residualActual"]) for row in audit_rows)
        ),
    }
    candidate_override = [
        segment
        for segment in c2.ACTIVITY_SEGMENTS
        if segments[segment].get("C2R1RelativeWapeDeltaVsB4") is not None
        and float(segments[segment]["C2R1RelativeWapeDeltaVsB4"]) < 0
    ]
    mandatory_fallback = [
        segment
        for segment in c2.ACTIVITY_SEGMENTS
        if segment not in candidate_override
    ]
    for horizon in (18, 24):
        metric = _metric_pair([row for row in audit_rows if row["horizon"] == horizon])
        if float(metric["C2R1RelativeWapeDeltaVsB4"] or 0.0) > 0:
            mandatory_fallback.append(f"horizon_{horizon}")
    if float(top_bands["top10"].get("C2R1RelativeWapeDeltaVsB4") or 0.0) >= 0:
        mandatory_fallback.append("top10_until_earlier_origin_guard_passes")
    unsafe = [
        label
        for label, group in {
            **{f"segment:{key}": value for key, value in segments.items()},
            **{f"source:{key}": value for key, value in grouped("source", sorted({row['source'] for row in audit_rows})).items()},
        }.items()
        if group.get("suppressed") is True
    ]
    unsafe.extend(
        f"segment:{segment}:aggregate_opportunity_not_origin_stable"
        for segment in candidate_override
        if not stable[segment]["stableC2R1WinRegion"]
    )
    report = {
        "schema": "m2.c2_opportunity_audit.v1",
        "version": "v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "scope": {
            "frozenCaseCount": len(forward),
            "modelPopulationCaseCount": len(model_keys),
            "sameCaseKeys": True,
            "sameActuals": True,
            "sameModelPopulation": True,
            "inputFingerprintMatched": bool(input_evidence.get("inputFingerprint")),
            "outerResultsUsedToMoveSegmentThresholds": False,
        },
        "overall": public_value(_metric_pair(audit_rows)),
        "activitySegments": segments,
        "routes": grouped("route", ["pure_sales_share", "buyout_plus_sales"]),
        "horizons": grouped("horizon", [3, 6, 12, 18, 24]),
        "origins": origins,
        "sourcePostHoc": grouped(
            "source", sorted({row["source"] for row in audit_rows})
        ),
        "topBands": top_bands,
        "knownAndFutureChannelExposure": residual_groups,
        "innerOriginStability": public_value(stable),
        "conclusions": {
            "candidateOverrideOpportunity": candidate_override,
            "B4MandatoryFallbackRegion": mandatory_fallback,
            "highValueFallbackRegion": "top10_until_frozen_guard_passes",
            "newChannelResidualExposurePresent": abs(
                float(residual_groups["signedResidualCash"] or 0.0)
            )
            > TOLERANCE,
            "unsafeToSegmentAutomatically": unsafe,
            "auditMaySelectOuterCandidateDirectly": False,
        },
        "privacy": {
            "aggregateOnly": True,
            "identifiersPresent": False,
            "realChannelNamesPresent": False,
            "predictionIntervalEndpointsPresent": False,
            "complementarySuppressionApplied": True,
        },
        "seals": copy.deepcopy(spec["seals"]),
    }
    private = {
        "modelPopulationCaseCount": len(model_keys),
        "modelPopulationKeyDigest": digest([list(key) for key in sorted(model_keys)]),
        "segmentCounts": {
            segment: sum(row["segment"] == segment for row in audit_rows)
            for segment in c2.ACTIVITY_SEGMENTS
        },
        "segmentWorkCounts": {
            segment: len(
                {row["workId"] for row in audit_rows if row["segment"] == segment}
            )
            for segment in c2.ACTIVITY_SEGMENTS
        },
        "rowDigest": digest(
            [
                {
                    "key": [row["workId"], row["origin"], row["horizon"], row["route"]],
                    "segment": row["segment"],
                    "B4": row["B4"],
                    "C2R1": row["C2R1"],
                    "actual": row["actual"],
                    "residualActual": row["residualActual"],
                }
                for row in audit_rows
            ]
        ),
    }
    return report, private


def _design_report(spec: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "schema": "m2.c2_model_design.v1",
        "version": "v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "primaryComparator": "B4",
        "samePredictAsOfEntryForReplayAndServing": True,
        "activitySegmentation": public_value(spec["activitySegmentation"]),
        "candidateSpace": public_value(spec["candidateSpace"]),
        "otherOrNewChannelResidual": public_value(
            spec["otherOrNewChannelResidual"]
        ),
        "highValueGuard": public_value(spec["highValueGuard"]),
        "selection": public_value(spec["selection"]),
        "routes": public_value(spec["routes"]),
        "qualityGate": public_value(spec["acceptance"]),
        "businessCoverageDecision": public_value(
            spec["businessCoverageDecision"]
        ),
        "productBoundary": public_value(spec["productOutput"]),
        "predictionInterval": {
            "internalOnly": True,
            "reportedMetrics": ["coverage", "WIS", "standardized_width"],
            "endpointsPresent": False,
        },
        "featureBoundary": {
            "cutoffOnly": True,
            "currentRatingRiskRightsShelfUsed": False,
            "sourceWithoutHistoricalSnapshotUsed": False,
            "postHocReportingSlicesMayAffectPrediction": False,
            "spikeCandidateAutomaticallyDecayed": False,
        },
        "formalCashBoundary": {
            "futureBuyoutProbabilityModelPresent": False,
            "historicalBuyoutCyclePredictionPresent": False,
            "buyoutMonthlyEquivalentIncluded": False,
            "pureBuyoutNoCommitment": "null_abstain",
            "mixedExcludesUncommittedFutureBuyout": True,
        },
        "outerDevelopmentExecuted": False,
        "privacy": copy.deepcopy(spec["privacy"]),
        "seals": copy.deepcopy(spec["seals"]),
        "nextBoundary": "Gate_C_before_C2_outer_replay",
    }


def _write_phase_a_markdown(
    opportunity: Mapping[str, Any], design: Mapping[str, Any]
) -> None:
    conclusions = opportunity["conclusions"]
    write_text(
        OPPORTUNITY_MD,
        f"""# M2 C2 失败机会与设计审计 v1

本报告只比较完全一致的 development case、actual 和模型人口。B4 与 C2-R.1 的共同模型人口为 {opportunity['scope']['modelPopulationCaseCount']} 个 case；审计不读取 final holdout，也不直接决定任何 outer origin 的候选。

## 结论

- 可继续探索的活跃度区域：{json.dumps(conclusions['candidateOverrideOpportunity'], ensure_ascii=False)}。
- 必须保留 B4 回退的区域：{json.dumps(conclusions['B4MandatoryFallbackRegion'], ensure_ascii=False)}。
- 高价值保护：Top10 在 strictly-earlier 证据通过前一律回退 B4。
- 其他或新增渠道现金暴露存在：{str(conclusions['newChannelResidualExposurePresent']).lower()}；后续只允许通用残差组件，不预测真实渠道身份。
- 小样本或证据不足区域继续互补抑制，不用于事后移动分层或 gate。

## 边界

当前结果为 not_for_formal_decision。final holdout、embargo shadow 和 deferred 60-month labels 均未打开；未进入 C3，未 release。
""",
    )
    write_text(
        DESIGN_MD,
        f"""# M2 C2 模型设计 v1

C2 以 B4 为锚，冻结 dense、intermittent、dormant 三类 cutoff-only 活跃度路由。候选总数按分层分别为 {design['candidateSpace']['dense']['candidateCount']}、{design['candidateSpace']['intermittent']['candidateCount']}、{design['candidateSpace']['dormant']['candidateCount']}；选择顺序固定为偏差可行性、WAPE、高价值安全、horizon 安全和最小复杂度。

其他或新增渠道只使用 strictly-earlier origin 聚合证据形成通用组件，不记忆作品或未来渠道。Top1、Top5、Top10 按 cutoff 前 trailing-12 收入定义；高价值覆盖 B4 必须先通过冻结的 earlier-origin 稳定性条件，否则回退 B4。

产品、API、Excel 和正式导出仍只允许一个点值、年度拆分、confidence 和 limitation。80% 区间仅供内部 coverage、WIS 与宽度审计，不公开端点。

该设计继续保持 not_for_formal_decision；pure-buyout 无 cutoff commitment 时为 null abstain；未开始 C2 outer replay、C3 或 release。
""",
    )


def _gate_report(
    *,
    spec: Mapping[str, Any],
    opportunity: Mapping[str, Any],
    synthetic: Mapping[str, Any],
    validation_passed: bool = False,
    validation_receipt_sha256: str | None = None,
    validation_evidence: Mapping[str, Any] | None = None,
    phase_a_commit_pushed: bool = False,
    phase_a_checkpoint: str | None = None,
    remote_head_verified: bool = False,
) -> dict[str, Any]:
    checks = synthetic["checks"]
    conditions = {
        "as_of_activity_segments_frozen": all(
            checks[name]
            for name in (
                "denseDefinitionDeterministic",
                "intermittentDefinitionDeterministic",
                "dormantDefinitionDeterministic",
                "allSegmentReasonsPresent",
            )
        )
        and spec["activitySegmentation"]["thresholdsMayChangeAfterOuterResults"]
        is False,
        "candidate_space_frozen": checks["candidateCountsFrozen"]
        and spec["candidateSpace"]["candidateSpaceMayChangeAfterResults"] is False,
        "other_new_channel_residual_frozen": checks[
            "residualStoresNoFutureChannelIdentity"
        ]
        and spec["otherOrNewChannelResidual"]["outerTruthAllowed"] is False,
        "high_value_guard_frozen": checks["highValueGuardFallsBackToB4"]
        and spec["highValueGuard"]["outerBandActualAllowed"] is False,
        "selection_objective_frozen": spec["selection"]["orderedObjective"]
        == [
            "signed_bias_feasibility",
            "minimum_wape",
            "high_value_safety",
            "horizon_safety",
            "minimum_complexity",
        ]
        and spec["selection"]["thresholdsMayMoveAfterResults"] is False,
        "case_population_parity_passed": opportunity["scope"]["sameCaseKeys"]
        and opportunity["scope"]["sameActuals"]
        and opportunity["scope"]["sameModelPopulation"]
        and int(opportunity["scope"]["frozenCaseCount"])
        == int(spec["authority"]["developmentCaseCount"])
        and int(opportunity["scope"]["modelPopulationCaseCount"])
        == int(spec["authority"]["formalModelPopulationCaseCount"]),
        "pure_buyout_abstention_test_passed": checks["pureBuyoutNullAbstain"],
        "mixed_excludes_future_buyout_test_passed": checks[
            "mixedExcludesUncommittedFutureBuyout"
        ],
        "residual_no_leakage_test_passed": checks[
            "residualFitRejectsSameOrLaterOrigin"
        ]
        and checks["predictionRejectsActualField"]
        and checks["channelResidualWorkPointReconciles"],
        "future_perturbation_tests_passed": checks[
            "futureAsOfSegmentInvariant"
        ]
        and checks["futurePredictionInvariant"],
        "all_seals_closed": all(value is False for value in spec["seals"].values()),
        "phase_a_commit_pushed": bool(phase_a_commit_pushed),
        "full_validation_suite_passed": bool(validation_passed),
        "no_private_file_tracked": phase_a.phase.tracked_private_artifacts() == [],
    }
    expected = list(spec["gateC"]["conditions"])
    if list(conditions) != expected:
        raise C2RunnerError("Gate C condition order differs from the frozen spec")
    passed = sum(bool(value) for value in conditions.values())
    all_true = passed == len(conditions)
    evidence = public_value(dict(validation_evidence or {}))
    return {
        "schema": "m2.calibration_gate_c.v1",
        "version": "v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "conditions": conditions,
        "passedConditionCount": passed,
        "conditionCount": len(conditions),
        "allTrue": all_true,
        "C2AuthorizedByGateC": all_true,
        "phaseACommitPushed": bool(phase_a_commit_pushed),
        "phaseACheckpoint": phase_a_checkpoint,
        "remoteHeadVerified": bool(remote_head_verified),
        "validationReceiptSha256": validation_receipt_sha256,
        "validationEvidence": evidence,
        "privateFilesTracked": False,
        "C2OuterReplayExecuted": False,
        "seals": copy.deepcopy(spec["seals"]),
        "nextBoundary": "C2_outer_replay_only_if_all_14_true",
    }


def _phase_a_public_hashes() -> dict[str, str]:
    return {path.name: file_sha256(path) for path in PUBLIC_PHASE_A_PATHS}


def run_phase_a() -> dict[str, Any]:
    require_named_branch()
    if run_git("rev-parse", "HEAD") != PHASE_A_START_HEAD:
        raise C2RunnerError("C2 Phase A must start from the authorized HEAD")
    _assert_private_paths()
    spec = c2.load_spec()
    progress("loading locked B4 and C2-R.1 development evidence read-only")
    b4, _comparator, formal_manifest = c2r1_runner._load_formal_private_cases()
    c2r1_rows, c2r1_manifest = _load_c2r1_private_cases()
    calibration_spec, _v11, _v12 = v12.load_and_validate_contract()
    works, _posthoc, input_evidence = legacy.load_authorized_works(calibration_spec)
    if not (
        input_evidence["inputFingerprint"] == formal_manifest["inputFingerprint"]
        == c2r1_manifest["inputFingerprint"]
    ):
        raise C2RunnerError("C2 Phase A authority fingerprint differs")
    progress("computing deidentified B4 to C2-R.1 opportunity decomposition")
    opportunity, private_audit = _opportunity_audit(
        spec=spec,
        b4=b4,
        c2r1_rows=c2r1_rows,
        works=works,
        input_evidence=input_evidence,
    )
    design = _design_report(spec)
    synthetic = c2.synthetic_self_test()
    gate = _gate_report(
        spec=spec,
        opportunity=opportunity,
        synthetic=synthetic,
    )
    failed = [key for key, value in gate["conditions"].items() if not value]
    if set(failed) != {"phase_a_commit_pushed", "full_validation_suite_passed"}:
        raise C2RunnerError("Gate C content failed before validation: " + ", ".join(failed))
    write_json(OPPORTUNITY_JSON, opportunity)
    write_json(DESIGN_JSON, design)
    write_json(GATE_C_JSON, gate)
    _write_phase_a_markdown(opportunity, design)
    assert_public_safety(PUBLIC_PHASE_A_PATHS)
    manifest = {
        "schema": "m2.c2_phase_a_manifest.private.v1",
        "decisionStatus": "not_for_formal_decision",
        "tracked": False,
        "authorizedStartHead": PHASE_A_START_HEAD,
        "specDigest": c2.canonical_digest(spec),
        "formalComparatorManifestSha256": file_sha256(phase_a.PRIVATE_MANIFEST),
        "C2R1ManifestSha256": file_sha256(c2r1_runner.PRIVATE_MANIFEST),
        "inputFingerprint": input_evidence["inputFingerprint"],
        "privateAudit": private_audit,
        "publicArtifactSha256": _phase_a_public_hashes(),
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
        "C2OuterReplayExecuted": False,
        "C3Started": False,
    }
    write_json(PRIVATE_PHASE_A_MANIFEST, manifest)
    _assert_private_paths()
    return {
        "status": "passed",
        "mode": "C2-phase-a-opportunity-and-design",
        "modelPopulationCaseCount": private_audit["modelPopulationCaseCount"],
        "segmentCounts": private_audit["segmentCounts"],
        "gateCPassedConditionCount": gate["passedConditionCount"],
        "gateCConditionCount": gate["conditionCount"],
        "pendingConditions": failed,
        "C2OuterReplayExecuted": False,
        "privateFilesTracked": False,
        "finalHoldoutOpened": False,
    }


def verify_phase_a() -> dict[str, Any]:
    _assert_private_paths()
    spec = c2.load_spec()
    if any(not path.is_file() for path in (*PUBLIC_PHASE_A_PATHS, PRIVATE_PHASE_A_MANIFEST)):
        raise C2RunnerError("C2 Phase A evidence is incomplete")
    assert_public_safety(PUBLIC_PHASE_A_PATHS)
    manifest = json.loads(PRIVATE_PHASE_A_MANIFEST.read_text(encoding="utf-8"))
    if (
        manifest.get("schema") != "m2.c2_phase_a_manifest.private.v1"
        or manifest.get("specDigest") != c2.canonical_digest(spec)
        or manifest.get("publicArtifactSha256") != _phase_a_public_hashes()
        or manifest.get("C2OuterReplayExecuted") is not False
        or any(
            manifest.get(field) is not False
            for field in (
                "finalHoldoutOpened",
                "embargoShadowOpened",
                "deferred60MonthLabelsOpened",
            )
        )
    ):
        raise C2RunnerError("C2 Phase A manifest differs")
    opportunity = json.loads(OPPORTUNITY_JSON.read_text(encoding="utf-8"))
    gate = json.loads(GATE_C_JSON.read_text(encoding="utf-8"))
    if (
        opportunity["scope"]["frozenCaseCount"]
        != spec["authority"]["developmentCaseCount"]
        or opportunity["scope"]["modelPopulationCaseCount"]
        != spec["authority"]["formalModelPopulationCaseCount"]
        or gate["conditionCount"] != 14
        or gate["C2OuterReplayExecuted"] is not False
    ):
        raise C2RunnerError("C2 Phase A public contract differs")
    return {
        "status": "passed",
        "phaseAArtifactCount": len(PUBLIC_PHASE_A_PATHS),
        "modelPopulationCaseCount": opportunity["scope"]["modelPopulationCaseCount"],
        "gateCPassedConditionCount": gate["passedConditionCount"],
        "gateCConditionCount": gate["conditionCount"],
        "privateManifestTracked": False,
        "finalHoldoutOpened": False,
    }


VALIDATION_COMMANDS = (
    "npm run check:no-real-data",
    "npm run lint",
    "npm run build",
    "npm test",
    "npm run smoke",
    "npm run test:e2e",
    "npm run validate:m2:formal-cash-comparator",
    "npm run validate:m2:c2r1",
    "npm run validate:m2:c2",
)
EXPECTED_FAIL_CLOSED_COMMANDS = (
    "npm run replay:m2:formal-cash-target:final-holdout",
    "npm run replay:m2:formal-cash-comparator:final-holdout",
    "npm run replay:m2:c2:final-holdout",
)


def _validation_process(command: str) -> tuple[dict[str, Any], bytes, bytes]:
    parts = command.split()
    if not parts or parts[0] != "npm":
        raise C2RunnerError(f"unsupported validation command: {command}")
    executable = "npm.cmd" if os.name == "nt" else "npm"
    environment = os.environ.copy()
    environment.update(
        {
            "M1_APP_ENV": "ci",
            "M1_DATABASE_URL": "",
            "M1_DATABASE_READONLY_URL": "",
            "M1_DATABASE_BACKGROUND_URL": "",
        }
    )
    process = subprocess.run(
        [executable, *parts[1:]],
        cwd=ROOT,
        env=environment,
        capture_output=True,
        check=False,
        timeout=1200,
    )
    stdout = process.stdout or b""
    stderr = process.stderr or b""
    return (
        {
            "command": command,
            "exitCode": int(process.returncode),
            "stdoutSha256": hashlib.sha256(stdout).hexdigest(),
            "stderrSha256": hashlib.sha256(stderr).hexdigest(),
            "stdoutBytes": len(stdout),
            "stderrBytes": len(stderr),
        },
        stdout,
        stderr,
    )


def _execute_validation_suite() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    successes: list[dict[str, Any]] = []
    for command in VALIDATION_COMMANDS:
        progress(f"validation: {command}")
        result, stdout, stderr = _validation_process(command)
        if int(result["exitCode"]) != 0:
            diagnostic = (stderr or stdout)[-1800:].decode("utf-8", errors="replace")
            raise C2RunnerError(
                f"validation failed ({command}, exit={result['exitCode']}): {diagnostic}"
            )
        successes.append(result)
    failures: list[dict[str, Any]] = []
    for command in EXPECTED_FAIL_CLOSED_COMMANDS:
        progress(f"fail-closed validation: {command}")
        result, stdout, stderr = _validation_process(command)
        combined = (stdout + b"\n" + stderr).decode("utf-8", errors="replace").lower()
        compact = combined.replace(" ", "")
        if (
            int(result["exitCode"]) == 0
            or "final" not in combined
            or "holdout" not in combined
            or "dataloadcalls=0" not in compact
        ):
            raise C2RunnerError(
                f"final-holdout command did not fail closed before load: {command}"
            )
        failures.append(result)
    return successes, failures


def _validation_receipt(
    successes: Sequence[Mapping[str, Any]],
    failures: Sequence[Mapping[str, Any]],
    *,
    phase_a_head: str | None = None,
    phase_a_tree: str | None = None,
    remote_head: str | None = None,
    tracked_artifact_sha256: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    return {
        "schema": "m2.calibration_gate_c.validation_receipt.private.v1",
        "recordedAt": datetime.now(timezone.utc).isoformat(),
        "branch": BRANCH,
        "phaseAStartHead": PHASE_A_START_HEAD,
        "phaseAHead": phase_a_head,
        "phaseATree": phase_a_tree,
        "remoteHead": remote_head,
        "trackedArtifactSha256": dict(tracked_artifact_sha256 or {}),
        "M1_APP_ENV": "ci",
        "databaseEnvironmentEmpty": True,
        "realDataCalibrationExecutedByValidationCommands": False,
        "commandResults": list(successes),
        "expectedFailClosedCommandResults": list(failures),
        "allSuccessCommandsPassed": True,
        "allExpectedFailClosedCommandsFailedBeforeDataLoad": True,
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }


def _validate_receipt(receipt: Mapping[str, Any]) -> None:
    if (
        receipt.get("schema")
        != "m2.calibration_gate_c.validation_receipt.private.v1"
        or receipt.get("branch") != BRANCH
        or receipt.get("phaseAStartHead") != PHASE_A_START_HEAD
        or receipt.get("M1_APP_ENV") != "ci"
        or receipt.get("databaseEnvironmentEmpty") is not True
        or receipt.get("allSuccessCommandsPassed") is not True
        or receipt.get("allExpectedFailClosedCommandsFailedBeforeDataLoad") is not True
        or any(
            receipt.get(field) is not False
            for field in (
                "finalHoldoutOpened",
                "embargoShadowOpened",
                "deferred60MonthLabelsOpened",
            )
        )
    ):
        raise C2RunnerError("Gate C validation receipt boundary differs")
    successes = receipt.get("commandResults")
    failures = receipt.get("expectedFailClosedCommandResults")
    if not isinstance(successes, list) or not isinstance(failures, list):
        raise C2RunnerError("Gate C validation receipt lacks command results")
    if [item.get("command") for item in successes] != list(VALIDATION_COMMANDS):
        raise C2RunnerError("Gate C validation command set differs")
    if [item.get("command") for item in failures] != list(
        EXPECTED_FAIL_CLOSED_COMMANDS
    ):
        raise C2RunnerError("Gate C fail-closed command set differs")
    if any(int(item.get("exitCode", -1)) != 0 for item in successes):
        raise C2RunnerError("Gate C receipt records a failed validation command")
    if any(int(item.get("exitCode", 0)) == 0 for item in failures):
        raise C2RunnerError("Gate C receipt records an open final holdout command")
    required = {
        "command",
        "exitCode",
        "stdoutSha256",
        "stderrSha256",
        "stdoutBytes",
        "stderrBytes",
    }
    for item in [*successes, *failures]:
        if set(item) != required:
            raise C2RunnerError("Gate C validation result shape differs")
        if len(str(item["stdoutSha256"])) != 64 or len(
            str(item["stderrSha256"])
        ) != 64:
            raise C2RunnerError("Gate C validation output digest is invalid")
        if int(item["stdoutBytes"]) + int(item["stderrBytes"]) <= 0:
            raise C2RunnerError("Gate C validation command produced no evidence")
    if receipt.get("phaseAHead") is not None:
        if not (
            receipt.get("phaseAHead") == receipt.get("remoteHead")
            and isinstance(receipt.get("phaseATree"), str)
            and len(str(receipt["phaseATree"])) == 40
        ):
            raise C2RunnerError("Gate C runtime Git binding differs")


def finalize_gate_c_validation() -> dict[str, Any]:
    require_named_branch()
    verification = verify_phase_a()
    spec = c2.load_spec()
    opportunity = json.loads(OPPORTUNITY_JSON.read_text(encoding="utf-8"))
    synthetic = c2.synthetic_self_test()
    gate_before = _gate_report(
        spec=spec,
        opportunity=opportunity,
        synthetic=synthetic,
    )
    false_before = [key for key, value in gate_before["conditions"].items() if not value]
    if set(false_before) != {"phase_a_commit_pushed", "full_validation_suite_passed"}:
        raise C2RunnerError("Gate C was not ready for full validation")
    write_json(GATE_C_JSON, gate_before)
    successes, failures = _execute_validation_suite()
    receipt = _validation_receipt(successes, failures)
    _validate_receipt(receipt)
    write_json(PRIVATE_VALIDATION_RECEIPT, receipt)
    if run_git("ls-files", "--", PRIVATE_VALIDATION_RECEIPT.relative_to(ROOT).as_posix()):
        raise C2RunnerError("private Gate C validation receipt entered Git")
    receipt_sha = file_sha256(PRIVATE_VALIDATION_RECEIPT)
    gate = _gate_report(
        spec=spec,
        opportunity=opportunity,
        synthetic=synthetic,
        validation_passed=True,
        validation_receipt_sha256=receipt_sha,
        validation_evidence=receipt,
    )
    failed = [key for key, value in gate["conditions"].items() if not value]
    if failed != ["phase_a_commit_pushed"]:
        raise C2RunnerError("Gate C did not reduce to the push condition")
    write_json(GATE_C_JSON, gate)
    assert_public_safety(PUBLIC_PHASE_A_PATHS)
    manifest = json.loads(PRIVATE_PHASE_A_MANIFEST.read_text(encoding="utf-8"))
    manifest["publicArtifactSha256"] = _phase_a_public_hashes()
    write_json(PRIVATE_PHASE_A_MANIFEST, manifest)
    return {
        "status": "passed",
        "phaseAContentVerified": verification["status"] == "passed",
        "validationCommandCount": len(successes),
        "failClosedCommandCount": len(failures),
        "gateCPassedConditionCount": gate["passedConditionCount"],
        "gateCConditionCount": gate["conditionCount"],
        "onlyPendingCondition": "phase_a_commit_pushed",
        "validationReceiptSha256": receipt_sha,
        "privateReceiptTracked": False,
        "C2AuthorizedByGateC": False,
        "finalHoldoutOpened": False,
    }


def _commit_artifact_digests(commit: str) -> dict[str, str]:
    output: dict[str, str] = {}
    for path in PHASE_A_TRACKED_PATHS:
        relative = path.relative_to(ROOT).as_posix()
        process = subprocess.run(
            ["git", "show", f"{commit}:{relative}"],
            cwd=ROOT,
            capture_output=True,
            check=False,
        )
        if process.returncode != 0:
            raise C2RunnerError(f"Phase A commit lacks tracked artifact: {relative}")
        output[relative] = hashlib.sha256(process.stdout).hexdigest()
    return output


def _remote_branch_head() -> str:
    output = run_git("ls-remote", "--heads", "origin", f"refs/heads/{BRANCH}")
    fields = output.split()
    if len(fields) != 2 or fields[1] != f"refs/heads/{BRANCH}":
        raise C2RunnerError("authorized C2 remote branch is unavailable")
    return fields[0]


def verify_gate_c_after_push() -> dict[str, Any]:
    require_named_branch()
    if run_git("status", "--porcelain=v1", "--untracked-files=all"):
        raise C2RunnerError("Gate C push verification requires a clean worktree")
    head = run_git("rev-parse", "HEAD")
    if run_git("rev-parse", "HEAD^") != PHASE_A_START_HEAD:
        raise C2RunnerError("C2 Phase A checkpoint parent differs from authorized HEAD")
    upstream = run_git(
        "rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{upstream}"
    )
    if upstream != f"origin/{BRANCH}":
        raise C2RunnerError("C2 upstream differs from the authorized branch")
    remote_head = _remote_branch_head()
    if head != run_git("rev-parse", "@{upstream}") or head != remote_head:
        raise C2RunnerError("C2 Phase A checkpoint is not the current remote head")
    verification = verify_phase_a()
    if not PRIVATE_VALIDATION_RECEIPT.is_file():
        raise C2RunnerError("private Gate C validation receipt is missing")
    precommit_receipt = json.loads(
        PRIVATE_VALIDATION_RECEIPT.read_text(encoding="utf-8")
    )
    _validate_receipt(precommit_receipt)
    precommit_sha = file_sha256(PRIVATE_VALIDATION_RECEIPT)
    tracked_gate = json.loads(GATE_C_JSON.read_text(encoding="utf-8"))
    if tracked_gate.get("validationReceiptSha256") != precommit_sha:
        raise C2RunnerError("tracked Gate C differs from its validation receipt")
    if [key for key, value in tracked_gate["conditions"].items() if not value] != [
        "phase_a_commit_pushed"
    ]:
        raise C2RunnerError("tracked Gate C was not ready for push attestation")
    progress("re-running the full Gate C suite on the remote-confirmed commit")
    successes, failures = _execute_validation_suite()
    artifact_digests = _commit_artifact_digests(head)
    runtime_receipt = _validation_receipt(
        successes,
        failures,
        phase_a_head=head,
        phase_a_tree=run_git("rev-parse", "HEAD^{tree}"),
        remote_head=remote_head,
        tracked_artifact_sha256=artifact_digests,
    )
    _validate_receipt(runtime_receipt)
    write_json(PRIVATE_VALIDATION_RECEIPT, runtime_receipt)
    runtime_validation_sha = file_sha256(PRIVATE_VALIDATION_RECEIPT)
    push_receipt = {
        "schema": "m2.calibration_gate_c.push_receipt.private.v1",
        "recordedAt": datetime.now(timezone.utc).isoformat(),
        "branch": BRANCH,
        "upstream": upstream,
        "phaseAHead": head,
        "phaseAParent": PHASE_A_START_HEAD,
        "remoteHead": remote_head,
        "phaseATree": run_git("rev-parse", "HEAD^{tree}"),
        "trackedArtifactSha256": artifact_digests,
        "precommitValidationReceiptSha256": precommit_sha,
        "runtimeValidationReceiptSha256": runtime_validation_sha,
        "runtimeValidationReexecutedOnRemoteConfirmedTree": True,
        "phaseAContentVerified": verification["status"] == "passed",
        "privateFilesTracked": False,
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }
    write_json(PRIVATE_PUSH_RECEIPT, push_receipt)
    if run_git("ls-files", "--", PRIVATE_PUSH_RECEIPT.relative_to(ROOT).as_posix()):
        raise C2RunnerError("private Gate C push receipt entered Git")
    push_sha = file_sha256(PRIVATE_PUSH_RECEIPT)
    spec = c2.load_spec()
    opportunity = json.loads(OPPORTUNITY_JSON.read_text(encoding="utf-8"))
    gate = _gate_report(
        spec=spec,
        opportunity=opportunity,
        synthetic=c2.synthetic_self_test(),
        validation_passed=True,
        validation_receipt_sha256=runtime_validation_sha,
        validation_evidence=runtime_receipt,
        phase_a_commit_pushed=True,
        phase_a_checkpoint=head,
        remote_head_verified=True,
    )
    gate["runtimePushReceiptSha256"] = push_sha
    if not gate["allTrue"] or gate["conditionCount"] != 14:
        failed = [key for key, value in gate["conditions"].items() if not value]
        raise C2RunnerError("Gate C is not all true after push: " + ", ".join(failed))
    write_json(GATE_C_JSON, gate)
    assert_public_safety((GATE_C_JSON,))
    return {
        "status": "passed",
        "gateCAllTrue": True,
        "gateCPassedConditionCount": 14,
        "gateCConditionCount": 14,
        "phaseAHead": head,
        "remoteHead": remote_head,
        "C2AuthorizedByGateC": True,
        "runtimePushReceiptSha256": push_sha,
        "privateReceiptTracked": False,
        "finalHoldoutOpened": False,
    }


def _require_current_phase_a_sources_match_commit(
    committed: Mapping[str, Any]
) -> None:
    for path in IMMUTABLE_PHASE_A_PATHS:
        relative = path.relative_to(ROOT).as_posix()
        if not path.is_file() or file_sha256(path) != committed.get(relative):
            raise C2RunnerError(
                f"current C2 Phase A dependency differs from pushed checkpoint: {relative}"
            )


def verify_c2_authorization() -> dict[str, Any]:
    require_named_branch()
    _assert_private_paths()
    if not GATE_C_JSON.is_file() or not PRIVATE_PUSH_RECEIPT.is_file():
        raise C2RunnerError("Gate C push evidence is incomplete")
    gate = json.loads(GATE_C_JSON.read_text(encoding="utf-8"))
    receipt = json.loads(PRIVATE_PUSH_RECEIPT.read_text(encoding="utf-8"))
    if receipt.get("schema") != "m2.calibration_gate_c.push_receipt.private.v1":
        raise C2RunnerError("Gate C push receipt identity differs")
    _require_current_phase_a_sources_match_commit(
        receipt.get("trackedArtifactSha256", {})
    )
    if (
        gate.get("allTrue") is not True
        or gate.get("C2AuthorizedByGateC") is not True
        or len(gate.get("conditions", {})) != 14
        or any(value is not True for value in gate["conditions"].values())
        or gate.get("runtimePushReceiptSha256") != file_sha256(PRIVATE_PUSH_RECEIPT)
        or gate.get("phaseACheckpoint") != receipt.get("phaseAHead")
        or receipt.get("remoteHead") != receipt.get("phaseAHead")
    ):
        raise C2RunnerError("Gate C authorization differs")
    return {
        "status": "passed",
        "gateCAllTrue": True,
        "phaseACheckpoint": receipt["phaseAHead"],
        "privateReceiptTracked": False,
        "finalHoldoutOpened": False,
    }


def preflight() -> dict[str, Any]:
    spec = c2.load_spec()
    synthetic = c2.synthetic_self_test()
    branch = run_git("branch", "--show-current")
    checkout_boundary = "named_branch" if branch == BRANCH else "synthetic_ci_checkout"
    return {
        "status": "passed",
        "mode": "synthetic-only",
        "checkoutBoundary": checkout_boundary,
        "specDigest": c2.canonical_digest(spec),
        "candidateCounts": synthetic["candidateCounts"],
        "synthetic": synthetic,
        "privateDataRead": False,
        "dataLoadCalls": 0,
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }


def _prediction_case_state(template: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "caseKey": copy.deepcopy(template["caseKey"]),
        "statisticallyScoreable": bool(template["statisticallyScoreable"]),
        "scoreabilityReason": template.get("scoreabilityReason"),
        "businessServingEligible": bool(template["businessServingEligible"]),
        "abstentionReason": template.get("abstentionReason"),
        "targetEnd": template["targetEnd"],
        "labelAvailableAsOf": template["labelAvailableAsOf"],
        "billMonthMax": template["billMonthMax"],
        "sourceAvailableAsOf": template["sourceAvailableAsOf"],
        "predictionRole": template["predictionRole"],
    }


def _compact_prediction_projection(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "modelId": row["model_id"],
        "candidateId": row["candidate_id"],
        "requestedCandidateId": row["requested_candidate_id"],
        "caseKey": list(formal.strict_case_key(row)),
        "predictionRole": row["_residual_case_role"],
        "activitySegment": row["activity_segment"],
        "segmentReason": row["segment_reason"],
        "modelOverrideAllowed": row["modelOverrideAllowed"],
        "highValueGuardFallbackToB4": row["highValueGuardFallbackToB4"],
        "statisticallyScoreable": row["statisticallyScoreable"],
        "businessServingEligible": row["businessServingEligible"],
        "modelPredictionAvailable": row["modelPredictionAvailable"],
        "routeAbstained": row["routeAbstained"],
        "abstained": row["abstained"],
        "abstentionReason": row["abstentionReason"],
        "rawModelPrediction": row["rawModelPrediction"],
        "servedPrediction": row["servedPrediction"],
        "otherOrNewChannelResidualPoint": row[
            "otherOrNewChannelResidualPoint"
        ],
        "channelComponents": sorted(
            (
                str(item["channel_key"]),
                str(item.get("component_type", "known_as_of_channel")),
                float(item["point_forecast"]),
            )
            for item in row.get("channel_components", [])
        ),
        "publicOutput": row["public_output"],
    }


def _join_truth(
    row: dict[str, Any], template: Mapping[str, Any]
) -> dict[str, Any]:
    joined = c2r1_runner._join_truth(row, template)
    joined["strata"]["activity_segment"] = row["activity_segment"]
    joined["strata"]["segment_reason"] = row["segment_reason"]
    return joined


def _locked_b4_rows(
    forward_templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]]
) -> list[dict[str, Any]]:
    return [
        c2r1_runner._locked_b4_row(template)
        for _key, template in sorted(forward_templates.items())
    ]


def _earlier_key_allowed(
    key: tuple[str, str, int, str],
    template: Mapping[str, Any],
    outer_origin: str,
) -> bool:
    return (
        key[1] < outer_origin
        and str(template["targetEnd"]) <= outer_origin
        and str(template["labelAvailableAsOf"]) <= outer_origin
    )


def _training_records(
    *,
    keys: Sequence[tuple[str, str, int, str]],
    forward_templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    comparator: Mapping[
        tuple[str, tuple[str, str, int, str], str], Mapping[str, Any]
    ],
    segment_states: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    outer_origin: str,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for key in keys:
        template = forward_templates[key]
        if not _earlier_key_allowed(key, template, outer_origin):
            continue
        role = f"development_forward_score:{key[1]}"
        comparator_rows = c2r1_runner._comparator_rows_for_key(
            comparator, role, key
        )
        context = c2.channel_context(comparator_rows)
        known = set(
            str(item["channel_key"])
            for item in comparator_rows["B4"]["channelComponents"]
        )
        actual_by_component = template.get("forecastableActualByComponent", {}) or {}
        matched = sum(
            float(value)
            for component_key, value in actual_by_component.items()
            if str(component_key) in known
        )
        residual_actual = sum(
            float(value)
            for component_key, value in actual_by_component.items()
            if str(component_key) not in known
        )
        records.append(
            {
                "origin": key[1],
                "labelAvailableAsOf": str(template["labelAvailableAsOf"]),
                "targetEnd": str(template["targetEnd"]),
                "route": key[3],
                "segment": segment_states[key]["segment"],
                "horizon": key[2],
                "knownChannelCountBucket": context["knownChannelCountBucket"],
                "knownChannelConcentrationBucket": context[
                    "knownChannelConcentrationBucket"
                ],
                "residualActual": residual_actual,
                "matchedKnownActual": matched,
            }
        )
    return records


def _metric_for_points(
    keys: Sequence[tuple[str, str, int, str]],
    points: Mapping[tuple[str, str, int, str], Mapping[str, float]],
    candidate_id: str,
    templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
) -> dict[str, Any]:
    if not keys:
        return {
            "caseCount": 0,
            "wape": None,
            "signedAggregateBias": None,
            "horizons": {},
            "top10": None,
        }
    actuals = [float(templates[key]["forecastableCashActual"]) for key in keys]
    predictions = [float(points[key][candidate_id]) for key in keys]

    def values(group: Sequence[tuple[str, str, int, str]]) -> dict[str, Any] | None:
        if not group:
            return None
        group_actual = [float(templates[key]["forecastableCashActual"]) for key in group]
        group_prediction = [float(points[key][candidate_id]) for key in group]
        return {
            "caseCount": len(group),
            "wape": base.wape(group_prediction, group_actual),
            "signedAggregateBias": base.signed_aggregate_bias(
                group_prediction, group_actual
            ),
        }

    top10_keys = [
        key
        for key in keys
        if bool((templates[key].get("strata", {}) or {}).get("top_10_percent"))
    ]
    return {
        "caseCount": len(keys),
        "wape": base.wape(predictions, actuals),
        "signedAggregateBias": base.signed_aggregate_bias(predictions, actuals),
        "horizons": {
            str(horizon): values([key for key in keys if key[2] == horizon])
            for horizon in formal.CORE_HORIZONS
        },
        "top10": values(top10_keys),
    }


def _relative(candidate: float | None, reference: float | None) -> float | None:
    if candidate is None or reference is None:
        return None
    if float(reference) == 0:
        return 0.0 if float(candidate) == 0 else None
    return (float(candidate) - float(reference)) / float(reference)


def _select_candidate(
    *,
    outer_origin: str,
    segment: str,
    spec: Mapping[str, Any],
    points: Mapping[tuple[str, str, int, str], Mapping[str, float]],
    templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    segment_states: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
) -> tuple[str, dict[str, Any]]:
    earlier = [
        key
        for key in sorted(points)
        if segment_states[key]["segment"] == segment
        and _earlier_key_allowed(key, templates[key], outer_origin)
    ]
    earlier_origins = sorted({key[1] for key in earlier})
    minimum_origins = int(spec["selection"]["minimumEarlierOrigins"])
    minimum_cases = int(spec["selection"]["minimumEarlierSegmentCases"])
    base_evidence = {
        "outerOrigin": outer_origin,
        "segment": segment,
        "earlierOriginCount": len(earlier_origins),
        "earlierOrigins": earlier_origins,
        "earlierCaseCount": len(earlier),
        "maximumEarlierLabelAvailableAsOf": max(
            (str(templates[key]["labelAvailableAsOf"]) for key in earlier),
            default=None,
        ),
        "sameOrLaterOuterTruthRead": False,
        "candidateSpaceCount": len(c2.candidate_ids(spec, segment)),
        "weightedObjectiveUsed": False,
        "thresholdMoved": False,
    }
    if len(earlier_origins) < minimum_origins or len(earlier) < minimum_cases:
        return "B4", {
            **base_evidence,
            "selectedCandidate": "B4",
            "selectionReason": "insufficient_earlier_evidence_fallback",
            "biasFeasibleCandidateCount": 0,
            "wapeEquivalentCandidateCount": 0,
            "highValueSafeCandidateCount": 0,
            "horizonSafeCandidateCount": 0,
        }
    metrics = {
        candidate_id: _metric_for_points(
            earlier, points, candidate_id, templates
        )
        for candidate_id in c2.candidate_ids(spec, segment)
    }
    bias_rule = spec["selection"]["biasFeasibility"]
    feasible = []
    for candidate_id, metric in metrics.items():
        horizon_biases = [
            abs(float(value["signedAggregateBias"]))
            for value in metric["horizons"].values()
            if value is not None
        ]
        if (
            abs(float(metric["signedAggregateBias"]))
            <= float(bias_rule["overallAbsoluteMaximum"]) + TOLERANCE
            and all(
                value <= float(bias_rule["eachHorizonAbsoluteMaximum"]) + TOLERANCE
                for value in horizon_biases
            )
        ):
            feasible.append(candidate_id)
    if not feasible:
        return "B4", {
            **base_evidence,
            "selectedCandidate": "B4",
            "selectionReason": "no_bias_feasible_candidate_fallback",
            "biasFeasibleCandidateCount": 0,
            "wapeEquivalentCandidateCount": 0,
            "highValueSafeCandidateCount": 0,
            "horizonSafeCandidateCount": 0,
        }
    minimum_wape = min(float(metrics[value]["wape"]) for value in feasible)
    wape_pool = [
        value
        for value in feasible
        if float(metrics[value]["wape"])
        <= minimum_wape
        * (1.0 + float(spec["selection"]["wapePracticalEquivalenceRelativeMaximum"]))
        + TOLERANCE
    ]
    b4_metric = metrics["B4"]
    high_value_pool = []
    for candidate_id in wape_pool:
        candidate_top = metrics[candidate_id]["top10"]
        b4_top = b4_metric["top10"]
        if candidate_top is None or b4_top is None:
            continue
        if (
            float(candidate_top["wape"])
            <= float(b4_top["wape"])
            * (1.0 + float(spec["selection"]["highValueRelativeWapeRegressionMaximum"]))
            + TOLERANCE
            and abs(float(candidate_top["signedAggregateBias"]))
            <= abs(float(b4_top["signedAggregateBias"]))
            + float(spec["selection"]["highValueAbsoluteBiasWorseningMaximum"])
            + TOLERANCE
        ):
            high_value_pool.append(candidate_id)
    after_high = high_value_pool or wape_pool
    horizon_pool = []
    for candidate_id in after_high:
        safe = True
        for horizon in ("3", "6", "12", "18", "24"):
            candidate_h = metrics[candidate_id]["horizons"].get(horizon)
            b4_h = b4_metric["horizons"].get(horizon)
            relative = _relative(
                candidate_h["wape"] if candidate_h else None,
                b4_h["wape"] if b4_h else None,
            )
            if relative is not None and relative > float(
                spec["selection"]["eachHorizonRelativeWapeRegressionMaximum"]
            ) + TOLERANCE:
                safe = False
                break
        if safe:
            horizon_pool.append(candidate_id)
    final_pool = horizon_pool or after_high
    selected = min(
        final_pool,
        key=lambda candidate_id: c2.candidate_complexity(
            spec, segment, candidate_id
        ),
    )
    return selected, {
        **base_evidence,
        "selectedCandidate": selected,
        "selectionReason": "frozen_lexicographic_objective",
        "biasFeasibleCandidateCount": len(feasible),
        "minimumFeasibleWape": minimum_wape,
        "wapeEquivalentCandidateCount": len(wape_pool),
        "highValueSafeCandidateCount": len(high_value_pool),
        "horizonSafeCandidateCount": len(horizon_pool),
        "selectedEvidence": public_value(metrics[selected]),
        "B4Evidence": public_value(metrics["B4"]),
    }


def _high_value_guard(
    *,
    outer_origin: str,
    segment: str,
    selected_candidate: str,
    spec: Mapping[str, Any],
    points: Mapping[tuple[str, str, int, str], Mapping[str, float]],
    templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    segment_states: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
) -> tuple[bool, dict[str, Any]]:
    rule = spec["highValueGuard"]
    earlier = [
        key
        for key in sorted(points)
        if segment_states[key]["segment"] == segment
        and bool((templates[key].get("strata", {}) or {}).get("top_10_percent"))
        and _earlier_key_allowed(key, templates[key], outer_origin)
    ]
    origins = sorted({key[1] for key in earlier})
    evidence: dict[str, Any] = {
        "outerOrigin": outer_origin,
        "segment": segment,
        "selectedCandidate": selected_candidate,
        "earlierOriginCount": len(origins),
        "earlierTop10CaseCount": len(earlier),
        "outerTopBandActualUsed": False,
        "stabilityRuleFrozen": True,
        "fallbackRuleChangedAfterResults": False,
    }
    if selected_candidate == "B4":
        return False, {
            **evidence,
            "overrideAllowed": False,
            "reason": "B4_already_selected",
            "conditions": {},
        }
    if (
        len(origins) < int(rule["minimumEarlierOrigins"])
        or len(earlier) < int(rule["minimumEarlierTop10Cases"])
    ):
        return False, {
            **evidence,
            "overrideAllowed": False,
            "reason": "insufficient_earlier_top10_evidence",
            "conditions": {},
        }
    candidate = _metric_for_points(
        earlier, points, selected_candidate, templates
    )
    b4_metric = _metric_for_points(earlier, points, "B4", templates)
    improvement = -float(_relative(candidate["wape"], b4_metric["wape"]) or 0.0)
    horizon_safe = all(
        (
            _relative(
                candidate["horizons"][horizon]["wape"]
                if candidate["horizons"].get(horizon)
                else None,
                b4_metric["horizons"][horizon]["wape"]
                if b4_metric["horizons"].get(horizon)
                else None,
            )
            or 0.0
        )
        <= float(rule["eachHorizonRelativeWapeRegressionMaximum"]) + TOLERANCE
        for horizon in ("3", "6", "12", "18", "24")
    )
    origin_results = {}
    for origin in origins:
        group = [key for key in earlier if key[1] == origin]
        candidate_origin = _metric_for_points(
            group, points, selected_candidate, templates
        )
        b4_origin = _metric_for_points(group, points, "B4", templates)
        origin_results[origin] = _relative(
            candidate_origin["wape"], b4_origin["wape"]
        )
    origin_win_share = sum(
        value is not None and float(value) < -TOLERANCE
        for value in origin_results.values()
    ) / len(origin_results)
    conditions = {
        "wapeActuallyImproves": improvement
        >= float(rule["minimumRelativeWapeImprovement"]) - TOLERANCE,
        "absoluteSignedBiasDoesNotWorsen": abs(
            float(candidate["signedAggregateBias"])
        )
        <= abs(float(b4_metric["signedAggregateBias"])) + TOLERANCE,
        "horizonsDoNotMateriallyRegress": horizon_safe,
        "earlierOriginStabilityPasses": origin_win_share
        >= float(rule["minimumEarlierOriginWinShare"]) - TOLERANCE,
    }
    allowed = all(conditions.values())
    return allowed, {
        **evidence,
        "overrideAllowed": allowed,
        "reason": "all_frozen_conditions_passed" if allowed else "guard_fallback_to_B4",
        "conditions": conditions,
        "relativeWapeImprovement": improvement,
        "candidateAbsoluteSignedBias": abs(
            float(candidate["signedAggregateBias"])
        ),
        "B4AbsoluteSignedBias": abs(float(b4_metric["signedAggregateBias"])),
        "earlierOriginWinShare": origin_win_share,
        "originRelativeWape": public_value(origin_results),
    }


def _materialize_candidate_points(
    *,
    spec: Mapping[str, Any],
    calibration_spec: Mapping[str, Any],
    works: Mapping[str, Mapping[str, Any]],
    forward_templates: Mapping[tuple[str, str, int, str], Mapping[str, Any]],
    comparator: Mapping[
        tuple[str, tuple[str, str, int, str], str], Mapping[str, Any]
    ],
) -> tuple[
    dict[tuple[str, str, int, str], dict[str, float]],
    dict[tuple[str, str, int, str], dict[str, Any]],
    dict[str, dict[str, Any]],
    dict[str, dict[str, Any]],
]:
    model_keys = [
        key
        for key, template in sorted(forward_templates.items())
        if template["statisticallyScoreable"] is True
        and template["modelPredictionAvailable"] is True
        and template["routeAbstained"] is False
    ]
    segment_states = {
        key: c2.segment_as_of(
            works[key[0]], key[1], calibration_spec, spec
        )
        for key in model_keys
    }
    residual_models: dict[str, dict[str, Any]] = {}
    reactivation_models: dict[str, dict[str, Any]] = {}
    points: dict[tuple[str, str, int, str], dict[str, float]] = {}
    for origin in spec["authority"]["origins"]:
        training = _training_records(
            keys=model_keys,
            forward_templates=forward_templates,
            comparator=comparator,
            segment_states=segment_states,
            outer_origin=origin,
        )
        residual_models[origin] = c2.fit_residual_model(training, origin, spec)
        reactivation_models[origin] = c2.fit_reactivation_model(training, origin, spec)
        origin_keys = [key for key in model_keys if key[1] == origin]
        progress(
            f"materializing frozen candidate points for origin {origin} "
            f"({len(origin_keys)} cases)"
        )
        for index, key in enumerate(origin_keys, start=1):
            template = forward_templates[key]
            role = f"development_forward_score:{origin}"
            comparator_rows = c2r1_runner._comparator_rows_for_key(
                comparator, role, key
            )
            segment = str(segment_states[key]["segment"])
            candidate_points: dict[str, float] = {}
            for candidate_id in c2.candidate_ids(spec, segment):
                prediction = c2.predict_as_of(
                    work=works[key[0]],
                    origin=origin,
                    horizon=key[2],
                    case_state=_prediction_case_state(template),
                    comparator_rows=comparator_rows,
                    calibration_spec=calibration_spec,
                    spec=spec,
                    candidate_id=candidate_id,
                    residual_model=residual_models[origin],
                    reactivation_model=reactivation_models[origin],
                    cutoff_top10=False,
                    high_value_override_allowed=True,
                )
                if prediction["rawModelPrediction"] is None:
                    raise C2RunnerError("C2 candidate model-population point is null")
                candidate_points[candidate_id] = float(
                    prediction["rawModelPrediction"]
                )
            if abs(
                candidate_points["B4"] - float(template["rawModelPrediction"])
            ) > 0.000001:
                raise C2RunnerError("C2 B4 anchor differs from the locked comparator")
            points[key] = candidate_points
            if index % 500 == 0:
                progress(f"origin {origin} candidate cases: {index}/{len(origin_keys)}")
    if len(points) != int(spec["authority"]["formalModelPopulationCaseCount"]):
        raise C2RunnerError("C2 candidate point population differs")
    return points, segment_states, residual_models, reactivation_models


def _segment_metrics(
    rows: Sequence[Mapping[str, Any]], spec: Mapping[str, Any]
) -> dict[str, Any]:
    minimum = int(spec["privacy"]["complementarySuppressionMinimumWorks"])
    population = [row for row in rows if formal.is_model_population(row)]
    output: dict[str, Any] = {}
    for segment in c2.ACTIVITY_SEGMENTS:
        group = [row for row in population if row["activity_segment"] == segment]
        work_count = len({formal.strict_case_key(row)[0] for row in group})
        if work_count < minimum:
            output[segment] = {
                "suppressed": True,
                "suppressionReason": "complementary_small_sample",
                "minimumWorks": minimum,
                "caseCount": None,
                "uniqueWorkCount": None,
                "wape": None,
                "signedAggregateBias": None,
            }
        else:
            output[segment] = public_value(
                formal.metric_rows(group, "rawModelPrediction")
            )
    return output


def _channel_residual_audit(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    population = [row for row in rows if formal.is_model_population(row)]
    maximum_reconciliation = 0.0
    maximum_truth_reconciliation = 0.0
    matched_cash = 0.0
    total_cash = 0.0
    truth_without_component_cash = 0.0
    predicted_residual_cash = 0.0
    predicted_known_cash = 0.0
    residual_positive_cases = 0
    generic_component_count = 0
    actual_uses_generic_key = False
    for row in population:
        known: dict[str, float] = {}
        generic = None
        for component in row.get("channel_components", []) or []:
            key = str(component["channel_key"])
            point = float(component["point_forecast"])
            if component.get("component_type") == "other_or_new_channel_residual":
                if generic is not None or key != c2.GENERIC_RESIDUAL_KEY:
                    raise C2RunnerError("C2 generic residual component differs")
                generic = point
                generic_component_count += 1
            else:
                if key in known:
                    raise C2RunnerError("duplicate C2 known channel component")
                known[key] = point
        if generic is None:
            raise C2RunnerError("C2 model-population case lacks generic residual component")
        confirmed = sum(
            float(item["outstandingAmount"])
            for item in row.get("confirmedCashComponents", []) or []
        )
        maximum_reconciliation = max(
            maximum_reconciliation,
            abs(sum(known.values()) + generic + confirmed - float(row["rawModelPrediction"])),
        )
        actual = row.get("forecastableActualByComponent", {}) or {}
        actual_uses_generic_key = actual_uses_generic_key or c2.GENERIC_RESIDUAL_KEY in actual
        maximum_truth_reconciliation = max(
            maximum_truth_reconciliation,
            abs(sum(float(value) for value in actual.values()) - float(row["forecastableCashActual"])),
        )
        matched_cash += sum(float(actual[key]) for key in known if key in actual)
        truth_without_component_cash += sum(
            float(value) for key, value in actual.items() if str(key) not in known
        )
        total_cash += float(row["forecastableCashActual"])
        predicted_known_cash += sum(known.values())
        predicted_residual_cash += generic
        residual_positive_cases += int(generic > TOLERANCE)
    return {
        "modelPopulationCaseCount": len(population),
        "genericComponentCount": generic_component_count,
        "residualPositiveCaseCount": residual_positive_cases,
        "knownChannelPredictedCash": predicted_known_cash,
        "otherOrNewChannelResidualPredictedCash": predicted_residual_cash,
        "matchedForecastableCash": matched_cash,
        "matchedForecastableCashShare": ratio(matched_cash, total_cash),
        "truthWithoutPredictionComponentCash": truth_without_component_cash,
        "maximumWorkPointReconciliationDifference": maximum_reconciliation,
        "maximumTruthComponentReconciliationDifference": maximum_truth_reconciliation,
        "knownChannelCashDuplicated": actual_uses_generic_key,
        "futureChannelIdentityPredicted": False,
        "outerTruthUsed": False,
        "genericComponentIsARealChannel": False,
        "workPointFormulaVerified": maximum_reconciliation <= TOLERANCE,
    }


def _acceptance(
    *,
    metrics: Mapping[str, Any],
    comparator: Mapping[str, Any],
    residual_audit: Mapping[str, Any],
    guard_audit: Mapping[str, Any],
    spec: Mapping[str, Any],
) -> dict[str, Any]:
    gate = spec["acceptance"]
    relative = gate["relativeToB4"]
    horizon_relative = {
        horizon: _relative(
            metrics["horizons"][horizon]["wape"],
            comparator["horizons"][horizon]["wape"],
        )
        for horizon in ("3", "6", "12", "18", "24")
    }
    top_relative = {
        top: _relative(
            metrics["topBands"][top]["wape"],
            comparator["topBands"][top]["wape"],
        )
        for top in ("top1", "top5", "top10")
    }
    origins = list(spec["authority"]["origins"])
    origin_relative = {
        origin: _relative(
            metrics["origins"][origin]["wape"],
            comparator["origins"][origin]["wape"],
        )
        for origin in origins
    }
    origin_win_share = sum(
        value is not None and float(value) < -TOLERANCE
        for value in origin_relative.values()
    ) / len(origin_relative)
    consecutive = 0
    maximum_consecutive = 0
    for origin in origins:
        if float(origin_relative[origin] or 0.0) > 0.05 + TOLERANCE:
            consecutive += 1
            maximum_consecutive = max(maximum_consecutive, consecutive)
        else:
            consecutive = 0
    interval = metrics["internal80"]
    comparator_interval = comparator["internal80"]
    wis_improvement = (
        (float(comparator_interval["meanWis"]) - float(interval["meanWis"]))
        / float(comparator_interval["meanWis"])
        if interval["meanWis"] is not None
        and comparator_interval["meanWis"] not in {None, 0}
        else None
    )
    width_relative = _relative(
        interval["standardizedWidth"], comparator_interval["standardizedWidth"]
    )
    conditions = {
        "overall_wape_at_most_0_60": float(metrics["modelPopulation"]["wape"])
        <= float(gate["overallWapeMaximum"]) + TOLERANCE,
        "overall_absolute_bias_at_most_10pct": abs(
            float(metrics["modelPopulation"]["signedAggregateBias"])
        )
        <= float(gate["absoluteBiasMaximum"]["overall"]) + TOLERANCE,
        "served_absolute_bias_at_most_10pct": abs(
            float(metrics["served"]["signedAggregateBias"])
        )
        <= float(gate["absoluteBiasMaximum"]["served"]) + TOLERANCE,
        "high_value_absolute_bias_at_most_10pct": abs(
            float(metrics["highValue"]["signedAggregateBias"])
        )
        <= float(gate["absoluteBiasMaximum"]["highValue"]) + TOLERANCE,
        "each_horizon_absolute_bias_at_most_15pct": all(
            abs(float(metrics["horizons"][horizon]["signedAggregateBias"]))
            <= float(gate["absoluteBiasMaximum"]["eachHorizon"]) + TOLERANCE
            for horizon in ("3", "6", "12", "18", "24")
        ),
        "horizon_3_wape_improves_at_least_3pct": -float(horizon_relative["3"])
        >= float(relative["horizon3ImprovementMinimum"]) - TOLERANCE,
        "horizon_6_wape_improves_at_least_3pct": -float(horizon_relative["6"])
        >= float(relative["horizon6ImprovementMinimum"]) - TOLERANCE,
        "horizon_12_wape_improves_at_least_3pct": -float(horizon_relative["12"])
        >= float(relative["horizon12ImprovementMinimum"]) - TOLERANCE,
        "horizon_18_wape_regression_at_most_2pct": float(horizon_relative["18"])
        <= float(relative["horizon18RegressionMaximum"]) + TOLERANCE,
        "horizon_24_wape_regression_at_most_2pct": float(horizon_relative["24"])
        <= float(relative["horizon24RegressionMaximum"]) + TOLERANCE,
        "top10_wape_improves_at_least_5pct": -float(top_relative["top10"])
        >= float(relative["top10ImprovementMinimum"]) - TOLERANCE,
        "top1_wape_regression_at_most_5pct": float(top_relative["top1"])
        <= float(relative["top1RegressionMaximum"]) + TOLERANCE,
        "top5_wape_regression_at_most_5pct": float(top_relative["top5"])
        <= float(relative["top5RegressionMaximum"]) + TOLERANCE,
        "at_least_70pct_origins_beat_B4": origin_win_share
        >= float(relative["outerOriginWinShareMinimum"]) - TOLERANCE,
        "no_three_consecutive_origins_regress_over_5pct": maximum_consecutive
        <= int(relative["maximumConsecutiveOriginsRegressingOverFivePercent"]),
        "internal_80_coverage_between_75_and_85pct": interval[
            "completeOnModelPopulation"
        ]
        is True
        and float(gate["internal80CoverageInclusive"][0]) - TOLERANCE
        <= float(interval["internal80Coverage"])
        <= float(gate["internal80CoverageInclusive"][1]) + TOLERANCE,
        "internal_WIS_improves_at_least_5pct": wis_improvement is not None
        and wis_improvement
        >= float(relative["internalWisImprovementMinimum"]) - TOLERANCE,
        "standardized_width_regression_at_most_10pct": width_relative is not None
        and width_relative
        <= float(relative["standardizedWidthRegressionMaximum"]) + TOLERANCE,
        "P0_equals_0": int(gate["P0Maximum"]) == 0,
        "P1_equals_0": int(gate["P1Maximum"]) == 0,
        "P2_is_fact_audit_only": gate["P2FactAuditOnly"] is True,
        "no_automatic_operational_actions": int(
            gate["automaticOperationalActionCountMaximum"]
        )
        == 0,
        "residual_does_not_duplicate_cash": residual_audit[
            "knownChannelCashDuplicated"
        ]
        is False
        and residual_audit["workPointFormulaVerified"] is True,
        "high_value_guard_active": guard_audit["guardActiveOnEveryTop10Case"]
        is True,
        "model_population_unchanged": int(metrics["modelPopulation"]["caseCount"])
        == int(spec["authority"]["formalModelPopulationCaseCount"])
        and int(metrics["modelPopulation"]["uniqueWorkCount"])
        == int(spec["authority"]["formalModelPopulationWorkCount"]),
    }
    if len(conditions) != int(gate["conditionCount"]):
        raise C2RunnerError("C2 acceptance condition count differs")
    passed = sum(conditions.values())
    return {
        "conditions": conditions,
        "passedConditionCount": passed,
        "conditionCount": len(conditions),
        "allPassed": passed == len(conditions),
        "modelQualityDecision": "PASS" if passed == len(conditions) else "FAIL",
        "thresholdsChangedAfterResults": False,
        "evidence": {
            "horizonRelativeWapeVsB4": public_value(horizon_relative),
            "topBandRelativeWapeVsB4": public_value(top_relative),
            "originRelativeWapeVsB4": public_value(origin_relative),
            "originWinShare": origin_win_share,
            "maximumConsecutiveOriginsRegressingOverFivePercent": maximum_consecutive,
            "internalWisImprovementVsB4": wis_improvement,
            "standardizedWidthRelativeDeltaVsB4": width_relative,
        },
    }


def _guard_audit(
    rows: Sequence[Mapping[str, Any]],
    guard_evidence: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    population = [row for row in rows if formal.is_model_population(row)]
    top10 = [
        row
        for row in population
        if bool((row.get("strata", {}) or {}).get("top_10_percent"))
    ]
    return {
        "top10CaseCount": len(top10),
        "top10WorkCount": len({formal.strict_case_key(row)[0] for row in top10}),
        "guardActiveOnEveryTop10Case": all(
            row.get("highValueGuardActive") is True for row in top10
        ),
        "guardFallbackCaseCount": sum(
            row.get("highValueGuardFallbackToB4") is True for row in top10
        ),
        "B4AlreadySelectedCaseCount": sum(
            row.get("candidate_id") == "B4"
            and row.get("highValueGuardFallbackToB4") is False
            for row in top10
        ),
        "nonB4OverrideCaseCount": sum(
            row.get("candidate_id") != "B4" for row in top10
        ),
        "outerTopBandActualUsed": False,
        "fallbackRuleChangedAfterResults": False,
        "guardEvidenceByOriginAndSegment": public_value(list(guard_evidence)),
        "stabilityEvidenceUsesEarlierOriginsOnly": True,
    }


def _residual_group(
    rows: Sequence[Mapping[str, Any]], spec: Mapping[str, Any]
) -> dict[str, Any]:
    works = {formal.strict_case_key(row)[0] for row in rows}
    minimum = int(spec["privacy"]["complementarySuppressionMinimumWorks"])
    if len(works) < minimum:
        return {
            "suppressed": True,
            "suppressionReason": "complementary_small_sample",
            "minimumWorks": minimum,
            "caseCount": None,
            "workCount": None,
            "predictedResidualCash": None,
            "actualCashWithoutKnownComponent": None,
        }
    predicted = 0.0
    actual_residual = 0.0
    for row in rows:
        known = {
            str(item["channel_key"])
            for item in row.get("channel_components", []) or []
            if item.get("component_type") != "other_or_new_channel_residual"
        }
        predicted += float(row.get("otherOrNewChannelResidualPoint", 0.0))
        actual_residual += sum(
            float(value)
            for component_key, value in (
                row.get("forecastableActualByComponent", {}) or {}
            ).items()
            if str(component_key) not in known
        )
    return {
        "caseCount": len(rows),
        "workCount": len(works),
        "predictedResidualCash": predicted,
        "actualCashWithoutKnownComponent": actual_residual,
        "signedAggregateDifference": predicted - actual_residual,
    }


def _segment_manifest(
    *,
    rows: Sequence[Mapping[str, Any]],
    selections: Sequence[Mapping[str, Any]],
    spec: Mapping[str, Any],
) -> dict[str, Any]:
    population = [row for row in rows if formal.is_model_population(row)]
    segments = {}
    for segment in c2.ACTIVITY_SEGMENTS:
        group = [row for row in population if row["activity_segment"] == segment]
        segments[segment] = {
            "definition": public_value(spec["activitySegmentation"][segment]),
            "metrics": _segment_metrics(rows, spec)[segment],
            "caseCount": len(group),
            "workCount": len({formal.strict_case_key(row)[0] for row in group}),
            "effectiveCandidateDistribution": {
                candidate: sum(row["candidate_id"] == candidate for row in group)
                for candidate in sorted({str(row["candidate_id"]) for row in group})
            },
        }
    return {
        "schema": "m2.c2_activity_segment_route_manifest.v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "segments": public_value(segments),
        "selectionByOriginAndSegment": public_value(list(selections)),
        "thresholdMovedAfterResults": False,
        "currentLifecycleRatingRiskRightsShelfUsed": False,
        "sourceWithoutHistoricalSnapshotUsed": False,
        "futurePerturbationInvariant": True,
        "samePredictAsOfEntryUsed": True,
        "seals": copy.deepcopy(spec["seals"]),
    }


def _residual_report(
    *,
    rows: Sequence[Mapping[str, Any]],
    audit: Mapping[str, Any],
    spec: Mapping[str, Any],
) -> dict[str, Any]:
    population = [row for row in rows if formal.is_model_population(row)]
    return {
        "schema": "m2.c2_other_new_channel_residual_audit.v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        **public_value(audit),
        "byActivitySegment": {
            segment: _residual_group(
                [row for row in population if row["activity_segment"] == segment],
                spec,
            )
            for segment in c2.ACTIVITY_SEGMENTS
        },
        "byHorizon": {
            str(horizon): _residual_group(
                [
                    row
                    for row in population
                    if formal.strict_case_key(row)[2] == horizon
                ],
                spec,
            )
            for horizon in formal.CORE_HORIZONS
        },
        "fitEvidence": "strictly_earlier_nested_origins_only",
        "futureChannelIdentityPredicted": False,
        "outerTruthUsed": False,
        "knownChannelCashDuplicated": False,
        "publicRealChannelIdentifiersPresent": False,
        "seals": copy.deepcopy(spec["seals"]),
    }


def _guard_report(
    audit: Mapping[str, Any], spec: Mapping[str, Any]
) -> dict[str, Any]:
    return {
        "schema": "m2.c2_high_value_guard_audit.v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "definition": public_value(spec["highValueGuard"]),
        **public_value(audit),
        "outerTopBandActualUsed": False,
        "fallbackRuleChangedAfterResults": False,
        "seals": copy.deepcopy(spec["seals"]),
    }


def _business_decision(spec: Mapping[str, Any]) -> dict[str, Any]:
    coverage = json.loads(phase_a.COVERAGE_JSON.read_text(encoding="utf-8"))
    overall = float(coverage["cashCoverage"]["forecastableCashShareOfLedgerCash"])
    top10 = float(coverage["topBands"]["top10"]["forecastableCashCoverage"])
    threshold = spec["businessCoverageDecision"]
    valid = (
        int(coverage["scope"]["standardWorkCount"])
        == int(spec["authority"]["standardWorkCount"])
        and int(coverage["scope"]["completeIncomeFactCount"])
        == int(spec["authority"]["completeIncomeFactCount"])
    )
    if not valid:
        decision = "FAIL"
    elif overall >= float(threshold["fullLibraryForecastableCashCoverageMinimum"]) and top10 >= float(
        threshold["top10ForecastableCashCoverageMinimum"]
    ):
        decision = "PASS"
    else:
        decision = "CONDITIONAL"
    return {
        "schema": "m2.c2_business_coverage_decision.v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "businessCoverageDecision": decision,
        "scope": {
            "standardWorkCount": coverage["scope"]["standardWorkCount"],
            "completeIncomeFactCount": coverage["scope"]["completeIncomeFactCount"],
            "nonOverlappingWorkLevelAggregation": coverage["scope"][
                "nonOverlappingWorkLevelAggregation"
            ],
        },
        "fullLibraryForecastableCashCoverage": overall,
        "top1ForecastableCashCoverage": coverage["topBands"]["top1"][
            "forecastableCashCoverage"
        ],
        "top5ForecastableCashCoverage": coverage["topBands"]["top5"][
            "forecastableCashCoverage"
        ],
        "top10ForecastableCashCoverage": top10,
        "fullLibraryThreshold": threshold[
            "fullLibraryForecastableCashCoverageMinimum"
        ],
        "top10Threshold": threshold["top10ForecastableCashCoverageMinimum"],
        "populationMovedToMeetCoverage": False,
        "surpriseBuyoutHidden": False,
        "mayAuthorizeRelease": False,
        "seals": copy.deepcopy(spec["seals"]),
    }


def _overall_decision(
    model_decision: str, business_decision: str, spec: Mapping[str, Any]
) -> str:
    mapping = spec["overallDecision"]
    if business_decision == "FAIL":
        return mapping["anyInvalidEvidence"]
    if model_decision == "PASS" and business_decision == "PASS":
        return mapping["modelPassBusinessPass"]
    if model_decision == "PASS":
        return mapping["modelPassBusinessConditional"]
    if business_decision == "PASS":
        return mapping["modelFailBusinessPass"]
    return mapping["modelFailBusinessConditional"]


def _compact_comparator_metrics(value: Mapping[str, Any]) -> dict[str, Any]:
    return public_value(
        {
            "modelPopulation": value["modelPopulation"],
            "highValue": value["highValue"],
            "horizons": value["horizons"],
            "topBands": value["topBands"],
            "routes": value["routes"],
            "origins": value["origins"],
            "internal80": value["internal80"],
        }
    )


def _build_reports(
    *,
    spec: Mapping[str, Any],
    metrics: Mapping[str, Any],
    b4_metrics: Mapping[str, Any],
    acceptance: Mapping[str, Any],
    bootstrap: Mapping[str, Any],
    selection_evidence: Sequence[Mapping[str, Any]],
    guard_evidence: Sequence[Mapping[str, Any]],
    rows: Sequence[Mapping[str, Any]],
    residual_audit: Mapping[str, Any],
    guard_audit: Mapping[str, Any],
    prediction_lock: Mapping[str, Any],
) -> tuple[dict[str, Any], ...]:
    business = _business_decision(spec)
    model_quality = {
        "schema": "m2.c2_model_quality_decision.v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "modelQualityDecision": acceptance["modelQualityDecision"],
        "acceptance": public_value(acceptance),
        "primaryComparator": "B4",
        "modelPopulationCaseCount": metrics["modelPopulation"]["caseCount"],
        "modelPopulationWorkCount": metrics["modelPopulation"]["uniqueWorkCount"],
        "populationMoved": False,
        "seals": copy.deepcopy(spec["seals"]),
    }
    overall = _overall_decision(
        model_quality["modelQualityDecision"],
        business["businessCoverageDecision"],
        spec,
    )
    comparator_bundle = json.loads(phase_a.BUNDLE_JSON.read_text(encoding="utf-8"))
    comparison = {
        model: _compact_comparator_metrics(comparator_bundle["metrics"][model])
        for model in ("B0b", "B1", "B3", "B4")
    }
    comparison["C2"] = _compact_comparator_metrics(metrics)
    validation = {
        "schema": "m2.c2_development_validation.v1",
        "version": "v1",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "modelQualityDecision": model_quality["modelQualityDecision"],
        "businessCoverageDecision": business["businessCoverageDecision"],
        "overallDecision": overall,
        "technicalSummary": {
            "C2Executed": True,
            "primaryComparator": "B4",
            "frozenCaseCount": len(rows),
            "statisticallyScoreableCaseCount": metrics["caseState"][
                "statisticallyScoreableCaseCount"
            ],
            "modelPopulationCaseCount": metrics["modelPopulation"]["caseCount"],
            "modelPopulationWorkCount": metrics["modelPopulation"]["uniqueWorkCount"],
            "modelPopulationUnchanged": metrics["modelPopulation"]["caseCount"]
            == spec["authority"]["formalModelPopulationCaseCount"],
            "pureBuyoutNullScoredAsZero": False,
            "thresholdMoved": False,
            "outerResultUsedForSelection": False,
            "samePredictAsOfEntryUsed": True,
        },
        "predictionIntegrity": public_value(prediction_lock),
        "metrics": public_value(
            {
                **metrics,
                "activitySegments": _segment_metrics(rows, spec),
            }
        ),
        "B4Metrics": _compact_comparator_metrics(b4_metrics),
        "acceptance": public_value(acceptance),
        "pairedBlockBootstrapVsB4": public_value(bootstrap),
        "comparatorBundle": {
            "reported": ["B0b", "B1", "B3", "B4", "C2"],
            "primaryComparator": "B4",
            "sameCaseKeys": True,
            "sameActuals": True,
            "sameModelPopulation": True,
            "sameOriginsHorizonsAndSeed": True,
            "metrics": comparison,
        },
        "selectionByOriginAndSegment": public_value(list(selection_evidence)),
        "highValueGuardByOriginAndSegment": public_value(list(guard_evidence)),
        "P0Count": 0,
        "P1Count": 0,
        "P2Boundary": "fact_audit_only",
        "automaticOperationalActionCount": 0,
        "C3Started": False,
        "M3Started": False,
        "seals": copy.deepcopy(spec["seals"]),
        "privacy": {
            "aggregateOnly": True,
            "identifiersPresent": False,
            "realChannelNamesPresent": False,
            "predictionIntervalEndpointsPresent": False,
        },
        "nextBoundary": "stop_before_C3_wait_for_user_authorization",
    }
    segment_report = _segment_manifest(
        rows=rows, selections=selection_evidence, spec=spec
    )
    residual_report = _residual_report(
        rows=rows, audit=residual_audit, spec=spec
    )
    guard_report = _guard_report(guard_audit, spec)
    return (
        validation,
        segment_report,
        residual_report,
        guard_report,
        model_quality,
        business,
    )


def _write_development_markdown(
    validation: Mapping[str, Any],
    segment: Mapping[str, Any],
    residual: Mapping[str, Any],
    guard: Mapping[str, Any],
    model_quality: Mapping[str, Any],
    business: Mapping[str, Any],
) -> None:
    metrics = validation["metrics"]["modelPopulation"]
    write_text(
        VALIDATION_MD,
        f"""# M2 C2 development 验证 v1

C2 已在冻结的 18615 个 development case、12223 个 statistically scoreable case 和 7851 个 formal-cash 模型人口 case 上执行。总体 WAPE 为 {metrics['wape']:.8f}，signed aggregate bias 为 {metrics['signedAggregateBias']:+.8f}。

模型质量判定为 {validation['modelQualityDecision']}；业务覆盖判定为 {validation['businessCoverageDecision']}；总判定为 {validation['overallDecision']}。所有选择只使用 strictly-earlier origin，B4 始终为锚，高价值证据不足时强制回退 B4。

pure-buyout 无 cutoff commitment 时继续 null abstain；mixed 只预测实销和 cutoff 已确认应收。通用其他或新增渠道残差不预测真实渠道身份，也不重复已知渠道现金。

结果继续为 not_for_formal_decision。final holdout、embargo shadow 和 deferred 60-month labels 均未打开；未进入 C3，未 release，未进入 M3。
""",
    )
    segment_counts = {
        key: value["caseCount"] for key, value in segment["segments"].items()
    }
    write_text(
        SEGMENT_MD,
        f"""# M2 C2 活跃度分层与路由清单 v1

dense、intermittent、dormant 均由 cutoff 及以前的实销完整月历史判定，case 分布为 {json.dumps(segment_counts, ensure_ascii=False)}。每个 case 保存 segmentReason；短历史或从未出现正实销证据时直接使用 B4。

分层阈值未按 outer 结果移动，当前生命周期、rating、risk、版权和货架状态均未作为历史预测特征。结果仅供 development 校准，继续 not_for_formal_decision。
""",
    )
    write_text(
        RESIDUAL_MD,
        f"""# M2 C2 其他或新增渠道残差审计 v1

通用 residual 预测现金合计为 {residual['otherOrNewChannelResidualPredictedCash']:.2f}；已知渠道 matched actual coverage 为 {residual['matchedForecastableCashShare']:.4%}。最大 work point 对账差为 {residual['maximumWorkPointReconciliationDifference']:.10f}。

residual 参数只来自 strictly-earlier origins，不读取当前 outer truth，不记忆作品未来渠道，不伪造真实渠道名称，也不与已知渠道现金重复。公开报告不含渠道标识或区间端点。
""",
    )
    write_text(
        GUARD_MD,
        f"""# M2 C2 高价值 B4 保护审计 v1

Top10 共有 {guard['top10CaseCount']} 个模型人口 case；所有 Top10 case 均经过冻结的 guard。guard 回退 B4 的 case 数为 {guard['guardFallbackCaseCount']}，通过 earlier-origin 证据保留非 B4 覆盖的 case 数为 {guard['nonB4OverrideCaseCount']}。

Top1、Top5、Top10 均按 cutoff 前 trailing-12 收入排名。当前 outer top-band actual 未用于选择，guard 规则未按结果修改。
""",
    )
    write_text(
        MODEL_DECISION_MD,
        f"""# M2 C2 模型质量判定 v1

固定 25 项模型质量条件通过 {model_quality['acceptance']['passedConditionCount']} 项，判定为 {model_quality['modelQualityDecision']}。人口、gate 和阈值均未移动；该判定不构成正式业务批准或 release。
""",
    )
    write_text(
        BUSINESS_DECISION_MD,
        f"""# M2 C2 业务覆盖判定 v1

全库 forecastable cash coverage 为 {business['fullLibraryForecastableCashCoverage']:.4%}，Top10 coverage 为 {business['top10ForecastableCashCoverage']:.4%}，观察线均为 90%。业务覆盖判定为 {business['businessCoverageDecision']}。

未通过移动人口、伪造预测或隐藏 surprise buyout 改善覆盖。该观察判定继续为 not_for_formal_decision，不能授权 release。
""",
    )


def _write_public_development_reports(reports: Sequence[Mapping[str, Any]]) -> None:
    validation, segment, residual, guard, model_quality, business = reports
    pairs = (
        (VALIDATION_JSON, validation),
        (SEGMENT_JSON, segment),
        (RESIDUAL_JSON, residual),
        (GUARD_JSON, guard),
        (MODEL_DECISION_JSON, model_quality),
        (BUSINESS_DECISION_JSON, business),
    )
    for path, value in pairs:
        write_json(path, value)
    _write_development_markdown(
        validation, segment, residual, guard, model_quality, business
    )
    assert_public_safety(
        (
            VALIDATION_JSON,
            VALIDATION_MD,
            SEGMENT_JSON,
            SEGMENT_MD,
            RESIDUAL_JSON,
            RESIDUAL_MD,
            GUARD_JSON,
            GUARD_MD,
            MODEL_DECISION_JSON,
            MODEL_DECISION_MD,
            BUSINESS_DECISION_JSON,
            BUSINESS_DECISION_MD,
        )
    )


def _private_case_payload(row: Mapping[str, Any]) -> dict[str, Any]:
    key = formal.strict_case_key(row)
    return {
        "caseKey": {
            "standard_work_id": key[0],
            "origin": key[1],
            "horizon_months": key[2],
            "route": key[3],
        },
        "statisticallyScoreable": row["statisticallyScoreable"],
        "businessServingEligible": row["businessServingEligible"],
        "modelPredictionAvailable": row["modelPredictionAvailable"],
        "routeAbstained": row["routeAbstained"],
        "abstentionReason": row["abstentionReason"],
        "rawModelPrediction": row["rawModelPrediction"],
        "servedPrediction": row["servedPrediction"],
        "forecastableCashActual": row["forecastableCashActual"],
        "uncommittedBuyoutSurpriseActual": row[
            "uncommittedBuyoutSurpriseActual"
        ],
        "totalLedgerCashActual": row["totalLedgerCashActual"],
        "activitySegment": row["activity_segment"],
        "segmentReason": row["segment_reason"],
        "requestedCandidate": row["requested_candidate_id"],
        "effectiveCandidate": row["candidate_id"],
        "highValueGuardFallbackToB4": row["highValueGuardFallbackToB4"],
        "otherOrNewChannelResidualPoint": row[
            "otherOrNewChannelResidualPoint"
        ],
        "channelComponents": copy.deepcopy(row.get("channel_components", [])),
        "forecastableActualByComponent": copy.deepcopy(
            row.get("forecastableActualByComponent", {})
        ),
        "strata": copy.deepcopy(row.get("strata", {})),
        "internalInterval": copy.deepcopy(row.get("_internal_interval", {})),
    }


def _write_private_cases(rows: Sequence[Mapping[str, Any]]) -> tuple[int, str]:
    hasher = hashlib.sha256()
    count = 0
    with PRIVATE_CASES.open("wb") as handle:
        for row in rows:
            payload = _private_case_payload(row)
            raw = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            handle.write(raw)
            hasher.update(raw)
            count += 1
    return count, hasher.hexdigest()


def _write_private_workbook(
    *,
    rows: Sequence[Mapping[str, Any]],
    validation: Mapping[str, Any],
    selections: Sequence[Mapping[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
        from openpyxl.styles import Font  # type: ignore[import-not-found]
    except ImportError as exc:
        raise C2RunnerError("openpyxl is required for the private C2 workbook") from exc
    workbook = Workbook()
    summary = workbook.active
    summary.title = "结论摘要"
    summary.append(["项目", "结果"])
    summary.append(["模型质量判定", validation["modelQualityDecision"]])
    summary.append(["业务覆盖判定", validation["businessCoverageDecision"]])
    summary.append(["总判定", validation["overallDecision"]])
    summary.append(["总体 WAPE", validation["metrics"]["modelPopulation"]["wape"]])
    summary.append(
        [
            "总体 signed bias",
            validation["metrics"]["modelPopulation"]["signedAggregateBias"],
        ]
    )
    summary.append(["final holdout 已打开", False])
    summary.append(["是否可正式决策", False])
    selection_sheet = workbook.create_sheet("分层选择")
    selection_sheet.append(
        [
            "outer origin",
            "活跃度分层",
            "选择候选",
            "选择原因",
            "earlier origin 数",
            "earlier case 数",
        ]
    )
    for item in selections:
        selection_sheet.append(
            [
                item["outerOrigin"],
                item["segment"],
                item["selectedCandidate"],
                item["selectionReason"],
                item["earlierOriginCount"],
                item["earlierCaseCount"],
            ]
        )
    cases = workbook.create_sheet("业务抽检样本")
    cases.append(
        [
            "标准作品ID",
            "origin",
            "horizon",
            "收入路由",
            "活跃度分层",
            "候选",
            "C2预测",
            "实际现金",
            "通用残差",
            "高价值回退B4",
        ]
    )
    sample = sorted(
        [row for row in rows if formal.is_model_population(row)],
        key=lambda row: abs(
            float(row["rawModelPrediction"]) - float(row["forecastableCashActual"])
        ),
        reverse=True,
    )[:500]
    for row in sample:
        key = formal.strict_case_key(row)
        cases.append(
            [
                key[0],
                key[1],
                key[2],
                key[3],
                row["activity_segment"],
                row["candidate_id"],
                row["rawModelPrediction"],
                row["forecastableCashActual"],
                row["otherOrNewChannelResidualPoint"],
                row["highValueGuardFallbackToB4"],
            ]
        )
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(PRIVATE_WORKBOOK)


def run_development() -> dict[str, Any]:
    require_named_branch()
    _assert_private_paths()
    progress("verifying Gate C 14/14 before any development input read")
    authorization = verify_c2_authorization()
    spec = c2.load_spec()
    gate = json.loads(GATE_C_JSON.read_text(encoding="utf-8"))
    if gate["passedConditionCount"] != 14 or gate["C2AuthorizedByGateC"] is not True:
        raise C2RunnerError("Gate C is not 14/14; C2 replay remains blocked")
    progress("loading locked formal comparator cases")
    b4, comparator, formal_manifest = c2r1_runner._load_formal_private_cases()
    calibration_spec, _v11, _v12 = v12.load_and_validate_contract()
    progress("loading authorized work histories read-only")
    works_list, _posthoc, input_evidence = legacy.load_authorized_works(calibration_spec)
    if input_evidence["inputFingerprint"] != formal_manifest["inputFingerprint"]:
        raise C2RunnerError("C2 authority fingerprint differs from formal comparator")
    works = {str(work["standard_work_id"]): work for work in works_list}
    forward_templates = {
        key: template
        for (role, key), template in b4.items()
        if role.startswith("development_forward_score:")
    }
    if len(forward_templates) != int(spec["authority"]["developmentCaseCount"]):
        raise C2RunnerError("C2 frozen development case count differs")
    points, segment_states, residual_models, reactivation_models = (
        _materialize_candidate_points(
            spec=spec,
            calibration_spec=calibration_spec,
            works=works,
            forward_templates=forward_templates,
            comparator=comparator,
        )
    )
    selections: dict[tuple[str, str], str] = {}
    selection_evidence: list[dict[str, Any]] = []
    guard_allowed: dict[tuple[str, str], bool] = {}
    guard_evidence: list[dict[str, Any]] = []
    progress("selecting each outer-origin activity route from strictly earlier origins")
    for origin in spec["authority"]["origins"]:
        for segment in c2.ACTIVITY_SEGMENTS:
            selected, evidence = _select_candidate(
                outer_origin=origin,
                segment=segment,
                spec=spec,
                points=points,
                templates=forward_templates,
                segment_states=segment_states,
            )
            selections[(origin, segment)] = selected
            selection_evidence.append(evidence)
            allowed, guard = _high_value_guard(
                outer_origin=origin,
                segment=segment,
                selected_candidate=selected,
                spec=spec,
                points=points,
                templates=forward_templates,
                segment_states=segment_states,
            )
            guard_allowed[(origin, segment)] = allowed
            guard_evidence.append(guard)
    predictions: list[dict[str, Any]] = []
    progress("materializing selected C2 projections through the single predict_as_of entry")
    for index, (key, template) in enumerate(sorted(forward_templates.items()), start=1):
        is_model = (
            template["statisticallyScoreable"] is True
            and template["modelPredictionAvailable"] is True
            and template["routeAbstained"] is False
        )
        if is_model:
            segment = str(segment_states[key]["segment"])
            candidate_id = selections[(key[1], segment)]
            allow_high = guard_allowed[(key[1], segment)]
            role = f"development_forward_score:{key[1]}"
            comparator_rows = c2r1_runner._comparator_rows_for_key(
                comparator, role, key
            )
        else:
            candidate_id = "B4"
            allow_high = False
            comparator_rows = None
        prediction = c2.predict_as_of(
            work=works[key[0]],
            origin=key[1],
            horizon=key[2],
            case_state=_prediction_case_state(template),
            comparator_rows=comparator_rows,
            calibration_spec=calibration_spec,
            spec=spec,
            candidate_id=candidate_id,
            residual_model=residual_models[key[1]],
            reactivation_model=reactivation_models[key[1]],
            cutoff_top10=bool(
                (template.get("strata", {}) or {}).get("top_10_percent")
            ),
            high_value_override_allowed=allow_high,
            cash_commitment_snapshots=[],
        )
        predictions.append(prediction)
        if index % 4000 == 0:
            progress(f"selected C2 projections: {index}/{len(forward_templates)}")
    projection_before = digest(
        [_compact_prediction_projection(row) for row in predictions]
    )
    forward = [
        _join_truth(row, forward_templates[formal.strict_case_key(row)])
        for row in predictions
    ]
    projection_after = digest(
        [_compact_prediction_projection(row) for row in forward]
    )
    if projection_before != projection_after:
        raise C2RunnerError("C2 prediction changed after the truth join")
    warmup = []
    for (role, _key), template in sorted(b4.items()):
        if role != "development_warmup_interval_calibration":
            continue
        row = c2r1_runner._build_warmup_row(template)
        row["model_id"] = "C2"
        row["candidate_id"] = "B4_warmup"
        warmup.append(row)
    progress("calibrating internal 80% intervals from earlier C2 residuals only")
    formal.apply_internal_intervals(forward, [*warmup, *forward], formal.load_spec())
    metrics = phase_a.metrics_for_model(forward)
    b4_forward = _locked_b4_rows(forward_templates)
    b4_metrics = phase_a.metrics_for_model(b4_forward)
    model_keys = {
        formal.strict_case_key(row) for row in forward if formal.is_model_population(row)
    }
    b4_keys = {
        formal.strict_case_key(row)
        for row in b4_forward
        if formal.is_model_population(row)
    }
    if model_keys != b4_keys or len(model_keys) != int(
        spec["authority"]["formalModelPopulationCaseCount"]
    ):
        raise C2RunnerError("C2 and B4 model-population keys differ")
    for candidate_row, b4_row in zip(forward, b4_forward):
        if (
            formal.strict_case_key(candidate_row) != formal.strict_case_key(b4_row)
            or float(candidate_row["forecastableCashActual"])
            != float(b4_row["forecastableCashActual"])
        ):
            raise C2RunnerError("C2 and B4 case key or actual differs")
    residual_audit = _channel_residual_audit(forward)
    guard_audit = _guard_audit(forward, guard_evidence)
    acceptance = _acceptance(
        metrics=metrics,
        comparator=b4_metrics,
        residual_audit=residual_audit,
        guard_audit=guard_audit,
        spec=spec,
    )
    progress("running paired work-by-origin block bootstrap")
    bootstrap = formal.paired_relative_block_bootstrap(
        [*b4_forward, *forward], "B4", ("B4", "C2"), formal.load_spec()
    )
    bootstrap.pop("clusterKeys", None)
    bootstrap["clusterDefinition"] = "deidentified_work_x_origin"
    prediction_lock = {
        "predictionLockedBeforeTruthJoin": True,
        "predictionProjectionDigest": projection_before,
        "postTruthProjectionMatchesLock": True,
        "actualFieldAbsentAtPredictionLock": True,
        "fullCaseKeyCount": len(forward),
        "modelPopulationKeyCount": len(model_keys),
        "modelPopulationMatchesPrimaryComparator": True,
        "futurePerturbationInvariant": True,
        "finalHoldoutOpened": False,
    }
    reports = _build_reports(
        spec=spec,
        metrics=metrics,
        b4_metrics=b4_metrics,
        acceptance=acceptance,
        bootstrap=bootstrap,
        selection_evidence=selection_evidence,
        guard_evidence=guard_evidence,
        rows=forward,
        residual_audit=residual_audit,
        guard_audit=guard_audit,
        prediction_lock=prediction_lock,
    )
    progress("writing deidentified reports and ignored private evidence")
    _write_public_development_reports(reports)
    case_count, case_sha = _write_private_cases(forward)
    _write_private_workbook(
        rows=forward,
        validation=reports[0],
        selections=selection_evidence,
    )
    public_paths = (
        VALIDATION_JSON,
        VALIDATION_MD,
        SEGMENT_JSON,
        SEGMENT_MD,
        RESIDUAL_JSON,
        RESIDUAL_MD,
        GUARD_JSON,
        GUARD_MD,
        MODEL_DECISION_JSON,
        MODEL_DECISION_MD,
        BUSINESS_DECISION_JSON,
        BUSINESS_DECISION_MD,
    )
    manifest = {
        "schema": "m2.c2_development_manifest.private.v1",
        "decisionStatus": "not_for_formal_decision",
        "tracked": False,
        "specDigest": c2.canonical_digest(spec),
        "gateCDigest": digest(gate),
        "phaseACheckpoint": authorization["phaseACheckpoint"],
        "inputFingerprint": input_evidence["inputFingerprint"],
        "privateCaseCount": case_count,
        "modelPopulationCaseCount": len(model_keys),
        "privateCaseSha256": case_sha,
        "privateWorkbookSha256": file_sha256(PRIVATE_WORKBOOK),
        "publicArtifactSha256": {
            path.name: file_sha256(path) for path in public_paths
        },
        "modelQualityDecision": reports[0]["modelQualityDecision"],
        "businessCoverageDecision": reports[0]["businessCoverageDecision"],
        "overallDecision": reports[0]["overallDecision"],
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
        "C3Started": False,
        "M3Started": False,
    }
    write_json(PRIVATE_DEVELOPMENT_MANIFEST, manifest)
    _assert_private_paths()
    return {
        "status": "passed",
        "modelQualityDecision": reports[0]["modelQualityDecision"],
        "businessCoverageDecision": reports[0]["businessCoverageDecision"],
        "overallDecision": reports[0]["overallDecision"],
        "acceptancePassedConditionCount": acceptance["passedConditionCount"],
        "acceptanceConditionCount": acceptance["conditionCount"],
        "modelPopulationCaseCount": metrics["modelPopulation"]["caseCount"],
        "overallWape": rounded(metrics["modelPopulation"]["wape"]),
        "overallSignedAggregateBias": rounded(
            metrics["modelPopulation"]["signedAggregateBias"]
        ),
        "privateWorkbookTracked": False,
        "decisionStatus": "not_for_formal_decision",
        "finalHoldoutOpened": False,
        "C3Started": False,
        "M3Started": False,
    }


def verify_development() -> dict[str, Any]:
    _assert_private_paths()
    spec = c2.load_spec()
    public_paths = (
        VALIDATION_JSON,
        VALIDATION_MD,
        SEGMENT_JSON,
        SEGMENT_MD,
        RESIDUAL_JSON,
        RESIDUAL_MD,
        GUARD_JSON,
        GUARD_MD,
        MODEL_DECISION_JSON,
        MODEL_DECISION_MD,
        BUSINESS_DECISION_JSON,
        BUSINESS_DECISION_MD,
    )
    if any(
        not path.is_file()
        for path in (
            *public_paths,
            PRIVATE_CASES,
            PRIVATE_DEVELOPMENT_MANIFEST,
            PRIVATE_WORKBOOK,
        )
    ):
        raise C2RunnerError("C2 development evidence is incomplete")
    assert_public_safety(public_paths)
    manifest = json.loads(
        PRIVATE_DEVELOPMENT_MANIFEST.read_text(encoding="utf-8")
    )
    if (
        manifest.get("schema") != "m2.c2_development_manifest.private.v1"
        or manifest.get("specDigest") != c2.canonical_digest(spec)
        or file_sha256(PRIVATE_CASES) != manifest.get("privateCaseSha256")
        or file_sha256(PRIVATE_WORKBOOK) != manifest.get("privateWorkbookSha256")
        or any(
            file_sha256(path)
            != manifest.get("publicArtifactSha256", {}).get(path.name)
            for path in public_paths
        )
    ):
        raise C2RunnerError("C2 development manifest differs")
    rows = []
    count = 0
    with PRIVATE_CASES.open("rb") as handle:
        for raw in handle:
            payload = json.loads(raw)
            canonical = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            if raw != canonical:
                raise C2RunnerError("C2 private case is not canonical")
            count += 1
            key = payload["caseKey"]
            rows.append(
                {
                    "case_key": copy.deepcopy(key),
                    "statisticallyScoreable": payload["statisticallyScoreable"],
                    "modelPredictionAvailable": payload["modelPredictionAvailable"],
                    "routeAbstained": payload["routeAbstained"],
                    "rawModelPrediction": payload["rawModelPrediction"],
                    "forecastableCashActual": payload["forecastableCashActual"],
                }
            )
    if count != int(manifest["privateCaseCount"]) or count != int(
        spec["authority"]["developmentCaseCount"]
    ):
        raise C2RunnerError("C2 private case count differs")
    population = [row for row in rows if formal.is_model_population(row)]
    recomputed = formal.metric_rows(population, "rawModelPrediction")
    validation = json.loads(VALIDATION_JSON.read_text(encoding="utf-8"))
    if (
        int(recomputed["caseCount"])
        != int(spec["authority"]["formalModelPopulationCaseCount"])
        or abs(
            float(recomputed["wape"])
            - float(validation["metrics"]["modelPopulation"]["wape"])
        )
        > 1e-8
        or abs(
            float(recomputed["signedAggregateBias"])
            - float(
                validation["metrics"]["modelPopulation"]["signedAggregateBias"]
            )
        )
        > 1e-8
        or any(value is not False for value in validation["seals"].values())
        or validation["decisionStatus"] != "not_for_formal_decision"
    ):
        raise C2RunnerError("C2 development metrics or boundary does not recompute")
    return {
        "status": "passed",
        "modelQualityDecision": validation["modelQualityDecision"],
        "businessCoverageDecision": validation["businessCoverageDecision"],
        "overallDecision": validation["overallDecision"],
        "privateCaseCount": count,
        "modelPopulationCaseCount": len(population),
        "overallWape": rounded(recomputed["wape"]),
        "overallSignedAggregateBias": rounded(
            recomputed["signedAggregateBias"]
        ),
        "privateWorkbookTracked": False,
        "decisionStatus": "not_for_formal_decision",
        "finalHoldoutOpened": False,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    modes = parser.add_mutually_exclusive_group()
    modes.add_argument("--preflight", action="store_true")
    modes.add_argument("--run-phase-a", action="store_true")
    modes.add_argument("--verify-phase-a", action="store_true")
    modes.add_argument("--finalize-gate-c-validation", action="store_true")
    modes.add_argument("--verify-gate-c-after-push", action="store_true")
    modes.add_argument("--verify-c2-authorization", action="store_true")
    modes.add_argument("--run-development", action="store_true")
    modes.add_argument("--verify-development", action="store_true")
    modes.add_argument("--run-final-holdout", action="store_true")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        if args.run_final_holdout:
            raise C2RunnerError(
                "final holdout is sealed in the C2 runner; dataLoadCalls=0"
            )
        if args.run_phase_a:
            result = run_phase_a()
        elif args.verify_phase_a:
            result = verify_phase_a()
        elif args.finalize_gate_c_validation:
            result = finalize_gate_c_validation()
        elif args.verify_gate_c_after_push:
            result = verify_gate_c_after_push()
        elif args.verify_c2_authorization:
            result = verify_c2_authorization()
        elif args.run_development:
            result = run_development()
        elif args.verify_development:
            result = verify_development()
        else:
            result = preflight()
        print(json.dumps(result, ensure_ascii=False, sort_keys=True, allow_nan=False))
        return 0
    except (
        C2RunnerError,
        c2.C2Error,
        formal.FormalComparatorError,
        phase_a.FormalReplayError,
        legacy.ReplayError,
        v12.CalibrationV12Error,
        RuntimeError,
        ValueError,
        AssertionError,
        KeyError,
    ) as exc:
        print(
            json.dumps({"status": "failed", "reason": str(exc)}, ensure_ascii=False),
            file=sys.stderr,
        )
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
