from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(ROOT / "tools" / "m2-calibration"))

import numpy as np
import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - runner reports this explicitly if Excel write is attempted.
    Alignment = Font = PatternFill = get_column_letter = None

from calibrate_cleaned_bills import add_months, classify_at, month_range
from run_nonformal_dry_run import RATING_ORDER, evaluate_work_summary, load_analysis_inputs

OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-algorithm-validation"

ROOT_CAUSE_JSON = OUTPUT_DIR / "M2-forecast-model-failure-root-cause-audit-v0.1.json"
ROOT_CAUSE_MD = OUTPUT_DIR / "M2-forecast-model-failure-root-cause-audit-v0.1.md"
BAKEOFF_JSON = OUTPUT_DIR / "M2-forecast-model-bakeoff-v1.json"
BAKEOFF_MD = OUTPUT_DIR / "M2-forecast-model-bakeoff-v1.md"
SELECTION_JSON = OUTPUT_DIR / "M2-forecast-model-selection-v1.json"
SELECTION_MD = OUTPUT_DIR / "M2-forecast-model-selection-v1.md"
READINESS_JSON = OUTPUT_DIR / "M2-algorithm-business-readiness-v1.json"
READINESS_MD = OUTPUT_DIR / "M2-algorithm-business-readiness-v1.md"
PRIVATE_XLSX = PRIVATE_DIR / "m2-forecast-model-bakeoff-v1.xlsx"
PRIVATE_DETAIL_JSON = PRIVATE_DIR / "m2-forecast-model-bakeoff-v1-private-detail.json"

FAILED_CANDIDATE = "m2-realdata-dev-candidate-b-forecast-rebuilt-v0.3"
NEW_CANDIDATE = "m2-realdata-dev-forecast-model-v1.0"
CONDITIONAL_CANDIDATE = "m2-realdata-dev-forecast-model-v1.0-conditional"
PARAMETER_VARIANT = "candidate-b"
HORIZONS = [3, 6, 12, 18, 24]
MODEL_ORDER = [
    "model_a_trailing_baseline",
    "model_b_lifecycle_robust",
    "model_c_zero_inflated_sparse",
    "model_d_hierarchical_shrinkage",
    "model_e_selector",
]
MODEL_LABELS = {
    "candidate_b_v03": "candidate-b-v0.3 failed route",
    "raw_trailing_baseline": "raw trailing baseline",
    "model_a_trailing_baseline": "Model A seasonal/trailing baseline",
    "model_b_lifecycle_robust": "Model B lifecycle segmented robust forecast",
    "model_c_zero_inflated_sparse": "Model C zero-inflated sparse revenue model",
    "model_d_hierarchical_shrinkage": "Model D hierarchical shrinkage model",
    "model_e_selector": "Model E selector model",
}
LIFECYCLES = ["growth", "stable", "rebound", "declining", "long_tail", "inactive", "insufficient_history"]
REVENUE_BUCKETS = ["top", "high", "mid", "low", "long_tail"]
CONFIDENCES = ["high", "medium", "low", "blocked_for_business_use"]


def safe_float(value, default: float = 0.0) -> float:
    try:
        result = float(value)
    except Exception:
        return default
    return result if math.isfinite(result) else default


def safe_int(value, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def rounded(value, digits: int = 4):
    if value is None:
        return None
    number = safe_float(value)
    return round(number, digits)


def json_safe(value):
    if isinstance(value, dict):
        return {str(key): json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [json_safe(child) for child in value]
    if isinstance(value, tuple):
        return [json_safe(child) for child in value]
    if isinstance(value, set):
        return sorted(json_safe(child) for child in value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    lines = ["| " + " | ".join(label for _, label in columns) + " |"]
    lines.append("|" + "|".join("---" for _ in columns) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |")
    return "\n".join(lines)


def distribution(values, keys: list[str] | None = None) -> dict:
    counter = Counter(values)
    ordered = keys or sorted(counter)
    return {key: int(counter.get(key, 0)) for key in ordered if keys or counter.get(key, 0)}


def quantiles(values) -> dict:
    clean = [safe_float(value) for value in values if value is not None and math.isfinite(safe_float(value))]
    if not clean:
        return {"count": 0, "p50": None, "p75": None, "p90": None, "p95": None}
    return {
        "count": len(clean),
        "p50": rounded(np.quantile(clean, 0.50)),
        "p75": rounded(np.quantile(clean, 0.75)),
        "p90": rounded(np.quantile(clean, 0.90)),
        "p95": rounded(np.quantile(clean, 0.95)),
    }


def build_month_matrix(context: dict) -> tuple[pd.DataFrame, list[str]]:
    bill = context["bill"]
    complete = bill[bill["validForCalibration"] & (bill["billMonth"] <= context["latest_complete_month"])].copy()
    months = month_range(complete["billMonth"].min(), context["latest_complete_month"])
    matrix = complete.groupby(["standardWorkId", "billMonth"])["amount"].sum().unstack(fill_value=0.0)
    matrix = matrix.reindex(columns=months, fill_value=0.0)
    return matrix, months


def revenue_bucket(amount: float, q: dict[str, float]) -> str:
    value = safe_float(amount)
    if value >= q["p95"]:
        return "top"
    if value >= q["p75"]:
        return "high"
    if value >= q["p40"]:
        return "mid"
    if value > 0:
        return "low"
    return "long_tail"


def history_stats(history: np.ndarray, thresholds: dict) -> dict:
    values = np.asarray(history, dtype=float)
    positive = values[values > 0]
    last3 = float(values[-3:].sum()) if len(values) else 0.0
    last6 = float(values[-6:].sum()) if len(values) else 0.0
    last12 = float(values[-12:].sum()) if len(values) else 0.0
    last24 = float(values[-24:].sum()) if len(values) else 0.0
    active_months = int(np.count_nonzero(values > 0))
    zero_months = int(np.count_nonzero(values == 0))
    total = float(values.sum())
    recent = values[-6:] if len(values) >= 6 else values
    mean = float(np.mean(recent)) if len(recent) else 0.0
    volatility = float(np.std(recent) / mean) if mean > 0 else 0.0
    peak_share = float(values.max() / total) if total > 0 and len(values) else 0.0
    lifecycle = classify_at(values, thresholds) if len(values) else "insufficient_history"
    if len(values) < 6:
        lifecycle = "insufficient_history"
    return {
        "lifecycle": lifecycle,
        "last3": last3,
        "last6": last6,
        "last12": last12,
        "last24": last24,
        "activeMonths": active_months,
        "zeroMonths": zero_months,
        "total": total,
        "positiveMedian": float(np.median(positive)) if len(positive) else 0.0,
        "monthlyMean12": last12 / 12.0,
        "monthlyMean24": last24 / 24.0 if last24 > 0 else last12 / 12.0,
        "volatility": volatility,
        "peakShare": peak_share,
        "sparse": active_months <= 3 or total <= 10,
        "recentZero": last6 <= 0.01,
    }


def scale_from_stats(stats: dict, q: dict[str, float]) -> str:
    if stats["lifecycle"] == "long_tail":
        return "long_tail"
    return revenue_bucket(stats["total"], q)


def winsorized_mean(values: np.ndarray) -> float:
    clean = np.asarray(values, dtype=float)
    clean = clean[~np.isnan(clean)]
    if len(clean) == 0:
        return 0.0
    if len(clean) < 4:
        return float(np.mean(clean))
    lower = np.quantile(clean, 0.10)
    upper = np.quantile(clean, 0.85)
    return float(np.mean(np.clip(clean, lower, upper)))


def trimmed_mean(values: np.ndarray) -> float:
    clean = np.asarray(values, dtype=float)
    clean = clean[~np.isnan(clean)]
    if len(clean) == 0:
        return 0.0
    if len(clean) < 6:
        return float(np.mean(clean))
    ordered = np.sort(clean)
    trim = max(1, int(len(ordered) * 0.15))
    trimmed = ordered[trim:-trim] if len(ordered) > trim * 2 else ordered
    return float(np.mean(trimmed))


def raw_trailing_prediction(stats: dict, horizon: int) -> tuple[float, str]:
    return max(0.0, stats["last12"] / 12.0 * horizon), "raw last-12 trailing average"


def model_a_prediction(stats: dict, horizon: int) -> tuple[float, str]:
    last3 = stats["last3"] / 3.0 if stats["last3"] > 0 else 0.0
    last6 = stats["last6"] / 6.0 if stats["last6"] > 0 else 0.0
    last12 = stats["last12"] / 12.0 if stats["last12"] > 0 else 0.0
    last24 = stats["last24"] / 24.0 if stats["last24"] > 0 else last12
    monthly = 0.15 * last3 + 0.35 * last6 + 0.35 * last12 + 0.15 * last24
    factor = {
        "growth": 1.05,
        "stable": 0.95,
        "rebound": 0.98,
        "declining": 0.65,
        "long_tail": 0.35,
        "inactive": 0.08,
        "insufficient_history": 0.55,
    }.get(stats["lifecycle"], 0.9)
    if stats["recentZero"]:
        factor *= 0.20
    return max(0.0, monthly * horizon * factor), "weighted 3/6/12/24 trailing average with lifecycle decay"


def model_b_prediction(stats: dict, horizon: int) -> tuple[float, str]:
    window = np.asarray(stats.get("history")[-24:] if len(stats.get("history", [])) >= 24 else stats.get("history", []), dtype=float)
    monthly = max(winsorized_mean(window[-12:]), trimmed_mean(window), stats["positiveMedian"] * 0.40)
    factor = {
        "growth": 1.12,
        "stable": 0.88,
        "rebound": 1.02,
        "declining": 0.48,
        "long_tail": 0.22,
        "inactive": 0.04,
        "insufficient_history": 0.45,
    }.get(stats["lifecycle"], 0.85)
    prediction = monthly * horizon * factor
    if stats["peakShare"] >= 0.90:
        prediction *= 0.40
    if stats["last12"] <= 10 or stats["activeMonths"] <= 2:
        prediction = min(prediction, max(stats["last12"] * horizon / 12.0, 1.0 if stats["last12"] > 0 else 0.0))
    return max(0.0, prediction), "winsorized/trimmed monthly signal with lifecycle-specific factor"


def model_c_prediction(stats: dict, horizon: int) -> tuple[float, str]:
    if stats["recentZero"]:
        return 0.0, "recent zero revenue and sparse tail treated as likely zero"
    if stats["lifecycle"] in {"inactive", "long_tail"} or stats["last12"] <= 10 or stats["activeMonths"] <= 3:
        base = min(
            stats["last6"] / 6.0 * horizon * 0.35,
            max(stats["last12"] / 12.0 * horizon * 0.35, 1.0 if stats["last12"] > 0 else 0.0),
        )
        return max(0.0, base), "zero-inflated low-revenue guard"
    base, _ = model_b_prediction(stats, horizon)
    return base * 0.75, "moderately sparse fallback to damped robust model"


def model_d_prediction(stats: dict, horizon: int, cohort_monthly: float) -> tuple[float, str]:
    individual, _ = model_b_prediction(stats, horizon)
    cohort = max(0.0, cohort_monthly * horizon)
    reliability = min(0.85, max(0.10, stats["activeMonths"] / 24.0))
    if stats["lifecycle"] in {"inactive", "long_tail", "insufficient_history"} or stats["last12"] <= 10:
        reliability *= 0.35
    if stats["volatility"] > 1.5 or stats["peakShare"] >= 0.90:
        reliability *= 0.55
    prediction = individual * reliability + cohort * (1 - reliability)
    if stats["last12"] <= 10:
        prediction = min(prediction, max(stats["last12"] * horizon / 12.0, cohort * 0.6, 1.0 if stats["last12"] > 0 else 0.0))
    return max(0.0, prediction), "work-level signal shrunk toward lifecycle/revenue cohort prior"


def select_model(stats: dict, rating: str, scale: str, spike: bool) -> tuple[str, str]:
    if stats["lifecycle"] in {"inactive", "long_tail"} or rating in {"D", "E"} or scale in {"low", "long_tail"}:
        return "model_c_zero_inflated_sparse", "inactive, long-tail, D/E, or low-revenue guard"
    if stats["lifecycle"] == "insufficient_history" or stats["activeMonths"] < 6:
        return "model_d_hierarchical_shrinkage", "insufficient history shrinkage"
    if spike:
        return "model_b_lifecycle_robust", "spike-sensitive work uses robust trimmed signal"
    if stats["lifecycle"] in {"stable", "growth", "rebound", "declining"}:
        return "model_b_lifecycle_robust", "established lifecycle signal"
    return "model_a_trailing_baseline", "fallback baseline"


def confidence_for(stats: dict, scale: str, model_id: str, data_gap: bool = False) -> str:
    if data_gap and stats["lifecycle"] in {"inactive", "long_tail", "insufficient_history"}:
        return "blocked_for_business_use"
    if stats["lifecycle"] in {"inactive", "long_tail", "insufficient_history"} or scale in {"low", "long_tail"}:
        return "low"
    if stats["activeMonths"] >= 18 and stats["volatility"] <= 0.60 and stats["peakShare"] < 0.75 and scale in {"top", "high"}:
        return "high"
    if model_id in {"model_d_hierarchical_shrinkage", "model_c_zero_inflated_sparse"}:
        return "low"
    return "medium"


def interval_ratio(confidence: str, stats: dict, scale: str, model_id: str) -> float:
    volatility = min(2.5, max(0.0, stats["volatility"]))
    if confidence == "high":
        return min(1.48, 1.28 + volatility * 0.12)
    if confidence == "medium":
        return min(1.95, 1.55 + volatility * 0.20)
    if confidence == "low":
        return min(3.00, 2.20 + volatility * 0.28 + (0.20 if scale in {"low", "long_tail"} else 0.0))
    return min(3.25, 2.75 + volatility * 0.20)


def build_forecast_output(base: float, confidence: str, stats: dict, scale: str, model_id: str) -> dict:
    ratio = interval_ratio(confidence, stats, scale, model_id)
    side = math.sqrt(max(1.05, ratio))
    pessimistic = max(0.0, base / side)
    optimistic = max(0.0, base * side)
    return {
        "base": rounded(base, 6),
        "pessimistic": rounded(pessimistic, 6),
        "optimistic": rounded(optimistic, 6),
        "confidence": confidence,
        "optimisticPessimisticRatio": rounded(ratio),
        "intervalReason": f"{confidence} confidence interval from volatility, revenue scale, lifecycle, and model family",
        "spikeAdjusted": stats["peakShare"] >= 0.90,
        "lowRevenueGuard": stats["last12"] <= 10 or scale in {"low", "long_tail"},
        "inactiveLongTailGuard": stats["lifecycle"] in {"inactive", "long_tail"},
        "dataSufficiencyFlag": "sufficient" if stats["activeMonths"] >= 6 else "insufficient_history",
    }


def predict_models(history: np.ndarray, horizon: int, thresholds: dict, q: dict[str, float], cohort_priors: dict, rating: str = "C", data_gap: bool = False) -> dict:
    stats = history_stats(history, thresholds)
    stats["history"] = np.asarray(history, dtype=float)
    scale = scale_from_stats(stats, q)
    spike = stats["peakShare"] >= 0.90
    model_values = {}

    raw, raw_reason = raw_trailing_prediction(stats, horizon)
    model_values["raw_trailing_baseline"] = {"base": raw, "selectedModel": "raw_trailing_baseline", "modelReason": raw_reason}

    a, reason = model_a_prediction(stats, horizon)
    model_values["model_a_trailing_baseline"] = {"base": a, "selectedModel": "model_a_trailing_baseline", "modelReason": reason}

    b, reason = model_b_prediction(stats, horizon)
    model_values["model_b_lifecycle_robust"] = {"base": b, "selectedModel": "model_b_lifecycle_robust", "modelReason": reason}

    c, reason = model_c_prediction(stats, horizon)
    model_values["model_c_zero_inflated_sparse"] = {"base": c, "selectedModel": "model_c_zero_inflated_sparse", "modelReason": reason}

    cohort_key = (stats["lifecycle"], scale)
    cohort_monthly = cohort_priors.get(cohort_key, cohort_priors.get((stats["lifecycle"], "all"), 0.0))
    d, reason = model_d_prediction(stats, horizon, cohort_monthly)
    model_values["model_d_hierarchical_shrinkage"] = {"base": d, "selectedModel": "model_d_hierarchical_shrinkage", "modelReason": reason}

    selected_model, selection_reason = select_model(stats, rating, scale, spike)
    selected = model_values[selected_model]["base"]
    model_values["model_e_selector"] = {
        "base": selected,
        "selectedModel": selected_model,
        "modelReason": selection_reason,
    }

    candidate_v03 = b
    if stats["lifecycle"] == "inactive":
        candidate_v03 = min(candidate_v03, max(stats["last6"] * horizon / 6 * 0.15, stats["last12"] * horizon / 12 * 0.08))
    if stats["lifecycle"] == "long_tail":
        candidate_v03 = min(candidate_v03, max(stats["last6"] * horizon / 6 * 0.45, stats["last12"] * horizon / 12 * 0.30, 1 if stats["last12"] > 0 else 0))
    if stats["last12"] <= 10 or stats["activeMonths"] <= 2:
        candidate_v03 = min(candidate_v03, max(stats["last12"] * horizon / 12, 1 if stats["last12"] > 0 else 0))
    if spike:
        candidate_v03 *= 0.35
    model_values["candidate_b_v03"] = {"base": max(0.0, candidate_v03), "selectedModel": "candidate_b_v03", "modelReason": "candidate-b v0.3 rebuilt forecast approximation"}

    for model_id, item in list(model_values.items()):
        confidence = confidence_for(stats, scale, item["selectedModel"], data_gap=data_gap)
        output = build_forecast_output(max(0.0, item["base"]), confidence, stats, scale, item["selectedModel"])
        output.update({
            "selectedModel": item["selectedModel"],
            "modelReason": item["modelReason"],
            "lifecycle": stats["lifecycle"],
            "revenueScale": scale,
            "activeMonths": stats["activeMonths"],
            "zeroMonths": stats["zeroMonths"],
            "last12": stats["last12"],
            "peakShare": stats["peakShare"],
        })
        model_values[model_id] = output
    return model_values


def build_quantile_reference(matrix: pd.DataFrame) -> dict[str, float]:
    totals = matrix.sum(axis=1).to_numpy(dtype=float)
    return {
        "p40": float(np.quantile(totals, 0.40)),
        "p75": float(np.quantile(totals, 0.75)),
        "p95": float(np.quantile(totals, 0.95)),
    }


def build_cohort_priors(matrix: pd.DataFrame, cutoff_idx: int, thresholds: dict, q: dict[str, float]) -> dict:
    priors: dict[tuple[str, str], list[float]] = defaultdict(list)
    for _, series in matrix.iloc[:, : cutoff_idx + 1].iterrows():
        history = series.to_numpy(dtype=float)
        stats = history_stats(history, thresholds)
        scale = scale_from_stats(stats, q)
        monthly = min(stats["last12"] / 12.0 if stats["last12"] > 0 else 0.0, stats["positiveMedian"] if stats["positiveMedian"] else stats["last12"] / 12.0)
        priors[(stats["lifecycle"], scale)].append(monthly)
        priors[(stats["lifecycle"], "all")].append(monthly)
    return {key: float(np.median(values)) if values else 0.0 for key, values in priors.items()}


def rolling_cutoff_indices(months: list[str], horizon: int) -> list[int]:
    max_cutoff = len(months) - horizon - 1
    indices = list(range(24, max_cutoff + 1, 6))
    if max_cutoff >= 24 and max_cutoff not in indices:
        indices.append(max_cutoff)
    return sorted(set(indices))


def smape(predicted: float, actual: float) -> float | None:
    denominator = abs(predicted) + abs(actual)
    if denominator <= 0:
        return None
    return 2 * abs(predicted - actual) / denominator


def build_backtest_cases(matrix: pd.DataFrame, months: list[str], thresholds: dict, q: dict[str, float], feature_lookup: dict[str, dict]) -> pd.DataFrame:
    records = []
    series_rows = [(str(standard_id), series.to_numpy(dtype=float)) for standard_id, series in matrix.iterrows()]
    cohort_priors_cache = {}
    for horizon in HORIZONS:
        for cutoff_idx in rolling_cutoff_indices(months, horizon):
            if cutoff_idx not in cohort_priors_cache:
                cohort_priors_cache[cutoff_idx] = build_cohort_priors(matrix, cutoff_idx, thresholds, q)
            cohort_priors = cohort_priors_cache[cutoff_idx]
            cutoff_month = months[cutoff_idx]
            for standard_id, values in series_rows:
                history = values[: cutoff_idx + 1]
                actual_window = values[cutoff_idx + 1 : cutoff_idx + 1 + horizon]
                if len(actual_window) < horizon:
                    continue
                actual = float(np.sum(actual_window))
                features = feature_lookup.get(str(standard_id), {})
                rating = str(features.get("rating", "C"))
                risk_codes = set(features.get("riskCodes") or [])
                data_gap = bool(features.get("forecastFallbackUsed")) or "missing_copyright_end" in risk_codes
                model_outputs = predict_models(history, horizon, thresholds, q, cohort_priors, rating=rating, data_gap=data_gap)
                raw_baseline = model_outputs["raw_trailing_baseline"]["base"]
                baseline_error = abs(raw_baseline - actual)
                for model_id in ["candidate_b_v03", *MODEL_ORDER]:
                    output = model_outputs[model_id]
                    prediction = safe_float(output["base"])
                    error = abs(prediction - actual)
                    records.append({
                        "workKey": str(standard_id),
                        "anonymousWorkId": None,
                        "modelId": model_id,
                        "cutoffMonth": cutoff_month,
                        "horizonMonths": horizon,
                        "predicted": prediction,
                        "actual": actual,
                        "absoluteError": error,
                        "baselinePredicted": raw_baseline,
                        "baselineAbsoluteError": baseline_error,
                        "betterThanBaseline": error <= baseline_error,
                        "smape": smape(prediction, actual),
                        "ape": error / actual if actual > 0 else None,
                        "intervalCoverage": safe_float(output["pessimistic"]) <= actual <= safe_float(output["optimistic"]),
                        "overForecast": prediction > actual,
                        "underForecast": prediction < actual,
                        "confidence": output["confidence"],
                        "optimisticPessimisticRatio": output["optimisticPessimisticRatio"],
                        "ratingAtCutoff": rating,
                        "lifecycleAtCutoff": output["lifecycle"],
                        "revenueScaleAtCutoff": output["revenueScale"],
                        "activeMonthsAtCutoff": output["activeMonths"],
                        "activeMonthsBucketAtCutoff": count_bucket(output["activeMonths"], [3, 6, 12, 18]),
                        "zeroMonthsAtCutoff": output["zeroMonths"],
                        "zeroMonthsBucketAtCutoff": count_bucket(output["zeroMonths"], [3, 6, 12, 24]),
                        "peakShareAtCutoff": output["peakShare"],
                        "abnormalSpikeAtCutoff": bool_bucket(output["peakShare"] >= 0.90),
                        "dataGapAtCutoff": bool_bucket(data_gap),
                        "suggestionBucket": features.get("suggestionBucket", "unknown"),
                        "remainingCopyrightBucket": features.get("remainingCopyrightBucket", "unknown"),
                        "selectedModel": output["selectedModel"],
                        "selectionReason": output["modelReason"],
                    })
    frame = pd.DataFrame(records)
    return frame


def aggregate_cases(cases: pd.DataFrame) -> dict:
    if cases.empty:
        return {
            "caseCount": 0,
            "actualTotal": 0,
            "predictedTotal": 0,
            "wape": None,
            "mae": None,
            "smape": None,
            "medianApe": None,
            "intervalCoverage": None,
            "overForecastRate": None,
            "underForecastRate": None,
            "betterThanBaselineRate": None,
        }
    actual_total = safe_float(cases["actual"].sum())
    abs_total = safe_float(cases["absoluteError"].sum())
    ape = cases["ape"].dropna()
    smape_values = cases["smape"].dropna()
    return {
        "caseCount": int(len(cases)),
        "actualTotal": rounded(actual_total, 2),
        "predictedTotal": rounded(cases["predicted"].sum(), 2),
        "baselinePredictedTotal": rounded(cases["baselinePredicted"].sum(), 2),
        "wape": None if actual_total <= 0 else rounded(abs_total / actual_total),
        "mae": rounded(abs_total / len(cases), 4),
        "smape": None if smape_values.empty else rounded(smape_values.mean()),
        "medianApe": None if ape.empty else rounded(ape.median()),
        "intervalCoverage": rounded(cases["intervalCoverage"].mean()),
        "overForecastRate": rounded(cases["overForecast"].mean()),
        "underForecastRate": rounded(cases["underForecast"].mean()),
        "betterThanBaselineRate": rounded(cases["betterThanBaseline"].mean()),
        "totalAbsErrorBetterThanBaseline": bool(abs_total <= cases["baselineAbsoluteError"].sum()),
    }


def segment_metrics(cases: pd.DataFrame, model_id: str, segment_field: str) -> list[dict]:
    subset = cases[cases["modelId"] == model_id]
    rows = []
    if subset.empty:
        return rows
    for segment, group in subset.groupby(segment_field):
        metrics = aggregate_cases(group)
        fail_rate = failure_rate_for_cases(group)
        rows.append({
            "modelId": model_id,
            "segmentType": segment_field,
            "segment": str(segment),
            "wape": metrics["wape"],
            "mae": metrics["mae"],
            "smape": metrics["smape"],
            "coverage": metrics["intervalCoverage"],
            "failRate": fail_rate,
            "betterThanBaselineRate": metrics["betterThanBaselineRate"],
            "caseCount": metrics["caseCount"],
        })
    return sorted(rows, key=lambda item: (item["segmentType"], item["segment"]))


def case_failure_label(row) -> str:
    actual = safe_float(row.actual)
    predicted = safe_float(row.predicted)
    error = safe_float(row.absoluteError)
    ape = row.ape
    smape_value = row.smape
    if actual <= 1 and predicted <= 3:
        return "pass"
    if actual <= 1 and predicted > 20:
        return "fail"
    if ape is not None and safe_float(ape) > 3.0 and error > 100:
        return "fail"
    if smape_value is not None and safe_float(smape_value) > 1.35 and error > 50:
        return "fail"
    if ape is not None and safe_float(ape) > 1.25:
        return "warning"
    if not bool(row.intervalCoverage):
        return "warning"
    return "pass"


def failure_rate_for_cases(cases: pd.DataFrame) -> float:
    if cases.empty:
        return 0.0
    labels = cases.apply(case_failure_label, axis=1)
    return rounded((labels == "fail").mean())


def model_scoreboard(cases: pd.DataFrame) -> list[dict]:
    rows = []
    for model_id in MODEL_ORDER:
        subset = cases[cases["modelId"] == model_id]
        metrics = aggregate_cases(subset)
        labels = subset.apply(case_failure_label, axis=1) if not subset.empty else pd.Series(dtype=str)
        spread = spread_summary(subset)
        rows.append({
            "modelId": model_id,
            "modelName": MODEL_LABELS[model_id],
            **metrics,
            "passCount": int((labels == "pass").sum()),
            "warningCount": int((labels == "warning").sum()),
            "failCount": int((labels == "fail").sum()),
            "failRate": rounded((labels == "fail").mean()) if len(labels) else None,
            "warningRate": rounded((labels == "warning").mean()) if len(labels) else None,
            "highConfidenceCoverage": confidence_coverage(subset, "high"),
            "allCoverage": metrics["intervalCoverage"],
            "highConfidenceSpreadP75": spread["highConfidenceSpread"]["p75"],
            "nonLowConfidenceSpreadP75": spread["nonLowConfidenceSpread"]["p75"],
        })
    return rows


def confidence_coverage(cases: pd.DataFrame, confidence: str) -> float | None:
    subset = cases[cases["confidence"] == confidence]
    if subset.empty:
        return None
    return rounded(subset["intervalCoverage"].mean())


def spread_summary(cases: pd.DataFrame) -> dict:
    high = cases[cases["confidence"] == "high"]["optimisticPessimisticRatio"].dropna().tolist()
    non_low = cases[~cases["confidence"].isin(["low", "blocked_for_business_use"])]["optimisticPessimisticRatio"].dropna().tolist()
    all_ratios = cases["optimisticPessimisticRatio"].dropna().tolist()
    return {
        "allSpread": quantiles(all_ratios),
        "highConfidenceSpread": quantiles(high),
        "nonLowConfidenceSpread": quantiles(non_low),
        "byConfidence": {
            confidence: quantiles(cases[cases["confidence"] == confidence]["optimisticPessimisticRatio"].dropna().tolist())
            for confidence in CONFIDENCES
        },
    }


def stable_sample(evaluated: pd.DataFrame, size: int) -> pd.DataFrame:
    required = []
    for rating in RATING_ORDER:
        required.append(("rating", rating))
    for lifecycle in LIFECYCLES:
        required.append(("lifecycle", lifecycle))
    for bucket in REVENUE_BUCKETS:
        required.append(("revenueBucket", bucket))
    selected = []
    covered = set()

    def sort_frame(frame: pd.DataFrame) -> pd.DataFrame:
        return frame.assign(_sortAmount=frame["ratingBasisAmount"].astype(float)).sort_values(["_sortAmount", "standardWorkId"], ascending=[False, True])

    for field, value in required:
        if (field, value) in covered:
            continue
        candidates = evaluated[(evaluated[field] == value) & (~evaluated.index.isin(selected))]
        if candidates.empty:
            continue
        idx = sort_frame(candidates).index[0]
        selected.append(idx)
        row = evaluated.loc[idx]
        covered.add(("rating", row.rating))
        covered.add(("lifecycle", row.lifecycle))
        covered.add(("revenueBucket", row.revenueBucket))
        if len(selected) >= size:
            break
    while len(selected) < min(size, len(evaluated)):
        remaining = evaluated[~evaluated.index.isin(selected)].copy()
        if remaining.empty:
            break
        remaining["coverageGain"] = remaining.apply(
            lambda row: len({("rating", row.rating), ("lifecycle", row.lifecycle), ("revenueBucket", row.revenueBucket)} - covered),
            axis=1,
        )
        idx = remaining.sort_values(["coverageGain", "ratingBasisAmount", "standardWorkId"], ascending=[False, False, True]).index[0]
        selected.append(idx)
        row = evaluated.loc[idx]
        covered.add(("rating", row.rating))
        covered.add(("lifecycle", row.lifecycle))
        covered.add(("revenueBucket", row.revenueBucket))
    return evaluated.loc[selected].reset_index(drop=True)


def final_predictions(evaluated: pd.DataFrame, matrix: pd.DataFrame, thresholds: dict, q: dict[str, float]) -> pd.DataFrame:
    cohort_priors = build_cohort_priors(matrix, len(matrix.columns) - 1, thresholds, q)
    rows = []
    for index, row in evaluated.reset_index(drop=True).iterrows():
        history = matrix.loc[row.standardWorkId].to_numpy(dtype=float)
        horizon = max(1, safe_int(row.remainingMonthsForForecast, 12))
        data_gap = bool(row.forecastFallbackUsed) or "missing_copyright_end" in set(row.riskCodes)
        outputs = predict_models(history, horizon, thresholds, q, cohort_priors, rating=str(row.rating), data_gap=data_gap)
        selected = outputs["model_e_selector"]
        rows.append({
            "workKey": row.standardWorkId,
            "anonymousId": f"F{index + 1:04d}",
            "rating": row.rating,
            "lifecycle": row.lifecycle,
            "revenueBucket": row.revenueBucket,
            "riskBucket": row.riskBucket,
            "suggestionBucket": row.suggestionBucket,
            "remainingCopyrightBucket": row.remainingCopyrightBucket,
            "baseForecast": selected["base"],
            "optimisticForecast": selected["optimistic"],
            "pessimisticForecast": selected["pessimistic"],
            "forecastConfidence": selected["confidence"],
            "selectedModel": selected["selectedModel"],
            "selectionReason": selected["modelReason"],
            "intervalReason": selected["intervalReason"],
            "dataSufficiencyFlag": selected["dataSufficiencyFlag"],
            "spikeAdjustedFlag": selected["spikeAdjusted"],
            "lowRevenueGuardFlag": selected["lowRevenueGuard"],
            "inactiveLongTailGuardFlag": selected["inactiveLongTailGuard"],
            "activeMonths": selected["activeMonths"],
            "activeMonthsBucket": count_bucket(selected["activeMonths"], [3, 6, 12, 18]),
            "zeroMonths": selected["zeroMonths"],
            "zeroMonthsBucket": count_bucket(selected["zeroMonths"], [3, 6, 12, 24]),
            "dataGapFlag": bool_bucket(data_gap),
            "abnormalSpikeFlag": bool_bucket(selected["spikeAdjusted"]),
            "remainingCopyrightMonths": None if pd.isna(row.remainingCopyrightMonths) else safe_int(row.remainingCopyrightMonths),
            "yearlyForecastBreakdown": yearly_breakdown(selected["base"], horizon, row.latestCompleteMonth),
        })
    return pd.DataFrame(rows)


def yearly_breakdown(base: float, horizon: int, latest_month: str) -> list[dict]:
    monthly = safe_float(base) / max(1, horizon)
    first_month = add_months(latest_month, 1)
    by_year = defaultdict(float)
    for offset in range(horizon):
        year = add_months(first_month, offset)[:4]
        by_year[year] += monthly
    return [{"year": year, "baseForecast": rounded(amount, 2)} for year, amount in sorted(by_year.items())]


def enrich_evaluated(evaluated: pd.DataFrame) -> pd.DataFrame:
    frame = evaluated.copy()
    total = frame["totalHistoricalRevenue"].astype(float).to_numpy()
    q = {"p40": float(np.quantile(total, 0.40)), "p75": float(np.quantile(total, 0.75)), "p95": float(np.quantile(total, 0.95))}
    frame["revenueBucket"] = frame.apply(lambda row: "long_tail" if row.lifecycle == "long_tail" else revenue_bucket(row.totalHistoricalRevenue, q), axis=1)
    frame["riskBucket"] = frame["riskCodes"].map(primary_risk_bucket)
    frame["suggestionBucket"] = frame["suggestionCodes"].map(lambda items: (items or ["observe_only"])[0])
    frame["remainingCopyrightBucket"] = frame["remainingCopyrightMonths"].map(remaining_bucket)
    return frame


def primary_risk_bucket(risks: list[str]) -> str:
    risk_set = set(risks or [])
    if risk_set.intersection({"abnormal_spike", "buyout_or_oneoff_income"}):
        return "abnormal_spike"
    if risk_set.intersection({"missing_copyright_end", "copyright_date_conflict", "aggregate_projection_gap"}):
        return "data_gap_or_copyright_fallback"
    if risk_set.intersection({"insufficient_history", "insufficient_revenue_history"}):
        return "insufficient_history"
    if "copyright_expiry" in risk_set:
        return "copyright_expiry"
    if risk_set.intersection({"channel_concentration", "channel_concentration_advisory"}):
        return "channel_concentration"
    if "inactive_tail" in risk_set:
        return "inactive_tail"
    if "revenue_decline" in risk_set:
        return "revenue_decline"
    return "no_major_risk" if not risk_set else "other_risk"


def remaining_bucket(value) -> str:
    if value is None or pd.isna(value):
        return "fallback"
    months = safe_int(value)
    if months <= 0:
        return "expired_or_zero"
    if months <= 12:
        return "0_to_12"
    if months <= 24:
        return "13_to_24"
    if months <= 60:
        return "25_to_60"
    return "gt_60"


def count_bucket(value, boundaries: list[int]) -> str:
    number = safe_int(value)
    previous = 0
    for boundary in boundaries:
        if number <= boundary:
            return f"{previous + 1}_to_{boundary}" if previous else f"0_to_{boundary}"
        previous = boundary
    return f"gt_{boundaries[-1]}"


def bool_bucket(value) -> str:
    return "yes" if bool(value) else "no"


def model_validation_summary(cases: pd.DataFrame, final_outputs: pd.DataFrame, sample_200: pd.DataFrame) -> dict:
    selected_cases = cases[cases["modelId"] == "model_e_selector"].copy()
    sample_keys = set(sample_200["standardWorkId"].astype(str))
    selected_sample_cases = selected_cases[selected_cases["workKey"].isin(sample_keys)]
    sample_labels = selected_sample_cases.apply(case_failure_label, axis=1)
    full_labels = selected_cases.apply(case_failure_label, axis=1)
    p0 = p1 = p2 = 0
    issue_rows = []
    for _, row in final_outputs.iterrows():
        issues = []
        if row["forecastConfidence"] == "high" and row["baseForecast"] <= 0 and row["rating"] in {"S+", "S", "A"}:
            issues.append(("P0", "high rating with zero high-confidence forecast"))
        if row["lifecycle"] in {"inactive", "long_tail"} and row["baseForecast"] > 100 and row["revenueBucket"] in {"low", "long_tail"}:
            issues.append(("P1", "inactive/long-tail elevated forecast"))
        if row["forecastConfidence"] in {"low", "blocked_for_business_use"} and row["suggestionBucket"] in {"promote", "downlist_or_suspend", "renewal_review"}:
            issues.append(("P2", "action-bearing suggestion has low-confidence forecast"))
        for severity, reason in issues:
            if severity == "P0":
                p0 += 1
            elif severity == "P1":
                p1 += 1
            else:
                p2 += 1
            issue_rows.append({
                "anonymousId": row["anonymousId"],
                "issueType": severity,
                "reason": reason,
                "selectedModel": row["selectedModel"],
                "requiredFix": "manual review or data/model correction before formal release",
            })
    high_cases = selected_cases[selected_cases["confidence"] == "high"]
    non_low_cases = selected_cases[~selected_cases["confidence"].isin(["low", "blocked_for_business_use"])]
    return {
        "p0": p0,
        "p1": p1,
        "p2": p2,
        "issueRows": issue_rows,
        "sample200PassWarningFail": distribution(sample_labels, ["pass", "warning", "fail"]),
        "fullPassWarningFail": distribution(full_labels, ["pass", "warning", "fail"]),
        "sample200FailRate": rounded((sample_labels == "fail").mean()) if len(sample_labels) else None,
        "sample200WarningRate": rounded((sample_labels == "warning").mean()) if len(sample_labels) else None,
        "fullFailRate": rounded((full_labels == "fail").mean()) if len(full_labels) else None,
        "highConfidenceCoverage": None if high_cases.empty else rounded(high_cases["intervalCoverage"].mean()),
        "allCoverage": rounded(selected_cases["intervalCoverage"].mean()) if not selected_cases.empty else None,
        "highConfidenceSpreadP75": quantiles(high_cases["optimisticPessimisticRatio"].dropna().tolist())["p75"],
        "nonLowConfidenceSpreadP75": quantiles(non_low_cases["optimisticPessimisticRatio"].dropna().tolist())["p75"],
    }


def classify_verdict(summary: dict) -> str:
    if (
        summary["p0"] == 0
        and summary["p1"] <= 3
        and safe_float(summary["sample200FailRate"], 1) <= 0.10
        and safe_float(summary["sample200WarningRate"], 1) <= 0.50
        and safe_float(summary["fullFailRate"], 1) <= 0.20
        and safe_float(summary["allCoverage"], 0) >= 0.45
        and (summary["highConfidenceCoverage"] is None or safe_float(summary["highConfidenceCoverage"]) >= 0.60)
        and (summary["highConfidenceSpreadP75"] is None or safe_float(summary["highConfidenceSpreadP75"]) <= 1.50)
        and (summary["nonLowConfidenceSpreadP75"] is None or safe_float(summary["nonLowConfidenceSpreadP75"]) <= 2.00)
    ):
        return "PASS"
    if (
        summary["p0"] == 0
        and summary["p1"] <= 10
        and safe_float(summary["sample200FailRate"], 1) <= 0.20
        and safe_float(summary["fullFailRate"], 1) <= 0.35
        and safe_float(summary["allCoverage"], 0) >= 0.35
    ):
        return "CONDITIONAL PASS"
    return "FAIL"


def root_cause_audit(candidate_cases: pd.DataFrame) -> dict:
    rows = []
    dimensions = [
        ("lifecycle", "lifecycleAtCutoff"),
        ("revenueScale", "revenueScaleAtCutoff"),
        ("confidence", "confidence"),
        ("horizon", "horizonMonths"),
    ]
    for dimension, field in dimensions:
        for segment, group in candidate_cases.groupby(field):
            metrics = aggregate_cases(group)
            rows.append({
                "dimension": dimension,
                "segment": str(segment),
                "count": metrics["caseCount"],
                "actualTotal": metrics["actualTotal"],
                "predictedTotal": metrics["predictedTotal"],
                "wape": metrics["wape"],
                "mape": metrics["medianApe"],
                "mae": metrics["mae"],
                "coverage": metrics["intervalCoverage"],
                "candidateBetterThanBaselineRate": metrics["betterThanBaselineRate"],
            })
    factor_flags = {
        "low_income_denominator_mape_explosion": bool(any(row["segment"] in {"low", "long_tail"} and safe_float(row["mape"]) > 3 for row in rows)),
        "inactive_long_tail_overforecast": bool(any(row["segment"] in {"inactive", "long_tail"} and safe_float(row["wape"]) > 1 for row in rows)),
        "spike_not_fully_removed": bool(candidate_cases["peakShareAtCutoff"].fillna(0).ge(0.90).any()),
        "trailing_baseline_better_for_many_cases": bool(safe_float(aggregate_cases(candidate_cases)["betterThanBaselineRate"]) < 0.60),
        "lifecycle_forecast_mismatch": bool(any(row["dimension"] == "lifecycle" and safe_float(row["wape"]) > 1.0 for row in rows)),
        "remaining_copyright_horizon_amplifies_error": "not directly measurable in rolling cutoff; current reports show fallback and horizon caps remain material",
        "sparse_data_instability": bool(candidate_cases["activeMonthsAtCutoff"].fillna(0).le(3).mean() > 0.05),
        "validation_low_amount_penalty": True,
    }
    return {
        "schema": "m2.forecast_model_failure_root_cause_audit.v0.1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "failedCandidate": FAILED_CANDIDATE,
        "conclusion": "candidate-b forecast route should be abandoned for M2 baseline selection; failures are structural rather than another patch target.",
        "factorFlags": factor_flags,
        "failureGroups": rows,
        "classification": {
            "algorithmIssues": [
                "inactive and long-tail rows remain hard to calibrate with lifecycle-factor extrapolation",
                "scenario intervals do not reach required coverage after fixed multiplier removal",
                "model route is too coupled to last-12 extrapolation and lifecycle caps",
            ],
            "metricIssues": [
                "MAPE becomes unstable on very low actual revenue denominators; WAPE/SMAPE/median APE are safer acceptance metrics",
            ],
            "dataSparsityIssues": [
                "low active-month count and zero-heavy histories need a zero-inflated or shrinkage model instead of point extrapolation",
            ],
        },
        "abandonCandidateBForecastRoute": True,
        "safeOutputBoundary": safe_boundary(),
    }


def safe_boundary() -> dict:
    return {
        "rawRowsWritten": False,
        "realWorkNamesWritten": False,
        "realAuthorNamesWritten": False,
        "realChannelNamesWritten": False,
        "connectionStringsWritten": False,
        "notFinalReleaseApproval": True,
        "m3Started": False,
    }


def write_reports(root_cause: dict, bakeoff: dict, selection: dict, readiness: dict) -> None:
    write_json(ROOT_CAUSE_JSON, root_cause)
    write_json(BAKEOFF_JSON, bakeoff)
    write_json(SELECTION_JSON, selection)
    write_json(READINESS_JSON, readiness)

    ROOT_CAUSE_MD.write_text(
        "\n".join([
            "# M2 Forecast Model Failure Root-Cause Audit v0.1",
            "",
            f"Failed candidate: `{FAILED_CANDIDATE}`",
            "",
            f"Conclusion: {root_cause['conclusion']}",
            "",
            "## Factor Flags",
            "",
            markdown_table([{"factor": key, "value": value} for key, value in root_cause["factorFlags"].items()], [("factor", "Factor"), ("value", "Value")]),
            "",
            "## Failure Groups",
            "",
            markdown_table(root_cause["failureGroups"][:40], [("dimension", "Dimension"), ("segment", "Segment"), ("count", "Count"), ("actualTotal", "Actual"), ("predictedTotal", "Predicted"), ("wape", "WAPE"), ("mape", "Median APE"), ("mae", "MAE"), ("coverage", "Coverage"), ("candidateBetterThanBaselineRate", "Better Than Baseline")]),
            "",
            "This report is sanitized and aggregate-only.",
            "",
        ]),
        encoding="utf-8",
    )

    BAKEOFF_MD.write_text(
        "\n".join([
            "# M2 Forecast Model Bake-Off v1",
            "",
            "candidate-b is treated as a failed route and is not patched further in this report.",
            "",
            "## Model Scoreboard",
            "",
            markdown_table(bakeoff["modelScoreboard"], [("modelId", "Model"), ("wape", "WAPE"), ("mae", "MAE"), ("smape", "SMAPE"), ("intervalCoverage", "Coverage"), ("betterThanBaselineRate", "Better Than Baseline"), ("failRate", "Fail Rate"), ("warningRate", "Warning Rate"), ("highConfidenceSpreadP75", "High Conf Spread P75"), ("nonLowConfidenceSpreadP75", "Non-low Spread P75")]),
            "",
            "## Selected Model Validation",
            "",
            markdown_table([selection["validationSummary"]], [("candidateVersion", "Candidate"), ("verdict", "Verdict"), ("sample200FailRate", "200 Fail Rate"), ("sample200WarningRate", "200 Warning Rate"), ("fullFailRate", "Full Fail Rate"), ("p0", "P0"), ("p1", "P1"), ("p2", "P2"), ("allCoverage", "All Coverage")]),
            "",
            "This report is sanitized and aggregate-only.",
            "",
        ]),
        encoding="utf-8",
    )

    SELECTION_MD.write_text(
        "\n".join([
            "# M2 Forecast Model Selection v1",
            "",
            f"Selected candidate: `{selection['validationSummary']['candidateVersion']}`",
            f"Verdict: `{selection['validationSummary']['verdict']}`",
            "",
            "## Why This Model",
            "",
            selection["selectionConclusion"],
            "",
            "## Segment Summary",
            "",
            markdown_table(selection["keySegmentSummary"], [("segmentType", "Segment Type"), ("segment", "Segment"), ("wape", "WAPE"), ("coverage", "Coverage"), ("failRate", "Fail Rate"), ("caseCount", "Cases")]),
            "",
            "This report is sanitized and aggregate-only.",
            "",
        ]),
        encoding="utf-8",
    )

    READINESS_MD.write_text(
        "\n".join([
            "# M2 Algorithm Business Readiness v1",
            "",
            f"Business review readiness: `{readiness['businessReviewReadiness']}`",
            f"M3 allowed: `{readiness['m3Allowed']}`",
            "",
            "## Readiness By Group",
            "",
            markdown_table(readiness["readinessGroups"], [("group", "Group"), ("status", "Status"), ("reason", "Reason")]),
            "",
            "This report is sanitized and aggregate-only. It is not final production release approval.",
            "",
        ]),
        encoding="utf-8",
    )


def write_excel(bakeoff: dict, cases: pd.DataFrame, evaluated: pd.DataFrame, final_outputs: pd.DataFrame, sample_200: pd.DataFrame, validation: dict) -> None:
    if Alignment is None:
        raise RuntimeError("openpyxl is required to write the private Excel workbook.")
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    selected_cases = cases[cases["modelId"].isin(MODEL_ORDER)].copy()
    detail = selected_cases.head(30000).copy()
    detail["anonymousSampleId"] = detail.groupby("workKey").ngroup().map(lambda value: f"B{value + 1:05d}")
    sample_200_outputs = final_outputs[final_outputs["workKey"].isin(set(sample_200["standardWorkId"].astype(str)))].copy()
    sample_200_outputs = sample_200_outputs.head(200).copy()
    sample_200_outputs["sampleId"] = [f"S{index + 1:03d}" for index in range(len(sample_200_outputs))]
    fail_cases = pd.DataFrame(validation["issueRows"])
    if fail_cases.empty:
        fail_cases = pd.DataFrame([{"anonymousId": "none", "issueType": "none", "reason": "no P0/P1/fail selected-model output issue", "selectedModel": "model_e_selector", "requiredFix": "none"}])

    with pd.ExcelWriter(PRIVATE_XLSX, engine="openpyxl") as writer:
        pd.DataFrame([
            {"item": "purpose", "value": "M2 forecast model replacement bake-off"},
            {"item": "candidate-b route", "value": "abandoned for M2 forecast baseline"},
            {"item": "pass standard", "value": "P0=0, P1<=3, 200 fail<=10%, full fail<=20%, coverage/spread guardrails"},
            {"item": "safety", "value": "anonymous IDs only; no real work names, authors, channels, or raw rows"},
        ]).to_excel(writer, sheet_name="00_read_me", index=False)
        pd.DataFrame(bakeoff["modelScoreboard"]).to_excel(writer, sheet_name="01_model_scoreboard", index=False)
        pd.DataFrame(bakeoff["segmentComparison"]).to_excel(writer, sheet_name="02_model_comparison_by_segment", index=False)
        pd.DataFrame(bakeoff["forecastSpreadRows"]).to_excel(writer, sheet_name="03_forecast_spread", index=False)
        detail[["anonymousSampleId", "modelId", "cutoffMonth", "horizonMonths", "predicted", "actual", "absoluteError", "baselinePredicted", "betterThanBaseline", "confidence"]].to_excel(writer, sheet_name="04_backtest_detail", index=False)
        sample_200_outputs[["sampleId", "selectedModel", "lifecycle", "rating", "revenueBucket", "baseForecast", "forecastConfidence", "selectionReason"]].to_excel(writer, sheet_name="05_200_sample_validation", index=False)
        pd.DataFrame(bakeoff["fullCohortSegments"]).to_excel(writer, sheet_name="06_full_cohort_segments", index=False)
        fail_cases.to_excel(writer, sheet_name="07_fail_cases", index=False)
        sample_200_outputs[["sampleId", "baseForecast", "optimisticForecast", "pessimisticForecast", "forecastConfidence", "selectedModel", "selectionReason", "intervalReason"]].to_excel(writer, sheet_name="08_selected_model_outputs", index=False)
        pd.DataFrame(bakeoff["businessReadinessRows"]).to_excel(writer, sheet_name="09_business_readiness", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, column in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in column:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(10, min(max_len + 3, 45))
            worksheet.auto_filter.ref = worksheet.dimensions


def business_readiness(validation: dict, final_outputs: pd.DataFrame) -> dict:
    groups = [
        {"group": "high_confidence_stable_growth_top_high", "status": "business_review_ready", "reason": "high/medium confidence and enough history"},
        {"group": "inactive_long_tail_D_E_low_revenue", "status": "low_confidence_manual_review", "reason": "zero-inflated guard prevents overforecast but business action remains manual"},
        {"group": "insufficient_history", "status": "manual_review_required", "reason": "model uses shrinkage prior and cannot be formal without more months"},
        {"group": "data_gap_or_copyright_fallback", "status": "manual_review_required", "reason": "forecast may exist but source readiness is incomplete"},
        {"group": "abnormal_spike_or_oneoff", "status": "manual_review_required", "reason": "spike damping is deterministic but business cause remains unconfirmed"},
    ]
    can_review = validation["verdict"] in {"PASS", "CONDITIONAL PASS"}
    return {
        "schema": "m2.algorithm_business_readiness.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "businessReviewReadiness": "conditional_business_review_ready" if can_review else "not_ready",
        "m3Allowed": False,
        "reason": "M3 remains blocked until model pass, review decisions, release gates, and formal approval are complete.",
        "readinessGroups": groups,
        "selectedOutputConfidenceDistribution": distribution(final_outputs["forecastConfidence"], CONFIDENCES),
        "safeOutputBoundary": safe_boundary(),
    }


def run() -> dict:
    context = load_analysis_inputs()
    evaluated = evaluate_work_summary(
        context["work_summary"],
        context["parameters"],
        context["latest_complete_month"],
        context["incomplete_work_ids"],
        PARAMETER_VARIANT,
    ).sort_values("standardWorkId").reset_index(drop=True)
    evaluated = enrich_evaluated(evaluated)
    matrix, months = build_month_matrix(context)
    q = build_quantile_reference(matrix)
    thresholds = context["parameters"]["lifecycle"]
    feature_lookup = {
        str(row.standardWorkId): {
            "rating": row.rating,
            "riskCodes": row.riskCodes,
            "forecastFallbackUsed": row.forecastFallbackUsed,
            "suggestionBucket": row.suggestionBucket,
            "remainingCopyrightBucket": row.remainingCopyrightBucket,
        }
        for _, row in evaluated.iterrows()
    }
    cases = build_backtest_cases(matrix, months, thresholds, q, feature_lookup)
    sample_200 = stable_sample(evaluated, 200)
    final_outputs = final_predictions(evaluated, matrix, thresholds, q)
    scoreboard = model_scoreboard(cases)
    selected_cases = cases[cases["modelId"] == "model_e_selector"]
    segment_comparison = []
    for model_id in MODEL_ORDER:
        for field in [
            "ratingAtCutoff",
            "lifecycleAtCutoff",
            "revenueScaleAtCutoff",
            "activeMonthsBucketAtCutoff",
            "zeroMonthsBucketAtCutoff",
            "abnormalSpikeAtCutoff",
            "dataGapAtCutoff",
            "suggestionBucket",
            "remainingCopyrightBucket",
            "horizonMonths",
            "confidence",
        ]:
            segment_comparison.extend(segment_metrics(cases, model_id, field))
    full_segments = []
    for field in [
        "rating",
        "lifecycle",
        "revenueBucket",
        "activeMonthsBucket",
        "zeroMonthsBucket",
        "abnormalSpikeFlag",
        "dataGapFlag",
        "riskBucket",
        "suggestionBucket",
        "remainingCopyrightBucket",
    ]:
        for segment, group in final_outputs.groupby(field):
            full_segments.append({
                "segmentType": field,
                "segment": str(segment),
                "count": int(len(group)),
                "baseForecastTotal": rounded(group["baseForecast"].sum(), 2),
                "confidenceDistribution": json.dumps(distribution(group["forecastConfidence"], CONFIDENCES), ensure_ascii=False),
            })
    spread_rows = []
    for model_id in MODEL_ORDER:
        subset = cases[cases["modelId"] == model_id]
        spread = spread_summary(subset)
        spread_rows.append({"modelId": model_id, "confidence": "all", **spread["allSpread"]})
        for confidence, values in spread["byConfidence"].items():
            spread_rows.append({"modelId": model_id, "confidence": confidence, **values})
    validation = model_validation_summary(cases, final_outputs, sample_200)
    verdict = classify_verdict(validation)
    candidate_version = NEW_CANDIDATE if verdict == "PASS" else CONDITIONAL_CANDIDATE if verdict == "CONDITIONAL PASS" else None
    validation_summary = {
        "candidateVersion": candidate_version,
        "verdict": verdict,
        **{key: validation[key] for key in ["p0", "p1", "p2", "sample200FailRate", "sample200WarningRate", "fullFailRate", "highConfidenceCoverage", "allCoverage", "highConfidenceSpreadP75", "nonLowConfidenceSpreadP75"]},
    }
    bakeoff = {
        "schema": "m2.forecast_model_bakeoff.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "failedCandidateRoute": FAILED_CANDIDATE,
        "candidateBSeriesAbandoned": True,
        "modelsCompared": MODEL_ORDER,
        "modelScoreboard": scoreboard,
        "segmentComparison": segment_comparison,
        "forecastSpreadRows": spread_rows,
        "fullCohortSegments": full_segments,
        "businessReadinessRows": business_readiness({**validation, "verdict": verdict}, final_outputs)["readinessGroups"],
        "selectedModelValidation": validation_summary,
        "safeOutputBoundary": safe_boundary(),
    }
    key_segment_summary = [
        row for row in segment_comparison
        if row["modelId"] == "model_e_selector"
        and row["segmentType"] in {"ratingAtCutoff", "lifecycleAtCutoff", "revenueScaleAtCutoff"}
    ][:60]
    selection = {
        "schema": "m2.forecast_model_selection.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateBSeriesAbandoned": True,
        "newCandidateVersion": candidate_version,
        "selectionConclusion": (
            "Model E selector wins because it routes sparse/inactive/long-tail works to zero-inflated guards, insufficient history to shrinkage, and established works to robust lifecycle forecasts."
            if verdict in {"PASS", "CONDITIONAL PASS"}
            else "No model meets M2 acceptance standards; do not promote a new forecast baseline."
        ),
        "validationSummary": validation_summary,
        "keySegmentSummary": key_segment_summary,
        "safeOutputBoundary": safe_boundary(),
    }
    readiness = business_readiness({**validation, "verdict": verdict}, final_outputs)
    root_cause = root_cause_audit(cases[cases["modelId"] == "candidate_b_v03"])

    write_reports(root_cause, bakeoff, selection, readiness)
    write_excel(bakeoff, cases, evaluated, final_outputs, sample_200, {**validation, "verdict": verdict})
    write_json(PRIVATE_DETAIL_JSON, {
        "schema": "m2.private.forecast_model_bakeoff_detail.v1",
        "notForCommit": True,
        "selectedModelValidation": validation_summary,
        "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
    })
    return {
        "status": verdict,
        "candidateBSeriesAbandoned": True,
        "newCandidateVersion": candidate_version,
        "modelBakeoffCompleted": True,
        "selectedModelValidation": validation_summary,
        "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
        "sanitizedReports": [
            str(ROOT_CAUSE_JSON.relative_to(ROOT)),
            str(BAKEOFF_JSON.relative_to(ROOT)),
            str(SELECTION_JSON.relative_to(ROOT)),
            str(READINESS_JSON.relative_to(ROOT)),
        ],
    }


if __name__ == "__main__":
    result = run()
    print(json.dumps(json_safe(result), ensure_ascii=False, indent=2))
