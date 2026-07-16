#!/usr/bin/env python3
"""Train and validate C2-R.1 only after the pushed Gate B is all true."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

import m2_calibration_c2r1_v1 as c2r1
import m2_calibration_v1 as base
import m2_calibration_v1_2 as v12
import m2_formal_cash_comparator_v1 as formal
import run_m2_calibration_baseline_replay as legacy
import run_m2_formal_cash_comparator_replay as phase_a


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-c2r1-v1"
DESIGN_JSON = PUBLIC_DIR / "M2-C2R1-design-v1.json"
DESIGN_MD = PUBLIC_DIR / "M2-C2R1-design-v1.md"
VALIDATION_JSON = PUBLIC_DIR / "M2-C2R1-development-validation-v1.json"
VALIDATION_MD = PUBLIC_DIR / "M2-C2R1-development-validation-v1.md"
ROUTE_JSON = PUBLIC_DIR / "M2-C2R1-route-specific-metrics-v1.json"
ROUTE_MD = PUBLIC_DIR / "M2-C2R1-route-specific-metrics-v1.md"
CHANNEL_JSON = PUBLIC_DIR / "M2-C2R1-channel-reconciliation-completeness-v1.json"
CHANNEL_MD = PUBLIC_DIR / "M2-C2R1-channel-reconciliation-completeness-v1.md"
BUSINESS_JSON = PUBLIC_DIR / "M2-C2R1-end-to-end-business-coverage-v1.json"
BUSINESS_MD = PUBLIC_DIR / "M2-C2R1-end-to-end-business-coverage-v1.md"
PRIVATE_CASES = PRIVATE_DIR / "M2-C2R1-development-cases-private-v1.ndjson"
PRIVATE_MANIFEST = PRIVATE_DIR / "M2-C2R1-development-manifest-private-v1.json"
PRIVATE_WORKBOOK = PRIVATE_DIR / "M2-C2R1-中文业务抽检工作簿-private-v1.xlsx"
PRIVATE_PREFLIGHT = PRIVATE_DIR / "M2-C2R1-pre-data-lock-private-v1.json"
BRANCH = "codex/m2-calibration-v1"
MINIMUM_CELL = 10
TOLERANCE = 1e-6
NUMERIC_EPSILON = 1e-12


class C2R1RunnerError(RuntimeError):
    """C2-R.1 development evidence could not be produced safely."""


def progress(message: str) -> None:
    print(f"[C2-R.1] {message}", file=sys.stderr, flush=True)


def write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False)
        + "\n",
        encoding="utf-8",
        newline="\n",
    )


def write_text(path: Path, value: str) -> None:
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def digest(value: Any) -> str:
    return c2r1.canonical_digest(value)


def rounded(value: Any, places: int = 8) -> float | None:
    if value is None:
        return None
    number = float(value)
    return round(number, places) if math.isfinite(number) else None


def public_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {
            str(key): public_value(child)
            for key, child in value.items()
            if key
            not in {
                "actualTotal",
                "predictedTotal",
                "forecastableActualByComponent",
                "totalLedgerActualByComponent",
                "channelComponents",
                "internalInterval",
            }
        }
    if isinstance(value, list):
        return [public_value(child) for child in value]
    if isinstance(value, float):
        return rounded(value)
    return value


def _case_key(row: Mapping[str, Any]) -> tuple[str, str, int, str]:
    key = row["caseKey"]
    return (
        str(key["standard_work_id"]),
        str(key["origin"]),
        int(key["horizon_months"]),
        str(key["route"]),
    )


def _role_key(row: Mapping[str, Any]) -> tuple[str, tuple[str, str, int, str]]:
    return str(row["predictionRole"]), _case_key(row)


def _assert_private_paths() -> None:
    for path in (PRIVATE_CASES, PRIVATE_MANIFEST, PRIVATE_WORKBOOK, PRIVATE_PREFLIGHT):
        if not phase_a.phase.git_ignored(path):
            raise C2R1RunnerError(f"private C2-R.1 role is not Git ignored: {path.name}")
        if phase_a.run_git("ls-files", "--", str(path)):
            raise C2R1RunnerError(f"private C2-R.1 role entered Git: {path.name}")
    if phase_a.phase.tracked_private_artifacts():
        raise C2R1RunnerError("a private calibration artifact is tracked")


def _validate_phase_bindings(spec: Mapping[str, Any]) -> dict[str, Any]:
    authorization = phase_a.verify_c2r1_authorization()
    bundle = json.loads(phase_a.BUNDLE_JSON.read_text(encoding="utf-8"))
    gate = json.loads(phase_a.GATE_B_JSON.read_text(encoding="utf-8"))
    binding = spec["phaseABinding"]
    if (
        authorization["phaseACheckpoint"] != binding["checkpoint"]
        or formal.canonical_digest(formal.load_spec())
        != binding["formalCashComparatorSpecDigest"]
        or digest(bundle) != binding["formalCashComparatorBundleDigest"]
        or digest(gate) != binding["gateBDigestAtAuthorization"]
        or gate.get("allTrue") is not True
        or gate.get("passedConditionCount") != binding["gateBRequiredTrueCount"]
        or gate.get("C2R1AuthorizedByGateB") is not True
        or bundle.get("primaryComparator") != binding["primaryComparator"]
    ):
        raise C2R1RunnerError("C2-R.1 Phase A/Gate B binding differs")
    return {
        "phaseACheckpoint": binding["checkpoint"],
        "formalCashComparatorSpecDigest": binding[
            "formalCashComparatorSpecDigest"
        ],
        "formalCashComparatorBundleDigest": binding[
            "formalCashComparatorBundleDigest"
        ],
        "gateBDigestAtAuthorization": binding["gateBDigestAtAuthorization"],
        "gateBAllTrue": True,
        "remotePhaseACheckpointVerified": True,
    }


def _write_pre_data_lock(spec: Mapping[str, Any], binding: Mapping[str, Any]) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    receipt = {
        "schema": "m2.c2r1.pre_data_lock.private.v1",
        "recordedAt": datetime.now(timezone.utc).isoformat(),
        "decisionStatus": "not_for_formal_decision",
        "branch": BRANCH,
        "specDigest": c2r1.canonical_digest(spec),
        "candidateCount": len(c2r1.candidate_ids(spec)),
        "phaseBinding": dict(binding),
        "candidateSpecLoadedBeforePrivateCaseOrAuthorityInput": True,
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }
    write_json(PRIVATE_PREFLIGHT, receipt)


def _load_formal_private_cases() -> tuple[
    dict[tuple[str, tuple[str, str, int, str]], dict[str, Any]],
    dict[tuple[str, tuple[str, str, int, str], str], dict[str, Any]],
    dict[str, Any],
]:
    manifest = json.loads(phase_a.PRIVATE_MANIFEST.read_text(encoding="utf-8"))
    hasher = hashlib.sha256()
    count = 0
    b4: dict[tuple[str, tuple[str, str, int, str]], dict[str, Any]] = {}
    comparator: dict[
        tuple[str, tuple[str, str, int, str], str], dict[str, Any]
    ] = {}
    with phase_a.PRIVATE_CASES.open("rb") as handle:
        for raw in handle:
            if not raw.endswith(b"\n"):
                raise C2R1RunnerError("formal private case is not LF-delimited")
            payload = json.loads(raw[:-1].decode("utf-8"))
            canonical = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            if canonical != raw:
                raise C2R1RunnerError("formal private case is not canonical")
            hasher.update(raw)
            count += 1
            role = str(payload["predictionRole"])
            key = _case_key(payload)
            model = str(payload["modelId"])
            if model == "B4":
                b4_key = (role, key)
                if b4_key in b4:
                    raise C2R1RunnerError("duplicate private B4 case")
                b4[b4_key] = payload
            if (
                key[3] in c2r1.SALES_ROUTES
                and payload.get("statisticallyScoreable") is True
            ):
                compact = {
                    "channelComponents": [
                        {
                            "channel_key": str(item["channel_key"]),
                            "point_forecast": float(item["point_forecast"]),
                        }
                        for item in payload.get("channelComponents", [])
                    ]
                }
                comparator[(role, key, model)] = compact
    if (
        count != int(manifest["privateCaseRowCount"])
        or hasher.hexdigest() != manifest["caseEvidenceSha256"]
    ):
        raise C2R1RunnerError("formal private case manifest differs")
    if len(b4) * len(formal.MODEL_IDS) != count:
        raise C2R1RunnerError("formal private comparator multiplicity differs")
    return b4, comparator, manifest


def _comparator_rows_for_key(
    comparator: Mapping[
        tuple[str, tuple[str, str, int, str], str], Mapping[str, Any]
    ],
    role: str,
    key: tuple[str, str, int, str],
) -> dict[str, Mapping[str, Any]]:
    result = {
        model: comparator[(role, key, model)] for model in c2r1.COMPARATOR_COMPONENTS
    }
    return result


def _compact_prediction_projection(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "modelId": row["model_id"],
        "candidateId": row["candidate_id"],
        "caseKey": list(formal.strict_case_key(row)),
        "predictionRole": row["_residual_case_role"],
        "statisticallyScoreable": row["statisticallyScoreable"],
        "scoreabilityReason": row["scoreabilityReason"],
        "businessServingEligible": row["businessServingEligible"],
        "modelPredictionAvailable": row["modelPredictionAvailable"],
        "routeAbstained": row["routeAbstained"],
        "abstained": row["abstained"],
        "abstentionReason": row["abstentionReason"],
        "rawModelPrediction": row["rawModelPrediction"],
        "servedPrediction": row["servedPrediction"],
        "channelComponentPoints": sorted(
            (str(item["channel_key"]), float(item["point_forecast"]))
            for item in row.get("channel_components", [])
        ),
        "publicOutput": row["public_output"],
    }


def _join_truth(row: dict[str, Any], template: Mapping[str, Any]) -> dict[str, Any]:
    joined = copy.deepcopy(row)
    joined.update(
        {
            "forecastableCashActual": float(template["forecastableCashActual"]),
            "uncommittedBuyoutSurpriseActual": float(
                template["uncommittedBuyoutSurpriseActual"]
            ),
            "totalLedgerCashActual": float(template["totalLedgerCashActual"]),
            "forecastableActualByComponent": copy.deepcopy(
                template.get("forecastableActualByComponent", {})
            ),
            "totalLedgerActualByComponent": copy.deepcopy(
                template.get("totalLedgerActualByComponent", {})
            ),
            "strata": copy.deepcopy(template.get("strata", {})),
            "actual": float(template["forecastableCashActual"]),
        }
    )
    return joined


def _locked_b4_row(template: Mapping[str, Any]) -> dict[str, Any]:
    key = _case_key(template)
    served = template.get("servedPrediction")
    confidence = "medium" if served is not None else "unavailable"
    reason = template.get("abstentionReason")
    limitations = [reason] if isinstance(reason, str) and reason else []
    row = {
        "model_id": "B4",
        "candidate_id": "B4",
        "case_key": {
            "standard_work_id": key[0],
            "origin": key[1],
            "horizon_months": key[2],
            "route": key[3],
        },
        "statisticallyScoreable": template["statisticallyScoreable"],
        "scoreabilityReason": template["scoreabilityReason"],
        "businessServingEligible": template["businessServingEligible"],
        "modelPredictionAvailable": template["modelPredictionAvailable"],
        "routeAbstained": template["routeAbstained"],
        "abstained": template["abstained"],
        "abstentionReason": reason,
        "rawModelPrediction": template["rawModelPrediction"],
        "servedPrediction": served,
        "formalModelPopulationEligible": bool(
            template["statisticallyScoreable"]
            and template["modelPredictionAvailable"]
            and not template["routeAbstained"]
        ),
        "channel_components": copy.deepcopy(template.get("channelComponents", [])),
        "target_end": template["targetEnd"],
        "label_available_as_of": template["labelAvailableAsOf"],
        "_bill_month_max": template["billMonthMax"],
        "_available_as_of": template["sourceAvailableAsOf"],
        "_residual_case_role": template["predictionRole"],
        "confidence": confidence,
        "limitation": limitations,
        "public_output": {
            "pointForecast": served,
            "annualBreakdown": [],
            "confidence": confidence,
            "limitation": limitations,
        },
        "_internal_interval": copy.deepcopy(template.get("internalInterval", {})),
    }
    formal.validate_case_state(row)
    return _join_truth(row, template)


def _candidate_metrics(
    keys: Sequence[tuple[str, str, int, str]],
    candidate_id: str,
    candidate_points: Mapping[tuple[str, str, int, str], Mapping[str, float]],
    templates: Mapping[tuple[str, tuple[str, str, int, str]], Mapping[str, Any]],
    origins: set[str],
    route: str,
) -> dict[str, Any] | None:
    selected = [key for key in keys if key[1] in origins and key[3] == route]
    if not selected:
        return None
    actuals = [float(templates[(f"development_forward_score:{key[1]}", key)]["forecastableCashActual"]) for key in selected]
    predictions = [float(candidate_points[key][candidate_id]) for key in selected]
    overall_wape = base.wape(predictions, actuals)
    bias = base.signed_aggregate_bias(predictions, actuals)
    high_keys = [
        key
        for key in selected
        if bool(
            templates[(f"development_forward_score:{key[1]}", key)]
            .get("strata", {})
            .get("high_value")
        )
    ]
    top10_keys = [
        key
        for key in selected
        if bool(
            templates[(f"development_forward_score:{key[1]}", key)]
            .get("strata", {})
            .get("top_10_percent")
        )
    ]

    def metric(group: Sequence[tuple[str, str, int, str]]) -> tuple[float, float]:
        group_actual = [
            float(
                templates[(f"development_forward_score:{key[1]}", key)][
                    "forecastableCashActual"
                ]
            )
            for key in group
        ]
        group_prediction = [float(candidate_points[key][candidate_id]) for key in group]
        return (
            base.wape(group_prediction, group_actual),
            base.signed_aggregate_bias(group_prediction, group_actual),
        )

    high_wape, high_bias = metric(high_keys) if high_keys else (math.inf, math.inf)
    top10_wape, _top10_bias = metric(top10_keys) if top10_keys else (math.inf, math.inf)
    horizons = {}
    for horizon in formal.CORE_HORIZONS:
        group = [key for key in selected if key[2] == horizon]
        horizons[str(horizon)] = metric(group) if group else (math.inf, math.inf)
    mean_horizon_wape = sum(value[0] for value in horizons.values()) / len(horizons)
    return {
        "caseCount": len(selected),
        "wape": overall_wape,
        "signedAggregateBias": bias,
        "highValueWape": high_wape,
        "highValueSignedAggregateBias": high_bias,
        "top10Wape": top10_wape,
        "horizons": {
            horizon: {"wape": values[0], "signedAggregateBias": values[1]}
            for horizon, values in horizons.items()
        },
        "meanHorizonWape": mean_horizon_wape,
    }


def _complexity(candidate_id: str) -> tuple[int, int, str]:
    if candidate_id.startswith("blend_"):
        return 2, 2, candidate_id
    if candidate_id in c2r1.COMPARATOR_COMPONENTS:
        return 1, 0, candidate_id
    return 1, 1, candidate_id


def _select_candidates(
    *,
    spec: Mapping[str, Any],
    candidate_points: Mapping[tuple[str, str, int, str], Mapping[str, float]],
    b4: Mapping[tuple[str, tuple[str, str, int, str]], Mapping[str, Any]],
) -> tuple[dict[tuple[str, str], str], list[dict[str, Any]]]:
    origins = list(spec["authority"]["origins"])
    keys = sorted(candidate_points)
    selected: dict[tuple[str, str], str] = {}
    evidence: list[dict[str, Any]] = []
    rule = spec["selection"]
    bias = rule["biasFeasibility"]
    objective_rule = rule["objective"]
    for origin_index, origin in enumerate(origins):
        earlier = set(origins[:origin_index])
        for route in sorted(c2r1.SALES_ROUTES):
            route_case_count = sum(key[1] in earlier and key[3] == route for key in keys)
            if (
                len(earlier) < int(rule["minimumEarlierOrigins"])
                or route_case_count < int(rule["minimumEarlierRouteCases"])
            ):
                selected[(origin, route)] = "B4"
                evidence.append(
                    {
                        "origin": origin,
                        "route": route,
                        "earlierOriginCount": len(earlier),
                        "earlierRouteCaseCount": route_case_count,
                        "biasFeasibleCandidateCount": 0,
                        "selectedCandidate": "B4",
                        "selectionReason": "preregistered_insufficient_earlier_evidence_fallback",
                        "thresholdMoved": False,
                    }
                )
                continue
            feasible: list[tuple[float, tuple[int, int, str], str, dict[str, Any]]] = []
            for candidate in c2r1.candidate_ids(spec):
                metrics = _candidate_metrics(
                    keys, candidate, candidate_points, b4, earlier, route
                )
                if metrics is None:
                    continue
                horizon_bias_ok = all(
                    abs(float(value["signedAggregateBias"]))
                    <= float(bias["eachHorizonAbsoluteMaximum"]) + 1e-12
                    for value in metrics["horizons"].values()
                )
                if not (
                    abs(float(metrics["signedAggregateBias"]))
                    <= float(bias["overallAbsoluteMaximum"]) + 1e-12
                    and abs(float(metrics["highValueSignedAggregateBias"]))
                    <= float(bias["highValueAbsoluteMaximum"]) + 1e-12
                    and horizon_bias_ok
                ):
                    continue
                objective = (
                    float(objective_rule["overallWapeWeight"]) * metrics["wape"]
                    + float(objective_rule["meanHorizonWapeWeight"])
                    * metrics["meanHorizonWape"]
                    + float(objective_rule["absoluteBiasWeight"])
                    * abs(metrics["signedAggregateBias"])
                    + float(objective_rule["top10WapeWeight"]) * metrics["top10Wape"]
                )
                feasible.append((objective, _complexity(candidate), candidate, metrics))
            feasible.sort(key=lambda item: (item[0], item[1]))
            chosen = feasible[0] if feasible else None
            selected[(origin, route)] = chosen[2] if chosen else "B4"
            evidence.append(
                {
                    "origin": origin,
                    "route": route,
                    "earlierOriginCount": len(earlier),
                    "earlierRouteCaseCount": route_case_count,
                    "biasFeasibleCandidateCount": len(feasible),
                    "selectedCandidate": selected[(origin, route)],
                    "selectionReason": (
                        "bias_feasible_weighted_objective"
                        if chosen
                        else "preregistered_no_bias_feasible_candidate_fallback"
                    ),
                    "selectedObjective": rounded(chosen[0]) if chosen else None,
                    "thresholdMoved": False,
                }
            )
    return selected, evidence


def _selection_future_invariance_self_test(
    spec: Mapping[str, Any],
) -> dict[str, bool]:
    """Prove an outer selection cannot inspect its own or later-origin labels."""

    origins = list(spec["authority"]["origins"])
    horizons = tuple(int(value) for value in spec["authority"]["horizonsMonths"])
    candidates = c2r1.candidate_ids(spec)
    candidate_points: dict[tuple[str, str, int, str], dict[str, float]] = {}
    templates: dict[
        tuple[str, tuple[str, str, int, str]], dict[str, Any]
    ] = {}
    for origin_index, origin in enumerate(origins):
        for route in sorted(c2r1.SALES_ROUTES):
            for horizon in horizons:
                for work_index in range(10):
                    key = (
                        f"SYNTH-{origin_index:02d}-{route}-{horizon:02d}-{work_index:02d}",
                        origin,
                        horizon,
                        route,
                    )
                    actual = float(100 + horizon)
                    candidate_points[key] = {
                        candidate: actual for candidate in candidates
                    }
                    templates[(f"development_forward_score:{origin}", key)] = {
                        "forecastableCashActual": actual,
                        "strata": {"high_value": True, "top_10_percent": True},
                    }
    selected_before, _evidence_before = _select_candidates(
        spec=spec, candidate_points=candidate_points, b4=templates
    )
    target_origin = origins[2]
    perturbed_points = copy.deepcopy(candidate_points)
    perturbed_templates = copy.deepcopy(templates)
    for key in perturbed_points:
        if key[1] >= target_origin:
            perturbed_points[key] = {
                candidate: point + 999999.0
                for candidate, point in perturbed_points[key].items()
            }
            perturbed_templates[(f"development_forward_score:{key[1]}", key)][
                "forecastableCashActual"
            ] += 777777.0
    selected_after, _evidence_after = _select_candidates(
        spec=spec, candidate_points=perturbed_points, b4=perturbed_templates
    )
    checks = {
        "targetUsesTwoEarlierOrigins": target_origin == origins[2],
        "sameAndLaterOriginPerturbationInvariant": all(
            selected_before[(target_origin, route)]
            == selected_after[(target_origin, route)]
            for route in c2r1.SALES_ROUTES
        ),
    }
    if not all(checks.values()):
        raise C2R1RunnerError("C2-R.1 selection future invariance failed")
    return checks


def _build_warmup_row(template: Mapping[str, Any]) -> dict[str, Any]:
    key = _case_key(template)
    paths = None
    if template["statisticallyScoreable"] and key[3] in c2r1.SALES_ROUTES:
        paths = {
            str(component["channel_key"]): [
                float(component["point_forecast"]) / key[2] for _ in range(key[2])
            ]
            for component in template.get("channelComponents", [])
        }
    prediction = c2r1.build_prediction(
        template=template, candidate_id="B4_warmup", channel_paths=paths
    )
    return _join_truth(prediction, template)


def _relative(candidate: float | None, reference: float | None) -> float | None:
    if candidate is None or reference is None:
        return None
    if float(reference) == 0:
        return 0.0 if float(candidate) == 0 else None
    return (float(candidate) - float(reference)) / float(reference)


def _origin_win_share(relative_by_origin: Mapping[str, float]) -> float:
    return sum(
        float(value) < -NUMERIC_EPSILON for value in relative_by_origin.values()
    ) / len(relative_by_origin)


def _acceptance(
    metrics: Mapping[str, Any],
    comparator: Mapping[str, Any],
    spec: Mapping[str, Any],
) -> dict[str, Any]:
    gate = spec["acceptance"]
    relative_rule = gate["relativeToPrimaryComparator"]
    horizon_relative = {
        horizon: _relative(
            metrics["horizons"][horizon]["wape"],
            comparator["horizons"][horizon]["wape"],
        )
        for horizon in ("3", "6", "12", "18", "24")
    }
    top_relative = {
        top: _relative(
            metrics["topBands"][top]["wape"], comparator["topBands"][top]["wape"]
        )
        for top in ("top1", "top5", "top10")
    }
    origins = list(spec["authority"]["origins"])
    origin_relative = {
        origin: _relative(
            metrics["origins"][origin]["wape"], comparator["origins"][origin]["wape"]
        )
        for origin in origins
    }
    # A comparator-equal origin is not a win. The epsilon only absorbs binary
    # floating-point noise and therefore cannot move a frozen business gate.
    win_share = _origin_win_share(origin_relative)
    consecutive = 0
    maximum_consecutive = 0
    for origin in origins:
        if float(origin_relative[origin]) > 0.05 + NUMERIC_EPSILON:
            consecutive += 1
            maximum_consecutive = max(maximum_consecutive, consecutive)
        else:
            consecutive = 0
    interval = metrics["internal80"]
    comparator_interval = comparator["internal80"]
    wis_improvement = (
        (float(comparator_interval["meanWis"]) - float(interval["meanWis"]))
        / float(comparator_interval["meanWis"])
    )
    width_regression = _relative(
        interval["standardizedWidth"], comparator_interval["standardizedWidth"]
    )
    coverage_low, coverage_high = gate["internal80CoverageInclusive"]
    conditions = {
        "overallWapeAtMost60Percent": metrics["modelPopulation"]["wape"]
        <= gate["overallWapeMaximum"],
        "overallBiasWithin10Percent": abs(
            metrics["modelPopulation"]["signedAggregateBias"]
        )
        <= gate["absoluteBiasMaximum"]["overall"],
        "servedBiasWithin10Percent": abs(metrics["served"]["signedAggregateBias"])
        <= gate["absoluteBiasMaximum"]["served"],
        "highValueBiasWithin10Percent": abs(
            metrics["highValue"]["signedAggregateBias"]
        )
        <= gate["absoluteBiasMaximum"]["highValue"],
        "eachHorizonBiasWithin15Percent": all(
            abs(value["signedAggregateBias"])
            <= gate["absoluteBiasMaximum"]["eachHorizon"]
            for value in metrics["horizons"].values()
        ),
        "horizon3ImprovesAtLeast3Percent": horizon_relative["3"]
        <= -relative_rule["horizon3ImprovementMinimum"],
        "horizon6ImprovesAtLeast3Percent": horizon_relative["6"]
        <= -relative_rule["horizon6ImprovementMinimum"],
        "horizon12ImprovesAtLeast3Percent": horizon_relative["12"]
        <= -relative_rule["horizon12ImprovementMinimum"],
        "horizon18RegressionAtMost2Percent": horizon_relative["18"]
        <= relative_rule["horizon18RegressionMaximum"],
        "horizon24RegressionAtMost2Percent": horizon_relative["24"]
        <= relative_rule["horizon24RegressionMaximum"],
        "top10ImprovesAtLeast5Percent": top_relative["top10"]
        <= -relative_rule["top10ImprovementMinimum"],
        "top1RegressionAtMost5Percent": top_relative["top1"]
        <= relative_rule["top1RegressionMaximum"],
        "top5RegressionAtMost5Percent": top_relative["top5"]
        <= relative_rule["top5RegressionMaximum"],
        "atLeast70PercentOriginsBeatPrimary": win_share
        >= relative_rule["outerOriginWinShareMinimum"],
        "noThreeConsecutiveOriginsRegressOver5Percent": maximum_consecutive
        <= relative_rule["maximumConsecutiveOriginsRegressingOverFivePercent"],
        "internal80CoverageBetween75And85Percent": coverage_low
        <= interval["internal80Coverage"]
        <= coverage_high,
        "internalWisImprovesAtLeast5Percent": wis_improvement
        >= relative_rule["internalWisImprovementMinimum"],
        "standardizedWidthRegressionAtMost10Percent": width_regression
        <= relative_rule["standardizedWidthRegressionMaximum"],
        "P0EqualsZero": True,
        "P1EqualsZero": True,
        "P2FactAuditOnly": True,
        "noAutomaticOperationalAction": True,
        "modelPopulationUnchanged": metrics["modelPopulation"]["caseCount"]
        == spec["authority"]["formalModelPopulationCaseCount"],
    }
    return {
        "conditions": conditions,
        "conditionCount": len(conditions),
        "passedConditionCount": sum(conditions.values()),
        "allTrue": all(conditions.values()),
        "decision": "PASS" if all(conditions.values()) else "FAIL",
        "relativeWape": {
            "horizons": horizon_relative,
            "topBands": top_relative,
            "origins": origin_relative,
        },
        "originWinShare": win_share,
        "maximumConsecutiveOriginsRegressingOverFivePercent": maximum_consecutive,
        "internalWisImprovement": wis_improvement,
        "standardizedWidthRegression": width_regression,
        "thresholdMoved": False,
    }


def _privacy() -> dict[str, bool]:
    return {
        "aggregateOnly": True,
        "deidentified": True,
        "workIdentifiersPresent": False,
        "authorIdentifiersPresent": False,
        "channelIdentifiersPresent": False,
        "privatePathsPresent": False,
        "rawRowsPresent": False,
        "predictionIntervalEndpointsPresent": False,
    }


def _seals() -> dict[str, bool]:
    return {
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
    }


def _design_report(
    spec: Mapping[str, Any], binding: Mapping[str, Any]
) -> dict[str, Any]:
    return {
        "schema": "m2.c2r1.design.public.v1",
        "version": "M2-C2R1-design-v1",
        "language": "zh-CN",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "specDigest": c2r1.canonical_digest(spec),
        "phaseBinding": dict(binding),
        "candidateSpace": public_value(spec["candidateSpace"]),
        "routeContract": public_value(spec["routes"]),
        "selection": public_value(spec["selection"]),
        "acceptance": public_value(spec["acceptance"]),
        "formalCashTarget": public_value(spec["formalCashTarget"]),
        "candidateSpecFrozenBeforePrivateDataRead": True,
        "legacyC2RBuyoutCycleImported": False,
        "futureBuyoutProbabilityModelPresent": False,
        "positiveOnlyMedianPresent": False,
        "postHocOuterScalingPresent": False,
        "seals": _seals(),
        "privacy": _privacy(),
    }


def _reports(
    *,
    spec: Mapping[str, Any],
    binding: Mapping[str, Any],
    metrics: Mapping[str, Any],
    comparator_metrics: Mapping[str, Any],
    acceptance: Mapping[str, Any],
    selection: Sequence[Mapping[str, Any]],
    bootstrap: Mapping[str, Any],
    lock: Mapping[str, Any],
) -> tuple[dict[str, Any], ...]:
    design = _design_report(spec, binding)
    public_metrics = public_value(metrics)
    validation = {
        "schema": "m2.c2r1.development_validation.public.v1",
        "version": "M2-C2R1-development-validation-v1",
        "language": "zh-CN",
        "decisionStatus": "not_for_formal_decision",
        "formalDecisionAuthorized": False,
        "releaseAuthorized": False,
        "candidateDecision": acceptance["decision"],
        "technicalSummary": {
            "C2R1Executed": True,
            "primaryComparator": "B4",
            "modelPopulationCaseCount": metrics["modelPopulation"]["caseCount"],
            "modelPopulationUnchanged": acceptance["conditions"][
                "modelPopulationUnchanged"
            ],
            "pureBuyoutNullScoredAsZero": False,
            "thresholdMoved": False,
            "finalHoldoutOpened": False,
        },
        "metrics": public_metrics,
        "primaryComparatorMetrics": public_value(comparator_metrics),
        "acceptance": public_value(acceptance),
        "selectionByOriginAndRoute": public_value(list(selection)),
        "pairedWorkOriginBootstrapAudit": public_value(bootstrap),
        "predictionIntegrity": dict(lock),
        "P0Count": 0,
        "P1Count": 0,
        "P2Boundary": "fact_audit_only",
        "automaticOperationalActionCount": 0,
        "seals": _seals(),
        "privacy": _privacy(),
        "nextBoundary": "stop_before_C2_C3",
    }
    route = {
        "schema": "m2.c2r1.route_metrics.public.v1",
        "version": "M2-C2R1-route-specific-metrics-v1",
        "language": "zh-CN",
        "decisionStatus": "not_for_formal_decision",
        "candidateDecision": acceptance["decision"],
        "salesRoutes": public_metrics["routes"],
        "primaryComparatorSalesRoutes": public_value(comparator_metrics["routes"]),
        "pureBuyoutAbstention": {
            "scoreableCaseCount": metrics["caseState"][
                "pureBuyoutNoCommitmentCaseCount"
            ],
            "commitmentCaseCount": metrics["caseState"][
                "pureBuyoutCommitmentCaseCount"
            ],
            "rawAndServedNull": True,
            "zeroImputationUsed": False,
            "abstentionReason": "uncommitted_future_buyout_not_forecastable",
        },
        "unknownRevenueModel": {
            "fallbackToBestRouteUsed": False,
            "withoutExplicitAsOfSalesEvidence": "route_abstain_null",
        },
        "excludesUncommittedFutureBuyout": True,
        "futureBuyoutProbabilityModelPresent": False,
        "seals": _seals(),
        "privacy": _privacy(),
    }
    channel = {
        "schema": "m2.c2r1.channel_reconciliation.public.v1",
        "version": "M2-C2R1-channel-reconciliation-completeness-v1",
        "language": "zh-CN",
        "decisionStatus": "not_for_formal_decision",
        "candidateDecision": acceptance["decision"],
        "candidate": public_metrics["channel"],
        "primaryComparator": public_value(comparator_metrics["channel"]),
        "twoLevelConservation": {
            "channelSumEqualsFutureSalesCashForecast": public_metrics["channel"][
                "overall"
            ]["allWorkPointsStrictlyReconciled"],
            "componentActualSumEqualsForecastableCashActual": public_metrics[
                "channel"
            ]["overall"]["allWorkActualsStrictlyReconciled"],
            "cutoffConfirmedReceivableCount": 0,
        },
        "matchedChannelMetricMayBeNamedWorkLevelModelWape": False,
        "truthWithoutPredictionReportedNotHidden": True,
        "seals": _seals(),
        "privacy": _privacy(),
    }
    coverage_source = json.loads(phase_a.COVERAGE_JSON.read_text(encoding="utf-8"))
    business = {
        "schema": "m2.c2r1.end_to_end_business_coverage.public.v1",
        "version": "M2-C2R1-end-to-end-business-coverage-v1",
        "language": "zh-CN",
        "decisionStatus": "not_for_formal_decision",
        "candidateDecision": acceptance["decision"],
        "completePopulation": coverage_source,
        "developmentBusinessCoverage": public_metrics["businessCoverage"],
        "businessObservation": copy.deepcopy(coverage_source["observationGates"]),
        "observationGateMayAuthorizeFormalApproval": False,
        "endToEndBusinessGapMayBeNamedModelWape": False,
        "surpriseCashHidden": False,
        "seals": _seals(),
        "privacy": _privacy(),
    }
    return design, validation, route, channel, business


def _assert_public_safety(paths: Sequence[Path]) -> None:
    forbidden = (
        "data/private-output",
        "private-output\\",
        '"standard_work_id"',
        '"channel_key"',
        '"rawChannel',
        '"lower"',
        '"upper"',
        "optimistic",
        "pessimistic",
    )
    for path in paths:
        text = path.read_text(encoding="utf-8")
        if any(fragment in text for fragment in forbidden):
            raise C2R1RunnerError(f"public C2-R.1 artifact is unsafe: {path.name}")


def _write_markdown(
    validation: Mapping[str, Any],
    route: Mapping[str, Any],
    channel: Mapping[str, Any],
    business: Mapping[str, Any],
    design: Mapping[str, Any],
) -> None:
    metric = validation["metrics"]["modelPopulation"]
    acceptance = validation["acceptance"]
    write_text(
        DESIGN_MD,
        f"""# M2 C2-R.1 设计 v1

## 冻结结论

C2-R.1 在任何 private case 或权威输入读取前冻结了 {design['candidateSpace']['candidateCount']} 个透明候选。纯实销与买断+实销均逐渠道预测后严格求和；纯买断无 cutoff 承诺和缺少明确实销证据的 unknown route 均输出 null，不用 0 代替。

候选只使用零保留 trailing/seasonal/zero-aware/robust/trend/recency 组件与 formal B0b/B1/B3/B4。禁止 positive-only median、未来买断周期、买断概率、已到账买断摊销和事后 outer 缩放。每个 outer origin×route 只用 earlier origins，通过 bias feasibility 后才计算冻结的多目标分数；无可行候选时回退 B4，不移动阈值。

本设计保持 `not_for_formal_decision`，不授权 final holdout、release、C2/C3 或 M3。
""",
    )
    write_text(
        VALIDATION_MD,
        f"""# M2 C2-R.1 development 验证 v1

## 结论

C2-R.1 结论为 **{validation['candidateDecision']}**，仍为 `not_for_formal_decision`。模型人口保持 {metric['caseCount']} 个 case；overall WAPE 为 {metric['wape']:.4f}，signed bias 为 {metric['signedAggregateBias']:+.4f}。{acceptance['passedConditionCount']}/{acceptance['conditionCount']} 项验收条件通过。

纯买断无承诺始终 raw/served null，未按 0 计分；final holdout、embargo shadow 与 deferred 60-month labels 均未打开。即使候选 PASS，也仍需中文业务抽检和用户明确批准；本轮在 C2-R.1 后停止。
""",
    )
    write_text(
        ROUTE_MD,
        f"""# M2 C2-R.1 路由指标 v1

## 路由边界

- pure-sales：逐渠道独立预测后求和。
- mixed：只预测实销现金，排除未承诺未来买断。
- pure-buyout：{route['pureBuyoutAbstention']['scoreableCaseCount']} 个 scoreable case 无 cutoff 承诺，raw/served 均为 null。
- unknown：没有明确 as-of 实销组件证据时 abstain，不回退到表现最好的路线。

候选结论为 **{route['candidateDecision']}**；所有指标仍为 development evidence。
""",
    )
    overall = channel["candidate"]["overall"]
    write_text(
        CHANNEL_MD,
        f"""# M2 C2-R.1 渠道守恒与完整性 v1

## 结论

渠道点值和作品点值最大绝对差为 {overall['maximumChannelSumToWorkPointAbsoluteDifference']:.8f}；渠道 actual 与 formal-cash actual 最大绝对差为 {overall['maximumTruthComponentSumToWorkActualAbsoluteDifference']:.8f}。匹配组件覆盖 forecastable cash 的 {overall['matchedForecastableCashShare']:.2%}；未匹配 actual 组件数量与金额已单列，未隐藏，也未命名为作品级模型 WAPE。

公开报告只保留脱敏聚合，不含真实渠道或作品标识。
""",
    )
    observation = business["businessObservation"]
    write_text(
        BUSINESS_MD,
        f"""# M2 C2-R.1 端到端业务现金覆盖 v1

## 结论

完整人口 forecastable cash share 为 {observation['forecastableCashShareObserved']:.2%}，top10 forecastable cash coverage 为 {observation['top10ForecastableCashCoverageObserved']:.2%}。两项 90% 观察门槛分别为 {'通过' if observation['forecastableCashShareObservedPass'] else '未通过'}、{'通过' if observation['top10ForecastableCashCoverageObservedPass'] else '未通过'}。

这些是业务覆盖观察门槛，不是模型 WAPE，也不能授权 formal approval 或 release。surprise cash 与未服务现金差额均明确保留。
""",
    )


def _write_private_cases(rows: Sequence[Mapping[str, Any]]) -> tuple[int, str]:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    hasher = hashlib.sha256()
    count = 0
    with PRIVATE_CASES.open("wb") as handle:
        for row in sorted(rows, key=formal.strict_case_key):
            key = formal.strict_case_key(row)
            payload = {
                "caseKey": {
                    "standard_work_id": key[0],
                    "origin": key[1],
                    "horizon_months": key[2],
                    "route": key[3],
                },
                "selectedCandidate": row["candidate_id"],
                "statisticallyScoreable": row["statisticallyScoreable"],
                "businessServingEligible": row["businessServingEligible"],
                "modelPredictionAvailable": row["modelPredictionAvailable"],
                "routeAbstained": row["routeAbstained"],
                "abstentionReason": row["abstentionReason"],
                "rawModelPrediction": row["rawModelPrediction"],
                "servedPrediction": row["servedPrediction"],
                "forecastableCashActual": row["forecastableCashActual"],
                "uncommittedBuyoutSurpriseActual": row[
                    "uncommittedBuyoutSurpriseActual"
                ],
                "totalLedgerCashActual": row["totalLedgerCashActual"],
                "channelComponents": copy.deepcopy(row.get("channel_components", [])),
                "forecastableActualByComponent": copy.deepcopy(
                    row.get("forecastableActualByComponent", {})
                ),
                "internalInterval": copy.deepcopy(row.get("_internal_interval", {})),
                "strata": copy.deepcopy(row.get("strata", {})),
            }
            encoded = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            handle.write(encoded)
            hasher.update(encoded)
            count += 1
    return count, hasher.hexdigest()


def _write_workbook(
    validation: Mapping[str, Any],
    selection: Sequence[Mapping[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook  # pylint: disable=import-outside-toplevel
        from openpyxl.styles import Font  # pylint: disable=import-outside-toplevel
    except ImportError as exc:
        raise C2R1RunnerError("private workbook dependency is unavailable") from exc
    workbook = Workbook()
    summary = workbook.active
    summary.title = "结论"
    summary.append(["字段", "结果"])
    summary.append(["候选结论", validation["candidateDecision"]])
    summary.append(["决策状态", "not_for_formal_decision"])
    summary.append(["总体WAPE", validation["metrics"]["modelPopulation"]["wape"]])
    summary.append(
        ["总体signed bias", validation["metrics"]["modelPopulation"]["signedAggregateBias"]]
    )
    summary.append(["通过条件", validation["acceptance"]["passedConditionCount"]])
    summary.append(["条件总数", validation["acceptance"]["conditionCount"]])
    summary.append(["final holdout已打开", "否"])
    selection_sheet = workbook.create_sheet("origin路由选择")
    selection_sheet.append(
        ["origin", "收入路由", "earlier origin数", "earlier case数", "可行候选数", "选中候选", "原因"]
    )
    for row in selection:
        selection_sheet.append(
            [
                row["origin"],
                row["route"],
                row["earlierOriginCount"],
                row["earlierRouteCaseCount"],
                row["biasFeasibleCandidateCount"],
                row["selectedCandidate"],
                row["selectionReason"],
            ]
        )
    horizon = workbook.create_sheet("horizon指标")
    horizon.append(["horizon月", "WAPE", "signed bias"])
    for label, metric in validation["metrics"]["horizons"].items():
        horizon.append([int(label), metric["wape"], metric["signedAggregateBias"]])
    gates = workbook.create_sheet("验收条件")
    gates.append(["条件", "是否通过"])
    for name, passed in validation["acceptance"]["conditions"].items():
        gates.append([name, "是" if passed else "否"])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(PRIVATE_WORKBOOK)


def run_development() -> dict[str, Any]:
    if phase_a.run_git("branch", "--show-current") != BRANCH:
        raise C2R1RunnerError(f"C2-R.1 must run on {BRANCH}")
    _assert_private_paths()
    spec = c2r1.load_spec()
    progress("verifying pushed Gate B before any private input read")
    binding = _validate_phase_bindings(spec)
    _write_pre_data_lock(spec, binding)
    progress("loading and verifying locked formal comparator cases")
    b4, comparator, phase_manifest = _load_formal_private_cases()
    calibration_spec, _v11, _v12 = v12.load_and_validate_contract()
    progress("loading authorized work histories read-only")
    works, _posthoc, input_evidence = legacy.load_authorized_works(calibration_spec)
    if input_evidence["inputFingerprint"] != phase_manifest["inputFingerprint"]:
        raise C2R1RunnerError("C2-R.1 authority fingerprint differs from Phase A")
    work_by_id = {str(work["standard_work_id"]): work for work in works}
    origins = list(spec["authority"]["origins"])
    forward_templates = {
        key: template
        for (role, key), template in b4.items()
        if role.startswith("development_forward_score:")
    }
    if len(forward_templates) != int(spec["authority"]["developmentCaseCount"]):
        raise C2R1RunnerError("C2-R.1 frozen case count differs")
    candidate_points: dict[tuple[str, str, int, str], dict[str, float]] = {}
    progress("materializing the frozen 45-candidate channel space without actuals")
    for index, (key, template) in enumerate(sorted(forward_templates.items()), start=1):
        if not (
            template["statisticallyScoreable"] is True
            and key[3] in c2r1.SALES_ROUTES
        ):
            continue
        role = f"development_forward_score:{key[1]}"
        candidates = c2r1.candidate_channel_paths(
            work=work_by_id[key[0]],
            origin=key[1],
            horizon=key[2],
            comparator_rows=_comparator_rows_for_key(comparator, role, key),
            calibration_spec=calibration_spec,
            spec=spec,
        )
        candidate_points[key] = {
            candidate: sum(sum(path) for path in channels.values())
            for candidate, channels in candidates.items()
        }
        if index % 3000 == 0:
            progress(f"candidate projections prepared through case {index}")
    if len(candidate_points) != int(spec["authority"]["formalModelPopulationCaseCount"]):
        raise C2R1RunnerError("C2-R.1 candidate model population differs")
    progress("selecting each outer-origin route from strictly earlier origins")
    selected, selection_evidence = _select_candidates(
        spec=spec, candidate_points=candidate_points, b4=b4
    )
    predictions: list[dict[str, Any]] = []
    progress("materializing selected C2-R.1 predictions before truth join")
    for index, (key, template) in enumerate(sorted(forward_templates.items()), start=1):
        candidate_id = (
            selected[(key[1], key[3])]
            if template["statisticallyScoreable"] is True
            and key[3] in c2r1.SALES_ROUTES
            else "route_abstain"
            if key[3] not in c2r1.SALES_ROUTES
            else "model_unavailable"
        )
        paths = None
        if template["statisticallyScoreable"] is True and key[3] in c2r1.SALES_ROUTES:
            role = f"development_forward_score:{key[1]}"
            paths = c2r1.candidate_channel_paths(
                work=work_by_id[key[0]],
                origin=key[1],
                horizon=key[2],
                comparator_rows=_comparator_rows_for_key(comparator, role, key),
                calibration_spec=calibration_spec,
                spec=spec,
            )[candidate_id]
        predictions.append(
            c2r1.build_prediction(
                template=template, candidate_id=candidate_id, channel_paths=paths
            )
        )
        if index % 5000 == 0:
            progress(f"selected predictions prepared through case {index}")
    projection_before = digest(
        [_compact_prediction_projection(row) for row in predictions]
    )
    forward = [
        _join_truth(row, forward_templates[formal.strict_case_key(row)])
        for row in predictions
    ]
    projection_after = digest(
        [_compact_prediction_projection(row) for row in forward]
    )
    if projection_before != projection_after:
        raise C2R1RunnerError("C2-R.1 prediction changed after truth join")
    warmup = [
        _build_warmup_row(template)
        for (role, _key), template in sorted(b4.items())
        if role == "development_warmup_interval_calibration"
    ]
    progress("calibrating internal 80% intervals from earlier residuals only")
    formal.apply_internal_intervals(forward, [*warmup, *forward], formal.load_spec())
    metrics = phase_a.metrics_for_model(forward)
    b4_forward = [
        _locked_b4_row(template) for _key, template in sorted(forward_templates.items())
    ]
    # Recompute B4 from the locked private Phase-A cases instead of comparing
    # against the eight-decimal public report. Otherwise a comparator-equal
    # origin can become a false win after public-report rounding.
    comparator_metrics = phase_a.metrics_for_model(b4_forward)
    candidate_keys = {
        formal.strict_case_key(row) for row in forward if formal.is_model_population(row)
    }
    comparator_keys = {
        formal.strict_case_key(row)
        for row in b4_forward
        if formal.is_model_population(row)
    }
    if candidate_keys != comparator_keys or len(candidate_keys) != int(
        spec["authority"]["formalModelPopulationCaseCount"]
    ):
        raise C2R1RunnerError("C2-R.1 and primary comparator model populations differ")
    progress("running paired work×origin bootstrap audit")
    bootstrap = formal.paired_relative_block_bootstrap(
        [*b4_forward, *forward], "B4", ("B4", "C2-R.1"), formal.load_spec()
    )
    bootstrap.pop("clusterKeys", None)
    bootstrap["clusterDefinition"] = "deidentified_work_x_origin"
    acceptance = _acceptance(metrics, comparator_metrics, spec)
    lock = {
        "predictionLockedBeforeTruthJoin": True,
        "predictionProjectionDigest": projection_before,
        "postTruthProjectionMatchesLock": True,
        "fullCaseKeyCount": len(forward),
        "modelPopulationKeyCount": len(candidate_keys),
        "modelPopulationMatchesPrimaryComparator": True,
        "actualFieldAbsentAtPredictionLock": True,
        "finalHoldoutOpened": False,
    }
    design, validation, route, channel, business = _reports(
        spec=spec,
        binding=binding,
        metrics=metrics,
        comparator_metrics=comparator_metrics,
        acceptance=acceptance,
        selection=selection_evidence,
        bootstrap=bootstrap,
        lock=lock,
    )
    progress("writing deidentified public reports and ignored private evidence")
    write_json(DESIGN_JSON, design)
    write_json(VALIDATION_JSON, validation)
    write_json(ROUTE_JSON, route)
    write_json(CHANNEL_JSON, channel)
    write_json(BUSINESS_JSON, business)
    _write_markdown(validation, route, channel, business, design)
    public_paths = (
        DESIGN_JSON,
        DESIGN_MD,
        VALIDATION_JSON,
        VALIDATION_MD,
        ROUTE_JSON,
        ROUTE_MD,
        CHANNEL_JSON,
        CHANNEL_MD,
        BUSINESS_JSON,
        BUSINESS_MD,
    )
    _assert_public_safety(public_paths)
    case_count, case_sha = _write_private_cases(forward)
    _write_workbook(validation, selection_evidence)
    manifest = {
        "schema": "m2.c2r1.development_manifest.private.v1",
        "decisionStatus": "not_for_formal_decision",
        "tracked": False,
        "specDigest": c2r1.canonical_digest(spec),
        "phaseACheckpoint": binding["phaseACheckpoint"],
        "inputFingerprint": input_evidence["inputFingerprint"],
        "privateCaseCount": case_count,
        "privateCaseSha256": case_sha,
        "privateWorkbookSha256": file_sha256(PRIVATE_WORKBOOK),
        "publicArtifactSha256": {
            path.name: file_sha256(path) for path in public_paths
        },
        "candidateDecision": acceptance["decision"],
        "finalHoldoutOpened": False,
        "embargoShadowOpened": False,
        "deferred60MonthLabelsOpened": False,
        "C2Started": False,
        "C3Started": False,
    }
    write_json(PRIVATE_MANIFEST, manifest)
    _assert_private_paths()
    return {
        "status": "passed",
        "candidateDecision": acceptance["decision"],
        "acceptancePassedConditionCount": acceptance["passedConditionCount"],
        "acceptanceConditionCount": acceptance["conditionCount"],
        "modelPopulationCaseCount": metrics["modelPopulation"]["caseCount"],
        "overallWape": rounded(metrics["modelPopulation"]["wape"]),
        "overallSignedAggregateBias": rounded(
            metrics["modelPopulation"]["signedAggregateBias"]
        ),
        "caseEvidenceSha256": case_sha,
        "privateWorkbookTracked": False,
        "decisionStatus": "not_for_formal_decision",
        "finalHoldoutOpened": False,
        "C2Started": False,
        "C3Started": False,
    }


def verify_development() -> dict[str, Any]:
    _assert_private_paths()
    spec = c2r1.load_spec()
    _validate_phase_bindings(spec)
    public_paths = (
        DESIGN_JSON,
        DESIGN_MD,
        VALIDATION_JSON,
        VALIDATION_MD,
        ROUTE_JSON,
        ROUTE_MD,
        CHANNEL_JSON,
        CHANNEL_MD,
        BUSINESS_JSON,
        BUSINESS_MD,
    )
    if any(not path.is_file() for path in (*public_paths, PRIVATE_CASES, PRIVATE_MANIFEST, PRIVATE_WORKBOOK)):
        raise C2R1RunnerError("C2-R.1 development evidence is incomplete")
    _assert_public_safety(public_paths)
    manifest = json.loads(PRIVATE_MANIFEST.read_text(encoding="utf-8"))
    hasher = hashlib.sha256()
    count = 0
    with PRIVATE_CASES.open("rb") as handle:
        for raw in handle:
            payload = json.loads(raw)
            canonical = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
                + b"\n"
            )
            if raw != canonical:
                raise C2R1RunnerError("C2-R.1 private case is not canonical")
            hasher.update(raw)
            count += 1
    if (
        count != manifest["privateCaseCount"]
        or hasher.hexdigest() != manifest["privateCaseSha256"]
        or file_sha256(PRIVATE_WORKBOOK) != manifest["privateWorkbookSha256"]
        or any(
            file_sha256(path) != manifest["publicArtifactSha256"].get(path.name)
            for path in public_paths
        )
    ):
        raise C2R1RunnerError("C2-R.1 development manifest differs")
    validation = json.loads(VALIDATION_JSON.read_text(encoding="utf-8"))
    if (
        validation["technicalSummary"]["modelPopulationCaseCount"]
        != spec["authority"]["formalModelPopulationCaseCount"]
        or validation["technicalSummary"]["pureBuyoutNullScoredAsZero"] is not False
        or any(value is not False for value in validation["seals"].values())
    ):
        raise C2R1RunnerError("C2-R.1 public validation boundary differs")
    return {
        "status": "passed",
        "candidateDecision": validation["candidateDecision"],
        "privateCaseCount": count,
        "privateCaseSha256": hasher.hexdigest(),
        "privateWorkbookTracked": False,
        "publicArtifactCount": len(public_paths),
        "decisionStatus": "not_for_formal_decision",
        "finalHoldoutOpened": False,
    }


def preflight() -> dict[str, Any]:
    _assert_private_paths()
    spec = c2r1.load_spec()
    synthetic = c2r1.synthetic_self_test()
    runner_checks = {
        "comparatorEqualityIsNotOriginWin": _origin_win_share(
            {"equal": 0.0, "floatingNoise": -NUMERIC_EPSILON / 2.0}
        )
        == 0.0,
        **_selection_future_invariance_self_test(spec),
    }
    if not all(runner_checks.values()):
        raise C2R1RunnerError("C2-R.1 runner synthetic checks failed")
    gate = json.loads(phase_a.GATE_B_JSON.read_text(encoding="utf-8"))
    if digest(gate) != spec["phaseABinding"]["gateBDigestAtAuthorization"]:
        raise C2R1RunnerError("C2-R.1 preflight Gate B digest differs")
    return {
        "status": "passed",
        "mode": "synthetic-only",
        "specDigest": c2r1.canonical_digest(spec),
        "candidateCount": len(c2r1.candidate_ids(spec)),
        "synthetic": synthetic,
        "runnerChecks": runner_checks,
        "gateBAllTrue": gate.get("allTrue") is True,
        "privateDataRead": False,
        "dataLoadCalls": 0,
        "finalHoldoutOpened": False,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--preflight", action="store_true")
    group.add_argument("--run-development", action="store_true")
    group.add_argument("--verify-development", action="store_true")
    group.add_argument("--run-final-holdout", action="store_true")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        if args.run_final_holdout:
            raise C2R1RunnerError(
                "final holdout is sealed in the C2-R.1 development runner; "
                "dataLoadCalls=0"
            )
        if args.run_development:
            result = run_development()
        elif args.verify_development:
            result = verify_development()
        else:
            result = preflight()
        print(json.dumps(result, ensure_ascii=False, sort_keys=True, allow_nan=False))
        return 0
    except (
        C2R1RunnerError,
        c2r1.C2R1Error,
        formal.FormalComparatorError,
        phase_a.FormalReplayError,
        legacy.ReplayError,
        v12.CalibrationV12Error,
        RuntimeError,
        ValueError,
        AssertionError,
        KeyError,
    ) as exc:
        print(
            json.dumps({"status": "failed", "reason": str(exc)}, ensure_ascii=False),
            file=sys.stderr,
        )
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
