from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing spreadsheet dependency. Install openpyxl in a temporary dependency "
        "directory and expose it through PYTHONPATH before rerunning."
    ) from exc


ROOT = Path(__file__).resolve().parents[2]
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
TAXONOMY_PATH = (
    ROOT
    / "src"
    / "domain"
    / "oldProductEvaluation"
    / "classificationTaxonomy.v1.json"
)
FILLED_WORKBOOK_PATH = (
    PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.xlsx"
)
BASELINE_PATH = (
    PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.json"
)
FIXED_WORKBOOK_PATH = (
    PRIVATE_DIR / "M2-classification-tag-foundation-local-fixed-cn-v1.xlsx"
)
FIXED_JSON_PATH = (
    PRIVATE_DIR / "M2-classification-tag-foundation-local-fixed-cn-v1.json"
)
PUBLIC_REPORT_MD_PATH = (
    DOCS_DIR / "M2-classification-tag-foundation-local-closeout-v1.md"
)
PUBLIC_REPORT_JSON_PATH = (
    DOCS_DIR / "M2-classification-tag-foundation-local-closeout-v1.json"
)

MAIN_SHEETS = ["01_出版物", "02_网文"]
MAIN_HEADERS = [
    "序号",
    "书名",
    "作品编号",
    "作者",
    "一级分类（最终采用）",
    "二级分类（最终采用）",
    "三级分类（最终采用）",
    "辅助标签（最终采用）",
    "备注",
]
USER_ADDED_LEVEL3 = ["科普", "教辅", "诗歌"]
USER_ADDED_TAGS = [
    "女性主义",
    "戛纳电影节",
    "冰岛",
    "动画",
    "柏林电影节",
    "奥地利",
    "广播剧",
    "晋江",
    "银河奖",
    "南非",
    "希腊",
]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_work_id(value) -> str:
    text = clean(value)
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def split_tags(value) -> list[str]:
    text = clean(value)
    if not text or text == "无":
        return []
    parts = [
        clean(item)
        for item in re.split(r"[；;、，,|/]+", text)
        if clean(item)
    ]
    if "无" in parts and len(parts) > 1:
        raise ValueError("The no-tag marker cannot be combined with other tags.")
    result = []
    for item in parts:
        if item != "无" and item not in result:
            result.append(item)
    return result


def relative(path: Path) -> str:
    return str(path.relative_to(ROOT)).replace("\\", "/")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def work_sort_key(work_id: str) -> tuple[int, int | str]:
    return (0, int(work_id)) if work_id.isdigit() else (1, work_id)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply the user-confirmed M2 classification/tag foundation workbook."
    )
    parser.add_argument(
        "--confirm-user-message",
        action="store_true",
        help="Use the current user message as the explicit whole-workbook confirmation.",
    )
    parser.add_argument(
        "--repair-out-of-scope-author-edits",
        action="store_true",
        help="Restore accidental author display edits from the private baseline before apply.",
    )
    return parser.parse_args()


def load_taxonomy() -> dict:
    return json.loads(TAXONOMY_PATH.read_text(encoding="utf-8"))


def load_baseline() -> tuple[dict[str, dict], dict]:
    payload = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
    records = {}
    for row in payload.get("records", []):
        work_id = normalize_work_id(row.get("作品编号"))
        if not work_id or work_id in records:
            raise SystemExit("Private baseline has a missing or duplicate work ID.")
        records[work_id] = row
    return records, payload.get("summary", {})


def load_filled_workbook() -> tuple[dict[str, dict], str]:
    workbook = load_workbook(FILLED_WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        rows = {}
        for sheet_name in MAIN_SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise SystemExit(f"Filled workbook is missing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            header_values = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            headers = [clean(value) for value in header_values]
            if headers != MAIN_HEADERS:
                raise SystemExit(f"Filled workbook headers changed in sheet: {sheet_name}")
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = {
                    header: clean(values[index] if index < len(values) else "")
                    for index, header in enumerate(headers)
                }
                work_id = normalize_work_id(row.get("作品编号"))
                if not work_id:
                    continue
                if work_id in rows:
                    raise SystemExit("Filled workbook contains duplicate work IDs.")
                row["作品编号"] = work_id
                row["原工作表"] = sheet_name
                rows[work_id] = row

        if "03_整体确认" not in workbook.sheetnames:
            raise SystemExit("Filled workbook is missing the whole-workbook confirmation sheet.")
        confirmation = clean(workbook["03_整体确认"]["C2"].value)
        return rows, confirmation
    finally:
        workbook.close()


def repair_out_of_scope_author_edits(baseline: dict[str, dict]) -> int:
    workbook = load_workbook(FILLED_WORKBOOK_PATH)
    repaired = 0
    try:
        for sheet_name in MAIN_SHEETS:
            sheet = workbook[sheet_name]
            headers = [clean(cell.value) for cell in sheet[1]]
            work_id_column = headers.index("作品编号") + 1
            author_column = headers.index("作者") + 1
            for row_index in range(2, sheet.max_row + 1):
                work_id = normalize_work_id(
                    sheet.cell(row=row_index, column=work_id_column).value
                )
                if not work_id or work_id not in baseline:
                    continue
                expected_author = clean(baseline[work_id].get("作者"))
                author_cell = sheet.cell(row=row_index, column=author_column)
                if clean(author_cell.value) != expected_author:
                    author_cell.value = expected_author
                    repaired += 1
        if repaired:
            workbook.save(FILLED_WORKBOOK_PATH)
        return repaired
    finally:
        workbook.close()


def validate_and_build(
    filled: dict[str, dict],
    baseline: dict[str, dict],
    taxonomy: dict,
    confirmation: str,
    confirm_user_message: bool,
    repaired_author_edits: int,
) -> tuple[list[dict], dict]:
    if confirmation != "确认作为后续基础表格" and not confirm_user_message:
        raise SystemExit(
            "Whole-workbook confirmation is missing. Fill 03_整体确认 or rerun with "
            "--confirm-user-message after an explicit user confirmation."
        )

    if set(filled) != set(baseline):
        raise SystemExit(
            "Filled workbook scope differs from the private baseline: "
            f"missing={len(set(baseline) - set(filled))}, "
            f"extra={len(set(filled) - set(baseline))}."
        )

    allowed_tags = {
        tag
        for tags in taxonomy["auxiliaryTagGroups"].values()
        for tag in tags
    }
    change_counts = Counter()
    changed_works = set()
    tag_additions = Counter()
    tag_removals = Counter()
    user_added_level3_counts = Counter()
    user_added_tag_counts = Counter()
    out_of_scope_author_edits = 0
    rows = []

    for work_id in sorted(filled, key=work_sort_key):
        incoming = filled[work_id]
        before = baseline[work_id]
        if incoming["书名"] != clean(before.get("书名")):
            raise SystemExit("A work title changed in the classification/tag workbook.")
        if incoming["作者"] != clean(before.get("作者")):
            out_of_scope_author_edits += 1

        level1 = incoming["一级分类（最终采用）"]
        level2 = incoming["二级分类（最终采用）"]
        level3 = incoming["三级分类（最终采用）"]
        if not all([level1, level2, level3]):
            raise SystemExit("Filled workbook contains an incomplete classification path.")
        branches = taxonomy["classificationTree"].get(level1, {})
        if level2 not in branches or level3 not in branches[level2]:
            raise SystemExit("Filled workbook contains a path outside the confirmed taxonomy.")

        try:
            tags = split_tags(incoming["辅助标签（最终采用）"])
        except ValueError as exc:
            raise SystemExit(str(exc)) from exc
        unknown_tags = [tag for tag in tags if tag not in allowed_tags]
        if unknown_tags:
            raise SystemExit(
                "Filled workbook contains tags outside the confirmed controlled list: "
                + "、".join(sorted(set(unknown_tags)))
            )

        before_tags = split_tags(before.get("辅助标签"))
        field_pairs = [
            ("一级分类", level1, clean(before.get("一级分类"))),
            ("二级分类", level2, clean(before.get("二级分类"))),
            ("三级分类", level3, clean(before.get("三级分类"))),
            ("辅助标签", set(tags), set(before_tags)),
        ]
        for field, after_value, before_value in field_pairs:
            if after_value != before_value:
                change_counts[field] += 1
                changed_works.add(work_id)

        tag_additions.update(set(tags) - set(before_tags))
        tag_removals.update(set(before_tags) - set(tags))
        if level3 in USER_ADDED_LEVEL3:
            user_added_level3_counts[level3] += 1
        user_added_tag_counts.update(tag for tag in tags if tag in USER_ADDED_TAGS)

        rows.append(
            {
                "作品编号": work_id,
                "书名": clean(before.get("书名")),
                "作者": clean(before.get("作者")),
                "一级分类": level1,
                "二级分类": level2,
                "三级分类": level3,
                "辅助标签": "；".join(tags) if tags else "无",
                "备注": incoming.get("备注", ""),
                "固定来源": "用户最终分类标签基础大表",
            }
        )

    tag_lists = [split_tags(row["辅助标签"]) for row in rows]
    level1_distribution = Counter(row["一级分类"] for row in rows)
    level2_distribution = Counter(
        f"{row['一级分类']}>{row['二级分类']}" for row in rows
    )
    level3_distribution = Counter(
        f"{row['一级分类']}>{row['二级分类']}>{row['三级分类']}"
        for row in rows
    )
    tag_distribution = Counter(tag for tags in tag_lists for tag in tags)
    confirmation_source = (
        "工作表整体确认"
        if confirmation == "确认作为后续基础表格"
        else "用户会话明确确认"
    )

    summary = {
        "schema": "m2.classification_tag_foundation_local_closeout.v1",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "candidateVersion": "m2-local-classification-tag-foundation-v1",
        "taxonomyVersion": taxonomy["version"],
        "confirmation": {
            "confirmed": True,
            "source": confirmation_source,
            "worksheetValue": confirmation,
        },
        "scope": {
            "workCount": len(rows),
            "uniqueWorkIds": len({row["作品编号"] for row in rows}),
            "level1Distribution": dict(level1_distribution),
        },
        "changesFromGeneratedBaseline": {
            "changedWorks": len(changed_works),
            "fieldChangeCounts": dict(change_counts),
            "tagAdditionAssignments": sum(tag_additions.values()),
            "tagRemovalAssignments": sum(tag_removals.values()),
        },
        "userAddedControlledValues": {
            "level3Counts": dict(user_added_level3_counts),
            "tagCounts": dict(user_added_tag_counts),
        },
        "finalDistribution": {
            "level2": dict(level2_distribution),
            "level3": dict(level3_distribution),
            "tags": dict(tag_distribution),
            "taggedWorks": sum(bool(tags) for tags in tag_lists),
            "tagAssignments": sum(len(tags) for tags in tag_lists),
        },
        "quality": {
            "missingWorkIds": 0,
            "duplicateWorkIds": 0,
            "missingRequiredFields": 0,
            "invalidClassificationPaths": 0,
            "invalidAuxiliaryTags": 0,
            "outOfScopeAuthorDisplayEdits": out_of_scope_author_edits,
            "outOfScopeAuthorEditsApplied": 0,
            "outOfScopeAuthorEditsConfirmedAccidentalAndRepaired": repaired_author_edits,
        },
        "closure": {
            "classificationAndTagsLocallyClosed": True,
            "manualFoundationFieldsLocallyClosed": True,
            "formalMasterDataWritten": False,
            "databaseWritten": False,
            "m2FormalComplete": False,
            "m3FormalExecutionStarted": False,
        },
        "inputs": {
            "filledWorkbook": relative(FILLED_WORKBOOK_PATH),
            "filledWorkbookSha256": sha256(FILLED_WORKBOOK_PATH),
            "privateBaseline": relative(BASELINE_PATH),
            "privateBaselineSha256": sha256(BASELINE_PATH),
        },
        "outputs": {
            "privateFixedWorkbook": relative(FIXED_WORKBOOK_PATH),
            "privateFixedJson": relative(FIXED_JSON_PATH),
            "publicReportMarkdown": relative(PUBLIC_REPORT_MD_PATH),
            "publicReportJson": relative(PUBLIC_REPORT_JSON_PATH),
        },
    }
    return rows, summary


def apply_style(sheet, final_columns: set[int] | None = None) -> None:
    final_columns = final_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    final_fill = PatternFill("solid", fgColor="E2F0D9")
    fixed_fill = PatternFill("solid", fgColor="F2F2F2")
    border_side = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side,
            )
            cell.fill = final_fill if cell.column in final_columns else fixed_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def write_fixed_workbook(records: list[dict], summary: dict, taxonomy: dict) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "00_固定说明"
    for row in [
        ["项目", "说明"],
        ["固定状态", "已由用户明确确认，作为后续本地分类与标签基础信息。"],
        ["覆盖范围", f"共 {summary['scope']['workCount']} 部作品。"],
        ["分类与标签", "以用户填写的最终基础大表为准。"],
        ["作者", "仅沿用此前已收口作者值；本表中的作者显示修改未应用。"],
        ["正式主数据", "尚未写入。"],
    ]:
        readme.append(row)
    apply_style(readme)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 88

    headers = MAIN_HEADERS
    for level1, sheet_name in [("出版物", "01_出版物"), ("网文", "02_网文")]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        subset = [row for row in records if row["一级分类"] == level1]
        for index, row in enumerate(subset, 1):
            sheet.append(
                [
                    index,
                    row["书名"],
                    row["作品编号"],
                    row["作者"],
                    row["一级分类"],
                    row["二级分类"],
                    row["三级分类"],
                    row["辅助标签"],
                    row["备注"],
                ]
            )
        apply_style(sheet, final_columns={5, 6, 7, 8, 9})
        for column, width in {
            "A": 8,
            "B": 34,
            "C": 16,
            "D": 22,
            "E": 18,
            "F": 20,
            "G": 24,
            "H": 34,
            "I": 30,
        }.items():
            sheet.column_dimensions[column].width = width

    fixed_summary = workbook.create_sheet("03_固定摘要")
    fixed_summary.append(["项目", "结果"])
    for key, value in [
        ("确认来源", summary["confirmation"]["source"]),
        ("覆盖作品", summary["scope"]["workCount"]),
        ("出版物", summary["scope"]["level1Distribution"].get("出版物", 0)),
        ("网文", summary["scope"]["level1Distribution"].get("网文", 0)),
        ("相对基线发生修改的作品", summary["changesFromGeneratedBaseline"]["changedWorks"]),
        ("含辅助标签作品", summary["finalDistribution"]["taggedWorks"]),
        ("辅助标签赋值", summary["finalDistribution"]["tagAssignments"]),
        ("分类与标签本地收口", "是"),
        ("正式主数据写入", "否"),
    ]:
        fixed_summary.append([key, value])
    apply_style(fixed_summary)
    fixed_summary.column_dimensions["A"].width = 36
    fixed_summary.column_dimensions["B"].width = 48

    options = workbook.create_sheet("04_分类标签选项")
    options.append(["一级分类", "二级分类", "三级分类", "辅助标签组", "辅助标签"])
    classification_rows = [
        [level1, level2, level3]
        for level1, branches in taxonomy["classificationTree"].items()
        for level2, level3_values in branches.items()
        for level3 in level3_values
    ]
    auxiliary_rows = [
        [group, tag]
        for group, tags in taxonomy["auxiliaryTagGroups"].items()
        for tag in tags
    ]
    for index in range(max(len(classification_rows), len(auxiliary_rows))):
        classification = (
            classification_rows[index]
            if index < len(classification_rows)
            else ["", "", ""]
        )
        auxiliary = (
            auxiliary_rows[index] if index < len(auxiliary_rows) else ["", ""]
        )
        options.append(classification + auxiliary)
    apply_style(options)
    for column, width in {"A": 18, "B": 22, "C": 28, "D": 20, "E": 28}.items():
        options.column_dimensions[column].width = width

    workbook.properties.title = "M2 分类与标签本地固定基础表"
    workbook.properties.creator = "KAtOReNA7/system local development"
    workbook.save(FIXED_WORKBOOK_PATH)


def write_outputs(records: list[dict], summary: dict, taxonomy: dict) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    FIXED_JSON_PATH.write_text(
        json.dumps({"summary": summary, "records": records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_fixed_workbook(records, summary, taxonomy)

    public_summary = {
        key: summary[key]
        for key in [
            "schema",
            "generatedAt",
            "candidateVersion",
            "taxonomyVersion",
            "confirmation",
            "scope",
            "changesFromGeneratedBaseline",
            "userAddedControlledValues",
            "finalDistribution",
            "quality",
            "closure",
            "outputs",
        ]
    }
    PUBLIC_REPORT_JSON_PATH.write_text(
        json.dumps(public_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    changes = summary["changesFromGeneratedBaseline"]
    quality = summary["quality"]
    closure = summary["closure"]
    markdown = f"""# M2 分类与标签本地基础信息收口 v1

## 结论

- 用户已明确确认完整基础大表，确认来源：`{summary['confirmation']['source']}`。
- 已固定 `{summary['scope']['workCount']}` 部作品：出版物 `{summary['scope']['level1Distribution'].get('出版物', 0)}` 部，网文 `{summary['scope']['level1Distribution'].get('网文', 0)}` 部。
- 分类路径、辅助标签、作品编号完整性校验均通过；分类与标签当前人工缺口为 `0`。
- 当前结果是本地文件级固定基础候选，不是正式主数据写入，也不是 M2 formal completion。

## 用户修改

- 相对生成基线发生修改的作品：`{changes['changedWorks']}` 部。
- 字段修改数：`{json.dumps(changes['fieldChangeCounts'], ensure_ascii=False)}`。
- 标签新增赋值：`{changes['tagAdditionAssignments']}`；标签删除赋值：`{changes['tagRemovalAssignments']}`。
- 用户新增三级分类采用分布：`{json.dumps(summary['userAddedControlledValues']['level3Counts'], ensure_ascii=False)}`。
- 用户新增辅助标签采用分布：`{json.dumps(summary['userAddedControlledValues']['tagCounts'], ensure_ascii=False)}`。

## 最终质量

- 唯一作品编号：`{summary['scope']['uniqueWorkIds']}`。
- 缺失必填字段：`{quality['missingRequiredFields']}`；重复编号：`{quality['duplicateWorkIds']}`。
- 无效分类路径：`{quality['invalidClassificationPaths']}`；无效辅助标签：`{quality['invalidAuxiliaryTags']}`。
- 含辅助标签作品：`{summary['finalDistribution']['taggedWorks']}`；标签赋值：`{summary['finalDistribution']['tagAssignments']}`。
- 用户已确认作者显示修改属于误操作；已恢复 `{quality['outOfScopeAuthorEditsConfirmedAccidentalAndRepaired']}` 个单元格，剩余越界作者修改 `{quality['outOfScopeAuthorDisplayEdits']}` 个。

## 人工数据字段状态

- 分类与标签本地收口：`{'是' if closure['classificationAndTagsLocallyClosed'] else '否'}`。
- 当前约定范围内的人工基础字段本地收口：`{'是' if closure['manualFoundationFieldsLocallyClosed'] else '否'}`。
- 正式主数据写入：`否`。
- M2 formal complete：`否`；M3 formal execution：`未开始`。

## 后续人工介入

- 分类与标签不再需要逐条人工补齐。
- 作者误操作已恢复，不再需要用户确认，也不会进入提交。
- 到期但仍有收入、版权有效但收入稀疏等属于后续业务判断，不是基础字段缺失。
- 正式主数据写入、M2 formal gate 和 M3 formal execution 仍需后续单独授权。

## 安全边界

- private 固定基础表和 JSON 位于 Git 忽略范围，禁止提交。
- 可提交报告只包含聚合统计，不包含作品名、作者名、渠道名或行级明细。
- 本轮未连接数据库、未写正式主数据、未进入 M3。
"""
    PUBLIC_REPORT_MD_PATH.write_text(markdown, encoding="utf-8")


def main() -> None:
    args = parse_arguments()
    required = [TAXONOMY_PATH, FILLED_WORKBOOK_PATH, BASELINE_PATH]
    missing = [relative(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit("Required local inputs are missing: " + ", ".join(missing))

    taxonomy = load_taxonomy()
    baseline, _ = load_baseline()
    repaired_author_edits = 0
    if args.repair_out_of_scope_author_edits:
        repaired_author_edits = repair_out_of_scope_author_edits(baseline)
    filled, confirmation = load_filled_workbook()
    records, summary = validate_and_build(
        filled,
        baseline,
        taxonomy,
        confirmation,
        args.confirm_user_message,
        repaired_author_edits,
    )
    write_outputs(records, summary, taxonomy)
    print(
        json.dumps(
            {
                "状态": "已固定",
                "确认来源": summary["confirmation"]["source"],
                "覆盖作品": summary["scope"]["workCount"],
                "出版物": summary["scope"]["level1Distribution"].get("出版物", 0),
                "网文": summary["scope"]["level1Distribution"].get("网文", 0),
                "发生修改的作品": summary["changesFromGeneratedBaseline"]["changedWorks"],
                "分类与标签人工缺口": 0,
                "越界作者显示修改未应用": summary["quality"]["outOfScopeAuthorDisplayEdits"],
                "作者误操作已恢复": summary["quality"]["outOfScopeAuthorEditsConfirmedAccidentalAndRepaired"],
                "输出基础表": relative(FIXED_WORKBOOK_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
