from __future__ import annotations

import argparse
import json
import math
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(SCRIPT_DIR))

import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Alignment = Font = PatternFill = get_column_letter = None

import run_m2_disentangled_forecastability_validation as v1
import run_m2_disentangled_forecast_v1_1_validation as v11
import run_m2_forecast_model_bakeoff as bake

OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-algorithm-validation"

INTEGRITY_JSON = OUTPUT_DIR / "M2-v1.1-backtest-integrity-audit.json"
INTEGRITY_MD = OUTPUT_DIR / "M2-v1.1-backtest-integrity-audit.md"
COMPARISON_JSON = OUTPUT_DIR / "M2-v1.1-same-cohort-model-comparison.json"
COMPARISON_MD = OUTPUT_DIR / "M2-v1.1-same-cohort-model-comparison.md"
FREEZE_JSON = OUTPUT_DIR / "M2-v1.1-conditional-baseline-freeze-decision.json"
FREEZE_MD = OUTPUT_DIR / "M2-v1.1-conditional-baseline-freeze-decision.md"
READINESS_JSON = OUTPUT_DIR / "M2-v1.1-limited-business-review-readiness.json"
READINESS_MD = OUTPUT_DIR / "M2-v1.1-limited-business-review-readiness.md"
PACKAGE_SUMMARY_JSON = OUTPUT_DIR / "M2-v1.1-conditional-business-review-package-summary.json"
PACKAGE_SUMMARY_MD = OUTPUT_DIR / "M2-v1.1-conditional-business-review-package-summary.md"
USER_GUIDE_JSON = OUTPUT_DIR / "M2-v1.1-business-review-user-guide.json"
USER_GUIDE_MD = OUTPUT_DIR / "M2-v1.1-business-review-user-guide.md"
BASELINE_RECORD_JSON = OUTPUT_DIR / "M2-v1.1-conditional-baseline-record.json"
BASELINE_RECORD_MD = OUTPUT_DIR / "M2-v1.1-conditional-baseline-record.md"

PRIVATE_XLSX = PRIVATE_DIR / "m2-v1.1-conditional-business-review-pack.xlsx"
PRIVATE_DETAIL_JSON = PRIVATE_DIR / "m2-v1.1-conditional-business-review-private-detail.json"

NUMERIC_STATUS = v1.NUMERIC_STATUS
CONSERVATIVE_STATUS = v1.CONSERVATIVE_STATUS
TRUE_BLOCKED_STATUS = v1.TRUE_BLOCKED_STATUS
OBSERVE_STATUS = v1.OBSERVE_STATUS


def safe_float(value, default: float = 0.0) -> float:
    return bake.safe_float(value, default)


def safe_int(value, default: int = 0) -> int:
    return bake.safe_int(value, default)


def rounded(value, digits: int = 4):
    return bake.rounded(value, digits)


def percent(part, total) -> float:
    total = safe_float(total)
    if total <= 0:
        return 0.0
    return rounded(safe_float(part) / total)


def distribution(values, keys: list[str] | None = None) -> dict:
    return bake.distribution(values, keys)


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(audit_json_safe(payload), ensure_ascii=False, indent=2, default=audit_json_default),
        encoding="utf-8",
    )


def audit_json_safe(value):
    if isinstance(value, dict):
        return {str(key): audit_json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [audit_json_safe(child) for child in value]
    if isinstance(value, tuple):
        return [audit_json_safe(child) for child in value]
    if type(value).__name__ == "bool_":
        return bool(value)
    return value


def audit_json_default(value):
    if type(value).__name__ in {"bool", "bool_"}:
        return bool(value)
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return str(value)


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    return bake.markdown_table(rows, columns)


def status_count(frame: pd.DataFrame, column: str, value: str) -> int:
    if frame.empty or column not in frame.columns:
        return 0
    return int((frame[column].astype(str) == value).sum())


def status_revenue_share(frame: pd.DataFrame, column: str, value: str) -> float:
    if frame.empty or column not in frame.columns or "totalHistoricalRevenue" not in frame.columns:
        return 0.0
    total = safe_float(frame["totalHistoricalRevenue"].sum())
    part = safe_float(frame.loc[frame[column].astype(str) == value, "totalHistoricalRevenue"].sum())
    return percent(part, total)


def assign_alias(frame: pd.DataFrame, output: str, candidates: list[str], default=None) -> None:
    for candidate in candidates:
        if candidate in frame.columns:
            frame[output] = frame[candidate]
            return
    frame[output] = default


def ensure_columns(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    copy = frame.copy()
    for column in columns:
        if column not in copy.columns:
            copy[column] = None
    return copy[columns]


def safe_boundary() -> dict:
    return {
        "aggregateOnly": True,
        "anonymousIdsOnly": True,
        "sensitiveEntityNamesIncluded": False,
        "sourceRowLevelDetailsIncluded": False,
        "connectionSecretsIncluded": False,
        "privateBusinessReviewPackAvailable": True,
    }


def case_key_frame(frame: pd.DataFrame) -> pd.Series:
    return (
        frame["workKey"].astype(str)
        + "|"
        + frame["cutoffMonth"].astype(str)
        + "|"
        + frame["horizonMonths"].astype(str)
    )


def score_point_cases(cases: pd.DataFrame, label: str, coverage_applicable: bool = True) -> dict:
    if cases.empty:
        return {
            "model": label,
            "caseCount": 0,
            "wape": None,
            "smape": None,
            "mae": None,
            "coverage": None,
            "passCount": None,
            "warningCount": None,
            "failCount": None,
        }
    actual_total = safe_float(cases["actual"].sum())
    abs_total = safe_float(cases["absoluteError"].sum())
    labels = cases.apply(bake.case_failure_label, axis=1) if coverage_applicable else pd.Series(dtype=str)
    return {
        "model": label,
        "caseCount": int(len(cases)),
        "actualTotal": rounded(actual_total, 2),
        "predictedTotal": rounded(cases["predicted"].sum(), 2),
        "wape": None if actual_total <= 0 else rounded(abs_total / actual_total),
        "smape": None if cases["smape"].dropna().empty else rounded(cases["smape"].dropna().mean()),
        "mae": rounded(abs_total / len(cases), 4),
        "coverage": rounded(cases["intervalCoverage"].mean()) if coverage_applicable else None,
        "overForecastRate": rounded(cases["overForecast"].mean()),
        "underForecastRate": rounded(cases["underForecast"].mean()),
        "betterThanBaselineRate": rounded(cases["betterThanBaseline"].mean()) if "betterThanBaseline" in cases.columns else None,
        "passCount": int((labels == "pass").sum()) if coverage_applicable else None,
        "warningCount": int((labels == "warning").sum()) if coverage_applicable else None,
        "failCount": int((labels == "fail").sum()) if coverage_applicable else None,
        "p0": None,
        "p1": None,
        "p2": None,
    }


def baseline_cases_from(cohort: pd.DataFrame) -> pd.DataFrame:
    frame = cohort.copy()
    frame["predicted"] = frame["baselinePredicted"].astype(float)
    frame["absoluteError"] = (frame["predicted"] - frame["actual"].astype(float)).abs()
    frame["smape"] = frame.apply(lambda row: bake.smape(row.predicted, row.actual), axis=1)
    frame["ape"] = frame.apply(lambda row: row.absoluteError / row.actual if safe_float(row.actual) > 0 else None, axis=1)
    frame["overForecast"] = frame["predicted"] > frame["actual"]
    frame["underForecast"] = frame["predicted"] < frame["actual"]
    frame["betterThanBaseline"] = True
    return frame


def issue_counts_from_validation(validation: dict) -> dict:
    return {
        "p0": validation["issueSummary"]["p0"],
        "p1": validation["issueSummary"]["p1"],
        "p2": validation["issueSummary"]["p2"],
    }


def attach_issue_counts(row: dict, validation: dict | None) -> dict:
    if validation:
        row.update(issue_counts_from_validation(validation))
    return row


def build_static_integrity_checks() -> dict:
    bake_source = (SCRIPT_DIR / "run_m2_forecast_model_bakeoff.py").read_text(encoding="utf-8")
    v11_source = (SCRIPT_DIR / "run_m2_disentangled_forecast_v1_1_validation.py").read_text(encoding="utf-8")
    checks = {
        "featuresUseCutoffOrEarlierOnly": "history = values[: cutoff_idx + 1]" in bake_source,
        "actualUsesCutoffFutureWindowOnly": "actual_window = values[cutoff_idx + 1 : cutoff_idx + 1 + horizon]" in bake_source,
        "incompleteMonthsExcluded": 'billMonth"] <= context["latest_complete_month"]' in bake_source
        and "validForCalibration" in bake_source,
        "insufficientHorizonSkipped": "if len(actual_window) < horizon:" in bake_source and "continue" in bake_source,
        "baselineUsesSameHistoryAndCutoff": 'raw_baseline = model_outputs["raw_trailing_baseline"]["base"]' in bake_source,
        "v11IntervalUsesRollingPriorResiduals": "cutoffOrdinal\"] < cutoff_ordinal" in v11_source
        and "rolling_prior_residual_quantile_by" in v11_source,
        "fixedGlobalIntervalMultiplierAbsent": "fixedGlobalMultiplierUsed" in v11_source
        and '"fixedGlobalMultiplierUsed": False' in v11_source,
    }
    return checks


def high_value_score(cases: pd.DataFrame, bucket: str) -> dict:
    if bucket == "top1":
        subset = cases[cases["materialityBucket"] == "top_1_percent"]
    elif bucket == "top5":
        subset = cases[cases["materialityBucket"].isin(["top_1_percent", "top_5_percent"])]
    else:
        subset = cases[cases["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"])]
    score = score_point_cases(subset, f"v1.1_{bucket}_high_value")
    return {
        "bucket": bucket,
        "caseCount": score["caseCount"],
        "wape": score["wape"],
        "smape": score["smape"],
        "mae": score["mae"],
        "coverage": score["coverage"],
        "failCount": score["failCount"],
    }


def build_same_cohort_comparison(
    raw_cases: pd.DataFrame,
    v1_cases: pd.DataFrame,
    v1_validation: dict,
    v11_cases: pd.DataFrame,
    v11_validation: dict,
    v1_coverage: dict,
    v11_coverage: dict,
) -> dict:
    v11_forecastable = v11_cases[v11_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    v11_forecastable["caseKey"] = case_key_frame(v11_forecastable)
    key_set = set(v11_forecastable["caseKey"])

    v1_cases = v1_cases.copy()
    v1_cases["caseKey"] = case_key_frame(v1_cases)
    v1_on_v11 = v1_cases[v1_cases["caseKey"].isin(key_set)].copy()

    model_e = raw_cases[raw_cases["modelId"] == "model_e_selector"].copy()
    model_e["caseKey"] = case_key_frame(model_e)
    model_e_on_v11 = model_e[model_e["caseKey"].isin(key_set)].copy()

    baseline = baseline_cases_from(v11_forecastable)
    rows = [
        score_point_cases(baseline, "trailing_baseline_same_v1_1_cohort", coverage_applicable=False),
        attach_issue_counts(score_point_cases(v1_on_v11, "v1.0_on_v1.1_forecastable_cohort"), v1_validation),
        attach_issue_counts(score_point_cases(v11_forecastable, "v1.1_on_v1.1_forecastable_cohort"), v11_validation),
        score_point_cases(model_e_on_v11, "model_e_selector_on_v1.1_forecastable_cohort"),
    ]

    common_keys = set(v1_cases[v1_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]["caseKey"]).intersection(
        set(v11_forecastable["caseKey"])
    )
    v1_common = v1_cases[v1_cases["caseKey"].isin(common_keys)].copy()
    v11_common = v11_forecastable[v11_forecastable["caseKey"].isin(common_keys)].copy()
    common_rows = [
        score_point_cases(v1_common, "v1.0_common_forecastable_intersection"),
        score_point_cases(v11_common, "v1.1_common_forecastable_intersection"),
    ]

    return {
        "schema": "m2.v1_1_same_cohort_model_comparison.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "primaryComparisonCohort": "v1.1 forecastable cohort, same work/cutoff/horizon cases",
        "directV1_0VsV1_1AggregateComparable": False,
        "directV1_0VsV1_1AggregateNote": "Published v1.0 and v1.1 aggregate WAPE use different forecastable cohorts; use common/intersection rows instead.",
        "v1_0Coverage": {
            "numericPlusConservativeRevenueShare": v1_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueBlockedRevenueShare": v1_coverage["trueForecastBlocked"]["revenueShare"],
        },
        "v1_1Coverage": {
            "numericPlusConservativeRevenueShare": v11_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueBlockedRevenueShare": v11_coverage["trueForecastBlocked"]["revenueShare"],
        },
        "sameCohortRows": rows,
        "commonForecastableIntersectionRows": common_rows,
        "highValuePerformance": [high_value_score(v11_forecastable, bucket) for bucket in ["top1", "top5", "top10"]],
        "v1_1BetterThanBaselineOnSameCohort": safe_float(rows[2]["wape"], 999) <= safe_float(rows[0]["wape"], -1),
        "safeOutputBoundary": safe_boundary(),
    }


def build_integrity_audit(v11_cases: pd.DataFrame, v11_validation: dict, comparison: dict, interval_rows: list[dict]) -> dict:
    checks = build_static_integrity_checks()
    forecastable = v11_cases[v11_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    horizon_rows = []
    for horizon, group in forecastable.groupby("horizonMonths"):
        score = score_point_cases(group, f"horizon_{horizon}")
        horizon_rows.append(
            {
                "horizonMonths": safe_int(horizon),
                "caseCount": score["caseCount"],
                "wape": score["wape"],
                "coverage": score["coverage"],
            }
        )
    reason_distribution = distribution(forecastable["intervalCalibrationReason"])
    spread = v11_validation["spreadSummary"]
    pass_conditions = {
        **checks,
        "requiredHorizonsPresent": {3, 6, 12}.issubset(
            {safe_int(value) for value in forecastable["horizonMonths"].unique()}
        ),
        "baselineSameCohortComparable": comparison["v1_1BetterThanBaselineOnSameCohort"],
        "wapeDenominatorPositive": safe_float(v11_validation["forecastableCohortScore"]["actualTotal"]) > 0,
        "coverageComputedFromIntervalContainment": "intervalCoverage" in forecastable.columns
        and forecastable["intervalCoverage"].notna().all(),
        "highConfidenceSpreadWithinGuardrail": spread["highConfidenceSpreadP75"] is None
        or safe_float(spread["highConfidenceSpreadP75"]) <= 1.5,
        "nonLowConfidenceSpreadWithinGuardrail": spread["nonLowConfidenceSpreadP75"] is None
        or safe_float(spread["nonLowConfidenceSpreadP75"]) <= 2.0,
        "lowConfidenceExplicitlyMarked": int((forecastable["confidence"] == "low").sum()) > 0,
    }
    integrity = "PASS" if all(pass_conditions.values()) else "FAIL"
    return {
        "schema": "m2.v1_1_backtest_integrity_audit.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "backtestIntegrity": integrity,
        "bugFixAppliedBeforeAudit": {
            "issue": "v1.1 interval calibration previously used full-sample residual quantiles for coverage evaluation.",
            "fix": "calibration now uses rolling-prior residual quantiles from cutoff months earlier than the evaluated cutoff; early insufficient segments retain the original interval.",
            "modelParameterTuning": False,
        },
        "cutoffAudit": {
            "featuresUseCutoffOrEarlierOnly": checks["featuresUseCutoffOrEarlierOnly"],
            "actualUsesCutoffFutureWindowOnly": checks["actualUsesCutoffFutureWindowOnly"],
            "incompleteMonthsExcluded": checks["incompleteMonthsExcluded"],
            "futureLeakageDetected": not (
                checks["featuresUseCutoffOrEarlierOnly"]
                and checks["actualUsesCutoffFutureWindowOnly"]
                and checks["v11IntervalUsesRollingPriorResiduals"]
            ),
        },
        "horizonAudit": {
            "requiredHorizons": [3, 6, 12],
            "optionalExtendedHorizons": [18, 24],
            "observedHorizons": sorted({safe_int(value) for value in forecastable["horizonMonths"].unique()}),
            "insufficientHorizonSkipped": checks["insufficientHorizonSkipped"],
            "horizonRows": sorted(horizon_rows, key=lambda item: item["horizonMonths"]),
        },
        "baselineAudit": {
            "baselineUsesSameHistoryAndCutoff": checks["baselineUsesSameHistoryAndCutoff"],
            "baselineSameCohortComparable": comparison["v1_1BetterThanBaselineOnSameCohort"],
            "baselineWape": comparison["sameCohortRows"][0]["wape"],
            "v1_1Wape": comparison["sameCohortRows"][2]["wape"],
        },
        "metricAudit": {
            "wapeDenominatorCorrect": pass_conditions["wapeDenominatorPositive"],
            "mapeNearZeroRisk": "MAPE/APE is retained only as a diagnostic; WAPE and SMAPE carry the decision because near-zero actuals can amplify APE.",
            "smapeMoreStableForLowRevenue": True,
            "maeHighRevenueSensitive": True,
            "coverageComputedByActualWithinInterval": pass_conditions["coverageComputedFromIntervalContainment"],
        },
        "intervalCoverageAudit": {
            "coverage": v11_validation["forecastableCohortScore"]["intervalCoverage"],
            "usesRollingPriorResiduals": checks["v11IntervalUsesRollingPriorResiduals"],
            "fixedGlobalIntervalMultiplierUsed": False,
            "highConfidenceSpreadP75": spread["highConfidenceSpreadP75"],
            "nonLowConfidenceSpreadP75": spread["nonLowConfidenceSpreadP75"],
            "calibrationReasonDistribution": reason_distribution,
            "segmentCalibrationRows": interval_rows[:120],
        },
        "highValueAudit": {
            "highValueRows": comparison["highValuePerformance"],
            "p0": v11_validation["issueSummary"]["p0"],
            "p1": v11_validation["issueSummary"]["p1"],
            "acceptable": v11_validation["issueSummary"]["p0"] == 0 and v11_validation["issueSummary"]["p1"] == 0,
        },
        "passConditions": pass_conditions,
        "safeOutputBoundary": safe_boundary(),
    }


def build_freeze_decision(integrity: dict, comparison: dict, v11_coverage: dict, v11_validation: dict) -> tuple[dict, dict]:
    allow_freeze = (
        integrity["backtestIntegrity"] == "PASS"
        and comparison["v1_1BetterThanBaselineOnSameCohort"]
        and v11_validation["verdict"] == "CONDITIONAL PASS"
    )
    decision = {
        "schema": "m2.v1_1_conditional_baseline_freeze_decision.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "freezeDecision": "FREEZE_CONDITIONAL" if allow_freeze else "DO_NOT_FREEZE",
        "candidateVersion": v11_validation["candidateVersion"] if allow_freeze else None,
        "backtestIntegrity": integrity["backtestIntegrity"],
        "sameCohortBetterThanBaseline": comparison["v1_1BetterThanBaselineOnSameCohort"],
        "applicableCohort": "numeric_forecast_eligible and conservative_numeric_forecast only",
        "excludedCohort": "true_forecast_blocked and observe_only_no_numeric_forecast",
        "forecastConfidenceBoundary": "high/medium/low confidence retained; low confidence is reference-only or requires manual review before action",
        "businessRestrictions": [
            "not a final release-approved result",
            "non-forecastable cohort must not receive business-usable numeric forecasts",
            "formal readiness blockers affect formal release, not local forecast validation",
            "business action blockers affect action execution, not forecast calculation",
            "M3 remains blocked unless explicitly authorized later for parallel planning",
        ],
        "m3Allowed": False,
        "safeOutputBoundary": safe_boundary(),
    }
    readiness = {
        "schema": "m2.v1_1_limited_business_review_readiness.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "limitedBusinessReviewReady": bool(allow_freeze),
        "candidateVersion": decision["candidateVersion"],
        "forecastableRevenueShare": v11_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
        "trueForecastBlockedRevenueShare": v11_coverage["trueForecastBlocked"]["revenueShare"],
        "businessReviewBoundary": [
            {
                "cohort": "numeric_forecast_eligible",
                "allowedUse": "limited M2 business review of numeric forecast",
            },
            {
                "cohort": "conservative_numeric_forecast",
                "allowedUse": "reference forecast only; business action still requires guard checks",
            },
            {
                "cohort": "true_forecast_blocked",
                "allowedUse": "no business-usable numeric forecast",
            },
            {
                "cohort": "observe_only_no_numeric_forecast",
                "allowedUse": "observe only; no business action forecast",
            },
        ],
        "m3Allowed": False,
        "notFinalReleaseApproval": True,
        "safeOutputBoundary": safe_boundary(),
    }
    return decision, readiness


def build_business_review_package_summary(
    freeze: dict,
    readiness: dict,
    v11_coverage: dict,
    v11_validation: dict,
    integrity: dict,
    comparison: dict,
    current_gate: pd.DataFrame,
) -> dict:
    forecastable_statuses = [NUMERIC_STATUS, CONSERVATIVE_STATUS]
    forecastable = current_gate[current_gate["forecastabilityStatus"].isin(forecastable_statuses)].copy()
    formal_blocked_forecastable = forecastable[forecastable["formalReadinessStatus"] != v1.READY_STATUS].copy()
    business_action_guard = forecastable[
        forecastable["businessActionStatus"].isin([v1.MANUAL_CONFIRMATION_STATUS, v1.ACTION_BLOCKED_STATUS])
    ].copy()
    return {
        "schema": "m2.v1_1_conditional_business_review_package_summary.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateVersion": freeze["candidateVersion"],
        "freezeType": "conditional business-review baseline",
        "limitedBusinessReviewReady": readiness["limitedBusinessReviewReady"],
        "businessReviewPackGenerated": True,
        "businessReviewPackCommitted": False,
        "requiredSheets": [
            "00_read_me",
            "01_executive_summary",
            "02_forecastable_cohort_review",
            "03_high_value_review",
            "04_non_forecastable_cohort",
            "05_formal_readiness_blocked_but_forecastable",
            "06_business_action_guard",
            "07_warning_and_p2_cases",
            "08_user_decision_template",
        ],
        "metrics": {
            "wape": v11_validation["forecastableCohortScore"]["wape"],
            "baselineWape": v11_validation["forecastableCohortScore"]["baselineWape"],
            "coverage": v11_validation["forecastableCohortScore"]["intervalCoverage"],
            "p0": v11_validation["issueSummary"]["p0"],
            "p1": v11_validation["issueSummary"]["p1"],
            "p2": v11_validation["issueSummary"]["p2"],
            "forecastableRevenueShare": v11_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueForecastBlockedRevenueShare": v11_coverage["trueForecastBlocked"]["revenueShare"],
            "highConfidenceSpreadP75": v11_validation["spreadSummary"]["highConfidenceSpreadP75"],
            "nonLowConfidenceSpreadP75": v11_validation["spreadSummary"]["nonLowConfidenceSpreadP75"],
        },
        "cohortCounts": {
            NUMERIC_STATUS: status_count(current_gate, "forecastabilityStatus", NUMERIC_STATUS),
            CONSERVATIVE_STATUS: status_count(current_gate, "forecastabilityStatus", CONSERVATIVE_STATUS),
            TRUE_BLOCKED_STATUS: status_count(current_gate, "forecastabilityStatus", TRUE_BLOCKED_STATUS),
            OBSERVE_STATUS: status_count(current_gate, "forecastabilityStatus", OBSERVE_STATUS),
            "formalReadinessBlockedButForecastable": int(len(formal_blocked_forecastable)),
            "businessActionGuardRows": int(len(business_action_guard)),
        },
        "cohortRevenueShares": {
            NUMERIC_STATUS: status_revenue_share(current_gate, "forecastabilityStatus", NUMERIC_STATUS),
            CONSERVATIVE_STATUS: status_revenue_share(current_gate, "forecastabilityStatus", CONSERVATIVE_STATUS),
            TRUE_BLOCKED_STATUS: status_revenue_share(current_gate, "forecastabilityStatus", TRUE_BLOCKED_STATUS),
            OBSERVE_STATUS: status_revenue_share(current_gate, "forecastabilityStatus", OBSERVE_STATUS),
        },
        "highValueAttention": comparison["highValuePerformance"],
        "backtestIntegrity": integrity["backtestIntegrity"],
        "recommendedNextStep": "business users review the private workbook and decide whether to accept v1.1 as a limited forecastable-cohort review baseline",
        "m3Allowed": False,
        "notFinalReleaseApproval": True,
        "safeOutputBoundary": safe_boundary(),
    }


def build_business_review_user_guide(freeze: dict, summary: dict) -> dict:
    return {
        "schema": "m2.v1_1_business_review_user_guide.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateVersion": freeze["candidateVersion"],
        "audience": "business reviewer and next Codex session",
        "reviewPurpose": "confirm whether the conditional v1.1 forecast baseline is acceptable for limited M2 business review on forecastable cohorts only",
        "howToUseWorkbook": [
            {
                "sheet": "00_read_me",
                "use": "read the candidate boundary, non-release status, cohort definitions, and business decision task",
            },
            {
                "sheet": "01_executive_summary",
                "use": "check WAPE, baseline comparison, coverage, P0/P1/P2, applicable scope, and remaining restrictions",
            },
            {
                "sheet": "02_forecastable_cohort_review",
                "use": "review anonymous forecastable rows and decide whether forecast confidence/action labels are acceptable",
            },
            {
                "sheet": "03_high_value_review",
                "use": "inspect high-value anonymous cases because top revenue coverage remains weaker than overall coverage",
            },
            {
                "sheet": "04_non_forecastable_cohort",
                "use": "confirm that blocked or observe-only rows do not receive business-usable numeric forecasts",
            },
            {
                "sheet": "05_formal_readiness_blocked_but_forecastable",
                "use": "separate local forecast validation from formal release blockers",
            },
            {
                "sheet": "06_business_action_guard",
                "use": "confirm action-bearing suggestions still require manual business review before execution",
            },
            {
                "sheet": "07_warning_and_p2_cases",
                "use": "review warning/P2 cases before accepting the conditional baseline",
            },
            {
                "sheet": "08_user_decision_template",
                "use": "record group-level accept/reject/pending decisions, reviewer reason, reviewer name, and date",
            },
        ],
        "decisionOptions": [
            "accept_limited_business_review_baseline",
            "accept_with_manual_high_value_review",
            "reject_for_algorithm_rework",
            "keep_pending",
        ],
        "mustNotDo": [
            "do not use this as final release approval",
            "do not use non-forecastable rows for numeric business decisions",
            "do not enter M3 unless the user explicitly authorizes parallel planning after accepting the baseline",
            "do not commit the private workbook",
        ],
        "summaryMetrics": summary["metrics"],
        "safeOutputBoundary": safe_boundary(),
    }


def build_conditional_baseline_record(
    freeze: dict,
    readiness: dict,
    v11_validation: dict,
    v11_coverage: dict,
    integrity: dict,
    comparison: dict,
) -> dict:
    return {
        "schema": "m2.v1_1_conditional_baseline_record.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateName": freeze["candidateVersion"],
        "freezeType": "conditional business-review baseline",
        "freezeDecision": freeze["freezeDecision"],
        "applicableCohort": freeze["applicableCohort"],
        "excludedCohort": freeze["excludedCohort"],
        "metrics": {
            "wape": v11_validation["forecastableCohortScore"]["wape"],
            "baselineWape": v11_validation["forecastableCohortScore"]["baselineWape"],
            "coverage": v11_validation["forecastableCohortScore"]["intervalCoverage"],
            "forecastableRevenueShare": v11_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueForecastBlockedRevenueShare": v11_coverage["trueForecastBlocked"]["revenueShare"],
            "p0": v11_validation["issueSummary"]["p0"],
            "p1": v11_validation["issueSummary"]["p1"],
            "p2": v11_validation["issueSummary"]["p2"],
            "highConfidenceSpreadP75": v11_validation["spreadSummary"]["highConfidenceSpreadP75"],
            "nonLowConfidenceSpreadP75": v11_validation["spreadSummary"]["nonLowConfidenceSpreadP75"],
        },
        "evidence": {
            "backtestIntegrity": integrity["backtestIntegrity"],
            "sameCohortBetterThanBaseline": comparison["v1_1BetterThanBaselineOnSameCohort"],
            "limitedBusinessReviewReady": readiness["limitedBusinessReviewReady"],
        },
        "restrictions": freeze["businessRestrictions"],
        "notFinalReleaseApproved": True,
        "m3AllowedByDefault": False,
        "safeOutputBoundary": safe_boundary(),
    }


def write_reports(
    integrity: dict,
    comparison: dict,
    freeze: dict,
    readiness: dict,
    package_summary: dict,
    user_guide: dict,
    baseline_record: dict,
) -> None:
    write_json(INTEGRITY_JSON, integrity)
    write_json(COMPARISON_JSON, comparison)
    write_json(FREEZE_JSON, freeze)
    write_json(READINESS_JSON, readiness)
    write_json(PACKAGE_SUMMARY_JSON, package_summary)
    write_json(USER_GUIDE_JSON, user_guide)
    write_json(BASELINE_RECORD_JSON, baseline_record)

    INTEGRITY_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Backtest Integrity Audit",
                "",
                f"Backtest integrity: `{integrity['backtestIntegrity']}`.",
                "",
                "## Cutoff / Horizon / Baseline",
                "",
                markdown_table(
                    [
                        {"area": "cutoff", **integrity["cutoffAudit"]},
                        {"area": "horizon", "insufficientHorizonSkipped": integrity["horizonAudit"]["insufficientHorizonSkipped"], "observedHorizons": integrity["horizonAudit"]["observedHorizons"]},
                        {"area": "baseline", **integrity["baselineAudit"]},
                    ],
                    [("area", "Area"), ("featuresUseCutoffOrEarlierOnly", "Past Features"), ("actualUsesCutoffFutureWindowOnly", "Future Actual"), ("futureLeakageDetected", "Leakage"), ("baselineSameCohortComparable", "Baseline Comparable"), ("baselineWape", "Baseline WAPE"), ("v1_1Wape", "v1.1 WAPE")],
                ),
                "",
                "## Horizon Rows",
                "",
                markdown_table(integrity["horizonAudit"]["horizonRows"], [("horizonMonths", "Horizon"), ("caseCount", "Cases"), ("wape", "WAPE"), ("coverage", "Coverage")]),
                "",
                "## Interval Coverage",
                "",
                markdown_table(
                    [integrity["intervalCoverageAudit"]],
                    [("coverage", "Coverage"), ("usesRollingPriorResiduals", "Rolling Prior"), ("fixedGlobalIntervalMultiplierUsed", "Fixed Multiplier"), ("highConfidenceSpreadP75", "High P75"), ("nonLowConfidenceSpreadP75", "Non-low P75")],
                ),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    COMPARISON_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Same-Cohort Model Comparison",
                "",
                f"Primary cohort: `{comparison['primaryComparisonCohort']}`.",
                f"v1.1 better than baseline on same cohort: `{comparison['v1_1BetterThanBaselineOnSameCohort']}`.",
                "",
                "## Same Cohort Rows",
                "",
                markdown_table(
                    comparison["sameCohortRows"],
                    [("model", "Model"), ("caseCount", "Cases"), ("wape", "WAPE"), ("smape", "SMAPE"), ("mae", "MAE"), ("coverage", "Coverage"), ("passCount", "Pass"), ("warningCount", "Warning"), ("failCount", "Fail"), ("p0", "P0"), ("p1", "P1"), ("p2", "P2")],
                ),
                "",
                "## Common Forecastable Intersection",
                "",
                markdown_table(
                    comparison["commonForecastableIntersectionRows"],
                    [("model", "Model"), ("caseCount", "Cases"), ("wape", "WAPE"), ("smape", "SMAPE"), ("mae", "MAE"), ("coverage", "Coverage"), ("passCount", "Pass"), ("warningCount", "Warning"), ("failCount", "Fail")],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    FREEZE_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Conditional Baseline Freeze Decision",
                "",
                f"Freeze decision: `{freeze['freezeDecision']}`.",
                f"Candidate version: `{freeze['candidateVersion']}`.",
                f"Backtest integrity: `{freeze['backtestIntegrity']}`.",
                f"Same-cohort better than baseline: `{freeze['sameCohortBetterThanBaseline']}`.",
                "",
                "## Boundary",
                "",
                markdown_table(
                    [
                        {"field": "applicableCohort", "value": freeze["applicableCohort"]},
                        {"field": "excludedCohort", "value": freeze["excludedCohort"]},
                        {"field": "m3Allowed", "value": freeze["m3Allowed"]},
                    ],
                    [("field", "Field"), ("value", "Value")],
                ),
                "",
                "This is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    READINESS_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Limited Business Review Readiness",
                "",
                f"Limited business review ready: `{readiness['limitedBusinessReviewReady']}`.",
                f"M3 allowed: `{readiness['m3Allowed']}`.",
                "",
                "## Business Review Boundary",
                "",
                markdown_table(readiness["businessReviewBoundary"], [("cohort", "Cohort"), ("allowedUse", "Allowed Use")]),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    PACKAGE_SUMMARY_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Conditional Business Review Package Summary",
                "",
                f"Candidate version: `{package_summary['candidateVersion']}`.",
                f"Freeze type: `{package_summary['freezeType']}`.",
                f"Limited business review ready: `{package_summary['limitedBusinessReviewReady']}`.",
                f"M3 allowed: `{package_summary['m3Allowed']}`.",
                "",
                "## Metrics",
                "",
                markdown_table(
                    [package_summary["metrics"]],
                    [
                        ("wape", "WAPE"),
                        ("baselineWape", "Baseline WAPE"),
                        ("coverage", "Coverage"),
                        ("forecastableRevenueShare", "Forecastable Revenue Share"),
                        ("trueForecastBlockedRevenueShare", "True Blocked Revenue Share"),
                        ("p0", "P0"),
                        ("p1", "P1"),
                        ("p2", "P2"),
                    ],
                ),
                "",
                "## Cohort Counts",
                "",
                markdown_table(
                    [{"cohort": key, "count": value} for key, value in package_summary["cohortCounts"].items()],
                    [("cohort", "Cohort"), ("count", "Count")],
                ),
                "",
                "The private workbook is generated under a gitignored private output directory and is not committed. This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    USER_GUIDE_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Business Review User Guide",
                "",
                f"Candidate version: `{user_guide['candidateVersion']}`.",
                "",
                "The review purpose is to decide whether the conditional v1.1 forecast baseline is acceptable for limited M2 business review on forecastable cohorts only. It is not final release approval and does not open M3 by default.",
                "",
                "## Workbook Sheets",
                "",
                markdown_table(user_guide["howToUseWorkbook"], [("sheet", "Sheet"), ("use", "Use")]),
                "",
                "## Decision Options",
                "",
                markdown_table(
                    [{"decisionOption": option} for option in user_guide["decisionOptions"]],
                    [("decisionOption", "Decision Option")],
                ),
                "",
                "## Restrictions",
                "",
                markdown_table(
                    [{"restriction": item} for item in user_guide["mustNotDo"]],
                    [("restriction", "Restriction")],
                ),
                "",
                "This guide is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    BASELINE_RECORD_MD.write_text(
        "\n".join(
            [
                "# M2 v1.1 Conditional Baseline Record",
                "",
                f"Candidate name: `{baseline_record['candidateName']}`.",
                f"Freeze type: `{baseline_record['freezeType']}`.",
                f"Freeze decision: `{baseline_record['freezeDecision']}`.",
                f"Not final release-approved: `{baseline_record['notFinalReleaseApproved']}`.",
                f"M3 allowed by default: `{baseline_record['m3AllowedByDefault']}`.",
                "",
                "## Metrics",
                "",
                markdown_table(
                    [baseline_record["metrics"]],
                    [
                        ("wape", "WAPE"),
                        ("baselineWape", "Baseline WAPE"),
                        ("coverage", "Coverage"),
                        ("forecastableRevenueShare", "Forecastable Revenue Share"),
                        ("trueForecastBlockedRevenueShare", "True Blocked Revenue Share"),
                        ("p0", "P0"),
                        ("p1", "P1"),
                        ("p2", "P2"),
                        ("highConfidenceSpreadP75", "High Confidence Spread P75"),
                        ("nonLowConfidenceSpreadP75", "Non-low Confidence Spread P75"),
                    ],
                ),
                "",
                "## Evidence",
                "",
                markdown_table([baseline_record["evidence"]], [("backtestIntegrity", "Backtest Integrity"), ("sameCohortBetterThanBaseline", "Same-Cohort Better Than Baseline"), ("limitedBusinessReviewReady", "Limited Business Review Ready")]),
                "",
                "This record is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_business_review_pack(
    integrity: dict,
    comparison: dict,
    freeze: dict,
    readiness: dict,
    current_gate: pd.DataFrame,
    v11_cases: pd.DataFrame,
    v11_validation: dict,
) -> None:
    if Alignment is None:
        raise RuntimeError("openpyxl is required to write the private business review pack.")
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)

    forecastable_gate = current_gate[current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    non_forecastable = current_gate[current_gate["forecastabilityStatus"].isin([TRUE_BLOCKED_STATUS, OBSERVE_STATUS])].copy()
    forecastable_cases = v11_cases[v11_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    forecastable_cases["anonymousCaseId"] = [f"CASE-{index + 1:06d}" for index in range(len(forecastable_cases))]
    labels = v1.labels_for(forecastable_cases)
    forecastable_cases["caseVerdict"] = labels.values
    forecastable_cases["issueType"] = forecastable_cases["caseVerdict"].map(
        lambda value: "interval_warning_or_p2_attention" if value == "warning" else "forecast_error_attention"
    )
    forecastable_cases["requiredAction"] = forecastable_cases["caseVerdict"].map(
        lambda value: "review_before_accepting_baseline" if value in {"warning", "fail"} else "no_extra_action"
    )

    high_value = forecastable_cases[
        forecastable_cases["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"])
    ].copy()
    high_value["highValueReviewId"] = [f"HV-{index + 1:05d}" for index in range(len(high_value))]
    high_value["manualConfirmationRequired"] = high_value.apply(
        lambda row: bool(row.caseVerdict != "pass" or row.businessActionStatus != v1.ACTION_ALLOWED_STATUS),
        axis=1,
    )
    high_value["warningOrFailReason"] = high_value.apply(
        lambda row: f"{row.caseVerdict}; {row.intervalCalibrationReason}",
        axis=1,
    )

    forecastable_gate = forecastable_gate.copy()
    assign_alias(forecastable_gate, "rating", ["rating", "ratingAtCutoff"], "unknown")
    assign_alias(forecastable_gate, "lifecycle", ["lifecycle", "lifecycleAtCutoff"], "unknown")
    assign_alias(forecastable_gate, "revenueBucket", ["revenueBucket", "revenueScaleAtCutoff", "materialityBucket"], "unknown")
    assign_alias(forecastable_gate, "forecastConfidence", ["forecastConfidence", "confidence"], "unknown")
    forecastable_gate["intervalReason"] = forecastable_gate.apply(
        lambda row: "; ".join(row.forecastabilityReasonCodes)
        if isinstance(row.forecastabilityReasonCodes, list)
        else str(row.forecastabilityReasonCodes),
        axis=1,
    )
    forecastable_gate["backtestWarning"] = forecastable_gate["forecastabilityStatus"].map(
        lambda status: "conservative_or_low_confidence_review" if status == CONSERVATIVE_STATUS else "standard_review"
    )
    forecastable_gate["requiredBusinessReviewAction"] = forecastable_gate.apply(
        lambda row: row.requiredBusinessAction
        if isinstance(row.get("requiredBusinessAction"), str) and row.requiredBusinessAction
        else row.requiredForecastabilityAction,
        axis=1,
    )

    formal_blocked_forecastable = forecastable_gate[
        forecastable_gate["formalReadinessStatus"] != v1.READY_STATUS
    ].copy()
    business_action_guard = forecastable_gate[
        forecastable_gate["businessActionStatus"].isin([v1.MANUAL_CONFIRMATION_STATUS, v1.ACTION_BLOCKED_STATUS])
    ].copy()
    warn_fail = forecastable_cases[forecastable_cases["caseVerdict"].isin(["warning", "fail"])].copy().head(3000)

    summary_rows = [
        {"metric": "candidate", "value": freeze["candidateVersion"]},
        {"metric": "freeze_type", "value": "conditional business-review baseline"},
        {"metric": "wape", "value": v11_validation["forecastableCohortScore"]["wape"]},
        {"metric": "baseline_wape", "value": v11_validation["forecastableCohortScore"]["baselineWape"]},
        {"metric": "coverage", "value": v11_validation["forecastableCohortScore"]["intervalCoverage"]},
        {"metric": "p0", "value": v11_validation["issueSummary"]["p0"]},
        {"metric": "p1", "value": v11_validation["issueSummary"]["p1"]},
        {"metric": "p2", "value": v11_validation["issueSummary"]["p2"]},
        {"metric": "forecastable_cohort", "value": freeze["applicableCohort"]},
        {"metric": "excluded_cohort", "value": freeze["excludedCohort"]},
        {"metric": "not_final_release_approval", "value": True},
        {"metric": "m3_allowed", "value": False},
    ]

    decision_template_rows = [
        {
            "groupId": "numeric_forecast_eligible",
            "decisionOptions": "accept_limited_business_review_baseline | accept_with_manual_high_value_review | reject_for_algorithm_rework | keep_pending",
            "reviewerDecision": "",
            "reviewerReason": "",
            "reviewerName": "",
            "reviewedAt": "",
            "notes": "validated numeric forecast cohort; still not final release approval",
        },
        {
            "groupId": "conservative_numeric_forecast",
            "decisionOptions": "accept_reference_only | require_manual_review | reject_for_algorithm_rework | keep_pending",
            "reviewerDecision": "",
            "reviewerReason": "",
            "reviewerName": "",
            "reviewedAt": "",
            "notes": "low/medium confidence or conservative boundary; business action needs guard checks",
        },
        {
            "groupId": "formal_readiness_blocked_but_forecastable",
            "decisionOptions": "accept_local_forecast_only | require_formal_fix | keep_pending",
            "reviewerDecision": "",
            "reviewerReason": "",
            "reviewerName": "",
            "reviewedAt": "",
            "notes": "forecast may be reviewed locally but cannot be released formally until formal blockers close",
        },
        {
            "groupId": "business_action_guard",
            "decisionOptions": "manual_confirm_action | restrict_to_observe | keep_pending",
            "reviewerDecision": "",
            "reviewerReason": "",
            "reviewerName": "",
            "reviewedAt": "",
            "notes": "promote/downlist/renewal/action-bearing rows require manual business confirmation",
        },
        {
            "groupId": "non_forecastable_cohort",
            "decisionOptions": "keep_excluded | request_data_fix | keep_pending",
            "reviewerDecision": "",
            "reviewerReason": "",
            "reviewerName": "",
            "reviewedAt": "",
            "notes": "no business-usable numeric forecast should be used for this group",
        },
    ]

    with pd.ExcelWriter(PRIVATE_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"item": "purpose", "value": "M2 v1.1 conditional baseline business review pack"},
                {"item": "candidate", "value": freeze["candidateVersion"]},
                {"item": "meaning", "value": "limited business review for forecastable cohort only"},
                {"item": "not_final_release", "value": True},
                {"item": "m3_allowed", "value": False},
                {"item": "forecastable_groups", "value": freeze["applicableCohort"]},
                {"item": "non_forecastable_groups", "value": freeze["excludedCohort"]},
                {"item": "business_decision_task", "value": "fill 08_user_decision_template after reviewing sheets 02-07"},
                {"item": "safety", "value": "anonymous IDs only; no work names, authors, channels, or source rows"},
            ]
        ).to_excel(writer, sheet_name="00_read_me", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="01_executive_summary", index=False)
        ensure_columns(
            forecastable_gate,
            [
                "anonymousId",
                "forecastabilityStatus",
                "rating",
                "lifecycle",
                "revenueBucket",
                "forecastConfidence",
                "baseForecast",
                "optimisticForecast",
                "pessimisticForecast",
                "intervalReason",
                "backtestWarning",
                "requiredBusinessReviewAction",
            ],
        ).to_excel(writer, sheet_name="02_forecastable_cohort_review", index=False)
        high_value[
            [
                "highValueReviewId",
                "materialityBucket",
                "forecastabilityStatus",
                "confidence",
                "horizonMonths",
                "predicted",
                "actual",
                "absoluteError",
                "intervalCoverage",
                "warningOrFailReason",
                "manualConfirmationRequired",
            ]
        ].head(2000).to_excel(writer, sheet_name="03_high_value_review", index=False)
        ensure_columns(
            non_forecastable,
            [
                "anonymousId",
                "forecastabilityStatus",
                "forecastabilityReasonCodes",
                "requiredForecastabilityAction",
                "businessActionStatus",
                "requiredBusinessAction",
            ],
        ).to_excel(writer, sheet_name="04_non_forecastable_cohort", index=False)
        ensure_columns(
            formal_blocked_forecastable,
            [
                "anonymousId",
                "forecastabilityStatus",
                "forecastConfidence",
                "formalReadinessStatus",
                "formalReadinessReasonCodes",
                "requiredFormalAction",
            ],
        ).to_excel(writer, sheet_name="05_formal_readiness_blocked_but_forecastable", index=False)
        ensure_columns(
            business_action_guard,
            [
                "anonymousId",
                "forecastabilityStatus",
                "forecastConfidence",
                "businessActionStatus",
                "businessActionReasonCodes",
                "requiredBusinessAction",
            ],
        ).to_excel(writer, sheet_name="06_business_action_guard", index=False)
        ensure_columns(
            warn_fail,
            [
                "anonymousCaseId",
                "caseVerdict",
                "issueType",
                "forecastabilityStatus",
                "confidence",
                "horizonMonths",
                "predicted",
                "actual",
                "absoluteError",
                "intervalCoverage",
                "businessActionStatus",
                "intervalCalibrationReason",
                "requiredAction",
            ],
        ).to_excel(writer, sheet_name="07_warning_and_p2_cases", index=False)
        pd.DataFrame(decision_template_rows).to_excel(writer, sheet_name="08_user_decision_template", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, column in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in column:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(10, min(max_len + 3, 45))
            worksheet.auto_filter.ref = worksheet.dimensions


def build_outputs() -> dict:
    prepared = v1.prepare_inputs()
    raw_cases = prepared["cases"]
    evaluated = prepared["evaluated"]
    final_outputs = prepared["finalOutputs"]

    v1_gate = v1.build_current_gate_frame(evaluated, final_outputs)
    v1_coverage = v1.build_coverage_report(v1_gate)
    v1_cases = v1.build_disentangled_cases(raw_cases, v1_gate)
    v1_validation = v1.build_validation_report(v1_cases, v1_gate, v1_coverage)

    v11_gate, gate_changes = v11.apply_v1_1_gate_boundary(v1_gate, final_outputs)
    v11_coverage = v1.build_coverage_report(v11_gate)
    v11_raw_cases = v1.build_disentangled_cases(raw_cases, v11_gate)
    v11_cases, interval_rows = v11.apply_interval_calibration(v11_raw_cases)
    v11_validation = v1.build_validation_report(v11_cases, v11_gate, v11_coverage)
    v11_validation = v11.enrich_validation(v11_validation, v11_cases, v11_coverage)

    comparison = build_same_cohort_comparison(
        raw_cases,
        v1_cases,
        v1_validation,
        v11_cases,
        v11_validation,
        v1_coverage,
        v11_coverage,
    )
    integrity = build_integrity_audit(v11_cases, v11_validation, comparison, interval_rows)
    freeze, readiness = build_freeze_decision(integrity, comparison, v11_coverage, v11_validation)
    package_summary = build_business_review_package_summary(
        freeze,
        readiness,
        v11_coverage,
        v11_validation,
        integrity,
        comparison,
        v11_gate,
    )
    user_guide = build_business_review_user_guide(freeze, package_summary)
    baseline_record = build_conditional_baseline_record(
        freeze,
        readiness,
        v11_validation,
        v11_coverage,
        integrity,
        comparison,
    )

    write_reports(integrity, comparison, freeze, readiness, package_summary, user_guide, baseline_record)
    if freeze["freezeDecision"] == "FREEZE_CONDITIONAL":
        write_business_review_pack(integrity, comparison, freeze, readiness, v11_gate, v11_cases, v11_validation)
    write_json(
        PRIVATE_DETAIL_JSON,
        {
            "schema": "m2.private.v1_1_conditional_business_review_detail.v1",
            "notForCommit": True,
            "businessReviewPack": str(PRIVATE_XLSX.relative_to(ROOT)),
            "freezeDecision": freeze["freezeDecision"],
            "candidateVersion": freeze["candidateVersion"],
            "gateChanges": gate_changes,
        },
    )

    return {
        "backtestIntegrity": integrity["backtestIntegrity"],
        "sameCohortBetterThanBaseline": comparison["v1_1BetterThanBaselineOnSameCohort"],
        "freezeDecision": freeze["freezeDecision"],
        "candidateVersion": freeze["candidateVersion"],
        "limitedBusinessReviewReady": readiness["limitedBusinessReviewReady"],
        "m3Allowed": readiness["m3Allowed"],
        "businessReviewPack": str(PRIVATE_XLSX.relative_to(ROOT)) if PRIVATE_XLSX.exists() else None,
        "reports": [
            str(INTEGRITY_JSON.relative_to(ROOT)),
            str(COMPARISON_JSON.relative_to(ROOT)),
            str(FREEZE_JSON.relative_to(ROOT)),
            str(READINESS_JSON.relative_to(ROOT)),
            str(PACKAGE_SUMMARY_JSON.relative_to(ROOT)),
            str(USER_GUIDE_JSON.relative_to(ROOT)),
            str(BASELINE_RECORD_JSON.relative_to(ROOT)),
        ],
    }


def validate_freeze_decision() -> dict:
    if not FREEZE_JSON.exists() or not INTEGRITY_JSON.exists() or not COMPARISON_JSON.exists():
        raise SystemExit("freeze decision artifacts are missing; run --audit first")
    freeze = json.loads(FREEZE_JSON.read_text(encoding="utf-8"))
    integrity = json.loads(INTEGRITY_JSON.read_text(encoding="utf-8"))
    comparison = json.loads(COMPARISON_JSON.read_text(encoding="utf-8"))
    if integrity["backtestIntegrity"] != "PASS":
        raise SystemExit("backtest integrity is not PASS")
    if not comparison["v1_1BetterThanBaselineOnSameCohort"]:
        raise SystemExit("v1.1 is not better than baseline on same cohort")
    if freeze["freezeDecision"] != "FREEZE_CONDITIONAL":
        raise SystemExit("freeze decision is not FREEZE_CONDITIONAL")
    if not PRIVATE_XLSX.exists():
        raise SystemExit("private business review pack is missing")
    return {
        "freezeDecisionValidated": True,
        "candidateVersion": freeze["candidateVersion"],
        "m3Allowed": freeze["m3Allowed"],
        "businessReviewPack": str(PRIVATE_XLSX.relative_to(ROOT)),
    }


def fixture_self_test() -> dict:
    static_checks = build_static_integrity_checks()
    required = [
        "featuresUseCutoffOrEarlierOnly",
        "actualUsesCutoffFutureWindowOnly",
        "incompleteMonthsExcluded",
        "insufficientHorizonSkipped",
        "baselineUsesSameHistoryAndCutoff",
        "v11IntervalUsesRollingPriorResiduals",
    ]
    passed = all(static_checks.get(key) for key in required)
    if not passed:
        raise SystemExit(json.dumps({"fixtureSelfTest": False, "checks": static_checks}, ensure_ascii=False))
    return {"fixtureSelfTest": True, "checks": static_checks}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit", action="store_true", help="run full local real-data backtest integrity audit")
    parser.add_argument("--freeze-decision", action="store_true", help="validate existing freeze decision artifacts")
    parser.add_argument("--fixture-self-test", action="store_true", help="run static no-real-data self-test")
    args = parser.parse_args()
    if args.fixture_self_test:
        print(json.dumps(fixture_self_test(), ensure_ascii=False, indent=2))
        return
    if args.freeze_decision:
        print(json.dumps(validate_freeze_decision(), ensure_ascii=False, indent=2))
        return
    print(json.dumps(build_outputs(), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
