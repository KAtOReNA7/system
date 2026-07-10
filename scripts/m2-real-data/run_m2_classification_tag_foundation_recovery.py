from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing spreadsheet dependency. Run through scripts/run-codex-python.mjs."
    ) from exc


ROOT = Path(__file__).resolve().parents[2]
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
SOURCE_WORKBOOK = PRIVATE_DIR / "M2-classification-tag-final-foundation-table-cn-v1.xlsx"
FIXED_WORKBOOK = PRIVATE_DIR / "M2-classification-tag-foundation-local-fixed-cn-v1.xlsx"
FIXED_JSON = PRIVATE_DIR / "M2-classification-tag-foundation-local-fixed-cn-v1.json"
TAXONOMY_PATH = (
    ROOT
    / "src"
    / "domain"
    / "oldProductEvaluation"
    / "classificationTaxonomy.v1.json"
)
PUBLIC_CLOSEOUT = (
    ROOT
    / "docs"
    / "analysis"
    / "m2-real-data"
    / "M2-classification-tag-foundation-local-closeout-v1.json"
)

MAIN_SHEETS = ["01_出版物", "02_网文"]
MAIN_HEADERS = [
    "序号",
    "书名",
    "作品编号",
    "作者",
    "一级分类（最终采用）",
    "二级分类（最终采用）",
    "三级分类（最终采用）",
    "辅助标签（最终采用）",
    "备注",
]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Recover the ignored M2 fixed classification/tag JSON from an already "
            "confirmed final foundation workbook."
        )
    )
    parser.add_argument(
        "--confirm-user-message",
        action="store_true",
        help="Record the current user message as explicit whole-workbook confirmation.",
    )
    parser.add_argument(
        "--source-workbook",
        type=Path,
        default=SOURCE_WORKBOOK,
        help="Private final foundation workbook path.",
    )
    return parser.parse_args()


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def canonical_work_id(value) -> str:
    text = clean(value)
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    if not match:
        return text
    return str(int(match.group(0)))


def split_tags(value) -> list[str]:
    text = clean(value)
    if not text or text == "无":
        return []
    tags = []
    for item in re.split(r"[；;、，,|/]+", text):
        tag = clean(item)
        if tag and tag != "无" and tag not in tags:
            tags.append(tag)
    return tags


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT.resolve())).replace("\\", "/")
    except ValueError:
        return "external-private-input/" + path.name


def read_workbook(path: Path) -> tuple[list[dict], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = [name for name in [*MAIN_SHEETS, "03_整体确认"] if name not in workbook.sheetnames]
        if missing:
            raise SystemExit("Final foundation workbook is missing sheets: " + ", ".join(missing))

        records = []
        seen = set()
        for sheet_name in MAIN_SHEETS:
            sheet = workbook[sheet_name]
            headers = [
                clean(cell.value)
                for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(MAIN_HEADERS)))
            ]
            if headers != MAIN_HEADERS:
                raise SystemExit(f"Final foundation headers changed in sheet: {sheet_name}")
            for values in sheet.iter_rows(
                min_row=2, max_col=len(MAIN_HEADERS), values_only=True
            ):
                row = {
                    header: clean(values[index])
                    for index, header in enumerate(MAIN_HEADERS)
                }
                if not any(row.values()):
                    continue
                work_id = canonical_work_id(row["作品编号"])
                if not work_id:
                    raise SystemExit("Final foundation workbook contains a missing work ID.")
                if work_id in seen:
                    raise SystemExit("Final foundation workbook contains duplicate canonical work IDs.")
                seen.add(work_id)
                records.append(
                    {
                        "作品编号": work_id,
                        "书名": row["书名"],
                        "作者": row["作者"],
                        "一级分类": row["一级分类（最终采用）"],
                        "二级分类": row["二级分类（最终采用）"],
                        "三级分类": row["三级分类（最终采用）"],
                        "辅助标签": "；".join(split_tags(row["辅助标签（最终采用）"])) or "无",
                        "备注": row["备注"],
                        "固定来源": "用户最终分类标签基础大表恢复",
                    }
                )
        confirmation = clean(workbook["03_整体确认"]["C2"].value)
        return records, confirmation
    finally:
        workbook.close()


def validate(records: list[dict], taxonomy: dict) -> dict:
    allowed_tags = {
        tag
        for tags in taxonomy["auxiliaryTagGroups"].values()
        for tag in tags
    }
    quality = Counter()
    level1 = Counter()
    level2 = Counter()
    level3 = Counter()
    tags = Counter()
    tagged_works = 0

    for row in records:
        required = [
            row["作品编号"],
            row["书名"],
            row["作者"],
            row["一级分类"],
            row["二级分类"],
            row["三级分类"],
        ]
        if not all(required):
            quality["missingRequiredFields"] += 1
        branch = taxonomy["classificationTree"].get(row["一级分类"], {}).get(
            row["二级分类"], []
        )
        if row["三级分类"] not in branch:
            quality["invalidClassificationPaths"] += 1
        row_tags = split_tags(row["辅助标签"])
        if any(tag not in allowed_tags for tag in row_tags):
            quality["invalidAuxiliaryTags"] += 1
        if row_tags:
            tagged_works += 1
            tags.update(row_tags)
        level1[row["一级分类"]] += 1
        level2[f"{row['一级分类']}>{row['二级分类']}"] += 1
        level3[
            f"{row['一级分类']}>{row['二级分类']}>{row['三级分类']}"
        ] += 1

    summary = {
        "scope": {
            "workCount": len(records),
            "uniqueWorkIds": len({row["作品编号"] for row in records}),
            "level1Distribution": dict(level1),
        },
        "finalDistribution": {
            "level2": dict(level2),
            "level3": dict(level3),
            "tags": dict(tags),
            "taggedWorks": tagged_works,
            "tagAssignments": sum(tags.values()),
        },
        "quality": {
            "missingWorkIds": 0,
            "duplicateWorkIds": len(records)
            - len({row["作品编号"] for row in records}),
            "missingRequiredFields": quality["missingRequiredFields"],
            "invalidClassificationPaths": quality["invalidClassificationPaths"],
            "invalidAuxiliaryTags": quality["invalidAuxiliaryTags"],
        },
    }
    return summary


def assert_matches_closeout(summary: dict, expected: dict) -> None:
    checks = {
        "workCount": (
            summary["scope"]["workCount"],
            expected["scope"]["workCount"],
        ),
        "uniqueWorkIds": (
            summary["scope"]["uniqueWorkIds"],
            expected["scope"]["uniqueWorkIds"],
        ),
        "level1Distribution": (
            summary["scope"]["level1Distribution"],
            expected["scope"]["level1Distribution"],
        ),
        "taggedWorks": (
            summary["finalDistribution"]["taggedWorks"],
            expected["finalDistribution"]["taggedWorks"],
        ),
        "tagAssignments": (
            summary["finalDistribution"]["tagAssignments"],
            expected["finalDistribution"]["tagAssignments"],
        ),
    }
    mismatches = [
        f"{key}: actual={actual}, expected={expected_value}"
        for key, (actual, expected_value) in checks.items()
        if actual != expected_value
    ]
    if mismatches:
        raise SystemExit(
            "Final foundation workbook differs from the committed closeout: "
            + "; ".join(mismatches)
        )
    if any(summary["quality"].values()):
        raise SystemExit(
            "Final foundation workbook failed quality checks: "
            + json.dumps(summary["quality"], ensure_ascii=False)
        )


def main() -> None:
    args = parse_arguments()
    source = args.source_workbook.resolve()
    required = [source, TAXONOMY_PATH, PUBLIC_CLOSEOUT]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit("Required recovery inputs are missing: " + ", ".join(missing))

    taxonomy = json.loads(TAXONOMY_PATH.read_text(encoding="utf-8"))
    expected = json.loads(PUBLIC_CLOSEOUT.read_text(encoding="utf-8"))
    records, worksheet_confirmation = read_workbook(source)
    confirmed = (
        worksheet_confirmation == "确认作为后续基础表格"
        or args.confirm_user_message
    )
    if not confirmed:
        raise SystemExit(
            "Whole-workbook confirmation is missing. Confirm in 03_整体确认 or rerun "
            "with --confirm-user-message after explicit user confirmation."
        )

    summary = validate(records, taxonomy)
    assert_matches_closeout(summary, expected)
    recovery_summary = {
        "schema": "m2.classification_tag_foundation_private_recovery.v1",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "candidateVersion": expected["candidateVersion"],
        "taxonomyVersion": taxonomy["version"],
        "confirmation": {
            "confirmed": True,
            "source": (
                "工作表整体确认"
                if worksheet_confirmation == "确认作为后续基础表格"
                else "用户会话明确确认"
            ),
            "worksheetValue": worksheet_confirmation,
        },
        **summary,
        "closure": expected["closure"],
        "recovery": {
            "recoveredFromConfirmedWorkbook": True,
            "sourceWorkbook": relative(source),
            "sourceWorkbookSha256": file_hash(source),
            "matchedCommittedCloseout": True,
            "publicCloseout": relative(PUBLIC_CLOSEOUT),
            "databaseWritten": False,
            "formalMasterDataWritten": False,
        },
    }

    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    FIXED_JSON.write_text(
        json.dumps(
            {"summary": recovery_summary, "records": records},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    if source != FIXED_WORKBOOK.resolve():
        shutil.copy2(source, FIXED_WORKBOOK)

    print(
        json.dumps(
            {
                "status": "recovered",
                "workCount": summary["scope"]["workCount"],
                "uniqueWorkIds": summary["scope"]["uniqueWorkIds"],
                "level1Distribution": summary["scope"]["level1Distribution"],
                "taggedWorks": summary["finalDistribution"]["taggedWorks"],
                "tagAssignments": summary["finalDistribution"]["tagAssignments"],
                "quality": summary["quality"],
                "confirmationSource": recovery_summary["confirmation"]["source"],
                "matchedCommittedCloseout": True,
                "privateFixedJson": relative(FIXED_JSON),
                "privateFixedWorkbook": relative(FIXED_WORKBOOK),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
