from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it into a local temp dependency path, "
        "for example: python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


CANDIDATE_VERSION = "m2-realdata-dev-commercial-rating-suggestion-v2.0"
PREVIOUS_CANDIDATE_VERSION = "m2-realdata-dev-rating-suggestion-calibrated-v1.0"
M2_TOTAL_WORKS = 3054

PRIVATE_M2 = ROOT / "data" / "private-output" / "m2-business-review"
PRIVATE_M1 = ROOT / "data" / "private-output" / "m1-master-data"
OPS_DIR = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"

INPUT_AFTER_STAGING_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-after-dual-source-staging-v2.xlsx"
INPUT_V1_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-after-rating-suggestion-calibration-v1.xlsx"
STAGING_TABLE_JSON = PRIVATE_M1 / "M1-dual-source-limited-staging-table-v1.json"
FORECAST_OUTPUT_JSON = DOCS_DIR / "M2-forecast-output-type-after-dual-source-staging-v2.json"
RATING_V1_SUMMARY_JSON = DOCS_DIR / "M2-rating-suggestion-calibration-v1-summary.json"
SUGGESTION_V1_ROOT_JSON = DOCS_DIR / "M2-suggestion-failure-root-cause-after-staging-v1.json"
RATING_V1_ROOT_JSON = DOCS_DIR / "M2-rating-failure-root-cause-after-staging-v1.json"

PRIVATE_VALIDATION_XLSX = PRIVATE_M2 / "m2-commercial-rating-suggestion-v2.0-validation.xlsx"
PRIVATE_OPERATOR_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-commercial-rating-v2.xlsx"

COMMERCIAL_AUDIT_JSON = DOCS_DIR / "M2-commercial-terms-source-audit-v1.json"
COMMERCIAL_AUDIT_MD = DOCS_DIR / "M2-commercial-terms-source-audit-v1.md"
RATING_SUMMARY_JSON = DOCS_DIR / "M2-commercial-rating-calibration-v2-summary.json"
RATING_SUMMARY_MD = DOCS_DIR / "M2-commercial-rating-calibration-v2-summary.md"
SUGGESTION_SUMMARY_JSON = DOCS_DIR / "M2-commercial-suggestion-calibration-v2-summary.json"
SUGGESTION_SUMMARY_MD = DOCS_DIR / "M2-commercial-suggestion-calibration-v2-summary.md"
OPERATOR_SUMMARY_JSON = DOCS_DIR / "M2-operator-task-pack-commercial-rating-v2-summary.json"
OPERATOR_SUMMARY_MD = DOCS_DIR / "M2-operator-task-pack-commercial-rating-v2-summary.md"

TASK_SHEET = "01_运营任务卡"
USER_RESERVED = "用户指定作品"

RATING_RANK = {"S+": 0, "S": 1, "A": 2, "B": 3, "C": 4, "D": 5, "E": 6}
REVENUE_SCORE = {
    "top": 62,
    "high": 58,
    "medium": 46,
    "mid": 46,
    "low": 30,
    "long_tail": 24,
    "near_zero": 18,
    "zero": 12,
}
LIFECYCLE_SCORE = {
    "growth": 10,
    "rebound": 8,
    "stable": 8,
    "declining": -6,
    "inactive": -14,
    "long_tail": -16,
    "insufficient_history": -10,
}
CONFIDENCE_SCORE = {"high": 6, "medium": 2, "low": -4, "blocked_for_business_use": -10}
FORECAST_SCORE = {
    "numeric_forecast_eligible": 8,
    "conservative_numeric_forecast": 3,
    "observe_only_no_numeric_forecast": -12,
    "true_forecast_blocked": -14,
}


def main() -> None:
    ensure_inputs()
    commercial_lookup, source_audit = build_commercial_lookup()
    staging_index = build_staging_work_index()
    commercial_audit = build_commercial_audit(commercial_lookup, source_audit, staging_index)
    rows = load_operator_rows(INPUT_AFTER_STAGING_XLSX)
    v1_rows = load_v1_rows(INPUT_V1_XLSX) if INPUT_V1_XLSX.exists() else {}
    calibrated = [calibrate_operator_row(row, commercial_lookup, v1_rows) for row in rows]
    reviewable = [row for row in calibrated if row["sampleSource"] != USER_RESERVED]

    rating_summary = build_rating_summary(reviewable)
    suggestion_summary = build_suggestion_summary(reviewable)
    operator_summary = build_operator_summary(calibrated, reviewable)

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    write_json(COMMERCIAL_AUDIT_JSON, public_envelope("m2.commercial_terms_source_audit.v1", commercial_audit))
    write_text(COMMERCIAL_AUDIT_MD, commercial_audit_md(commercial_audit))
    write_json(RATING_SUMMARY_JSON, public_envelope("m2.commercial_rating_calibration_v2.summary", rating_summary))
    write_text(RATING_SUMMARY_MD, rating_summary_md(rating_summary))
    write_json(SUGGESTION_SUMMARY_JSON, public_envelope("m2.commercial_suggestion_calibration_v2.summary", suggestion_summary))
    write_text(SUGGESTION_SUMMARY_MD, suggestion_summary_md(suggestion_summary))
    write_json(OPERATOR_SUMMARY_JSON, public_envelope("m2.operator_task_pack_commercial_rating_v2.summary", operator_summary))
    write_text(OPERATOR_SUMMARY_MD, operator_summary_md(operator_summary))

    write_private_workbooks(calibrated, reviewable, commercial_audit, rating_summary, suggestion_summary)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "commercialMappedWorks": commercial_audit["m2Mapping"]["knownCommercialModelWorks"],
                "buyoutWorks": commercial_audit["m2Mapping"]["commercialModelDistribution"].get("buyout", 0),
                "unknownWorks": commercial_audit["m2Mapping"]["commercialModelDistribution"].get("unknown", 0),
                "reviewableRows": len(reviewable),
                "expiredRows": rating_summary["rightsStatusDistribution"].get("expired", 0),
                "expiredDisplayedAsOnlyE": rating_summary["expiredDisplayedAsOnlyE"],
                "suggestionsDeletedOrDowngraded": suggestion_summary["deletedOrDowngradedRows"],
                "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
                "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    missing = [
        path
        for path in [
            INPUT_AFTER_STAGING_XLSX,
            STAGING_TABLE_JSON,
            FORECAST_OUTPUT_JSON,
            RATING_V1_SUMMARY_JSON,
            RATING_V1_ROOT_JSON,
            SUGGESTION_V1_ROOT_JSON,
            ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
            ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
            ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
            ROOT / "src" / "domain" / "oldProductEvaluation" / "forecastabilityGate.js",
            ROOT / "src" / "domain" / "oldProductEvaluation" / "forecastScenarioCalibration.js",
            ROOT / "package.json",
            ROOT / ".gitignore",
        ]
        if not path.exists()
    ]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def build_commercial_lookup() -> tuple[dict[str, dict], dict]:
    source_paths = [
        OPS_DIR / "ops-confirmation-v2.3-data.json",
        OPS_DIR / "ops-confirmation-v2-data.json",
    ]
    source_path = next((path for path in source_paths if path.exists()), None)
    if not source_path:
        return {}, {
            "sourceFiles": [],
            "fieldFindings": [],
            "sourceRows": 0,
            "sourceWithCommercialTextRows": 0,
            "commercialModelDistributionInSource": {},
            "sourceGap": "未找到运营确认 JSON，无法可靠识别买断/版税/分成字段。",
        }

    data = read_json(source_path)
    target_sheet = None
    for sheet in data.get("sheets", []):
        if sheet.get("name") == "标准作品基础信息补全":
            target_sheet = sheet
            break
    rows = target_sheet.get("rows", []) if target_sheet else []
    lookup: dict[str, dict] = {}
    model_counter = Counter()
    nonempty = 0
    field_counter = Counter()

    for row in rows:
        sid = clean(row.get("标准作品ID"))
        if not sid:
            continue
        values = []
        for key in ["标签候选", "运营备注", "需要补充的字段", "补全状态"]:
            value = clean(row.get(key))
            if value:
                values.append(value)
                if has_commercial_keyword(value):
                    field_counter[key] += 1
        parsed = parse_commercial_terms(values, source="ops-confirmation:标准作品基础信息补全")
        if parsed["commercialModel"] != "unknown":
            nonempty += 1
        model_counter[parsed["commercialModel"]] += 1
        existing = lookup.get(sid)
        if not existing or confidence_rank(parsed["commercialModelConfidence"]) > confidence_rank(existing["commercialModelConfidence"]):
            lookup[sid] = parsed

    audit = {
        "sourceFiles": [
            {
                "sourceAlias": "OPS_CONFIRMATION_V2_3_JSON" if "v2.3" in source_path.name else "OPS_CONFIRMATION_V2_JSON",
                "relativePath": rel(source_path),
                "sheetOrObject": "sheets[].标准作品基础信息补全.rows",
                "usableForCommercialTerms": True,
            }
        ],
        "fieldFindings": [
            {
                "fieldName": "标签候选",
                "containsSignals": ["合作方式", "买断", "分成", "版税", "预付", "保底"],
                "nonEmptyCommercialSignalRows": int(field_counter.get("标签候选", 0)),
                "rawContractTextCommitted": False,
            },
            {
                "fieldName": "运营备注/需要补充的字段/补全状态",
                "containsSignals": ["人工备注中的商业模式线索"],
                "nonEmptyCommercialSignalRows": int(sum(v for k, v in field_counter.items() if k != "标签候选")),
                "rawContractTextCommitted": False,
            },
        ],
        "sourceRows": len(rows),
        "sourceWithCommercialTextRows": nonempty,
        "commercialModelDistributionInSource": dict(sorted(model_counter.items())),
        "sourceGap": None if nonempty else "当前精简台账没有商业模式字段，需依赖运营确认 JSON；若该 JSON 不可用则无法可靠识别买断。",
    }
    return lookup, audit


def parse_commercial_terms(values: list[str], source: str = "unknown") -> dict:
    texts = [clean(value) for value in values if clean(value)]
    models = [model_from_text(text) for text in texts]
    non_unknown = [model for model in models if model != "unknown"]
    text = "\n".join(texts)
    buyout = bool(re.search(r"买断|著作权转让|全部著作权归|版权归[^，。,；;\n]*所有", text))
    royalty = "版税" in text
    prepaid = "预付" in text
    revenue_share = bool(re.search(r"分成|收益分配|收入分配|按比例结算", text))
    guarantee = "保底" in text

    if len(set(non_unknown)) > 1:
        model = "mixed"
        confidence = "medium"
        manual = True
        reason = ["多来源或同一字段中出现冲突商业模式信号"]
    elif prepaid and royalty:
        model = "prepaid_royalty"
        confidence = "high"
        manual = False
        reason = ["字段同时包含预付和版税"]
    elif buyout and (royalty or revenue_share):
        model = "mixed"
        confidence = "medium"
        manual = True
        reason = ["买断与版税/分成信号同时出现"]
    elif buyout:
        model = "buyout"
        confidence = "high"
        manual = False
        reason = ["字段包含买断或著作权转让信号"]
    elif revenue_share:
        model = "revenue_share"
        confidence = "high"
        manual = False
        reason = ["字段包含分成/收益分配信号"]
    elif royalty:
        model = "royalty"
        confidence = "high"
        manual = False
        reason = ["字段包含版税信号"]
    elif prepaid:
        model = "mixed"
        confidence = "low"
        manual = True
        reason = ["仅出现预付信号，缺少结算上下文"]
    else:
        model = "unknown"
        confidence = "none"
        manual = True
        reason = ["未发现可用商业模式字段"]
    if guarantee:
        reason.append("字段包含保底信号，需要人工确认结算口径")
    return {
        "commercialModel": model,
        "commercialModelChinese": commercial_model_cn(model),
        "commercialModelConfidence": confidence,
        "commercialTermsSource": source,
        "commercialTermsReason": reason,
        "buyoutFlag": buyout or model == "buyout",
        "royaltyFlag": royalty or model in {"royalty", "prepaid_royalty"},
        "prepaidFlag": prepaid,
        "revenueShareFlag": revenue_share or model == "revenue_share",
        "guaranteeFlag": guarantee,
        "requiresManualCommercialReview": manual or model in {"mixed", "unknown"},
        "commercialRiskLevel": "high" if model == "mixed" else ("medium" if model == "unknown" or guarantee else "low"),
    }


def model_from_text(text: str) -> str:
    buyout = bool(re.search(r"买断|著作权转让|全部著作权归|版权归[^，。,；;\n]*所有", text))
    royalty = "版税" in text
    prepaid = "预付" in text
    revenue_share = bool(re.search(r"分成|收益分配|收入分配|按比例结算", text))
    if prepaid and royalty:
        return "prepaid_royalty"
    if buyout and (royalty or revenue_share):
        return "mixed"
    if buyout:
        return "buyout"
    if revenue_share:
        return "revenue_share"
    if royalty:
        return "royalty"
    return "unknown"


def build_staging_work_index() -> dict[str, dict]:
    records = read_json(STAGING_TABLE_JSON).get("records", [])
    works: dict[str, dict] = {}
    for record in records:
        sid = clean(record.get("standardWorkId"))
        if not sid:
            continue
        works.setdefault(sid, {"standardWorkId": sid, "totalHistoricalRevenue": 0.0, "fields": set(), "records": 0})
        works[sid]["records"] += 1
        works[sid]["fields"].add(clean(record.get("fieldName")))
        works[sid]["totalHistoricalRevenue"] = max(
            works[sid]["totalHistoricalRevenue"],
            number_or_zero(record.get("totalHistoricalRevenue")),
        )
    for value in works.values():
        value["fields"] = sorted(value["fields"])
    return works


def build_commercial_audit(commercial_lookup: dict[str, dict], source_audit: dict, staging_index: dict[str, dict]) -> dict:
    model_counter = Counter()
    revenue_by_model = Counter()
    matched = 0
    known = 0
    total_revenue = sum(item["totalHistoricalRevenue"] for item in staging_index.values())
    for sid, item in staging_index.items():
        terms = commercial_lookup.get(sid) or unknown_terms()
        if sid in commercial_lookup:
            matched += 1
        model = terms["commercialModel"]
        if model != "unknown":
            known += 1
        model_counter[model] += 1
        revenue_by_model[model] += item["totalHistoricalRevenue"]

    top_distribution = {}
    sorted_works = sorted(staging_index.values(), key=lambda item: item["totalHistoricalRevenue"], reverse=True)
    for label, pct in [("top1Percent", 0.01), ("top5Percent", 0.05), ("top10Percent", 0.10)]:
        count = max(1, math.ceil(M2_TOTAL_WORKS * pct))
        cohort = sorted_works[:count]
        counter = Counter()
        revenue_counter = Counter()
        for item in cohort:
            terms = commercial_lookup.get(item["standardWorkId"]) or unknown_terms()
            counter[terms["commercialModel"]] += 1
            revenue_counter[terms["commercialModel"]] += item["totalHistoricalRevenue"]
        top_distribution[label] = {
            "workCount": len(cohort),
            "distribution": dict(sorted(counter.items())),
            "revenueDistribution": money_counter(revenue_counter),
        }

    return {
        "candidateVersion": CANDIDATE_VERSION,
        "forecastModelChanged": False,
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "m3Entered": False,
        **source_audit,
        "m2Mapping": {
            "m2TotalWorks": M2_TOTAL_WORKS,
            "stagingRevenueMappedWorks": len(staging_index),
            "unmappedByStagingRevenueIndexWorks": M2_TOTAL_WORKS - len(staging_index),
            "commercialSourceMatchedWorks": matched,
            "knownCommercialModelWorks": known,
            "commercialCoverageRateOfM2": round(matched / M2_TOTAL_WORKS, 6),
            "knownCommercialCoverageRateOfM2": round(known / M2_TOTAL_WORKS, 6),
            "commercialCoverageRateOfStagingRevenueMapped": round(matched / len(staging_index), 6) if staging_index else 0,
            "knownCommercialCoverageRateOfStagingRevenueMapped": round(known / len(staging_index), 6) if staging_index else 0,
            "commercialModelDistribution": dict(sorted(model_counter.items())),
            "commercialRevenueDistribution": money_counter(revenue_by_model),
            "buyoutRevenueShareOfMappedRevenue": round(float(revenue_by_model["buyout"]) / total_revenue, 6) if total_revenue else 0,
            "topRevenueCommercialDistribution": top_distribution,
        },
        "safeOutputBoundary": safe_boundary(),
    }


def calibrate_operator_row(row: dict, commercial_lookup: dict[str, dict], v1_rows: dict[str, dict]) -> dict:
    sid = clean(row.get("standard_work_id"))
    sample_source = clean(row.get("样本来源"))
    if sample_source == USER_RESERVED or not sid:
        return {
            **row,
            "sampleSource": sample_source,
            "isReviewable": False,
            "commercialTerms": unknown_terms(),
            "displayRatingCode": "",
            "displayRatingExplanationCn": "",
            "suggestionType": "",
            "suggestionChinese": "",
            "suggestionDeletedOrDowngraded": False,
            "requiresManualConfirmation": True,
        }

    commercial_terms = commercial_lookup.get(sid) or unknown_terms()
    base = {
        "oldRating": clean(row.get("评级")),
        "v1Rating": clean(v1_rows.get(sid, {}).get("新评级")),
        "oldSuggestion": clean(row.get("运营建议")),
        "v1Suggestion": clean(v1_rows.get(sid, {}).get("新运营建议")),
        "revenueBucket": clean(row.get("辅助原始收入层级code")),
        "lifecycle": clean(row.get("辅助原始生命周期code")),
        "forecastabilityStatus": clean(row.get("辅助原始预测状态code")),
        "businessActionStatus": clean(row.get("辅助原始业务动作状态code")),
        "forecastOutputType": clean(row.get("辅助原始forecastOutputType")),
        "forecastConfidence": confidence_code(clean(row.get("预测置信度"))),
        "remainingCopyrightMonths": number_or_none(row.get("剩余版权月数")),
        "totalHistoricalRevenue": number_or_zero(row.get("历史总收入")),
        "last12Revenue": number_or_zero(row.get("过去12个月收入")),
        "activeMonthCount": None,
        "copyrightEndDate": clean(row.get("版权到期日期")),
        "commercialTerms": commercial_terms,
    }
    rating = calibrate_rating_py(base)
    suggestion = calibrate_suggestion_py(base, rating)
    suggestion_deleted_or_downgraded = suggestion["suggestionType"] in {"manual_review_required", "observe_only"} and base["oldSuggestion"]
    return {
        **row,
        "sampleSource": sample_source,
        "isReviewable": True,
        "commercialTerms": commercial_terms,
        "commercialModel": commercial_terms["commercialModel"],
        "commercialModelChinese": commercial_terms["commercialModelChinese"],
        "commercialModelConfidence": commercial_terms["commercialModelConfidence"],
        "buyoutFlag": commercial_terms["buyoutFlag"],
        "historicalPerformanceRating": rating["historicalPerformanceRating"],
        "currentRightsStatus": rating["currentRightsStatus"],
        "forecastValueRating": rating["forecastValueRating"],
        "operationalDecisionRating": rating["operationalDecisionRating"],
        "displayRatingCode": rating["displayRatingCode"],
        "displayRatingExplanationCn": rating["displayRatingExplanationCn"],
        "ratingRationaleCn": "；".join(rating["rationaleCn"]),
        "oldRating": base["oldRating"],
        "v1Rating": base["v1Rating"],
        "ratingChangedFromV1": base["v1Rating"] != rating["displayRatingCode"],
        "expiredButRevenuePresent": rating["expiredButRevenuePresent"],
        "requiresRightsAudit": rating["requiresRightsAudit"],
        "suggestionType": suggestion["suggestionType"],
        "suggestionChinese": suggestion["suggestionChinese"],
        "suggestionQualityLevel": suggestion["suggestionQualityLevel"],
        "suggestionEvidence": "；".join(suggestion["evidenceSignals"]),
        "whyNotOtherSuggestions": "；".join(suggestion["whyNotOtherSuggestions"]),
        "suggestionDeletedOrDowngraded": suggestion_deleted_or_downgraded,
        "requiresManualConfirmation": bool(suggestion["requiredManualChecks"] or rating["requiresRightsAudit"] or commercial_terms["requiresManualCommercialReview"]),
        "requiredManualChecks": "；".join(suggestion["requiredManualChecks"]),
        "v1Suggestion": base["v1Suggestion"],
    }


def calibrate_rating_py(base: dict) -> dict:
    historical = historical_rating(base)
    rights = rights_status(base)
    forecast = forecast_rating(base)
    operational = operational_rating(base, historical, rights, forecast)
    display_code = historical["rating"]
    display = f"历史表现 {historical['rating']}；当前版权状态：{rights['statusCn']}；预测价值 {forecast['ratingCn']}；运营决策：{operational['ratingCn']}"
    rationale = historical["rationale"] + rights["rationale"] + forecast["rationale"] + operational["rationale"]
    return {
        "historicalPerformanceRating": historical["rating"],
        "currentRightsStatus": rights["status"],
        "forecastValueRating": forecast["rating"],
        "operationalDecisionRating": operational["rating"],
        "displayRatingCode": display_code,
        "displayRatingExplanationCn": display,
        "rationaleCn": rationale,
        "expiredButRevenuePresent": rights["expiredButRevenuePresent"],
        "requiresRightsAudit": rights["requiresRightsAudit"],
    }


def historical_rating(base: dict) -> dict:
    score = REVENUE_SCORE.get(base["revenueBucket"], 42)
    score += LIFECYCLE_SCORE.get(base["lifecycle"], 0)
    if base["last12Revenue"] > 0 and base["lifecycle"] in {"stable", "growth", "rebound"}:
        score += 4
    rating = rating_from_score(score)
    return {
        "rating": rating,
        "score": round(score, 2),
        "rationale": [
            f"历史评级基于收入层级={base['revenueBucket'] or '未知'}",
            f"生命周期={base['lifecycle'] or '未知'}",
            "版权到期不直接清空历史价值评级",
        ],
    }


def rights_status(base: dict) -> dict:
    months = base["remainingCopyrightMonths"]
    end = base["copyrightEndDate"]
    if months is not None and months < 0:
        status = "expired"
    elif months is not None and months >= 0:
        status = "active"
    elif "无限" in end:
        status = "perpetual"
    elif not end:
        status = "unknown"
    else:
        status = "pending_review"
    expired_revenue = status == "expired" and (base["totalHistoricalRevenue"] > 0 or base["last12Revenue"] > 0)
    status_cn = {
        "active": "版权有效",
        "expired": "版权已到期",
        "perpetual": "无限期或长期有效",
        "unknown": "版权状态未知",
        "pending_review": "版权状态待复核",
    }.get(status, "版权状态未知")
    return {
        "status": status,
        "statusCn": status_cn,
        "expiredButRevenuePresent": expired_revenue,
        "requiresRightsAudit": status in {"expired", "unknown", "pending_review"} or expired_revenue,
        "rationale": [
            f"当前版权状态={status_cn}",
            "版权状态影响运营动作，不改写历史表现评级",
        ],
    }


def forecast_rating(base: dict) -> dict:
    status = base["forecastabilityStatus"]
    if status in {"true_forecast_blocked", "observe_only_no_numeric_forecast"}:
        return {"rating": "not_applicable", "ratingCn": "暂不适用", "rationale": ["预测价值评级不适用：当前不可输出业务可用数值预测"]}
    score = REVENUE_SCORE.get(base["revenueBucket"], 42)
    score += CONFIDENCE_SCORE.get(base["forecastConfidence"], 0)
    score += FORECAST_SCORE.get(status, 0)
    if base["remainingCopyrightMonths"] is not None and base["remainingCopyrightMonths"] <= 12:
        score -= 4
    rating = rating_from_score(score)
    return {
        "rating": rating,
        "ratingCn": rating,
        "rationale": [f"预测评级基于预测状态={status or '未知'}", f"预测置信度={base['forecastConfidence'] or '未知'}"],
    }


def operational_rating(base: dict, historical: dict, rights: dict, forecast: dict) -> dict:
    model = base["commercialTerms"]["commercialModel"]
    if rights["status"] == "expired":
        rating = "renewal_review_required"
        label = "需续约/权利复核"
        reason = "版权已到期，当前不可直接运营"
    elif rights["status"] in {"unknown", "pending_review"}:
        rating = "rights_review_required"
        label = "需权利核查"
        reason = "版权状态未知或待复核"
    elif model in {"unknown", "mixed"}:
        rating = "commercial_review_required"
        label = "需商业模式复核"
        reason = "商业模式未知或冲突"
    elif base["businessActionStatus"] == "action_blocked" or base["forecastabilityStatus"] == "true_forecast_blocked":
        rating = "manual_review_required"
        label = "需人工复核"
        reason = "业务动作或预测状态仍阻断"
    else:
        rating = "operable"
        label = "可运营"
        reason = "权利、商业模式和预测状态未形成硬阻断"
    return {
        "rating": rating,
        "ratingCn": label,
        "rationale": [f"运营决策={label}", reason, f"商业模式={model}"],
    }


def calibrate_suggestion_py(base: dict, rating: dict) -> dict:
    model = base["commercialTerms"]["commercialModel"]
    rights = rating["currentRightsStatus"]
    hist = rating["historicalPerformanceRating"]
    lifecycle = base["lifecycle"]
    forecastability = base["forecastabilityStatus"]
    value_supported = hist in {"S+", "S", "A", "B"} or base["revenueBucket"] in {"top", "high", "medium", "mid"}

    if rights == "expired":
        if value_supported:
            return suggestion("renewal_review", "先做续约价值复核和权利核查", "需人工确认", ["版权已到期", "历史价值有支撑"], base, ["确认版权是否可续约", "确认到期后收入是否合规"], "有证据但需人工确认")
        return suggestion("observe_only", "仅归档观察，暂不建议续约或运营动作", "仅供参考", ["版权已到期", "缺少价值支撑"], base, ["如业务仍关注，人工复核权利状态"], "低动作价值", deleted=True)
    if rights in {"unknown", "pending_review"}:
        return no_auto_suggestion(base, ["版权状态未知或待复核"])
    if model in {"unknown", "mixed"} or base["commercialTerms"]["requiresManualCommercialReview"]:
        return no_auto_suggestion(base, ["商业模式未知或冲突"])
    if forecastability in {"true_forecast_blocked", "observe_only_no_numeric_forecast"} or base["businessActionStatus"] in {"action_blocked", "observe_only"}:
        return no_auto_suggestion(base, ["预测或业务动作阻断"])
    if model == "buyout" and rights == "active" and lifecycle in {"stable", "growth", "rebound"} and value_supported:
        return suggestion("maintain", "维持运营，可人工评估分发加强", "可执行", ["商业模式=买断", "版权有效", "历史或预测价值有支撑"], base, ["确认买断权利范围是否覆盖当前渠道"], "有证据")
    if should_promote(base, hist, rights, model):
        return suggestion("promote", "可考虑加大推广或重点分发", "需人工确认", ["可预测", "历史或预测价值高", "生命周期支撑", "版权有效"], base, ["确认近期收入不是异常峰值"], "有证据但需人工确认")
    if base["forecastOutputType"] == "copyright_term_forecast" and base["remainingCopyrightMonths"] is not None and base["remainingCopyrightMonths"] <= 12 and value_supported:
        return suggestion("renewal_review", "版权续约价值复核", "需人工确认", ["版权期偏短", "历史或预测价值有支撑"], base, ["确认续约成本", "确认权利范围"], "有证据但需人工确认")
    if hist in {"S+", "S", "A", "B", "C"} and lifecycle in {"growth", "stable", "rebound"}:
        return suggestion("maintain", "维持当前运营", "可执行", [f"历史评级={hist}", f"生命周期={lifecycle}", "版权有效"], base, [], "有证据")
    if hist in {"B", "C", "D"} and lifecycle in {"declining", "inactive", "long_tail", "insufficient_history"} and model != "buyout" and base["revenueBucket"] in {"top", "high", "medium", "mid"}:
        return suggestion("reduce_investment", "降低增量投入，保留观察", "仅供参考", [f"历史评级={hist}", f"生命周期={lifecycle}"], base, [], "保守建议")
    if hist in {"D", "E"} and lifecycle in {"inactive", "long_tail"} and model != "buyout" and base["revenueBucket"] in {"low", "long_tail", "near_zero", "zero"}:
        return suggestion("downlist_or_suspend", "下架或暂停运营候选", "需人工确认", ["低历史价值", "低收入", "长期沉寂"], base, ["确认无近期收入", "确认无买断保留价值"], "高风险动作需人工确认")
    return no_auto_suggestion(base, ["结构化信号不足以支持自动动作"])


def should_promote(base: dict, hist_rating: str, rights: str, model: str) -> bool:
    return (
        hist_rating in {"S+", "S", "A"}
        and base["lifecycle"] in {"growth", "rebound"}
        and base["forecastabilityStatus"] == "numeric_forecast_eligible"
        and base["forecastConfidence"] in {"high", "medium", ""}
        and base["businessActionStatus"] == "action_allowed"
        and rights == "active"
        and model not in {"unknown", "mixed"}
    )


def no_auto_suggestion(base: dict, evidence: list[str]) -> dict:
    return suggestion(
        "manual_review_required",
        "暂无自动运营建议，仅建议人工复核/观察",
        "不建议自动动作",
        evidence,
        base,
        ["人工复核商业模式、权利状态和近期收入"],
        "证据不足",
        deleted=True,
    )


def suggestion(kind: str, chinese: str, actionability: str, evidence: list[str], base: dict, checks: list[str], quality: str, deleted: bool = False) -> dict:
    return {
        "suggestionType": kind,
        "suggestionChinese": chinese,
        "actionabilityLevel": actionability,
        "evidenceSignals": evidence,
        "commercialTermsImpact": commercial_impact(base["commercialTerms"]),
        "rightsImpact": rights_impact(base),
        "forecastImpact": forecast_impact(base),
        "lifecycleImpact": f"生命周期={base['lifecycle'] or '未知'}",
        "whyThisSuggestion": "基于商业模式、版权状态、历史价值、预测状态和生命周期的结构化信号",
        "whyNotOtherSuggestions": why_not_other(kind),
        "confidence": "medium" if actionability != "不建议自动动作" else "low",
        "requiredManualChecks": checks,
        "suggestionQualityLevel": quality,
        "automaticSuggestionDeleted": deleted,
    }


def build_rating_summary(reviewable: list[dict]) -> dict:
    expired_rows = [row for row in reviewable if row["currentRightsStatus"] == "expired"]
    changed = [row for row in reviewable if row["v1Rating"] and row["v1Rating"] != row["displayRatingCode"]]
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "previousCandidateVersion": PREVIOUS_CANDIDATE_VERSION,
        "reviewableRows": len(reviewable),
        "historicalPerformanceDistribution": dict(sorted(Counter(row["historicalPerformanceRating"] for row in reviewable).items())),
        "rightsStatusDistribution": dict(sorted(Counter(row["currentRightsStatus"] for row in reviewable).items())),
        "forecastValueRatingDistribution": dict(sorted(Counter(row["forecastValueRating"] for row in reviewable).items())),
        "operationalDecisionRatingDistribution": dict(sorted(Counter(row["operationalDecisionRating"] for row in reviewable).items())),
        "displayRatingDistribution": dict(sorted(Counter(row["displayRatingCode"] for row in reviewable).items())),
        "ratingChangedFromV1Rows": len(changed),
        "expiredRows": len(expired_rows),
        "expiredDisplayedAsOnlyE": sum(1 for row in expired_rows if row["displayRatingCode"] == "E"),
        "expiredButRevenuePresentRows": sum(1 for row in expired_rows if row["expiredButRevenuePresent"]),
        "rightsAuditRequiredRows": sum(1 for row in reviewable if row["requiresRightsAudit"]),
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "m3Entered": False,
        "safeOutputBoundary": safe_boundary(),
    }


def build_suggestion_summary(reviewable: list[dict]) -> dict:
    deleted = [row for row in reviewable if row["suggestionDeletedOrDowngraded"]]
    evidence = [row for row in reviewable if row["suggestionEvidence"]]
    manual = [row for row in reviewable if row["requiresManualConfirmation"]]
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "reviewableRows": len(reviewable),
        "suggestionDistribution": dict(sorted(Counter(row["suggestionType"] for row in reviewable).items())),
        "actionabilityDistribution": dict(sorted(Counter(row["suggestionQualityLevel"] for row in reviewable).items())),
        "deletedOrDowngradedRows": len(deleted),
        "evidenceBackedSuggestionRows": len(evidence),
        "manualConfirmationRows": len(manual),
        "buyoutInfluencedRows": sum(1 for row in reviewable if row["commercialModel"] == "buyout"),
        "expiredRightsInfluencedRows": sum(1 for row in reviewable if row["currentRightsStatus"] == "expired"),
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "m3Entered": False,
        "safeOutputBoundary": safe_boundary(),
    }


def build_operator_summary(calibrated: list[dict], reviewable: list[dict]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "taskRows": len(calibrated),
        "reviewableRows": len(reviewable),
        "privateWorkbookPath": rel(PRIVATE_OPERATOR_XLSX),
        "validationWorkbookPath": rel(PRIVATE_VALIDATION_XLSX),
        "containsCommercialModelColumns": True,
        "containsSplitRatingColumns": True,
        "containsSuggestionEvidenceColumns": True,
        "expiredRowsDisplayedAsOnlyE": sum(1 for row in reviewable if row["currentRightsStatus"] == "expired" and row["displayRatingCode"] == "E"),
        "buyoutRowsVisible": sum(1 for row in reviewable if row["buyoutFlag"]),
        "manualConfirmationRows": sum(1 for row in reviewable if row["requiresManualConfirmation"]),
        "privateWorkbookGitignored": True,
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "databaseWritten": False,
        "m3Entered": False,
        "safeOutputBoundary": safe_boundary(),
    }


def write_private_workbooks(calibrated: list[dict], reviewable: list[dict], audit: dict, rating: dict, suggestion_summary: dict) -> None:
    PRIVATE_M2.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_sheet(wb.active, "00_阅读说明", [["项目", "说明"], ["候选版本", CANDIDATE_VERSION], ["边界", "private，包含真实作品字段，不提交"], ["M3", "不进入"]])
    write_sheet(wb.create_sheet("01_商业模式审计"), dict_rows([
        {"项目": "M2总作品数", "值": M2_TOTAL_WORKS},
        {"项目": "已识别商业模式作品数", "值": audit["m2Mapping"]["knownCommercialModelWorks"]},
        {"项目": "买断作品数", "值": audit["m2Mapping"]["commercialModelDistribution"].get("buyout", 0)},
        {"项目": "未知作品数", "值": audit["m2Mapping"]["commercialModelDistribution"].get("unknown", 0)},
    ]))
    write_sheet(wb.create_sheet("02_旧评级问题"), dict_rows([
        {"问题": "版权到期直接显示 E", "处理": "拆分历史表现和当前版权状态"},
        {"问题": "缺商业模式", "处理": "引入买断/版税/分成/未知字段"},
    ]))
    write_sheet(wb.create_sheet("03_新评级体系"), dict_rows([
        {"评级": "历史表现评级", "说明": "反映过去收入价值，不因到期直接置 E"},
        {"评级": "当前版权状态", "说明": "active/expired/perpetual/unknown/pending_review"},
        {"评级": "预测价值评级", "说明": "仅对 forecastable cohort 输出"},
        {"评级": "运营决策级别", "说明": "结合版权、商业模式和动作状态"},
    ]))
    write_sheet(wb.create_sheet("04_版权到期作品处理"), dict_rows([
        {"规则": "高历史价值+到期", "处理": "续约价值复核/权利核查"},
        {"规则": "低历史价值+到期", "处理": "仅归档观察"},
    ]))
    write_sheet(wb.create_sheet("05_买断作品处理"), dict_rows([
        {"规则": "买断+有效+稳定收入", "处理": "维持运营，可人工评估分发加强"},
        {"规则": "买断+到期/未知", "处理": "先做权利核查"},
    ]))
    write_sheet(wb.create_sheet("06_运营建议质量门槛"), dict_rows([
        {"门槛": "无证据", "处理": "不输出强建议"},
        {"门槛": "推广", "处理": "必须同时满足可预测、高价值、生命周期支撑、权利可运营"},
        {"门槛": "下架", "处理": "仅低价值、长尾/沉寂、非买断，且需人工确认"},
    ]))
    write_sheet(wb.create_sheet("07_30样本新旧对比"), dict_rows(private_compare_rows(calibrated)))
    write_sheet(wb.create_sheet("08_建议被删除或降级的样本"), dict_rows(private_compare_rows([row for row in reviewable if row["suggestionDeletedOrDowngraded"]])))
    write_sheet(wb.create_sheet("09_仍需人工复核"), dict_rows(private_compare_rows([row for row in reviewable if row["requiresManualConfirmation"]])))
    write_sheet(wb.create_sheet("10_M4校准候选"), dict_rows(private_compare_rows([row for row in reviewable if row["requiresManualConfirmation"] or row["ratingChangedFromV1"]])))
    wb.save(PRIVATE_VALIDATION_XLSX)

    task_wb = Workbook()
    write_sheet(task_wb.active, "00_阅读说明", [["项目", "说明"], ["候选版本", CANDIDATE_VERSION], ["填写对象", "请复核商业模式、拆分评级和建议证据"]])
    write_sheet(task_wb.create_sheet("01_运营任务卡"), dict_rows(private_task_rows(calibrated)))
    write_sheet(task_wb.create_sheet("02_字段说明"), dict_rows([
        {"字段": "商业模式", "说明": "买断/版税/预付+版税/分成/混合/未知"},
        {"字段": "历史表现评级", "说明": "不因版权到期直接置 E"},
        {"字段": "当前运营决策级别", "说明": "结合版权状态、商业模式和预测可用性"},
        {"字段": "建议证据", "说明": "无证据则不输出强建议"},
    ]))
    write_sheet(task_wb.create_sheet("03_填写选项"), [["字段", "可选项"], ["评级是否合理", "合理 / 不合理 / 不确定"], ["建议是否可执行", "可执行 / 不可执行 / 不适用 / 不确定"], ["商业模式是否正确", "正确 / 不正确 / 不确定"]])
    task_wb.save(PRIVATE_OPERATOR_XLSX)


def private_compare_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "样本编号": row.get("样本编号"),
            "standard_work_id": row.get("standard_work_id"),
            "作品名": row.get("作品名"),
            "作者": row.get("作者"),
            "商业模式": row.get("commercialModelChinese"),
            "商业模式置信度": row.get("commercialModelConfidence"),
            "是否买断": "是" if row.get("buyoutFlag") else "否",
            "旧评级": row.get("评级"),
            "v1评级": row.get("v1Rating"),
            "历史表现评级": row.get("historicalPerformanceRating"),
            "当前版权状态": row.get("currentRightsStatus"),
            "预测价值评级": row.get("forecastValueRating"),
            "当前运营决策级别": row.get("operationalDecisionRating"),
            "展示评级": row.get("displayRatingCode"),
            "展示评级说明": row.get("displayRatingExplanationCn"),
            "旧建议": row.get("运营建议"),
            "v1建议": row.get("v1Suggestion"),
            "v2建议": row.get("suggestionChinese"),
            "建议质量等级": row.get("suggestionQualityLevel"),
            "建议证据": row.get("suggestionEvidence"),
            "为什么不给其他建议": row.get("whyNotOtherSuggestions"),
            "是否删除或降级自动建议": "是" if row.get("suggestionDeletedOrDowngraded") else "否",
            "是否需要人工确认": "是" if row.get("requiresManualConfirmation") else "否",
        }
        for row in rows
    ]


def private_task_rows(rows: list[dict]) -> list[dict]:
    output = []
    for row in rows:
        output.append({
            "样本编号": row.get("样本编号"),
            "样本来源": row.get("样本来源"),
            "standard_work_id": row.get("standard_work_id"),
            "raw_work_id": row.get("raw_work_id"),
            "作品名": row.get("作品名"),
            "作者": row.get("作者"),
            "商业模式": row.get("commercialModelChinese"),
            "商业模式置信度": row.get("commercialModelConfidence"),
            "是否买断": "是" if row.get("buyoutFlag") else "否",
            "历史表现评级": row.get("historicalPerformanceRating"),
            "当前版权状态": row.get("currentRightsStatus"),
            "预测价值评级": row.get("forecastValueRating"),
            "当前运营决策级别": row.get("operationalDecisionRating"),
            "展示评级": row.get("displayRatingCode"),
            "展示评级说明": row.get("displayRatingExplanationCn"),
            "建议质量等级": row.get("suggestionQualityLevel"),
            "运营建议": row.get("suggestionChinese"),
            "建议证据": row.get("suggestionEvidence"),
            "为什么不给其他建议": row.get("whyNotOtherSuggestions"),
            "是否删除自动建议": "是" if row.get("suggestionDeletedOrDowngraded") else "否",
            "是否需要人工确认": "是" if row.get("requiresManualConfirmation") else "否",
            "需要人工确认事项": row.get("requiredManualChecks"),
            "运营判断：预测是否可信": row.get("运营判断：预测是否可信"),
            "运营判断：评级是否合理": row.get("运营判断：评级是否合理"),
            "运营判断：建议是否可执行": row.get("运营判断：建议是否可执行"),
            "运营发现的问题类型": row.get("运营发现的问题类型"),
            "运营建议修正": row.get("运营建议修正"),
            "运营判断：商业模式是否正确": "",
            "运营判断：展示评级是否合理": "",
            "运营判断：v2建议是否可执行": "",
            "运营补充说明": "",
        })
    return output


def load_operator_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb[TASK_SHEET] if TASK_SHEET in wb.sheetnames else wb.worksheets[1]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    output = []
    for idx, values in enumerate(rows, start=2):
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        if any(clean(value) for value in row.values()):
            row["_rowIndex"] = idx
            output.append(row)
    return output


def load_v1_rows(path: Path) -> dict[str, dict]:
    rows = load_operator_rows(path)
    output = {}
    for row in rows:
        sid = clean(row.get("standard_work_id"))
        if sid:
            output[sid] = row
    return output


def write_sheet(ws, title_or_rows, rows: list[list] | list[dict] | None = None) -> None:
    if rows is None:
        rows = title_or_rows
    else:
        ws.title = str(title_or_rows)[:31]
    if not rows:
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            ws.append(row)


def dict_rows(rows: list[dict]) -> list[dict]:
    return rows


def commercial_audit_md(payload: dict) -> str:
    rows = [{"model": key, "count": value} for key, value in payload["m2Mapping"]["commercialModelDistribution"].items()]
    return "\n".join([
        "# M2 Commercial Terms Source Audit v1",
        "",
        f"- Candidate: `{payload['candidateVersion']}`",
        f"- M2 total works: `{payload['m2Mapping']['m2TotalWorks']}`",
        f"- Commercial source matched works: `{payload['m2Mapping']['commercialSourceMatchedWorks']}`",
        f"- Known commercial model works: `{payload['m2Mapping']['knownCommercialModelWorks']}`",
        f"- Known commercial coverage of M2: `{payload['m2Mapping']['knownCommercialCoverageRateOfM2']}`",
        f"- Buyout revenue share of mapped revenue: `{payload['m2Mapping']['buyoutRevenueShareOfMappedRevenue']}`",
        "",
        markdown_table(rows, [("model", "commercial model"), ("count", "work count")]),
        "",
        "- Reports are aggregate-only and do not include real work names, authors, channels, contract text, or raw ledger rows.",
        "- M3 entered: `false`",
    ])


def rating_summary_md(payload: dict) -> str:
    return "\n".join([
        "# M2 Commercial Rating Calibration v2 Summary",
        "",
        f"- Candidate: `{payload['candidateVersion']}`",
        f"- Reviewable rows: `{payload['reviewableRows']}`",
        f"- Rating changed from v1 rows: `{payload['ratingChangedFromV1Rows']}`",
        f"- Expired rows: `{payload['expiredRows']}`",
        f"- Expired rows displayed as only E: `{payload['expiredDisplayedAsOnlyE']}`",
        f"- Rights audit required rows: `{payload['rightsAuditRequiredRows']}`",
        "- Historical performance, current rights status, forecast value, operational decision, and display rating are split.",
        "- M3 entered: `false`",
    ])


def suggestion_summary_md(payload: dict) -> str:
    rows = [{"suggestion": key, "count": value} for key, value in payload["suggestionDistribution"].items()]
    return "\n".join([
        "# M2 Commercial Suggestion Calibration v2 Summary",
        "",
        f"- Candidate: `{payload['candidateVersion']}`",
        f"- Reviewable rows: `{payload['reviewableRows']}`",
        f"- Deleted or downgraded rows: `{payload['deletedOrDowngradedRows']}`",
        f"- Evidence-backed suggestion rows: `{payload['evidenceBackedSuggestionRows']}`",
        f"- Manual confirmation rows: `{payload['manualConfirmationRows']}`",
        "",
        markdown_table(rows, [("suggestion", "suggestion"), ("count", "count")]),
        "",
        "- No strong suggestion is emitted without structured evidence.",
        "- M3 entered: `false`",
    ])


def operator_summary_md(payload: dict) -> str:
    return "\n".join([
        "# M2 Commercial Rating v2 Operator Task Pack Summary",
        "",
        f"- Candidate: `{payload['candidateVersion']}`",
        f"- Task rows: `{payload['taskRows']}`",
        f"- Reviewable rows: `{payload['reviewableRows']}`",
        f"- Private workbook: `{payload['privateWorkbookPath']}`",
        f"- Contains commercial model columns: `{payload['containsCommercialModelColumns']}`",
        f"- Contains split rating columns: `{payload['containsSplitRatingColumns']}`",
        f"- Contains suggestion evidence columns: `{payload['containsSuggestionEvidenceColumns']}`",
        f"- Expired rows displayed as only E: `{payload['expiredRowsDisplayedAsOnlyE']}`",
        f"- Private workbook gitignored: `{payload['privateWorkbookGitignored']}`",
        "- M3 entered: `false`",
    ])


def public_envelope(schema: str, payload: dict) -> dict:
    return {
        "schema": schema,
        "generatedAt": now(),
        "currentHead": git(["rev-parse", "HEAD"]),
        "originMain": git(["rev-parse", "origin/main"]),
        "payload": payload,
        "safeOutputBoundary": safe_boundary(),
    }


def safe_boundary() -> dict:
    return {
        "sanitizedAggregateOnly": True,
        "realWorkNamesIncluded": False,
        "authorNamesIncluded": False,
        "channelNamesIncluded": False,
        "rawLedgerRowsIncluded": False,
        "rawContractTextIncluded": False,
        "standardWorkIdDetailsIncluded": False,
        "privateDetailsStoredOnlyInGitignoredOutput": True,
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "databaseWritten": False,
        "forecastModelChanged": False,
        "m3Entered": False,
    }


def unknown_terms() -> dict:
    return {
        "commercialModel": "unknown",
        "commercialModelChinese": "未知",
        "commercialModelConfidence": "none",
        "commercialTermsSource": "none",
        "commercialTermsReason": ["未找到可用商业模式字段"],
        "buyoutFlag": False,
        "royaltyFlag": False,
        "prepaidFlag": False,
        "revenueShareFlag": False,
        "guaranteeFlag": False,
        "requiresManualCommercialReview": True,
        "commercialRiskLevel": "medium",
    }


def commercial_model_cn(model: str) -> str:
    return {
        "buyout": "买断",
        "royalty": "版税",
        "prepaid_royalty": "预付+版税",
        "revenue_share": "分成",
        "mixed": "混合/冲突",
        "unknown": "未知",
    }.get(model, "未知")


def confidence_rank(value: str) -> int:
    return {"none": 0, "low": 1, "medium": 2, "high": 3}.get(value, 0)


def has_commercial_keyword(value: str) -> bool:
    return any(keyword in value for keyword in ["买断", "版税", "预付", "合作方式", "合同类型", "授权方式", "版权费用", "分成", "保底", "结算方式"])


def confidence_code(value: str) -> str:
    if "高" in value:
        return "high"
    if "中" in value:
        return "medium"
    if "低" in value:
        return "low"
    if "阻断" in value:
        return "blocked_for_business_use"
    return clean(value)


def rating_from_score(score: float) -> str:
    if score >= 96:
        return "S+"
    if score >= 88:
        return "S"
    if score >= 70:
        return "A"
    if score >= 52:
        return "B"
    if score >= 38:
        return "C"
    if score >= 24:
        return "D"
    return "E"


def commercial_impact(terms: dict) -> str:
    model = terms["commercialModel"]
    if model == "buyout":
        return "买断作品不按普通版税/续约模板处理"
    if model == "revenue_share":
        return "分成模式需关注后续结算口径"
    if model == "prepaid_royalty":
        return "预付+版税模式需关注预付回收和版税结算"
    if model == "royalty":
        return "版税模式需考虑后续结算成本"
    if model == "mixed":
        return "商业模式冲突，不能自动强动作"
    return "商业模式未知，建议人工复核"


def rights_impact(base: dict) -> str:
    if base["remainingCopyrightMonths"] is not None and base["remainingCopyrightMonths"] < 0:
        return "版权已到期"
    if base["remainingCopyrightMonths"] is None:
        return "版权状态不完整"
    return "版权期仍可支持运营"


def forecast_impact(base: dict) -> str:
    if base["forecastabilityStatus"] == "true_forecast_blocked":
        return "预测不可用，不能支撑自动动作"
    if base["forecastabilityStatus"] == "conservative_numeric_forecast":
        return "仅可保守预测"
    if base["forecastabilityStatus"] == "numeric_forecast_eligible":
        return "可数值预测"
    return "预测证据不足或仅观察"


def why_not_other(kind: str) -> list[str]:
    if kind == "promote":
        return ["不自动执行推广，仍需确认资源与非一次性收入"]
    if kind == "renewal_review":
        return ["不直接推广：权利状态需要先确认", "不下架：历史价值仍需判断"]
    if kind == "downlist_or_suspend":
        return ["必须人工确认，避免误伤高收入、稳定或买断作品"]
    if kind == "maintain":
        return ["未达到强推广门槛，也没有下架证据"]
    return ["证据不足，不输出强建议"]


def money_counter(counter: Counter) -> dict:
    return {key: round(float(value), 2) for key, value in sorted(counter.items())}


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    lines = ["| " + " | ".join(label for _, label in columns) + " |"]
    lines.append("|" + "|".join("---" for _ in columns) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |")
    return "\n".join(lines)


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text + "\n", encoding="utf-8")


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def json_safe(value):
    if isinstance(value, dict):
        return {str(key): json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [json_safe(child) for child in value]
    if isinstance(value, set):
        return sorted(value)
    return value


def number_or_none(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isfinite(number):
        return number
    return None


def number_or_zero(value) -> float:
    number = number_or_none(value)
    return float(number or 0)


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def git(args: list[str]) -> str | None:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return None


if __name__ == "__main__":
    main()
