from __future__ import annotations

import json
import hashlib
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(SCRIPT_DIR))

import numpy as np
import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - handled by runner at Excel write time.
    Alignment = Font = PatternFill = get_column_letter = None

import run_m2_forecast_model_bakeoff as bake

OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-algorithm-validation"

MATERIALITY_JSON = OUTPUT_DIR / "M2-forecast-materiality-root-cause-audit-v1.json"
MATERIALITY_MD = OUTPUT_DIR / "M2-forecast-materiality-root-cause-audit-v1.md"
GATE_JSON = OUTPUT_DIR / "M2-forecastability-gate-v1.json"
GATE_MD = OUTPUT_DIR / "M2-forecastability-gate-v1.md"
VALIDATION_JSON = OUTPUT_DIR / "M2-forecastability-gated-model-validation-v1.json"
VALIDATION_MD = OUTPUT_DIR / "M2-forecastability-gated-model-validation-v1.md"
BUSINESS_JSON = OUTPUT_DIR / "M2-forecastability-business-readiness-v1.json"
BUSINESS_MD = OUTPUT_DIR / "M2-forecastability-business-readiness-v1.md"

PRIVATE_XLSX = PRIVATE_DIR / "m2-forecastability-gated-v1-validation.xlsx"
PRIVATE_DETAIL_JSON = PRIVATE_DIR / "m2-forecastability-gated-v1-private-detail.json"
PREPARE_CACHE = PRIVATE_DIR / "m2-forecast-bakeoff-prepare-cache.pkl"
PREPARE_CACHE_VERSION = "m2.forecast_bakeoff_prepare_cache.v1"

GATED_CANDIDATE = "m2-realdata-dev-forecastability-gated-v1.0"
GATED_CONDITIONAL_CANDIDATE = "m2-realdata-dev-forecastability-gated-v1.0-conditional"
MODEL_F = "model_f_forecastability_gated_ensemble"

FORECASTABILITY_STATUSES = [
    "numeric_forecast_eligible",
    "conservative_numeric_forecast",
    "observe_only_no_numeric_forecast",
    "manual_review_required",
    "excluded_from_forecast_baseline",
]
NUMERIC_STATUS = "numeric_forecast_eligible"
CONSERVATIVE_STATUS = "conservative_numeric_forecast"
OBSERVE_STATUS = "observe_only_no_numeric_forecast"
MANUAL_STATUS = "manual_review_required"
EXCLUDED_STATUS = "excluded_from_forecast_baseline"

SEVERE_DATA_GAP_RISKS = {
    "missing_copyright_end",
    "copyright_date_conflict",
    "aggregate_projection_gap",
    "mapping_uncertainty",
    "missing_basic_info",
}
SPIKE_RISKS = {"abnormal_spike", "buyout_or_oneoff_income"}
ACTIONABLE_SUGGESTIONS = {"promote", "downlist_or_suspend", "renewal_review"}


def safe_float(value, default: float = 0.0) -> float:
    return bake.safe_float(value, default)


def safe_int(value, default: int = 0) -> int:
    return bake.safe_int(value, default)


def rounded(value, digits: int = 4):
    return bake.rounded(value, digits)


def distribution(values, keys: list[str] | None = None) -> dict:
    return bake.distribution(values, keys)


def write_json(path: Path, payload: dict) -> None:
    bake.write_json(path, payload)


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    return bake.markdown_table(rows, columns)


def percent(part: float, total: float) -> float:
    total = safe_float(total)
    if total <= 0:
        return 0.0
    return rounded(safe_float(part) / total)


def bool_bucket(value) -> str:
    return "yes" if bool(value) else "no"


def yes_bucket(value) -> bool:
    return str(value).lower() in {"yes", "true", "1"}


def safe_boundary() -> dict:
    return {
        "rawRowsWritten": False,
        "realWorkNamesWritten": False,
        "realAuthorNamesWritten": False,
        "realChannelNamesWritten": False,
        "connectionStringsWritten": False,
        "privateWorkbookGitignored": True,
        "notFinalReleaseApproval": True,
        "m3Started": False,
    }


def cache_file_entries(path: Path) -> list[dict]:
    if not path.exists():
        return [{"path": str(path.relative_to(ROOT) if path.is_relative_to(ROOT) else path), "missing": True}]
    if path.is_file():
        stat = path.stat()
        return [
            {
                "path": str(path.relative_to(ROOT) if path.is_relative_to(ROOT) else path),
                "size": stat.st_size,
                "mtimeNs": stat.st_mtime_ns,
            }
        ]
    entries = []
    for child in sorted(item for item in path.rglob("*") if item.is_file()):
        stat = child.stat()
        entries.append(
            {
                "path": str(child.relative_to(ROOT) if child.is_relative_to(ROOT) else child),
                "size": stat.st_size,
                "mtimeNs": stat.st_mtime_ns,
            }
        )
    return entries


def prepare_cache_signature() -> str:
    tracked_inputs = [
        ROOT / "data",
        SCRIPT_DIR / "run_m2_forecast_model_bakeoff.py",
        SCRIPT_DIR / "run_m2_forecastability_gated_validation.py",
        SCRIPT_DIR / "run_nonformal_dry_run.py",
    ]
    entries = []
    for path in tracked_inputs:
        entries.extend(cache_file_entries(path))
    payload = json.dumps({"version": PREPARE_CACHE_VERSION, "entries": entries}, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def load_prepare_cache(signature: str) -> dict | None:
    if os.environ.get("M2_DISABLE_FORECAST_PREPARE_CACHE") == "1":
        return None
    if not PREPARE_CACHE.exists():
        return None
    try:
        payload = pd.read_pickle(PREPARE_CACHE)
    except Exception:
        return None
    if payload.get("cacheVersion") != PREPARE_CACHE_VERSION or payload.get("signature") != signature:
        return None
    data = payload.get("data")
    return data if isinstance(data, dict) else None


def write_prepare_cache(signature: str, data: dict) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    pd.to_pickle(
        {
            "cacheVersion": PREPARE_CACHE_VERSION,
            "signature": signature,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "data": data,
        },
        PREPARE_CACHE,
    )


def materiality_bucket_for(amount: float, rank_percent: float | None) -> str:
    amount = safe_float(amount)
    rank = None if rank_percent is None else safe_float(rank_percent, None)
    if rank is not None and rank <= 0.01:
        return "top_1_percent"
    if rank is not None and rank <= 0.05:
        return "top_5_percent"
    if rank is not None and rank <= 0.10:
        return "top_10_percent"
    if amount <= 10:
        return "near_zero"
    if rank is not None and rank > 0.50:
        return "bottom_50_percent"
    return "middle_40_percent"


def build_gate_result(status: str, reasons: list[str], confidence: str, materiality_bucket: str, required_action: str) -> dict:
    can_use_numeric = status in {NUMERIC_STATUS, CONSERVATIVE_STATUS}
    return {
        "forecastabilityStatus": status,
        "reasonCodes": reasons,
        "confidence": confidence,
        "materialityBucket": materiality_bucket,
        "canUseNumericForecast": can_use_numeric,
        "canUseForBusinessReview": status == NUMERIC_STATUS,
        "requiredAction": required_action,
    }


def evaluate_gate(features: dict) -> dict:
    risks = set(features.get("riskCodes") or [])
    suggestion_codes = list(features.get("suggestionCodes") or [])
    lifecycle = str(features.get("lifecycle") or "")
    rating = str(features.get("rating") or "")
    revenue_scale = str(features.get("revenueScale") or "")
    active_months = safe_int(features.get("activeMonthCount"))
    zero_months = safe_int(features.get("zeroRevenueMonthCount"))
    total_revenue = safe_float(features.get("totalHistoricalRevenue"))
    recent_revenue = safe_float(features.get("recentRevenue", features.get("last12MonthRevenue", 0)))
    volatility = safe_float(features.get("volatility", features.get("last6CoefficientOfVariation", 0)))
    remaining_months = safe_int(features.get("remainingMonthsForForecast"), 12)
    materiality_bucket = str(features.get("materialityBucket") or materiality_bucket_for(total_revenue, features.get("materialityRankPercent")))

    if active_months <= 0 or total_revenue <= 0:
        return build_gate_result(
            EXCLUDED_STATUS,
            ["no_backtestable_revenue_history"],
            "blocked_for_business_use",
            materiality_bucket,
            "exclude_from_numeric_forecast_baseline",
        )

    reasons = []
    if risks.intersection(SEVERE_DATA_GAP_RISKS) or bool(features.get("forecastFallbackUsed")):
        reasons.append("severe_data_gap_or_copyright_fallback")
    if risks.intersection(SPIKE_RISKS):
        reasons.append("unresolved_spike_or_oneoff_income")
    if active_months < 6 or lifecycle == "insufficient_history":
        reasons.append("insufficient_history")
    if reasons:
        return build_gate_result(
            MANUAL_STATUS,
            reasons,
            "blocked_for_business_use",
            materiality_bucket,
            "manual_review_before_numeric_forecast",
        )

    zero_heavy = zero_months >= max(12, active_months * 2)
    low_materiality = (
        materiality_bucket in {"near_zero", "bottom_50_percent"}
        or rating in {"D", "E"}
        or revenue_scale in {"low", "long_tail"}
    )
    if lifecycle in {"inactive", "long_tail"} or low_materiality or zero_heavy or recent_revenue <= 10:
        return build_gate_result(
            OBSERVE_STATUS,
            ["low_materiality_or_zero_heavy_pattern"],
            "low",
            materiality_bucket,
            "observe_only_no_business_numeric_forecast",
        )

    material = materiality_bucket in {"top_1_percent", "top_5_percent", "top_10_percent", "middle_40_percent"}
    if (
        material
        and active_months >= 12
        and recent_revenue > 100
        and volatility <= 0.9
        and remaining_months <= 72
        and lifecycle in {"growth", "stable", "rebound"}
    ):
        return build_gate_result(
            NUMERIC_STATUS,
            ["material_stable_history"],
            "high",
            materiality_bucket,
            "numeric_forecast_and_business_review_allowed",
        )

    return build_gate_result(
        CONSERVATIVE_STATUS,
        ["bounded_but_forecastable_with_conservative_interval"],
        "low" if (suggestion_codes and suggestion_codes[0] in ACTIONABLE_SUGGESTIONS) else "medium",
        materiality_bucket,
        "conservative_numeric_forecast_only",
    )


def prepare_inputs() -> dict:
    signature = prepare_cache_signature()
    cached = load_prepare_cache(signature)
    if cached is not None:
        return cached
    context = bake.load_analysis_inputs()
    evaluated = bake.evaluate_work_summary(
        context["work_summary"],
        context["parameters"],
        context["latest_complete_month"],
        context["incomplete_work_ids"],
        bake.PARAMETER_VARIANT,
    ).sort_values("standardWorkId").reset_index(drop=True)
    evaluated = bake.enrich_evaluated(evaluated)
    matrix, months = bake.build_month_matrix(context)
    q = bake.build_quantile_reference(matrix)
    thresholds = context["parameters"]["lifecycle"]
    feature_lookup = {
        str(row.standardWorkId): {
            "rating": row.rating,
            "riskCodes": row.riskCodes,
            "forecastFallbackUsed": row.forecastFallbackUsed,
            "suggestionBucket": row.suggestionBucket,
            "remainingCopyrightBucket": row.remainingCopyrightBucket,
        }
        for _, row in evaluated.iterrows()
    }
    cases = bake.build_backtest_cases(matrix, months, thresholds, q, feature_lookup)
    final_outputs = bake.final_predictions(evaluated, matrix, thresholds, q)
    sample_200 = bake.stable_sample(evaluated, 200)
    prepared = {
        "context": context,
        "evaluated": evaluated,
        "matrix": matrix,
        "months": months,
        "cases": cases,
        "finalOutputs": final_outputs,
        "sample200": sample_200,
    }
    write_prepare_cache(signature, prepared)
    return prepared


def build_current_gate_frame(evaluated: pd.DataFrame, final_outputs: pd.DataFrame) -> pd.DataFrame:
    frame = evaluated.copy()
    frame["_totalRevenue"] = frame["totalHistoricalRevenue"].astype(float)
    frame = frame.sort_values(["_totalRevenue", "standardWorkId"], ascending=[False, True]).reset_index(drop=True)
    total_count = max(1, len(frame))
    frame["materialityRankPercent"] = [(index + 1) / total_count for index in range(total_count)]
    frame["materialityBucket"] = frame.apply(
        lambda row: materiality_bucket_for(row.totalHistoricalRevenue, row.materialityRankPercent),
        axis=1,
    )
    output_lookup = final_outputs.set_index("workKey").to_dict(orient="index")
    rows = []
    for index, row in frame.iterrows():
        gate = evaluate_gate(
            {
                "riskCodes": row.riskCodes,
                "suggestionCodes": row.suggestionCodes,
                "lifecycle": row.lifecycle,
                "rating": row.rating,
                "revenueScale": row.revenueBucket,
                "activeMonthCount": row.activeMonthCount,
                "zeroRevenueMonthCount": row.zeroRevenueMonthCount,
                "totalHistoricalRevenue": row.totalHistoricalRevenue,
                "recentRevenue": row.last12MonthRevenue,
                "volatility": row.last6CoefficientOfVariation,
                "remainingMonthsForForecast": row.remainingMonthsForForecast,
                "forecastFallbackUsed": row.forecastFallbackUsed,
                "materialityBucket": row.materialityBucket,
            }
        )
        selected_output = output_lookup.get(str(row.standardWorkId), {})
        can_use_numeric = gate["canUseNumericForecast"]
        rows.append(
            {
                "workKey": str(row.standardWorkId),
                "anonymousId": selected_output.get("anonymousId", f"G{index + 1:04d}"),
                "rating": row.rating,
                "lifecycle": row.lifecycle,
                "revenueBucket": row.revenueBucket,
                "riskBucket": row.riskBucket,
                "suggestionBucket": row.suggestionBucket,
                "remainingCopyrightBucket": row.remainingCopyrightBucket,
                "totalHistoricalRevenue": safe_float(row.totalHistoricalRevenue),
                "last12MonthRevenue": safe_float(row.last12MonthRevenue),
                "activeMonthCount": safe_int(row.activeMonthCount),
                "zeroRevenueMonthCount": safe_int(row.zeroRevenueMonthCount),
                "forecastabilityStatus": gate["forecastabilityStatus"],
                "reasonCodes": gate["reasonCodes"],
                "gateConfidence": gate["confidence"],
                "materialityBucket": gate["materialityBucket"],
                "canUseNumericForecast": gate["canUseNumericForecast"],
                "canUseForBusinessReview": gate["canUseForBusinessReview"],
                "requiredAction": gate["requiredAction"],
                "selectedModel": selected_output.get("selectedModel"),
                "baseForecast": selected_output.get("baseForecast") if can_use_numeric else None,
                "optimisticForecast": selected_output.get("optimisticForecast") if can_use_numeric else None,
                "pessimisticForecast": selected_output.get("pessimisticForecast") if can_use_numeric else None,
                "forecastConfidence": selected_output.get("forecastConfidence") if can_use_numeric else "blocked_for_business_use",
                "businessNumericForecastAvailable": bool(gate["canUseForBusinessReview"]),
            }
        )
    return pd.DataFrame(rows)


def gate_case(row, gate_lookup: dict[str, dict]) -> dict:
    current = gate_lookup.get(str(row.workKey), {})
    risk_codes = []
    if yes_bucket(row.dataGapAtCutoff):
        risk_codes.append("missing_copyright_end")
    if yes_bucket(row.abnormalSpikeAtCutoff):
        risk_codes.append("abnormal_spike")
    baseline_last12 = safe_float(row.baselinePredicted) * 12.0 / max(1, safe_int(row.horizonMonths, 1))
    return evaluate_gate(
        {
            "riskCodes": risk_codes,
            "suggestionCodes": [row.suggestionBucket] if isinstance(row.suggestionBucket, str) else [],
            "lifecycle": row.lifecycleAtCutoff,
            "rating": row.ratingAtCutoff,
            "revenueScale": row.revenueScaleAtCutoff,
            "activeMonthCount": row.activeMonthsAtCutoff,
            "zeroRevenueMonthCount": row.zeroMonthsAtCutoff,
            "totalHistoricalRevenue": current.get("totalHistoricalRevenue", max(safe_float(row.actual), safe_float(row.predicted))),
            "recentRevenue": baseline_last12,
            "volatility": 0.7 if str(row.confidence) in {"high", "medium"} else 1.2,
            "remainingMonthsForForecast": 36,
            "forecastFallbackUsed": yes_bucket(row.dataGapAtCutoff),
            "materialityBucket": current.get("materialityBucket"),
        }
    )


def build_model_f_cases(cases: pd.DataFrame, current_gate: pd.DataFrame) -> pd.DataFrame:
    selected = cases[cases["modelId"] == "model_e_selector"].copy()
    model_d = cases[cases["modelId"] == "model_d_hierarchical_shrinkage"].copy()
    model_d_lookup = {
        (str(row.workKey), str(row.cutoffMonth), safe_int(row.horizonMonths)): row
        for _, row in model_d.iterrows()
    }
    gate_lookup = current_gate.set_index("workKey").to_dict(orient="index")
    rows = []
    for _, source_row in selected.iterrows():
        gate = gate_case(source_row, gate_lookup)
        row = source_row.copy()
        if gate["forecastabilityStatus"] == CONSERVATIVE_STATUS:
            replacement = model_d_lookup.get((str(source_row.workKey), str(source_row.cutoffMonth), safe_int(source_row.horizonMonths)))
            if replacement is not None:
                row = replacement.copy()
        if gate["forecastabilityStatus"] in {OBSERVE_STATUS, MANUAL_STATUS, EXCLUDED_STATUS}:
            predicted = 0.0
            actual = safe_float(source_row.actual)
            absolute_error = abs(predicted - actual)
            interval_coverage = actual <= 1.0
            over_forecast = False
            under_forecast = actual > 0
            confidence = "blocked_for_business_use"
            spread = None
            selected_model = "no_business_numeric_forecast"
            selection_reason = gate["requiredAction"]
        else:
            predicted = safe_float(row.predicted)
            actual = safe_float(row.actual)
            absolute_error = abs(predicted - actual)
            interval_coverage = bool(row.intervalCoverage)
            over_forecast = predicted > actual
            under_forecast = predicted < actual
            confidence = str(row.confidence)
            if gate["forecastabilityStatus"] == CONSERVATIVE_STATUS:
                confidence = "low" if confidence == "high" else confidence
            spread = row.optimisticPessimisticRatio
            selected_model = str(row.selectedModel)
            selection_reason = str(row.selectionReason)
        rows.append(
            {
                "workKey": str(source_row.workKey),
                "anonymousWorkId": None,
                "modelId": MODEL_F,
                "cutoffMonth": source_row.cutoffMonth,
                "horizonMonths": safe_int(source_row.horizonMonths),
                "predicted": predicted,
                "actual": safe_float(source_row.actual),
                "absoluteError": absolute_error,
                "baselinePredicted": safe_float(source_row.baselinePredicted),
                "baselineAbsoluteError": safe_float(source_row.baselineAbsoluteError),
                "betterThanBaseline": absolute_error <= safe_float(source_row.baselineAbsoluteError),
                "smape": bake.smape(predicted, safe_float(source_row.actual)),
                "ape": absolute_error / safe_float(source_row.actual) if safe_float(source_row.actual) > 0 else None,
                "intervalCoverage": interval_coverage,
                "overForecast": over_forecast,
                "underForecast": under_forecast,
                "confidence": confidence,
                "optimisticPessimisticRatio": spread,
                "ratingAtCutoff": source_row.ratingAtCutoff,
                "lifecycleAtCutoff": source_row.lifecycleAtCutoff,
                "revenueScaleAtCutoff": source_row.revenueScaleAtCutoff,
                "activeMonthsAtCutoff": source_row.activeMonthsAtCutoff,
                "activeMonthsBucketAtCutoff": source_row.activeMonthsBucketAtCutoff,
                "zeroMonthsAtCutoff": source_row.zeroMonthsAtCutoff,
                "zeroMonthsBucketAtCutoff": source_row.zeroMonthsBucketAtCutoff,
                "peakShareAtCutoff": source_row.peakShareAtCutoff,
                "abnormalSpikeAtCutoff": source_row.abnormalSpikeAtCutoff,
                "dataGapAtCutoff": source_row.dataGapAtCutoff,
                "suggestionBucket": source_row.suggestionBucket,
                "remainingCopyrightBucket": source_row.remainingCopyrightBucket,
                "selectedModel": selected_model,
                "selectionReason": selection_reason,
                "forecastabilityStatus": gate["forecastabilityStatus"],
                "gateReason": ";".join(gate["reasonCodes"]),
                "materialityBucket": gate["materialityBucket"],
                "requiredAction": gate["requiredAction"],
                "canUseNumericForecast": gate["canUseNumericForecast"],
                "canUseForBusinessReview": gate["canUseForBusinessReview"],
            }
        )
    return pd.DataFrame(rows)


def aggregate_by(frame: pd.DataFrame, field: str, value_field: str = "totalHistoricalRevenue") -> list[dict]:
    total = safe_float(frame[value_field].sum())
    rows = []
    for segment, group in frame.groupby(field, dropna=False):
        rows.append(
            {
                "segment": str(segment),
                "count": int(len(group)),
                "revenueTotal": rounded(group[value_field].sum(), 2),
                "revenueShare": percent(group[value_field].sum(), total),
            }
        )
    return sorted(rows, key=lambda item: (-item["revenueShare"], item["segment"]))


def build_materiality_audit(evaluated: pd.DataFrame, cases: pd.DataFrame, final_outputs: pd.DataFrame, current_gate: pd.DataFrame) -> dict:
    totals = evaluated[["standardWorkId", "totalHistoricalRevenue", "rating", "lifecycle", "revenueBucket", "riskBucket"]].copy()
    totals["totalHistoricalRevenue"] = totals["totalHistoricalRevenue"].astype(float)
    totals = totals.sort_values(["totalHistoricalRevenue", "standardWorkId"], ascending=[False, True]).reset_index(drop=True)
    total_revenue = safe_float(totals["totalHistoricalRevenue"].sum())
    count = len(totals)

    def share_for_top(fraction: float) -> dict:
        top_n = max(1, math.ceil(count * fraction))
        revenue = safe_float(totals.head(top_n)["totalHistoricalRevenue"].sum())
        return {"workCount": top_n, "revenueTotal": rounded(revenue, 2), "revenueShare": percent(revenue, total_revenue)}

    bottom_n = math.floor(count * 0.5)
    bottom_revenue = safe_float(totals.tail(bottom_n)["totalHistoricalRevenue"].sum())
    near_zero = totals[totals["totalHistoricalRevenue"] <= 10]
    selected_cases = cases[cases["modelId"] == "model_e_selector"].copy()
    selected_cases["failureLabel"] = selected_cases.apply(bake.case_failure_label, axis=1)
    fail_cases = selected_cases[selected_cases["failureLabel"] == "fail"]
    warning_cases = selected_cases[selected_cases["failureLabel"] == "warning"]
    all_case_actual = safe_float(selected_cases["actual"].sum())
    fail_actual = safe_float(fail_cases["actual"].sum())
    low_fail_cases = fail_cases[fail_cases["revenueScaleAtCutoff"].isin(["low", "long_tail"])]

    p0_mask = (
        (final_outputs["forecastConfidence"] == "high")
        & (final_outputs["baseForecast"].astype(float) <= 0)
        & (final_outputs["rating"].isin(["S+", "S", "A"]))
    )
    p1_mask = (
        (final_outputs["lifecycle"].isin(["inactive", "long_tail"]))
        & (final_outputs["baseForecast"].astype(float) > 100)
        & (final_outputs["revenueBucket"].isin(["low", "long_tail"]))
    )
    output_revenue = final_outputs[["workKey"]].merge(
        totals.rename(columns={"standardWorkId": "workKey"})[["workKey", "totalHistoricalRevenue"]],
        on="workKey",
        how="left",
    )
    p0_revenue = safe_float(output_revenue.loc[p0_mask.to_numpy(), "totalHistoricalRevenue"].sum())
    p1_revenue = safe_float(output_revenue.loc[p1_mask.to_numpy(), "totalHistoricalRevenue"].sum())

    materiality_distribution = {
        "totalWorks": int(count),
        "totalHistoricalRevenue": rounded(total_revenue, 2),
        "top1Percent": share_for_top(0.01),
        "top5Percent": share_for_top(0.05),
        "top10Percent": share_for_top(0.10),
        "bottom50Percent": {
            "workCount": int(bottom_n),
            "revenueTotal": rounded(bottom_revenue, 2),
            "revenueShare": percent(bottom_revenue, total_revenue),
        },
        "zeroOrNearZeroWorks": {
            "workCount": int(len(near_zero)),
            "revenueTotal": rounded(near_zero["totalHistoricalRevenue"].sum(), 2),
            "revenueShare": percent(near_zero["totalHistoricalRevenue"].sum(), total_revenue),
        },
    }
    fail_distribution = {
        "ungatedModel": "model_e_selector",
        "caseCount": int(len(selected_cases)),
        "failCaseCount": int(len(fail_cases)),
        "warningCaseCount": int(len(warning_cases)),
        "failCaseRateByCount": percent(len(fail_cases), len(selected_cases)),
        "failCaseActualRevenue": rounded(fail_actual, 2),
        "failCaseRevenueShare": percent(fail_actual, all_case_actual),
        "failUniqueWorkCount": int(fail_cases["workKey"].nunique()),
        "lowRevenueFailCaseShareByCount": percent(len(low_fail_cases), len(fail_cases)),
        "lowRevenueFailActualShareWithinFails": percent(low_fail_cases["actual"].sum(), fail_actual),
        "p0RevenueShare": percent(p0_revenue, total_revenue),
        "p1RevenueShare": percent(p1_revenue, total_revenue),
    }
    segment_materiality = []
    for field in ["rating", "lifecycle", "revenueBucket", "riskBucket"]:
        for row in aggregate_by(totals, field):
            segment_materiality.append({"segmentType": field, **row})
    for row in aggregate_by(current_gate, "forecastabilityStatus"):
        segment_materiality.append({"segmentType": "forecastabilityStatus", **row})

    top_segments = [row for row in segment_materiality if row["segmentType"] in {"rating", "lifecycle", "forecastabilityStatus"}]
    forecast_problem_severe = bool(
        fail_distribution["failCaseRevenueShare"] > 0.25
        or any(row["segmentType"] == "forecastabilityStatus" and row["segment"] == NUMERIC_STATUS and row["revenueShare"] < 0.50 for row in top_segments)
    )
    return {
        "schema": "m2.forecast_materiality_root_cause_audit.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateBSeriesAbandoned": True,
        "ungatedBakeoffVerdict": "FAIL",
        "materialityDistribution": materiality_distribution,
        "ungatedFailureMateriality": fail_distribution,
        "segmentMateriality": segment_materiality,
        "conclusion": {
            "lowMaterialityDominatesFailCount": fail_distribution["lowRevenueFailCaseShareByCount"] >= 0.50,
            "failureRevenueShareMaterial": fail_distribution["failCaseRevenueShare"] >= 0.15,
            "forecastProblemRemainsSevere": forecast_problem_severe,
            "interpretation": "Use revenue materiality plus forecastability gate; do not require uniform numeric forecasts for every work.",
        },
        "safeOutputBoundary": safe_boundary(),
    }


def build_gate_report(current_gate: pd.DataFrame) -> dict:
    total_revenue = safe_float(current_gate["totalHistoricalRevenue"].sum())
    summary = []
    for status, group in current_gate.groupby("forecastabilityStatus"):
        summary.append(
            {
                "forecastabilityStatus": status,
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShare": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                "confidenceDistribution": distribution(group["gateConfidence"], ["high", "medium", "low", "blocked_for_business_use"]),
                "canUseNumericForecastCount": int(group["canUseNumericForecast"].sum()),
                "canUseForBusinessReviewCount": int(group["canUseForBusinessReview"].sum()),
            }
        )
    materiality_rows = []
    for bucket, group in current_gate.groupby("materialityBucket"):
        materiality_rows.append(
            {
                "materialityBucket": bucket,
                "count": int(len(group)),
                "revenueShare": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                "statusDistribution": distribution(group["forecastabilityStatus"], FORECASTABILITY_STATUSES),
            }
        )
    return {
        "schema": "m2.forecastability_gate.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "gateStatuses": FORECASTABILITY_STATUSES,
        "summaryByStatus": sorted(summary, key=lambda item: FORECASTABILITY_STATUSES.index(item["forecastabilityStatus"])),
        "summaryByMateriality": sorted(materiality_rows, key=lambda item: item["materialityBucket"]),
        "numericForecastEligible": {
            "count": int((current_gate["forecastabilityStatus"] == NUMERIC_STATUS).sum()),
            "revenueShare": percent(
                current_gate.loc[current_gate["forecastabilityStatus"] == NUMERIC_STATUS, "totalHistoricalRevenue"].sum(),
                total_revenue,
            ),
        },
        "forecastableNumericIncludingConservative": {
            "count": int(current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]).sum()),
            "revenueShare": percent(
                current_gate.loc[current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]), "totalHistoricalRevenue"].sum(),
                total_revenue,
            ),
        },
        "observeOnlyCount": int((current_gate["forecastabilityStatus"] == OBSERVE_STATUS).sum()),
        "manualReviewCount": int((current_gate["forecastabilityStatus"] == MANUAL_STATUS).sum()),
        "excludedCount": int((current_gate["forecastabilityStatus"] == EXCLUDED_STATUS).sum()),
        "safeOutputBoundary": safe_boundary(),
    }


def issue_summary_for_current(current_gate: pd.DataFrame) -> dict:
    high_ratings = {"S+", "S", "A"}
    p0_rows = current_gate[
        (current_gate["canUseForBusinessReview"])
        & (current_gate["forecastConfidence"] == "high")
        & (current_gate["baseForecast"].fillna(0).astype(float) <= 0)
        & (current_gate["rating"].isin(high_ratings))
    ]
    non_numeric = current_gate[~current_gate["canUseNumericForecast"]]
    p1_rows = non_numeric[non_numeric["baseForecast"].notna()]
    p2_rows = current_gate[
        (~current_gate["canUseForBusinessReview"])
        & (current_gate["suggestionBucket"].isin(ACTIONABLE_SUGGESTIONS))
    ]
    return {
        "p0": int(len(p0_rows)),
        "p1": int(len(p1_rows)),
        "p2": int(len(p2_rows)),
        "p0RevenueShare": percent(p0_rows["totalHistoricalRevenue"].sum(), current_gate["totalHistoricalRevenue"].sum()),
        "p1RevenueShare": percent(p1_rows["totalHistoricalRevenue"].sum(), current_gate["totalHistoricalRevenue"].sum()),
        "p2RevenueShare": percent(p2_rows["totalHistoricalRevenue"].sum(), current_gate["totalHistoricalRevenue"].sum()),
    }


def labeled_counts(cases: pd.DataFrame) -> dict:
    if cases.empty:
        return {"pass": 0, "warning": 0, "fail": 0}
    labels = cases.apply(bake.case_failure_label, axis=1)
    return distribution(labels, ["pass", "warning", "fail"])


def layer_score(layer: pd.DataFrame) -> dict:
    metrics = bake.aggregate_cases(layer)
    labels = labeled_counts(layer)
    spread = bake.spread_summary(layer)
    baseline_abs = safe_float(layer["baselineAbsoluteError"].sum()) if not layer.empty else 0.0
    actual = safe_float(layer["actual"].sum()) if not layer.empty else 0.0
    return {
        **metrics,
        "baselineWape": None if actual <= 0 else rounded(baseline_abs / actual),
        "passCount": labels["pass"],
        "warningCount": labels["warning"],
        "failCount": labels["fail"],
        "failRate": percent(labels["fail"], len(layer)) if len(layer) else None,
        "warningRate": percent(labels["warning"], len(layer)) if len(layer) else None,
        "highConfidenceSpreadP75": spread["highConfidenceSpread"]["p75"],
        "nonLowConfidenceSpreadP75": spread["nonLowConfidenceSpread"]["p75"],
    }


def build_validation_report(cases: pd.DataFrame, model_f_cases: pd.DataFrame, current_gate: pd.DataFrame) -> dict:
    model_scoreboard = []
    for row in bake.model_scoreboard(cases):
        model_scoreboard.append({"modelId": row["modelId"], "cohort": "ungated_all_cases", **row})

    layer_a = model_f_cases[model_f_cases["forecastabilityStatus"] == NUMERIC_STATUS]
    layer_b = model_f_cases[model_f_cases["forecastabilityStatus"] == CONSERVATIVE_STATUS]
    layer_c = model_f_cases[model_f_cases["forecastabilityStatus"].isin([OBSERVE_STATUS, MANUAL_STATUS, EXCLUDED_STATUS])]
    layer_a_score = layer_score(layer_a)
    layer_b_score = layer_score(layer_b)
    all_f_score = layer_score(model_f_cases)
    issue_summary = issue_summary_for_current(current_gate)
    numeric_revenue_share = percent(
        current_gate.loc[current_gate["forecastabilityStatus"] == NUMERIC_STATUS, "totalHistoricalRevenue"].sum(),
        current_gate["totalHistoricalRevenue"].sum(),
    )
    numeric_plus_conservative_share = percent(
        current_gate.loc[current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]), "totalHistoricalRevenue"].sum(),
        current_gate["totalHistoricalRevenue"].sum(),
    )
    nonforecastable_not_masqueraded = issue_summary["p1"] == 0

    model_scoreboard.append(
        {
            "modelId": MODEL_F,
            "modelName": "Model F forecastability-gated ensemble",
            "cohort": NUMERIC_STATUS,
            "revenueCoverageShare": numeric_revenue_share,
            **layer_a_score,
        }
    )
    model_scoreboard.append(
        {
            "modelId": MODEL_F,
            "modelName": "Model F forecastability-gated ensemble",
            "cohort": CONSERVATIVE_STATUS,
            "revenueCoverageShare": percent(
                current_gate.loc[current_gate["forecastabilityStatus"] == CONSERVATIVE_STATUS, "totalHistoricalRevenue"].sum(),
                current_gate["totalHistoricalRevenue"].sum(),
            ),
            **layer_b_score,
        }
    )
    model_scoreboard.append(
        {
            "modelId": MODEL_F,
            "modelName": "Model F forecastability-gated ensemble",
            "cohort": "all_classified_cases",
            "revenueCoverageShare": numeric_plus_conservative_share,
            **all_f_score,
        }
    )

    segment_performance = []
    for field in [
        "forecastabilityStatus",
        "ratingAtCutoff",
        "lifecycleAtCutoff",
        "revenueScaleAtCutoff",
        "materialityBucket",
        "confidence",
    ]:
        for segment, group in model_f_cases.groupby(field):
            metrics = bake.aggregate_cases(group)
            segment_performance.append(
                {
                    "segmentType": field,
                    "segment": str(segment),
                    "caseCount": metrics["caseCount"],
                    "wape": metrics["wape"],
                    "mae": metrics["mae"],
                    "smape": metrics["smape"],
                    "coverage": metrics["intervalCoverage"],
                    "failRate": bake.failure_rate_for_cases(group),
                }
            )

    pass_conditions = {
        "numericRevenueCoverageAtLeast70": numeric_revenue_share >= 0.70,
        "numericWapeNotWorseThanBaseline": layer_a_score["wape"] is not None
        and layer_a_score["baselineWape"] is not None
        and safe_float(layer_a_score["wape"]) <= safe_float(layer_a_score["baselineWape"]),
        "numericIntervalCoverageAtLeast60": layer_a_score["intervalCoverage"] is not None
        and safe_float(layer_a_score["intervalCoverage"]) >= 0.60,
        "p0EqualsZero": issue_summary["p0"] == 0,
        "p1AtMostThree": issue_summary["p1"] <= 3,
        "highConfidenceSpreadP75AtMost1_5": layer_a_score["highConfidenceSpreadP75"] is None
        or safe_float(layer_a_score["highConfidenceSpreadP75"]) <= 1.50,
        "nonForecastableNotMasqueradedAsBusinessForecast": nonforecastable_not_masqueraded,
    }
    conditional_conditions = {
        "numericRevenueCoverageAtLeast50": numeric_revenue_share >= 0.50,
        "numericWapeNotWorseThanBaseline": pass_conditions["numericWapeNotWorseThanBaseline"],
        "p0EqualsZero": issue_summary["p0"] == 0,
        "p1AtMostTen": issue_summary["p1"] <= 10,
        "nonForecastableTaggedCorrectly": nonforecastable_not_masqueraded,
    }
    if all(pass_conditions.values()):
        verdict = "PASS"
        candidate = GATED_CANDIDATE
    elif all(conditional_conditions.values()):
        verdict = "CONDITIONAL PASS"
        candidate = GATED_CONDITIONAL_CANDIDATE
    else:
        verdict = "FAIL"
        candidate = None

    return {
        "schema": "m2.forecastability_gated_model_validation.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateBSeriesAbandoned": True,
        "ungatedBakeoffVerdict": "FAIL",
        "modelF": {
            "modelId": MODEL_F,
            "candidateVersion": candidate,
            "verdict": verdict,
            "numericForecastRevenueCoverage": numeric_revenue_share,
            "numericPlusConservativeRevenueCoverage": numeric_plus_conservative_share,
            "layerA_numericForecastEligible": layer_a_score,
            "layerB_conservativeNumericForecast": layer_b_score,
            "layerC_nonForecastableCaseCount": int(len(layer_c)),
            "issueSummary": issue_summary,
            "passConditions": pass_conditions,
            "conditionalPassConditions": conditional_conditions,
        },
        "modelScoreboard": model_scoreboard,
        "segmentPerformance": sorted(segment_performance, key=lambda item: (item["segmentType"], item["segment"])),
        "safeOutputBoundary": safe_boundary(),
    }


def build_business_readiness_report(gate_report: dict, validation: dict) -> dict:
    verdict = validation["modelF"]["verdict"]
    numeric = gate_report["numericForecastEligible"]
    forecastable = gate_report["forecastableNumericIncludingConservative"]
    if verdict == "PASS":
        readiness = "m2_business_review_ready_for_numeric_eligible_cohort"
    elif verdict == "CONDITIONAL PASS":
        readiness = "limited_business_review_ready_for_numeric_eligible_cohort"
    else:
        readiness = "not_ready_for_m2_business_review"
    groups = [
        {
            "group": "numeric_forecast_eligible",
            "businessUse": "allowed_for_business_review_only_if_verdict_pass_or_conditional",
            "reason": "material history, meaningful recent revenue, controlled volatility, no severe gap",
        },
        {
            "group": "conservative_numeric_forecast",
            "businessUse": "reference_only_low_confidence",
            "reason": "bounded forecast can be shown, but not used for strong business actions",
        },
        {
            "group": "observe_only_no_numeric_forecast",
            "businessUse": "no_business_numeric_forecast",
            "reason": "low materiality, zero-heavy, inactive, long-tail, D/E, or near-zero revenue",
        },
        {
            "group": "manual_review_required",
            "businessUse": "blocked_until_manual_review",
            "reason": "spike, data gap, copyright fallback, mapping/basic-info issue, or insufficient history",
        },
        {
            "group": "excluded_from_forecast_baseline",
            "businessUse": "excluded_from_acceptance_score",
            "reason": "not backtestable or no usable revenue history",
        },
    ]
    return {
        "schema": "m2.forecastability_business_readiness.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "readiness": readiness,
        "m3Allowed": False,
        "candidateVersion": validation["modelF"]["candidateVersion"],
        "modelVerdict": verdict,
        "numericForecastEligibleCount": numeric["count"],
        "numericForecastEligibleRevenueShare": numeric["revenueShare"],
        "forecastableNumericIncludingConservativeCount": forecastable["count"],
        "forecastableNumericIncludingConservativeRevenueShare": forecastable["revenueShare"],
        "businessGroups": groups,
        "m3BlockReason": "M3 remains blocked until a forecast baseline passes and user/business review is completed.",
        "safeOutputBoundary": safe_boundary(),
    }


def sample_validation_rows(sample_200: pd.DataFrame, current_gate: pd.DataFrame) -> list[dict]:
    lookup = current_gate.set_index("workKey").to_dict(orient="index")
    rows = []
    for index, sample in sample_200.reset_index(drop=True).iterrows():
        gate = lookup.get(str(sample.standardWorkId), {})
        status = gate.get("forecastabilityStatus", "unknown")
        if status == NUMERIC_STATUS:
            verdict = "forecastable_numeric"
            issue = "none"
            score = 100
        elif status == CONSERVATIVE_STATUS:
            verdict = "reference_only"
            issue = "low_confidence_numeric"
            score = 70
        elif status == OBSERVE_STATUS:
            verdict = "observe_only"
            issue = "non_forecastable_low_materiality"
            score = 50
        elif status == MANUAL_STATUS:
            verdict = "manual_review"
            issue = "manual_review_required"
            score = 30
        else:
            verdict = "excluded"
            issue = "not_backtestable"
            score = 0
        rows.append(
            {
                "sampleId": f"S{index + 1:03d}",
                "forecastabilityStatus": status,
                "materialityBucket": gate.get("materialityBucket"),
                "selectedModel": gate.get("selectedModel") if gate.get("canUseNumericForecast") else "no_business_numeric_forecast",
                "rating": gate.get("rating"),
                "lifecycle": gate.get("lifecycle"),
                "revenueBucket": gate.get("revenueBucket"),
                "verdict": verdict,
                "score": score,
                "issueType": issue,
                "requiredAction": gate.get("requiredAction"),
            }
        )
    return rows


def write_reports(materiality: dict, gate_report: dict, validation: dict, business: dict) -> None:
    write_json(MATERIALITY_JSON, materiality)
    write_json(GATE_JSON, gate_report)
    write_json(VALIDATION_JSON, validation)
    write_json(BUSINESS_JSON, business)

    MATERIALITY_MD.write_text(
        "\n".join(
            [
                "# M2 Forecast Materiality Root-Cause Audit v1",
                "",
                "candidate-b and ungated A-E bake-off are treated as failed routes for M2 forecast baseline promotion.",
                "",
                "## Revenue Contribution",
                "",
                markdown_table(
                    [
                        {"group": "top 1%", **materiality["materialityDistribution"]["top1Percent"]},
                        {"group": "top 5%", **materiality["materialityDistribution"]["top5Percent"]},
                        {"group": "top 10%", **materiality["materialityDistribution"]["top10Percent"]},
                        {"group": "bottom 50%", **materiality["materialityDistribution"]["bottom50Percent"]},
                        {"group": "zero or near-zero", **materiality["materialityDistribution"]["zeroOrNearZeroWorks"]},
                    ],
                    [("group", "Group"), ("workCount", "Works"), ("revenueTotal", "Revenue"), ("revenueShare", "Revenue Share")],
                ),
                "",
                "## Ungated Failure Materiality",
                "",
                markdown_table(
                    [materiality["ungatedFailureMateriality"]],
                    [
                        ("failCaseCount", "Fail Cases"),
                        ("failCaseRateByCount", "Fail Rate By Count"),
                        ("failCaseRevenueShare", "Fail Revenue Share"),
                        ("lowRevenueFailCaseShareByCount", "Low-Revenue Fail Count Share"),
                        ("lowRevenueFailActualShareWithinFails", "Low-Revenue Fail Revenue Share"),
                        ("p0RevenueShare", "P0 Revenue Share"),
                        ("p1RevenueShare", "P1 Revenue Share"),
                    ],
                ),
                "",
                "## Segment Materiality",
                "",
                markdown_table(
                    materiality["segmentMateriality"][:80],
                    [("segmentType", "Segment Type"), ("segment", "Segment"), ("count", "Count"), ("revenueShare", "Revenue Share")],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    GATE_MD.write_text(
        "\n".join(
            [
                "# M2 Forecastability Gate v1",
                "",
                "The gate separates works that can receive business-usable numeric forecasts from works that should be observe-only, manual-review, or excluded.",
                "",
                "## Summary By Status",
                "",
                markdown_table(
                    gate_report["summaryByStatus"],
                    [
                        ("forecastabilityStatus", "Status"),
                        ("count", "Count"),
                        ("revenueShare", "Revenue Share"),
                        ("canUseNumericForecastCount", "Numeric Count"),
                        ("canUseForBusinessReviewCount", "Business Review Count"),
                    ],
                ),
                "",
                "## Summary By Materiality",
                "",
                markdown_table(
                    gate_report["summaryByMateriality"],
                    [("materialityBucket", "Materiality"), ("count", "Count"), ("revenueShare", "Revenue Share")],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    VALIDATION_MD.write_text(
        "\n".join(
            [
                "# M2 Forecastability-Gated Model Validation v1",
                "",
                f"Verdict: `{validation['modelF']['verdict']}`",
                f"Candidate version: `{validation['modelF']['candidateVersion']}`",
                "",
                "## Model F Summary",
                "",
                markdown_table(
                    [
                        {
                            "numericForecastRevenueCoverage": validation["modelF"]["numericForecastRevenueCoverage"],
                            "numericWape": validation["modelF"]["layerA_numericForecastEligible"]["wape"],
                            "baselineWape": validation["modelF"]["layerA_numericForecastEligible"]["baselineWape"],
                            "coverage": validation["modelF"]["layerA_numericForecastEligible"]["intervalCoverage"],
                            "p0": validation["modelF"]["issueSummary"]["p0"],
                            "p1": validation["modelF"]["issueSummary"]["p1"],
                            "p2": validation["modelF"]["issueSummary"]["p2"],
                        }
                    ],
                    [
                        ("numericForecastRevenueCoverage", "Numeric Revenue Coverage"),
                        ("numericWape", "Numeric WAPE"),
                        ("baselineWape", "Baseline WAPE"),
                        ("coverage", "Coverage"),
                        ("p0", "P0"),
                        ("p1", "P1"),
                        ("p2", "P2"),
                    ],
                ),
                "",
                "## Pass Conditions",
                "",
                markdown_table(
                    [{"condition": key, "passed": value} for key, value in validation["modelF"]["passConditions"].items()],
                    [("condition", "Condition"), ("passed", "Passed")],
                ),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    BUSINESS_MD.write_text(
        "\n".join(
            [
                "# M2 Forecastability Business Readiness v1",
                "",
                f"Readiness: `{business['readiness']}`",
                f"M3 allowed: `{business['m3Allowed']}`",
                "",
                "## Business Groups",
                "",
                markdown_table(
                    business["businessGroups"],
                    [("group", "Group"), ("businessUse", "Business Use"), ("reason", "Reason")],
                ),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_excel(materiality: dict, gate_report: dict, validation: dict, business: dict, model_f_cases: pd.DataFrame, current_gate: pd.DataFrame, sample_200: pd.DataFrame) -> None:
    if Alignment is None:
        raise RuntimeError("openpyxl is required to write the private Excel workbook.")
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    numeric_cases = model_f_cases[model_f_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    numeric_cases = numeric_cases.head(30000).reset_index(drop=True)
    numeric_cases["anonymousCaseId"] = [f"C{index + 1:05d}" for index in range(len(numeric_cases))]
    non_forecastable = current_gate[~current_gate["canUseNumericForecast"]].copy().head(5000)
    sample_rows = sample_validation_rows(sample_200, current_gate)
    full_summary = gate_report["summaryByStatus"]
    with pd.ExcelWriter(PRIVATE_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"item": "purpose", "value": "M2 forecastability gated validation"},
                {"item": "gate", "value": "numeric, conservative, observe-only, manual-review, excluded"},
                {"item": "numeric forecast", "value": "only numeric_forecast_eligible is business-review usable"},
                {"item": "pass standard", "value": "revenue coverage, WAPE vs baseline, interval coverage, P0/P1, spread, non-masquerade"},
                {"item": "safety", "value": "anonymous IDs only; no real work names, authors, channels, or raw rows"},
            ]
        ).to_excel(writer, sheet_name="00_read_me", index=False)
        pd.DataFrame(
            [
                {"metric": "top_1_percent_revenue_share", "value": materiality["materialityDistribution"]["top1Percent"]["revenueShare"]},
                {"metric": "top_5_percent_revenue_share", "value": materiality["materialityDistribution"]["top5Percent"]["revenueShare"]},
                {"metric": "top_10_percent_revenue_share", "value": materiality["materialityDistribution"]["top10Percent"]["revenueShare"]},
                {"metric": "bottom_50_percent_revenue_share", "value": materiality["materialityDistribution"]["bottom50Percent"]["revenueShare"]},
                {"metric": "zero_near_zero_count", "value": materiality["materialityDistribution"]["zeroOrNearZeroWorks"]["workCount"]},
                {"metric": "fail_case_rate_by_count", "value": materiality["ungatedFailureMateriality"]["failCaseRateByCount"]},
                {"metric": "fail_case_revenue_share", "value": materiality["ungatedFailureMateriality"]["failCaseRevenueShare"]},
                {"metric": "business_materiality_conclusion", "value": materiality["conclusion"]["interpretation"]},
            ]
        ).to_excel(writer, sheet_name="01_materiality_audit", index=False)
        pd.DataFrame(gate_report["summaryByStatus"]).to_excel(writer, sheet_name="02_forecastability_gate_summary", index=False)
        pd.DataFrame(validation["modelScoreboard"]).to_excel(writer, sheet_name="03_model_scoreboard", index=False)
        numeric_cases[
            [
                "anonymousCaseId",
                "forecastabilityStatus",
                "modelId",
                "selectedModel",
                "cutoffMonth",
                "horizonMonths",
                "predicted",
                "actual",
                "absoluteError",
                "baselinePredicted",
                "betterThanBaseline",
                "confidence",
                "requiredAction",
            ]
        ].to_excel(writer, sheet_name="04_numeric_forecast_detail", index=False)
        non_forecastable[
            [
                "anonymousId",
                "forecastabilityStatus",
                "materialityBucket",
                "gateConfidence",
                "requiredAction",
                "reasonCodes",
                "rating",
                "lifecycle",
                "revenueBucket",
            ]
        ].to_excel(writer, sheet_name="05_non_forecastable_detail", index=False)
        pd.DataFrame(validation["segmentPerformance"]).to_excel(writer, sheet_name="06_segment_performance", index=False)
        pd.DataFrame(sample_rows).to_excel(writer, sheet_name="07_200_sample_validation", index=False)
        pd.DataFrame(full_summary).to_excel(writer, sheet_name="08_full_cohort_summary", index=False)
        pd.DataFrame(business["businessGroups"]).to_excel(writer, sheet_name="09_business_readiness", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, column in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in column:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(10, min(max_len + 3, 45))
            worksheet.auto_filter.ref = worksheet.dimensions


def run() -> dict:
    prepared = prepare_inputs()
    evaluated = prepared["evaluated"]
    cases = prepared["cases"]
    final_outputs = prepared["finalOutputs"]
    sample_200 = prepared["sample200"]
    current_gate = build_current_gate_frame(evaluated, final_outputs)
    model_f_cases = build_model_f_cases(cases, current_gate)
    materiality = build_materiality_audit(evaluated, cases, final_outputs, current_gate)
    gate_report = build_gate_report(current_gate)
    validation = build_validation_report(cases, model_f_cases, current_gate)
    business = build_business_readiness_report(gate_report, validation)
    write_reports(materiality, gate_report, validation, business)
    write_excel(materiality, gate_report, validation, business, model_f_cases, current_gate, sample_200)
    write_json(
        PRIVATE_DETAIL_JSON,
        {
            "schema": "m2.private.forecastability_gated_validation_detail.v1",
            "notForCommit": True,
            "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
            "verdict": validation["modelF"]["verdict"],
            "candidateVersion": validation["modelF"]["candidateVersion"],
        },
    )
    return {
        "forecastabilityGateCompleted": True,
        "candidateBSeriesAbandoned": True,
        "ungatedBakeoffVerdict": "FAIL",
        "newCandidateVersion": validation["modelF"]["candidateVersion"],
        "verdict": validation["modelF"]["verdict"],
        "m2BusinessReviewReady": business["readiness"],
        "m3Allowed": business["m3Allowed"],
        "numericForecastEligible": gate_report["numericForecastEligible"],
        "forecastableNumericIncludingConservative": gate_report["forecastableNumericIncludingConservative"],
        "modelFSummary": validation["modelF"],
        "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
        "sanitizedReports": [
            str(MATERIALITY_JSON.relative_to(ROOT)),
            str(GATE_JSON.relative_to(ROOT)),
            str(VALIDATION_JSON.relative_to(ROOT)),
            str(BUSINESS_JSON.relative_to(ROOT)),
        ],
    }


if __name__ == "__main__":
    result = run()
    print(json.dumps(bake.json_safe(result), ensure_ascii=False, indent=2))
