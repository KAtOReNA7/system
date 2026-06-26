from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


INPUT_XLSX = ROOT / "data" / "private-output" / "m2-business-review" / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v2.xlsx"
OUTPUT_JSON = ROOT / "docs" / "analysis" / "m2-real-data" / "M2-rating-standard-v2-operator-validation-summary-v1.json"
OUTPUT_MD = ROOT / "docs" / "analysis" / "m2-real-data" / "M2-rating-standard-v2-operator-validation-summary-v1.md"

RATING_ORDER = ["S+", "S", "A", "B", "C", "D", "E"]


def main() -> None:
    if not INPUT_XLSX.exists():
        raise SystemExit(f"Missing required input workbook: {relative(INPUT_XLSX)}")

    rows = read_operator_rows(INPUT_XLSX)
    summary = build_summary(rows)
    write_json(OUTPUT_JSON, envelope(summary))
    write_text(OUTPUT_MD, render_markdown(summary))

    print(
        json.dumps(
            {
                "inputWorkbook": relative(INPUT_XLSX),
                "rows": summary["sampleRows"],
                "feedbackCompletionRate": summary["feedbackCompletionRate"],
                "revenueModelFeedback": summary["feedbackSummary"]["revenueModelReasonable"],
                "ratingFeedback": summary["feedbackSummary"]["ratingReasonable"],
                "suggestionFeedback": summary["feedbackSummary"]["suggestionActionable"],
                "ratingThresholdMismatchCount": summary["ratingThresholdCheck"]["mismatchCount"],
                "m4CalibrationCandidateCount": summary["m4CalibrationCandidates"]["count"],
                "baselineDecision": summary["baselineDecision"],
                "outputJson": relative(OUTPUT_JSON),
                "outputMarkdown": relative(OUTPUT_MD),
            },
            ensure_ascii=False,
        )
    )


def read_operator_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["01_运营任务卡"] if "01_运营任务卡" in workbook.sheetnames else workbook.worksheets[-1]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = []
    for index, values in enumerate(iterator, start=1):
        if not any(clean(value) for value in values):
            continue
        source = {headers[pos]: values[pos] if pos < len(values) else "" for pos in range(len(headers))}
        rows.append(
            {
                "anonymousId": f"RSV2-{index:03d}",
                "revenueModel": clean(source.get("收入模式")),
                "salesRevenue12m": to_number(source.get("实销12月收入")),
                "annualizedSalesRevenue": to_number(source.get("年化实销收入")),
                "buyoutEstimatedAmount": to_number(source.get("买断估计金额")),
                "salesRating": clean(source.get("实销评级")),
                "buyoutHistoricalValueRating": clean(source.get("买断历史价值评级")),
                "historicalPerformanceRating": clean(source.get("历史表现评级")),
                "currentRightsStatus": clean(source.get("当前版权状态")),
                "shelfStatus": clean(source.get("下架状态")),
                "forecastValueRating": clean(source.get("预测价值评级")),
                "operationalRating": clean(source.get("当前运营评级")),
                "displayRating": clean(source.get("展示评级")),
                "suggestion": clean(source.get("运营建议")),
                "reviewPrompt": clean(source.get("复核提示")),
                "suggestionEvidence": clean(source.get("建议证据")),
                "manualConfirmation": clean(source.get("是否需要人工确认")),
                "revenueModelFeedback": clean(source.get("用户反馈：收入模式是否合理")),
                "ratingFeedback": clean(source.get("用户反馈：评级是否合理")),
                "suggestionFeedback": clean(source.get("用户反馈：建议是否可执行")),
                "issueNote": clean(source.get("用户反馈：问题说明")),
            }
        )
    return rows


def build_summary(rows: list[dict]) -> dict:
    feedback_rows = [row for row in rows if has_feedback(row)]
    revenue_model_feedback = summarize_feedback(row["revenueModelFeedback"] for row in rows)
    rating_feedback = summarize_feedback(row["ratingFeedback"] for row in rows)
    suggestion_feedback = summarize_feedback(row["suggestionFeedback"] for row in rows)
    issue_categories = categorize_issues(rows)
    rating_threshold_check = check_rating_thresholds(rows)
    buyout_summary = summarize_buyout(rows)
    expired_summary = summarize_expired(rows)
    suggestion_summary = summarize_suggestions(rows, issue_categories)
    m4_candidates = identify_m4_candidates(rows, rating_threshold_check["mismatchAnonymousIds"])
    baseline_decision = decide_baseline(
        rows=rows,
        feedback_rows=feedback_rows,
        revenue_model_feedback=revenue_model_feedback,
        rating_feedback=rating_feedback,
        suggestion_feedback=suggestion_feedback,
        rating_threshold_check=rating_threshold_check,
        suggestion_summary=suggestion_summary,
    )

    return {
        "inputWorkbook": relative(INPUT_XLSX),
        "onlyReadRatingStandardV2Workbook": True,
        "formalMasterDataWritten": False,
        "m3Entered": False,
        "sampleRows": len(rows),
        "feedbackRows": len(feedback_rows),
        "feedbackCompletionRate": rate(len(feedback_rows), len(rows)),
        "feedbackSummary": {
            "revenueModelReasonable": revenue_model_feedback,
            "ratingReasonable": rating_feedback,
            "suggestionActionable": suggestion_feedback,
        },
        "revenueModelDistribution": dict(Counter(row["revenueModel"] for row in rows)),
        "shelfStatusDistribution": dict(Counter(row["shelfStatus"] for row in rows)),
        "salesRatingDistribution": ordered_distribution(rows, "salesRating"),
        "historicalRatingDistribution": ordered_distribution(rows, "historicalPerformanceRating"),
        "buyoutSummary": buyout_summary,
        "expiredRightsSummary": expired_summary,
        "ratingThresholdCheck": rating_threshold_check,
        "suggestionSummary": suggestion_summary,
        "issueCategoryCounts": dict(issue_categories),
        "m4CalibrationCandidates": m4_candidates,
        "baselineDecision": baseline_decision,
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
    }


def summarize_feedback(values) -> dict:
    counter = Counter()
    normalized = Counter()
    for value in values:
        text = clean(value)
        if not text:
            continue
        counter[text] += 1
        normalized[classify_feedback(text)] += 1
    total = sum(normalized.values())
    return {
        "filled": total,
        "positive": int(normalized.get("positive", 0)),
        "negative": int(normalized.get("negative", 0)),
        "uncertain": int(normalized.get("uncertain", 0)),
        "other": int(normalized.get("other", 0)),
        "positiveRate": rate(normalized.get("positive", 0), total),
        "negativeRate": rate(normalized.get("negative", 0), total),
        "rawValueDistribution": dict(counter),
    }


def classify_feedback(text: str) -> str:
    lowered = text.lower()
    positive_phrases = ["合理", "正确", "符合", "可以", "可执行", "接受", "认可", "ok", "通过", "没问题", "无问题"]
    negative_phrases = ["不合理", "不正确", "不符合", "不可以", "不可执行", "错误", "不准", "偏高", "偏低", "需改", "需要修改", "模板", "无启发", "没有意义"]
    uncertain_phrases = ["不确定", "待确认", "需人工", "存疑", "看情况"]
    if any(phrase in lowered for phrase in ["没问题", "无问题"]):
        return "positive"
    if any(phrase in lowered for phrase in negative_phrases):
        return "negative"
    if any(phrase in lowered for phrase in uncertain_phrases):
        return "uncertain"
    if any(phrase in lowered for phrase in positive_phrases):
        return "positive"
    return "other"


def categorize_issues(rows: list[dict]) -> Counter:
    counter = Counter()
    for row in rows:
        note = row["issueNote"]
        if not note:
            continue
        categories = issue_categories(note)
        for category in categories:
            counter[category] += 1
    return counter


def issue_categories(text: str) -> list[str]:
    categories = []
    checks = [
        ("收入模式/买断实销拆分", ["收入", "买断", "实销", "拆分", "模式", "分成"]),
        ("货架/版权状态", ["下架", "在架", "版权", "到期", "续约", "权利"]),
        ("评级标准", ["评级", "档位", "S+", "S", "A", "B", "C", "D", "E"]),
        ("运营建议质量", ["建议", "运营", "模板", "启发", "可执行", "推广", "维持", "观察", "核查"]),
        ("M4校准候选", ["M4", "校准", "案例", "样本"]),
    ]
    for category, keywords in checks:
        if any(keyword in text for keyword in keywords):
            categories.append(category)
    return categories or ["其他/需人工归类"]


def check_rating_thresholds(rows: list[dict]) -> dict:
    mismatches = []
    for row in rows:
        basis = row["annualizedSalesRevenue"] if row["annualizedSalesRevenue"] is not None else row["salesRevenue12m"]
        expected = rating_from_sales(basis)
        actual = row["salesRating"]
        if actual and expected and actual != expected:
            mismatches.append(
                {
                    "anonymousId": row["anonymousId"],
                    "expected": expected,
                    "actual": actual,
                }
            )
    return {
        "basis": "年化实销收入优先；为空时使用实销12月收入；买断金额不进入实销评级",
        "mismatchCount": len(mismatches),
        "mismatchAnonymousIds": [item["anonymousId"] for item in mismatches],
        "mismatchesByExpectedActual": dict(Counter(f"{item['expected']}->{item['actual']}" for item in mismatches)),
    }


def summarize_buyout(rows: list[dict]) -> dict:
    buyout_rows = [row for row in rows if "买断" in row["revenueModel"]]
    buyout_plus_sales = [row for row in rows if "买断+实销" in row["revenueModel"]]
    pure_buyout = [row for row in buyout_rows if row not in buyout_plus_sales]
    split_complete = [
        row
        for row in buyout_plus_sales
        if row["buyoutEstimatedAmount"] is not None
        and row["salesRevenue12m"] is not None
        and row["buyoutHistoricalValueRating"]
        and row["salesRating"]
    ]
    return {
        "buyoutRelatedRows": len(buyout_rows),
        "pureBuyoutRows": len(pure_buyout),
        "buyoutPlusSalesRows": len(buyout_plus_sales),
        "buyoutPlusSalesSplitCompleteRows": len(split_complete),
        "buyoutHistoricalRatingDistribution": dict(Counter(row["buyoutHistoricalValueRating"] for row in buyout_rows)),
        "ratingFeedbackOnBuyoutRows": summarize_feedback(row["ratingFeedback"] for row in buyout_rows),
    }


def summarize_expired(rows: list[dict]) -> dict:
    expired_rows = [row for row in rows if "expired" in row["currentRightsStatus"].lower() or "到期" in row["currentRightsStatus"]]
    simple_e_rows = [row for row in expired_rows if row["historicalPerformanceRating"] == "E" and row["displayRating"].strip() in {"E", "历史表现 E"}]
    non_simple_e_rows = [row for row in expired_rows if row not in simple_e_rows]
    return {
        "expiredRows": len(expired_rows),
        "simpleEOnlyRows": len(simple_e_rows),
        "notSimplyRatedERows": len(non_simple_e_rows),
        "expiredHistoricalRatingDistribution": dict(Counter(row["historicalPerformanceRating"] for row in expired_rows)),
    }


def summarize_suggestions(rows: list[dict], issue_categories: Counter) -> dict:
    rows_with_evidence = [row for row in rows if row["suggestionEvidence"]]
    rows_with_suggestion = [row for row in rows if row["suggestion"] and "暂无自动" not in row["suggestion"]]
    template_issue_count = issue_categories.get("运营建议质量", 0)
    return {
        "rowsWithSuggestionEvidence": len(rows_with_evidence),
        "rowsWithConcreteSuggestion": len(rows_with_suggestion),
        "rowsWithReviewPrompt": sum(1 for row in rows if row["reviewPrompt"]),
        "userMarkedSuggestionNegative": count_feedback_class(rows, "suggestionFeedback", "negative"),
        "templateLikeIssueCountFromUserNotes": int(template_issue_count),
        "suggestionFeedback": summarize_feedback(row["suggestionFeedback"] for row in rows),
    }


def identify_m4_candidates(rows: list[dict], rating_mismatch_ids: list[str]) -> dict:
    candidates = {}
    for row in rows:
        reasons = []
        if row["anonymousId"] in rating_mismatch_ids:
            reasons.append("实销评级阈值复核")
        if classify_feedback(row["revenueModelFeedback"]) == "negative":
            reasons.append("收入模式判断复核")
        if classify_feedback(row["ratingFeedback"]) == "negative":
            reasons.append("评级判断复核")
        if classify_feedback(row["suggestionFeedback"]) == "negative":
            reasons.append("运营建议复核")
        for category in issue_categories(row["issueNote"]):
            if category != "其他/需人工归类":
                reasons.append(category)
        if reasons:
            candidates[row["anonymousId"]] = sorted(set(reasons))
    by_reason = defaultdict(int)
    for reasons in candidates.values():
        for reason in reasons:
            by_reason[reason] += 1
    return {
        "count": len(candidates),
        "anonymousIds": sorted(candidates),
        "reasonDistribution": dict(sorted(by_reason.items())),
        "rawWorkIdentifiersExcluded": True,
    }


def decide_baseline(
    rows: list[dict],
    feedback_rows: list[dict],
    revenue_model_feedback: dict,
    rating_feedback: dict,
    suggestion_feedback: dict,
    rating_threshold_check: dict,
    suggestion_summary: dict,
) -> dict:
    if not rows or len(feedback_rows) / len(rows) < 0.8:
        verdict = "NOT_READY"
        reason = "用户反馈完成率不足 80%，不能判断为候选基线"
    elif rating_threshold_check["mismatchCount"] > 0:
        verdict = "NOT_READY"
        reason = "存在实销评级阈值机械不一致，需要先修正"
    elif (
        revenue_model_feedback["negativeRate"] <= 0.15
        and rating_feedback["negativeRate"] <= 0.15
        and suggestion_feedback["negativeRate"] <= 0.2
        and suggestion_summary["templateLikeIssueCountFromUserNotes"] == 0
    ):
        verdict = "ACCEPT_AS_LIMITED_M2_CANDIDATE_BASELINE"
        reason = "用户反馈负向率在阈值内，且未发现机械评级错误或模板化建议问题"
    else:
        verdict = "CONDITIONAL_REWORK_REQUIRED"
        reason = "存在较多负向反馈或建议质量问题，应进入 M4 校准样本池后再定版"
    return {
        "verdict": verdict,
        "reason": reason,
        "notFinalFormalReleaseApproval": True,
        "m3Allowed": False,
    }


def count_feedback_class(rows: list[dict], field: str, expected: str) -> int:
    return sum(1 for row in rows if classify_feedback(row[field]) == expected)


def ordered_distribution(rows: list[dict], field: str) -> dict:
    counter = Counter(row[field] for row in rows)
    return {rating: int(counter.get(rating, 0)) for rating in RATING_ORDER}


def rating_from_sales(amount: float | None) -> str:
    if amount is None or amount < 100:
        return "E"
    if amount > 100000:
        return "S+"
    if amount >= 10000:
        return "S"
    if amount >= 5000:
        return "A"
    if amount >= 1000:
        return "B"
    if amount >= 500:
        return "C"
    return "D"


def render_markdown(summary: dict) -> str:
    lines = [
        "# M2 Rating Standard v2 Operator Validation Summary",
        "",
        "This report is sanitized. It contains aggregate counts and anonymous sample ids only.",
        "",
        "## Conclusion",
        "",
        f"- Feedback completion rate: {summary['feedbackCompletionRate']}",
        f"- Baseline decision: `{summary['baselineDecision']['verdict']}`",
        f"- Decision reason: {summary['baselineDecision']['reason']}",
        f"- M3 allowed: {str(summary['baselineDecision']['m3Allowed']).lower()}",
        "",
        "## Feedback Summary",
        "",
        markdown_table(
            [
                {"item": "Revenue model reasonable", **summary["feedbackSummary"]["revenueModelReasonable"]},
                {"item": "Rating reasonable", **summary["feedbackSummary"]["ratingReasonable"]},
                {"item": "Suggestion actionable", **summary["feedbackSummary"]["suggestionActionable"]},
            ],
            ["item", "filled", "positive", "negative", "uncertain", "positiveRate", "negativeRate"],
        ),
        "## Rating Threshold Check",
        "",
        f"- Mismatch count: {summary['ratingThresholdCheck']['mismatchCount']}",
        f"- Basis: {summary['ratingThresholdCheck']['basis']}",
        "",
        "## M4 Calibration Candidates",
        "",
        f"- Count: {summary['m4CalibrationCandidates']['count']}",
        f"- Anonymous ids: {', '.join(summary['m4CalibrationCandidates']['anonymousIds']) or 'none'}",
        "",
        "## Safety",
        "",
        "- Formal master data written: false",
        "- M3 entered: false",
        "- Private workbook committed: false",
        "- Real work titles/authors/channels/raw billing rows included: false",
    ]
    return "\n".join(lines) + "\n"


def markdown_table(rows: list[dict], columns: list[str]) -> str:
    lines = ["| " + " | ".join(columns) + " |", "|" + "|".join("---" for _ in columns) + "|"]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(column, "")) for column in columns) + " |")
    return "\n".join(lines) + "\n"


def envelope(payload: dict) -> dict:
    return {
        "reportId": "m2.rating_standard_v2_operator_validation_summary.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "payload": payload,
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def has_feedback(row: dict) -> bool:
    return any(row[field] for field in ["revenueModelFeedback", "ratingFeedback", "suggestionFeedback", "issueNote"])


def to_number(value) -> float | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except Exception:
        return None
    return number if math.isfinite(number) else None


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def rate(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


if __name__ == "__main__":
    main()
