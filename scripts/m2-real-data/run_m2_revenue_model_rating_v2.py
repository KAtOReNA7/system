from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(ROOT / "tools" / "m2-calibration"))
sys.path.insert(0, str(ROOT / "scripts" / "m2-real-data"))

try:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency. Install local temp deps, for example: "
        "python -m pip install --target %TEMP%\\codex-system-pydeps pandas numpy openpyxl"
    ) from exc

from calibrate_cleaned_bills import month_range
from run_nonformal_dry_run import evaluate_work_summary, load_analysis_inputs


CANDIDATE_VERSION = "m2-realdata-dev-revenue-model-rating-v2.0"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-business-review"

INPUT_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v1.xlsx"
STAGING_TABLE_JSON = ROOT / "data" / "private-output" / "m1-master-data" / "M1-dual-source-limited-staging-table-v1.json"
PRIVATE_VALIDATION_XLSX = PRIVATE_DIR / "m2-revenue-model-rating-v2-validation.xlsx"
PRIVATE_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-rating-standard-v2.xlsx"

PER_CHANNEL_JSON = DOCS_DIR / "M2-per-channel-revenue-pattern-audit-v2.json"
PER_CHANNEL_MD = DOCS_DIR / "M2-per-channel-revenue-pattern-audit-v2.md"
REVENUE_MODEL_JSON = DOCS_DIR / "M2-revenue-model-classification-v2.json"
REVENUE_MODEL_MD = DOCS_DIR / "M2-revenue-model-classification-v2.md"
SHELF_JSON = DOCS_DIR / "M2-shelf-status-inference-v1.json"
SHELF_MD = DOCS_DIR / "M2-shelf-status-inference-v1.md"
THRESHOLDS_JSON = DOCS_DIR / "M2-rating-standard-v2-business-thresholds.json"
THRESHOLDS_MD = DOCS_DIR / "M2-rating-standard-v2-business-thresholds.md"
RATING_JSON = DOCS_DIR / "M2-rating-calibration-v5-summary.json"
RATING_MD = DOCS_DIR / "M2-rating-calibration-v5-summary.md"
SUGGESTION_JSON = DOCS_DIR / "M2-suggestion-calibration-v5-summary.json"
SUGGESTION_MD = DOCS_DIR / "M2-suggestion-calibration-v5-summary.md"
OPERATOR_JSON = DOCS_DIR / "M2-operator-task-pack-rating-standard-v2-summary.json"
OPERATOR_MD = DOCS_DIR / "M2-operator-task-pack-rating-standard-v2-summary.md"

CHANNEL_MODELS = ["sales_share_channel", "buyout_channel", "mixed_channel", "unknown_channel"]
WORK_MODELS = ["pure_sales_share", "pure_buyout", "buyout_plus_sales", "unknown_revenue_model"]
SHELF_STATUSES = [
    "active_on_shelf",
    "likely_off_shelf",
    "rights_expired_likely_off_shelf",
    "off_shelf_but_tail_revenue",
    "unknown_shelf_status",
]
RATINGS = ["S+", "S", "A", "B", "C", "D", "E"]
RANK = {rating: index for index, rating in enumerate(RATINGS)}
MIN_BUYOUT_NO_SALES_MONTHS = 12

WORK_MODEL_CN = {
    "pure_sales_share": "纯实销/纯分成",
    "pure_buyout": "纯买断",
    "buyout_plus_sales": "买断+实销",
    "unknown_revenue_model": "收入模式未知",
}
SHELF_CN = {
    "active_on_shelf": "仍在架或可运营",
    "likely_off_shelf": "大概率下架",
    "rights_expired_likely_off_shelf": "版权到期，大概率下架",
    "off_shelf_but_tail_revenue": "已下架但仍有存量会员/已购用户尾部收入",
    "unknown_shelf_status": "无法判断",
}
CHANNEL_CN = {
    "sales_share_channel": "实销/分成渠道",
    "buyout_channel": "买断渠道",
    "mixed_channel": "混合渠道",
    "unknown_channel": "未知渠道",
}


def main() -> None:
    ensure_inputs()
    context = load_analysis_inputs()
    bill = context["bill"]
    complete = bill[bill["validForCalibration"] & (bill["billMonth"] <= context["latest_complete_month"])].copy()
    if not {"standardWorkId", "channelKey", "businessForm", "billMonth", "amount"}.issubset(complete.columns):
        raise SystemExit("Cannot locate standardWorkId + channelKey + businessForm + billMonth + amount income facts.")
    if complete.empty:
        raise SystemExit("No cleaned bill income facts available for per-channel classification.")

    evaluated = evaluate_work_summary(
        context["work_summary"],
        context["parameters"],
        context["latest_complete_month"],
        context["incomplete_work_ids"],
        "candidate-b",
    ).sort_values("standardWorkId")

    staging_index = build_staging_index()
    months = month_range(complete["billMonth"].min(), context["latest_complete_month"])
    cluster_lookup = build_same_amount_clusters(complete).set_index(["standardWorkId", "channelKey", "businessForm"]).to_dict("index")
    channel_rows = build_channel_rows(complete, months, cluster_lookup, context["latest_complete_month"])
    work_rows = build_work_rows(channel_rows, evaluated, months, context["latest_complete_month"], staging_index)
    operator_rows = build_operator_rows(work_rows)

    per_channel_report = build_per_channel_report(channel_rows, context, months)
    revenue_model_report = build_revenue_model_report(work_rows)
    shelf_report = build_shelf_report(work_rows)
    threshold_report = build_threshold_report(work_rows)
    rating_report = build_rating_report(work_rows)
    suggestion_report = build_suggestion_report(work_rows)
    operator_report = build_operator_report(operator_rows)

    write_public_reports(
        per_channel_report,
        revenue_model_report,
        shelf_report,
        threshold_report,
        rating_report,
        suggestion_report,
        operator_report,
    )
    write_private_workbooks(
        channel_rows,
        work_rows,
        operator_rows,
        per_channel_report,
        revenue_model_report,
        shelf_report,
        threshold_report,
        rating_report,
        suggestion_report,
    )

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "totalWorks": len(work_rows),
                "channelGroups": len(channel_rows),
                "revenueModelDistribution": revenue_model_report["revenueModelDistribution"],
                "revenueModelRevenueShare": revenue_model_report["revenueShareByRevenueModel"],
                "shelfStatusDistribution": shelf_report["shelfStatusDistribution"],
                "salesRatingDistribution": rating_report["salesPerformanceRatingDistribution"],
                "suggestionDistribution": suggestion_report["suggestionTypeDistribution"],
                "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
                "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    required = [
        ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
        ROOT / "docs" / "analysis" / "m2-real-data" / "M2-billing-pattern-feature-audit-v1.json",
        ROOT / "docs" / "analysis" / "m2-real-data" / "M2-revenue-model-classification-audit-v1.json",
        ROOT / "docs" / "analysis" / "m2-real-data" / "M2-rating-standard-definition-v1.json",
        ROOT / "docs" / "analysis" / "m2-real-data" / "M2-rating-standard-v1-validation-summary.json",
        STAGING_TABLE_JSON,
        INPUT_OPERATOR_XLSX,
        ROOT / "src" / "domain" / "oldProductEvaluation" / "revenueModelClassifier.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "shelfStatusInference.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "forecastabilityGate.js",
        ROOT / "package.json",
        ROOT / ".gitignore",
    ]
    missing = [path for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def build_staging_index() -> dict[str, dict]:
    payload = read_json(STAGING_TABLE_JSON)
    records = payload.get("records", []) if isinstance(payload, dict) else payload
    by_work: dict[str, dict] = defaultdict(dict)
    for record in records:
        sid = clean(record.get("standardWorkId"))
        field = clean(record.get("fieldName"))
        if not sid or field not in {"copyrightStartDate", "copyrightEndDate"}:
            continue
        parsed = parse_date_value(record.get("applyValue"))
        if not parsed:
            continue
        by_work[sid][field] = parsed
        by_work[sid][f"{field}Source"] = clean(record.get("candidateSource")) or clean(record.get("applySourceType")) or "dual_source_staging"
    return dict(by_work)


def parse_date_value(value) -> str:
    text = clean(value)
    if not text:
        return ""
    try:
        numeric = float(text)
        if numeric > 1000:
            return (datetime(1899, 12, 30) + timedelta(days=int(numeric))).date().isoformat()
    except Exception:
        pass
    for fmt, width in (
        ("%Y-%m-%d", 10),
        ("%Y/%m/%d", 10),
        ("%Y.%m.%d", 10),
        ("%Y-%m", 7),
        ("%Y/%m", 7),
        ("%Y.%m", 7),
    ):
        try:
            parsed = datetime.strptime(text[:width], fmt)
            return parsed.date().isoformat()
        except Exception:
            continue
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.date().isoformat()
    except Exception:
        pass
    return ""


def build_same_amount_clusters(complete: pd.DataFrame) -> pd.DataFrame:
    data = complete[complete["amount"] > 0].copy()
    data["amountRounded2"] = data["amount"].round(2)
    clusters = data.groupby(["billMonth", "amountRounded2"])["standardWorkId"].nunique().reset_index(name="clusterSize")
    merged = data[["standardWorkId", "channelKey", "businessForm", "billMonth", "amountRounded2"]].merge(
        clusters, on=["billMonth", "amountRounded2"], how="left"
    )
    aggregated = (
        merged.groupby(["standardWorkId", "channelKey", "businessForm"], dropna=False)
        .agg(
            sameMonthSameAmountClusterSize=("clusterSize", "max"),
            adjacentRowsSameAmountSignal=("clusterSize", lambda values: bool((values >= 2).any())),
        )
        .reset_index()
    )
    return aggregated


def build_channel_rows(complete: pd.DataFrame, months: list[str], cluster_lookup: pd.DataFrame, latest_month: str) -> list[dict]:
    rows: list[dict] = []
    grouped = complete.groupby(["standardWorkId", "channelKey", "businessForm"], dropna=False)
    for (sid, channel, form), frame in grouped:
        monthly_series = frame.groupby("billMonth")["amount"].sum()
        positives = monthly_series[monthly_series > 0]
        if positives.empty:
            continue
        observed_months = month_range(str(positives.index.min()), latest_month)
        values = monthly_series.reindex(observed_months, fill_value=0.0).to_numpy(dtype=float)
        cluster = cluster_lookup.get((sid, channel, form), {"sameMonthSameAmountClusterSize": 1, "adjacentRowsSameAmountSignal": False})
        max_cluster = int(cluster.get("sameMonthSameAmountClusterSize", 1) or 1)
        adjacent_signal = bool(cluster.get("adjacentRowsSameAmountSignal", False))
        features = channel_features(values, max_cluster, adjacent_signal)
        model = classify_channel(features)
        rows.append(
            {
                "standardWorkId": clean(sid),
                "channelKey": clean(channel),
                "businessForm": clean(form),
                "monthlyValues": values,
                "observedMonths": observed_months,
                **features,
                **model,
            }
        )
    return rows


def channel_features(values: np.ndarray, max_cluster: int, adjacent_signal: bool) -> dict:
    positive = values[values > 0]
    total = float(positive.sum())
    largest = float(positive.max()) if len(positive) else 0.0
    peak_index = int(np.argmax(values)) if len(values) else 0
    tail_values = values[peak_index + 1 :] if len(values) else np.array([])
    tail = tail_values[(tail_values > 0) & (tail_values < largest * 0.35)] if largest > 0 else np.array([])
    post_positive_count = int(len(tail_values[tail_values > 0])) if len(tail_values) else 0
    post_observed_count = int(len(tail_values))
    post_no_sales_count = post_observed_count if post_positive_count == 0 else 0
    post_no_sales_signal = post_no_sales_count >= MIN_BUYOUT_NO_SALES_MONTHS
    positive_count = int(len(positive))
    observed_count = int(len(values))
    active_ratio = positive_count / observed_count if observed_count else 0.0
    continuity_score = continuity(values)
    amount_variation = variation_score(positive)
    integer_ratio = ratio(positive, lambda value: abs(value - round(value)) <= 0.01)
    round_ratio = ratio(positive, is_round_amount)
    nonstandard_ratio = ratio(positive, lambda value: abs(value - round(value)) > 0.01)
    large_share = largest / total if total > 0 else 0.0
    equal_split = max_cluster >= 3 or adjacent_signal
    natural_sales = positive_count >= 4 and continuity_score >= 0.45 and amount_variation >= 0.12 and large_share < 0.88
    buyout_score = score_buyout_channel(
        positive_count=positive_count,
        large_share=large_share,
        integer_ratio=integer_ratio,
        round_ratio=round_ratio,
        equal_split=equal_split,
        adjacent_signal=adjacent_signal,
        tail_count=len(tail),
        post_positive_count=post_positive_count,
        post_no_sales_signal=post_no_sales_signal,
    )
    sales_score = score_sales_channel(
        positive_count=positive_count,
        active_ratio=active_ratio,
        continuity_score=continuity_score,
        amount_variation=amount_variation,
        nonstandard_ratio=nonstandard_ratio,
        large_share=large_share,
        tail_count=len(tail),
    )
    return {
        "positiveMonthCount": positive_count,
        "observedMonthCount": observed_count,
        "activeMonthRatio": round_float(active_ratio),
        "monthlyContinuityScore": round_float(continuity_score),
        "amountVariationScore": round_float(amount_variation),
        "nonStandardAmountRatio": round_float(nonstandard_ratio),
        "integerAmountRatio": round_float(integer_ratio),
        "roundAmountRatio": round_float(round_ratio),
        "singleLargeMonthShare": round_float(large_share),
        "sameMonthSameAmountClusterSize": max_cluster,
        "adjacentRowsSameAmountSignal": adjacent_signal,
        "equalSplitBatchSignal": equal_split,
        "postLargePaymentTailMonthCount": int(len(tail)),
        "postLargePaymentTailRevenue": round_float(float(tail.sum()), 2),
        "postLargePaymentObservedMonthCount": post_observed_count,
        "postLargePaymentPositiveMonthCount": post_positive_count,
        "postLargePaymentNoSalesMonthCount": post_no_sales_count,
        "postLargePaymentNoSalesSignal": post_no_sales_signal,
        "largeIntegerPaymentSignal": largest >= 1000 and abs(largest - round(largest)) <= 0.01,
        "largeRoundPaymentSignal": largest >= 1000 and is_round_amount(largest),
        "naturalSalesSequenceSignal": natural_sales,
        "buyoutSignalScore": round_float(buyout_score),
        "salesSignalScore": round_float(sales_score),
        "positiveIncomeTotal": round_float(total, 2),
        "largestMonthIncome": round_float(largest, 2),
    }


def classify_channel(features: dict) -> dict:
    buyout_score = features["buyoutSignalScore"]
    sales_score = features["salesSignalScore"]
    reasons = []
    if features["positiveIncomeTotal"] <= 0 or features["positiveMonthCount"] <= 0:
        model = "unknown_channel"
        confidence = "low"
        reasons.append("渠道没有正收入事实")
    elif has_any_buyout_signal(features):
        model = "buyout_channel"
        confidence = "medium" if buyout_score >= 0.68 else "low"
        reasons.append("渠道命中大额整数/同批次同额/买断后无实销任一买断信号，按买断渠道处理")
    elif (
        features["naturalSalesSequenceSignal"]
        and features["positiveMonthCount"] >= 4
        and features["postLargePaymentTailMonthCount"] >= 2
    ):
        model = "sales_share_channel"
        confidence = "high" if features["positiveMonthCount"] >= 6 else "medium"
        reasons.append("渠道存在连续多月自然实销序列，优先按实销/分成处理，不因整额或同批次信号直接判买断")
    elif sales_score >= 0.5 and buyout_score < 0.68:
        model = "sales_share_channel"
        confidence = "high" if sales_score >= 0.68 else "medium"
        reasons.append("渠道收入连续或半连续，金额自然波动")
    elif features["positiveMonthCount"] >= 4 and features["postLargePaymentTailMonthCount"] >= 3:
        model = "sales_share_channel"
        confidence = "medium"
        reasons.append("渠道大额收入后仍有持续实销，按上线前期大卖或自然实销序列处理，不判买断")
    elif features["positiveMonthCount"] >= 3:
        model = "sales_share_channel"
        confidence = "low"
        reasons.append("多月收入更接近实销，不保守归为 unknown")
    elif features["positiveMonthCount"] >= 1:
        model = "sales_share_channel"
        confidence = "low"
        reasons.append("有效账单收入未命中任一买断信号，按单月实销样本计入实销口径")
    else:
        model = "unknown_channel"
        confidence = "low"
        reasons.append("渠道数据稀少或信号冲突")
    return {
        "channelRevenueModel": model,
        "channelRevenueModelChinese": CHANNEL_CN[model],
        "channelRevenueModelConfidence": confidence,
        "channelClassificationReasonChinese": reasons,
    }


def has_any_buyout_signal(features: dict) -> bool:
    return (
        has_large_amount_buyout_signal(features)
        or has_same_batch_buyout_signal(features)
        or has_no_sales_after_candidate_buyout_signal(features)
    )


def has_large_amount_buyout_signal(features: dict) -> bool:
    return bool(
        float(features.get("positiveIncomeTotal") or 0) >= 1000
        and int(features.get("postLargePaymentPositiveMonthCount") or 0) == 0
        and (features.get("largeRoundPaymentSignal") or features.get("largeIntegerPaymentSignal"))
    )


def has_same_batch_buyout_signal(features: dict) -> bool:
    return bool(
        features.get("equalSplitBatchSignal")
        or features.get("adjacentRowsSameAmountSignal")
    )


def has_no_sales_after_candidate_buyout_signal(features: dict) -> bool:
    return float(features.get("positiveIncomeTotal") or 0) >= 1000 and has_post_buyout_no_sales_signal(features)


def has_post_buyout_no_sales_signal(features: dict) -> bool:
    return bool(features.get("postLargePaymentNoSalesSignal")) or (
        int(features.get("postLargePaymentNoSalesMonthCount") or 0) >= MIN_BUYOUT_NO_SALES_MONTHS
        and int(features.get("postLargePaymentPositiveMonthCount") or 0) == 0
    )


def build_work_rows(
    channel_rows: list[dict],
    evaluated: pd.DataFrame,
    months: list[str],
    latest_month: str,
    staging_index: dict[str, dict],
) -> list[dict]:
    evaluated_index = evaluated.set_index("standardWorkId", drop=False)
    channels_by_work: dict[str, list[dict]] = defaultdict(list)
    for row in channel_rows:
        channels_by_work[row["standardWorkId"]].append(row)

    rows = []
    for sid in sorted(evaluated_index.index.astype(str)):
        summary = evaluated_index.loc[sid]
        if isinstance(summary, pd.DataFrame):
            summary = summary.iloc[0]
        channels = channels_by_work.get(sid, [])
        row = aggregate_work(sid, channels, summary, months, latest_month, staging_index)
        rows.append(row)
    return rows


def aggregate_work(sid: str, channels: list[dict], summary, months: list[str], latest_month: str, staging_index: dict[str, dict]) -> dict:
    channel_counts = Counter(channel["channelRevenueModel"] for channel in channels)
    has_sales = channel_counts["sales_share_channel"] > 0
    has_buyout = channel_counts["buyout_channel"] > 0
    has_mixed = channel_counts["mixed_channel"] > 0
    total_revenue = sum(channel["positiveIncomeTotal"] for channel in channels)
    sales_by_month = defaultdict(float)
    buyout_estimate = 0.0
    sales_tail_estimate = 0.0

    for channel in channels:
        values = channel["monthlyValues"]
        observed_months = channel["observedMonths"]
        model = channel["channelRevenueModel"]
        if model == "buyout_channel":
            buyout_estimate += channel["positiveIncomeTotal"]
        elif model == "mixed_channel":
            peak_index = int(np.argmax(values)) if len(values) else 0
            buyout_estimate += float(values[peak_index]) if len(values) else 0.0
            for idx, value in enumerate(values):
                if idx != peak_index and value > 0:
                    sales_by_month[observed_months[idx]] += float(value)
                    sales_tail_estimate += float(value)
        elif model == "sales_share_channel":
            for idx, value in enumerate(values):
                if value > 0:
                    sales_by_month[observed_months[idx]] += float(value)
                    sales_tail_estimate += float(value)

    if (has_buyout or has_mixed) and (has_sales or has_mixed):
        revenue_model = "buyout_plus_sales"
        confidence = "medium"
        reason = "作品存在买断渠道和实销渠道，或存在混合渠道"
    elif has_sales and not has_buyout and not has_mixed:
        revenue_model = "pure_sales_share"
        confidence = "medium"
        reason = "作品至少存在一个实销渠道，且没有强买断渠道"
    elif has_buyout and not has_sales and not has_mixed:
        revenue_model = "pure_buyout"
        confidence = "medium"
        reason = "作品存在买断渠道，且没有持续实销渠道"
    elif total_revenue > 0:
        fallback_model, fallback_reason = fallback_work_model(channels)
        revenue_model = fallback_model
        confidence = "low"
        reason = fallback_reason
        if revenue_model == "pure_sales_share":
            sales_tail_estimate = total_revenue
        elif revenue_model == "pure_buyout":
            buyout_estimate = total_revenue
    else:
        revenue_model = "unknown_revenue_model"
        confidence = "low"
        reason = "作品缺少可判定收入事实"

    last12 = months[-12:]
    sales_revenue12m = sum(sales_by_month.get(month, 0.0) for month in last12)
    first_sales_month = min(sales_by_month) if sales_by_month else None
    if first_sales_month:
        observed_sales_months = min(12, len(month_range(first_sales_month, latest_month)))
    else:
        observed_sales_months = 12
    sales_annualized = sales_revenue12m if observed_sales_months >= 12 else sales_revenue12m / max(1, observed_sales_months) * 12
    recent3 = sum(sales_by_month.get(month, 0.0) for month in months[-3:])
    recent6 = sum(sales_by_month.get(month, 0.0) for month in months[-6:])
    recent_positive_6 = sum(1 for month in months[-6:] if sales_by_month.get(month, 0.0) > 0)
    latest_income_month = clean(summary.get("latestIncomeMonth"))
    months_since_latest_income = month_distance(latest_income_month, latest_month) if latest_income_month else None
    staging_item = staging_index.get(sid, {})
    copyright_end_date = staging_item.get("copyrightEndDate", "")
    remaining = safe_float(summary.get("remainingCopyrightMonths"), None)
    if remaining is None and copyright_end_date:
        remaining = remaining_months_until(copyright_end_date, latest_month)
    current_rights = "expired" if remaining is not None and remaining < 0 else "active" if remaining is not None else "unknown"
    shelf = infer_shelf_status(
        current_rights_status=current_rights,
        remaining_months=remaining,
        recent3=recent3,
        recent6=recent6,
        sales12=sales_revenue12m,
        recent_positive_6=recent_positive_6,
        months_since_latest_income=months_since_latest_income,
        revenue_model=revenue_model,
    )
    sales_rating = rating_from_sales(sales_revenue12m if first_sales_month is None or observed_sales_months >= 12 else sales_annualized)
    buyout_rating = rating_from_sales(buyout_estimate) if revenue_model in {"pure_buyout", "buyout_plus_sales"} and buyout_estimate > 0 else "not_applicable"
    historical_rating = best_rating([sales_rating, buyout_rating if buyout_rating != "not_applicable" else None])
    forecast_rating = forecast_rating_from_summary(summary)
    operational = operational_rating(current_rights, shelf["shelfStatus"], revenue_model)
    suggestion = suggestion_for(
        revenue_model=revenue_model,
        sales_rating=sales_rating,
        buyout_rating=buyout_rating,
        shelf_status=shelf["shelfStatus"],
        current_rights=current_rights,
        forecastability=infer_forecastability_status(summary),
        confidence=infer_forecast_confidence(summary),
        lifecycle=clean(summary.get("lifecycle")),
        sales12=sales_revenue12m,
    )
    display = (
        f"历史表现 {historical_rating}；实销评级 {sales_rating}；"
        f"买断历史价值 {buyout_rating}；版权状态 {current_rights}；"
        f"下架状态 {SHELF_CN[shelf['shelfStatus']]}；预测评级 {forecast_rating}；"
        f"当前运营限制 {operational['operationalDecisionRatingChinese']}"
    )
    return {
        "standardWorkId": sid,
        "totalHistoricalRevenue": round_float(total_revenue, 2),
        "revenueModel": revenue_model,
        "revenueModelChinese": WORK_MODEL_CN[revenue_model],
        "revenueModelConfidence": confidence,
        "channelModelSummary": {model: int(channel_counts.get(model, 0)) for model in CHANNEL_MODELS},
        "channelCount": len(channels),
        "buyoutSignalScore": round_float(max([channel["buyoutSignalScore"] for channel in channels], default=0.0)),
        "salesSignalScore": round_float(max([channel["salesSignalScore"] for channel in channels], default=0.0)),
        "equalSplitSignalScore": round_float(1.0 if any(channel["equalSplitBatchSignal"] for channel in channels) else 0.0),
        "buyoutEstimatedAmount": round_float(buyout_estimate, 2),
        "salesRevenue12m": round_float(sales_revenue12m, 2),
        "salesRevenueAnnualized": round_float(sales_annualized, 2),
        "salesTailEstimatedAmount": round_float(sales_tail_estimate, 2),
        "classificationReasonChinese": reason,
        "manualReviewRequired": revenue_model == "unknown_revenue_model" or confidence == "low",
        "currentRightsStatus": current_rights,
        "remainingCopyrightMonths": round_float(remaining, 2) if remaining is not None else None,
        "copyrightEndDateSource": staging_item.get("copyrightEndDateSource", ""),
        **shelf,
        "salesPerformanceRating": sales_rating,
        "buyoutHistoricalValueRating": buyout_rating,
        "historicalPerformanceRating": historical_rating,
        "forecastValueRating": forecast_rating,
        **operational,
        **suggestion,
        "displayRating": display,
        "lifecycle": clean(summary.get("lifecycle")),
        "forecastabilityStatus": infer_forecastability_status(summary),
        "forecastConfidence": infer_forecast_confidence(summary),
    }


def fallback_work_model(channels: list[dict]) -> tuple[str, str]:
    positive_months = sum(channel["positiveMonthCount"] for channel in channels)
    total = sum(channel["positiveIncomeTotal"] for channel in channels)
    has_any_buyout = any(has_any_buyout_signal(channel) for channel in channels)
    if has_any_buyout:
        return "pure_buyout", "命中大额整数/同批次同额/买断后无实销任一买断信号"
    if positive_months >= 1:
        return "pure_sales_share", "有效账单收入未命中任一买断信号，按实销口径处理"
    return "unknown_revenue_model", "仅极少收入或渠道证据冲突，保留 unknown"


def infer_shelf_status(
    current_rights_status: str,
    remaining_months: float | None,
    recent3: float,
    recent6: float,
    sales12: float,
    recent_positive_6: int,
    months_since_latest_income: int | None,
    revenue_model: str,
) -> dict:
    expired = current_rights_status == "expired" or (remaining_months is not None and remaining_months < 0)
    active_rights = current_rights_status == "active" or (remaining_months is not None and remaining_months >= 0)
    if expired and sales12 > 0:
        status = "off_shelf_but_tail_revenue"
        reason = "版权已到期但仍有尾部实销收入"
    elif expired:
        status = "rights_expired_likely_off_shelf"
        reason = "版权已到期且未见尾部实销收入"
    elif active_rights and recent_positive_6 >= 2 and recent6 > 0:
        status = "active_on_shelf"
        reason = "版权有效且近 6 个月仍有持续实销"
    elif active_rights and sales12 > 0 and (months_since_latest_income is None or months_since_latest_income <= 6):
        status = "active_on_shelf"
        reason = "版权有效且近 12 个月有实销"
    elif not active_rights and months_since_latest_income is not None and months_since_latest_income >= 12 and revenue_model != "pure_buyout":
        status = "likely_off_shelf"
        reason = "近期收入断流且权利状态不明"
    elif sales12 == 0 and revenue_model == "pure_buyout":
        status = "unknown_shelf_status"
        reason = "买断后无持续实销不等于下架"
    else:
        status = "unknown_shelf_status"
        reason = "收入为 0 不能单独判断下架，证据不足"
    return {
        "shelfStatus": status,
        "shelfStatusChinese": SHELF_CN[status],
        "shelfStatusReasonChinese": reason,
    }


def operational_rating(current_rights: str, shelf_status: str, revenue_model: str) -> dict:
    if current_rights == "expired":
        return {"operationalDecisionRating": "renewal_review_required", "operationalDecisionRatingChinese": "需续约/权利复核"}
    if shelf_status == "off_shelf_but_tail_revenue":
        return {"operationalDecisionRating": "rights_audit_required", "operationalDecisionRatingChinese": "需权利/尾部收入核查"}
    if shelf_status in {"likely_off_shelf", "rights_expired_likely_off_shelf"}:
        return {"operationalDecisionRating": "shelf_review_required", "operationalDecisionRatingChinese": "需下架状态复核"}
    if revenue_model == "unknown_revenue_model":
        return {"operationalDecisionRating": "revenue_model_review_required", "operationalDecisionRatingChinese": "需收入模式复核"}
    return {"operationalDecisionRating": "operable", "operationalDecisionRatingChinese": "可运营/可观察"}


def suggestion_for(
    revenue_model: str,
    sales_rating: str,
    buyout_rating: str,
    shelf_status: str,
    current_rights: str,
    forecastability: str,
    confidence: str,
    lifecycle: str,
    sales12: float,
) -> dict:
    evidence = [
        f"收入模式={WORK_MODEL_CN[revenue_model]}",
        f"实销评级={sales_rating}",
        f"买断历史价值={buyout_rating}",
        f"下架状态={SHELF_CN[shelf_status]}",
        f"预测状态={forecastability}",
    ]
    operating = ""
    review = ""
    no_auto = ""
    manual = True
    if shelf_status == "off_shelf_but_tail_revenue":
        suggestion_type = "rights_audit"
        review = "先做权利核查和尾部收入归因"
    elif current_rights == "expired" or shelf_status == "rights_expired_likely_off_shelf":
        suggestion_type = "renewal_review" if sales_rating in {"S+", "S", "A", "B"} or buyout_rating in {"S+", "S", "A"} else "observe_only"
        review = "先做续约价值复核" if suggestion_type == "renewal_review" else "仅归档观察"
    elif shelf_status != "active_on_shelf":
        suggestion_type = "observe_only"
        review = "下架状态不明，仅观察"
        no_auto = "下架状态证据不足"
    elif (
        revenue_model in {"pure_sales_share", "buyout_plus_sales"}
        and sales_rating in {"S+", "S", "A"}
        and forecastability == "numeric_forecast_eligible"
        and confidence in {"high", "medium"}
    ):
        suggestion_type = "promote"
        operating = "可考虑加强分发或重点推广"
        manual = True
    elif shelf_status == "active_on_shelf" and sales_rating in {"B", "C"}:
        suggestion_type = "maintain"
        operating = "维持当前运营"
        manual = False
    elif shelf_status == "active_on_shelf" and sales_rating in {"C", "D"} and sales12 > 0:
        suggestion_type = "reduce"
        operating = "降低增量投入，保留观察"
        manual = True
    elif sales_rating == "E" and revenue_model not in {"pure_buyout", "buyout_plus_sales"} and sales12 <= 0:
        suggestion_type = "downlist_suspend"
        review = "下架/暂停候选，必须人工确认"
        no_auto = "下架是高风险动作，不自动执行"
    elif revenue_model == "unknown_revenue_model" or confidence == "low":
        suggestion_type = "observe_only"
        review = "收入模式或预测置信度不足，仅观察"
        no_auto = "证据不足，不输出强建议"
    else:
        suggestion_type = "observe_only"
        review = "暂无自动运营建议，仅建议人工复核/观察"
        no_auto = "缺少足够结构化证据"
    return {
        "suggestionType": suggestion_type,
        "operatingSuggestion": operating,
        "reviewPrompt": review,
        "noAutomaticSuggestionReason": no_auto,
        "suggestionEvidenceChinese": "；".join(evidence),
        "requiredManualChecks": "；".join(required_checks_for(suggestion_type)),
        "actionabilityLevel": "可执行" if operating and not manual else "需人工确认" if manual else "仅供参考",
        "requiresManualConfirmation": manual,
    }


def required_checks_for(suggestion_type: str) -> list[str]:
    return {
        "promote": ["确认渠道资源", "确认近期收入非异常"],
        "maintain": ["常规监控"],
        "reduce": ["确认是否有保留运营价值"],
        "renewal_review": ["确认续约成本", "确认权利窗口"],
        "rights_audit": ["确认权利状态", "确认尾部收入来源"],
        "downlist_suspend": ["确认无续约价值", "确认非买断高价值作品"],
        "observe_only": ["人工观察"],
    }.get(suggestion_type, ["人工复核"])


def build_operator_rows(work_rows: list[dict]) -> list[dict]:
    work_by_id = {row["standardWorkId"]: row for row in work_rows}
    workbook = load_workbook(INPUT_OPERATOR_XLSX, read_only=True, data_only=True)
    sheet = workbook["01_运营任务卡"] if "01_运营任务卡" in workbook.sheetnames else workbook.worksheets[-1]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = []
    for values in iterator:
        source = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if not any(clean(value) for value in source.values()):
            continue
        sid = find_value(source, ["standardWorkId", "standard_work_id", "标准作品ID"])
        work = work_by_id.get(sid, {})
        rows.append(
            {
                "standardWorkId": sid,
                "作品名": find_value(source, ["作品名", "作品名称"]),
                "作者": find_value(source, ["作者"]),
                "样本来源": find_value(source, ["样本来源"]),
                "收入模式": work.get("revenueModelChinese", ""),
                "渠道级模式摘要": json.dumps(work.get("channelModelSummary", {}), ensure_ascii=False),
                "实销12月收入": work.get("salesRevenue12m", ""),
                "年化实销收入": work.get("salesRevenueAnnualized", ""),
                "买断估计金额": work.get("buyoutEstimatedAmount", ""),
                "实销评级": work.get("salesPerformanceRating", ""),
                "买断历史价值评级": work.get("buyoutHistoricalValueRating", ""),
                "历史表现评级": work.get("historicalPerformanceRating", ""),
                "当前版权状态": work.get("currentRightsStatus", ""),
                "下架状态": work.get("shelfStatusChinese", ""),
                "预测价值评级": work.get("forecastValueRating", ""),
                "当前运营评级": work.get("operationalDecisionRatingChinese", ""),
                "展示评级": work.get("displayRating", ""),
                "运营建议": work.get("operatingSuggestion", "") or "暂无自动运营建议",
                "复核提示": work.get("reviewPrompt", ""),
                "建议证据": work.get("suggestionEvidenceChinese", ""),
                "是否需要人工确认": work.get("requiresManualConfirmation", True),
                "用户反馈：收入模式是否合理": "",
                "用户反馈：评级是否合理": "",
                "用户反馈：建议是否可执行": "",
                "用户反馈：问题说明": "",
            }
        )
    return rows


def build_per_channel_report(channel_rows: list[dict], context: dict, months: list[str]) -> dict:
    counter = Counter(row["channelRevenueModel"] for row in channel_rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "inputBoundary": "standard_work_id + channel + business_form + month income facts; no contract/commercial-term lookup",
        "latestCompleteMonth": context["latest_complete_month"],
        "channelGroups": len(channel_rows),
        "observableMonthSpan": len(months),
        "channelRevenueModelDistribution": {model: int(counter.get(model, 0)) for model in CHANNEL_MODELS},
        "featureDefinitions": [
            "positiveMonthCount",
            "observedMonthCount",
            "activeMonthRatio",
            "monthlyContinuityScore",
            "amountVariationScore",
            "nonStandardAmountRatio",
            "integerAmountRatio",
            "roundAmountRatio",
            "singleLargeMonthShare",
            "sameMonthSameAmountClusterSize",
            "adjacentRowsSameAmountSignal",
            "equalSplitBatchSignal",
            "postLargePaymentTailMonthCount",
            "postLargePaymentTailRevenue",
            "naturalSalesSequenceSignal",
            "buyoutSignalScore",
            "salesSignalScore",
            "channelRevenueModel",
        ],
        "naturalSalesSequenceSignalCount": sum(1 for row in channel_rows if row["naturalSalesSequenceSignal"]),
        "equalSplitBatchSignalCount": sum(1 for row in channel_rows if row["equalSplitBatchSignal"]),
        "sanitized": True,
        "m3Entered": False,
    }


def build_revenue_model_report(work_rows: list[dict]) -> dict:
    counter = Counter(row["revenueModel"] for row in work_rows)
    total_revenue = sum(row["totalHistoricalRevenue"] for row in work_rows)
    revenue_by_model = defaultdict(float)
    unknown_reasons = Counter()
    for row in work_rows:
        revenue_by_model[row["revenueModel"]] += row["totalHistoricalRevenue"]
        if row["revenueModel"] == "unknown_revenue_model":
            unknown_reasons[row["classificationReasonChinese"]] += 1
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "classificationBoundary": "work model is aggregated from per-channel models",
        "totalWorks": len(work_rows),
        "revenueModelDistribution": {model: int(counter.get(model, 0)) for model in WORK_MODELS},
        "revenueShareByRevenueModel": {
            model: round_float(revenue_by_model[model] / total_revenue if total_revenue else 0.0) for model in WORK_MODELS
        },
        "unknownReasons": dict(unknown_reasons),
        "manualReviewRequiredCount": sum(1 for row in work_rows if row["manualReviewRequired"]),
        "buyoutPlusSalesCount": int(counter.get("buyout_plus_sales", 0)),
        "unknownReducedFromV1": int(counter.get("unknown_revenue_model", 0)) < 409,
        "sanitized": True,
        "m3Entered": False,
    }


def build_shelf_report(work_rows: list[dict]) -> dict:
    counter = Counter(row["shelfStatus"] for row in work_rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "shelfStatusDistribution": {status: int(counter.get(status, 0)) for status in SHELF_STATUSES},
        "ruleBoundary": "shelf status affects operational decision and suggestion, not historicalPerformanceRating",
        "expiredTailRevenueCount": int(counter.get("off_shelf_but_tail_revenue", 0)),
        "zeroRevenueAloneNotOffShelf": True,
        "sanitized": True,
        "m3Entered": False,
    }


def build_threshold_report(work_rows: list[dict]) -> dict:
    sales_counter = Counter(row["salesPerformanceRating"] for row in work_rows)
    previous = read_json(DOCS_DIR / "M2-rating-standard-v1-validation-summary.json")["payload"]["ratingStandard"][
        "displayRatingDistribution"
    ]
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "userThresholds": {
            "S+": "100000 以上",
            "S": "10000 - 100000",
            "A": "5000 - 10000",
            "B": "1000 - 5000",
            "C": "500 - 1000",
            "D": "100 - 500",
            "E": "100 以内",
        },
        "revenuePath": [
            "salesRevenue12m after excluding buyout revenue",
            "salesRevenueAnnualized when less than 12 months are observable",
            "buyoutEstimatedAmount separately displayed as buyoutHistoricalValueRating",
        ],
        "salesPerformanceRatingDistribution": {rating: int(sales_counter.get(rating, 0)) for rating in RATINGS},
        "sAndSPlusCount": int(sales_counter.get("S+", 0) + sales_counter.get("S", 0)),
        "previousV1DisplayRatingDistribution": previous,
        "thresholdWidthCheck": threshold_width_check(work_rows),
        "usesUserThresholdsAsBaseline": True,
        "sanitized": True,
        "m3Entered": False,
    }


def build_rating_report(work_rows: list[dict]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "salesPerformanceRatingDistribution": distribution(work_rows, "salesPerformanceRating", RATINGS),
        "buyoutHistoricalValueRatingDistribution": dict(Counter(row["buyoutHistoricalValueRating"] for row in work_rows)),
        "historicalPerformanceRatingDistribution": distribution(work_rows, "historicalPerformanceRating", RATINGS),
        "forecastValueRatingDistribution": dict(Counter(row["forecastValueRating"] for row in work_rows)),
        "operationalDecisionRatingDistribution": dict(Counter(row["operationalDecisionRating"] for row in work_rows)),
        "expiredDoesNotRewriteHistoricalE": True,
        "shelfDoesNotRewriteHistoricalE": True,
        "sanitized": True,
        "m3Entered": False,
    }


def build_suggestion_report(work_rows: list[dict]) -> dict:
    counter = Counter(row["suggestionType"] for row in work_rows)
    operating = sum(1 for row in work_rows if row["operatingSuggestion"])
    review = sum(1 for row in work_rows if row["reviewPrompt"])
    no_auto = sum(1 for row in work_rows if row["noAutomaticSuggestionReason"])
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "operatingSuggestionCount": operating,
        "reviewPromptCount": review,
        "noAutomaticSuggestionCount": no_auto,
        "suggestionTypeDistribution": dict(counter),
        "usesEvidenceGate": True,
        "templateLikeSuggestionSuppressed": True,
        "sanitized": True,
        "m3Entered": False,
    }


def build_operator_report(operator_rows: list[dict]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "inputWorkbook": rel(INPUT_OPERATOR_XLSX),
        "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
        "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
        "rows": len(operator_rows),
        "containsChannelRevenueModelSummary": True,
        "containsShelfStatus": True,
        "containsUserThresholdRating": True,
        "containsUserFeedbackColumns": True,
        "privateOnly": True,
        "sanitizedPublicSummary": True,
        "m3Entered": False,
    }


def write_public_reports(*reports: dict) -> None:
    paths = [
        (PER_CHANNEL_JSON, PER_CHANNEL_MD, "m2.per_channel_revenue_pattern_audit.v2", per_channel_md),
        (REVENUE_MODEL_JSON, REVENUE_MODEL_MD, "m2.revenue_model_classification.v2", revenue_model_md),
        (SHELF_JSON, SHELF_MD, "m2.shelf_status_inference.v1", shelf_md),
        (THRESHOLDS_JSON, THRESHOLDS_MD, "m2.rating_standard_v2_business_thresholds", thresholds_md),
        (RATING_JSON, RATING_MD, "m2.rating_calibration_v5_summary", rating_md),
        (SUGGESTION_JSON, SUGGESTION_MD, "m2.suggestion_calibration_v5_summary", suggestion_md),
        (OPERATOR_JSON, OPERATOR_MD, "m2.operator_task_pack_rating_standard_v2_summary", operator_md),
    ]
    for payload, (json_path, md_path, report_id, md_func) in zip(reports, paths):
        write_json(json_path, envelope(report_id, payload))
        write_text(md_path, md_func(payload))


def write_private_workbooks(
    channel_rows: list[dict],
    work_rows: list[dict],
    operator_rows: list[dict],
    per_channel_report: dict,
    revenue_model_report: dict,
    shelf_report: dict,
    threshold_report: dict,
    rating_report: dict,
    suggestion_report: dict,
) -> None:
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "00_阅读说明"
    append_rows(
        ws,
        [
            ["项目", "说明"],
            ["候选版本", CANDIDATE_VERSION],
            ["边界", "按作品+渠道+业务形态先判定收入模式，再聚合作品级；不使用合同商业条款。"],
            ["注意", "本文件包含本地真实作品信息，仅供用户复核，不提交。"],
        ],
    )
    add_summary_sheet(wb, "01_渠道收入模式", per_channel_report)
    add_summary_sheet(wb, "02_作品收入模式", revenue_model_report)
    add_summary_sheet(wb, "03_下架状态", shelf_report)
    add_summary_sheet(wb, "04_评级档位", threshold_report)
    add_summary_sheet(wb, "05_评级汇总", rating_report)
    add_summary_sheet(wb, "06_建议汇总", suggestion_report)
    add_dict_rows_sheet(wb, "07_作品匿名明细", anonymized_work_rows(work_rows[:800]), list(anonymized_work_rows(work_rows[:1])[0].keys()))
    finalize_workbook(wb)
    wb.save(PRIVATE_VALIDATION_XLSX)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "00_阅读说明"
    append_rows(ws2, [["项目", "说明"], ["用途", "30部任务包 v2，供用户复核收入模式、下架状态、评级和建议。"], ["边界", "private，不提交。"]])
    add_dict_rows_sheet(wb2, "01_运营任务卡", operator_rows, list(operator_rows[0].keys()) if operator_rows else [])
    finalize_workbook(wb2)
    wb2.save(PRIVATE_OPERATOR_XLSX)


def per_channel_md(payload: dict) -> str:
    rows = [{"model": model, "count": payload["channelRevenueModelDistribution"].get(model, 0)} for model in CHANNEL_MODELS]
    return "# M2 Per-Channel Revenue Pattern Audit v2\n\n" + markdown_table(rows, [("model", "Channel model"), ("count", "Count")])


def revenue_model_md(payload: dict) -> str:
    rows = [
        {
            "model": model,
            "count": payload["revenueModelDistribution"].get(model, 0),
            "share": payload["revenueShareByRevenueModel"].get(model, 0),
        }
        for model in WORK_MODELS
    ]
    return "# M2 Revenue Model Classification v2\n\n" + markdown_table(
        rows, [("model", "Revenue model"), ("count", "Works"), ("share", "Revenue share")]
    )


def shelf_md(payload: dict) -> str:
    rows = [{"status": status, "count": payload["shelfStatusDistribution"].get(status, 0)} for status in SHELF_STATUSES]
    return "# M2 Shelf Status Inference v1\n\n" + markdown_table(rows, [("status", "Shelf status"), ("count", "Works")])


def thresholds_md(payload: dict) -> str:
    rows = [{"rating": rating, "range": payload["userThresholds"][rating], "count": payload["salesPerformanceRatingDistribution"][rating]} for rating in RATINGS]
    return "# M2 Rating Standard v2 Business Thresholds\n\n" + markdown_table(
        rows, [("rating", "Rating"), ("range", "User range"), ("count", "Works")]
    )


def rating_md(payload: dict) -> str:
    rows = [{"rating": rating, "count": payload["historicalPerformanceRatingDistribution"].get(rating, 0)} for rating in RATINGS]
    return "# M2 Rating Calibration v5 Summary\n\n" + markdown_table(rows, [("rating", "Historical rating"), ("count", "Works")])


def suggestion_md(payload: dict) -> str:
    rows = [{"type": key, "count": value} for key, value in sorted(payload["suggestionTypeDistribution"].items())]
    return "# M2 Suggestion Calibration v5 Summary\n\n" + markdown_table(rows, [("type", "Suggestion type"), ("count", "Works")])


def operator_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 Operator Task Pack Rating Standard v2 Summary",
            "",
            f"- Rows: {payload['rows']}",
            f"- Private operator workbook: `{payload['privateOperatorWorkbook']}`",
            f"- Private validation workbook: `{payload['privateValidationWorkbook']}`",
            "- M3 entered: false",
        ]
    )


def continuity(values: np.ndarray) -> float:
    indexes = np.where(values > 0)[0]
    if len(indexes) <= 1:
        return 0.1 if len(indexes) == 1 else 0.0
    adjacent = sum(1 for idx in range(1, len(indexes)) if indexes[idx] - indexes[idx - 1] <= 1)
    return adjacent / max(1, len(indexes) - 1)


def variation_score(values: np.ndarray) -> float:
    if len(values) < 3:
        return 0.0
    mean = float(values.mean())
    if mean <= 0:
        return 0.0
    cv = float(values.std() / mean)
    if cv < 0.05:
        return 0.1
    return min(1.0, cv / 1.2)


def ratio(values: np.ndarray, predicate) -> float:
    if len(values) == 0:
        return 0.0
    return sum(1 for value in values if predicate(float(value))) / len(values)


def is_round_amount(value: float) -> bool:
    abs_value = abs(float(value))
    if abs_value < 10:
        return False
    return any(abs(abs_value / base - round(abs_value / base)) <= 0.01 for base in [10, 100, 1000])


def score_buyout_channel(**kwargs) -> float:
    score = 0.0
    large_share = kwargs["large_share"]
    if large_share >= 0.82:
        score += 0.3
    elif large_share >= 0.62:
        score += 0.2
    positive_count = kwargs["positive_count"]
    if positive_count <= 1:
        score += 0.22
    elif positive_count <= 3:
        score += 0.14
    if kwargs["round_ratio"] >= 0.5:
        score += 0.13
    if kwargs["integer_ratio"] >= 0.7:
        score += 0.09
    if kwargs["equal_split"]:
        score += 0.18
    if kwargs["adjacent_signal"]:
        score += 0.08
    if kwargs["post_no_sales_signal"]:
        score += 0.16
    if kwargs["post_positive_count"] > 0:
        score -= 0.24
    return max(0.0, min(1.0, score))


def score_sales_channel(**kwargs) -> float:
    score = 0.0
    score += kwargs["continuity_score"] * 0.42
    score += min(1.0, kwargs["active_ratio"] * 2.4) * 0.22
    score += kwargs["amount_variation"] * 0.18
    score += min(1.0, kwargs["positive_count"] / 12) * 0.12
    score += min(1.0, kwargs["nonstandard_ratio"]) * 0.06
    if kwargs["large_share"] > 0.9 and kwargs["tail_count"] < 3:
        score -= 0.18
    return max(0.0, min(1.0, score))


def rating_from_sales(amount: float | None) -> str:
    value = safe_float(amount)
    if value > 100000:
        return "S+"
    if value >= 10000:
        return "S"
    if value >= 5000:
        return "A"
    if value >= 1000:
        return "B"
    if value >= 500:
        return "C"
    if value >= 100:
        return "D"
    return "E"


def best_rating(ratings: list[str | None]) -> str:
    clean_ratings = [rating for rating in ratings if rating in RANK]
    if not clean_ratings:
        return "E"
    return sorted(clean_ratings, key=lambda rating: RANK[rating])[0]


def forecast_rating_from_summary(summary) -> str:
    rating = clean(summary.get("rating"))
    return rating if rating in RANK else "not_applicable"


def infer_forecastability_status(summary) -> str:
    output_type = clean(summary.get("forecastOutputType"))
    if output_type == "copyright_term_forecast":
        return "numeric_forecast_eligible"
    if output_type == "operating_window_forecast":
        return "conservative_numeric_forecast"
    if output_type == "observe_only":
        return "observe_only_no_numeric_forecast"
    status = clean(summary.get("forecastabilityStatus"))
    if status:
        return status
    forecast_base = safe_float(summary.get("forecastBase"), 0.0)
    forecast_months = safe_float(summary.get("remainingMonthsForForecast"), 0.0)
    manual_mode = clean(summary.get("manualReviewMode"))
    risk_severity = clean(summary.get("riskSeverity"))
    if forecast_base and forecast_base > 0 and forecast_months and forecast_months > 0:
        if manual_mode == "blocking" or risk_severity in {"high", "critical"}:
            return "conservative_numeric_forecast"
        return "numeric_forecast_eligible"
    return "observe_only_no_numeric_forecast"


def infer_forecast_confidence(summary) -> str:
    value = clean(summary.get("forecastConfidence") or summary.get("confidence"))
    if value:
        return value
    rating = clean(summary.get("rating"))
    if rating in {"S+", "S", "A"}:
        return "medium"
    if rating in {"D", "E"}:
        return "low"
    return "medium"


def threshold_width_check(work_rows: list[dict]) -> dict:
    amounts = [row["salesRevenue12m"] if row["salesRevenue12m"] > 0 else row["salesRevenueAnnualized"] for row in work_rows]
    positive = [amount for amount in amounts if amount > 0]
    if not positive:
        return {"positiveCount": 0, "recommendation": "no_positive_sales"}
    q = {f"p{int(p*100)}": round_float(float(np.quantile(positive, p)), 2) for p in [0.5, 0.75, 0.9, 0.95, 0.99]}
    recommendation = "use_user_thresholds_as_baseline"
    if q["p99"] < 10000:
        recommendation = "thresholds_may_be_too_high_for_current_sales_distribution"
    return {"positiveCount": len(positive), "quantiles": q, "recommendation": recommendation}


def distribution(rows: list[dict], field: str, keys: list[str]) -> dict:
    counter = Counter(row[field] for row in rows)
    return {key: int(counter.get(key, 0)) for key in keys}


def anonymized_work_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "匿名ID": f"W{index:04d}",
            "收入模式": row["revenueModelChinese"],
            "渠道数": row["channelCount"],
            "实销12月收入": row["salesRevenue12m"],
            "年化实销收入": row["salesRevenueAnnualized"],
            "买断估计金额": row["buyoutEstimatedAmount"],
            "实销评级": row["salesPerformanceRating"],
            "买断历史价值评级": row["buyoutHistoricalValueRating"],
            "历史表现评级": row["historicalPerformanceRating"],
            "下架状态": row["shelfStatusChinese"],
            "运营建议类型": row["suggestionType"],
        }
        for index, row in enumerate(rows, start=1)
    ]


def find_value(row: dict, preferred_headers: list[str]) -> str:
    for header in preferred_headers:
        if header in row and clean(row.get(header)):
            return clean(row.get(header))
    for key, value in row.items():
        key_text = clean(key)
        if any(header in key_text for header in preferred_headers) and clean(value):
            return clean(value)
    return ""


def add_summary_sheet(wb: Workbook, title: str, payload: dict) -> None:
    ws = wb.create_sheet(title[:31])
    append_rows(ws, [["指标", "值"], *flatten_payload(payload)])


def add_dict_rows_sheet(wb: Workbook, title: str, rows: list[dict], headers: list[str]) -> None:
    ws = wb.create_sheet(title[:31])
    append_rows(ws, [headers])
    for row in rows:
        append_rows(ws, [[row.get(header, "") for header in headers]])


def flatten_payload(payload: dict, prefix: str = "") -> list[list[str]]:
    rows = []
    for key, value in payload.items():
        label = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten_payload(value, label))
        elif isinstance(value, list):
            rows.append([label, json.dumps(value, ensure_ascii=False)])
        else:
            rows.append([label, value])
    return rows


def append_rows(ws, rows: list[list]) -> None:
    for row in rows:
        ws.append(row)


def finalize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column_cells in enumerate(ws.columns, start=1):
            max_len = min(54, max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2)
            ws.column_dimensions[get_column_letter(index)].width = max(12, max_len)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def envelope(report_id: str, payload: dict) -> dict:
    return {
        "reportId": report_id,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sanitized": True,
        "containsRealWorkTitles": False,
        "containsAuthors": False,
        "containsChannels": False,
        "containsRawBillingRows": False,
        "payload": json_safe(payload),
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    lines = ["| " + " | ".join(label for _, label in columns) + " |"]
    lines.append("|" + "|".join("---" for _ in columns) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |")
    return "\n".join(lines) + "\n\nNo real work titles, authors, channel names, or raw billing rows are included."


def month_distance(start: str, end: str) -> int | None:
    if not start or not end:
        return None
    try:
        sy, sm = [int(part) for part in start[:7].split("-")]
        ey, em = [int(part) for part in end[:7].split("-")]
        return (ey - sy) * 12 + (em - sm)
    except Exception:
        return None


def remaining_months_until(end_date: str, latest_month: str) -> int | None:
    if not end_date or not latest_month:
        return None
    try:
        end_year, end_month = [int(part) for part in end_date[:7].split("-")]
        latest_year, latest_month_number = [int(part) for part in latest_month[:7].split("-")]
        return (end_year - latest_year) * 12 + (end_month - latest_month_number)
    except Exception:
        return None


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def safe_float(value, default: float | None = 0.0) -> float | None:
    try:
        number = float(value)
    except Exception:
        return default
    return number if math.isfinite(number) else default


def round_float(value, digits: int = 4):
    try:
        number = float(value)
    except Exception:
        return None
    return round(number, digits) if math.isfinite(number) else None


def json_safe(value):
    if isinstance(value, dict):
        return {str(key): json_safe(child) for key, child in value.items()}
    if isinstance(value, list):
        return [json_safe(child) for child in value]
    if isinstance(value, tuple):
        return [json_safe(child) for child in value]
    if isinstance(value, set):
        return sorted(json_safe(child) for child in value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, np.ndarray):
        return [json_safe(item) for item in value.tolist()]
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


if __name__ == "__main__":
    main()
