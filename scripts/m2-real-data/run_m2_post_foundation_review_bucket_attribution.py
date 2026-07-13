from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts" / "m2-real-data"))
sys.path.insert(0, str(ROOT / "tools" / "m2-calibration"))

FIXTURE_SELF_TEST = "--fixture-self-test" in sys.argv

if not FIXTURE_SELF_TEST:
    from openpyxl import Workbook  # noqa: E402
    from openpyxl.formatting.rule import FormulaRule  # noqa: E402
    from openpyxl.styles import (  # noqa: E402
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

    import run_m2_post_foundation_readiness as readiness  # noqa: E402
    from run_nonformal_dry_run import load_analysis_inputs  # noqa: E402


PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-readiness"
PRIVATE_XLSX = (
    PRIVATE_DIR / "M2-post-foundation-review-buckets-user-confirmation-cn-v1.xlsx"
)
PRIVATE_JSON = (
    PRIVATE_DIR / "M2-post-foundation-review-buckets-user-confirmation-cn-v1.json"
)
STATUS_PATH = PRIVATE_DIR / "M2-status-local-staging-apply-result-cn-v1.json"
CORE_PATH = PRIVATE_DIR / "M2-five-source-local-staging-apply-result-cn-v1.json"
PUBLIC_JSON = (
    ROOT
    / "docs"
    / "analysis"
    / "m2-real-data"
    / "M2-post-foundation-review-bucket-attribution-summary-v1.json"
)
PUBLIC_MD = PUBLIC_JSON.with_suffix(".md")

EXPECTED_EXPIRED = 146
EXPECTED_SPARSE = 92

EXPIRED_OPTIONS = [
    "采用系统候选",
    "确认已续约并填写新到期时间",
    "确认是结算或渠道滞后",
    "确认是权利期外收入需审计",
    "仍不确定",
    "其他（备注说明）",
]
SPARSE_OPTIONS = [
    "采用系统候选",
    "确认保持已上架",
    "确认改为已下架",
    "确认版权有效但暂停运营",
    "仍不确定",
    "其他（备注说明）",
]

if not FIXTURE_SELF_TEST:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    SYSTEM_FILL = PatternFill("solid", fgColor="E2F0D9")
    USER_FILL = PatternFill("solid", fgColor="FFF2CC")
    INFO_FILL = PatternFill("solid", fgColor="D9EAF7")
    WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def canonical_work_id(value) -> str:
    text = clean(value)
    if text.upper().startswith("Y"):
        text = text[1:]
    match = re.search(r"\d+", text)
    return str(int(match.group(0))) if match else text


def month_distance(earlier: str, later: str) -> int | None:
    if not re.fullmatch(r"\d{4}-\d{2}", clean(earlier)) or not re.fullmatch(
        r"\d{4}-\d{2}", clean(later)
    ):
        return None
    early_year, early_month = map(int, earlier.split("-"))
    late_year, late_month = map(int, later.split("-"))
    return (late_year - early_year) * 12 + late_month - early_month


def date_month(value) -> str:
    text = clean(value)
    match = re.match(r"(\d{4})[-/.年](\d{1,2})", text)
    if not match:
        return ""
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def load_field_index(path: Path) -> dict[str, dict[str, str]]:
    index: dict[str, dict[str, str]] = defaultdict(dict)
    for row in read_json(path).get("records", []):
        work_id = canonical_work_id(row.get("作品编号"))
        field = clean(row.get("字段"))
        if work_id and field:
            index[work_id][field] = clean(row.get("采用值"))
    return dict(index)


def monthly_evidence(mapped_bill, latest_complete_month: str) -> dict[str, dict]:
    complete = mapped_bill[
        mapped_bill["validForCalibration"]
        & (mapped_bill["billMonth"] <= latest_complete_month)
    ].copy()
    grouped = (
        complete.groupby(["standardWorkId", "billMonth"], dropna=False)["amount"]
        .sum()
        .reset_index()
    )
    by_work: dict[str, dict[str, float]] = defaultdict(dict)
    for row in grouped.itertuples(index=False):
        work_id = canonical_work_id(row.standardWorkId)
        by_work[work_id][clean(row.billMonth)] = float(row.amount)
    result = {}
    for work_id, values in by_work.items():
        positive = {month: amount for month, amount in values.items() if amount > 0}
        last12 = sorted(values)[-12:]
        result[work_id] = {
            "monthly": values,
            "latestIncomeMonth": max(positive) if positive else "",
            "positiveMonthCount12": sum(values.get(month, 0) > 0 for month in last12),
            "positiveRevenueTotal": sum(positive.values()),
        }
    return result


def income_scale(share: float) -> str:
    if share <= 0:
        return "无到期后收入"
    if share <= 0.01:
        return "微量（不超过历史正收入的百分之一）"
    if share <= 0.10:
        return "少量（不超过历史正收入的十分之一）"
    return "明显（超过历史正收入的十分之一）"


def attribute_expired_case(case: dict) -> dict:
    post_months = int(case.get("postExpiryPositiveMonths") or 0)
    share = float(case.get("postExpiryRevenueShare") or 0)
    work_status = clean(case.get("workStatus"))
    months_after_expiry = case.get("monthsAfterExpiryToLatestIncome")

    if post_months == 0:
        return {
            "attribution": "收入发生在到期前，因近十二个月窗口与到期日重叠进入复核桶",
            "candidate": "保持版权已到期；无需调整",
        }
    if share <= 0.01 and (months_after_expiry is None or months_after_expiry <= 3):
        return {
            "attribution": "到期后仅有微量短尾收入，优先按结算或渠道滞后核查",
            "candidate": "按结算或渠道滞后处理；保持现有版权状态",
        }
    if work_status == "已上架" and (
        months_after_expiry is None or months_after_expiry >= 4
    ):
        return {
            "attribution": "作品仍标记已上架且到期后持续有收入，疑似续约信息尚未回填",
            "candidate": "核查是否已续约；如已续约请填写新的版权到期时间",
        }
    if work_status == "已下架" and share <= 0.10:
        return {
            "attribution": "作品已下架且到期后收入占比较低，优先核查尾款或渠道延迟",
            "candidate": "按结算或渠道滞后处理；保持版权已到期和已下架",
        }
    return {
        "attribution": "到期后收入较明显或持续时间较长，不能仅按尾款自动关闭",
        "candidate": "发起权利期外收入审计；确认前不修改版权状态",
    }


def attribute_sparse_case(case: dict) -> dict:
    work_status = clean(case.get("workStatus"))
    months_since_income = case.get("monthsSinceLatestIncome")
    if work_status == "已下架":
        return {
            "attribution": "状态表已确认下架，版权仍有效不等于当前仍在运营",
            "candidate": "保持版权有效；作品状态保持已下架",
        }
    if work_status == "已上架" and months_since_income is not None:
        if months_since_income <= 3:
            return {
                "attribution": "作品已上架且近期仍有偶发收入，属于低频实销",
                "candidate": "保持已上架；标记为低频收入并继续观察",
            }
        if months_since_income <= 6:
            return {
                "attribution": "作品已上架，但收入间隔偏长，暂不足以自动判断下架",
                "candidate": "保持已上架；继续观察后续账单",
            }
        return {
            "attribution": "作品标记已上架，但较长时间未再产生收入，需要确认实际货架状态",
            "candidate": "确认是否仍上架；系统暂保留已上架状态",
        }
    return {
        "attribution": "已有收入信号，但作品状态证据不足",
        "candidate": "人工确认作品当前是已上架、已下架还是暂停运营",
    }


def build_case_rows() -> tuple[list[dict], list[dict], dict]:
    foundation, _ = readiness.load_foundation()
    core = load_field_index(CORE_PATH)
    status = load_field_index(STATUS_PATH)
    raw_mapping, standard_mapping = readiness.load_historical_mappings()
    context = load_analysis_inputs()
    mapped_bill, scope = readiness.apply_foundation_scope(
        context["bill"], set(foundation), raw_mapping, standard_mapping
    )
    work_rows, _ = readiness.build_current_work_rows(context, mapped_bill, foundation)
    evidence = monthly_evidence(mapped_bill, context["latest_complete_month"])

    expired_rows = []
    sparse_rows = []
    for work in work_rows:
        prompts = set(work.get("shelfStatusReviewPrompts", []))
        if not prompts:
            continue
        work_id = canonical_work_id(work.get("standardWorkId"))
        base = foundation[work_id]
        fields = core.get(work_id, {})
        states = status.get(work_id, {})
        monthly = evidence.get(work_id, {})
        latest_income = clean(monthly.get("latestIncomeMonth"))
        expiry = clean(fields.get("版权到期"))
        expiry_month = date_month(expiry)
        monthly_values = monthly.get("monthly", {})
        post_expiry = {
            month: amount
            for month, amount in monthly_values.items()
            if expiry_month and month > expiry_month and amount > 0
        }
        positive_total = float(monthly.get("positiveRevenueTotal") or 0)
        post_total = sum(post_expiry.values())
        post_share = post_total / positive_total if positive_total > 0 else 0.0
        common = {
            "workId": work_id,
            "title": clean(base.get("书名")),
            "workStatus": clean(states.get("作品状态")),
            "audioRightsStatus": clean(states.get("音频版权状态")),
            "copyrightEnd": expiry,
            "latestIncomeMonth": latest_income,
            "positiveMonthCount12": int(monthly.get("positiveMonthCount12") or 0),
        }

        if "expired_with_tail_revenue_review" in prompts:
            case = {
                **common,
                "postExpiryPositiveMonths": len(post_expiry),
                "postExpiryRevenueShare": round(post_share, 6),
                "postExpiryRevenueScale": income_scale(post_share),
                "monthsAfterExpiryToLatestIncome": month_distance(
                    expiry_month, latest_income
                ),
            }
            case.update(attribute_expired_case(case))
            expired_rows.append(case)

        if "active_rights_sparse_revenue_review" in prompts:
            case = {
                **common,
                "monthsSinceLatestIncome": month_distance(
                    latest_income, context["latest_complete_month"]
                ),
            }
            case.update(attribute_sparse_case(case))
            sparse_rows.append(case)

    if len(expired_rows) != EXPECTED_EXPIRED or len(sparse_rows) != EXPECTED_SPARSE:
        raise SystemExit(
            "Review bucket membership changed unexpectedly: "
            f"expired={len(expired_rows)}, sparse={len(sparse_rows)}."
        )
    return expired_rows, sparse_rows, {
        "latestCompleteMonth": context["latest_complete_month"],
        "scope": scope,
    }


def style_table(sheet, widths: list[int], user_columns: set[int]) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = USER_FILL if index in user_columns else SYSTEM_FILL
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions


def add_dropdown(sheet, column_letter: str, row_count: int, option_range: str) -> None:
    validation = DataValidation(type="list", formula1=option_range, allow_blank=True)
    validation.error = "请从下拉候选中选择；其他情况请选择“其他（备注说明）”。"
    validation.errorTitle = "确认项不在候选范围"
    validation.prompt = "选择“采用系统候选”或其他业务确认项。"
    validation.promptTitle = "请选择确认结果"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{row_count + 1}")


def write_workbook(expired_rows: list[dict], sparse_rows: list[dict]) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "00_填写说明"
    readme.merge_cells("A1:H1")
    readme["A1"] = "M2 两类复核桶人工确认表"
    readme["A1"].fill = HEADER_FILL
    readme["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    readme["A1"].alignment = Alignment(horizontal="center", vertical="center")
    readme.row_dimensions[1].height = 32
    instructions = [
        ("你需要填写什么", "只填写黄色列：“你的确认”、必要时填写新版权到期时间和备注。"),
        ("如何采用系统结论", "确认系统判断合理时，在“你的确认”选择“采用系统候选”。"),
        ("到期仍有收入", "重点确认是续约未回填、结算/渠道滞后，还是权利期外收入需要审计。"),
        ("版权有效但收入稀疏", "重点确认作品当前是已上架、已下架，还是版权有效但暂停运营。"),
        ("不会发生的操作", "本表填写不会自动写正式主数据，不会自动激活映射，也不会进入 M3。"),
        ("运营建议", "M2 已确定不输出运营建议；本表只有状态与权利复核候选。"),
    ]
    for row_index, (title, detail) in enumerate(instructions, start=3):
        readme.cell(row_index, 1, title)
        readme.cell(row_index, 2, detail)
        readme.cell(row_index, 1).font = Font(bold=True)
        readme.cell(row_index, 1).fill = INFO_FILL
        readme.cell(row_index, 2).fill = INFO_FILL
        readme.cell(row_index, 1).border = THIN_BORDER
        readme.cell(row_index, 2).border = THIN_BORDER
        readme.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 92

    option_sheet = workbook.create_sheet("下拉选项")
    option_sheet.append(["到期仍有收入确认项", "收入稀疏确认项"])
    for index in range(max(len(EXPIRED_OPTIONS), len(SPARSE_OPTIONS))):
        option_sheet.append(
            [
                EXPIRED_OPTIONS[index] if index < len(EXPIRED_OPTIONS) else "",
                SPARSE_OPTIONS[index] if index < len(SPARSE_OPTIONS) else "",
            ]
        )
    option_sheet.sheet_state = "hidden"

    expired_sheet = workbook.create_sheet("01_到期仍有收入")
    expired_headers = [
        "序号",
        "作品编号",
        "书名",
        "当前作品状态",
        "当前音频版权状态",
        "版权到期时间",
        "最后收入月份",
        "到期后有收入月份数",
        "到期后收入规模",
        "系统归因",
        "系统候选",
        "你的确认",
        "新的版权到期时间（如续约）",
        "你的备注",
    ]
    expired_sheet.append(expired_headers)
    for index, row in enumerate(expired_rows, start=1):
        expired_sheet.append(
            [
                index,
                row["workId"],
                row["title"],
                row["workStatus"],
                row["audioRightsStatus"],
                row["copyrightEnd"],
                row["latestIncomeMonth"],
                row["postExpiryPositiveMonths"],
                row["postExpiryRevenueScale"],
                row["attribution"],
                row["candidate"],
                "",
                "",
                "",
            ]
        )
    style_table(
        expired_sheet,
        [8, 14, 30, 15, 18, 16, 14, 18, 24, 45, 45, 24, 25, 35],
        {12, 13, 14},
    )
    add_dropdown(
        expired_sheet,
        "L",
        len(expired_rows),
        f"'下拉选项'!$A$2:$A${len(EXPIRED_OPTIONS) + 1}",
    )

    sparse_sheet = workbook.create_sheet("02_版权有效收入稀疏")
    sparse_headers = [
        "序号",
        "作品编号",
        "书名",
        "当前作品状态",
        "当前音频版权状态",
        "最后收入月份",
        "距最后收入月数",
        "近十二个月有收入月份数",
        "系统归因",
        "系统候选",
        "你的确认",
        "你的备注",
    ]
    sparse_sheet.append(sparse_headers)
    for index, row in enumerate(sparse_rows, start=1):
        sparse_sheet.append(
            [
                index,
                row["workId"],
                row["title"],
                row["workStatus"],
                row["audioRightsStatus"],
                row["latestIncomeMonth"],
                row["monthsSinceLatestIncome"],
                row["positiveMonthCount12"],
                row["attribution"],
                row["candidate"],
                "",
                "",
            ]
        )
    style_table(
        sparse_sheet,
        [8, 14, 30, 15, 18, 14, 16, 22, 45, 45, 24, 35],
        {11, 12},
    )
    add_dropdown(
        sparse_sheet,
        "K",
        len(sparse_rows),
        f"'下拉选项'!$B$2:$B${len(SPARSE_OPTIONS) + 1}",
    )

    for sheet, column in [(expired_sheet, "L"), (sparse_sheet, "K")]:
        sheet.conditional_formatting.add(
            f"{column}2:{column}{sheet.max_row}",
            FormulaRule(formula=[f'LEN({column}2)=0'], fill=WARNING_FILL),
        )

    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(PRIVATE_XLSX)


def sanitized_summary(expired_rows: list[dict], sparse_rows: list[dict], run: dict) -> dict:
    return {
        "schema": "m2.post_foundation_review_bucket_attribution_summary.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "scope": {
            "workCount": run["scope"]["foundationWorkCount"],
            "expiredWithRevenueReviewCount": len(expired_rows),
            "activeRightsSparseRevenueReviewCount": len(sparse_rows),
            "totalReviewCount": len(expired_rows) + len(sparse_rows),
            "latestCompleteMonth": run["latestCompleteMonth"],
        },
        "expiredWithRevenue": {
            "attributionDistribution": dict(Counter(row["attribution"] for row in expired_rows)),
            "candidateDistribution": dict(Counter(row["candidate"] for row in expired_rows)),
        },
        "activeRightsSparseRevenue": {
            "attributionDistribution": dict(Counter(row["attribution"] for row in sparse_rows)),
            "candidateDistribution": dict(Counter(row["candidate"] for row in sparse_rows)),
        },
        "userConfirmation": {
            "required": True,
            "confirmedCount": 0,
            "remainingCount": len(expired_rows) + len(sparse_rows),
            "privatePackReady": True,
        },
        "productDecision": {
            "automaticOperatingSuggestionsEnabled": False,
            "reviewPromptsOnly": True,
        },
        "formalAuthorization": {
            "m2FormalOperationsGranted": True,
            "executionPendingReviewConfirmation": True,
            "m3FormalExecutionGranted": False,
        },
        "privacy": {
            "publicReportAggregateOnly": True,
            "privateRowsCommitted": False,
            "privateOutputRole": "gitignored_local_only",
        },
    }


def render_markdown(summary: dict) -> str:
    scope = summary["scope"]
    expired = summary["expiredWithRevenue"]
    sparse = summary["activeRightsSparseRevenue"]
    lines = [
        "# M2 post-foundation 两类复核桶自动归因摘要 v1",
        "",
        "## 结论",
        "",
        f"- 已对 `{scope['expiredWithRevenueReviewCount']}` 个到期仍有收入样本和 `{scope['activeRightsSparseRevenueReviewCount']}` 个版权有效但收入稀疏样本完成系统初步归因。",
        "- 已生成 Git 忽略的中文人工确认表；逐作品结果尚未应用到正式主数据或正式评估。",
        "- M2 不输出运营建议，本次仅提供状态、权利和收入事实的复核候选。",
        "- 用户已授权 M2 formal 操作，但执行仍等待本复核表确认和逐作品输入内容契约通过；M3 formal 未获授权。",
        "",
        "## 到期仍有收入",
        "",
        f"- 自动归因分布：`{json.dumps(expired['attributionDistribution'], ensure_ascii=False)}`",
        f"- 系统候选分布：`{json.dumps(expired['candidateDistribution'], ensure_ascii=False)}`",
        "",
        "## 版权有效但收入稀疏",
        "",
        f"- 自动归因分布：`{json.dumps(sparse['attributionDistribution'], ensure_ascii=False)}`",
        f"- 系统候选分布：`{json.dumps(sparse['candidateDistribution'], ensure_ascii=False)}`",
        "",
        "## 边界",
        "",
        "- 公共报告不包含书名、作者、作品编号、渠道或逐作品收入。",
        "- private 表格只留在 Git 忽略区域，用户确认前不写正式主数据、不激活映射、不执行 formal evaluation。",
        "- 两类复核完成后，系统再执行正式输入快照、mapping、formal evaluation 和 task/export/release/audit 的受控链路。",
    ]
    return "\n".join(lines) + "\n"


def fixture_self_test() -> dict:
    expired = [
        attribute_expired_case(
            {
                "postExpiryPositiveMonths": 0,
                "postExpiryRevenueShare": 0,
                "workStatus": "已下架",
                "monthsAfterExpiryToLatestIncome": -1,
            }
        ),
        attribute_expired_case(
            {
                "postExpiryPositiveMonths": 1,
                "postExpiryRevenueShare": 0.005,
                "workStatus": "已下架",
                "monthsAfterExpiryToLatestIncome": 2,
            }
        ),
        attribute_expired_case(
            {
                "postExpiryPositiveMonths": 5,
                "postExpiryRevenueShare": 0.2,
                "workStatus": "已上架",
                "monthsAfterExpiryToLatestIncome": 8,
            }
        ),
    ]
    sparse = [
        attribute_sparse_case({"workStatus": "已下架", "monthsSinceLatestIncome": 5}),
        attribute_sparse_case({"workStatus": "已上架", "monthsSinceLatestIncome": 2}),
        attribute_sparse_case({"workStatus": "已上架", "monthsSinceLatestIncome": 8}),
    ]
    return {
        "fixtureSelfTest": True,
        "expiredBranches": len({row["attribution"] for row in expired}),
        "sparseBranches": len({row["attribution"] for row in sparse}),
        "noAutomaticFinalDecision": True,
        "automaticOperatingSuggestionsEnabled": False,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--fixture-self-test", action="store_true")
    args = parser.parse_args()
    if args.fixture_self_test:
        print(json.dumps(fixture_self_test(), ensure_ascii=False))
        return

    required = [
        readiness.FOUNDATION_PATH,
        CORE_PATH,
        STATUS_PATH,
        readiness.MAPPING_PAYLOAD,
        readiness.MAPPING_OVERLAY,
    ]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
    if missing:
        raise SystemExit("缺少生成复核表所需的本地 private 输入：" + "，".join(missing))

    expired_rows, sparse_rows, run = build_case_rows()
    write_workbook(expired_rows, sparse_rows)
    summary = sanitized_summary(expired_rows, sparse_rows, run)
    PRIVATE_JSON.write_text(
        json.dumps(
            {
                "summary": summary,
                "expiredWithRevenue": expired_rows,
                "activeRightsSparseRevenue": sparse_rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    PUBLIC_JSON.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_JSON.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    PUBLIC_MD.write_text(render_markdown(summary), encoding="utf-8")
    print(
        json.dumps(
            {
                "status": "ready_for_user_confirmation",
                "expiredWithRevenue": len(expired_rows),
                "activeRightsSparseRevenue": len(sparse_rows),
                "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)).replace("\\", "/"),
                "publicReport": str(PUBLIC_MD.relative_to(ROOT)).replace("\\", "/"),
                "formalDatabaseWritten": False,
                "mappingActivated": False,
                "m3Entered": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
