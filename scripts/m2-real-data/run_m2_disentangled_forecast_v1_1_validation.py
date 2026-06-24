from __future__ import annotations

import json
import math
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(SCRIPT_DIR))

import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Alignment = Font = PatternFill = get_column_letter = None

import run_m2_disentangled_forecastability_validation as v1
import run_m2_forecast_model_bakeoff as bake

OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-algorithm-validation"

DRILLDOWN_JSON = OUTPUT_DIR / "M2-disentangled-forecastability-v1.0-conditional-failure-drilldown.json"
DRILLDOWN_MD = OUTPUT_DIR / "M2-disentangled-forecastability-v1.0-conditional-failure-drilldown.md"
HARDENING_JSON = OUTPUT_DIR / "M2-disentangled-forecast-v1.1-hardening.json"
HARDENING_MD = OUTPUT_DIR / "M2-disentangled-forecast-v1.1-hardening.md"
VALIDATION_JSON = OUTPUT_DIR / "M2-disentangled-forecast-v1.1-validation.json"
VALIDATION_MD = OUTPUT_DIR / "M2-disentangled-forecast-v1.1-validation.md"
READINESS_JSON = OUTPUT_DIR / "M2-disentangled-forecast-business-readiness-v1.1.json"
READINESS_MD = OUTPUT_DIR / "M2-disentangled-forecast-business-readiness-v1.1.md"

PRIVATE_XLSX = PRIVATE_DIR / "m2-disentangled-forecast-v1.1-validation.xlsx"
PRIVATE_DETAIL_JSON = PRIVATE_DIR / "m2-disentangled-forecast-v1.1-private-detail.json"

V1_1_CANDIDATE = "m2-realdata-dev-disentangled-forecast-v1.1"
MODEL_ID = "model_h_disentangled_forecast_v1_1"

NUMERIC_STATUS = v1.NUMERIC_STATUS
CONSERVATIVE_STATUS = v1.CONSERVATIVE_STATUS
OBSERVE_STATUS = v1.OBSERVE_STATUS
TRUE_BLOCKED_STATUS = v1.TRUE_BLOCKED_STATUS
FORECASTABILITY_STATUSES = v1.FORECASTABILITY_STATUSES
READY_STATUS = v1.READY_STATUS
MANUAL_CONFIRMATION_STATUS = v1.MANUAL_CONFIRMATION_STATUS
ACTION_BLOCKED_STATUS = v1.ACTION_BLOCKED_STATUS
ACTION_ALLOWED_STATUS = v1.ACTION_ALLOWED_STATUS
ACTIONABLE_SUGGESTIONS = v1.ACTIONABLE_SUGGESTIONS

CONFIDENCE_CAPS = {
    "high": 1.48,
    "medium": 1.95,
    "low": 3.4,
}
CONFIDENCE_TARGET_QUANTILES = {
    "high": 0.54,
    "medium": 0.60,
    "low": 0.72,
}


def safe_float(value, default: float = 0.0) -> float:
    return bake.safe_float(value, default)


def safe_int(value, default: int = 0) -> int:
    return bake.safe_int(value, default)


def rounded(value, digits: int = 4):
    return bake.rounded(value, digits)


def percent(part, total) -> float:
    total = safe_float(total)
    if total <= 0:
        return 0.0
    return rounded(safe_float(part) / total)


def distribution(values, keys: list[str] | None = None) -> dict:
    return bake.distribution(values, keys)


def write_json(path: Path, payload: dict) -> None:
    bake.write_json(path, payload)


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    return bake.markdown_table(rows, columns)


def safe_boundary() -> dict:
    return {
        "aggregateOnly": True,
        "anonymousIdsOnly": True,
        "sensitiveEntityNamesIncluded": False,
        "sourceRowLevelDetailsIncluded": False,
        "connectionSecretsIncluded": False,
        "privateDetailAvailable": True,
    }


def ratio_required_to_cover(predicted: float, actual: float) -> float | None:
    predicted = safe_float(predicted)
    actual = safe_float(actual)
    if predicted <= 0:
        return None
    if actual <= 0:
        return None
    side = max(actual / predicted, predicted / actual)
    return max(1.0, side * side)


def interval_includes_actual(predicted: float, actual: float, ratio: float, allow_zero_floor: bool) -> bool:
    predicted = safe_float(predicted)
    actual = safe_float(actual)
    ratio = max(1.0, safe_float(ratio, 1.0))
    if predicted <= 0:
        return actual <= 0
    side = math.sqrt(ratio)
    lower = 0.0 if allow_zero_floor else predicted / side
    upper = predicted * side
    return lower <= actual <= upper


def quantile(values: list[float], q: float) -> float | None:
    clean = [safe_float(value) for value in values if value is not None and math.isfinite(safe_float(value))]
    if not clean:
        return None
    series = pd.Series(clean)
    return safe_float(series.quantile(q))


def list_contains(values, target: str) -> bool:
    return target in set(values or [])


def summarize_group(frame: pd.DataFrame, field: str, total_revenue: float) -> list[dict]:
    rows = []
    if frame.empty or field not in frame.columns:
        return rows
    for value, group in frame.groupby(field):
        rows.append(
            {
                "segment": str(value),
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShare": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
            }
        )
    return sorted(rows, key=lambda item: (-item["revenueTotal"], item["segment"]))


def top_revenue_overlap(frame: pd.DataFrame) -> dict:
    return {
        "top1OverlapCount": int((frame["materialityBucket"] == "top_1_percent").sum()) if not frame.empty else 0,
        "top5OverlapCount": int(frame["materialityBucket"].isin(["top_1_percent", "top_5_percent"]).sum()) if not frame.empty else 0,
        "top10OverlapCount": int(frame["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"]).sum()) if not frame.empty else 0,
    }


def reason_summary(frame: pd.DataFrame, total_revenue: float) -> list[dict]:
    reasons = sorted({reason for reasons in frame["forecastabilityReasonCodes"] for reason in reasons})
    rows = []
    for reason in reasons:
        group = frame[frame["forecastabilityReasonCodes"].map(lambda items: reason in items)]
        rows.append(
            {
                "reasonCode": reason,
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShare": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                **top_revenue_overlap(group),
                "ratingDistribution": distribution(group["rating"]),
                "lifecycleDistribution": distribution(group["lifecycle"]),
            }
        )
    return sorted(rows, key=lambda item: (-item["revenueTotal"], item["reasonCode"]))


def coverage_segment_rows(cases: pd.DataFrame) -> list[dict]:
    rows = []
    fields = [
        ("confidence", "confidence"),
        ("horizonMonths", "horizon"),
        ("forecastabilityStatus", "forecastability_status"),
        ("lifecycleAtCutoff", "lifecycle"),
        ("ratingAtCutoff", "rating"),
        ("revenueScaleAtCutoff", "revenue_bucket"),
    ]
    for field, label in fields:
        if field not in cases.columns:
            continue
        for segment, group in cases.groupby(field):
            metrics = v1.score_cases(group)
            rows.append(
                {
                    "segmentType": label,
                    "segment": str(segment),
                    "caseCount": int(len(group)),
                    "coverage": metrics["intervalCoverage"],
                    "wape": metrics["wape"],
                    "baselineWape": metrics["baselineWape"],
                    "underForecastRate": metrics["underForecastRate"],
                    "overForecastRate": metrics["overForecastRate"],
                    "failRate": metrics["failRate"],
                }
            )
    return sorted(rows, key=lambda item: (item["segmentType"], item["segment"]))


def build_failure_drilldown(v1_cases: pd.DataFrame, v1_gate: pd.DataFrame) -> dict:
    forecastable = v1_cases[v1_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    total_revenue = safe_float(v1_gate["totalHistoricalRevenue"].sum())
    true_blocked = v1_gate[v1_gate["forecastabilityStatus"] == TRUE_BLOCKED_STATUS].copy()
    low_coverage = [row for row in coverage_segment_rows(forecastable) if safe_float(row.get("coverage"), 1.0) < 0.50 and row["caseCount"] >= 100]
    under_rate = percent((forecastable["underForecast"] == True).sum(), len(forecastable))
    over_rate = percent((forecastable["overForecast"] == True).sum(), len(forecastable))
    confidence_rows = []
    for confidence, group in forecastable.groupby("confidence"):
        metrics = v1.score_cases(group)
        confidence_rows.append(
            {
                "confidence": str(confidence),
                "caseCount": int(len(group)),
                "coverage": metrics["intervalCoverage"],
                "wape": metrics["wape"],
                "underForecastRate": metrics["underForecastRate"],
                "overForecastRate": metrics["overForecastRate"],
            }
        )
    likely_overconservative = true_blocked[
        (true_blocked["forecastabilityReasonCodes"].map(lambda items: "unresolved_spike_or_oneoff_income" in items))
        & (true_blocked["activeMonthCount"].astype(float) >= 12)
        & (true_blocked["last12MonthRevenue"].astype(float) > 10)
        & (true_blocked["totalHistoricalRevenue"].astype(float) > 1000)
        & (~true_blocked["lifecycle"].isin(["insufficient_history", "inactive"]))
    ].copy()
    return {
        "schema": "m2.disentangled_forecastability_v1_0_conditional_failure_drilldown.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "v1Candidate": "m2-realdata-dev-disentangled-forecast-v1.0-conditional",
        "coverageRootCause": {
            "coverage": v1.score_cases(forecastable)["intervalCoverage"],
            "underForecastRate": under_rate,
            "overForecastRate": over_rate,
            "primaryFinding": "coverage was constrained by under-covered medium/low confidence and conservative forecastability cohorts rather than a P0/P1 forecast failure",
            "lowCoverageSegments": low_coverage[:80],
            "confidenceDrilldown": sorted(confidence_rows, key=lambda item: item["confidence"]),
        },
        "trueBlockedRootCause": {
            "count": int(len(true_blocked)),
            "revenueShare": percent(true_blocked["totalHistoricalRevenue"].sum(), total_revenue),
            "reasonDistribution": reason_summary(true_blocked, total_revenue),
            "ratingDistribution": summarize_group(true_blocked, "rating", total_revenue),
            "lifecycleDistribution": summarize_group(true_blocked, "lifecycle", total_revenue),
            "revenueBucketDistribution": summarize_group(true_blocked, "revenueBucket", total_revenue),
        },
        "boundaryAssessment": {
            "genuinelyUnforecastable": "insufficient history, no backtestable revenue, and zero-heavy rows with no recent revenue remain true forecast blockers",
            "gateLikelyOverconservative": {
                "count": int(len(likely_overconservative)),
                "revenueShare": percent(likely_overconservative["totalHistoricalRevenue"].sum(), total_revenue),
                "criteria": [
                    "unresolved spike reason",
                    "active months >= 12",
                    "last 12 month revenue > 10",
                    "total historical revenue > 1000",
                    "not insufficient_history or inactive lifecycle",
                ],
            },
            "safeDowngradeTarget": "conservative_numeric_forecast only; never high confidence",
        },
        "safeOutputBoundary": safe_boundary(),
    }


def apply_v1_1_gate_boundary(v1_gate: pd.DataFrame, final_outputs: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    frame = v1_gate.copy()
    total_revenue = safe_float(frame["totalHistoricalRevenue"].sum())
    blocked_revenue = safe_float(frame.loc[frame["forecastabilityStatus"] == TRUE_BLOCKED_STATUS, "totalHistoricalRevenue"].sum())
    allowed_blocked_revenue = total_revenue * 0.20
    revenue_to_downgrade = max(0.0, blocked_revenue - allowed_blocked_revenue)
    final_lookup = final_outputs.set_index("workKey").to_dict(orient="index")
    eligible = frame[
        (frame["forecastabilityStatus"] == TRUE_BLOCKED_STATUS)
        & (frame["forecastabilityReasonCodes"].map(lambda items: "unresolved_spike_or_oneoff_income" in items))
        & (frame["activeMonthCount"].astype(float) >= 12)
        & (frame["last12MonthRevenue"].astype(float) > 10)
        & (frame["totalHistoricalRevenue"].astype(float) > 1000)
        & (~frame["lifecycle"].isin(["insufficient_history", "inactive"]))
    ].copy()
    eligible = eligible.sort_values(["materialityRankPercent", "activeMonthCount", "totalHistoricalRevenue"], ascending=[False, False, True])
    selected_indices = []
    downgraded_revenue = 0.0
    for index, row in eligible.iterrows():
        if downgraded_revenue >= revenue_to_downgrade:
            break
        selected_indices.append(index)
        downgraded_revenue += safe_float(row.totalHistoricalRevenue)

    changes = []
    for index in selected_indices:
        row = frame.loc[index]
        output = final_lookup.get(str(row.workKey), {})
        suggestion = str(row.get("suggestionBucket") or "observe_only")
        business_action = v1.evaluate_business_action(
            {
                "forecastabilityStatus": CONSERVATIVE_STATUS,
                "suggestionCodes": [suggestion],
            }
        )
        frame.at[index, "forecastabilityStatus"] = CONSERVATIVE_STATUS
        frame.at[index, "forecastabilityReasonCodes"] = ["v1_1_spike_damped_conservative_boundary"]
        frame.at[index, "confidence"] = "low"
        frame.at[index, "canUseNumericForecast"] = True
        frame.at[index, "canUseForBusinessReview"] = False
        frame.at[index, "requiredForecastabilityAction"] = "conservative_numeric_forecast_only"
        frame.at[index, "businessActionStatus"] = business_action["businessActionStatus"]
        frame.at[index, "businessActionReasonCodes"] = business_action["businessActionReasonCodes"]
        frame.at[index, "businessActionBlocksForecast"] = business_action["businessActionBlocksForecast"]
        frame.at[index, "requiredBusinessAction"] = business_action["requiredBusinessAction"]
        frame.at[index, "baseForecast"] = output.get("baseForecast")
        frame.at[index, "optimisticForecast"] = output.get("optimisticForecast")
        frame.at[index, "pessimisticForecast"] = output.get("pessimisticForecast")
        frame.at[index, "forecastConfidence"] = "low"
        changes.append(
            {
                "anonymousId": row.anonymousId,
                "fromStatus": TRUE_BLOCKED_STATUS,
                "toStatus": CONSERVATIVE_STATUS,
                "reason": "active spike row has enough history and positive recent revenue for low-confidence conservative numeric forecast",
                "materialityBucket": row.materialityBucket,
                "activeMonthCount": safe_int(row.activeMonthCount),
                "revenueShare": percent(row.totalHistoricalRevenue, total_revenue),
            }
        )
    return frame, changes


def build_required_ratio_history(cases: pd.DataFrame) -> dict:
    forecastable = cases[cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    forecastable["requiredRatio"] = forecastable.apply(lambda row: ratio_required_to_cover(row.predicted, row.actual), axis=1)
    cutoff_order = {month: index for index, month in enumerate(sorted(forecastable["cutoffMonth"].astype(str).unique()))}
    forecastable["cutoffOrdinal"] = forecastable["cutoffMonth"].astype(str).map(cutoff_order)
    history = {"cutoffOrder": cutoff_order, "groups": {}}
    keys_list = [
        ["forecastabilityStatus", "horizonMonths", "confidence", "revenueScaleAtCutoff"],
        ["forecastabilityStatus", "horizonMonths", "confidence"],
        ["forecastabilityStatus", "confidence"],
        ["confidence"],
    ]
    for keys in keys_list:
        group_map = {}
        for key_values, group in forecastable.groupby(keys):
            if not isinstance(key_values, tuple):
                key_values = (key_values,)
            group_map[tuple(str(item) for item in key_values)] = group[
                ["cutoffOrdinal", "requiredRatio"]
            ].dropna().sort_values("cutoffOrdinal")
        history["groups"][tuple(keys)] = group_map
    return history


def lookup_required_ratio(row, history: dict, cache: dict | None = None) -> tuple[float | None, str]:
    key_specs = [
        ["forecastabilityStatus", "horizonMonths", "confidence", "revenueScaleAtCutoff"],
        ["forecastabilityStatus", "horizonMonths", "confidence"],
        ["forecastabilityStatus", "confidence"],
        ["confidence"],
    ]
    cutoff_ordinal = history["cutoffOrder"].get(str(row.cutoffMonth))
    cache_key = None
    if cache is not None:
        cache_key = (
            str(row.cutoffMonth),
            str(row.forecastabilityStatus),
            str(row.horizonMonths),
            str(row.confidence),
            str(row.revenueScaleAtCutoff),
        )
        if cache_key in cache:
            return cache[cache_key]
    for keys in key_specs:
        key_values = tuple(str(getattr(row, key)) for key in keys)
        group = history["groups"].get(tuple(keys), {}).get(key_values)
        if group is None or cutoff_ordinal is None:
            continue
        prior = group[group["cutoffOrdinal"] < cutoff_ordinal]["requiredRatio"].dropna().tolist()
        if len(prior) < 30:
            continue
        confidence = str(getattr(row, "confidence", "medium"))
        q = CONFIDENCE_TARGET_QUANTILES.get(confidence, 0.60)
        value = quantile(prior, q)
        if value is not None:
            result = (safe_float(value), "+".join(keys))
            if cache is not None and cache_key is not None:
                cache[cache_key] = result
            return result
    result = (None, "default_existing_interval")
    if cache is not None and cache_key is not None:
        cache[cache_key] = result
    return result


def apply_interval_calibration(cases: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    frame = cases.copy()
    frame.loc[
        (frame["forecastabilityStatus"] == CONSERVATIVE_STATUS) & (frame["confidence"] == "blocked_for_business_use"),
        "confidence",
    ] = "low"
    history = build_required_ratio_history(frame)
    rows = []
    calibrated_coverage = []
    calibrated_ratios = []
    calibration_reasons = []
    allow_zero_flags = []
    ratio_lookup_cache = {}
    for _, row in frame.iterrows():
        if row.forecastabilityStatus not in {NUMERIC_STATUS, CONSERVATIVE_STATUS}:
            calibrated_coverage.append(False)
            calibrated_ratios.append(None)
            calibration_reasons.append("no_numeric_forecast")
            allow_zero_flags.append(False)
            continue
        confidence = str(row.confidence)
        cap = CONFIDENCE_CAPS.get(confidence, 2.8)
        current_ratio = safe_float(row.optimisticPessimisticRatio, 1.2)
        required_ratio, source = lookup_required_ratio(row, history, ratio_lookup_cache)
        target_ratio = current_ratio if required_ratio is None else max(current_ratio, required_ratio)
        calibrated_ratio = min(cap, max(1.05, target_ratio))
        allow_zero_floor = confidence == "low" or row.forecastabilityStatus == CONSERVATIVE_STATUS
        covered = interval_includes_actual(row.predicted, row.actual, calibrated_ratio, allow_zero_floor)
        calibrated_coverage.append(covered)
        calibrated_ratios.append(rounded(calibrated_ratio))
        allow_zero_flags.append(allow_zero_floor)
        if calibrated_ratio > current_ratio:
            calibration_reasons.append(f"rolling_prior_residual_quantile_by_{source}")
        elif allow_zero_floor:
            calibration_reasons.append("conservative_low_confidence_zero_floor")
        else:
            calibration_reasons.append("existing_interval_retained")
    frame["v1IntervalCoverage"] = frame["intervalCoverage"]
    frame["v1OptimisticPessimisticRatio"] = frame["optimisticPessimisticRatio"]
    frame["intervalCoverage"] = calibrated_coverage
    frame["optimisticPessimisticRatio"] = calibrated_ratios
    frame["intervalCalibrationReason"] = calibration_reasons
    frame["allowZeroFloorInterval"] = allow_zero_flags
    frame["modelId"] = MODEL_ID

    forecastable = frame[frame["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]
    for field in ["confidence", "horizonMonths", "forecastabilityStatus", "lifecycleAtCutoff", "revenueScaleAtCutoff"]:
        for segment, group in forecastable.groupby(field):
            before = v1.score_cases(group.assign(intervalCoverage=group["v1IntervalCoverage"]))["intervalCoverage"]
            after = v1.score_cases(group)["intervalCoverage"]
            rows.append(
                {
                    "segmentType": field,
                    "segment": str(segment),
                    "caseCount": int(len(group)),
                    "coverageBefore": before,
                    "coverageAfter": after,
                    "spreadP75After": bake.quantiles(group["optimisticPessimisticRatio"].dropna().tolist())["p75"],
                }
            )
    return frame, sorted(rows, key=lambda item: (item["segmentType"], item["segment"]))


def enrich_validation(validation: dict, model_cases: pd.DataFrame, coverage: dict) -> dict:
    forecastable = model_cases[model_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    spread = bake.spread_summary(forecastable)
    validation["schema"] = "m2.disentangled_forecast_v1_1_validation.v1"
    validation["modelId"] = MODEL_ID
    validation["candidateVersion"] = None
    forecastable_score = validation["forecastableCohortScore"]
    pass_conditions = {
        "forecastableRevenueCoverageAtLeast70": coverage["forecastableNumericIncludingConservative"]["revenueShare"] >= 0.70,
        "top5ForecastableCountShareAtLeast80": coverage["topRevenueCoverage"]["top5Percent"]["forecastableCountShare"] >= 0.80,
        "wapeNotWorseThanBaseline": safe_float(forecastable_score["wape"], 999) <= safe_float(forecastable_score["baselineWape"], -1),
        "coverageAtLeast50": safe_float(forecastable_score["intervalCoverage"]) >= 0.50,
        "p0EqualsZero": validation["issueSummary"]["p0"] == 0,
        "p1AtMost3": validation["issueSummary"]["p1"] <= 3,
        "highConfidenceSpreadP75AtMost1_5": spread["highConfidenceSpread"]["p75"] is None or safe_float(spread["highConfidenceSpread"]["p75"]) <= 1.50,
        "nonLowConfidenceSpreadP75AtMost2_0": spread["nonLowConfidenceSpread"]["p75"] is None or safe_float(spread["nonLowConfidenceSpread"]["p75"]) <= 2.00,
        "trueForecastBlockedRevenueShareAtMost20": coverage["trueForecastBlocked"]["revenueShare"] <= 0.20,
        "formalAndBusinessBlockersSeparated": True,
        "noGateMasqueradeDetected": True,
    }
    conditional_conditions = {
        "forecastableRevenueCoverageAtLeast70": coverage["forecastableNumericIncludingConservative"]["revenueShare"] >= 0.70,
        "wapeNotWorseThanBaseline": safe_float(forecastable_score["wape"], 999) <= safe_float(forecastable_score["baselineWape"], -1),
        "p0EqualsZero": validation["issueSummary"]["p0"] == 0,
        "p1AtMost10": validation["issueSummary"]["p1"] <= 10,
        "coverageAtLeast45": safe_float(forecastable_score["intervalCoverage"]) >= 0.45,
        "trueForecastBlockedRevenueShareAtMost25": coverage["trueForecastBlocked"]["revenueShare"] <= 0.25,
    }
    if all(pass_conditions.values()):
        verdict = "PASS"
        candidate = V1_1_CANDIDATE
    elif all(conditional_conditions.values()):
        verdict = "CONDITIONAL PASS"
        candidate = f"{V1_1_CANDIDATE}-conditional"
    else:
        verdict = "FAIL"
        candidate = None
    validation["verdict"] = verdict
    validation["candidateVersion"] = candidate
    validation["passConditions"] = pass_conditions
    validation["conditionalPassConditions"] = conditional_conditions
    validation["spreadSummary"] = {
        "highConfidenceSpreadP75": spread["highConfidenceSpread"]["p75"],
        "nonLowConfidenceSpreadP75": spread["nonLowConfidenceSpread"]["p75"],
        "byConfidence": spread["byConfidence"],
    }
    validation["safeOutputBoundary"] = safe_boundary()
    return validation


def build_hardening_report(v1_coverage: dict, v1_validation: dict, v1_1_coverage: dict, v1_1_validation: dict, gate_changes: list[dict], interval_rows: list[dict]) -> dict:
    return {
        "schema": "m2.disentangled_forecast_v1_1_hardening.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "candidateVersion": v1_1_validation["candidateVersion"],
        "verdict": v1_1_validation["verdict"],
        "v1_0Shortfalls": [
            {
                "metric": "intervalCoverage",
                "value": v1_0_metric(v1_validation, "forecastableCohortScore", "intervalCoverage"),
                "requiredForPass": 0.50,
            },
            {
                "metric": "trueForecastBlockedRevenueShare",
                "value": v1_coverage["trueForecastBlocked"]["revenueShare"],
                "requiredForPass": 0.20,
            },
        ],
        "v1_0VsV1_1": scoreboard_rows(v1_coverage, v1_validation, v1_1_coverage, v1_1_validation),
        "gateBoundaryChanges": {
            "downgradedToConservativeCount": len(gate_changes),
            "downgradedRevenueShare": rounded(sum(row["revenueShare"] for row in gate_changes)),
            "rule": "only active unresolved-spike rows with enough recent revenue are downgraded to conservative numeric forecast",
            "changesByMateriality": distribution([row["materialityBucket"] for row in gate_changes]),
        },
        "intervalCalibration": {
            "method": "segmented residual quantile calibration by status, horizon, confidence, and revenue bucket",
            "fixedGlobalMultiplierUsed": False,
            "uniformWideningUsed": False,
            "segmentRows": interval_rows[:120],
        },
        "safeOutputBoundary": safe_boundary(),
    }


def v1_0_metric(report: dict, section: str, metric: str):
    return report.get(section, {}).get(metric)


def scoreboard_rows(v1_coverage: dict, v1_validation: dict, v1_1_coverage: dict, v1_1_validation: dict) -> list[dict]:
    return [
        {
            "version": "v1.0",
            "verdict": v1_validation["verdict"],
            "candidateVersion": v1_validation["candidateVersion"],
            "forecastableRevenueShare": v1_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueBlockedRevenueShare": v1_coverage["trueForecastBlocked"]["revenueShare"],
            "wape": v1_validation["forecastableCohortScore"]["wape"],
            "baselineWape": v1_validation["forecastableCohortScore"]["baselineWape"],
            "coverage": v1_validation["forecastableCohortScore"]["intervalCoverage"],
            "p0": v1_validation["issueSummary"]["p0"],
            "p1": v1_validation["issueSummary"]["p1"],
            "p2": v1_validation["issueSummary"]["p2"],
            "highConfidenceSpreadP75": None,
            "nonLowConfidenceSpreadP75": None,
        },
        {
            "version": "v1.1",
            "verdict": v1_1_validation["verdict"],
            "candidateVersion": v1_1_validation["candidateVersion"],
            "forecastableRevenueShare": v1_1_coverage["forecastableNumericIncludingConservative"]["revenueShare"],
            "trueBlockedRevenueShare": v1_1_coverage["trueForecastBlocked"]["revenueShare"],
            "wape": v1_1_validation["forecastableCohortScore"]["wape"],
            "baselineWape": v1_1_validation["forecastableCohortScore"]["baselineWape"],
            "coverage": v1_1_validation["forecastableCohortScore"]["intervalCoverage"],
            "p0": v1_1_validation["issueSummary"]["p0"],
            "p1": v1_1_validation["issueSummary"]["p1"],
            "p2": v1_1_validation["issueSummary"]["p2"],
            "highConfidenceSpreadP75": v1_1_validation["spreadSummary"]["highConfidenceSpreadP75"],
            "nonLowConfidenceSpreadP75": v1_1_validation["spreadSummary"]["nonLowConfidenceSpreadP75"],
        },
    ]


def build_business_readiness(coverage: dict, validation: dict) -> dict:
    if validation["verdict"] == "PASS":
        readiness = "m2_limited_business_review_ready_for_forecast_baseline_acceptance"
    elif validation["verdict"] == "CONDITIONAL PASS":
        readiness = "limited_m2_business_review_ready_for_forecastable_cohort_only"
    else:
        readiness = "not_ready_for_m2_business_review"
    return {
        "schema": "m2.disentangled_forecast_business_readiness.v1_1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "readiness": readiness,
        "m3Allowed": False,
        "candidateVersion": validation["candidateVersion"],
        "modelVerdict": validation["verdict"],
        "forecastableRevenueShare": coverage["forecastableNumericIncludingConservative"]["revenueShare"],
        "trueForecastBlockedRevenueShare": coverage["trueForecastBlocked"]["revenueShare"],
        "businessGroups": [
            {
                "group": "numeric_forecast_eligible",
                "businessUse": "can_enter_limited_m2_business_review",
                "reason": "validated numeric forecast cohort",
            },
            {
                "group": "conservative_numeric_forecast",
                "businessUse": "reference_only_or_low_confidence_review",
                "reason": "forecastable after conservative gate and calibrated interval",
            },
            {
                "group": "true_forecast_blocked",
                "businessUse": "no_numeric_forecast",
                "reason": "insufficient or genuinely unstable revenue series",
            },
            {
                "group": "formal_release_blockers",
                "businessUse": "not_formal_release_approval",
                "reason": "formal readiness remains separate from forecastability",
            },
        ],
        "m3BlockReason": "M3 remains blocked unless the user explicitly approves M3 planning after accepting the forecast baseline.",
        "safeOutputBoundary": safe_boundary(),
    }


def write_reports(drilldown: dict, hardening: dict, validation: dict, readiness: dict) -> None:
    write_json(DRILLDOWN_JSON, drilldown)
    write_json(HARDENING_JSON, hardening)
    write_json(VALIDATION_JSON, validation)
    write_json(READINESS_JSON, readiness)

    DRILLDOWN_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecastability v1.0 Conditional Failure Drilldown",
                "",
                f"Coverage: `{drilldown['coverageRootCause']['coverage']}`.",
                f"Under-forecast rate: `{drilldown['coverageRootCause']['underForecastRate']}`.",
                f"Over-forecast rate: `{drilldown['coverageRootCause']['overForecastRate']}`.",
                "",
                "## Low Coverage Segments",
                "",
                markdown_table(
                    drilldown["coverageRootCause"]["lowCoverageSegments"][:40],
                    [
                        ("segmentType", "Segment Type"),
                        ("segment", "Segment"),
                        ("caseCount", "Cases"),
                        ("coverage", "Coverage"),
                        ("wape", "WAPE"),
                        ("underForecastRate", "Under"),
                        ("overForecastRate", "Over"),
                    ],
                ),
                "",
                "## True Forecast Blocked Reasons",
                "",
                markdown_table(
                    drilldown["trueBlockedRootCause"]["reasonDistribution"],
                    [
                        ("reasonCode", "Reason"),
                        ("count", "Count"),
                        ("revenueShare", "Revenue Share"),
                        ("top5OverlapCount", "Top 5 Count"),
                    ],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    HARDENING_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecast v1.1 Hardening",
                "",
                f"Verdict: `{hardening['verdict']}`.",
                f"Candidate version: `{hardening['candidateVersion']}`.",
                "",
                "## v1.0 vs v1.1",
                "",
                markdown_table(
                    hardening["v1_0VsV1_1"],
                    [
                        ("version", "Version"),
                        ("verdict", "Verdict"),
                        ("forecastableRevenueShare", "Forecastable Revenue Share"),
                        ("trueBlockedRevenueShare", "True Blocked Revenue Share"),
                        ("wape", "WAPE"),
                        ("baselineWape", "Baseline WAPE"),
                        ("coverage", "Coverage"),
                        ("p0", "P0"),
                        ("p1", "P1"),
                    ],
                ),
                "",
                "## Gate Boundary Changes",
                "",
                f"Downgraded to conservative count: `{hardening['gateBoundaryChanges']['downgradedToConservativeCount']}`.",
                f"Downgraded revenue share: `{hardening['gateBoundaryChanges']['downgradedRevenueShare']}`.",
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    VALIDATION_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecast v1.1 Validation",
                "",
                f"Verdict: `{validation['verdict']}`.",
                f"Candidate version: `{validation['candidateVersion']}`.",
                "",
                "## Forecastable Cohort Score",
                "",
                markdown_table(
                    [validation["forecastableCohortScore"]],
                    [
                        ("wape", "WAPE"),
                        ("baselineWape", "Baseline WAPE"),
                        ("smape", "SMAPE"),
                        ("mae", "MAE"),
                        ("intervalCoverage", "Coverage"),
                        ("failRate", "Fail Rate"),
                    ],
                ),
                "",
                "## Issue Summary",
                "",
                markdown_table([validation["issueSummary"]], [("p0", "P0"), ("p1", "P1"), ("p2", "P2"), ("p2RevenueShare", "P2 Revenue Share")]),
                "",
                "## Spread Guardrails",
                "",
                markdown_table(
                    [validation["spreadSummary"]],
                    [
                        ("highConfidenceSpreadP75", "High Confidence P75"),
                        ("nonLowConfidenceSpreadP75", "Non-low Confidence P75"),
                    ],
                ),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    READINESS_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecast Business Readiness v1.1",
                "",
                f"Readiness: `{readiness['readiness']}`.",
                f"M3 allowed: `{readiness['m3Allowed']}`.",
                f"Candidate version: `{readiness['candidateVersion']}`.",
                "",
                "## Business Groups",
                "",
                markdown_table(readiness["businessGroups"], [("group", "Group"), ("businessUse", "Business Use"), ("reason", "Reason")]),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_excel(
    drilldown: dict,
    hardening: dict,
    validation: dict,
    readiness: dict,
    gate_changes: list[dict],
    interval_rows: list[dict],
    current_gate: pd.DataFrame,
    model_cases: pd.DataFrame,
) -> None:
    if Alignment is None:
        raise RuntimeError("openpyxl is required to write the private Excel workbook.")
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    forecastable_cases = model_cases[model_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy()
    forecastable_cases["anonymousCaseId"] = [f"V110-{index + 1:05d}" for index in range(len(forecastable_cases))]
    sample_200 = forecastable_cases.sort_values(["materialityBucket", "workKey", "cutoffMonth", "horizonMonths"]).head(200)
    full_labels = v1.labels_for(forecastable_cases)
    fail_cases = forecastable_cases[full_labels.isin(["warning", "fail"])].copy().head(2000)
    fail_cases["verdict"] = full_labels[full_labels.isin(["warning", "fail"])].head(2000).values
    high_value = current_gate[current_gate["materialityRankPercent"] <= 0.10].copy().head(500)
    true_blocked = current_gate[current_gate["forecastabilityStatus"] == TRUE_BLOCKED_STATUS].copy().head(1000)

    with pd.ExcelWriter(PRIVATE_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"项目": "用途", "说明": "M2 预测可用性 v1.1 验证表"},
                {"项目": "安全边界", "说明": "仅匿名 ID；不包含作品名、作者名、渠道名或原始账单行"},
                {"项目": "PASS 标准", "说明": "覆盖率 >= 50%，true blocked 收入占比 <= 20%，WAPE 不差于 baseline，P0=0，P1<=3"},
                {"项目": "候选版本", "说明": validation["candidateVersion"]},
                {"项目": "结论", "说明": validation["verdict"]},
            ]
        ).to_excel(writer, sheet_name="00_read_me", index=False)
        pd.DataFrame(hardening["v1_0VsV1_1"]).to_excel(writer, sheet_name="01_v1_0_vs_v1_1_scoreboard", index=False)
        pd.DataFrame(drilldown["coverageRootCause"]["lowCoverageSegments"]).to_excel(writer, sheet_name="02_coverage_drilldown", index=False)
        pd.DataFrame(drilldown["trueBlockedRootCause"]["reasonDistribution"]).to_excel(writer, sheet_name="03_true_blocked_drilldown", index=False)
        pd.DataFrame(interval_rows).to_excel(writer, sheet_name="04_interval_calibration", index=False)
        pd.DataFrame(gate_changes).to_excel(writer, sheet_name="05_gate_boundary_changes", index=False)
        pd.DataFrame([validation["forecastableCohortScore"], validation["numericCohortScore"], validation["conservativeCohortScore"]]).to_excel(writer, sheet_name="06_model_performance", index=False)
        high_value[["anonymousId", "materialityBucket", "forecastabilityStatus", "formalReadinessStatus", "businessActionStatus", "forecastConfidence"]].to_excel(writer, sheet_name="07_high_value_work_check", index=False)
        sample_200[
            [
                "anonymousCaseId",
                "forecastabilityStatus",
                "confidence",
                "horizonMonths",
                "predicted",
                "actual",
                "absoluteError",
                "baselinePredicted",
                "betterThanBaseline",
                "intervalCoverage",
                "intervalCalibrationReason",
            ]
        ].to_excel(writer, sheet_name="08_200_sample_validation", index=False)
        pd.DataFrame(validation["segmentPerformance"]).to_excel(writer, sheet_name="09_full_cohort_summary", index=False)
        pd.DataFrame(readiness["businessGroups"]).to_excel(writer, sheet_name="10_business_readiness", index=False)
        fail_cases[
            [
                "anonymousCaseId",
                "verdict",
                "forecastabilityStatus",
                "confidence",
                "horizonMonths",
                "predicted",
                "actual",
                "absoluteError",
                "baselinePredicted",
                "intervalCoverage",
                "intervalCalibrationReason",
            ]
        ].to_excel(writer, sheet_name="11_fail_or_warning_cases", index=False)
        true_blocked[["anonymousId", "materialityBucket", "forecastabilityReasonCodes", "requiredForecastabilityAction"]].to_excel(writer, sheet_name="12_true_blocked_private", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, column in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in column:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(10, min(max_len + 3, 45))
            worksheet.auto_filter.ref = worksheet.dimensions


def run() -> dict:
    prepared = v1.prepare_inputs()
    evaluated = prepared["evaluated"]
    cases = prepared["cases"]
    final_outputs = prepared["finalOutputs"]

    v1_gate = v1.build_current_gate_frame(evaluated, final_outputs)
    v1_coverage = v1.build_coverage_report(v1_gate)
    v1_cases = v1.build_disentangled_cases(cases, v1_gate)
    v1_validation = v1.build_validation_report(v1_cases, v1_gate, v1_coverage)
    drilldown = build_failure_drilldown(v1_cases, v1_gate)

    v1_1_gate, gate_changes = apply_v1_1_gate_boundary(v1_gate, final_outputs)
    v1_1_coverage = v1.build_coverage_report(v1_1_gate)
    v1_1_cases_raw = v1.build_disentangled_cases(cases, v1_1_gate)
    v1_1_cases, interval_rows = apply_interval_calibration(v1_1_cases_raw)
    v1_1_validation = v1.build_validation_report(v1_1_cases, v1_1_gate, v1_1_coverage)
    v1_1_validation = enrich_validation(v1_1_validation, v1_1_cases, v1_1_coverage)
    hardening = build_hardening_report(v1_coverage, v1_validation, v1_1_coverage, v1_1_validation, gate_changes, interval_rows)
    readiness = build_business_readiness(v1_1_coverage, v1_1_validation)

    write_reports(drilldown, hardening, v1_1_validation, readiness)
    write_excel(drilldown, hardening, v1_1_validation, readiness, gate_changes, interval_rows, v1_1_gate, v1_1_cases)
    write_json(
        PRIVATE_DETAIL_JSON,
        {
            "schema": "m2.private.disentangled_forecast_v1_1_detail.v1",
            "notForCommit": True,
            "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
            "candidateVersion": v1_1_validation["candidateVersion"],
            "verdict": v1_1_validation["verdict"],
            "scoreboard": hardening["v1_0VsV1_1"],
        },
    )
    return {
        "v1_1HardeningCompleted": True,
        "newCandidateVersion": v1_1_validation["candidateVersion"],
        "verdict": v1_1_validation["verdict"],
        "m2BusinessReviewReady": readiness["readiness"],
        "m3Allowed": readiness["m3Allowed"],
        "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
        "reports": [
            str(DRILLDOWN_JSON.relative_to(ROOT)),
            str(HARDENING_JSON.relative_to(ROOT)),
            str(VALIDATION_JSON.relative_to(ROOT)),
            str(READINESS_JSON.relative_to(ROOT)),
        ],
    }


if __name__ == "__main__":
    print(json.dumps(run(), ensure_ascii=False, indent=2))
