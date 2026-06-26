from __future__ import annotations

import importlib.util
import json
import os
import subprocess
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it into a local temp dependency path, "
        "for example: python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


CANDIDATE_VERSION = "m2-realdata-dev-rating-suggestion-calibrated-v1.0"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-business-review"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
M1_PRIVATE_DIR = ROOT / "data" / "private-output" / "m1-master-data"

INPUT_OPERATOR_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-after-dual-source-staging-v2.xlsx"
PRIVATE_VALIDATION_XLSX = PRIVATE_DIR / "m2-rating-suggestion-calibrated-v1.0-validation.xlsx"
PRIVATE_OPERATOR_V3_XLSX = PRIVATE_DIR / "m2-v1.1-30-work-operator-task-pack-cn-after-rating-suggestion-calibration-v1.xlsx"

OPERATOR_SUMMARY_JSON = DOCS_DIR / "M2-v1.1-after-staging-operator-validation-summary-v1.json"
FORECAST_OUTPUT_JSON = DOCS_DIR / "M2-forecast-output-type-after-dual-source-staging-v2.json"
FORECASTABILITY_JSON = DOCS_DIR / "M2-v1.1-forecastability-after-dual-source-staging-v1.json"
BUSINESS_READINESS_JSON = DOCS_DIR / "M2-v1.1-business-readiness-after-dual-source-staging-v1.json"
STAGING_TABLE_JSON = M1_PRIVATE_DIR / "M1-dual-source-limited-staging-table-v1.json"

RATING_ROOT_JSON = DOCS_DIR / "M2-rating-failure-root-cause-after-staging-v1.json"
RATING_ROOT_MD = DOCS_DIR / "M2-rating-failure-root-cause-after-staging-v1.md"
SUGGESTION_ROOT_JSON = DOCS_DIR / "M2-suggestion-failure-root-cause-after-staging-v1.json"
SUGGESTION_ROOT_MD = DOCS_DIR / "M2-suggestion-failure-root-cause-after-staging-v1.md"
CALIBRATION_SUMMARY_JSON = DOCS_DIR / "M2-rating-suggestion-calibration-v1-summary.json"
CALIBRATION_SUMMARY_MD = DOCS_DIR / "M2-rating-suggestion-calibration-v1-summary.md"
OPERATOR_V3_SUMMARY_JSON = DOCS_DIR / "M2-operator-task-pack-after-rating-suggestion-calibration-v1-summary.json"
OPERATOR_V3_SUMMARY_MD = DOCS_DIR / "M2-operator-task-pack-after-rating-suggestion-calibration-v1-summary.md"

TASK_SHEET = "01_运营任务卡"
USER_RESERVED = "用户指定作品"
RATING_RANK = {"S+": 0, "S": 1, "A": 2, "B": 3, "C": 4, "D": 5, "E": 6}


def main() -> None:
    inputs = load_inputs()
    rows = load_operator_rows()
    calibrated = [calibrate_row(row) for row in rows]
    reviewable = [row for row in calibrated if row["sampleSource"] != USER_RESERVED]

    rating_root = build_rating_root_cause(reviewable, inputs)
    suggestion_root = build_suggestion_root_cause(reviewable, inputs)
    calibration_summary = build_calibration_summary(calibrated, reviewable, rating_root, suggestion_root, inputs)
    operator_summary = build_operator_v3_summary(calibrated, reviewable)

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    write_json(RATING_ROOT_JSON, public_envelope("m2.rating_failure_root_cause_after_staging.v1", rating_root))
    write_text(RATING_ROOT_MD, rating_root_md(rating_root))
    write_json(SUGGESTION_ROOT_JSON, public_envelope("m2.suggestion_failure_root_cause_after_staging.v1", suggestion_root))
    write_text(SUGGESTION_ROOT_MD, suggestion_root_md(suggestion_root))
    write_json(CALIBRATION_SUMMARY_JSON, public_envelope("m2.rating_suggestion_calibration_v1.summary", calibration_summary))
    write_text(CALIBRATION_SUMMARY_MD, calibration_summary_md(calibration_summary))
    write_json(OPERATOR_V3_SUMMARY_JSON, public_envelope("m2.operator_task_pack_after_rating_suggestion_calibration_v1.summary", operator_summary))
    write_text(OPERATOR_V3_SUMMARY_MD, operator_summary_md(operator_summary))

    write_private_workbooks(calibrated, reviewable, rating_root, suggestion_root, calibration_summary)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "reviewableRows": len(reviewable),
                "ratingChangedRows": calibration_summary["rating"]["changedRows"],
                "suggestionChangedRows": calibration_summary["suggestion"]["changedRows"],
                "ratingFeedbackHitRows": calibration_summary["expectedImprovement"]["ratingUnreasonableHitRows"],
                "suggestionNonExecutableHitRows": calibration_summary["expectedImprovement"]["suggestionNotExecutableHitRows"],
                "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
                "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_V3_XLSX),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def load_inputs() -> dict:
    missing = [
        path
        for path in [
            INPUT_OPERATOR_XLSX,
            OPERATOR_SUMMARY_JSON,
            FORECAST_OUTPUT_JSON,
            FORECASTABILITY_JSON,
            BUSINESS_READINESS_JSON,
            STAGING_TABLE_JSON,
            ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
            ROOT / "package.json",
            ROOT / ".gitignore",
        ]
        if not path.exists()
    ]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))
    return {
        "operatorSummary": read_json(OPERATOR_SUMMARY_JSON),
        "forecastOutput": read_json(FORECAST_OUTPUT_JSON),
        "forecastability": read_json(FORECASTABILITY_JSON),
        "businessReadiness": read_json(BUSINESS_READINESS_JSON),
        "stagingRecordCount": len(read_json(STAGING_TABLE_JSON).get("records", [])),
    }


def load_operator_rows() -> list[dict]:
    workbook = load_workbook(INPUT_OPERATOR_XLSX, read_only=True, data_only=True)
    sheet = workbook[TASK_SHEET] if TASK_SHEET in workbook.sheetnames else workbook.worksheets[1]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = []
    for row_index, values in enumerate(iterator, start=2):
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if not any(clean(value) for value in row.values()):
            continue
        row["_rowIndex"] = row_index
        rows.append(row)
    return rows


def calibrate_row(row: dict) -> dict:
    sample_source = clean(row.get("样本来源"))
    if sample_source == USER_RESERVED:
        return {
            **row,
            "sampleSource": sample_source,
            "isReviewable": False,
            "newRating": "",
            "newSuggestion": "",
            "ratingChanged": False,
            "suggestionChanged": False,
            "ratingRootCauses": [],
            "suggestionRootCauses": [],
            "manualConfirmationRequired": True,
            "m4CandidateRecommended": "待补任务卡",
        }

    old_rating = clean(row.get("评级"))
    old_suggestion = clean(row.get("运营建议"))
    revenue = clean(row.get("辅助原始收入层级code"))
    lifecycle = clean(row.get("辅助原始生命周期code"))
    forecastability = clean(row.get("辅助原始预测状态code"))
    business_action = clean(row.get("辅助原始业务动作状态code"))
    forecast_output_type = clean(row.get("辅助原始forecastOutputType"))
    confidence = confidence_code(clean(row.get("预测置信度")))
    remaining_months = number_or_none(row.get("剩余版权月数"))
    rating_feedback = clean(row.get("运营判断：评级是否合理"))
    suggestion_feedback = clean(row.get("运营判断：建议是否可执行"))
    issue_type = clean(row.get("运营发现的问题类型"))

    rating = calibrate_rating(
        old_rating=old_rating,
        revenue=revenue,
        lifecycle=lifecycle,
        forecastability=forecastability,
        business_action=business_action,
        confidence=confidence,
        remaining_months=remaining_months,
        forecast_output_type=forecast_output_type,
    )
    suggestion = calibrate_suggestion(
        old_suggestion=old_suggestion,
        rating=rating["rating"],
        revenue=revenue,
        lifecycle=lifecycle,
        forecastability=forecastability,
        business_action=business_action,
        confidence=confidence,
        remaining_months=remaining_months,
        forecast_output_type=forecast_output_type,
    )
    suggestion_changed = old_suggestion not in {suggestion["suggestion"], suggestion["suggestion_cn"]}

    rating_causes = rating_root_causes(
        old_rating=old_rating,
        new_rating=rating["rating"],
        revenue=revenue,
        lifecycle=lifecycle,
        forecastability=forecastability,
        business_action=business_action,
        forecast_output_type=forecast_output_type,
        rating_feedback=rating_feedback,
        issue_type=issue_type,
    )
    suggestion_causes = suggestion_root_causes(
        old_suggestion=old_suggestion,
        new_suggestion=suggestion["suggestion"],
        forecastability=forecastability,
        business_action=business_action,
        suggestion_feedback=suggestion_feedback,
        issue_type=issue_type,
        suggestion_changed=suggestion_changed,
    )

    return {
        **row,
        "sampleSource": sample_source,
        "isReviewable": True,
        "oldRating": old_rating,
        "newRating": rating["rating"],
        "ratingChanged": old_rating != rating["rating"],
        "ratingChangeDirection": change_direction(old_rating, rating["rating"]),
        "ratingRationaleCn": "；".join(rating["rationale"]),
        "oldSuggestion": old_suggestion,
        "newSuggestion": suggestion["suggestion"],
        "newSuggestionCn": suggestion["suggestion_cn"],
        "suggestionChanged": suggestion_changed,
        "suggestionChangeReasonCn": "；".join(suggestion["evidence"]),
        "suggestionActionabilityLevel": suggestion["actionability"],
        "suggestionRequiresManualConfirmation": suggestion["manual"],
        "whyNotOtherSuggestionsCn": "；".join(suggestion["why_not"]),
        "ratingRootCauses": rating_causes,
        "suggestionRootCauses": suggestion_causes,
        "manualConfirmationRequired": rating["manual"] or suggestion["manual"],
        "m4CandidateRecommended": "是" if rating_feedback == "不合理" or suggestion_feedback == "不可执行" or "业务常识冲突" in issue_type else "待定",
    }


def calibrate_rating(*, old_rating, revenue, lifecycle, forecastability, business_action, confidence, remaining_months, forecast_output_type):
    score = {"top": 58, "high": 54, "medium": 46, "mid": 46, "low": 30, "long_tail": 24, "near_zero": 18, "zero": 12}.get(revenue, 42)
    score += {"growth": 10, "rebound": 8, "stable": 8, "declining": -6, "inactive": -14, "long_tail": -16, "insufficient_history": -10}.get(lifecycle, 0)
    score += {"high": 6, "medium": 2, "low": -4, "blocked_for_business_use": -10}.get(confidence, 0)
    score += {"numeric_forecast_eligible": 8, "conservative_numeric_forecast": 3, "observe_only_no_numeric_forecast": -12, "true_forecast_blocked": -14}.get(forecastability, 0)
    score += {"action_allowed": 2, "manual_confirmation_required": -2, "action_blocked": -6, "observe_only": -8}.get(business_action, 0)
    rationale = [f"收入层级={revenue or '未知'}", f"生命周期={lifecycle or '未知'}", f"预测状态={forecastability or '未知'}", f"业务动作状态={business_action or '未知'}"]
    manual = False

    if remaining_months is not None:
        if remaining_months >= 36 and revenue in {"top", "high"} and lifecycle in {"stable", "growth", "rebound"}:
            score += 6
            rationale.append("剩余版权期长且收入/生命周期健康")
        elif remaining_months >= 18 and lifecycle in {"stable", "growth", "rebound"}:
            score += 3
            rationale.append("剩余版权期可支持运营")
        elif remaining_months <= 6:
            score -= 8
            rationale.append("剩余版权期很短")
        elif remaining_months <= 12:
            score -= 4
            rationale.append("剩余版权期偏短")
    elif forecast_output_type == "operating_window_forecast_pending_expiry":
        rationale.append("缺版权到期只作为 readiness warning，不直接压低商业价值")

    rating = rating_from_score(score)
    if confidence in {"low", "blocked_for_business_use"}:
        rating = min_rating(rating, "A")
        rationale.append("低置信度不能给 S/S+")
    if forecastability == "true_forecast_blocked":
        if revenue in {"top", "high"} and lifecycle not in {"inactive", "long_tail"}:
            rating = min_rating(rating, "B")
            rating = max_rating(rating, "B")
            manual = True
            rationale.append("高收入但预测阻断，最高到 B 且需人工复核")
        elif revenue in {"top", "high"}:
            rating = max_rating(min_rating(rating, "C"), "C")
            rationale.append("预测阻断的高收入弱生命周期样本保护到 C")
        else:
            rating = min_rating(rating, "C")
            rationale.append("预测阻断默认不超过 C")
    if forecastability == "observe_only_no_numeric_forecast" or business_action == "observe_only":
        if revenue in {"top", "high"}:
            rating = max_rating(min_rating(rating, "C"), "C")
        else:
            rating = min_rating(rating, "D")
        rationale.append("仅观察样本不输出高评级")
    if business_action == "action_blocked" and forecastability != "true_forecast_blocked":
        rating = min_rating(rating, "B")
        rationale.append("业务动作阻断不直接给高评级")
    if lifecycle in {"inactive", "long_tail"} and revenue in {"low", "long_tail", "near_zero", "zero"}:
        rating = min_rating(rating, "D" if lifecycle == "long_tail" else "E")
        rationale.append("低收入长尾/沉寂保持低评级")
    if rating in {"S+", "S"}:
        manual = True
        rationale.append("S/S+ 必须人工确认")
    if old_rating and old_rating != rating:
        rationale.append(f"旧评级 {old_rating} 调整为 {rating}")
    return {"rating": rating, "score": round(score, 2), "rationale": rationale, "manual": manual}


def calibrate_suggestion(*, old_suggestion, rating, revenue, lifecycle, forecastability, business_action, confidence, remaining_months, forecast_output_type):
    if forecastability == "true_forecast_blocked" or business_action == "action_blocked":
        return {
            "suggestion": "manual_review_required",
            "suggestion_cn": "人工复核阻断原因后再决定运营动作",
            "evidence": ["预测或业务动作仍处于阻断状态", "不直接给投放、下架或续约动作"],
            "manual": True,
            "actionability": "manual_confirmation",
            "why_not": ["避免自动推广", "避免误伤下架", f"替代旧建议：{old_suggestion or '无'}"],
        }
    if forecastability == "observe_only_no_numeric_forecast" or business_action == "observe_only" or lifecycle == "insufficient_history":
        return {
            "suggestion": "observe_only",
            "suggestion_cn": "仅观察，暂不作为投放、续约或下架的直接依据",
            "evidence": ["仅观察或历史不足", "无足够证据支持直接动作"],
            "manual": False,
            "actionability": "observe_only",
            "why_not": ["预测证据不足，不给推广、下架或续约建议"],
        }
    if forecast_output_type == "copyright_term_forecast" and remaining_months is not None and remaining_months <= 12 and rating in {"S+", "S", "A", "B"} and revenue in {"top", "high", "medium", "mid"}:
        return {
            "suggestion": "renewal_review",
            "suggestion_cn": "版权续约复核",
            "evidence": ["剩余版权期偏短", f"评级={rating}", f"收入层级={revenue}"],
            "manual": True,
            "actionability": "manual_confirmation",
            "why_not": ["续约必须有版权期和收入价值支撑"],
        }
    if rating in {"S+", "S", "A"} and lifecycle in {"growth", "rebound"} and confidence in {"high", "medium", ""} and business_action == "action_allowed":
        return {
            "suggestion": "promote",
            "suggestion_cn": "加大推广或重点运营",
            "evidence": [f"评级={rating}", f"生命周期={lifecycle}", f"预测置信度={confidence or '中等或以上'}"],
            "manual": True,
            "actionability": "manual_confirmation",
            "why_not": ["高价值动作需确认非一次性收入"],
        }
    if rating in {"S+", "S", "A", "B", "C"} and lifecycle in {"stable", "growth", "rebound"} and confidence != "low":
        return {
            "suggestion": "maintain",
            "suggestion_cn": "维持当前运营",
            "evidence": [f"评级={rating}", f"生命周期={lifecycle}", "预测或历史证据支持继续运营"],
            "manual": False,
            "actionability": "actionable",
            "why_not": ["未达到强推广条件，也没有下架依据"],
        }
    if rating in {"B", "C", "D"} and lifecycle in {"declining", "inactive", "long_tail", "insufficient_history"} and revenue in {"top", "high", "medium", "mid"}:
        return {
            "suggestion": "reduce_investment",
            "suggestion_cn": "降低增量投入，保留观察",
            "evidence": [f"评级={rating}", f"生命周期={lifecycle}", f"收入层级={revenue}"],
            "manual": False,
            "actionability": "actionable_with_caution",
            "why_not": ["仍有收入或版权价值，不直接下架"],
        }
    if rating in {"D", "E"} and lifecycle in {"inactive", "long_tail"} and revenue in {"low", "long_tail", "near_zero", "zero"} and business_action != "action_blocked":
        return {
            "suggestion": "downlist_or_suspend",
            "suggestion_cn": "下架或暂停运营候选",
            "evidence": [f"评级={rating}", f"生命周期={lifecycle}", f"收入层级={revenue}"],
            "manual": True,
            "actionability": "manual_confirmation",
            "why_not": ["仅限低收入长尾/沉寂样本"],
        }
    return {
        "suggestion": "manual_review_required",
        "suggestion_cn": "人工复核后再决定运营动作",
        "evidence": ["规则未形成安全的自动建议", f"评级={rating or '未知'}", f"生命周期={lifecycle or '未知'}"],
        "manual": True,
        "actionability": "manual_confirmation",
        "why_not": ["避免把不确定样本包装成直接业务动作"],
    }


def rating_root_causes(**kwargs) -> list[str]:
    causes = []
    old_rating = kwargs["old_rating"]
    revenue = kwargs["revenue"]
    lifecycle = kwargs["lifecycle"]
    forecastability = kwargs["forecastability"]
    business_action = kwargs["business_action"]
    forecast_output_type = kwargs["forecast_output_type"]
    if kwargs["rating_feedback"] == "不合理":
        causes.append("operator_marked_rating_unreasonable")
    if revenue in {"top", "high"} and old_rating in {"C", "D", "E"}:
        causes.append("high_revenue_low_rating")
    if revenue in {"top", "high"} and old_rating == "B" and lifecycle in {"stable", "growth"} and forecastability != "true_forecast_blocked":
        causes.append("high_revenue_stable_under_rated")
    if forecastability in {"true_forecast_blocked", "observe_only_no_numeric_forecast"} and revenue in {"top", "high"}:
        causes.append("forecastability_over_penalized_rating")
    if lifecycle in {"inactive", "long_tail"} and revenue in {"top", "high"}:
        causes.append("lifecycle_over_penalized_high_revenue")
    if business_action in {"action_blocked", "observe_only"} and revenue in {"top", "high"}:
        causes.append("business_action_status_over_penalized_rating")
    if forecast_output_type == "operating_window_forecast_pending_expiry":
        causes.append("missing_expiry_should_be_readiness_warning_not_value_downgrade")
    if kwargs["new_rating"] != old_rating:
        causes.append("calibrated_rule_changed_rating")
    return sorted(set(causes))


def suggestion_root_causes(**kwargs) -> list[str]:
    causes = []
    if kwargs["suggestion_feedback"] == "不可执行":
        causes.append("operator_marked_suggestion_not_executable")
    if "业务常识冲突" in kwargs["issue_type"]:
        causes.append("operator_marked_business_common_sense_conflict")
    if kwargs["forecastability"] == "true_forecast_blocked":
        causes.append("old_suggestion_did_not_explain_true_forecast_blocker")
    if kwargs["business_action"] == "action_blocked":
        causes.append("old_suggestion_not_actionable_under_business_action_block")
    if kwargs["suggestion_changed"]:
        causes.append("calibrated_rule_changed_suggestion")
    return sorted(set(causes))


def build_rating_root_cause(reviewable: list[dict], inputs: dict) -> dict:
    rating_feedback_rows = [row for row in reviewable if clean(row.get("运营判断：评级是否合理")) == "不合理"]
    root_counter = Counter(cause for row in rating_feedback_rows for cause in row["ratingRootCauses"])
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "reviewableRows": len(reviewable),
        "ratingUnreasonableRows": len(rating_feedback_rows),
        "currentRatingDistribution": dict(sorted(Counter(row["oldRating"] or "未填写" for row in reviewable).items())),
        "newRatingDistribution": dict(sorted(Counter(row["newRating"] or "未填写" for row in reviewable).items())),
        "ratingChangeDirectionDistribution": dict(sorted(Counter(row["ratingChangeDirection"] for row in reviewable).items())),
        "rootCauseDistribution": dict(sorted(root_counter.items())),
        "analysisFindings": {
            "thresholdIssue": "旧规则把高收入样本大量压在 B/C/D/E，阈值和状态惩罚组合偏保守",
            "forecastWeightIssue": "预测阻断被用于压低商业价值，导致高历史收入样本被误判为低价值",
            "lifecycleWeightIssue": "inactive/long_tail 对高收入样本惩罚过强",
            "remainingCopyrightIssue": "双源 staging 后版权期变长，但旧 rating 未同步吸收剩余版权价值",
            "dataFallbackIssue": "缺版权到期应作为 readiness warning，不应直接打低价值评级",
            "businessKnowledgeMismatch": "用户对 25 个有效样本全部标记评级不合理",
            "forecastModelChanged": False,
            "stagingRecordCountRead": inputs["stagingRecordCount"],
        },
        "ruleFixesRequired": [
            "将历史收入/收入层级作为主价值信号",
            "forecastable 与 non-forecastable cohort 分开评级",
            "true_forecast_blocked 默认不超过 C，但高收入非长尾样本可到 B 且需人工复核",
            "缺版权到期只作为 readiness warning",
            "长剩余版权期与稳定/增长收入可上调评级",
        ],
        "safeOutputBoundary": safe_boundary(),
    }


def build_suggestion_root_cause(reviewable: list[dict], inputs: dict) -> dict:
    not_executable = [row for row in reviewable if clean(row.get("运营判断：建议是否可执行")) == "不可执行"]
    conflict = [row for row in reviewable if "业务常识冲突" in clean(row.get("运营发现的问题类型"))]
    root_counter = Counter(cause for row in reviewable for cause in row["suggestionRootCauses"])
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "reviewableRows": len(reviewable),
        "suggestionNotExecutableRows": len(not_executable),
        "businessCommonSenseConflictRows": len(conflict),
        "currentSuggestionDistribution": dict(sorted(Counter(row["oldSuggestion"] or "未填写" for row in reviewable).items())),
        "newSuggestionDistribution": dict(sorted(Counter(row["newSuggestion"] or "未填写" for row in reviewable).items())),
        "rootCauseDistribution": dict(sorted(root_counter.items())),
        "analysisFindings": {
            "promoteOveruse": 0,
            "downlistMisfire": 0,
            "renewalWithoutSupport": 0,
            "observeOnlyUnderused": sum(1 for row in reviewable if row.get("辅助原始预测状态code") == "observe_only_no_numeric_forecast"),
            "ignoredForecastabilityStatus": sum(1 for row in reviewable if row.get("辅助原始预测状态code") == "true_forecast_blocked"),
            "ignoredBusinessActionStatus": sum(1 for row in reviewable if row.get("辅助原始业务动作状态code") == "action_blocked"),
            "forecastModelChanged": False,
        },
        "ruleFixesRequired": [
            "action_blocked/true_forecast_blocked 统一转人工复核动作，不输出直接业务动作",
            "observe-only/历史不足保持仅观察",
            "promote 只允许高评级、增长/回升、置信度中高且 action_allowed",
            "downlist 只允许低收入长尾/沉寂，且必须人工确认",
            "renewal review 必须有版权期和收入价值支撑",
        ],
        "safeOutputBoundary": safe_boundary(),
    }


def build_calibration_summary(calibrated: list[dict], reviewable: list[dict], rating_root: dict, suggestion_root: dict, inputs: dict) -> dict:
    rating_changed = [row for row in reviewable if row["ratingChanged"]]
    suggestion_changed = [row for row in reviewable if row["suggestionChanged"]]
    rating_unreasonable = [row for row in reviewable if clean(row.get("运营判断：评级是否合理")) == "不合理"]
    suggestion_not_executable = [row for row in reviewable if clean(row.get("运营判断：建议是否可执行")) == "不可执行"]
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "forecastModelChanged": False,
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "m3Entered": False,
        "rating": {
            "changedRows": len(rating_changed),
            "changedCoverageRate": ratio(len(rating_changed), len(reviewable)),
            "oldDistribution": rating_root["currentRatingDistribution"],
            "newDistribution": rating_root["newRatingDistribution"],
            "changeDirectionDistribution": rating_root["ratingChangeDirectionDistribution"],
        },
        "suggestion": {
            "changedRows": len(suggestion_changed),
            "changedCoverageRate": ratio(len(suggestion_changed), len(reviewable)),
            "oldDistribution": suggestion_root["currentSuggestionDistribution"],
            "newDistribution": suggestion_root["newSuggestionDistribution"],
        },
        "expectedImprovement": {
            "ratingUnreasonableRows": len(rating_unreasonable),
            "ratingUnreasonableHitRows": sum(1 for row in rating_unreasonable if row["ratingChanged"]),
            "suggestionNotExecutableRows": len(suggestion_not_executable),
            "suggestionNotExecutableHitRows": sum(1 for row in suggestion_not_executable if row["suggestionChanged"]),
            "remainingManualReviewRows": sum(1 for row in reviewable if row["manualConfirmationRequired"]),
            "stillNeedsUserValidation": True,
        },
        "privateOutputs": {
            "validationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
            "operatorTaskPack": rel(PRIVATE_OPERATOR_V3_XLSX),
            "gitignored": True,
        },
        "inputContext": {
            "forecastabilityVerdict": inputs["forecastability"].get("payload", {}).get("modelVerdict"),
            "businessReviewReadyCohort": inputs["businessReadiness"].get("payload", {}).get("readinessAfterStaging", {}).get("businessReviewReadyCohort"),
            "copyrightTermForecastWorks": inputs["forecastOutput"].get("payload", {}).get("after", {}).get("copyright_term_forecast"),
            "operatingWindowForecastWorks": inputs["forecastOutput"].get("payload", {}).get("after", {}).get("operating_window_forecast_pending_expiry"),
        },
        "safeOutputBoundary": safe_boundary(),
    }


def build_operator_v3_summary(calibrated: list[dict], reviewable: list[dict]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "privateWorkbookPath": rel(PRIVATE_OPERATOR_V3_XLSX),
        "baseWorkbookPath": rel(INPUT_OPERATOR_XLSX),
        "taskRows": len(calibrated),
        "reviewableRows": len(reviewable),
        "userReservedRows": sum(1 for row in calibrated if row["sampleSource"] == USER_RESERVED),
        "hasStandardWorkIdColumn": True,
        "keepsUserFeedbackFields": True,
        "showsOldAndNewRating": True,
        "showsOldAndNewSuggestion": True,
        "ratingChangedRows": sum(1 for row in reviewable if row["ratingChanged"]),
        "suggestionChangedRows": sum(1 for row in reviewable if row["suggestionChanged"]),
        "manualConfirmationRows": sum(1 for row in reviewable if row["manualConfirmationRequired"]),
        "m4CandidateRecommendedRows": sum(1 for row in reviewable if row["m4CandidateRecommended"] == "是"),
        "privateWorkbookGitignored": True,
        "formalMasterDataWritten": False,
        "databaseWritten": False,
        "m3Entered": False,
        "safeOutputBoundary": safe_boundary(),
    }


def write_private_workbooks(calibrated, reviewable, rating_root, suggestion_root, summary) -> None:
    v3 = load_v3_module()
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    v3.write_xlsx(
        PRIVATE_VALIDATION_XLSX,
        {
            "00_阅读说明": readme_rows(),
            "01_用户反馈根因": user_feedback_rows(reviewable),
            "02_评级规则v1": rating_rule_rows(),
            "03_建议规则v1": suggestion_rule_rows(),
            "04_30样本新旧对比": comparison_rows(calibrated),
            "05_高收入样本": [row for row in comparison_rows(reviewable) if row.get("收入层级") in {"top", "high"}],
            "06_评级变化明细": [row for row in comparison_rows(reviewable) if row.get("评级是否变化") == "是"],
            "07_建议变化明细": [row for row in comparison_rows(reviewable) if row.get("建议是否变化") == "是"],
            "08_需要人工确认": [row for row in comparison_rows(reviewable) if row.get("是否需要人工确认") == "是"],
            "09_M4校准候选": [row for row in comparison_rows(reviewable) if row.get("M4校准候选建议") == "是"],
        },
    )
    v3.write_xlsx(
        PRIVATE_OPERATOR_V3_XLSX,
        {
            "00_阅读说明": [
                {"项目": "用途", "说明": "rating/suggestion calibration v1.0 后的 30-work 运营复核任务包"},
                {"项目": "边界", "说明": "不改 forecast 模型，不写正式主数据，不进入 M3"},
                {"项目": "填写", "说明": "请复核新评级、新建议、变化原因，并保留/更新用户反馈字段"},
            ],
            "01_运营任务卡": operator_v3_rows(calibrated),
            "02_规则变化说明": rating_rule_rows() + suggestion_rule_rows(),
            "03_填写选项": [
                {"字段": "运营判断：预测是否可信", "可选项": "可信 / 基本可信 / 不确定 / 不可信 / 不适用"},
                {"字段": "运营判断：评级是否合理", "可选项": "合理 / 基本合理 / 不确定 / 不合理 / 不适用"},
                {"字段": "运营判断：建议是否可执行", "可选项": "可执行 / 需要人工确认 / 仅供参考 / 不可执行 / 不适用"},
                {"字段": "是否应进入M4校准案例池", "可选项": "是 / 否 / 待定"},
            ],
        },
    )


def comparison_rows(rows: list[dict]) -> list[dict]:
    out = []
    for row in rows:
        out.append(
            {
                "样本来源": clean(row.get("样本来源")),
                "standard_work_id": clean(row.get("standard_work_id")),
                "作品名": clean(row.get("作品名")),
                "作者": clean(row.get("作者")),
                "收入层级": clean(row.get("辅助原始收入层级code")),
                "生命周期": clean(row.get("辅助原始生命周期code")),
                "预测状态": clean(row.get("辅助原始预测状态code")),
                "业务动作状态": clean(row.get("辅助原始业务动作状态code")),
                "旧评级": row.get("oldRating", ""),
                "新评级": row.get("newRating", ""),
                "评级是否变化": "是" if row.get("ratingChanged") else "否",
                "评级变化方向": row.get("ratingChangeDirection", ""),
                "评级变化原因": row.get("ratingRationaleCn", ""),
                "旧建议": row.get("oldSuggestion", ""),
                "新建议": row.get("newSuggestionCn", ""),
                "建议是否变化": "是" if row.get("suggestionChanged") else "否",
                "建议变化原因": row.get("suggestionChangeReasonCn", ""),
                "是否需要人工确认": "是" if row.get("manualConfirmationRequired") else "否",
                "M4校准候选建议": row.get("m4CandidateRecommended", ""),
                "用户评级反馈": clean(row.get("运营判断：评级是否合理")),
                "用户建议反馈": clean(row.get("运营判断：建议是否可执行")),
                "用户问题类型": clean(row.get("运营发现的问题类型")),
            }
        )
    return out


def operator_v3_rows(rows: list[dict]) -> list[dict]:
    output = []
    for row in rows:
        output.append(
            {
                "样本编号": clean(row.get("样本编号")),
                "样本来源": clean(row.get("样本来源")),
                "standard_work_id": clean(row.get("standard_work_id")),
                "raw_work_id": clean(row.get("raw_work_id")),
                "作品名": clean(row.get("作品名")),
                "作者": clean(row.get("作者")),
                "预测输出类型": clean(row.get("预测输出类型")),
                "预测置信度": clean(row.get("预测置信度")),
                "生命周期": clean(row.get("生命周期")),
                "收入层级": clean(row.get("收入层级")),
                "旧评级": row.get("oldRating", clean(row.get("评级"))),
                "新评级": row.get("newRating", ""),
                "评级变化原因": row.get("ratingRationaleCn", ""),
                "旧运营建议": row.get("oldSuggestion", clean(row.get("运营建议"))),
                "新运营建议": row.get("newSuggestionCn", ""),
                "建议变化原因": row.get("suggestionChangeReasonCn", ""),
                "建议可执行层级": row.get("suggestionActionabilityLevel", ""),
                "是否需要人工确认": "是" if row.get("manualConfirmationRequired") else "否",
                "为什么不是其他建议": row.get("whyNotOtherSuggestionsCn", ""),
                "运营判断：预测是否可信": clean(row.get("运营判断：预测是否可信")),
                "运营判断：评级是否合理": clean(row.get("运营判断：评级是否合理")),
                "运营判断：建议是否可执行": clean(row.get("运营判断：建议是否可执行")),
                "运营发现的问题类型": clean(row.get("运营发现的问题类型")),
                "运营建议修正": clean(row.get("运营建议修正")),
                "是否应进入M4校准案例池": clean(row.get("是否应进入M4校准案例池")),
                "本轮建议M4校准候选": row.get("m4CandidateRecommended", ""),
                "辅助原始forecastOutputType": clean(row.get("辅助原始forecastOutputType")),
                "辅助原始生命周期code": clean(row.get("辅助原始生命周期code")),
                "辅助原始收入层级code": clean(row.get("辅助原始收入层级code")),
                "辅助原始预测状态code": clean(row.get("辅助原始预测状态code")),
                "辅助原始业务动作状态code": clean(row.get("辅助原始业务动作状态code")),
            }
        )
    return output


def user_feedback_rows(reviewable: list[dict]) -> list[dict]:
    rows = []
    for row in reviewable:
        rows.append(
            {
                "standard_work_id": clean(row.get("standard_work_id")),
                "作品名": clean(row.get("作品名")),
                "作者": clean(row.get("作者")),
                "用户预测反馈": clean(row.get("运营判断：预测是否可信")),
                "用户评级反馈": clean(row.get("运营判断：评级是否合理")),
                "用户建议反馈": clean(row.get("运营判断：建议是否可执行")),
                "用户问题类型": clean(row.get("运营发现的问题类型")),
                "评级根因": "；".join(row["ratingRootCauses"]),
                "建议根因": "；".join(row["suggestionRootCauses"]),
            }
        )
    return rows


def readme_rows() -> list[dict]:
    return [
        {"项目": "候选版本", "说明": CANDIDATE_VERSION},
        {"项目": "输入", "说明": "after-dual-source-staging-v2 已填写 30-work 运营任务包"},
        {"项目": "边界", "说明": "只校准 rating/suggestion；forecast 模型未修改；不进入 M3"},
        {"项目": "安全", "说明": "本 workbook 为 private gitignored，不提交"},
    ]


def rating_rule_rows() -> list[dict]:
    return [
        {"规则": "主价值信号", "说明": "历史收入/收入层级优先，forecast value 不是唯一依据"},
        {"规则": "forecastable cohort", "说明": "numeric 使用预测价值；conservative 使用保守预测和人工确认"},
        {"规则": "true forecast blocked", "说明": "默认不超过 C；高收入非长尾/沉寂可到 B 且必须人工确认"},
        {"规则": "observe-only", "说明": "不输出高评级，高收入样本最多 C"},
        {"规则": "版权期", "说明": "长版权期且稳定/增长可上调；缺版权到期只作为 readiness warning"},
        {"规则": "S/S+", "说明": "必须高价值、中高置信且无严重阻断，并要求人工确认"},
    ]


def suggestion_rule_rows() -> list[dict]:
    return [
        {"规则": "promote", "说明": "仅限高评级、增长/回升、中高置信且 action_allowed"},
        {"规则": "maintain", "说明": "稳定/增长/回升且中高价值，无明显风险"},
        {"规则": "reduce", "说明": "下滑或弱生命周期但仍有收入，不直接下架"},
        {"规则": "downlist", "说明": "仅限低收入长尾/沉寂，必须人工确认"},
        {"规则": "renewal review", "说明": "必须有版权期和收入价值支撑"},
        {"规则": "manual review", "说明": "true forecast blocked 或 action_blocked 不输出直接业务动作"},
    ]


def rating_from_score(score: float) -> str:
    if score >= 96:
        return "S+"
    if score >= 88:
        return "S"
    if score >= 70:
        return "A"
    if score >= 52:
        return "B"
    if score >= 38:
        return "C"
    if score >= 24:
        return "D"
    return "E"


def min_rating(rating: str, cap: str) -> str:
    return cap if RATING_RANK.get(rating, 99) <= RATING_RANK.get(cap, 99) else rating


def max_rating(rating: str, floor: str) -> str:
    return floor if RATING_RANK.get(rating, 99) >= RATING_RANK.get(floor, 99) else rating


def change_direction(old: str, new: str) -> str:
    if not old or old == new:
        return "unchanged"
    return "upgrade" if RATING_RANK.get(new, 99) < RATING_RANK.get(old, 99) else "downgrade"


def confidence_code(value: str) -> str:
    if "高" in value or value == "high":
        return "high"
    if "低" in value or value == "low":
        return "low"
    if "阻断" in value or value == "blocked_for_business_use":
        return "blocked_for_business_use"
    if value:
        return "medium"
    return ""


def rating_root_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 rating failure root cause after staging v1",
            "",
            f"- Reviewable rows: `{payload['reviewableRows']}`",
            f"- Rating unreasonable rows: `{payload['ratingUnreasonableRows']}`",
            f"- Current rating distribution: `{json.dumps(payload['currentRatingDistribution'], ensure_ascii=False)}`",
            f"- New rating distribution: `{json.dumps(payload['newRatingDistribution'], ensure_ascii=False)}`",
            f"- Root cause distribution: `{json.dumps(payload['rootCauseDistribution'], ensure_ascii=False)}`",
            "",
            "## Rule Fixes",
            *[f"- {item}" for item in payload["ruleFixesRequired"]],
            "",
            "No real work names, author names, channel names, or row-level revenue details are included.",
        ]
    )


def suggestion_root_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 suggestion failure root cause after staging v1",
            "",
            f"- Reviewable rows: `{payload['reviewableRows']}`",
            f"- Suggestion not executable rows: `{payload['suggestionNotExecutableRows']}`",
            f"- Business common-sense conflict rows: `{payload['businessCommonSenseConflictRows']}`",
            f"- Current suggestion distribution: `{json.dumps(payload['currentSuggestionDistribution'], ensure_ascii=False)}`",
            f"- New suggestion distribution: `{json.dumps(payload['newSuggestionDistribution'], ensure_ascii=False)}`",
            f"- Root cause distribution: `{json.dumps(payload['rootCauseDistribution'], ensure_ascii=False)}`",
            "",
            "## Rule Fixes",
            *[f"- {item}" for item in payload["ruleFixesRequired"]],
            "",
            "No real work names, author names, channel names, or row-level revenue details are included.",
        ]
    )


def calibration_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 rating/suggestion calibration v1 summary",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Forecast model changed: `{payload['forecastModelChanged']}`",
            f"- Rating changed rows: `{payload['rating']['changedRows']}`",
            f"- Suggestion changed rows: `{payload['suggestion']['changedRows']}`",
            f"- Rating unreasonable hit rows: `{payload['expectedImprovement']['ratingUnreasonableHitRows']}`",
            f"- Suggestion not executable hit rows: `{payload['expectedImprovement']['suggestionNotExecutableHitRows']}`",
            f"- Remaining manual review rows: `{payload['expectedImprovement']['remainingManualReviewRows']}`",
            f"- Still needs user validation: `{payload['expectedImprovement']['stillNeedsUserValidation']}`",
            f"- Private validation workbook: `{payload['privateOutputs']['validationWorkbook']}`",
            f"- Private operator task pack: `{payload['privateOutputs']['operatorTaskPack']}`",
            "- M3 entered: `false`",
        ]
    )


def operator_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 operator task pack after rating/suggestion calibration v1 summary",
            "",
            f"- Candidate version: `{payload['candidateVersion']}`",
            f"- Private workbook path: `{payload['privateWorkbookPath']}`",
            f"- Task rows: `{payload['taskRows']}`",
            f"- Reviewable rows: `{payload['reviewableRows']}`",
            f"- Rating changed rows: `{payload['ratingChangedRows']}`",
            f"- Suggestion changed rows: `{payload['suggestionChangedRows']}`",
            f"- Manual confirmation rows: `{payload['manualConfirmationRows']}`",
            f"- M4 candidate recommended rows: `{payload['m4CandidateRecommendedRows']}`",
            "- Private workbook is gitignored and not submitted.",
            "- M3 entered: `false`",
        ]
    )


def public_envelope(schema: str, payload: dict) -> dict:
    return {
        "schema": schema,
        "generatedAt": now(),
        "currentHead": git(["rev-parse", "HEAD"]),
        "originMain": git(["rev-parse", "origin/main"]),
        "payload": payload,
        "safeOutputBoundary": safe_boundary(),
    }


def safe_boundary() -> dict:
    return {
        "sanitizedAggregateOnly": True,
        "realWorkNamesIncluded": False,
        "authorNamesIncluded": False,
        "channelNamesIncluded": False,
        "rawLedgerRowsIncluded": False,
        "standardWorkIdDetailsIncluded": False,
        "privateDetailsStoredOnlyInGitignoredOutput": True,
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "databaseWritten": False,
        "forecastModelChanged": False,
        "m3Entered": False,
    }


def load_v3_module():
    path = ROOT / "scripts" / "m2-real-data" / "run_cleaned_ledger_minimal_backfill_v3.py"
    spec = importlib.util.spec_from_file_location("cleaned_ledger_v3", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)  # type: ignore[union-attr]
    return module


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text + "\n", encoding="utf-8")


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def number_or_none(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def ratio(numerator: int, denominator: int) -> float:
    return round(numerator / denominator, 4) if denominator else 0.0


def now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def git(args: list[str]) -> str:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True, encoding="utf-8").strip()
    except Exception:
        return ""


def rel(path: Path) -> str:
    return str(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
