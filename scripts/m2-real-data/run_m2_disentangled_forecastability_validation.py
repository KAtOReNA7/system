from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))
sys.path.insert(0, str(SCRIPT_DIR))

import numpy as np
import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Alignment = Font = PatternFill = get_column_letter = None

import run_m2_forecast_model_bakeoff as bake
import run_m2_forecastability_gated_validation as gated

OUTPUT_DIR = ROOT / "docs" / "analysis" / "m2-real-data"
PRIVATE_DIR = ROOT / "data" / "private-output" / "m2-algorithm-validation"

DISENTANGLE_JSON = OUTPUT_DIR / "M2-manual-review-forecastability-disentanglement-v1.json"
DISENTANGLE_MD = OUTPUT_DIR / "M2-manual-review-forecastability-disentanglement-v1.md"
COVERAGE_JSON = OUTPUT_DIR / "M2-disentangled-forecastability-coverage-v1.json"
COVERAGE_MD = OUTPUT_DIR / "M2-disentangled-forecastability-coverage-v1.md"
VALIDATION_JSON = OUTPUT_DIR / "M2-disentangled-forecast-model-validation-v1.json"
VALIDATION_MD = OUTPUT_DIR / "M2-disentangled-forecast-model-validation-v1.md"
READINESS_JSON = OUTPUT_DIR / "M2-disentangled-algorithm-business-readiness-v1.json"
READINESS_MD = OUTPUT_DIR / "M2-disentangled-algorithm-business-readiness-v1.md"

PRIVATE_XLSX = PRIVATE_DIR / "m2-disentangled-forecastability-v1-validation.xlsx"
PRIVATE_DETAIL_JSON = PRIVATE_DIR / "m2-disentangled-forecastability-v1-private-detail.json"

DISENTANGLED_CANDIDATE = "m2-realdata-dev-disentangled-forecast-v1.0"
MODEL_ID = "model_g_disentangled_forecast"

NUMERIC_STATUS = "numeric_forecast_eligible"
CONSERVATIVE_STATUS = "conservative_numeric_forecast"
OBSERVE_STATUS = "observe_only_no_numeric_forecast"
TRUE_BLOCKED_STATUS = "true_forecast_blocked"
FORECASTABILITY_STATUSES = [NUMERIC_STATUS, CONSERVATIVE_STATUS, OBSERVE_STATUS, TRUE_BLOCKED_STATUS]

READY_STATUS = "ready_for_local_algorithm_validation"
FORMAL_BLOCKED_STATUS = "formal_release_blocked"
WAIVER_STATUS = "waiver_required"
DATA_FIX_STATUS = "data_fix_required"
MAPPING_STATUS = "mapping_activation_required"
FORMAL_STATUSES = [READY_STATUS, FORMAL_BLOCKED_STATUS, WAIVER_STATUS, DATA_FIX_STATUS, MAPPING_STATUS]

ACTION_ALLOWED_STATUS = "action_allowed"
MANUAL_CONFIRMATION_STATUS = "manual_confirmation_required"
ACTION_BLOCKED_STATUS = "action_blocked"
ACTION_OBSERVE_STATUS = "observe_only"
BUSINESS_ACTION_STATUSES = [
    ACTION_ALLOWED_STATUS,
    MANUAL_CONFIRMATION_STATUS,
    ACTION_BLOCKED_STATUS,
    ACTION_OBSERVE_STATUS,
]

SPIKE_RISKS = {"abnormal_spike", "buyout_or_oneoff_income"}
DATA_FIX_RISKS = {
    "copyright_date_conflict",
    "aggregate_projection_gap",
    "missing_basic_info",
    "metadata_gap",
}
WAIVER_RISKS = {"missing_copyright_end", "copyright_expiry", "expiry_high_value"}
MAPPING_RISKS = {"mapping_uncertainty", "mapping_not_active", "mapping_version_inactive"}
ACTIONABLE_SUGGESTIONS = {"promote", "downlist_or_suspend", "renewal_review"}


def safe_float(value, default: float = 0.0) -> float:
    return bake.safe_float(value, default)


def safe_int(value, default: int = 0) -> int:
    return bake.safe_int(value, default)


def rounded(value, digits: int = 4):
    return bake.rounded(value, digits)


def write_json(path: Path, payload: dict) -> None:
    bake.write_json(path, payload)


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    return bake.markdown_table(rows, columns)


def percent(part, total) -> float:
    total = safe_float(total)
    if total <= 0:
        return 0.0
    return rounded(safe_float(part) / total)


def distribution(values, keys: list[str] | None = None) -> dict:
    return bake.distribution(values, keys)


def materiality_bucket_for(amount: float, rank_percent: float | None) -> str:
    return gated.materiality_bucket_for(amount, rank_percent)


def safe_boundary() -> dict:
    return {
        "rawRowsWritten": False,
        "realWorkNamesWritten": False,
        "realAuthorNamesWritten": False,
        "realChannelNamesWritten": False,
        "connectionStringsWritten": False,
        "privateWorkbookGitignored": True,
        "notFinalReleaseApproval": True,
        "m3Started": False,
    }


def build_forecastability(status: str, reasons: list[str], confidence: str, materiality_bucket: str, action: str) -> dict:
    can_use = status in {NUMERIC_STATUS, CONSERVATIVE_STATUS}
    return {
        "forecastabilityStatus": status,
        "forecastabilityReasonCodes": reasons,
        "forecastabilityConfidence": confidence,
        "materialityBucket": materiality_bucket,
        "canUseNumericForecast": can_use,
        "canUseForBusinessReview": status == NUMERIC_STATUS,
        "requiredForecastabilityAction": action,
    }


def evaluate_disentangled_forecastability(features: dict) -> dict:
    risks = set(features.get("riskCodes") or [])
    lifecycle = str(features.get("lifecycle") or "")
    rating = str(features.get("rating") or "")
    revenue_scale = str(features.get("revenueScale") or "")
    active_months = safe_int(features.get("activeMonthCount"))
    zero_months = safe_int(features.get("zeroRevenueMonthCount"))
    total_revenue = safe_float(features.get("totalHistoricalRevenue"))
    recent_revenue = safe_float(features.get("recentRevenue", features.get("last12MonthRevenue", 0)))
    volatility = safe_float(features.get("volatility", features.get("last6CoefficientOfVariation", 0)))
    peak_share = safe_float(features.get("peakShare", features.get("peakMonthShare", 0)))
    remaining_months = safe_int(features.get("remainingMonthsForForecast"), 12)
    materiality_bucket = str(features.get("materialityBucket") or materiality_bucket_for(total_revenue, features.get("materialityRankPercent")))

    if active_months <= 0 or total_revenue <= 0:
        return build_forecastability(
            TRUE_BLOCKED_STATUS,
            ["no_backtestable_revenue_history"],
            "blocked_for_business_use",
            materiality_bucket,
            "exclude_from_numeric_forecast_baseline",
        )
    if active_months < 6 or lifecycle == "insufficient_history":
        return build_forecastability(
            TRUE_BLOCKED_STATUS,
            ["insufficient_revenue_time_series"],
            "blocked_for_business_use",
            materiality_bucket,
            "collect_more_revenue_history_before_numeric_forecast",
        )

    spike_risk = bool(risks.intersection(SPIKE_RISKS))
    spike_damped_backtest_passed = bool(features.get("spikeDampedBacktestPassed")) or (
        spike_risk and active_months >= 12 and recent_revenue > 100 and peak_share < 0.85
    )
    if spike_risk and not spike_damped_backtest_passed:
        return build_forecastability(
            TRUE_BLOCKED_STATUS,
            ["unresolved_spike_or_oneoff_income"],
            "blocked_for_business_use",
            materiality_bucket,
            "manual_review_or_spike_damped_backtest_required",
        )

    zero_heavy = zero_months >= max(12, active_months * 2)
    near_zero = materiality_bucket == "near_zero" or recent_revenue <= 10
    low_materiality = materiality_bucket == "bottom_50_percent" or rating in {"D", "E"} or revenue_scale in {"low", "long_tail"}
    tail_pattern = lifecycle in {"inactive", "long_tail"} or zero_heavy or near_zero or low_materiality
    material = materiality_bucket in {"top_1_percent", "top_5_percent", "top_10_percent", "middle_40_percent"}

    if tail_pattern:
        if material and total_revenue > 1000 and active_months >= 12 and recent_revenue > 10:
            return build_forecastability(
                CONSERVATIVE_STATUS,
                ["material_tail_or_zero_heavy_but_backtestable"],
                "low",
                materiality_bucket,
                "conservative_numeric_forecast_only",
            )
        return build_forecastability(
            OBSERVE_STATUS,
            ["low_materiality_or_zero_heavy_pattern"],
            "low",
            materiality_bucket,
            "observe_only_no_business_numeric_forecast",
        )

    if (
        material
        and active_months >= 12
        and recent_revenue > 100
        and volatility <= 1.2
        and remaining_months <= 120
        and lifecycle in {"growth", "stable", "rebound"}
    ):
        return build_forecastability(
            NUMERIC_STATUS,
            ["material_stable_history"],
            "high",
            materiality_bucket,
            "numeric_forecast_and_business_review_allowed",
        )

    return build_forecastability(
        CONSERVATIVE_STATUS,
        ["bounded_but_forecastable_with_conservative_interval"],
        "low" if spike_risk else "medium",
        materiality_bucket,
        "conservative_numeric_forecast_only",
    )


def classify_formal_readiness(features: dict) -> dict:
    risks = set(features.get("riskCodes") or [])
    if bool(features.get("mappingActivationRequired")) or risks.intersection(MAPPING_RISKS):
        return {
            "formalReadinessStatus": MAPPING_STATUS,
            "formalReadinessReasonCodes": ["mapping_activation_required"],
            "formalReadinessBlocksRelease": True,
            "formalReadinessBlocksLocalForecast": False,
            "requiredFormalAction": "prepare_or_activate_mapping_version_before_formal_release",
        }
    if risks.intersection(DATA_FIX_RISKS):
        return {
            "formalReadinessStatus": DATA_FIX_STATUS,
            "formalReadinessReasonCodes": ["formal_data_fix_required"],
            "formalReadinessBlocksRelease": True,
            "formalReadinessBlocksLocalForecast": False,
            "requiredFormalAction": "fix_formal_source_or_metadata_before_release",
        }
    if bool(features.get("forecastFallbackUsed")) or bool(features.get("waiverRequired")) or risks.intersection(WAIVER_RISKS):
        return {
            "formalReadinessStatus": WAIVER_STATUS,
            "formalReadinessReasonCodes": ["formal_waiver_required"],
            "formalReadinessBlocksRelease": True,
            "formalReadinessBlocksLocalForecast": False,
            "requiredFormalAction": "record_formal_waiver_before_release",
        }
    if bool(features.get("releaseApprovalMissing")) or bool(features.get("formalFlagsIncomplete")):
        return {
            "formalReadinessStatus": FORMAL_BLOCKED_STATUS,
            "formalReadinessReasonCodes": ["formal_release_approval_missing"],
            "formalReadinessBlocksRelease": True,
            "formalReadinessBlocksLocalForecast": False,
            "requiredFormalAction": "complete_formal_release_approval",
        }
    return {
        "formalReadinessStatus": READY_STATUS,
        "formalReadinessReasonCodes": ["formal_readiness_not_blocking_local_algorithm_validation"],
        "formalReadinessBlocksRelease": False,
        "formalReadinessBlocksLocalForecast": False,
        "requiredFormalAction": "local_algorithm_validation_allowed",
    }


def evaluate_business_action(features: dict) -> dict:
    forecastability_status = str(features.get("forecastabilityStatus") or "")
    suggestions = set(features.get("suggestionCodes") or [])
    if forecastability_status == TRUE_BLOCKED_STATUS:
        return {
            "businessActionStatus": ACTION_BLOCKED_STATUS,
            "businessActionReasonCodes": ["true_forecast_blocked_before_action"],
            "businessActionBlocksForecast": False,
            "requiredBusinessAction": "resolve_forecast_blocker_before_business_action",
        }
    if forecastability_status == OBSERVE_STATUS:
        return {
            "businessActionStatus": ACTION_OBSERVE_STATUS,
            "businessActionReasonCodes": ["observe_only_forecastability_status"],
            "businessActionBlocksForecast": False,
            "requiredBusinessAction": "observe_without_promote_downlist_or_renewal_action",
        }
    manual = sorted(suggestions.intersection(ACTIONABLE_SUGGESTIONS))
    if manual or bool(features.get("manualConfirmationRequired")):
        return {
            "businessActionStatus": MANUAL_CONFIRMATION_STATUS,
            "businessActionReasonCodes": manual or ["manual_confirmation_required"],
            "businessActionBlocksForecast": False,
            "requiredBusinessAction": "manual_confirmation_before_business_action",
        }
    return {
        "businessActionStatus": ACTION_ALLOWED_STATUS,
        "businessActionReasonCodes": ["no_business_action_blocker"],
        "businessActionBlocksForecast": False,
        "requiredBusinessAction": "business_action_allowed_if_formal_policy_allows",
    }


def prepare_inputs() -> dict:
    return gated.prepare_inputs()


def build_current_gate_frame(evaluated: pd.DataFrame, final_outputs: pd.DataFrame) -> pd.DataFrame:
    frame = evaluated.copy()
    frame["_totalRevenue"] = frame["totalHistoricalRevenue"].astype(float)
    frame = frame.sort_values(["_totalRevenue", "standardWorkId"], ascending=[False, True]).reset_index(drop=True)
    total_count = max(1, len(frame))
    frame["materialityRankPercent"] = [(index + 1) / total_count for index in range(total_count)]
    frame["materialityBucket"] = frame.apply(
        lambda row: materiality_bucket_for(row.totalHistoricalRevenue, row.materialityRankPercent),
        axis=1,
    )
    output_lookup = final_outputs.set_index("workKey").to_dict(orient="index")
    rows = []
    for index, row in frame.iterrows():
        features = {
            "riskCodes": row.riskCodes,
            "suggestionCodes": row.suggestionCodes,
            "lifecycle": row.lifecycle,
            "rating": row.rating,
            "revenueScale": row.revenueBucket,
            "activeMonthCount": row.activeMonthCount,
            "zeroRevenueMonthCount": row.zeroRevenueMonthCount,
            "totalHistoricalRevenue": row.totalHistoricalRevenue,
            "recentRevenue": row.last12MonthRevenue,
            "volatility": row.last6CoefficientOfVariation,
            "peakShare": row.peakMonthShare,
            "remainingMonthsForForecast": row.remainingMonthsForForecast,
            "forecastFallbackUsed": row.forecastFallbackUsed,
            "materialityBucket": row.materialityBucket,
        }
        old_gate = gated.evaluate_gate(features)
        forecastability = evaluate_disentangled_forecastability(features)
        formal = classify_formal_readiness(features)
        action = evaluate_business_action(
            {
                "forecastabilityStatus": forecastability["forecastabilityStatus"],
                "suggestionCodes": row.suggestionCodes,
            }
        )
        output = output_lookup.get(str(row.standardWorkId), {})
        rows.append(
            {
                "workKey": str(row.standardWorkId),
                "anonymousId": output.get("anonymousId", f"D{index + 1:04d}"),
                "materialityRankPercent": row.materialityRankPercent,
                "materialityBucket": row.materialityBucket,
                "rating": row.rating,
                "lifecycle": row.lifecycle,
                "revenueBucket": row.revenueBucket,
                "riskBucket": row.riskBucket,
                "suggestionBucket": row.suggestionBucket,
                "remainingCopyrightBucket": row.remainingCopyrightBucket,
                "totalHistoricalRevenue": safe_float(row.totalHistoricalRevenue),
                "last12MonthRevenue": safe_float(row.last12MonthRevenue),
                "activeMonthCount": safe_int(row.activeMonthCount),
                "zeroRevenueMonthCount": safe_int(row.zeroRevenueMonthCount),
                "oldForecastabilityStatus": old_gate["forecastabilityStatus"],
                "oldReasonCodes": old_gate["reasonCodes"],
                **forecastability,
                **formal,
                **action,
                "selectedModel": output.get("selectedModel"),
                "baseForecast": output.get("baseForecast") if forecastability["canUseNumericForecast"] else None,
                "optimisticForecast": output.get("optimisticForecast") if forecastability["canUseNumericForecast"] else None,
                "pessimisticForecast": output.get("pessimisticForecast") if forecastability["canUseNumericForecast"] else None,
                "forecastConfidence": output.get("forecastConfidence") if forecastability["canUseNumericForecast"] else "blocked_for_business_use",
            }
        )
    return pd.DataFrame(rows)


def primary_manual_category(row) -> str:
    if row.forecastabilityStatus == TRUE_BLOCKED_STATUS:
        return "true_forecastability_blocker"
    if row.formalReadinessStatus != READY_STATUS:
        return "formal_readiness_blocker_only"
    if row.businessActionStatus in {MANUAL_CONFIRMATION_STATUS, ACTION_BLOCKED_STATUS, ACTION_OBSERVE_STATUS}:
        return "business_action_blocker_only"
    return "forecastable_no_manual_blocker_after_disentanglement"


def build_manual_review_disentanglement(current_gate: pd.DataFrame) -> dict:
    old_manual = current_gate[current_gate["oldForecastabilityStatus"] == "manual_review_required"].copy()
    old_manual["disentangledCategory"] = old_manual.apply(primary_manual_category, axis=1)
    total_revenue = safe_float(current_gate["totalHistoricalRevenue"].sum())
    old_manual_revenue = safe_float(old_manual["totalHistoricalRevenue"].sum())
    reason_rows = []
    for reason in sorted({reason for reasons in old_manual["oldReasonCodes"] for reason in reasons}):
        group = old_manual[old_manual["oldReasonCodes"].map(lambda items: reason in items)]
        reason_rows.append(
            {
                "reasonCode": reason,
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShareTotal": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                "revenueShareWithinManualReview": percent(group["totalHistoricalRevenue"].sum(), old_manual_revenue),
                "top1OverlapCount": int((group["materialityBucket"] == "top_1_percent").sum()),
                "top5OverlapCount": int(group["materialityBucket"].isin(["top_1_percent", "top_5_percent"]).sum()),
                "top10OverlapCount": int(group["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"]).sum()),
                "ratingDistribution": distribution(group["rating"]),
                "lifecycleDistribution": distribution(group["lifecycle"]),
                "recentRevenueDistribution": distribution(group["revenueBucket"]),
            }
        )

    category_rows = []
    for category, group in old_manual.groupby("disentangledCategory"):
        category_rows.append(
            {
                "category": category,
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShareTotal": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                "revenueShareWithinManualReview": percent(group["totalHistoricalRevenue"].sum(), old_manual_revenue),
            }
        )

    high_revenue_rows = []
    for bucket in ["top_1_percent", "top_5_percent", "top_10_percent"]:
        if bucket == "top_1_percent":
            subset = old_manual[old_manual["materialityBucket"] == "top_1_percent"]
        elif bucket == "top_5_percent":
            subset = old_manual[old_manual["materialityBucket"].isin(["top_1_percent", "top_5_percent"])]
        else:
            subset = old_manual[old_manual["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"])]
        for category, group in subset.groupby("disentangledCategory"):
            high_revenue_rows.append(
                {
                    "bucket": bucket,
                    "category": category,
                    "count": int(len(group)),
                    "revenueShareWithinBucketManualReview": percent(group["totalHistoricalRevenue"].sum(), subset["totalHistoricalRevenue"].sum()),
                    "revenueShareTotal": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
                }
            )

    true_forecast_share = next((row["revenueShareWithinManualReview"] for row in category_rows if row["category"] == "true_forecastability_blocker"), 0.0)
    over_blocked = true_forecast_share < 0.5
    return {
        "schema": "m2.manual_review_forecastability_disentanglement.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "oldManualReview": {
            "count": int(len(old_manual)),
            "revenueTotal": rounded(old_manual_revenue, 2),
            "revenueShare": percent(old_manual_revenue, total_revenue),
        },
        "reasonCodeDistribution": reason_rows,
        "disentangledCategoryDistribution": sorted(category_rows, key=lambda item: item["category"]),
        "highRevenueManualReviewBreakdown": high_revenue_rows,
        "previousGateOverBlockedForecastability": over_blocked,
        "interpretation": (
            "previous gate over-blocked forecastability"
            if over_blocked
            else "previous gate captured a material true forecastability blocker share"
        ),
        "safeOutputBoundary": safe_boundary(),
    }


def status_summary(frame: pd.DataFrame, field: str, keys: list[str]) -> list[dict]:
    total_revenue = safe_float(frame["totalHistoricalRevenue"].sum())
    rows = []
    for key in keys:
        group = frame[frame[field] == key]
        rows.append(
            {
                "status": key,
                "count": int(len(group)),
                "revenueTotal": rounded(group["totalHistoricalRevenue"].sum(), 2),
                "revenueShare": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
            }
        )
    return rows


def top_bucket_coverage(frame: pd.DataFrame, max_rank: float) -> dict:
    subset = frame[frame["materialityRankPercent"] <= max_rank]
    forecastable = subset[subset["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]
    return {
        "workCount": int(len(subset)),
        "forecastableCount": int(len(forecastable)),
        "forecastableCountShare": percent(len(forecastable), len(subset)),
        "forecastableRevenueShareWithinBucket": percent(forecastable["totalHistoricalRevenue"].sum(), subset["totalHistoricalRevenue"].sum()),
    }


def build_coverage_report(current_gate: pd.DataFrame) -> dict:
    total_revenue = safe_float(current_gate["totalHistoricalRevenue"].sum())
    old_summary = status_summary(
        current_gate,
        "oldForecastabilityStatus",
        [
            "numeric_forecast_eligible",
            "conservative_numeric_forecast",
            "observe_only_no_numeric_forecast",
            "manual_review_required",
            "excluded_from_forecast_baseline",
        ],
    )
    forecast_summary = status_summary(current_gate, "forecastabilityStatus", FORECASTABILITY_STATUSES)
    formal_summary = status_summary(current_gate, "formalReadinessStatus", FORMAL_STATUSES)
    business_summary = status_summary(current_gate, "businessActionStatus", BUSINESS_ACTION_STATUSES)
    forecastable = current_gate[current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]
    true_blocked = current_gate[current_gate["forecastabilityStatus"] == TRUE_BLOCKED_STATUS]
    old_manual = current_gate[current_gate["oldForecastabilityStatus"] == "manual_review_required"].copy()
    old_manual["disentangledCategory"] = old_manual.apply(primary_manual_category, axis=1)
    manual_reclassified = []
    for category, group in old_manual.groupby("disentangledCategory"):
        manual_reclassified.append(
            {
                "category": category,
                "count": int(len(group)),
                "revenueShareWithinOldManual": percent(group["totalHistoricalRevenue"].sum(), old_manual["totalHistoricalRevenue"].sum()),
                "revenueShareTotal": percent(group["totalHistoricalRevenue"].sum(), total_revenue),
            }
        )
    return {
        "schema": "m2.disentangled_forecastability_coverage.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "oldGateSummary": old_summary,
        "disentangledForecastabilitySummary": forecast_summary,
        "formalReadinessSummary": formal_summary,
        "businessActionSummary": business_summary,
        "forecastableNumericIncludingConservative": {
            "count": int(len(forecastable)),
            "revenueShare": percent(forecastable["totalHistoricalRevenue"].sum(), total_revenue),
        },
        "numericForecastEligible": {
            "count": int((current_gate["forecastabilityStatus"] == NUMERIC_STATUS).sum()),
            "revenueShare": percent(
                current_gate.loc[current_gate["forecastabilityStatus"] == NUMERIC_STATUS, "totalHistoricalRevenue"].sum(),
                total_revenue,
            ),
        },
        "conservativeForecast": {
            "count": int((current_gate["forecastabilityStatus"] == CONSERVATIVE_STATUS).sum()),
            "revenueShare": percent(
                current_gate.loc[current_gate["forecastabilityStatus"] == CONSERVATIVE_STATUS, "totalHistoricalRevenue"].sum(),
                total_revenue,
            ),
        },
        "trueForecastBlocked": {
            "count": int(len(true_blocked)),
            "revenueShare": percent(true_blocked["totalHistoricalRevenue"].sum(), total_revenue),
        },
        "topRevenueCoverage": {
            "top1Percent": top_bucket_coverage(current_gate, 0.01),
            "top5Percent": top_bucket_coverage(current_gate, 0.05),
            "top10Percent": top_bucket_coverage(current_gate, 0.10),
        },
        "oldManualReviewReclassified": manual_reclassified,
        "safeOutputBoundary": safe_boundary(),
    }


def build_disentangled_cases(cases: pd.DataFrame, current_gate: pd.DataFrame) -> pd.DataFrame:
    selected = cases[cases["modelId"] == "model_e_selector"].copy()
    model_d_columns = [
        "workKey",
        "cutoffMonth",
        "horizonMonths",
        "predicted",
        "intervalCoverage",
        "confidence",
        "optimisticPessimisticRatio",
        "selectedModel",
        "selectionReason",
    ]
    model_d = cases[cases["modelId"] == "model_d_hierarchical_shrinkage"][model_d_columns].copy()
    model_d = model_d.rename(
        columns={
            "predicted": "predictedModelD",
            "intervalCoverage": "intervalCoverageModelD",
            "confidence": "confidenceModelD",
            "optimisticPessimisticRatio": "optimisticPessimisticRatioModelD",
            "selectedModel": "selectedModelD",
            "selectionReason": "selectionReasonModelD",
        }
    )
    gate_columns = [
        "workKey",
        "forecastabilityStatus",
        "formalReadinessStatus",
        "businessActionStatus",
        "materialityBucket",
        "requiredForecastabilityAction",
    ]
    gate = current_gate[gate_columns].copy()
    frame = selected.merge(model_d, on=["workKey", "cutoffMonth", "horizonMonths"], how="left")
    frame = frame.merge(gate, on="workKey", how="left")
    frame["forecastabilityStatus"] = frame["forecastabilityStatus"].fillna(TRUE_BLOCKED_STATUS)

    conservative = frame["forecastabilityStatus"] == CONSERVATIVE_STATUS
    can_forecast = frame["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])

    model_d_available = frame["predictedModelD"].notna()
    use_model_d = conservative & model_d_available
    predicted_source = frame["predictedModelD"].where(use_model_d, frame["predicted"]).astype(float)
    predicted = predicted_source.where(can_forecast, 0.0)
    actual = frame["actual"].astype(float)
    absolute_error = (predicted - actual).abs()
    baseline_error = frame["baselineAbsoluteError"].astype(float)

    confidence_source = frame["confidenceModelD"].where(use_model_d, frame["confidence"]).astype(str)
    confidence_source = confidence_source.where(~(conservative & (confidence_source == "high")), "medium")
    confidence = confidence_source.where(can_forecast, "blocked_for_business_use")

    selected_model = frame["selectedModelD"].where(use_model_d, frame["selectedModel"]).astype(str)
    selected_model = selected_model.where(can_forecast, "no_business_numeric_forecast")
    selection_reason = frame["selectionReasonModelD"].where(use_model_d, frame["selectionReason"]).astype(str)
    selection_reason = selection_reason.where(
        can_forecast,
        frame["requiredForecastabilityAction"].fillna("no_business_numeric_forecast").astype(str),
    )
    interval_coverage = frame["intervalCoverageModelD"].where(use_model_d, frame["intervalCoverage"]).fillna(False)
    interval_coverage = interval_coverage.where(can_forecast, False).astype(bool)
    spread = frame["optimisticPessimisticRatioModelD"].where(use_model_d, frame["optimisticPessimisticRatio"])

    smape_denominator = predicted.abs() + actual.abs()
    smape = (2 * absolute_error / smape_denominator).where(smape_denominator > 0, None)
    ape = (absolute_error / actual).where(actual > 0, None)

    return pd.DataFrame(
        {
            "workKey": frame["workKey"].astype(str),
            "modelId": MODEL_ID,
            "cutoffMonth": frame["cutoffMonth"],
            "horizonMonths": frame["horizonMonths"].astype(int),
            "predicted": predicted,
            "actual": actual,
            "absoluteError": absolute_error,
            "baselinePredicted": frame["baselinePredicted"].astype(float),
            "baselineAbsoluteError": baseline_error,
            "betterThanBaseline": absolute_error <= baseline_error,
            "smape": smape,
            "ape": ape,
            "intervalCoverage": interval_coverage,
            "overForecast": predicted > actual,
            "underForecast": predicted < actual,
            "confidence": confidence,
            "optimisticPessimisticRatio": spread,
            "ratingAtCutoff": frame["ratingAtCutoff"],
            "lifecycleAtCutoff": frame["lifecycleAtCutoff"],
            "revenueScaleAtCutoff": frame["revenueScaleAtCutoff"],
            "selectedModel": selected_model,
            "selectionReason": selection_reason,
            "forecastabilityStatus": frame["forecastabilityStatus"],
            "formalReadinessStatus": frame["formalReadinessStatus"],
            "businessActionStatus": frame["businessActionStatus"],
            "materialityBucket": frame["materialityBucket"],
        }
    )


def labels_for(cases: pd.DataFrame) -> pd.Series:
    if cases.empty:
        return pd.Series(dtype=str)
    actual = pd.to_numeric(cases["actual"], errors="coerce").fillna(0)
    predicted = pd.to_numeric(cases["predicted"], errors="coerce").fillna(0)
    error = pd.to_numeric(cases["absoluteError"], errors="coerce").fillna(0)
    ape = pd.to_numeric(cases["ape"], errors="coerce")
    smape_value = pd.to_numeric(cases["smape"], errors="coerce")
    interval_coverage = cases["intervalCoverage"].astype(bool)
    labels = pd.Series("pass", index=cases.index)
    labels = labels.mask(~interval_coverage, "warning")
    labels = labels.mask(ape.gt(1.25), "warning")
    labels = labels.mask(((ape.gt(3.0)) & error.gt(100)) | ((smape_value.gt(1.35)) & error.gt(50)), "fail")
    labels = labels.mask((actual.le(1)) & predicted.gt(20), "fail")
    labels = labels.mask((actual.le(1)) & predicted.le(3), "pass")
    return labels


def score_cases(cases: pd.DataFrame) -> dict:
    metrics = bake.aggregate_cases(cases)
    labels = labels_for(cases)
    actual = safe_float(cases["actual"].sum()) if not cases.empty else 0.0
    baseline_abs = safe_float(cases["baselineAbsoluteError"].sum()) if not cases.empty else 0.0
    return {
        **metrics,
        "baselineWape": None if actual <= 0 else rounded(baseline_abs / actual),
        "passCount": int((labels == "pass").sum()),
        "warningCount": int((labels == "warning").sum()),
        "failCount": int((labels == "fail").sum()),
        "failRate": percent((labels == "fail").sum(), len(cases)) if len(cases) else None,
        "warningRate": percent((labels == "warning").sum(), len(cases)) if len(cases) else None,
    }


def build_validation_report(model_cases: pd.DataFrame, current_gate: pd.DataFrame, coverage: dict) -> dict:
    forecastable_cases = model_cases[model_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]
    numeric_cases = model_cases[model_cases["forecastabilityStatus"] == NUMERIC_STATUS]
    conservative_cases = model_cases[model_cases["forecastabilityStatus"] == CONSERVATIVE_STATUS]
    high_value_cases = model_cases[model_cases["materialityBucket"].isin(["top_1_percent", "top_5_percent", "top_10_percent"])]
    forecastable_score = score_cases(forecastable_cases)
    high_value_score = score_cases(high_value_cases[high_value_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])])

    total_revenue = safe_float(current_gate["totalHistoricalRevenue"].sum())
    p0_rows = current_gate[
        (current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]))
        & (current_gate["rating"].isin(["S+", "S", "A"]))
        & (current_gate["baseForecast"].fillna(0).astype(float) <= 0)
    ]
    p1_rows = current_gate[
        (~current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]))
        & (current_gate["baseForecast"].notna())
    ]
    p2_rows = current_gate[
        (current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]))
        & (current_gate["businessActionStatus"] == MANUAL_CONFIRMATION_STATUS)
    ]

    segment_rows = []
    for field in ["forecastabilityStatus", "formalReadinessStatus", "businessActionStatus", "materialityBucket", "ratingAtCutoff", "lifecycleAtCutoff", "revenueScaleAtCutoff"]:
        for segment, group in model_cases.groupby(field):
            if field in {"ratingAtCutoff", "lifecycleAtCutoff", "revenueScaleAtCutoff"}:
                group = group[group["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])]
            metrics = score_cases(group)
            segment_rows.append(
                {
                    "segmentType": field,
                    "segment": str(segment),
                    "caseCount": metrics["caseCount"],
                    "wape": metrics["wape"],
                    "smape": metrics["smape"],
                    "mae": metrics["mae"],
                    "coverage": metrics["intervalCoverage"],
                    "failRate": metrics["failRate"],
                }
            )

    forecastable_revenue_share = coverage["forecastableNumericIncludingConservative"]["revenueShare"]
    true_blocked_share = coverage["trueForecastBlocked"]["revenueShare"]
    top5_count_share = coverage["topRevenueCoverage"]["top5Percent"]["forecastableCountShare"]
    interval_coverage = safe_float(forecastable_score["intervalCoverage"], 0)
    wape_not_worse = (
        forecastable_score["wape"] is not None
        and forecastable_score["baselineWape"] is not None
        and safe_float(forecastable_score["wape"]) <= safe_float(forecastable_score["baselineWape"])
    )
    pass_conditions = {
        "forecastableRevenueCoverageAtLeast70": forecastable_revenue_share >= 0.70,
        "top5ForecastableCountShareAtLeast80": top5_count_share >= 0.80,
        "wapeNotWorseThanBaseline": wape_not_worse,
        "highRevenueP0EqualsZero": int(len(p0_rows)) == 0,
        "p1AtMost3": int(len(p1_rows)) <= 3,
        "intervalCoverageAtLeast50": interval_coverage >= 0.50,
        "trueForecastBlockedRevenueShareAtMost20": true_blocked_share <= 0.20,
        "formalBlockersSeparated": True,
        "businessActionBlockersSeparated": True,
    }
    conditional_conditions = {
        "forecastableRevenueCoverageAtLeast50": forecastable_revenue_share >= 0.50,
        "top5ForecastableCountShareAtLeast70": top5_count_share >= 0.70,
        "wapeNotWorseThanBaseline": wape_not_worse,
        "p0EqualsZero": int(len(p0_rows)) == 0,
        "p1AtMost10": int(len(p1_rows)) <= 10,
        "trueForecastBlockedRevenueShareAtMost35": true_blocked_share <= 0.35,
    }
    if all(pass_conditions.values()):
        verdict = "PASS"
        candidate = DISENTANGLED_CANDIDATE
    elif all(conditional_conditions.values()):
        verdict = "CONDITIONAL PASS"
        candidate = f"{DISENTANGLED_CANDIDATE}-conditional"
    else:
        verdict = "FAIL"
        candidate = None

    return {
        "schema": "m2.disentangled_forecast_model_validation.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "modelId": MODEL_ID,
        "candidateVersion": candidate,
        "verdict": verdict,
        "forecastableCohortScore": forecastable_score,
        "numericCohortScore": score_cases(numeric_cases),
        "conservativeCohortScore": score_cases(conservative_cases),
        "highValueForecastableScore": high_value_score,
        "issueSummary": {
            "p0": int(len(p0_rows)),
            "p1": int(len(p1_rows)),
            "p2": int(len(p2_rows)),
            "p0RevenueShare": percent(p0_rows["totalHistoricalRevenue"].sum(), total_revenue),
            "p1RevenueShare": percent(p1_rows["totalHistoricalRevenue"].sum(), total_revenue),
            "p2RevenueShare": percent(p2_rows["totalHistoricalRevenue"].sum(), total_revenue),
        },
        "passWarningFail": distribution(labels_for(forecastable_cases), ["pass", "warning", "fail"]),
        "segmentPerformance": sorted(segment_rows, key=lambda item: (item["segmentType"], item["segment"])),
        "passConditions": pass_conditions,
        "conditionalPassConditions": conditional_conditions,
        "safeOutputBoundary": safe_boundary(),
    }


def build_business_readiness(coverage: dict, validation: dict) -> dict:
    verdict = validation["verdict"]
    if verdict == "PASS":
        readiness = "m2_business_review_ready_for_forecastable_cohort"
    elif verdict == "CONDITIONAL PASS":
        readiness = "limited_m2_business_review_ready_for_forecastable_cohort"
    else:
        readiness = "not_ready_for_m2_business_review"
    groups = [
        {
            "group": "numeric_forecast_eligible",
            "businessUse": "can_enter_business_review_if_validation_passes",
            "reason": "forecastability is based only on local revenue time-series evidence",
        },
        {
            "group": "conservative_numeric_forecast",
            "businessUse": "reference_forecast_only",
            "reason": "forecastable but lower confidence or tail/volatile pattern",
        },
        {
            "group": "formal_release_blocked_but_forecastable",
            "businessUse": "can_validate_algorithm_locally_but_cannot_release_formally",
            "reason": "formal readiness blockers are separated from forecastability",
        },
        {
            "group": "business_action_blocked_but_forecastable",
            "businessUse": "forecast_can_be_reviewed_but_action_requires_confirmation",
            "reason": "promote/downlist/renewal controls block action, not forecast",
        },
        {
            "group": "true_forecast_blocked",
            "businessUse": "no_numeric_forecast",
            "reason": "insufficient series, severe unresolved spike, or not backtestable",
        },
    ]
    return {
        "schema": "m2.disentangled_algorithm_business_readiness.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "readiness": readiness,
        "m3Allowed": False,
        "candidateVersion": validation["candidateVersion"],
        "modelVerdict": verdict,
        "forecastableRevenueShare": coverage["forecastableNumericIncludingConservative"]["revenueShare"],
        "trueForecastBlockedRevenueShare": coverage["trueForecastBlocked"]["revenueShare"],
        "businessGroups": groups,
        "m3BlockReason": "M3 remains blocked unless user explicitly authorizes parallel planning after forecast baseline acceptance.",
        "safeOutputBoundary": safe_boundary(),
    }


def write_reports(disentangle: dict, coverage: dict, validation: dict, readiness: dict) -> None:
    write_json(DISENTANGLE_JSON, disentangle)
    write_json(COVERAGE_JSON, coverage)
    write_json(VALIDATION_JSON, validation)
    write_json(READINESS_JSON, readiness)

    DISENTANGLE_MD.write_text(
        "\n".join(
            [
                "# M2 Manual Review Forecastability Disentanglement v1",
                "",
                f"Old manual-review count: `{disentangle['oldManualReview']['count']}`.",
                f"Old manual-review revenue share: `{disentangle['oldManualReview']['revenueShare']}`.",
                f"Conclusion: `{disentangle['interpretation']}`.",
                "",
                "## Reason Code Distribution",
                "",
                markdown_table(
                    disentangle["reasonCodeDistribution"],
                    [
                        ("reasonCode", "Reason"),
                        ("count", "Count"),
                        ("revenueShareTotal", "Revenue Share"),
                        ("top1OverlapCount", "Top 1 Count"),
                        ("top5OverlapCount", "Top 5 Count"),
                        ("top10OverlapCount", "Top 10 Count"),
                    ],
                ),
                "",
                "## Disentangled Categories",
                "",
                markdown_table(
                    disentangle["disentangledCategoryDistribution"],
                    [("category", "Category"), ("count", "Count"), ("revenueShareTotal", "Revenue Share"), ("revenueShareWithinManualReview", "Share Within Old Manual Review")],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    COVERAGE_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecastability Coverage v1",
                "",
                "## Forecastability Summary",
                "",
                markdown_table(
                    coverage["disentangledForecastabilitySummary"],
                    [("status", "Status"), ("count", "Count"), ("revenueShare", "Revenue Share")],
                ),
                "",
                "## Top Revenue Coverage",
                "",
                markdown_table(
                    [{"bucket": key, **value} for key, value in coverage["topRevenueCoverage"].items()],
                    [("bucket", "Bucket"), ("workCount", "Works"), ("forecastableCount", "Forecastable"), ("forecastableCountShare", "Forecastable Count Share"), ("forecastableRevenueShareWithinBucket", "Forecastable Revenue Share")],
                ),
                "",
                "This report is sanitized and aggregate-only.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    VALIDATION_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Forecast Model Validation v1",
                "",
                f"Verdict: `{validation['verdict']}`",
                f"Candidate version: `{validation['candidateVersion']}`",
                "",
                "## Forecastable Cohort Score",
                "",
                markdown_table(
                    [validation["forecastableCohortScore"]],
                    [("wape", "WAPE"), ("baselineWape", "Baseline WAPE"), ("smape", "SMAPE"), ("mae", "MAE"), ("intervalCoverage", "Coverage"), ("failRate", "Fail Rate")],
                ),
                "",
                "## Issues",
                "",
                markdown_table([validation["issueSummary"]], [("p0", "P0"), ("p1", "P1"), ("p2", "P2"), ("p2RevenueShare", "P2 Revenue Share")]),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    READINESS_MD.write_text(
        "\n".join(
            [
                "# M2 Disentangled Algorithm Business Readiness v1",
                "",
                f"Readiness: `{readiness['readiness']}`",
                f"M3 allowed: `{readiness['m3Allowed']}`",
                "",
                "## Business Groups",
                "",
                markdown_table(readiness["businessGroups"], [("group", "Group"), ("businessUse", "Business Use"), ("reason", "Reason")]),
                "",
                "This report is sanitized and aggregate-only. It is not final production release approval.",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_excel(disentangle: dict, coverage: dict, validation: dict, readiness: dict, current_gate: pd.DataFrame, model_cases: pd.DataFrame) -> None:
    if Alignment is None:
        raise RuntimeError("openpyxl is required to write the private Excel workbook.")
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    high_value = current_gate[current_gate["materialityRankPercent"] <= 0.10].copy().head(500)
    true_blocked = current_gate[current_gate["forecastabilityStatus"] == TRUE_BLOCKED_STATUS].copy().head(2000)
    formal_blocked_forecastable = current_gate[
        (current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]))
        & (current_gate["formalReadinessStatus"] != READY_STATUS)
    ].copy().head(2000)
    action_blocked_forecastable = current_gate[
        (current_gate["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS]))
        & (current_gate["businessActionStatus"].isin([MANUAL_CONFIRMATION_STATUS, ACTION_BLOCKED_STATUS]))
    ].copy().head(2000)
    forecastable_cases = model_cases[model_cases["forecastabilityStatus"].isin([NUMERIC_STATUS, CONSERVATIVE_STATUS])].copy().head(30000)
    forecastable_cases["anonymousCaseId"] = [f"D{index + 1:05d}" for index in range(len(forecastable_cases))]

    with pd.ExcelWriter(PRIVATE_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"item": "purpose", "value": "M2 disentangled forecastability validation"},
                {"item": "forecastability", "value": "revenue time-series predictability only"},
                {"item": "formal readiness", "value": "release blockers separated from local forecast validation"},
                {"item": "business action", "value": "promote/downlist/renewal confirmations separated from forecast validation"},
                {"item": "safety", "value": "anonymous IDs only; no work names, authors, channels, or raw rows"},
            ]
        ).to_excel(writer, sheet_name="00_read_me", index=False)
        pd.DataFrame(disentangle["reasonCodeDistribution"]).to_excel(writer, sheet_name="01_manual_review_disentangle", index=False)
        pd.DataFrame(
            coverage["oldGateSummary"]
            + [{"status": f"new::{row['status']}", "count": row["count"], "revenueShare": row["revenueShare"]} for row in coverage["disentangledForecastabilitySummary"]]
        ).to_excel(writer, sheet_name="02_gate_comparison", index=False)
        pd.DataFrame(
            [
                {"metric": "numeric_plus_conservative_revenue_share", "value": coverage["forecastableNumericIncludingConservative"]["revenueShare"]},
                {"metric": "numeric_revenue_share", "value": coverage["numericForecastEligible"]["revenueShare"]},
                {"metric": "conservative_revenue_share", "value": coverage["conservativeForecast"]["revenueShare"]},
                {"metric": "true_forecast_blocked_revenue_share", "value": coverage["trueForecastBlocked"]["revenueShare"]},
                {"metric": "top1_forecastable_count_share", "value": coverage["topRevenueCoverage"]["top1Percent"]["forecastableCountShare"]},
                {"metric": "top5_forecastable_count_share", "value": coverage["topRevenueCoverage"]["top5Percent"]["forecastableCountShare"]},
                {"metric": "top10_forecastable_count_share", "value": coverage["topRevenueCoverage"]["top10Percent"]["forecastableCountShare"]},
            ]
        ).to_excel(writer, sheet_name="03_forecastable_coverage", index=False)
        pd.DataFrame([validation["forecastableCohortScore"], validation["numericCohortScore"], validation["conservativeCohortScore"]]).to_excel(writer, sheet_name="04_model_performance", index=False)
        high_value[
            [
                "anonymousId",
                "materialityBucket",
                "forecastabilityStatus",
                "formalReadinessStatus",
                "businessActionStatus",
                "selectedModel",
                "forecastConfidence",
            ]
        ].to_excel(writer, sheet_name="05_high_value_work_check", index=False)
        true_blocked[
            ["anonymousId", "materialityBucket", "forecastabilityStatus", "forecastabilityReasonCodes", "requiredForecastabilityAction"]
        ].to_excel(writer, sheet_name="06_true_forecast_blocked", index=False)
        formal_blocked_forecastable[
            ["anonymousId", "forecastabilityStatus", "formalReadinessStatus", "formalReadinessReasonCodes", "requiredFormalAction"]
        ].to_excel(writer, sheet_name="07_formal_blocked_forecastable", index=False)
        action_blocked_forecastable[
            ["anonymousId", "forecastabilityStatus", "businessActionStatus", "businessActionReasonCodes", "requiredBusinessAction"]
        ].to_excel(writer, sheet_name="08_action_blocked_forecastable", index=False)
        pd.DataFrame(readiness["businessGroups"]).to_excel(writer, sheet_name="09_business_readiness", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, column in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in column:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(text), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(column_index)].width = max(10, min(max_len + 3, 45))
            worksheet.auto_filter.ref = worksheet.dimensions


def run() -> dict:
    prepared = prepare_inputs()
    evaluated = prepared["evaluated"]
    cases = prepared["cases"]
    final_outputs = prepared["finalOutputs"]
    current_gate = build_current_gate_frame(evaluated, final_outputs)
    disentangle = build_manual_review_disentanglement(current_gate)
    coverage = build_coverage_report(current_gate)
    model_cases = build_disentangled_cases(cases, current_gate)
    validation = build_validation_report(model_cases, current_gate, coverage)
    readiness = build_business_readiness(coverage, validation)
    write_reports(disentangle, coverage, validation, readiness)
    write_excel(disentangle, coverage, validation, readiness, current_gate, model_cases)
    write_json(
        PRIVATE_DETAIL_JSON,
        {
            "schema": "m2.private.disentangled_forecastability_detail.v1",
            "notForCommit": True,
            "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
            "verdict": validation["verdict"],
            "candidateVersion": validation["candidateVersion"],
        },
    )
    return {
        "threeGateDisentanglementCompleted": True,
        "newCandidateVersion": validation["candidateVersion"],
        "verdict": validation["verdict"],
        "m2BusinessReviewReady": readiness["readiness"],
        "m3Allowed": readiness["m3Allowed"],
        "oldManualReview": disentangle["oldManualReview"],
        "previousGateOverBlockedForecastability": disentangle["previousGateOverBlockedForecastability"],
        "forecastableNumericIncludingConservative": coverage["forecastableNumericIncludingConservative"],
        "trueForecastBlocked": coverage["trueForecastBlocked"],
        "topRevenueCoverage": coverage["topRevenueCoverage"],
        "forecastableCohortScore": validation["forecastableCohortScore"],
        "issueSummary": validation["issueSummary"],
        "privateWorkbook": str(PRIVATE_XLSX.relative_to(ROOT)),
        "sanitizedReports": [
            str(DISENTANGLE_JSON.relative_to(ROOT)),
            str(COVERAGE_JSON.relative_to(ROOT)),
            str(VALIDATION_JSON.relative_to(ROOT)),
            str(READINESS_JSON.relative_to(ROOT)),
        ],
    }


if __name__ == "__main__":
    print(json.dumps(bake.json_safe(run()), ensure_ascii=False, indent=2))
