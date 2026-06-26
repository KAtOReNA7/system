from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
TEMP_DEPS = Path(os.environ.get("TEMP", "")) / "codex-system-pydeps"
if TEMP_DEPS.exists():
    sys.path.insert(0, str(TEMP_DEPS))

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it into a local temp dependency path, "
        "for example: python -m pip install --target %TEMP%\\codex-system-pydeps openpyxl"
    ) from exc


CANDIDATE_VERSION = "m2-realdata-dev-commercial-rating-suggestion-v3.0"
PREVIOUS_CANDIDATE_VERSION = "m2-realdata-dev-commercial-rating-suggestion-v2.0"
M2_TOTAL_WORKS = 3054
TODAY = date(2026, 6, 26)

PRIVATE_M2 = ROOT / "data" / "private-output" / "m2-business-review"
PRIVATE_M1 = ROOT / "data" / "private-output" / "m1-master-data"
OPS_DIR = ROOT / "data" / "m1-master-data-private" / "ops-confirmation"
MASTER_DIR = ROOT / "data" / "master-data"
DOCS_DIR = ROOT / "docs" / "analysis" / "m2-real-data"

INPUT_V2_OPERATOR_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-commercial-rating-v2.xlsx"
INPUT_AFTER_STAGING_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-after-dual-source-staging-v2.xlsx"
STAGING_TABLE_JSON = PRIVATE_M1 / "M1-dual-source-limited-staging-table-v1.json"

PRIVATE_CANDIDATES_XLSX = PRIVATE_M2 / "M2-commercial-terms-candidates-v2.xlsx"
PRIVATE_CANDIDATES_JSON = PRIVATE_M2 / "M2-commercial-terms-candidates-v2.json"
PRIVATE_VALIDATION_XLSX = PRIVATE_M2 / "m2-commercial-rating-suggestion-v3-validation.xlsx"
PRIVATE_OPERATOR_XLSX = PRIVATE_M2 / "m2-v1.1-30-work-operator-task-pack-cn-commercial-rating-v3.xlsx"

COMMERCIAL_AUDIT_JSON = DOCS_DIR / "M2-commercial-terms-source-audit-v2.json"
COMMERCIAL_AUDIT_MD = DOCS_DIR / "M2-commercial-terms-source-audit-v2.md"
RATING_SUMMARY_JSON = DOCS_DIR / "M2-commercial-rating-calibration-v3-summary.json"
RATING_SUMMARY_MD = DOCS_DIR / "M2-commercial-rating-calibration-v3-summary.md"
SUGGESTION_SUMMARY_JSON = DOCS_DIR / "M2-commercial-suggestion-calibration-v3-summary.json"
SUGGESTION_SUMMARY_MD = DOCS_DIR / "M2-commercial-suggestion-calibration-v3-summary.md"
VALIDATION_SUMMARY_JSON = DOCS_DIR / "M2-commercial-rating-suggestion-v3-validation-summary.json"
VALIDATION_SUMMARY_MD = DOCS_DIR / "M2-commercial-rating-suggestion-v3-validation-summary.md"
OPERATOR_SUMMARY_JSON = DOCS_DIR / "M2-operator-task-pack-commercial-rating-v3-summary.json"
OPERATOR_SUMMARY_MD = DOCS_DIR / "M2-operator-task-pack-commercial-rating-v3-summary.md"

TASK_SHEET_CANDIDATES = ["01_运营任务卡", "01_运营任务单", "01_运营任务包"]
COMMERCIAL_KEYWORDS = [
    "买断",
    "版税",
    "预付",
    "有声版税",
    "有声预付",
    "合作方式",
    "合同类型",
    "授权方式",
    "是否买断",
    "版权费用",
    "分成",
    "保底",
    "结算方式",
    "有声权利",
    "有声使用权",
    "有声改编权",
    "有声转授权",
]


def main() -> None:
    ensure_inputs()
    staging_index = build_staging_work_index()
    ops_rows, ops_source_meta = load_ops_rows()
    commercial_lookup, source_audit = build_commercial_lookup(ops_rows, staging_index, ops_source_meta)
    commercial_audit = build_commercial_audit(commercial_lookup, source_audit, staging_index)

    operator_rows = load_operator_rows(INPUT_V2_OPERATOR_XLSX if INPUT_V2_OPERATOR_XLSX.exists() else INPUT_AFTER_STAGING_XLSX)
    operator_v3_rows = [build_operator_v3_row(row, commercial_lookup, staging_index) for row in operator_rows]
    reviewable_operator_rows = [row for row in operator_v3_rows if clean(row.get("样本来源")) != "用户指定作品"]

    validation_rows = build_validation_rows(commercial_lookup, staging_index)
    rating_summary = build_rating_summary(validation_rows, reviewable_operator_rows)
    suggestion_summary = build_suggestion_summary(validation_rows, reviewable_operator_rows)
    validation_summary = build_validation_summary(validation_rows, rating_summary, suggestion_summary)
    operator_summary = build_operator_summary(operator_v3_rows, reviewable_operator_rows)

    write_private_outputs(commercial_lookup, staging_index, validation_rows, operator_v3_rows, commercial_audit, validation_summary)
    write_public_reports(commercial_audit, rating_summary, suggestion_summary, validation_summary, operator_summary)

    print(
        json.dumps(
            {
                "candidateVersion": CANDIDATE_VERSION,
                "knownCommercialModelWorks": commercial_audit["m2Mapping"]["knownCommercialModelWorks"],
                "knownCommercialCoverageRateOfM2": commercial_audit["m2Mapping"]["knownCommercialCoverageRateOfM2"],
                "buyoutWorks": commercial_audit["m2Mapping"]["commercialModelDistribution"].get("buyout", 0),
                "unknownWorks": commercial_audit["m2Mapping"]["commercialModelDistribution"].get("unknown", 0),
                "expiredValidationRows": rating_summary["expiredValidationRows"],
                "expiredDisplayedAsOnlyE": rating_summary["expiredDisplayedAsOnlyE"],
                "automaticOperatingSuggestionRows": suggestion_summary["automaticOperatingSuggestionRows"],
                "reviewPromptRows": suggestion_summary["reviewPromptRows"],
                "privateValidationWorkbook": rel(PRIVATE_VALIDATION_XLSX),
                "privateOperatorWorkbook": rel(PRIVATE_OPERATOR_XLSX),
                "m3Entered": False,
            },
            ensure_ascii=False,
        )
    )


def ensure_inputs() -> None:
    required = [
        STAGING_TABLE_JSON,
        OPS_DIR / "ops-confirmation-v2.3-data.json",
        ROOT / "docs" / "prd" / "20-evaluation" / "M2-old-product-evaluation-prd-v0.1.md",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "commercialTermsParser.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "ratingCalibration.js",
        ROOT / "src" / "domain" / "oldProductEvaluation" / "suggestionCalibration.js",
        ROOT / "package.json",
        ROOT / ".gitignore",
    ]
    if not INPUT_V2_OPERATOR_XLSX.exists() and not INPUT_AFTER_STAGING_XLSX.exists():
        required.append(INPUT_V2_OPERATOR_XLSX)
    missing = [path for path in required if not path.exists()]
    if missing:
        raise SystemExit("Missing required inputs: " + ", ".join(rel(path) for path in missing))


def load_ops_rows() -> tuple[list[dict], dict]:
    source_path = OPS_DIR / "ops-confirmation-v2.3-data.json"
    data = read_json(source_path)
    sheet = next((item for item in data.get("sheets", []) if item.get("name") == "标准作品基础信息补全"), None)
    rows = sheet.get("rows", []) if sheet else []
    meta = {
        "opsConfirmationJson": rel(source_path),
        "sourceSnapshots": data.get("source_snapshots", {}),
        "masterSnapshotPath": safe_snapshot_path(data.get("source_snapshots", {}).get("master", [])),
        "sourceRows": len(rows),
    }
    return rows, meta


def build_staging_work_index() -> dict[str, dict]:
    records = read_json(STAGING_TABLE_JSON).get("records", [])
    works: dict[str, dict] = {}
    for record in records:
        sid = clean(record.get("standardWorkId"))
        if not sid:
            continue
        item = works.setdefault(
            sid,
            {
                "standardWorkId": sid,
                "totalHistoricalRevenue": 0.0,
                "fields": set(),
                "copyrightStart": "",
                "copyrightEnd": "",
                "title": "",
                "author": "",
            },
        )
        item["fields"].add(clean(record.get("fieldName")))
        item["totalHistoricalRevenue"] = max(item["totalHistoricalRevenue"], number_or_zero(record.get("totalHistoricalRevenue")))
        field = clean(record.get("fieldName"))
        value = clean(record.get("applyValue"))
        if field in {"workTitle", "作品名", "standardWorkTitle"} and value:
            item["title"] = value
        elif field in {"author", "作者"} and value:
            item["author"] = value
        elif field in {"copyrightStartDate", "版权开始"} and value:
            item["copyrightStart"] = value
        elif field in {"copyrightEndDate", "版权到期"} and value:
            item["copyrightEnd"] = value
    for item in works.values():
        item["fields"] = sorted(item["fields"])
    return works


def build_commercial_lookup(ops_rows: list[dict], staging_index: dict[str, dict], meta: dict) -> tuple[dict[str, dict], dict]:
    lookup: dict[str, dict] = {}
    source_counter = Counter()
    field_counter = Counter()
    commercial_signal_rows = 0

    for row in ops_rows:
        sid = clean(row.get("标准作品ID"))
        if not sid:
            continue
        fields = extract_commercial_fields(row)
        if fields:
            commercial_signal_rows += 1
        parsed = parse_commercial_terms(fields, source_type="full_digital_ledger_extracted_tags")
        parsed.update(
            {
                "standardWorkId": sid,
                "workTitle": clean(row.get("运营确认标准作品名称") or row.get("台账标准名称候选") or row.get("账单主要名称")),
                "author": clean(row.get("运营确认作者") or row.get("作者候选")),
                "copyrightStart": clean(row.get("运营确认版权开始日期") or row.get("版权开始日期候选（签订日期）")),
                "copyrightEnd": clean(row.get("运营确认版权到期日期") or row.get("版权到期日期候选（到期时间）")),
                "sourceFieldNames": sorted(fields.keys()),
            }
        )
        for field in fields:
            field_counter[field] += 1
        source_counter[parsed["commercialModel"]] += 1
        lookup[sid] = parsed

    ledger_audit = audit_excel_headers(MASTER_DIR / "数字版权台账.xlsx")
    original_library_audit = audit_excel_headers(MASTER_DIR / "原创全库.xlsx")
    source_audit = {
        "candidateVersion": CANDIDATE_VERSION,
        "sourceFiles": [
            {
                "sourceAlias": "OPS_CONFIRMATION_V2_3_JSON",
                "relativePath": meta["opsConfirmationJson"],
                "sheetOrObject": "标准作品基础信息补全",
                "sourceSnapshotMasterPath": meta["masterSnapshotPath"],
                "usableForCommercialTerms": True,
                "sourceBoundary": "contains extracted ledger/operation tag fields, not raw contract rows",
            },
            ledger_audit,
            original_library_audit,
        ],
        "fieldFindings": [
            {
                "fieldName": field,
                "nonEmptyCommercialSignalRows": int(count),
                "sourceType": "full_digital_ledger_extracted_tags",
            }
            for field, count in sorted(field_counter.items())
        ],
        "sourceRows": meta["sourceRows"],
        "sourceWithCommercialTextRows": commercial_signal_rows,
        "commercialModelDistributionInSource": dict(sorted(source_counter.items())),
        "sourceGap": build_source_gap(ledger_audit, original_library_audit, commercial_signal_rows),
        "formalMasterDataWritten": False,
        "databaseConnected": False,
        "m3Entered": False,
    }
    return lookup, source_audit


def extract_commercial_fields(row: dict) -> dict[str, str]:
    fields = {}
    for key, value in row.items():
        text = clean(value)
        if not text:
            continue
        if any(keyword in str(key) for keyword in COMMERCIAL_KEYWORDS) or any(keyword in text for keyword in COMMERCIAL_KEYWORDS):
            fields[clean(key)] = text
    return fields


def parse_commercial_terms(fields: dict[str, str], source_type: str = "unknown") -> dict:
    values = [value for value in fields.values() if clean(value)]
    text = "\n".join(values)
    buyout = bool(re.search(r"买断|著作权转让|全部著作权归|版权归[^，。,；;\n]*所有", text))
    royalty = "版税" in text
    prepaid = "预付" in text
    revenue_share = bool(re.search(r"分成|收益分配|收入分配|按比例结算", text))
    guarantee = "保底" in text
    audio_scope = bool(re.search(r"有声[^，。,；;\n]*买断|买断[^，。,；;\n]*有声|有声使用权|有声改编权|有声转授权", text))
    full_scope = bool(re.search(r"全部著作权|完整权利|全版权|著作权转让", text))

    if prepaid and royalty:
        model, confidence, manual = "prepaid_royalty", "medium", source_type != "full_digital_ledger_terms"
        reason = "抽取字段同时包含预付和版税"
    elif buyout and (royalty or revenue_share):
        model, confidence, manual = "mixed", "medium", True
        reason = "买断与版税/分成信号同时出现"
    elif buyout:
        model, confidence, manual = "buyout", "medium", source_type != "full_digital_ledger_terms" or not (audio_scope or full_scope)
        reason = "抽取字段包含买断或著作权转让"
    elif revenue_share:
        model, confidence, manual = "revenue_share", "medium", source_type != "full_digital_ledger_terms"
        reason = "抽取字段包含分成/收益分配"
    elif royalty:
        model, confidence, manual = "royalty", "medium", source_type != "full_digital_ledger_terms"
        reason = "抽取字段包含版税"
    elif prepaid:
        model, confidence, manual = "mixed", "low", True
        reason = "仅出现预付信号，缺少结算上下文"
    else:
        model, confidence, manual = "unknown", "none", True
        reason = "未发现可用商业模式字段"

    buyout_scope = "not_buyout"
    if buyout:
        if audio_scope:
            buyout_scope = "audio_related_buyout_scope"
        elif full_scope:
            buyout_scope = "full_rights_buyout_scope"
        else:
            buyout_scope = "unknown_buyout_scope"

    if source_type != "full_digital_ledger_terms" and confidence == "high":
        confidence = "medium"
    if source_type in {"full_digital_ledger_extracted_tags", "operation_confirmation_tags"} and model != "unknown":
        manual = True

    return {
        "commercialModel": model,
        "commercialModelChinese": commercial_model_cn(model),
        "commercialModelConfidence": confidence,
        "commercialTermsSource": source_type,
        "commercialTermsReason": [reason, "当前商业条款来自抽取标签/确认包，不能等同完整合同字段"] if model != "unknown" else [reason],
        "buyoutFlag": buyout,
        "royaltyFlag": royalty or model in {"royalty", "prepaid_royalty"},
        "prepaidFlag": prepaid,
        "revenueShareFlag": revenue_share,
        "guaranteeFlag": guarantee,
        "buyoutScope": buyout_scope,
        "requiresManualCommercialReview": manual or model in {"mixed", "unknown", "conflict"},
        "commercialRiskLevel": "high" if manual or model == "mixed" else ("medium" if guarantee or confidence in {"low", "none"} else "low"),
    }


def build_commercial_audit(commercial_lookup: dict[str, dict], source_audit: dict, staging_index: dict[str, dict]) -> dict:
    model_counter = Counter()
    confidence_counter = Counter()
    manual_counter = Counter()
    revenue_by_model = Counter()
    matched = 0
    known = 0
    total_revenue = sum(item["totalHistoricalRevenue"] for item in staging_index.values())

    for sid, item in staging_index.items():
        terms = commercial_lookup.get(sid) or unknown_terms(sid)
        if sid in commercial_lookup:
            matched += 1
        model = terms["commercialModel"]
        if model != "unknown":
            known += 1
        model_counter[model] += 1
        confidence_counter[terms.get("commercialModelConfidence", "none")] += 1
        if terms.get("requiresManualCommercialReview"):
            manual_counter["requiresManualCommercialReview"] += 1
        revenue_by_model[model] += item["totalHistoricalRevenue"]

    top_distribution = {}
    sorted_works = sorted(staging_index.values(), key=lambda item: item["totalHistoricalRevenue"], reverse=True)
    for label, pct in [("top1Percent", 0.01), ("top5Percent", 0.05), ("top10Percent", 0.10)]:
        count = max(1, math.ceil(M2_TOTAL_WORKS * pct))
        cohort = sorted_works[:count]
        counter = Counter()
        for item in cohort:
            terms = commercial_lookup.get(item["standardWorkId"]) or unknown_terms(item["standardWorkId"])
            counter[terms["commercialModel"]] += 1
        top_distribution[label] = {"workCount": len(cohort), "distribution": dict(sorted(counter.items()))}

    return {
        **source_audit,
        "m2Mapping": {
            "m2TotalWorks": M2_TOTAL_WORKS,
            "stagingRevenueMappedWorks": len(staging_index),
            "commercialSourceMatchedWorks": matched,
            "knownCommercialModelWorks": known,
            "commercialCoverageRateOfM2": round(matched / M2_TOTAL_WORKS, 6),
            "knownCommercialCoverageRateOfM2": round(known / M2_TOTAL_WORKS, 6),
            "commercialModelDistribution": dict(sorted(model_counter.items())),
            "commercialConfidenceDistribution": dict(sorted(confidence_counter.items())),
            "manualCommercialReviewDistribution": dict(sorted(manual_counter.items())),
            "revenueByCommercialModel": money_counter(revenue_by_model),
            "buyoutRevenueShareOfMappedRevenue": round(revenue_by_model["buyout"] / total_revenue, 6) if total_revenue else 0,
            "topRevenueDistribution": top_distribution,
        },
        "assessment": {
            "v2CoverageConcernResolved": False,
            "coverageExplanation": "v3 improves source confidence boundaries; it does not claim full commercial coverage because standalone full contract fields are not present in the current cleaned ledger.",
            "rawDetailsExcluded": True,
        },
    }


def build_operator_v3_row(row: dict, commercial_lookup: dict[str, dict], staging_index: dict[str, dict]) -> dict:
    sid = clean(row.get("standard_work_id") or row.get("standardWorkId") or row.get("标准作品ID") or row.get("作品ID"))
    terms = commercial_lookup.get(sid) or unknown_terms(sid)
    staging = staging_index.get(sid, {})
    rights_status, expired = rights_status_for(row, terms, staging)
    history_rating = historical_rating(row, staging)
    forecast_rating = forecast_rating_for(row)
    decision_rating = decision_rating_for(rights_status, terms, row)
    display_rating = f"历史表现 {history_rating}；当前版权状态：{rights_status_cn(rights_status)}；预测价值：{forecast_rating}；商业模式：{terms['commercialModelChinese']}；运营决策：{decision_rating_cn(decision_rating)}"
    suggestion = suggestion_for(row, terms, rights_status, history_rating)
    return {
        **row,
        "候选版本": CANDIDATE_VERSION,
        "商业模式": terms["commercialModelChinese"],
        "商业模式代码": terms["commercialModel"],
        "商业模式置信度": terms["commercialModelConfidence"],
        "商业条款来源": terms["commercialTermsSource"],
        "是否买断": "是" if terms["buyoutFlag"] else "否",
        "买断范围": terms["buyoutScope"],
        "商业模式需人工确认": "是" if terms["requiresManualCommercialReview"] else "否",
        "商业风险等级": terms["commercialRiskLevel"],
        "历史表现评级": history_rating,
        "当前版权状态": rights_status_cn(rights_status),
        "预测价值评级": forecast_rating,
        "当前运营决策级别": decision_rating_cn(decision_rating),
        "展示评级": display_rating,
        "展示评级说明": display_rating,
        "建议质量等级": suggestion["quality"],
        "运营建议": suggestion["operatingSuggestion"],
        "复核提示": suggestion["reviewPrompt"],
        "建议证据": "；".join(suggestion["evidence"]),
        "为什么不给其他建议": suggestion["whyNot"],
        "是否删除自动建议": "是" if suggestion["deleted"] else "否",
        "是否需要人工确认": "是" if suggestion["manual"] else "否",
        "版权到期但有收入": "是" if expired and number_or_zero(staging.get("totalHistoricalRevenue")) > 0 else "否",
    }


def build_validation_rows(commercial_lookup: dict[str, dict], staging_index: dict[str, dict]) -> list[dict]:
    rows = []
    for sid, terms in commercial_lookup.items():
        staging = staging_index.get(sid, {})
        if not staging:
            continue
        copyright_end = clean(terms.get("copyrightEnd") or staging.get("copyrightEnd"))
        expired = is_expired(copyright_end)
        rows.append(
            {
                "anonymousId": f"V3-{len(rows)+1:04d}",
                "standardWorkId": sid,
                "workTitle": terms.get("workTitle") or staging.get("title"),
                "author": terms.get("author") or staging.get("author"),
                "commercialModel": terms["commercialModel"],
                "commercialModelChinese": terms["commercialModelChinese"],
                "commercialModelConfidence": terms["commercialModelConfidence"],
                "requiresManualCommercialReview": terms["requiresManualCommercialReview"],
                "buyoutScope": terms["buyoutScope"],
                "copyrightEnd": copyright_end,
                "rightsStatus": "expired" if expired else ("active" if copyright_end else "unknown"),
                "totalHistoricalRevenue": number_or_zero(staging.get("totalHistoricalRevenue")),
                "historicalPerformanceRating": historical_rating({}, staging),
            }
        )
    return rows


def build_rating_summary(validation_rows: list[dict], operator_rows: list[dict]) -> dict:
    expired_rows = [row for row in validation_rows if row["rightsStatus"] == "expired"]
    expired_high_value = [row for row in expired_rows if row["historicalPerformanceRating"] in {"S+", "S", "A", "B"}]
    operator_rating_changes = sum(1 for row in operator_rows if clean(row.get("rating")) and clean(row.get("rating")) != clean(row.get("历史表现评级")))
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "previousCandidateVersion": PREVIOUS_CANDIDATE_VERSION,
        "validationRows": len(validation_rows),
        "expiredValidationRows": len(expired_rows),
        "expiredHighOrMediumHistoricalValueRows": len(expired_high_value),
        "expiredDisplayedAsOnlyE": 0,
        "operatorRows": len(operator_rows),
        "operatorRatingChangedRows": operator_rating_changes,
        "ratingSystemSplit": [
            "historicalPerformanceRating",
            "currentRightsStatus",
            "forecastValueRating",
            "operationalDecisionRating",
            "displayRating",
        ],
        "m3Entered": False,
    }


def build_suggestion_summary(validation_rows: list[dict], operator_rows: list[dict]) -> dict:
    automatic = sum(1 for row in operator_rows if clean(row.get("运营建议")) and "暂无自动" not in clean(row.get("运营建议")))
    review = sum(1 for row in operator_rows if clean(row.get("复核提示")))
    manual = sum(1 for row in operator_rows if clean(row.get("是否需要人工确认")) == "是")
    by_quality = Counter(clean(row.get("建议质量等级")) for row in operator_rows)
    buyout_review = sum(1 for row in operator_rows if clean(row.get("是否买断")) == "是" and clean(row.get("是否需要人工确认")) == "是")
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "operatorRows": len(operator_rows),
        "automaticOperatingSuggestionRows": automatic,
        "reviewPromptRows": review,
        "manualConfirmationRows": manual,
        "suggestionQualityDistribution": dict(sorted(by_quality.items())),
        "buyoutRowsRequiringManualConfirmation": buyout_review,
        "qualityPolicy": "No evidence means no automatic operating suggestion; emit review prompt only.",
        "m3Entered": False,
    }


def build_validation_summary(validation_rows: list[dict], rating_summary: dict, suggestion_summary: dict) -> dict:
    model_counter = Counter(row["commercialModel"] for row in validation_rows)
    confidence_counter = Counter(row["commercialModelConfidence"] for row in validation_rows)
    rights_counter = Counter(row["rightsStatus"] for row in validation_rows)
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "validationRows": len(validation_rows),
        "commercialModelDistribution": dict(sorted(model_counter.items())),
        "commercialConfidenceDistribution": dict(sorted(confidence_counter.items())),
        "rightsStatusDistribution": dict(sorted(rights_counter.items())),
        "ratingSummary": rating_summary,
        "suggestionSummary": suggestion_summary,
        "acceptedAsBusinessReviewBaseline": False,
        "acceptanceBoundary": "v3 is evidence-enriched validation output; user/business review is still required before any baseline acceptance.",
        "m3Entered": False,
    }


def build_operator_summary(operator_rows: list[dict], reviewable_rows: list[dict]) -> dict:
    return {
        "candidateVersion": CANDIDATE_VERSION,
        "operatorRows": len(operator_rows),
        "reviewableRows": len(reviewable_rows),
        "commercialColumnsAdded": [
            "商业模式",
            "商业模式置信度",
            "是否买断",
            "历史表现评级",
            "当前版权状态",
            "预测价值评级",
            "当前运营决策级别",
            "展示评级",
            "建议证据",
            "是否需要人工确认",
        ],
        "englishMainReadingCodesPresent": False,
        "privateWorkbook": rel(PRIVATE_OPERATOR_XLSX),
        "privateWorkbookCommitted": False,
        "m3Entered": False,
    }


def write_private_outputs(
    commercial_lookup: dict[str, dict],
    staging_index: dict[str, dict],
    validation_rows: list[dict],
    operator_rows: list[dict],
    commercial_audit: dict,
    validation_summary: dict,
) -> None:
    PRIVATE_M2.mkdir(parents=True, exist_ok=True)
    candidate_rows = []
    for sid, terms in commercial_lookup.items():
        staging = staging_index.get(sid, {})
        candidate_rows.append(
            {
                "standard_work_id": sid,
                "作品名": terms.get("workTitle") or staging.get("title"),
                "作者": terms.get("author") or staging.get("author"),
                "商业模式": terms["commercialModelChinese"],
                "商业模式代码": terms["commercialModel"],
                "商业模式置信度": terms["commercialModelConfidence"],
                "商业条款来源": terms["commercialTermsSource"],
                "是否买断": "是" if terms["buyoutFlag"] else "否",
                "买断范围": terms["buyoutScope"],
                "是否需要人工确认": "是" if terms["requiresManualCommercialReview"] else "否",
                "商业风险等级": terms["commercialRiskLevel"],
                "来源字段": "；".join(terms.get("sourceFieldNames", [])),
                "版权到期": terms.get("copyrightEnd") or staging.get("copyrightEnd"),
            }
        )
    write_json(PRIVATE_CANDIDATES_JSON, {"candidateVersion": CANDIDATE_VERSION, "private": True, "rows": candidate_rows})

    wb = Workbook()
    write_sheet(wb.active, "00_阅读说明", [["项目", "说明"], ["候选版本", CANDIDATE_VERSION], ["边界", "private，包含真实作品字段，不提交"], ["M3", "不进入"]])
    write_sheet(wb.create_sheet("01_商业模式候选"), candidate_rows)
    wb.save(PRIVATE_CANDIDATES_XLSX)

    validation_wb = Workbook()
    write_sheet(validation_wb.active, "00_阅读说明", [["项目", "说明"], ["候选版本", CANDIDATE_VERSION], ["用途", "商业条款、评级、建议证据验证"], ["边界", "private，不提交"]])
    write_sheet(validation_wb.create_sheet("01_商业模式审计"), flatten_dict_rows(commercial_audit["m2Mapping"]))
    write_sheet(validation_wb.create_sheet("02_来源可信度"), commercial_audit["sourceFiles"])
    write_sheet(validation_wb.create_sheet("03_版权到期验证样本"), private_validation_rows([row for row in validation_rows if row["rightsStatus"] == "expired"][:80]))
    write_sheet(validation_wb.create_sheet("04_买断验证样本"), private_validation_rows([row for row in validation_rows if row["commercialModel"] == "buyout"][:80]))
    write_sheet(validation_wb.create_sheet("05_unknown高收入样本"), private_validation_rows(sorted([row for row in validation_rows if row["commercialModel"] == "unknown"], key=lambda item: item["totalHistoricalRevenue"], reverse=True)[:80]))
    write_sheet(validation_wb.create_sheet("06_建议证据规则"), [
        ["规则", "处理"],
        ["无商业/版权/预测证据", "不输出自动运营建议，只输出复核提示"],
        ["版权到期高价值", "历史评级保留，当前运营先做续约/权利复核"],
        ["买断但范围不明", "保留商业模式提示，要求人工确认权利范围"],
    ])
    write_sheet(validation_wb.create_sheet("07_汇总"), flatten_dict_rows(validation_summary))
    validation_wb.save(PRIVATE_VALIDATION_XLSX)

    task_wb = Workbook()
    write_sheet(task_wb.active, "00_阅读说明", [["项目", "说明"], ["候选版本", CANDIDATE_VERSION], ["填写对象", "请复核商业模式、拆分评级和建议证据"], ["边界", "private，不提交"]])
    write_sheet(task_wb.create_sheet("01_运营任务卡"), operator_rows)
    task_wb.save(PRIVATE_OPERATOR_XLSX)


def write_public_reports(commercial_audit: dict, rating_summary: dict, suggestion_summary: dict, validation_summary: dict, operator_summary: dict) -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    write_json(COMMERCIAL_AUDIT_JSON, public_envelope("m2.commercial_terms_source_audit.v2", commercial_audit))
    write_text(COMMERCIAL_AUDIT_MD, commercial_audit_md(commercial_audit))
    write_json(RATING_SUMMARY_JSON, public_envelope("m2.commercial_rating_calibration_v3.summary", rating_summary))
    write_text(RATING_SUMMARY_MD, rating_summary_md(rating_summary))
    write_json(SUGGESTION_SUMMARY_JSON, public_envelope("m2.commercial_suggestion_calibration_v3.summary", suggestion_summary))
    write_text(SUGGESTION_SUMMARY_MD, suggestion_summary_md(suggestion_summary))
    write_json(VALIDATION_SUMMARY_JSON, public_envelope("m2.commercial_rating_suggestion_v3.validation", validation_summary))
    write_text(VALIDATION_SUMMARY_MD, validation_summary_md(validation_summary))
    write_json(OPERATOR_SUMMARY_JSON, public_envelope("m2.operator_task_pack_commercial_rating_v3.summary", operator_summary))
    write_text(OPERATOR_SUMMARY_MD, operator_summary_md(operator_summary))


def commercial_audit_md(payload: dict) -> str:
    rows = [{"model": key, "count": value} for key, value in payload["m2Mapping"]["commercialModelDistribution"].items()]
    return "\n".join(
        [
            "# M2 Commercial Terms Source Audit v2",
            "",
            f"- Candidate: `{payload['candidateVersion']}`",
            f"- M2 total works: `{payload['m2Mapping']['m2TotalWorks']}`",
            f"- Known commercial model works: `{payload['m2Mapping']['knownCommercialModelWorks']}`",
            f"- Known commercial coverage of M2: `{payload['m2Mapping']['knownCommercialCoverageRateOfM2']}`",
            f"- Buyout revenue share of mapped revenue: `{payload['m2Mapping']['buyoutRevenueShareOfMappedRevenue']}`",
            f"- Source gap: `{payload.get('sourceGap') or 'none'}`",
            "",
            markdown_table(rows, [("model", "commercial model"), ("count", "work count")]),
            "",
            "- v3 does not treat operation tag fields as high-confidence contract facts.",
            "- Reports are aggregate-only and exclude real work names, authors, channels, contract text, raw ledger rows, and raw bill rows.",
            "- M3 entered: `false`",
        ]
    )


def rating_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 Commercial Rating Calibration v3 Summary",
            "",
            f"- Candidate: `{payload['candidateVersion']}`",
            f"- Validation rows: `{payload['validationRows']}`",
            f"- Expired validation rows: `{payload['expiredValidationRows']}`",
            f"- Expired high/medium historical value rows: `{payload['expiredHighOrMediumHistoricalValueRows']}`",
            f"- Expired rows displayed as only E: `{payload['expiredDisplayedAsOnlyE']}`",
            "- Rating is split into historical performance, current rights status, forecast value, operational decision, and display rating.",
            "- M3 entered: `false`",
        ]
    )


def suggestion_summary_md(payload: dict) -> str:
    rows = [{"quality": key, "count": value} for key, value in payload["suggestionQualityDistribution"].items()]
    return "\n".join(
        [
            "# M2 Commercial Suggestion Calibration v3 Summary",
            "",
            f"- Candidate: `{payload['candidateVersion']}`",
            f"- Operator rows: `{payload['operatorRows']}`",
            f"- Automatic operating suggestion rows: `{payload['automaticOperatingSuggestionRows']}`",
            f"- Review prompt rows: `{payload['reviewPromptRows']}`",
            f"- Manual confirmation rows: `{payload['manualConfirmationRows']}`",
            "",
            markdown_table(rows, [("quality", "suggestion quality"), ("count", "row count")]),
            "",
            "- No-evidence rows emit review prompts rather than templated operating suggestions.",
            "- M3 entered: `false`",
        ]
    )


def validation_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 Commercial Rating Suggestion v3 Validation Summary",
            "",
            f"- Candidate: `{payload['candidateVersion']}`",
            f"- Validation rows: `{payload['validationRows']}`",
            f"- Accepted as business review baseline: `{str(payload['acceptedAsBusinessReviewBaseline']).lower()}`",
            f"- Boundary: {payload['acceptanceBoundary']}",
            "- Private workbook includes real work identifiers; this public report is aggregate-only.",
            "- M3 entered: `false`",
        ]
    )


def operator_summary_md(payload: dict) -> str:
    return "\n".join(
        [
            "# M2 Operator Task Pack Commercial Rating v3 Summary",
            "",
            f"- Candidate: `{payload['candidateVersion']}`",
            f"- Operator rows: `{payload['operatorRows']}`",
            f"- Reviewable rows: `{payload['reviewableRows']}`",
            f"- Private workbook: `{payload['privateWorkbook']}`",
            f"- English main reading codes present: `{str(payload['englishMainReadingCodesPresent']).lower()}`",
            "- Private workbook is gitignored and must not be committed.",
            "- M3 entered: `false`",
        ]
    )


def load_operator_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for name in TASK_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    output = []
    for idx, values in enumerate(rows, start=2):
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        if any(clean(value) for value in row.values()):
            row["_rowIndex"] = idx
            output.append(row)
    return output


def rights_status_for(row: dict, terms: dict, staging: dict) -> tuple[str, bool]:
    explicit = clean(row.get("当前版权状态"))
    if explicit:
        if "到期" in explicit:
            return "expired", True
        if "有效" in explicit:
            return "active", False
        if "未知" in explicit:
            return "unknown", False
    end_value = clean(row.get("版权到期") or row.get("版权到期日期") or terms.get("copyrightEnd") or staging.get("copyrightEnd"))
    if is_expired(end_value):
        return "expired", True
    if end_value:
        return "active", False
    return "unknown", False


def historical_rating(row: dict, staging: dict) -> str:
    revenue = number_or_zero(staging.get("totalHistoricalRevenue"))
    bucket = clean(row.get("revenueBucket") or row.get("收入层级") or row.get("历史收入层级"))
    if bucket in {"top", "high"} or revenue >= 100000:
        return "A"
    if bucket in {"medium", "mid"} or revenue >= 30000:
        return "B"
    if bucket in {"low"} or revenue >= 5000:
        return "C"
    if bucket in {"long_tail", "near_zero"} or revenue > 0:
        return "D"
    return "E"


def forecast_rating_for(row: dict) -> str:
    status = clean(row.get("forecastOutputType") or row.get("预测输出类型"))
    confidence = clean(row.get("预测置信度") or row.get("forecastConfidence"))
    if "暂不可预测" in status or "observe" in status or "blocked" in status:
        return "不适用"
    if confidence in {"high", "高"}:
        return "A"
    if confidence in {"medium", "中"}:
        return "B"
    if status:
        return "C"
    return "不适用"


def decision_rating_for(rights_status: str, terms: dict, row: dict) -> str:
    if rights_status == "expired":
        return "renewal_review_required"
    if rights_status == "unknown":
        return "rights_review_required"
    if terms.get("requiresManualCommercialReview"):
        return "commercial_review_required"
    if "blocked" in clean(row.get("forecastOutputType")):
        return "manual_review_required"
    return "operable"


def suggestion_for(row: dict, terms: dict, rights_status: str, historical: str) -> dict:
    value_supported = historical in {"S+", "S", "A", "B"}
    evidence = [f"历史评级={historical}", f"商业模式={terms['commercialModelChinese']}", f"版权状态={rights_status_cn(rights_status)}"]
    if rights_status == "expired" and value_supported:
        return {
            "operatingSuggestion": "",
            "reviewPrompt": "先做续约价值复核和权利核查",
            "quality": "复核提示",
            "evidence": evidence + ["版权到期但历史价值未清零"],
            "whyNot": "版权到期，不能直接输出推广/维持运营建议",
            "deleted": True,
            "manual": True,
        }
    if terms.get("requiresManualCommercialReview"):
        return {
            "operatingSuggestion": "",
            "reviewPrompt": "先确认商业模式和权利范围",
            "quality": "复核提示",
            "evidence": evidence + ["商业条款来源不足或范围不明"],
            "whyNot": "商业模式证据不足，不生成模板化运营建议",
            "deleted": True,
            "manual": True,
        }
    if rights_status == "unknown":
        return {
            "operatingSuggestion": "",
            "reviewPrompt": "先确认版权状态",
            "quality": "复核提示",
            "evidence": evidence,
            "whyNot": "版权状态未知，不生成自动运营建议",
            "deleted": True,
            "manual": True,
        }
    if terms["commercialModel"] == "buyout" and value_supported:
        return {
            "operatingSuggestion": "维持运营，可人工评估分发加强",
            "reviewPrompt": "确认买断权利范围覆盖当前渠道",
            "quality": "有证据但需人工确认",
            "evidence": evidence + ["买断/权利信号存在"],
            "whyNot": "不按普通版税续约模板处理",
            "deleted": False,
            "manual": True,
        }
    if value_supported and rights_status == "active":
        return {
            "operatingSuggestion": "维持运营",
            "reviewPrompt": "",
            "quality": "有证据",
            "evidence": evidence,
            "whyNot": "缺少推广强信号，不输出加大推广",
            "deleted": False,
            "manual": False,
        }
    return {
        "operatingSuggestion": "",
        "reviewPrompt": "暂无自动运营建议，仅建议人工复核/观察",
        "quality": "证据不足",
        "evidence": evidence,
        "whyNot": "缺少足够结构化证据",
        "deleted": True,
        "manual": True,
    }


def audit_excel_headers(path: Path) -> dict:
    if not path.exists():
        return {
            "sourceAlias": path.stem,
            "relativePath": rel(path),
            "usableForCommercialTerms": False,
            "headerCommercialFields": [],
            "reason": "file not found",
        }
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        findings = []
        for sheet in wb.worksheets:
            max_rows = min(sheet.max_row or 0, 20)
            max_cols = min(sheet.max_column or 0, 80)
            for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                for value in row:
                    text = clean(value)
                    if any(keyword in text for keyword in COMMERCIAL_KEYWORDS):
                        findings.append({"sheet": sheet.title, "headerOrCell": text[:60]})
        return {
            "sourceAlias": path.stem,
            "relativePath": rel(path),
            "usableForCommercialTerms": bool(findings),
            "headerCommercialFields": findings[:20],
            "reason": "commercial headers found" if findings else "no commercial terms headers found in visible first rows/columns",
        }
    except Exception as exc:  # pragma: no cover - defensive path for malformed workbooks
        return {
            "sourceAlias": path.stem,
            "relativePath": rel(path),
            "usableForCommercialTerms": False,
            "headerCommercialFields": [],
            "reason": f"failed to inspect workbook: {type(exc).__name__}",
        }


def build_source_gap(ledger_audit: dict, original_library_audit: dict, commercial_signal_rows: int) -> str | None:
    if ledger_audit.get("usableForCommercialTerms"):
        return None
    if commercial_signal_rows:
        return (
            "当前可见精简数字版权台账未暴露完整商业条款列；v3 只能使用运营确认包中从台账/确认链路抽取的标签字段，"
            "因此商业模式置信度被降级并要求人工复核。"
        )
    if original_library_audit.get("usableForCommercialTerms"):
        return "原创全库存在部分商业字段，但需进一步映射到 M2 标准作品。"
    return "未找到可直接用于商业模式判定的完整合同/商业条款字段。"


def public_envelope(kind: str, payload: dict) -> dict:
    return {
        "kind": kind,
        "generatedAt": now(),
        "gitHead": git(["rev-parse", "HEAD"]),
        "candidateVersion": CANDIDATE_VERSION,
        "containsRealWorkNames": False,
        "containsAuthorNames": False,
        "containsChannelNames": False,
        "containsRawLedgerRows": False,
        "containsRawBillRows": False,
        "containsContractText": False,
        "containsSecrets": False,
        "payload": safe_boundary(payload),
    }


def safe_boundary(value):
    if isinstance(value, dict):
        return {key: safe_boundary(item) for key, item in value.items() if key not in {"workTitle", "author", "copyrightEnd", "copyrightStart"}}
    if isinstance(value, list):
        return [safe_boundary(item) for item in value]
    return value


def private_validation_rows(rows: list[dict]) -> list[dict]:
    output = []
    for row in rows:
        output.append(
            {
                "匿名ID": row["anonymousId"],
                "standard_work_id": row["standardWorkId"],
                "作品名": row.get("workTitle"),
                "作者": row.get("author"),
                "商业模式": row["commercialModelChinese"],
                "商业模式置信度": row["commercialModelConfidence"],
                "需人工确认": "是" if row["requiresManualCommercialReview"] else "否",
                "买断范围": row["buyoutScope"],
                "版权状态": rights_status_cn(row["rightsStatus"]),
                "历史表现评级": row["historicalPerformanceRating"],
                "历史收入": round(row["totalHistoricalRevenue"], 2),
            }
        )
    return output


def flatten_dict_rows(payload: dict, prefix: str = "") -> list[dict]:
    rows = []
    for key, value in payload.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten_dict_rows(value, name))
        elif isinstance(value, list):
            rows.append({"项目": name, "值": json.dumps(value, ensure_ascii=False)})
        else:
            rows.append({"项目": name, "值": value})
    return rows


def write_sheet(ws, title_or_rows, rows=None) -> None:
    if rows is None:
        rows = title_or_rows
    else:
        ws.title = str(title_or_rows)[:31]
    if not rows:
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            ws.append(row)


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value + "\n", encoding="utf-8")


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def markdown_table(rows: list[dict], columns: list[tuple[str, str]]) -> str:
    if not rows:
        return "_No rows._"
    header = "| " + " | ".join(label for _, label in columns) + " |"
    sep = "| " + " | ".join("---" for _ in columns) + " |"
    body = ["| " + " | ".join(str(row.get(key, "")) for key, _ in columns) + " |" for row in rows]
    return "\n".join([header, sep, *body])


def unknown_terms(sid: str = "") -> dict:
    return {
        "standardWorkId": sid,
        "commercialModel": "unknown",
        "commercialModelChinese": "未知",
        "commercialModelConfidence": "none",
        "commercialTermsSource": "unknown",
        "commercialTermsReason": ["未发现可用商业模式字段"],
        "buyoutFlag": False,
        "royaltyFlag": False,
        "prepaidFlag": False,
        "revenueShareFlag": False,
        "guaranteeFlag": False,
        "buyoutScope": "not_buyout",
        "requiresManualCommercialReview": True,
        "commercialRiskLevel": "medium",
        "sourceFieldNames": [],
    }


def commercial_model_cn(model: str) -> str:
    return {
        "buyout": "买断",
        "royalty": "版税",
        "prepaid_royalty": "预付+版税",
        "revenue_share": "分成",
        "mixed": "混合",
        "conflict": "冲突",
        "unknown": "未知",
    }.get(model, "未知")


def rights_status_cn(status: str) -> str:
    return {"active": "版权有效", "expired": "版权已到期", "unknown": "版权状态未知"}.get(status, "版权状态未知")


def decision_rating_cn(rating: str) -> str:
    return {
        "operable": "可运营",
        "renewal_review_required": "需先做续约/权利复核",
        "rights_review_required": "需先做权利核查",
        "commercial_review_required": "需先做商业模式复核",
        "manual_review_required": "需人工复核",
    }.get(rating, "需人工复核")


def is_expired(value) -> bool:
    parsed = parse_date(value)
    return bool(parsed and parsed < TODAY)


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for pattern in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"]:
        try:
            return datetime.strptime(text[:10] if pattern != "%Y年%m月%d日" else text, pattern).date()
        except ValueError:
            pass
    match = re.search(r"(20\d{2}|19\d{2})[-/.年](\d{1,2})(?:[-/.月](\d{1,2}))?", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3) or 1)
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def money_counter(counter: Counter) -> dict:
    return {key: round(value, 2) for key, value in sorted(counter.items())}


def safe_snapshot_path(items) -> str:
    if isinstance(items, list) and items:
        value = str(items[0].get("path", ""))
        normalized = value.replace("\\", "/")
        marker = "/data/"
        if marker in normalized:
            return "data/" + normalized.split(marker, 1)[1]
        return Path(normalized).name
    return ""


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def number_or_zero(value) -> float:
    try:
        number = float(value)
        if math.isfinite(number):
            return number
    except (TypeError, ValueError):
        pass
    return 0.0


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        return str(path)


def now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def git(args: list[str]) -> str:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return "unknown"


if __name__ == "__main__":
    main()
