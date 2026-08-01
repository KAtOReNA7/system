#!/usr/bin/env python3
"""Build the scope-aware source-authority ledger for HPSR01/02.

The audit validates field presence, exact cross-source equality and amount
sign without aggregating or exposing an outcome. It separates the active
WORK_TOTAL development gate from the PARTIAL_NOT_ACTIVE WORK_CHANNEL gate.
No model is fitted, predicted, scored or bootstrapped here. Missing derived
caches and historical receipts never masquerade as missing source authority.
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import tempfile
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import export_m2_reversal_authority as reversal


ROOT = Path(__file__).resolve().parents[2]
READINESS_CONFIG = (
    ROOT / "config" / "m2-current-human-anchored-later-origin.v0.1.json"
)
HPSR_CONFIG = (
    ROOT / "config" / "m2-current-head-protected-segmented-router.v0.1.json"
)
FEATURE_CACHE = (
    ROOT
    / "data"
    / "private-output"
    / "m2-core-legacy-horizon-amount"
    / "M2-core-legacy-horizon-amount-feature-rows-private-v0.1.ndjson"
)
FROZEN_LG01_CACHE = (
    ROOT
    / "data"
    / "private-output"
    / "m2-core-legacy-horizon-amount"
    / "M2-core-legacy-horizon-amount-frozen-lg01-rows-private-v0.1.ndjson"
)
PRIVATE_LEDGER = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-head-protected-segmented-router-opened-origin-semantics-private-v0.2.json"
)
SOURCE_RECONCILIATION = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-source-authority-reconciliation-private-v0.2.json"
)
WORK_TOTAL_AUTHORITY_FACTS = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-work-total-authority-private-v0.2.ndjson"
)
WORK_TOTAL_AUTHORITY_RECEIPT = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-work-total-authority-receipt-private-v0.2.json"
)
ORIGIN_VISIBLE_AUTHORITY_FACTS = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-origin-visible-authority-private-v0.2.ndjson"
)
ORIGIN_VISIBLE_AUTHORITY_RECEIPT = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-origin-visible-authority-receipt-private-v0.2.json"
)
WORK_TOTAL_SCOPE_ASSESSMENT = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-work-total-scope-assessment-private-v0.2.json"
)
SOURCE_AUTHORITY_DECISION_TABLE = (
    ROOT
    / "data"
    / "private-output"
    / "m2-head-protected-segmented-router"
    / "M2-hpsr02-source-authority-decision-table-private-v0.2.json"
)
CHAM_RECEIPT = (
    ROOT
    / "data"
    / "private-output"
    / "m2-core-legacy-horizon-amount"
    / "M2-core-legacy-horizon-amount-attempt-receipt-private-v0.1.json"
)
CHAM_MANIFEST = (
    ROOT
    / "data"
    / "private-output"
    / "m2-core-legacy-horizon-amount"
    / "M2-core-legacy-horizon-amount-manifest-private-v0.1.json"
)
MONTH_PATTERN = re.compile(r"^\d{4}-(?:0[1-9]|1[0-2])$")
HPSR02_ORIGIN = "2026-03"
WORK_TOTAL_MAPPING_WARNING = (
    "WORK_TOTAL_CANONICAL_MAPPING_WARNING_"
    "WORK_CHANNEL_REMAINS_PARTIAL"
)
NON_EVALUATION_METADATA_WARNING = (
    "NON_EVALUATION_METADATA_DIFFERENCE_WARNING"
)


class DateAuditError(RuntimeError):
    """The date-only audit cannot prove its frozen public assertions."""


def _repository_relative(path: Path) -> str:
    return path.resolve().relative_to(ROOT).as_posix()


def _month_index(value: str) -> int:
    if not MONTH_PATTERN.fullmatch(value):
        raise DateAuditError(f"invalid month key: {value!r}")
    year, month = (int(part) for part in value.split("-"))
    return year * 12 + month - 1


def _add_months(value: str, offset: int) -> str:
    target = _month_index(value) + offset
    year, month_index = divmod(target, 12)
    return f"{year:04d}-{month_index + 1:02d}"


def _normalize_month(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        result = value.strftime("%Y-%m")
    else:
        text = str(value).strip()
        match = re.match(r"^(\d{4})[-/.年](\d{1,2})", text)
        if not match:
            return None
        result = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return result if MONTH_PATTERN.fullmatch(result) else None


def _scan_opened_cache(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise DateAuditError(f"rebuildable opened-origin cache missing: {path.name}")
    origins: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "rowCount": 0,
            "nonNullExistingActualCount": 0,
            "horizonsMonths": set(),
            "labelDateKeys": set(),
        }
    )
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                raise DateAuditError(
                    f"invalid NDJSON at {path.name}:{line_number}"
                ) from exc
            origin = str(row.get("origin", ""))
            if not MONTH_PATTERN.fullmatch(origin):
                raise DateAuditError(
                    f"missing origin at {path.name}:{line_number}"
                )
            entry = origins[origin]
            entry["rowCount"] += 1
            entry["nonNullExistingActualCount"] += row.get("actual") is not None
            horizon = row.get("horizonMonths")
            if horizon is not None:
                entry["horizonsMonths"].add(int(horizon))
            for key in ("labelAvailableAsOf", "maximumLabelAvailableAsOf"):
                value = row.get(key)
                if value is not None and MONTH_PATTERN.fullmatch(str(value)):
                    entry["labelDateKeys"].add(str(value))
    rows = []
    for origin, entry in sorted(origins.items()):
        rows.append(
            {
                "origin": origin,
                "rowCount": entry["rowCount"],
                "nonNullExistingActualCount": entry[
                    "nonNullExistingActualCount"
                ],
                "horizonsMonths": sorted(entry["horizonsMonths"]),
                "labelDateKeys": sorted(entry["labelDateKeys"]),
            }
        )
    if not rows:
        raise DateAuditError(f"opened-origin cache empty: {path.name}")
    return {
        "role": (
            "frozen-development-feature-rows"
            if path == FEATURE_CACHE
            else "frozen-lg01-same-case-rows"
        ),
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "repositoryRelativePath": _repository_relative(path),
        "scannedFields": [
            "origin",
            "horizonMonths",
            "labelAvailableAsOf_or_maximumLabelAvailableAsOf",
            "actual_nullness_only",
        ],
        "originCount": len(rows),
        "minimumOrigin": rows[0]["origin"],
        "maximumOrigin": rows[-1]["origin"],
        "origins": rows,
    }


def _scan_opened_cache_if_present(path: Path) -> dict[str, Any]:
    if path.is_file():
        return _scan_opened_cache(path)
    return {
        "role": (
            "frozen-development-feature-rows"
            if path == FEATURE_CACHE
            else "frozen-lg01-same-case-rows"
        ),
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "repositoryRelativePath": _repository_relative(path),
        "status": "CACHE_MISS_REBUILDABLE",
        "scannedFields": [],
        "originCount": None,
        "minimumOrigin": None,
        "maximumOrigin": None,
        "origins": [],
    }


def _metadata_row_key(values: tuple[Any, ...]) -> tuple[str, ...]:
    return tuple(reversal.clean_text(value) for value in values)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _stable_hash(value: Any) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _exact_amount(value: Any) -> tuple[str | None, Decimal | None, int]:
    if value is None or isinstance(value, bool):
        return None, None, 0
    if isinstance(value, float) and not math.isfinite(value):
        return None, None, 0
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None, None, 0
    if not amount.is_finite():
        return None, None, 0
    if amount == 0:
        return "0", Decimal("0"), 0
    raw_text = format(amount, "f")
    fraction = raw_text.partition(".")[2].rstrip("0")
    canonical_text = raw_text.rstrip("0").rstrip(".") \
        if "." in raw_text else raw_text
    return canonical_text, amount, len(fraction)


def _source_identity(
    raw_channel_id: str,
    raw_channel_name: str,
) -> tuple[str, str] | None:
    if raw_channel_id:
        return ("RAW_CHANNEL_ID", raw_channel_id)
    if raw_channel_name:
        return ("RAW_CHANNEL_NAME", raw_channel_name)
    return None


def _identity_token(identity: tuple[str, str] | None) -> str | None:
    if identity is None:
        return None
    return f"{identity[0]}:{_stable_hash(identity)[:24]}"


def _cash_category(split_type: str) -> str | None:
    if split_type in {"", "分成"}:
        return "sales_share"
    if split_type == "买断":
        return "buyout"
    return None


def _canonical_alias_index(
    channel_master: dict[tuple[str, str], dict],
) -> dict[tuple[str, str], set[str]]:
    output: dict[tuple[str, str], set[str]] = defaultdict(set)
    for (raw_id, raw_name), row in channel_master.items():
        identity = _source_identity(
            reversal.canonical.clean(raw_id),
            reversal.canonical.clean(raw_name),
        )
        channel_uid = reversal.canonical.clean(row.get("channelUid"))
        if identity is not None and channel_uid:
            output[identity].add(channel_uid)
    return output


def _scan_scope_aware_ledger(
    path: Path,
    *,
    expected_types: set[str],
    ledger_role: str,
    mapping: dict[str, str] | None = None,
    channel_master: dict[tuple[str, str], dict] | None = None,
    alias_index: dict[tuple[str, str], set[str]] | None = None,
) -> dict[str, Any]:
    if not path.is_file():
        raise DateAuditError(f"source authority missing: {path.name}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [
            cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        required_headers = [
            *reversal.REAL_BILL_COLUMNS,
            reversal.TYPE_COLUMN,
        ]
        if headers != required_headers:
            raise DateAuditError(f"ledger schema mismatch: {path.name}")
        if len(required_headers) != 8 or required_headers[6] != "实销金额":
            raise DateAuditError("standard ledger amount column position changed")
        partial_markers = re.compile(
            r"partial|incomplete|部分导入|未完整",
            re.IGNORECASE,
        )
        if any(
            partial_markers.search(str(value))
            for value in [*workbook.sheetnames, *headers]
        ):
            raise DateAuditError(
                f"explicit partial-import marker present: {path.name}"
            )
        counts: Counter[str] = Counter()
        metadata_keys: Counter[tuple[str, ...]] = Counter()
        presentation_keys: Counter[tuple[str, ...]] = Counter()
        presentation_locations: dict[
            tuple[str, ...], list[int]
        ] = defaultdict(list)
        critical_keys: Counter[tuple[str, ...]] = Counter()
        nonzero_critical_keys: Counter[tuple[str, ...]] = Counter()
        critical_amounts: dict[tuple[str, ...], Decimal] = defaultdict(
            lambda: Decimal("0")
        )
        positive_occurrences: set[tuple[str, str]] = set()
        first_positive_month_by_work_identity: dict[
            tuple[str, str], str
        ] = {}
        scope_key_by_presentation: dict[
            tuple[str, ...], tuple[str, str]
        ] = {}
        canonical_member_by_presentation: dict[
            tuple[str, ...], str | None
        ] = {}
        raw_ids_by_name: dict[str, set[str]] = defaultdict(set)
        raw_names_by_id: dict[str, set[str]] = defaultdict(set)
        missing_pairs: dict[tuple[str, str], dict[str, Any]] = {}
        missing_month_count = 0
        missing_work_mapping_count = 0
        missing_source_identity_count = 0
        invalid_amount_count = 0
        amount_cell_read_count = 0
        maximum_amount_scale_power = 0
        missing_channel_mapping_count = 0
        missing_channel_pairs: set[tuple[str, str]] = set()
        missing_channel_months: Counter[str] = Counter()
        type_counts: Counter[str] = Counter()
        for row_ordinal, row in enumerate(worksheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=8,
            values_only=True,
        ), start=2):
            if all(value is None for value in row):
                continue
            amount_cell_read_count += 1
            metadata = _metadata_row_key(row[:6])
            split_type = reversal.clean_text(row[7])
            type_counts[split_type] += 1
            amount_text, amount, scale_power = _exact_amount(row[6])
            maximum_amount_scale_power = max(
                maximum_amount_scale_power,
                scale_power,
            )
            if amount_text is None or amount is None:
                invalid_amount_count += 1
                amount_text = "INVALID_AMOUNT"
            raw_channel_id = reversal.canonical.clean(metadata[1])
            raw_channel_name = reversal.canonical.clean(metadata[2])
            raw_pair = (raw_channel_id, raw_channel_name)
            identity = _source_identity(
                raw_channel_id,
                raw_channel_name,
            )
            identity_token = _identity_token(identity)
            if identity_token is None:
                missing_source_identity_count += 1
                identity_token = "MISSING_SOURCE_IDENTITY"
            if raw_channel_name:
                raw_ids_by_name[raw_channel_name].add(raw_channel_id)
            if raw_channel_id:
                raw_names_by_id[raw_channel_id].add(raw_channel_name)
            month = _normalize_month(metadata[0])
            if month is None:
                missing_month_count += 1
                month = "INVALID_MONTH"
            else:
                counts[month] += 1
            raw_work_id = reversal.normalize_raw_work_id(metadata[4])
            standard_work_id = None
            if mapping is not None:
                standard_work_id = mapping.get(raw_work_id)
                if not standard_work_id:
                    standard_work_id = reversal.derive_standard_work_id(
                        raw_work_id
                    )
                if not standard_work_id:
                    missing_work_mapping_count += 1
            else:
                standard_work_id = raw_work_id or None
            standard_work_key = standard_work_id or "MISSING_STANDARD_WORK"
            cash_category = _cash_category(split_type)
            if cash_category is None:
                cash_category = "INVALID_CASH_CATEGORY"
            reversal_class = (
                "NEGATIVE_REVERSAL"
                if amount is not None and amount < 0
                else "POSITIVE_CASH"
                if amount is not None and amount > 0
                else "ZERO_CASH"
            )
            metadata_keys[metadata] += 1
            presentation_key = (
                *metadata,
                amount_text,
                cash_category,
            )
            presentation_keys[presentation_key] += 1
            presentation_locations[presentation_key].append(row_ordinal)
            critical_key = (
                month,
                standard_work_key,
                identity_token,
                cash_category,
                "authority_ledger_native_monetary_unit",
                reversal_class,
                amount_text,
            )
            critical_keys[critical_key] += 1
            scope_key_by_presentation[presentation_key] = (
                standard_work_key,
                identity_token,
            )
            if amount is not None and amount != 0:
                nonzero_critical_keys[critical_key] += 1
                aggregation_key = critical_key[:-1]
                critical_amounts[aggregation_key] += amount
                if amount > 0 and month <= HPSR02_ORIGIN:
                    scope_key = (standard_work_key, identity_token)
                    positive_occurrences.add(scope_key)
                    previous_first = (
                        first_positive_month_by_work_identity.get(
                            scope_key
                        )
                    )
                    if previous_first is None or month < previous_first:
                        first_positive_month_by_work_identity[
                            scope_key
                        ] = month
            if channel_master is not None:
                channel_mapping = channel_master.get(raw_pair)
                canonical_member_by_presentation[presentation_key] = (
                    reversal.canonical.clean(
                        (channel_mapping or {}).get("channelUid")
                    ) or None
                )
                if not channel_mapping or not channel_mapping.get(
                    "channelUid"
                ):
                    missing_channel_mapping_count += 1
                    missing_channel_pairs.add(raw_pair)
                    if month is not None:
                        missing_channel_months[month] += 1
                    issue = missing_pairs.setdefault(raw_pair, {
                        "rawPair": raw_pair,
                        "identity": identity,
                        "identityToken": identity_token,
                        "rowCount": 0,
                        "workIds": set(),
                        "months": set(),
                        "rowOrdinals": [],
                        "criticalKeys": Counter(),
                        "invalidAmountCount": 0,
                        "missingWorkCount": 0,
                    })
                    issue["rowCount"] += 1
                    issue["workIds"].add(standard_work_key)
                    issue["months"].add(month)
                    if len(issue["rowOrdinals"]) < 10:
                        issue["rowOrdinals"].append(row_ordinal)
                    issue["criticalKeys"][critical_key] += 1
                    issue["invalidAmountCount"] += int(amount is None)
                    issue["missingWorkCount"] += int(
                        standard_work_id is None
                    )
    finally:
        workbook.close()
    if not counts:
        raise DateAuditError(f"ledger has no bill-month values: {path.name}")
    if set(type_counts) != expected_types:
        raise DateAuditError(f"ledger split type invalid: {path.name}")
    if sum(type_counts.values()) != sum(metadata_keys.values()):
        raise DateAuditError(f"ledger split type row mismatch: {path.name}")
    metadata_collision_count = sum(
        count - 1 for count in metadata_keys.values() if count > 1
    )
    profile = {
        "role": "reviewed-ledger-authority",
        "ledgerRole": ledger_role,
        "artifactClass": "PRIVATE_SOURCE_AUTHORITY",
        "repositoryRelativePath": _repository_relative(path),
        "scannedFields": [
            "billMonth",
            "rawChannelId",
            "rawChannelName",
            "authorizationCategory",
            "rawWorkId",
            "workName",
            "splitType",
            "amount_validity_and_exact_equality_only",
            "currencyScope",
            "stableRawSourceIdentity",
            "reversalSign",
        ],
        "amountColumnHeaderValidated": True,
        "amountCellReadCount": amount_cell_read_count,
        "amountUse": (
            "SOURCE_AUTHORITY_VALIDITY_EQUALITY_AND_SIGN_ONLY_"
            "NOT_OUTCOME_AGGREGATION"
        ),
        "invalidAmountCount": invalid_amount_count,
        "maximumAmountScalePower": maximum_amount_scale_power,
        "schemaValid": True,
        "splitTypeValues": sorted(type_counts),
        "rowCount": sum(metadata_keys.values()),
        "distinctMetadataKeyCount": len(metadata_keys),
        "metadataCollisionCount": metadata_collision_count,
        "missingMonthCount": missing_month_count,
        "missingWorkMappingCount": missing_work_mapping_count,
        "missingStableSourceIdentityCount": missing_source_identity_count,
        "missingCanonicalChannelMappingCount":
            missing_channel_mapping_count,
        "missingCanonicalRawPairCount": len(missing_channel_pairs),
        "missingCanonicalChannelMonths": [
            {
                "billMonth": month,
                "rowCount": missing_channel_months[month],
            }
            for month in sorted(missing_channel_months)
        ],
        "minimumBillMonth": min(counts),
        "maximumBillMonth": max(counts),
        "monthlyFactCounts": [
            {"billMonth": month, "factCount": counts[month]}
            for month in sorted(counts)
        ],
        "explicitPartialImportMarkerPresent": False,
    }
    return {
        "profile": profile,
        "presentationKeys": presentation_keys,
        "presentationLocations": presentation_locations,
        "criticalKeys": critical_keys,
        "nonzeroCriticalKeys": nonzero_critical_keys,
        "criticalAmounts": dict(critical_amounts),
        "positiveOccurrences": positive_occurrences,
        "firstPositiveMonthByWorkIdentity": (
            first_positive_month_by_work_identity
        ),
        "scopeKeyByPresentation": scope_key_by_presentation,
        "canonicalMemberByPresentation": (
            canonical_member_by_presentation
        ),
        "rawIdsByName": raw_ids_by_name,
        "rawNamesById": raw_names_by_id,
        "missingPairs": missing_pairs,
        "sourceDigest": _sha256_file(path),
        "aliasIndex": alias_index or {},
    }


def _counter_difference_count(
    left: Counter[tuple[str, ...]],
    right: Counter[tuple[str, ...]],
) -> int:
    return sum((left - right).values())


def _amount_aggregate_mismatch_count(
    total: dict[tuple[str, ...], Decimal],
    combined: dict[tuple[str, ...], Decimal],
) -> int:
    keys = set(total) | set(combined)
    return sum(
        total.get(key, Decimal("0"))
        != combined.get(key, Decimal("0"))
        for key in keys
    )


def _audit_partition_fields(
    total: dict[str, Any],
    sales_share: dict[str, Any],
    buyout: dict[str, Any],
    scope_assessment: dict[str, Any] | None = None,
) -> dict[str, Any]:
    combined_presentation = (
        sales_share["presentationKeys"]
        + buyout["presentationKeys"]
    )
    combined_nonzero_critical = (
        sales_share["nonzeroCriticalKeys"]
        + buyout["nonzeroCriticalKeys"]
    )
    combined_amounts = dict(sales_share["criticalAmounts"])
    for key, value in buyout["criticalAmounts"].items():
        combined_amounts[key] = (
            combined_amounts.get(key, Decimal("0")) + value
        )
    presentation_missing = _counter_difference_count(
        total["presentationKeys"],
        combined_presentation,
    )
    presentation_extra = _counter_difference_count(
        combined_presentation,
        total["presentationKeys"],
    )
    nonzero_critical_missing = _counter_difference_count(
        total["nonzeroCriticalKeys"],
        combined_nonzero_critical,
    )
    nonzero_critical_extra = _counter_difference_count(
        combined_nonzero_critical,
        total["nonzeroCriticalKeys"],
    )
    aggregate_mismatch_count = _amount_aggregate_mismatch_count(
        total["criticalAmounts"],
        combined_amounts,
    )
    zero_cash_difference_only = (
        nonzero_critical_missing == 0
        and nonzero_critical_extra == 0
        and aggregate_mismatch_count == 0
    )
    critical_fields_consistent = all([
        zero_cash_difference_only,
        total["profile"]["invalidAmountCount"] == 0,
        sales_share["profile"]["invalidAmountCount"] == 0,
        buyout["profile"]["invalidAmountCount"] == 0,
        total["profile"]["missingMonthCount"] == 0,
        sales_share["profile"]["missingMonthCount"] == 0,
        buyout["profile"]["missingMonthCount"] == 0,
        sales_share["profile"]["missingWorkMappingCount"] == 0,
        sales_share["profile"]["missingStableSourceIdentityCount"] == 0,
    ])
    has_presentation_difference = (
        presentation_missing > 0 or presentation_extra > 0
    )
    private_difference_rows = []
    for difference_index, (key, count) in enumerate(sorted(
        (combined_presentation - total["presentationKeys"]).items(),
        key=lambda item: _stable_hash(item[0]),
    ), start=1):
        try:
            difference_amount = Decimal(key[-2])
        except InvalidOperation:
            difference_amount = None
        source = (
            sales_share
            if sales_share["presentationKeys"].get(key, 0)
            > total["presentationKeys"].get(key, 0)
            else buyout
        )
        scope_key = source["scopeKeyByPresentation"].get(key)
        first_positive_month = (
            source["firstPositiveMonthByWorkIdentity"].get(scope_key)
            if scope_key is not None else None
        )
        origin_mature_source = (
            first_positive_month is not None
            and _month_index(HPSR02_ORIGIN)
            - _month_index(first_positive_month)
            + 1 >= 3
        )
        work_total_scope_relevant = (
            key[-1] == "sales_share" and origin_mature_source
        )
        private_difference_rows.append({
            "issueId": f"HPSR02-FIELD-{difference_index:02d}",
            "redactedFactIdentity": f"fact-{_stable_hash(key)[:12]}",
            "billMonth": _normalize_month(key[0]),
            "redactedWorkIdentity": (
                f"work-{_stable_hash(key[4])[:12]}"
            ),
            "redactedSourceIdentity": (
                f"source-{_stable_hash([key[1], key[2]])[:12]}"
            ),
            "cashCategory": key[-1],
            "amountSign": (
                "NEGATIVE_REVERSAL"
                if difference_amount is not None
                and difference_amount < 0
                else "POSITIVE_CASH"
                if difference_amount is not None
                and difference_amount > 0
                else "INVALID_AMOUNT"
                if difference_amount is None
                else "ZERO_CASH"
            ),
            "firstPositiveMonth": first_positive_month,
            "originObservedBeforeOrAt2026_03": (
                first_positive_month is not None
            ),
            "originMatureThreeCompleteMonths": origin_mature_source,
            "workTotalScopeRelevantBeforeCore80": (
                work_total_scope_relevant
            ),
            "extraMultiplicity": count,
            "totalLedgerMultiplicity": total[
                "presentationKeys"
            ].get(key, 0),
            "splitLedgerMultiplicity": combined_presentation.get(key, 0),
            "privateRawPresentationValues": {
                "billMonth": key[0],
                "rawChannelId": key[1],
                "rawChannelName": key[2],
                "authorizationCategory": key[3],
                "rawWorkId": key[4],
                "workName": key[5],
                "amount": key[6],
                "cashCategory": key[7],
            },
            "privateScopeIdentity": {
                "standardWorkId": (
                    scope_key[0] if scope_key is not None else None
                ),
                "channelMemberId": source[
                    "canonicalMemberByPresentation"
                ].get(key),
            },
            "localLocations": {
                "totalLedger": {
                    "repositoryRelativeWorkbook": total["profile"][
                        "repositoryRelativePath"
                    ],
                    "worksheetRowOrdinals": total[
                        "presentationLocations"
                    ].get(key, []),
                },
                "splitLedger": {
                    "repositoryRelativeWorkbook": source["profile"][
                        "repositoryRelativePath"
                    ],
                    "worksheetRowOrdinals": source[
                        "presentationLocations"
                    ].get(key, []),
                },
            },
            "allNonTypePresentationFieldsMatch": True,
            "evaluationImpact": [
                "amount_conservation",
                "deduplication",
                "reversal_restatement_scope",
                *(
                    ["work_total_actual_if_work_is_dynamic_core80"]
                    if work_total_scope_relevant else []
                ),
            ],
        })
    scope_items = {
        item.get("redactedFactIdentity"): item
        for item in (scope_assessment or {}).get("items", [])
        if item.get("redactedFactIdentity")
    }
    assessment_complete = (
        scope_assessment is not None
        and scope_assessment.get("status")
        == "WORK_TOTAL_SCOPE_ASSESSMENT_COMPLETE"
        and set(scope_items) == {
            item["redactedFactIdentity"]
            for item in private_difference_rows
        }
    )
    for item in private_difference_rows:
        assessed = scope_items.get(item["redactedFactIdentity"], {})
        item["pairEligibleAtOrigin"] = (
            assessed.get("pairEligibleAtOrigin")
            if assessment_complete else None
        )
        item["workInDynamicCore80"] = (
            assessed.get("workInDynamicCore80")
            if assessment_complete else None
        )
        item["workTotalScopeRelevant"] = (
            item["workTotalScopeRelevantBeforeCore80"]
            and item["workInDynamicCore80"] is True
            if assessment_complete else None
        )
    work_total_scope_relevant_before_core_count = sum(
        item["extraMultiplicity"]
        for item in private_difference_rows
        if item["workTotalScopeRelevantBeforeCore80"]
    )
    work_total_scope_relevant_difference_count = (
        sum(
            item["extraMultiplicity"]
            for item in private_difference_rows
            if item["workTotalScopeRelevant"] is True
        )
        if assessment_complete else None
    )
    raw_critical_fields_consistent = critical_fields_consistent
    critical_fields_consistent = (
        critical_fields_consistent
        or (
            assessment_complete
            and presentation_missing == 0
            and len(private_difference_rows) > 0
            and work_total_scope_relevant_difference_count == 0
            and all(
                item["amountSign"] != "INVALID_AMOUNT"
                for item in private_difference_rows
            )
        )
    )
    status = (
        NON_EVALUATION_METADATA_WARNING
        if raw_critical_fields_consistent and has_presentation_difference
        else "OUT_OF_WORK_TOTAL_SCOPE_FACT_DIFFERENCE_WARNING"
        if critical_fields_consistent and has_presentation_difference
        else "WORK_TOTAL_SCOPE_ASSESSMENT_REQUIRED"
        if (
            not raw_critical_fields_consistent
            and has_presentation_difference
            and not assessment_complete
        )
        else "SOURCE_AUTHORITY_FIELDS_CONSISTENT"
        if critical_fields_consistent
        else "SOURCE_AUTHORITY_CONFLICT"
    )
    return {
        "status": status,
        "evaluationCriticalFields": [
            "standardWorkId",
            "billMonth",
            "amount",
            "currencyScope",
            "cashCategory",
            "stableRawSourceIdentity",
            "nonzeroRowIdentity",
            "reversalSignAndScope",
        ],
        "nonEvaluationFields": [
            "displayWorkName",
            "authorizationDisplayCategory",
            "sourceDisplayNameWhenStableRawIdExists",
            "importPresentationMetadata",
            "zeroCashPresentationRows",
        ],
        "presentationMissingFromSplitRowCount": presentation_missing,
        "presentationExtraInSplitRowCount": presentation_extra,
        "nonzeroCriticalMissingFromSplitRowCount": (
            nonzero_critical_missing
        ),
        "nonzeroCriticalExtraInSplitRowCount": nonzero_critical_extra,
        "criticalAmountAggregateMismatchCount": (
            aggregate_mismatch_count
        ),
        "zeroCashOrDisplayOnlyDifference": zero_cash_difference_only,
        "rawLedgerEvaluationCriticalFieldsConsistent": (
            raw_critical_fields_consistent
        ),
        "workTotalScopeRelevantDifferenceRowCount": (
            work_total_scope_relevant_difference_count
        ),
        "workTotalScopeRelevantBeforeCore80DifferenceRowCount": (
            work_total_scope_relevant_before_core_count
        ),
        "workTotalScopeAssessmentStatus": (
            "WORK_TOTAL_SCOPE_ASSESSMENT_COMPLETE"
            if assessment_complete
            else "WORK_TOTAL_SCOPE_ASSESSMENT_MISSING_OR_STALE"
        ),
        "evaluationCriticalFieldsConsistent": (
            critical_fields_consistent
        ),
        "sourceAuthorityPriority": (
            "USER_REVIEWED_SPLIT_MEMBERSHIP_FOR_CASH_CATEGORY_"
            "TOTAL_LEDGER_FOR_CONSERVATION_ONLY"
        ),
        "workTotalBlocking": not critical_fields_consistent,
        "privateDifferenceRows": private_difference_rows[:6],
    }


def _audit_missing_channel_pairs(
    sales_share: dict[str, Any],
) -> dict[str, Any]:
    issues = []
    hard_block_count = 0
    alias_wiring_count = 0
    raw_identity_fallback_count = 0
    for index, (raw_pair, value) in enumerate(sorted(
        sales_share["missingPairs"].items(),
        key=lambda item: _stable_hash(item[0]),
    ), start=1):
        identity = value["identity"]
        alias_targets = (
            sales_share["aliasIndex"].get(identity, set())
            if identity is not None else set()
        )
        identity_stable = identity is not None
        ambiguous_alias = len(alias_targets) > 1
        affected_work_identity_pairs = {
            (work_id, value["identityToken"])
            for work_id in value["workIds"]
        }
        observed_before_origin_count = sum(
            pair in sales_share["positiveOccurrences"]
            for pair in affected_work_identity_pairs
        )
        future_first_count = (
            len(affected_work_identity_pairs)
            - observed_before_origin_count
        )
        duplicate_nonzero_count = sum(
            count - 1
            for key, count in value["criticalKeys"].items()
            if key[-2] != "ZERO_CASH" and count > 1
        )
        raw_channel_id, raw_channel_name = raw_pair
        same_name_different_id = (
            bool(raw_channel_name)
            and len(sales_share["rawIdsByName"].get(
                raw_channel_name,
                set(),
            )) > 1
        )
        stable_id_prevents_name_merge = bool(raw_channel_id)
        hard_block_reasons = []
        if not identity_stable:
            hard_block_reasons.append("STABLE_SOURCE_IDENTITY_MISSING")
        if ambiguous_alias:
            hard_block_reasons.append(
                "STABLE_SOURCE_IDENTITY_MAPS_MULTIPLE_CANONICAL_MEMBERS"
            )
        if value["invalidAmountCount"] > 0:
            hard_block_reasons.append("AMOUNT_INVALID_OR_MISSING")
        if value["missingWorkCount"] > 0:
            hard_block_reasons.append("STANDARD_WORK_ID_MISSING")
        if duplicate_nonzero_count > 0:
            hard_block_reasons.append("NONZERO_DUPLICATE_RISK")
        if same_name_different_id and not stable_id_prevents_name_merge:
            hard_block_reasons.append(
                "SAME_NAME_DIFFERENT_CHANNEL_IDENTITY_AMBIGUOUS"
            )
        resolution = (
            "AUTHORITATIVE_ALIAS_WIRING_AVAILABLE"
            if len(alias_targets) == 1
            else "STABLE_RAW_SOURCE_IDENTITY_WORK_TOTAL_ONLY"
        )
        alias_wiring_count += int(
            resolution == "AUTHORITATIVE_ALIAS_WIRING_AVAILABLE"
        )
        raw_identity_fallback_count += int(
            resolution
            == "STABLE_RAW_SOURCE_IDENTITY_WORK_TOTAL_ONLY"
        )
        hard_block_count += int(bool(hard_block_reasons))
        issues.append({
            "issueId": f"HPSR02-CHANNEL-{index:02d}",
            "redactedSourceIdentity": (
                f"source-{_stable_hash(raw_pair)[:12]}"
            ),
            "rowCount": value["rowCount"],
            "workCount": len(value["workIds"]),
            "months": sorted(value["months"]),
            "stableSourceIdentity": identity_stable,
            "sourceIdentityKind": (
                identity[0] if identity is not None else None
            ),
            "salesShareMembershipConfirmed": True,
            "workMonthCurrencyAmountComplete": (
                value["invalidAmountCount"] == 0
                and value["missingWorkCount"] == 0
            ),
            "originVisibilityJudgeableFromSourceIdentity": (
                identity_stable
            ),
            "originObservedWorkSourceCount": (
                observed_before_origin_count
            ),
            "futureFirstObservedWorkSourceCount": future_first_count,
            "nonzeroDuplicateRiskCount": duplicate_nonzero_count,
            "sameNameDifferentChannelRiskIsolatedByStableId": (
                not same_name_different_id
                or stable_id_prevents_name_merge
            ),
            "resolution": resolution,
            "changesWorkTotalActual": False,
            "changesDynamicCore80": False,
            "changesCashBands": False,
            "changesR0R2SameCaseRelation": False,
            "hardBlockReasons": hard_block_reasons,
            "workTotalBlocking": bool(hard_block_reasons),
        })
    status = (
        "SOURCE_AUTHORITY_CONFLICT"
        if hard_block_count > 0
        else WORK_TOTAL_MAPPING_WARNING
        if issues
        else "CANONICAL_MAPPING_COMPLETE"
    )
    return {
        "status": status,
        "missingCanonicalRawPairCount": len(issues),
        "missingCanonicalMappingRowCount": sum(
            item["rowCount"] for item in issues
        ),
        "authoritativeAliasWiringCount": alias_wiring_count,
        "stableRawIdentityFallbackCount": raw_identity_fallback_count,
        "hardBlockIssueCount": hard_block_count,
        "workTotalUsesCanonicalChannelAsFeature": False,
        "workTotalUsesCanonicalChannelForCashBand": False,
        "workChannelStatus": "PARTIAL_NOT_ACTIVE",
        "canonicalMappingGuessedOrBackfilled": False,
        "issues": issues,
        "workTotalBlocking": hard_block_count > 0,
    }


def _write_source_decision_table_if_blocked(
    *,
    mapping_audit: dict[str, Any],
    partition_audit: dict[str, Any],
    sales_share: dict[str, Any],
) -> None:
    if (
        not mapping_audit["workTotalBlocking"]
        and partition_audit["status"] != "SOURCE_AUTHORITY_CONFLICT"
    ):
        return
    rows = []
    for issue in mapping_audit["issues"]:
        if not issue["workTotalBlocking"]:
            continue
        raw_value = next(
            value
            for value in sales_share["missingPairs"].values()
            if (
                f"source-{_stable_hash(value['rawPair'])[:12]}"
                == issue["redactedSourceIdentity"]
            )
        )
        rows.append({
            "stableIssueId": issue["issueId"],
            "issueTypeZh": "渠道映射 / 原始来源身份冲突",
            "redactedIdentity": issue["redactedSourceIdentity"],
            "privateRawIdentity": list(raw_value["rawPair"]),
            "localLocation": {
                "repositoryRelativeWorkbook": sales_share["profile"][
                    "repositoryRelativePath"
                ],
                "worksheetRowOrdinals": raw_value["rowOrdinals"],
            },
            "rowCount": issue["rowCount"],
            "workCount": issue["workCount"],
            "months": issue["months"],
            "candidateInterpretations": [
                "保持为独立稳定原始来源身份",
                "选择已有人工权威 canonical member（需用户确认）",
            ],
            "workTotalImpact": issue["hardBlockReasons"],
            "codexRecommendation": (
                "先确认源身份或去重关系，再恢复作品总额评价"
            ),
            "confidence": "HIGH",
            "userFinalChoice": None,
        })
    if (
        partition_audit["status"] == "SOURCE_AUTHORITY_CONFLICT"
        and len(rows) < 6
    ):
        for item in partition_audit["privateDifferenceRows"]:
            if len(rows) >= 6:
                break
            rows.append({
                "stableIssueId": item["issueId"],
                "issueTypeZh": "总表与分表非零事实行数冲突",
                "redactedIdentity": item["redactedFactIdentity"],
                "privateRawCriticalValues": item[
                    "privateRawPresentationValues"
                ],
                "localLocation": item["localLocations"],
                "rowCount": item["extraMultiplicity"],
                "workCount": 1,
                "months": [item["billMonth"]],
                "candidateInterpretations": [
                    (
                        "分表多出的事实是重复行，应只计一次并维持总表守恒"
                    ),
                    (
                        "分表多出的事实是权威新增现金，总表漏行且应修订总表"
                    ),
                ],
                "workTotalImpact": item["evaluationImpact"],
                "codexRecommendation": (
                    "逐项确认该非零事实应计一次还是两次；确认前不读取 outcome"
                ),
                "confidence": "HIGH",
                "userFinalChoice": None,
            })
    _write_atomic(SOURCE_AUTHORITY_DECISION_TABLE, {
        "schema": "m2.current.hpsr02.source_decision_table.private.v0.2",
        "artifactClass": "PRIVATE_RUN_PROVENANCE",
        "tracked": False,
        "status": (
            "M2_HPSR02_BLOCKED_ACTIONABLE_"
            "SOURCE_AUTHORITY_DECISION_REQUIRED"
        ),
        "instructionsZh": "仅填写 userFinalChoice；最多六项。",
        "items": rows[:6],
    })


def _read_historical_outcome_provenance() -> dict[str, Any]:
    if not CHAM_RECEIPT.is_file() or not CHAM_MANIFEST.is_file():
        return {
            "status": "OPTIONAL_PROVENANCE_MISSING",
            "completeOutcomePreviouslyProduced": None,
            "featureRowsBoundByFrozenManifest": None,
            "failedAttemptTouchedMetadataOnly": None,
            "failedAttemptOpenedOutcome": None,
            "evidenceRefs": [
                "PUBLIC_FROZEN_AUTHORITY:"
                "M2-head-protected-segmented-router-opened-origin-"
                "semantics-v0.2"
            ],
        }
    receipt = json.loads(CHAM_RECEIPT.read_text(encoding="utf-8"))
    manifest = json.loads(CHAM_MANIFEST.read_text(encoding="utf-8"))
    if (
        receipt.get("status") != "COMPLETE_RESULT_FROZEN"
        or receipt.get("completeMetricsProduced") is not True
        or receipt.get("validCompleteInterpretableResultProduced") is not True
        or manifest.get("status") != "COMPLETE_RESULT_FROZEN"
        or manifest.get("resultFrozen") is not True
        or manifest.get("outputBindings", {})
        .get("featureRows", {})
        .get("rowCount", 0)
        <= 0
    ):
        raise DateAuditError(
            "historical outcome provenance cannot prove actual-opened boundary"
        )
    recovery = receipt.get("recovery", {})
    return {
        "status": "PRIVATE_RUN_PROVENANCE_AVAILABLE",
        "completeOutcomePreviouslyProduced": True,
        "featureRowsBoundByFrozenManifest": True,
        "failedAttemptTouchedMetadataOnly": (
            recovery.get("partialOutcomeInspected") is False
            and recovery.get("priorAttemptId") is not None
        ),
        "failedAttemptOpenedOutcome": False,
        "evidenceRefs": [
            "PRIVATE_RUN_PROVENANCE:core-horizon-amount-complete-receipt",
            "PRIVATE_DERIVED_CACHE:core-horizon-amount-frozen-manifest",
            "PUBLIC_REPORT:M2-core-legacy-horizon-amount-development-v0.1",
        ],
    }


def _write_atomic(path: Path, value: dict[str, Any]) -> None:
    serialized = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
    if (
        path.exists()
        and path.read_text(encoding="utf-8").replace("\r\n", "\n")
        == serialized
    ):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(serialized)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    finally:
        temporary = Path(temporary_name)
        if temporary.exists():
            temporary.unlink()


def _load_current_scope_assessment(
    *,
    sales_share_digest: str,
    channel_master_digest: str,
) -> dict[str, Any] | None:
    if not WORK_TOTAL_SCOPE_ASSESSMENT.is_file():
        return None
    try:
        value = json.loads(
            WORK_TOTAL_SCOPE_ASSESSMENT.read_text(encoding="utf-8")
        )
    except (OSError, json.JSONDecodeError):
        return None
    if (
        value.get("schema")
        != "m2.current.hpsr02.work_total_scope_assessment.private.v0.2"
        or value.get("status")
        != "WORK_TOTAL_SCOPE_ASSESSMENT_COMPLETE"
        or value.get("origin") != HPSR02_ORIGIN
        or value.get("sourceDigests", {}).get("salesShare")
        != sales_share_digest
        or value.get("sourceDigests", {}).get("channelMaster")
        != channel_master_digest
        or value.get("futureActualOutcomeRead") is not False
        or value.get("finalHoldoutOutcomeRead") is not False
    ):
        return None
    return value


def run() -> dict[str, Any]:
    readiness = json.loads(READINESS_CONFIG.read_text(encoding="utf-8"))
    hpsr = json.loads(HPSR_CONFIG.read_text(encoding="utf-8"))
    capability, _ = reversal.load_contracts()
    sources, mapping_directory, channel_master_path = (
        reversal.authority_paths(capability)
    )
    expected_sales_share_path = (
        ROOT / readiness["privateInputs"]["salesShareLedger"]
    ).resolve()
    if sources.sales_share.resolve() != expected_sales_share_path:
        raise DateAuditError("sales-share authority path contract mismatch")
    mapping, _ = reversal.load_mapping(mapping_directory)
    channel_contract = json.loads(
        (
            ROOT / "config" / "m2-current-canonical-channel.v0.1.json"
        ).read_text(encoding="utf-8")
    )
    channel_master, channel_evidence = (
        reversal.canonical.load_channel_master(
            channel_contract,
            channel_master_path,
        )
    )
    alias_index = _canonical_alias_index(channel_master)
    total_scan = _scan_scope_aware_ledger(
        sources.total_ledger,
        expected_types={"", "买断"},
        ledger_role="total-ledger-authority",
        mapping=mapping,
        alias_index=alias_index,
    )
    bill_scan = _scan_scope_aware_ledger(
        sources.sales_share,
        expected_types={"分成"},
        ledger_role="sales-share-ledger-authority",
        mapping=mapping,
        channel_master=channel_master,
        alias_index=alias_index,
    )
    buyout_scan = _scan_scope_aware_ledger(
        sources.buyout,
        expected_types={"买断"},
        ledger_role="buyout-ledger-authority",
        mapping=mapping,
        alias_index=alias_index,
    )
    total_profile = total_scan["profile"]
    bill_profile = bill_scan["profile"]
    buyout_profile = buyout_scan["profile"]
    channel_master_digest = _sha256_file(channel_master_path)
    scope_assessment = _load_current_scope_assessment(
        sales_share_digest=bill_scan["sourceDigest"],
        channel_master_digest=channel_master_digest,
    )
    partition_audit = _audit_partition_fields(
        total_scan,
        bill_scan,
        buyout_scan,
        scope_assessment,
    )
    mapping_audit = _audit_missing_channel_pairs(bill_scan)
    split_multiset_conserved = (
        partition_audit["presentationMissingFromSplitRowCount"] == 0
        and partition_audit["presentationExtraInSplitRowCount"] == 0
    )
    metadata_split_missing_row_count = partition_audit[
        "presentationMissingFromSplitRowCount"
    ]
    metadata_split_extra_row_count = partition_audit[
        "presentationExtraInSplitRowCount"
    ]
    _write_source_decision_table_if_blocked(
        mapping_audit=mapping_audit,
        partition_audit=partition_audit,
        sales_share=bill_scan,
    )
    cache_profiles = [
        _scan_opened_cache_if_present(FEATURE_CACHE),
        _scan_opened_cache_if_present(FROZEN_LG01_CACHE),
    ]
    month_count = {
        row["billMonth"]: row["factCount"]
        for row in bill_profile["monthlyFactCounts"]
    }
    required_months = hpsr["laterOriginQualification"][
        "earliestIndependentFutureBillMonths"
    ]
    missing_required_months = [
        month
        for month in required_months
        if month_count.get(month, 0) <= 0
    ]
    work_total_source_checks_pass = all([
        total_profile["schemaValid"],
        bill_profile["schemaValid"],
        buyout_profile["schemaValid"],
        partition_audit["evaluationCriticalFieldsConsistent"],
        bill_profile["missingWorkMappingCount"] == 0,
        bill_profile["missingStableSourceIdentityCount"] == 0,
        mapping_audit["workTotalBlocking"] is False,
        not bill_profile["explicitPartialImportMarkerPresent"],
    ])
    bill_month_window_complete = all([
        not missing_required_months,
        not bill_profile["explicitPartialImportMarkerPresent"],
    ])
    complete_authoritative_bill_month_through = (
        required_months[-1]
        if bill_month_window_complete
        else hpsr["openedOriginSemantics"][
            "completeAuthoritativeBillMonthThrough"
        ]
    )
    source_authority_complete = work_total_source_checks_pass
    source_authority_status = (
        "SOURCE_AUTHORITY_AVAILABLE_FOR_WORK_TOTAL"
        if source_authority_complete
        else "SOURCE_AUTHORITY_WORK_TOTAL_SCOPE_ASSESSMENT_REQUIRED"
        if partition_audit["status"]
        == "WORK_TOTAL_SCOPE_ASSESSMENT_REQUIRED"
        else "SOURCE_AUTHORITY_CONFLICT"
    )
    provenance = _read_historical_outcome_provenance()
    max_availability_inspected_origin = hpsr["openedOriginSemantics"][
        "maxAvailabilityInspectedOrigin"
    ]
    max_actual_value_opened_origin = hpsr["openedOriginSemantics"][
        "maxActualValueOpenedOrigin"
    ]
    actual_value_opened_through = hpsr["openedOriginSemantics"][
        "actualValueOpenedThrough"
    ]
    availability_inspected_through = bill_profile["maximumBillMonth"]
    earliest_independent_origin = _add_months(
        max_actual_value_opened_origin,
        1,
    )
    earliest_future_months = [
        _add_months(earliest_independent_origin, offset)
        for offset in range(1, 4)
    ]
    prospective_final_holdout_origin = _add_months(
        earliest_independent_origin,
        3,
    )
    prospective_final_holdout_months = [
        _add_months(prospective_final_holdout_origin, offset)
        for offset in range(1, 4)
    ]
    date_window_ready = (
        _month_index(earliest_future_months[-1])
        <= _month_index(complete_authoritative_bill_month_through)
    )
    earliest_independent_ready = (
        date_window_ready and source_authority_complete
    )
    incomplete_months = [
        {
            "billMonth": month,
            "factCount": month_count.get(month, 0),
        }
        for month in missing_required_months
    ]
    ledger = {
        "schema": (
            "m2.current.head_protected_segmented_router."
            "opened_origin_semantics.private.v0.2"
        ),
        "tracked": False,
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "rebuildable": True,
        "auditMode": {
            "newFutureActualOutcomeOpened": False,
            "newModelMetricsRead": False,
            "modelFitRun": False,
            "modelEvaluationRun": False,
            "historicalCacheActualFieldUse": (
                "NULLNESS_ONLY_WHEN_CACHE_PRESENT_OTHERWISE_PUBLIC_"
                "FROZEN_BOUNDARY"
            ),
            "sourceLedgerFieldUse": (
                "FIELD_LEVEL_SOURCE_AUTHORITY_VALIDITY_EQUALITY_AND_"
                "SIGN_ONLY"
            ),
            "sourceLedgerAmountCellReadCount": (
                total_profile["amountCellReadCount"]
                + bill_profile["amountCellReadCount"]
                + buyout_profile["amountCellReadCount"]
            ),
            "sourceLedgerAmountValuesPublished": False,
            "cashAggregatedForOutcome": False,
        },
        "historicalCacheProfiles": cache_profiles,
        "sourceAuthorityReadiness": {
            "sourceAuthorityStatus": source_authority_status,
            "totalLedger": total_profile,
            "salesShareLedger": bill_profile,
            "buyoutLedger": buyout_profile,
            "splitRowMultisetConserved": split_multiset_conserved,
            "metadataSplitMissingRowCount":
                metadata_split_missing_row_count,
            "metadataSplitExtraRowCount":
                metadata_split_extra_row_count,
            "scopeAwareFieldAudit": partition_audit,
            "canonicalMappingScopeAudit": mapping_audit,
            "workMappingValid":
                bill_profile["missingWorkMappingCount"] == 0,
            "canonicalChannelMappingValid":
                bill_profile[
                    "missingCanonicalChannelMappingCount"
                ] == 0,
            "canonicalChannelCount":
                channel_evidence["canonicalChannelCount"],
            "workTotalGate": {
                "grain": "WORK_TOTAL",
                "status": (
                    "ACTIVE_FOR_AUTHORIZED_DEVELOPMENT_EVALUATION_"
                    "WITH_SOURCE_WARNINGS"
                    if source_authority_complete
                    else "BLOCKED_SOURCE_AUTHORITY_CONFLICT"
                ),
                "sourceAuthorityReady": source_authority_complete,
                "canonicalMappingStatus": mapping_audit["status"],
                "metadataDifferenceStatus": partition_audit["status"],
            },
            "workChannelGate": {
                "grain": "WORK_CHANNEL",
                "status": "PARTIAL_NOT_ACTIVE",
                "canonicalMappingComplete": False,
                "activated": False,
            },
            "standardMetadataChecksPass": (
                work_total_source_checks_pass
            ),
            "workTotalSourceAuthorityChecksPass": (
                work_total_source_checks_pass
            ),
            "billMonthWindowComplete": bill_month_window_complete,
            "dateWindowReady": date_window_ready,
            "completenessAuthorityMode": (
                "REQUIRED_MONTHS_PRESENT_STANDARD_METADATA_CHECKS_PASS_"
                "NO_PARTIAL_IMPORT_MARKER"
            ),
            "stalePriorIncompleteMonthSnapshotUsed": False,
        },
        "openedSemantics": {
            "maxAvailabilityInspectedOrigin":
                max_availability_inspected_origin,
            "maxActualValueOpenedOrigin": max_actual_value_opened_origin,
            "availabilityInspectedThrough":
                availability_inspected_through,
            "actualValueOpenedThrough": actual_value_opened_through,
            "completeAuthoritativeBillMonthThrough":
                complete_authoritative_bill_month_through,
            "failedAttemptTouchedMetadataOnly": provenance.get(
                "failedAttemptTouchedMetadataOnly"
            ),
            "failedAttemptOpenedOutcome": provenance.get(
                "failedAttemptOpenedOutcome"
            ),
            "unknownOrAmbiguous": False,
            "evidenceRefs": provenance["evidenceRefs"],
            "historicalReceiptStatus": provenance["status"],
        },
        "billMonthAvailability": {
            **bill_profile,
            "completeAuthoritativeBillMonthThrough":
                complete_authoritative_bill_month_through,
            "incompleteMonths": incomplete_months,
            "requiredMonths": required_months,
            "requiredMonthsPresent": not missing_required_months,
            "standardMetadataChecksPass": work_total_source_checks_pass,
            "workTotalSourceAuthorityChecksPass": (
                work_total_source_checks_pass
            ),
        },
        "prospectiveReservation": {
            "firstIndependentLaterOrigin": earliest_independent_origin,
            "firstIndependentFutureBillMonths": earliest_future_months,
            "firstIndependentRequiredCompleteThrough":
                earliest_future_months[-1],
            "firstIndependentLaterOriginReady":
                earliest_independent_ready,
            "prospectiveFinalHoldoutOrigin":
                prospective_final_holdout_origin,
            "prospectiveFinalHoldoutFutureBillMonths":
                prospective_final_holdout_months,
            "prospectiveFinalHoldoutOpened": False,
            "prospectiveFinalHoldoutOutcomeRead": False,
        },
        "semanticRevision": {
            "priorEarliestPotentialLaterOrigin": "2026-05",
            "correctedEarliestIndependentLaterOrigin":
                earliest_independent_origin,
            "changedEarliestLaterOrigin": (
                earliest_independent_origin != "2026-05"
            ),
            "reason": (
                "AVAILABILITY_METADATA_NO_LONGER_COUNTS_AS_ACTUAL_VALUE_OPENED"
            ),
            "historicalRawReceiptModified": False,
        },
        "decision": (
            "M2_HPSR02_WAITING_FOR_WORK_TOTAL_SCOPE_ASSESSMENT"
            if partition_audit["status"]
            == "WORK_TOTAL_SCOPE_ASSESSMENT_REQUIRED"
            else
            "M2_HPSR02_BLOCKED_ACTIONABLE_"
            "SOURCE_AUTHORITY_DECISION_REQUIRED"
            if not source_authority_complete
            else (
                "M2_HPSR02_WORK_TOTAL_SOURCE_AUTHORITY_RECONCILED_"
                "READY_FOR_AUTHORIZED_FIRST_INDEPENDENT_EVALUATION"
                if earliest_independent_ready
                else "M2_HPSR02_WAITING_FOR_COMPLETE_AUTHORITATIVE_BILLS"
            )
        ),
    }
    _write_atomic(SOURCE_RECONCILIATION, {
        "schema": "m2.current.hpsr02.source_authority_reconciliation.private.v0.2",
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "tracked": False,
        "sourceDigests": {
            "totalLedger": total_scan["sourceDigest"],
            "salesShare": bill_scan["sourceDigest"],
            "buyout": buyout_scan["sourceDigest"],
            "channelMaster": _sha256_file(channel_master_path),
        },
        "auditBoundary": {
            "sourceAuthorityCriticalAmountCellsValidated": True,
            "futureActualOutcomeOpened": False,
            "modelRun": False,
            "scoreComputed": False,
            "amountValuesPublished": False,
        },
        "partitionAudit": partition_audit,
        "canonicalMappingAudit": mapping_audit,
        "workTotalGate": ledger["sourceAuthorityReadiness"][
            "workTotalGate"
        ],
        "workChannelGate": ledger["sourceAuthorityReadiness"][
            "workChannelGate"
        ],
        "decision": ledger["decision"],
    })
    _write_atomic(PRIVATE_LEDGER, ledger)
    return {
        "schema": "m2.current.hpsr_date_audit_stdout.v0.2",
        "maxAvailabilityInspectedOrigin":
            max_availability_inspected_origin,
        "maxActualValueOpenedOrigin": max_actual_value_opened_origin,
        "availabilityInspectedThrough": availability_inspected_through,
        "actualValueOpenedThrough": actual_value_opened_through,
        "completeAuthoritativeBillMonthThrough":
            complete_authoritative_bill_month_through,
        "incompleteMonths": incomplete_months,
        "earliestIndependentLaterOrigin": earliest_independent_origin,
        "earliestIndependentRequiredCompleteThrough":
            earliest_future_months[-1],
        "prospectiveFinalHoldoutOrigin":
            prospective_final_holdout_origin,
        "earliestIndependentLaterOriginReady":
            earliest_independent_ready,
        "decision": ledger["decision"],
        "newFutureActualOutcomeOpened": False,
        "newModelMetricsRead": False,
        "sourceAuthorityStatus": source_authority_status,
        "derivedCacheStatus": (
            "CACHE_MISS_REBUILDABLE"
            if any(
                profile.get("status") == "CACHE_MISS_REBUILDABLE"
                for profile in cache_profiles
            )
            else "CACHE_READY"
        ),
        "historicalReceiptStatus": provenance["status"],
        "standardMetadataChecksPass": work_total_source_checks_pass,
        "workTotalSourceAuthorityChecksPass": (
            work_total_source_checks_pass
        ),
        "workTotalCanonicalMappingStatus": mapping_audit["status"],
        "metadataDifferenceStatus": partition_audit["status"],
        "workChannelGateStatus": "PARTIAL_NOT_ACTIVE",
        "billMonthWindowComplete": bill_month_window_complete,
        "dateWindowReady": date_window_ready,
        "requiredBillMonths": required_months,
        "sourceAuthorityCriticalAmountCellsValidated": (
            total_profile["amountCellReadCount"]
            + bill_profile["amountCellReadCount"]
            + buyout_profile["amountCellReadCount"]
        ),
        "amountValuesPublished": False,
    }


def _write_ndjson_atomic(path: Path, rows: list[dict[str, Any]]) -> None:
    serialized = "".join(
        json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n"
        for row in rows
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(serialized)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    finally:
        temporary = Path(temporary_name)
        if temporary.exists():
            temporary.unlink()


def export_origin_visible_authority() -> dict[str, Any]:
    """Export only origin-visible rows for a no-outcome population audit."""
    capability, _ = reversal.load_contracts()
    sources, mapping_directory, channel_master_path = (
        reversal.authority_paths(capability)
    )
    mapping, mapping_digests = reversal.load_mapping(mapping_directory)
    channel_contract = json.loads(
        (
            ROOT / "config" / "m2-current-canonical-channel.v0.1.json"
        ).read_text(encoding="utf-8")
    )
    channel_master, _ = reversal.canonical.load_channel_master(
        channel_contract,
        channel_master_path,
    )
    source_digest = _sha256_file(sources.sales_share)
    rows = []
    maximum_scale_power = 0
    negative_count = 0
    workbook = load_workbook(
        sources.sales_share,
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [
            cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        if headers != [*reversal.REAL_BILL_COLUMNS, reversal.TYPE_COLUMN]:
            raise DateAuditError("origin-visible ledger schema mismatch")
        for row_ordinal, row in enumerate(worksheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=8,
            values_only=True,
        ), start=2):
            if all(value is None for value in row):
                continue
            if reversal.clean_text(row[7]) != "分成":
                raise DateAuditError(
                    "origin-visible ledger sales-share type invalid"
                )
            month = _normalize_month(row[0])
            if month is None:
                raise DateAuditError(
                    "origin-visible ledger month invalid"
                )
            if month > HPSR02_ORIGIN:
                continue
            raw_work_id = reversal.normalize_raw_work_id(row[4])
            standard_work_id = mapping.get(raw_work_id)
            if not standard_work_id:
                standard_work_id = reversal.derive_standard_work_id(
                    raw_work_id
                )
            raw_pair = (
                reversal.canonical.clean(row[1]),
                reversal.canonical.clean(row[2]),
            )
            channel_mapping = channel_master.get(raw_pair)
            channel_member_id = reversal.canonical.clean(
                (channel_mapping or {}).get("channelUid")
            )
            amount_text, amount, scale_power = _exact_amount(row[6])
            if (
                not standard_work_id
                or not channel_member_id
                or amount_text is None
                or amount is None
            ):
                raise DateAuditError(
                    "origin-visible authority critical field missing"
                )
            maximum_scale_power = max(maximum_scale_power, scale_power)
            negative_count += int(amount < 0)
            rows.append({
                "authorityRecordId": _stable_hash([
                    source_digest,
                    row_ordinal,
                    month,
                    standard_work_id,
                    channel_member_id,
                    amount_text,
                ]),
                "authorityRowOrdinal": row_ordinal,
                "billMonth": f"{month}-01",
                "recordedAt": f"{month}-01",
                "standardWorkId": str(standard_work_id),
                "channelMemberId": channel_member_id,
                "actualSalesAmount": amount_text,
                "cashCategory": "sales_share",
                "cashCategoryAuthority": (
                    "user_reviewed_workbook_membership"
                ),
                "currencyScope": (
                    "authority_ledger_native_monetary_unit"
                ),
            })
    finally:
        workbook.close()
    if not rows:
        raise DateAuditError("origin-visible authority rows empty")
    _write_ndjson_atomic(ORIGIN_VISIBLE_AUTHORITY_FACTS, rows)
    receipt = {
        "schema": (
            "m2.current.hpsr02.origin_visible_authority_receipt.private.v0.2"
        ),
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "status": "READY_ORIGIN_VISIBLE_ONLY_NO_FUTURE_OUTCOME",
        "origin": HPSR02_ORIGIN,
        "sourceDigest": source_digest,
        "channelMasterDigest": _sha256_file(channel_master_path),
        "mappingArtifactDigests": mapping_digests,
        "factsDigest": _sha256_file(ORIGIN_VISIBLE_AUTHORITY_FACTS),
        "rowCount": len(rows),
        "negativeRowCount": negative_count,
        "amountScalePower": maximum_scale_power,
        "maximumBillMonth": HPSR02_ORIGIN,
        "futureActualOutcomeRead": False,
        "finalHoldoutOutcomeRead": False,
        "purpose": "ORIGIN_VISIBLE_DYNAMIC_CORE80_SCOPE_AUDIT_ONLY",
    }
    _write_atomic(ORIGIN_VISIBLE_AUTHORITY_RECEIPT, receipt)
    return {
        "status": receipt["status"],
        "origin": HPSR02_ORIGIN,
        "rowCount": len(rows),
        "futureActualOutcomeRead": False,
        "privateAmountValuesPrinted": False,
    }


def export_work_total_authority() -> dict[str, Any]:
    """Export the authorized WORK_TOTAL authority with stable raw fallback."""
    gate = run()
    if (
        gate.get("decision")
        != "M2_HPSR02_WORK_TOTAL_SOURCE_AUTHORITY_RECONCILED_"
        "READY_FOR_AUTHORIZED_FIRST_INDEPENDENT_EVALUATION"
        or gate.get("workTotalSourceAuthorityChecksPass") is not True
        or gate.get("workChannelGateStatus") != "PARTIAL_NOT_ACTIVE"
    ):
        raise DateAuditError("work-total source authority gate not ready")
    capability, _ = reversal.load_contracts()
    sources, mapping_directory, channel_master_path = (
        reversal.authority_paths(capability)
    )
    mapping, mapping_digests = reversal.load_mapping(mapping_directory)
    channel_contract = json.loads(
        (
            ROOT / "config" / "m2-current-canonical-channel.v0.1.json"
        ).read_text(encoding="utf-8")
    )
    channel_master, _ = reversal.canonical.load_channel_master(
        channel_contract,
        channel_master_path,
    )
    alias_index = _canonical_alias_index(channel_master)
    source_digest = _sha256_file(sources.sales_share)
    rows = []
    maximum_scale_power = 0
    negative_count = 0
    raw_identity_fallback_count = 0
    alias_wiring_count = 0
    workbook = load_workbook(
        sources.sales_share,
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [
            cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        if headers != [*reversal.REAL_BILL_COLUMNS, reversal.TYPE_COLUMN]:
            raise DateAuditError("work-total ledger schema mismatch")
        for row_ordinal, row in enumerate(worksheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=8,
            values_only=True,
        ), start=2):
            if all(value is None for value in row):
                continue
            if reversal.clean_text(row[7]) != "分成":
                raise DateAuditError("work-total ledger type invalid")
            month = _normalize_month(row[0])
            raw_work_id = reversal.normalize_raw_work_id(row[4])
            standard_work_id = mapping.get(raw_work_id)
            if not standard_work_id:
                standard_work_id = reversal.derive_standard_work_id(
                    raw_work_id
                )
            raw_pair = (
                reversal.canonical.clean(row[1]),
                reversal.canonical.clean(row[2]),
            )
            channel_mapping = channel_master.get(raw_pair)
            channel_member_id = reversal.canonical.clean(
                (channel_mapping or {}).get("channelUid")
            )
            source_identity = _source_identity(*raw_pair)
            if not channel_member_id and source_identity is not None:
                alias_targets = alias_index.get(source_identity, set())
                if len(alias_targets) == 1:
                    channel_member_id = next(iter(alias_targets))
                    alias_wiring_count += 1
                elif len(alias_targets) == 0:
                    channel_member_id = (
                        "HPSR02-RAW-SOURCE-"
                        + _stable_hash(source_identity)[:32]
                    )
                    raw_identity_fallback_count += 1
                else:
                    raise DateAuditError(
                        "work-total source identity alias ambiguous"
                    )
            amount_text, amount, scale_power = _exact_amount(row[6])
            if (
                month is None
                or not standard_work_id
                or not channel_member_id
                or amount_text is None
                or amount is None
            ):
                raise DateAuditError(
                    "work-total authority critical field missing"
                )
            maximum_scale_power = max(maximum_scale_power, scale_power)
            negative_count += int(amount < 0)
            rows.append({
                "authorityRecordId": _stable_hash([
                    source_digest,
                    row_ordinal,
                    month,
                    standard_work_id,
                    channel_member_id,
                    amount_text,
                ]),
                "authorityRowOrdinal": row_ordinal,
                "billMonth": f"{month}-01",
                "recordedAt": f"{month}-01",
                "standardWorkId": str(standard_work_id),
                "channelMemberId": channel_member_id,
                "actualSalesAmount": amount_text,
                "cashCategory": "sales_share",
                "cashCategoryAuthority": (
                    "user_reviewed_workbook_membership"
                ),
                "currencyScope": (
                    "authority_ledger_native_monetary_unit"
                ),
            })
    finally:
        workbook.close()
    if not rows:
        raise DateAuditError("work-total authority rows empty")
    _write_ndjson_atomic(WORK_TOTAL_AUTHORITY_FACTS, rows)
    receipt = {
        "schema": (
            "m2.current.hpsr02.work_total_authority_receipt.private.v0.2"
        ),
        "artifactClass": "PRIVATE_DERIVED_CACHE",
        "status": "READY_WORK_TOTAL_SCOPE_AWARE_AUTHORITY",
        "sourceDigests": {
            "salesShare": source_digest,
            "channelMaster": _sha256_file(channel_master_path),
        },
        "mappingArtifactDigests": mapping_digests,
        "factsDigest": _sha256_file(WORK_TOTAL_AUTHORITY_FACTS),
        "rowCount": len(rows),
        "negativeRowCount": negative_count,
        "amountScalePower": maximum_scale_power,
        "canonicalAliasWiringRowCount": alias_wiring_count,
        "stableRawSourceIdentityFallbackRowCount": (
            raw_identity_fallback_count
        ),
        "channelScopeMode": (
            "CANONICAL_OR_STABLE_RAW_SOURCE_IDENTITY_WORK_TOTAL_ONLY"
        ),
        "workTotalSourceAuthorityStatus": (
            "SOURCE_AUTHORITY_AVAILABLE_FOR_WORK_TOTAL"
        ),
        "workChannelGateStatus": "PARTIAL_NOT_ACTIVE",
        "futureActualOutcomeRead": True,
        "finalHoldoutOutcomeRead": False,
        "amountValuesPrinted": False,
    }
    _write_atomic(WORK_TOTAL_AUTHORITY_RECEIPT, receipt)
    return {
        "status": receipt["status"],
        "rowCount": receipt["rowCount"],
        "workTotalSourceAuthorityStatus": receipt[
            "workTotalSourceAuthorityStatus"
        ],
        "workChannelGateStatus": receipt["workChannelGateStatus"],
        "futureActualOutcomeRead": True,
        "amountValuesPrinted": False,
    }


def main() -> None:
    try:
        arguments = os.sys.argv[1:]
        result = (
            export_origin_visible_authority()
            if arguments == ["--export-origin-visible-authority"]
            else export_work_total_authority()
            if arguments == ["--export-work-total-authority"]
            else run()
        )
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    except Exception as exc:  # noqa: BLE001
        print(f"[M2_HPSR_DATE_AUDIT_ERROR] {exc}", file=os.sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
